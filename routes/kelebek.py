import os
import re
import json
import logging
import random
from io import BytesIO
import unicodedata
from datetime import datetime
from flask import Blueprint, request, jsonify, render_template, send_file
from flask_login import login_required, current_user
import psycopg2
from psycopg2.extras import RealDictCursor
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

logger = logging.getLogger(__name__)

try:
    pdfmetrics.registerFont(TTFont('DejaVuSans', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
    pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
except Exception:
    pass

PDF_FONT = 'DejaVuSans'
PDF_FONT_BOLD = 'DejaVuSans-Bold'


def tr_normalize(s):
    """Türkçe büyük/küçük harf farklarını (İ vs I, ı vs i) yok sayarak karşılaştırma anahtarı üretir."""
    if not s:
        return ''
    return s.upper().replace('İ', 'I').replace('I', 'I')


kelebek_bp = Blueprint('kelebek', __name__, url_prefix='/admin/kelebek')

DEFAULT_ROOM_CAPACITIES = {
    '5A': {'desks': 17, 'capacity': 34},
    '5B': {'desks': 17, 'capacity': 34},
    '5C': {'desks': 17, 'capacity': 34},
    '5D': {'desks': 17, 'capacity': 34},
    '6A': {'desks': 16, 'capacity': 32},
    '6B': {'desks': 16, 'capacity': 32},
    '6C': {'desks': 16, 'capacity': 32},
    '6D': {'desks': 16, 'capacity': 32},
    '7A': {'desks': 14, 'capacity': 28},
    '7B': {'desks': 15, 'capacity': 30},
    '7C': {'desks': 15, 'capacity': 30},
    '7D': {'desks': 16, 'capacity': 32},
    '8A': {'desks': 14, 'capacity': 28},
    '8B': {'desks': 14, 'capacity': 28},
    '8C': {'desks': 14, 'capacity': 28},
    '8D': {'desks': 13, 'capacity': 26},
    'Yedek Sınıf': {'desks': 14, 'capacity': 28},
}

PRIORITY_EXAM_ROOMS = ['8A', '8B', '7A', '7B', '6A', '6C']
ALWAYS_STUDY_ROOMS = ['8D', '7C']

def extract_grade_from_room(room_name):
    m = re.match(r'^(\d)[A-ZÇĞİÖŞÜa-zçğıöşü]$', room_name.strip())
    if m:
        return int(m.group(1))
    return None


def get_db():
    from app import get_db as app_get_db
    return app_get_db()


def init_kelebek_tables(conn):
    cur = conn.cursor()
    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS kelebek_room_config (
                id SERIAL PRIMARY KEY,
                room_name VARCHAR(50) NOT NULL UNIQUE,
                desks INTEGER NOT NULL,
                capacity INTEGER NOT NULL,
                default_type VARCHAR(20) DEFAULT 'auto',
                sort_order INTEGER DEFAULT 0
            );

            CREATE TABLE IF NOT EXISTS kelebek_plans (
                id SERIAL PRIMARY KEY,
                plan_name VARCHAR(200) NOT NULL,
                exam_date DATE,
                grade_levels JSONB NOT NULL,
                status VARCHAR(20) DEFAULT 'draft',
                created_by INTEGER REFERENCES users(id),
                created_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul'),
                updated_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul')
            );

            CREATE TABLE IF NOT EXISTS kelebek_rooms (
                id SERIAL PRIMARY KEY,
                plan_id INTEGER REFERENCES kelebek_plans(id) ON DELETE CASCADE,
                room_name VARCHAR(50) NOT NULL,
                desk_count INTEGER NOT NULL,
                capacity INTEGER NOT NULL,
                room_type VARCHAR(20) NOT NULL,
                grade_for_study INTEGER
            );

            CREATE TABLE IF NOT EXISTS kelebek_participants (
                id SERIAL PRIMARY KEY,
                plan_id INTEGER REFERENCES kelebek_plans(id) ON DELETE CASCADE,
                student_name VARCHAR(200) NOT NULL,
                student_no VARCHAR(20),
                class_name VARCHAR(10) NOT NULL,
                grade_level INTEGER NOT NULL,
                is_exam_taker BOOLEAN DEFAULT TRUE
            );

            CREATE TABLE IF NOT EXISTS kelebek_assignments (
                id SERIAL PRIMARY KEY,
                plan_id INTEGER REFERENCES kelebek_plans(id) ON DELETE CASCADE,
                room_id INTEGER REFERENCES kelebek_rooms(id) ON DELETE CASCADE,
                participant_id INTEGER REFERENCES kelebek_participants(id) ON DELETE CASCADE,
                desk_number INTEGER,
                seat_position VARCHAR(5),
                row_number INTEGER
            );
        """)
        conn.commit()

        cur.execute("SELECT COUNT(*) FROM kelebek_room_config")
        count = cur.fetchone()[0]
        if count == 0:
            sort_order = 0
            for room_name, info in DEFAULT_ROOM_CAPACITIES.items():
                default_type = 'auto'
                if room_name in ALWAYS_STUDY_ROOMS:
                    default_type = 'study'
                elif room_name == 'Yedek Sınıf':
                    default_type = 'exam'
                cur.execute("""
                    INSERT INTO kelebek_room_config (room_name, desks, capacity, default_type, sort_order)
                    VALUES (%s, %s, %s, %s, %s) ON CONFLICT (room_name) DO NOTHING
                """, (room_name, info['desks'], info['capacity'], default_type, sort_order))
                sort_order += 1
            conn.commit()
            logger.info("Kelebek room config seeded with defaults")

        logger.info("Kelebek tables created successfully")
    except Exception as e:
        conn.rollback()
        logger.error(f"Error creating kelebek tables: {e}")
        raise
    finally:
        cur.close()


@kelebek_bp.route('/', strict_slashes=False)
@login_required
def kelebek_page():
    if current_user.role not in ['admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    return render_template('kelebek.html')


@kelebek_bp.route('/api/room-config', methods=['GET'])
@login_required
def get_room_config():
    if current_user.role not in ['admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT * FROM kelebek_room_config ORDER BY sort_order, room_name")
        rooms = cur.fetchall()
        return jsonify({"success": True, "rooms": rooms})
    except Exception as e:
        logger.error(f"Error getting room config: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()


@kelebek_bp.route('/api/room-config', methods=['POST'])
@login_required
def add_room_config():
    if current_user.role not in ['admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    conn = None
    try:
        data = request.get_json()
        room_name = data.get('room_name', '').strip()
        desks = int(data.get('desks', 0))
        capacity = int(data.get('capacity', 0))
        default_type = data.get('default_type', 'auto')

        if not room_name:
            return jsonify({"error": "Sınıf adı gereklidir"}), 400
        if desks <= 0 or capacity <= 0:
            return jsonify({"error": "Sıra ve kapasite 0'dan büyük olmalıdır"}), 400
        if default_type not in ('auto', 'exam', 'study'):
            default_type = 'auto'

        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT COALESCE(MAX(sort_order), 0) + 1 as next_order FROM kelebek_room_config")
        next_order = cur.fetchone()['next_order']
        cur.execute("""
            INSERT INTO kelebek_room_config (room_name, desks, capacity, default_type, sort_order)
            VALUES (%s, %s, %s, %s, %s) RETURNING *
        """, (room_name, desks, capacity, default_type, next_order))
        room = cur.fetchone()
        conn.commit()
        return jsonify({"success": True, "room": room})
    except psycopg2.IntegrityError:
        if conn:
            conn.rollback()
        return jsonify({"error": f"'{room_name}' zaten mevcut"}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"Error adding room config: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()


@kelebek_bp.route('/api/room-config/<int:room_id>', methods=['PUT'])
@login_required
def update_room_config(room_id):
    if current_user.role not in ['admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    conn = None
    try:
        data = request.get_json()
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        updates = []
        params = []
        if 'room_name' in data:
            updates.append("room_name = %s")
            params.append(data['room_name'].strip())
        if 'desks' in data:
            updates.append("desks = %s")
            params.append(int(data['desks']))
        if 'capacity' in data:
            updates.append("capacity = %s")
            params.append(int(data['capacity']))
        if 'default_type' in data:
            dt = data['default_type']
            if dt not in ('auto', 'exam', 'study'):
                dt = 'auto'
            updates.append("default_type = %s")
            params.append(dt)

        if not updates:
            return jsonify({"error": "Güncellenecek alan yok"}), 400

        params.append(room_id)
        cur.execute(f"UPDATE kelebek_room_config SET {', '.join(updates)} WHERE id = %s RETURNING *", params)
        room = cur.fetchone()
        if not room:
            return jsonify({"error": "Sınıf bulunamadı"}), 404
        conn.commit()
        return jsonify({"success": True, "room": room})
    except psycopg2.IntegrityError:
        if conn:
            conn.rollback()
        return jsonify({"error": "Bu isimde bir sınıf zaten mevcut"}), 400
    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"Error updating room config: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()


@kelebek_bp.route('/api/room-config/<int:room_id>', methods=['DELETE'])
@login_required
def delete_room_config(room_id):
    if current_user.role not in ['admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM kelebek_room_config WHERE id = %s", (room_id,))
        if cur.rowcount == 0:
            return jsonify({"error": "Sınıf bulunamadı"}), 404
        conn.commit()
        return jsonify({"success": True})
    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"Error deleting room config: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()


@kelebek_bp.route('/api/create-plan', methods=['POST'])
@login_required
def create_plan():
    if current_user.role not in ['admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403

    conn = None
    try:
        data = request.get_json()
        plan_name = data.get('plan_name', '').strip()
        exam_date = data.get('exam_date')
        grade_levels = data.get('grade_levels', [])

        if not plan_name:
            return jsonify({"error": "Plan adı gereklidir"}), 400
        if not grade_levels:
            return jsonify({"error": "En az bir sınıf seviyesi seçilmelidir"}), 400

        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            INSERT INTO kelebek_plans (plan_name, exam_date, grade_levels, created_by)
            VALUES (%s, %s, %s, %s)
            RETURNING id, plan_name, exam_date, grade_levels, status, created_at
        """, (plan_name, exam_date, json.dumps(grade_levels), current_user.id))
        plan = cur.fetchone()
        conn.commit()

        for key in ['exam_date', 'created_at']:
            if plan.get(key) and hasattr(plan[key], 'isoformat'):
                plan[key] = plan[key].isoformat()

        return jsonify({"success": True, "plan": plan})
    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"Error creating plan: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()


@kelebek_bp.route('/api/plans')
@login_required
def list_plans():
    if current_user.role not in ['admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT p.*, u.full_name as created_by_name,
                   (SELECT COUNT(*) FROM kelebek_participants WHERE plan_id = p.id) as participant_count,
                   (SELECT COUNT(*) FROM kelebek_rooms WHERE plan_id = p.id) as room_count
            FROM kelebek_plans p
            LEFT JOIN users u ON p.created_by = u.id
            ORDER BY p.created_at DESC
        """)
        plans = cur.fetchall()

        for plan in plans:
            for key in ['exam_date', 'created_at', 'updated_at']:
                if plan.get(key) and hasattr(plan[key], 'isoformat'):
                    plan[key] = plan[key].isoformat()

        return jsonify({"success": True, "plans": plans})
    except Exception as e:
        logger.error(f"Error listing plans: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()


@kelebek_bp.route('/api/plan/<int:plan_id>')
@login_required
def get_plan(plan_id):
    if current_user.role not in ['admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        cur.execute("SELECT * FROM kelebek_plans WHERE id = %s", (plan_id,))
        plan = cur.fetchone()
        if not plan:
            return jsonify({"error": "Plan bulunamadı"}), 404

        for key in ['exam_date', 'created_at', 'updated_at']:
            if plan.get(key) and hasattr(plan[key], 'isoformat'):
                plan[key] = plan[key].isoformat()

        cur.execute("SELECT * FROM kelebek_rooms WHERE plan_id = %s ORDER BY room_name", (plan_id,))
        rooms = cur.fetchall()

        cur.execute("SELECT * FROM kelebek_participants WHERE plan_id = %s ORDER BY grade_level, class_name, student_name", (plan_id,))
        participants = cur.fetchall()

        cur.execute("""
            SELECT a.*, p.student_name, p.class_name, p.grade_level, p.student_no,
                   r.room_name, r.room_type
            FROM kelebek_assignments a
            JOIN kelebek_participants p ON a.participant_id = p.id
            JOIN kelebek_rooms r ON a.room_id = r.id
            WHERE a.plan_id = %s
            ORDER BY r.room_name, a.desk_number, a.seat_position
        """, (plan_id,))
        assignments = cur.fetchall()

        return jsonify({
            "success": True,
            "plan": plan,
            "rooms": rooms,
            "participants": participants,
            "assignments": assignments
        })
    except Exception as e:
        logger.error(f"Error getting plan: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()


@kelebek_bp.route('/api/plan/<int:plan_id>', methods=['DELETE'])
@login_required
def delete_plan(plan_id):
    if current_user.role not in ['admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM kelebek_plans WHERE id = %s", (plan_id,))
        conn.commit()
        return jsonify({"success": True, "message": "Plan silindi"})
    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"Error deleting plan: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()


@kelebek_bp.route('/api/excel-template')
@login_required
def download_excel_template():
    if current_user.role not in ['admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403

    wb = Workbook()
    ws = wb.active
    ws.title = "Sınav Katılımcıları"

    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['Sınıf', 'Okul Numarası', 'Adı', 'Soyadı']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    sample_data = [
        ('5A', '501', 'Ahmet', 'YILMAZ'),
        ('5A', '502', 'Ayşe', 'KAYA'),
        ('6B', '601', 'Mehmet', 'DEMİR'),
        ('7A', '701', 'Fatma', 'ÖZ'),
        ('8C', '801', 'Ali', 'CAN'),
    ]

    sample_font = Font(color="999999", italic=True)
    for row_idx, (sinif, no, adi, soyadi) in enumerate(sample_data, 2):
        for col_idx, value in enumerate([sinif, no, adi, soyadi], 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = sample_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center' if col_idx < 3 else 'left', vertical='center')

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    ws.row_dimensions[1].height = 30

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='kelebek_katilimci_sablonu.xlsx'
    )


@kelebek_bp.route('/api/upload-participants/<int:plan_id>', methods=['POST'])
@login_required
def upload_participants(plan_id):
    if current_user.role not in ['admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403

    if 'file' not in request.files:
        return jsonify({"error": "Dosya bulunamadı"}), 400

    file = request.files['file']
    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"error": "Sadece Excel dosyaları kabul edilir (.xlsx)"}), 400

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        cur.execute("SELECT * FROM kelebek_plans WHERE id = %s", (plan_id,))
        plan = cur.fetchone()
        if not plan:
            return jsonify({"error": "Plan bulunamadı"}), 404

        grade_levels = plan['grade_levels']
        if isinstance(grade_levels, str):
            grade_levels = json.loads(grade_levels)

        cur.execute("DELETE FROM kelebek_assignments WHERE plan_id = %s", (plan_id,))
        cur.execute("DELETE FROM kelebek_participants WHERE plan_id = %s", (plan_id,))
        cur.execute("DELETE FROM kelebek_rooms WHERE plan_id = %s", (plan_id,))

        wb = load_workbook(file)
        ws = wb.active

        exam_takers = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            class_name = str(row[0]).strip()
            student_no = str(row[1]).strip() if row[1] else ''
            first_name = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            last_name = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            student_name = (first_name + ' ' + last_name).strip()

            if not student_name:
                continue

            grade_level = None
            for ch in class_name:
                if ch.isdigit():
                    grade_level = int(ch)
                    break

            if grade_level is None or grade_level not in grade_levels:
                continue

            exam_takers.append({
                'student_name': student_name,
                'student_no': student_no,
                'class_name': class_name,
                'grade_level': grade_level
            })

        # Sınava giren öğrencileri hem öğrenci no hem isim bazlı indeksle
        exam_taker_keys_by_name = set()   # (normalize_isim, normalize_sınıf)
        exam_taker_keys_by_no = set()     # öğrenci_no (username)

        for et in exam_takers:
            key = (tr_normalize(et['student_name']), tr_normalize(et['class_name']))
            exam_taker_keys_by_name.add(key)
            if et['student_no']:
                exam_taker_keys_by_no.add(et['student_no'].strip())
            cur.execute("""
                INSERT INTO kelebek_participants (plan_id, student_name, student_no, class_name, grade_level, is_exam_taker)
                VALUES (%s, %s, %s, %s, %s, TRUE)
            """, (plan_id, et['student_name'], et['student_no'], et['class_name'], et['grade_level']))

        grade_placeholders = ','.join(['%s'] * len(grade_levels))
        cur.execute(f"""
            SELECT full_name, username, class_name
            FROM users
            WHERE role = 'student' AND class_name IS NOT NULL
            AND CAST(SUBSTRING(class_name FROM 1 FOR 1) AS INTEGER) IN ({grade_placeholders})
        """, grade_levels)
        all_students = cur.fetchall()

        non_exam_count = 0
        for student in all_students:
            # Önce öğrenci numarasıyla eşleştir (isim hatalarını önler)
            student_no = (student['username'] or '').strip()
            if student_no and student_no in exam_taker_keys_by_no:
                continue  # Bu öğrenci zaten sınava giriyor

            # Öğrenci no yoksa veya eşleşmediyse isimle kontrol et
            key = (tr_normalize(student['full_name']), tr_normalize(student['class_name']))
            if key in exam_taker_keys_by_name:
                continue  # İsim eşleşti, zaten sınava giriyor

            # Ne no ne isim eşleşti → ders çalışacak öğrenci
            grade_level = None
            for ch in student['class_name']:
                if ch.isdigit():
                    grade_level = int(ch)
                    break
            if grade_level is None:
                continue
            cur.execute("""
                INSERT INTO kelebek_participants (plan_id, student_name, student_no, class_name, grade_level, is_exam_taker)
                VALUES (%s, %s, %s, %s, %s, FALSE)
            """, (plan_id, student['full_name'], student['username'], student['class_name'], grade_level))
            non_exam_count += 1

        cur.execute("UPDATE kelebek_plans SET status = 'draft', updated_at = (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul') WHERE id = %s", (plan_id,))
        conn.commit()

        return jsonify({
            "success": True,
            "message": f"{len(exam_takers)} sınava giren, {non_exam_count} ders çalışacak öğrenci yüklendi",
            "exam_takers": len(exam_takers),
            "non_exam_takers": non_exam_count,
            "total": len(exam_takers) + non_exam_count
        })
    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"Error uploading participants: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()


def _run_butterfly_algorithm(plan_id, conn):
    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute("SELECT * FROM kelebek_plans WHERE id = %s", (plan_id,))
    plan = cur.fetchone()
    grade_levels = plan['grade_levels']
    if isinstance(grade_levels, str):
        grade_levels = json.loads(grade_levels)

    cur.execute("SELECT * FROM kelebek_participants WHERE plan_id = %s", (plan_id,))
    participants = cur.fetchall()

    exam_students = [p for p in participants if p['is_exam_taker']]
    non_exam_students = [p for p in participants if not p['is_exam_taker']]

    exam_by_grade = {}
    for s in exam_students:
        exam_by_grade.setdefault(s['grade_level'], []).append(s)

    exam_by_class = {}
    for s in exam_students:
        exam_by_class.setdefault(s['class_name'], []).append(s)

    non_exam_by_grade = {}
    for s in non_exam_students:
        non_exam_by_grade.setdefault(s['grade_level'], []).append(s)

    exam_count_by_class = {}
    for s in exam_students:
        exam_count_by_class[s['class_name']] = exam_count_by_class.get(s['class_name'], 0) + 1

    cur.execute("DELETE FROM kelebek_assignments WHERE plan_id = %s", (plan_id,))
    cur.execute("DELETE FROM kelebek_rooms WHERE plan_id = %s", (plan_id,))

    cur.execute("SELECT * FROM kelebek_room_config ORDER BY sort_order, room_name")
    all_room_configs = cur.fetchall()

    relevant_rooms = {}
    room_default_types = {}
    for rc in all_room_configs:
        rname = rc['room_name']
        room_grade = extract_grade_from_room(rname)
        is_relevant = (room_grade and room_grade in grade_levels) or (room_grade is None)
        if is_relevant:
            relevant_rooms[rname] = {'desks': rc['desks'], 'capacity': rc['capacity']}
            room_default_types[rname] = rc['default_type']

    exam_rooms = []
    study_rooms = []
    used_rooms = set()

    total_exam_students = len(exam_students)

    for room in relevant_rooms:
        if room_default_types.get(room) == 'study':
            study_rooms.append(room)
            used_rooms.add(room)

    candidate_exam_rooms = []
    for room in relevant_rooms:
        if room_default_types.get(room) == 'exam' and room not in used_rooms:
            candidate_exam_rooms.append(room)
            used_rooms.add(room)

    remaining_rooms = [r for r in relevant_rooms if r not in used_rooms]
    remaining_rooms.sort(key=lambda r: exam_count_by_class.get(r, 0), reverse=True)
    for room in remaining_rooms:
        candidate_exam_rooms.append(room)
        used_rooms.add(room)

    running_capacity = 0
    for room in candidate_exam_rooms:
        if running_capacity < total_exam_students:
            exam_rooms.append(room)
            running_capacity += relevant_rooms[room]['capacity']
        else:
            study_rooms.append(room)

    for room in relevant_rooms:
        if room not in set(exam_rooms) and room not in set(study_rooms):
            study_rooms.append(room)

    room_ids = {}
    for room_name in exam_rooms:
        info = relevant_rooms[room_name]
        cur.execute("""
            INSERT INTO kelebek_rooms (plan_id, room_name, desk_count, capacity, room_type, grade_for_study)
            VALUES (%s, %s, %s, %s, 'exam', NULL)
            RETURNING id
        """, (plan_id, room_name, info['desks'], info['capacity']))
        room_ids[room_name] = cur.fetchone()['id']

    for room_name in study_rooms:
        info = relevant_rooms[room_name]
        room_grade = extract_grade_from_room(room_name)
        cur.execute("""
            INSERT INTO kelebek_rooms (plan_id, room_name, desk_count, capacity, room_type, grade_for_study)
            VALUES (%s, %s, %s, %s, 'study', %s)
            RETURNING id
        """, (plan_id, room_name, info['desks'], info['capacity'], room_grade))
        room_ids[room_name] = cur.fetchone()['id']

    study_room_by_grade = {}
    for room_name in study_rooms:
        room_grade = extract_grade_from_room(room_name)
        if room_grade:
            study_room_by_grade.setdefault(room_grade, []).append(room_name)

    non_exam_by_class = {}
    for s in non_exam_students:
        non_exam_by_class.setdefault(s['class_name'], []).append(s)
    for cls in non_exam_by_class:
        random.shuffle(non_exam_by_class[cls])

    study_room_capacity = {}
    study_room_filled = {}
    for room_name in study_rooms:
        study_room_capacity[room_name] = relevant_rooms[room_name]['capacity']
        study_room_filled[room_name] = 0

    def fill_study_room(room_name, students_to_place):
        placed = []
        room_info = relevant_rooms[room_name]
        current_filled = study_room_filled[room_name]
        current_desk = (current_filled // 2) + 1
        current_pos_idx = current_filled % 2
        positions = ['left', 'right']
        s_idx = 0
        while s_idx < len(students_to_place) and current_desk <= room_info['desks']:
            start_pos = current_pos_idx if current_desk == (current_filled // 2) + 1 else 0
            for pos_i in range(start_pos, 2):
                if s_idx >= len(students_to_place):
                    break
                s = students_to_place[s_idx]
                cur.execute("""
                    INSERT INTO kelebek_assignments (plan_id, room_id, participant_id, desk_number, seat_position, row_number)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (plan_id, room_ids[room_name], s['id'], current_desk, positions[pos_i], current_desk))
                study_room_filled[room_name] += 1
                placed.append(s)
                s_idx += 1
            current_desk += 1
        return placed

    for room_name in study_rooms:
        home_students = list(non_exam_by_class.get(room_name, []))
        if home_students:
            placed = fill_study_room(room_name, home_students)
            placed_ids = set(s['id'] for s in placed)
            non_exam_by_class[room_name] = [s for s in non_exam_by_class.get(room_name, []) if s['id'] not in placed_ids]

    for room_name in study_rooms:
        remaining_cap = study_room_capacity[room_name] - study_room_filled[room_name]
        if remaining_cap <= 0:
            continue

        room_grade = extract_grade_from_room(room_name)

        same_grade_classes = sorted(
            [cls for cls in non_exam_by_class if cls != room_name and len(non_exam_by_class[cls]) > 0
             and extract_grade_from_room(cls) == room_grade],
            key=lambda cls: len(non_exam_by_class[cls]), reverse=True
        ) if room_grade else []

        for cls in same_grade_classes:
            if study_room_filled[room_name] >= study_room_capacity[room_name]:
                break
            available = non_exam_by_class[cls]
            if not available:
                continue
            placed = fill_study_room(room_name, available)
            placed_ids = set(s['id'] for s in placed)
            non_exam_by_class[cls] = [s for s in non_exam_by_class[cls] if s['id'] not in placed_ids]

    all_unplaced_study = []
    for cls, students in non_exam_by_class.items():
        all_unplaced_study.extend(students)
    random.shuffle(all_unplaced_study)

    if all_unplaced_study:
        empty_study_rooms = [rn for rn in study_rooms if study_room_filled[rn] == 0]
        for room_name in empty_study_rooms:
            if not all_unplaced_study:
                break
            cur.execute("UPDATE kelebek_rooms SET grade_for_study = NULL WHERE id = %s", (room_ids[room_name],))
            placed = fill_study_room(room_name, all_unplaced_study)
            placed_ids = set(s['id'] for s in placed)
            all_unplaced_study = [s for s in all_unplaced_study if s['id'] not in placed_ids]

    for cls in exam_by_class:
        random.shuffle(exam_by_class[cls])

    class_queues = {}
    for cls, students in exam_by_class.items():
        class_queues[cls] = list(students)

    overflow_students = []

    # --- Yardımcı fonksiyonlar: sınıf seviyesi eşleştirme ve şube harfi ---
    exam_grades_in_plan = set(s['grade_level'] for s in exam_students)

    def get_section(class_name):
        """Sınıf adından şube harfini çıkar: '7A' -> 'A', '8B' -> 'B'"""
        if class_name and len(class_name) >= 2:
            last = class_name[-1].upper()
            if last.isalpha():
                return last
        return None

    def get_preferred_grades(grade):
        """
        Hangi sınıf seviyeleri yanyana oturmalı?
        - 5. sınıf varsa: 5↔6, 7↔8
        - 5. sınıf yoksa: herhangi farklı sınıf
        """
        has_5 = 5 in exam_grades_in_plan
        has_6 = 6 in exam_grades_in_plan
        if has_5 and has_6:
            if grade == 5:
                return [6]
            if grade == 6:
                return [5]
            if grade == 7:
                return [8]
            if grade == 8:
                return [7]
        return sorted([g for g in exam_grades_in_plan if g != grade])

    def pick_right_student(left_student, rbg):
        """
        Sol koltukta oturan öğrenciye en uygun sağ koltuk öğrencisini seç.
        Öncelik sırası:
          1. Tercih edilen sınıf seviyesi + aynı şube harfi
          2. Tercih edilen sınıf seviyesi (herhangi şube)
          3. Herhangi farklı sınıf seviyesi + aynı şube harfi (fallback)
          4. Herhangi farklı sınıf seviyesi
          5. Aynı sınıf seviyesi (son çare)
        """
        lg = left_student.get('grade_level')
        ls = get_section(left_student.get('class_name', ''))
        preferred = get_preferred_grades(lg)

        if ls:
            for pg in preferred:
                for i, c in enumerate(rbg.get(pg, [])):
                    if get_section(c.get('class_name', '')) == ls:
                        return rbg[pg].pop(i)

        for pg in preferred:
            if rbg.get(pg):
                return rbg[pg].pop(0)

        if ls:
            for g in sorted(rbg.keys()):
                if g == lg:
                    continue
                for i, c in enumerate(rbg.get(g, [])):
                    if get_section(c.get('class_name', '')) == ls:
                        return rbg[g].pop(i)

        for g in sorted(rbg.keys()):
            if g != lg and rbg.get(g):
                return rbg[g].pop(0)

        for g in sorted(rbg.keys()):
            if rbg.get(g):
                return rbg[g].pop(0)

        return None

    # --- 1. Tur: Ev sınıfındaki öğrencileri sol sıralara yerleştir ---
    for room_name in exam_rooms:
        room_info = relevant_rooms[room_name]
        room_id = room_ids[room_name]
        num_desks = room_info['desks']

        home_students = list(class_queues.get(room_name, []))
        home_for_left = home_students[:num_desks]
        home_overflow = home_students[num_desks:]

        if room_name in class_queues:
            class_queues[room_name] = []

        overflow_students.extend(home_overflow)

        for desk_num, student in enumerate(home_for_left, 1):
            cur.execute("""
                INSERT INTO kelebek_assignments (plan_id, room_id, participant_id, desk_number, seat_position, row_number)
                VALUES (%s, %s, %s, %s, 'left', %s)
            """, (plan_id, room_id, student['id'], desk_num, desk_num))

    # --- Kalan öğrencileri topla: sınıf seviyesine göre grupla, şubeye göre sırala ---
    all_remaining = []
    for cls, students in class_queues.items():
        all_remaining.extend(students)
    all_remaining.extend(overflow_students)

    remaining_by_grade = {}
    for s in all_remaining:
        remaining_by_grade.setdefault(s['grade_level'], []).append(s)

    for g in remaining_by_grade:
        remaining_by_grade[g].sort(key=lambda s: (get_section(s.get('class_name', '')) or 'Z'))

    # --- 2. Tur: Sınav odalarındaki sıraları doldur ---
    for room_name in exam_rooms:
        room_info = relevant_rooms[room_name]
        room_id = room_ids[room_name]
        num_desks = room_info['desks']

        cur.execute("""
            SELECT a.desk_number, a.seat_position, p.grade_level, p.class_name
            FROM kelebek_assignments a
            JOIN kelebek_participants p ON a.participant_id = p.id
            WHERE a.room_id = %s
        """, (room_id,))
        existing_seats = {}
        for row in cur.fetchall():
            existing_seats[(row['desk_number'], row['seat_position'])] = {
                'grade_level': row['grade_level'],
                'class_name': row['class_name']
            }

        for desk_num in range(1, num_desks + 1):
            has_left = (desk_num, 'left') in existing_seats
            has_right = (desk_num, 'right') in existing_seats

            if has_left and not has_right:
                left_info = existing_seats[(desk_num, 'left')]
                right_student = pick_right_student(left_info, remaining_by_grade)
                if right_student:
                    cur.execute("""
                        INSERT INTO kelebek_assignments (plan_id, room_id, participant_id, desk_number, seat_position, row_number)
                        VALUES (%s, %s, %s, %s, 'right', %s)
                    """, (plan_id, room_id, right_student['id'], desk_num, desk_num))

            elif not has_left and not has_right:
                any_left = [g for g in remaining_by_grade if remaining_by_grade[g]]
                if not any_left:
                    continue
                any_left.sort(key=lambda g: len(remaining_by_grade[g]), reverse=True)
                left_student = remaining_by_grade[any_left[0]].pop(0)
                cur.execute("""
                    INSERT INTO kelebek_assignments (plan_id, room_id, participant_id, desk_number, seat_position, row_number)
                    VALUES (%s, %s, %s, %s, 'left', %s)
                """, (plan_id, room_id, left_student['id'], desk_num, desk_num))

                right_student = pick_right_student(left_student, remaining_by_grade)
                if right_student:
                    cur.execute("""
                        INSERT INTO kelebek_assignments (plan_id, room_id, participant_id, desk_number, seat_position, row_number)
                        VALUES (%s, %s, %s, %s, 'right', %s)
                    """, (plan_id, room_id, right_student['id'], desk_num, desk_num))

    cur.execute("UPDATE kelebek_plans SET status = 'generated', updated_at = (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul') WHERE id = %s", (plan_id,))
    conn.commit()


@kelebek_bp.route('/api/generate/<int:plan_id>', methods=['POST'])
@login_required
def generate_seating(plan_id):
    if current_user.role not in ['admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        cur.execute("SELECT COUNT(*) as cnt FROM kelebek_participants WHERE plan_id = %s", (plan_id,))
        count = cur.fetchone()['cnt']
        if count == 0:
            return jsonify({"error": "Önce katılımcıları yükleyin"}), 400

        _run_butterfly_algorithm(plan_id, conn)

        return jsonify({"success": True, "message": "Kelebek yerleşim planı oluşturuldu"})
    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"Error generating seating: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()


@kelebek_bp.route('/api/regenerate/<int:plan_id>', methods=['POST'])
@login_required
def regenerate_seating(plan_id):
    if current_user.role not in ['admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403

    conn = None
    try:
        conn = get_db()
        _run_butterfly_algorithm(plan_id, conn)
        return jsonify({"success": True, "message": "Yerleşim planı yeniden oluşturuldu (yeni karışım)"})
    except Exception as e:
        if conn:
            conn.rollback()
        logger.error(f"Error regenerating seating: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()


@kelebek_bp.route('/api/summary/<int:plan_id>')
@login_required
def get_summary(plan_id):
    if current_user.role not in ['admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        cur.execute("SELECT * FROM kelebek_plans WHERE id = %s", (plan_id,))
        plan = cur.fetchone()
        if not plan:
            return jsonify({"error": "Plan bulunamadı"}), 404

        cur.execute("""
            SELECT grade_level, is_exam_taker, COUNT(*) as cnt
            FROM kelebek_participants WHERE plan_id = %s
            GROUP BY grade_level, is_exam_taker
            ORDER BY grade_level
        """, (plan_id,))
        grade_stats = cur.fetchall()

        per_grade = {}
        total_exam = 0
        total_non_exam = 0
        for row in grade_stats:
            grade = row['grade_level']
            if grade not in per_grade:
                per_grade[grade] = {'exam_takers': 0, 'non_exam_takers': 0, 'total': 0}
            if row['is_exam_taker']:
                per_grade[grade]['exam_takers'] = row['cnt']
                total_exam += row['cnt']
            else:
                per_grade[grade]['non_exam_takers'] = row['cnt']
                total_non_exam += row['cnt']
            per_grade[grade]['total'] = per_grade[grade]['exam_takers'] + per_grade[grade]['non_exam_takers']

        cur.execute("""
            SELECT r.id, r.room_name, r.room_type, r.capacity, r.desk_count, r.grade_for_study,
                   COUNT(a.id) as assigned_count
            FROM kelebek_rooms r
            LEFT JOIN kelebek_assignments a ON a.room_id = r.id
            WHERE r.plan_id = %s
            GROUP BY r.id, r.room_name, r.room_type, r.capacity, r.desk_count, r.grade_for_study
            ORDER BY r.room_name
        """, (plan_id,))
        rooms = cur.fetchall()

        assigned_ids = set()
        cur.execute("SELECT DISTINCT participant_id FROM kelebek_assignments WHERE plan_id = %s", (plan_id,))
        for row in cur.fetchall():
            assigned_ids.add(row['participant_id'])

        cur.execute("SELECT id, student_name, class_name, is_exam_taker FROM kelebek_participants WHERE plan_id = %s", (plan_id,))
        all_parts = cur.fetchall()
        unassigned = [p for p in all_parts if p['id'] not in assigned_ids]
        unassigned_exam = [p for p in unassigned if p['is_exam_taker']]
        unassigned_study = [p for p in unassigned if not p['is_exam_taker']]

        return jsonify({
            "success": True,
            "summary": {
                "total_students": total_exam + total_non_exam,
                "exam_takers": total_exam,
                "non_exam_takers": total_non_exam,
                "per_grade": per_grade,
                "rooms": rooms,
                "unassigned_count": len(unassigned),
                "unassigned_exam_count": len(unassigned_exam),
                "unassigned_study_count": len(unassigned_study),
                "unassigned_exam_students": [{"name": u['student_name'], "class": u['class_name']} for u in unassigned_exam],
                "unassigned_study_students": [{"name": u['student_name'], "class": u['class_name']} for u in unassigned_study]
            }
        })
    except Exception as e:
        logger.error(f"Error getting summary: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()


@kelebek_bp.route('/api/room-list-pdf/<int:plan_id>')
@login_required
def download_room_list_pdf(plan_id):
    if current_user.role not in ['admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        cur.execute("SELECT * FROM kelebek_plans WHERE id = %s", (plan_id,))
        plan = cur.fetchone()
        if not plan:
            return jsonify({"error": "Plan bulunamadı"}), 404

        cur.execute("SELECT * FROM kelebek_rooms WHERE plan_id = %s ORDER BY room_type DESC, room_name", (plan_id,))
        rooms = cur.fetchall()

        cur.execute("""
            SELECT a.*, p.student_name, p.class_name, p.grade_level, p.student_no
            FROM kelebek_assignments a
            JOIN kelebek_participants p ON a.participant_id = p.id
            WHERE a.plan_id = %s
            ORDER BY a.room_id, a.desk_number, a.seat_position
        """, (plan_id,))
        all_assignments = cur.fetchall()

        assignments_by_room = {}
        for a in all_assignments:
            assignments_by_room.setdefault(a['room_id'], []).append(a)

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*cm, bottomMargin=1*cm, leftMargin=1.5*cm, rightMargin=1.5*cm)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('KelebekTitle', parent=styles['Title'], fontName=PDF_FONT_BOLD, fontSize=16, alignment=TA_CENTER, spaceAfter=6)
        subtitle_style = ParagraphStyle('KelebekSubtitle', parent=styles['Normal'], fontName=PDF_FONT, fontSize=11, alignment=TA_CENTER, spaceAfter=4)
        room_title_style = ParagraphStyle('RoomTitle', parent=styles['Heading2'], fontName=PDF_FONT_BOLD, fontSize=14, alignment=TA_CENTER, spaceAfter=8)
        cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'], fontName=PDF_FONT, fontSize=9, alignment=TA_CENTER)
        cell_style_left = ParagraphStyle('CellStyleLeft', parent=styles['Normal'], fontName=PDF_FONT, fontSize=9, alignment=TA_LEFT)
        header_cell_style = ParagraphStyle('HeaderCell', parent=styles['Normal'], fontName=PDF_FONT_BOLD, fontSize=10, alignment=TA_CENTER, textColor=colors.white)

        elements = []

        for room_idx, room in enumerate(rooms):
            room_id = room['id']
            room_assignments = assignments_by_room.get(room_id, [])
            room_type_label = 'SINAV SALONU' if room['room_type'] == 'exam' else 'DERS SALONU'

            elements.append(Paragraph(f"{room['room_name']} - {room_type_label}", room_title_style))

            exam_date_str = ''
            if plan.get('exam_date'):
                ed = plan['exam_date']
                if hasattr(ed, 'strftime'):
                    exam_date_str = ed.strftime('%d.%m.%Y')
                else:
                    exam_date_str = str(ed)

            info_text = f"{plan['plan_name']}"
            if exam_date_str:
                info_text += f" | Tarih: {exam_date_str}"
            info_text += f" | Kapasite: {room['capacity']} | Sıra: {room['desk_count']}"
            if room['room_type'] == 'study':
                if room.get('grade_for_study'):
                    info_text += f" | {room['grade_for_study']}. Sınıf Ders Çalışma"
                else:
                    info_text += " | Karma Ders Sınıfı"
            elements.append(Paragraph(info_text, subtitle_style))
            elements.append(Spacer(1, 10))

            if room['room_type'] == 'exam':
                desks = {}
                for a in room_assignments:
                    desk_num = a['desk_number']
                    if desk_num not in desks:
                        desks[desk_num] = {'left': None, 'right': None}
                    desks[desk_num][a['seat_position']] = a

                table_data = [[
                    Paragraph('Sıra', header_cell_style),
                    Paragraph('Sol Öğrenci', header_cell_style),
                    Paragraph('Sınıf', header_cell_style),
                    Paragraph('Sağ Öğrenci', header_cell_style),
                    Paragraph('Sınıf', header_cell_style),
                ]]

                for desk_num in sorted(desks.keys()):
                    desk = desks[desk_num]
                    left = desk.get('left')
                    right = desk.get('right')
                    table_data.append([
                        Paragraph(str(desk_num), cell_style),
                        Paragraph(left['student_name'] if left else '-', cell_style_left),
                        Paragraph(left['class_name'] if left else '-', cell_style),
                        Paragraph(right['student_name'] if right else '-', cell_style_left),
                        Paragraph(right['class_name'] if right else '-', cell_style),
                    ])

                col_widths = [1.2*cm, 6.5*cm, 2*cm, 6.5*cm, 2*cm]
                t = Table(table_data, colWidths=col_widths, repeatRows=1)
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTNAME', (0, 1), (-1, -1), PDF_FONT),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4f8')]),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]))
                elements.append(t)

            else:
                table_data = [[
                    Paragraph('No', header_cell_style),
                    Paragraph('Ad Soyad', header_cell_style),
                    Paragraph('Sınıf', header_cell_style),
                ]]

                for idx, a in enumerate(room_assignments, 1):
                    table_data.append([
                        Paragraph(str(idx), cell_style),
                        Paragraph(a['student_name'], cell_style_left),
                        Paragraph(a['class_name'], cell_style),
                    ])

                if not room_assignments:
                    table_data.append([
                        Paragraph('-', cell_style),
                        Paragraph('Atanmış öğrenci yok', cell_style_left),
                        Paragraph('-', cell_style),
                    ])

                col_widths = [1.5*cm, 12*cm, 3*cm]
                t = Table(table_data, colWidths=col_widths, repeatRows=1)
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2d6a4f')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('FONTNAME', (0, 1), (-1, -1), PDF_FONT),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0fdf4')]),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]))
                elements.append(t)

            if room_idx < len(rooms) - 1:
                elements.append(PageBreak())

        if not rooms:
            elements.append(Paragraph("Henüz oda ataması yapılmamış", subtitle_style))

        assigned_ids = set(a['participant_id'] for a in all_assignments)
        cur.execute("SELECT id, student_name, class_name, is_exam_taker FROM kelebek_participants WHERE plan_id = %s", (plan_id,))
        all_participants = cur.fetchall()
        unassigned_exam = [p for p in all_participants if p['id'] not in assigned_ids and p['is_exam_taker']]
        unassigned_study = [p for p in all_participants if p['id'] not in assigned_ids and not p['is_exam_taker']]

        if unassigned_exam or unassigned_study:
            elements.append(PageBreak())
            elements.append(Paragraph("YERLEŞTİRİLEMEYEN ÖĞRENCİLER", room_title_style))
            elements.append(Spacer(1, 10))

            warn_style = ParagraphStyle('WarnStyle', parent=styles['Normal'], fontName=PDF_FONT, fontSize=10, alignment=TA_LEFT, textColor=colors.HexColor('#dc2626'))

            if unassigned_exam:
                elements.append(Paragraph(f"⚠ Sınava Giren - Yerleştirilemeyen: {len(unassigned_exam)} öğrenci", warn_style))
                elements.append(Spacer(1, 6))
                u_table = [[
                    Paragraph('No', header_cell_style),
                    Paragraph('Ad Soyad', header_cell_style),
                    Paragraph('Sınıf', header_cell_style),
                ]]
                for idx, p in enumerate(sorted(unassigned_exam, key=lambda x: x['class_name']), 1):
                    u_table.append([
                        Paragraph(str(idx), cell_style),
                        Paragraph(p['student_name'], cell_style_left),
                        Paragraph(p['class_name'], cell_style),
                    ])
                col_widths = [1.5*cm, 12*cm, 3*cm]
                t = Table(u_table, colWidths=col_widths, repeatRows=1)
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc2626')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD),
                    ('FONTNAME', (0, 1), (-1, -1), PDF_FONT),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fef2f2')]),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]))
                elements.append(t)
                elements.append(Spacer(1, 16))

            if unassigned_study:
                warn_study_style = ParagraphStyle('WarnStudy', parent=styles['Normal'], fontName=PDF_FONT, fontSize=10, alignment=TA_LEFT, textColor=colors.HexColor('#92400e'))
                elements.append(Paragraph(f"⚠ Ders Çalışacak - Yerleştirilemeyen: {len(unassigned_study)} öğrenci", warn_study_style))
                elements.append(Spacer(1, 6))
                u_table = [[
                    Paragraph('No', header_cell_style),
                    Paragraph('Ad Soyad', header_cell_style),
                    Paragraph('Sınıf', header_cell_style),
                ]]
                for idx, p in enumerate(sorted(unassigned_study, key=lambda x: x['class_name']), 1):
                    u_table.append([
                        Paragraph(str(idx), cell_style),
                        Paragraph(p['student_name'], cell_style_left),
                        Paragraph(p['class_name'], cell_style),
                    ])
                col_widths = [1.5*cm, 12*cm, 3*cm]
                t = Table(u_table, colWidths=col_widths, repeatRows=1)
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#92400e')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD),
                    ('FONTNAME', (0, 1), (-1, -1), PDF_FONT),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fffbeb')]),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]))
                elements.append(t)

        doc.build(elements)
        buffer.seek(0)

        plan_name_safe = plan['plan_name'].replace(' ', '_')
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'kelebek_{plan_name_safe}.pdf'
        )
    except Exception as e:
        logger.error(f"Error generating PDF: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()


@kelebek_bp.route('/api/class-report-pdf/<int:plan_id>')
@login_required
def download_class_report_pdf(plan_id):
    if current_user.role not in ['admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        cur.execute("SELECT * FROM kelebek_plans WHERE id = %s", (plan_id,))
        plan = cur.fetchone()
        if not plan:
            return jsonify({"error": "Plan bulunamadı"}), 404

        cur.execute("SELECT * FROM kelebek_rooms WHERE plan_id = %s", (plan_id,))
        rooms = {r['id']: r for r in cur.fetchall()}

        cur.execute("""
            SELECT p.id, p.student_name, p.class_name, p.grade_level, p.student_no, p.is_exam_taker,
                   a.room_id, r.room_name, r.room_type
            FROM kelebek_participants p
            LEFT JOIN kelebek_assignments a ON a.participant_id = p.id AND a.plan_id = p.plan_id
            LEFT JOIN kelebek_rooms r ON r.id = a.room_id
            WHERE p.plan_id = %s
            ORDER BY p.class_name, p.student_name
        """, (plan_id,))
        all_participants = cur.fetchall()

        if not all_participants:
            return jsonify({"error": "Bu planda katılımcı bulunamadı"}), 404

        by_class = {}
        for p in all_participants:
            cls = p['class_name']
            by_class.setdefault(cls, []).append(p)

        exam_date_str = ''
        if plan.get('exam_date'):
            ed = plan['exam_date']
            if hasattr(ed, 'strftime'):
                exam_date_str = ed.strftime('%d.%m.%Y')
            else:
                exam_date_str = str(ed)

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                                topMargin=1.0*cm, bottomMargin=1.0*cm,
                                leftMargin=1.2*cm, rightMargin=1.2*cm)

        styles = getSampleStyleSheet()
        plan_title_style = ParagraphStyle('PlanTitle', parent=styles['Normal'],
                                          fontName=PDF_FONT_BOLD, fontSize=8,
                                          alignment=TA_CENTER, textColor=colors.HexColor('#4b5563'),
                                          spaceAfter=2)
        class_title_style = ParagraphStyle('ClassTitle', parent=styles['Title'],
                                           fontName=PDF_FONT_BOLD, fontSize=14,
                                           alignment=TA_CENTER, spaceAfter=2)
        class_subtitle_style = ParagraphStyle('ClassSub', parent=styles['Normal'],
                                              fontName=PDF_FONT, fontSize=8,
                                              alignment=TA_CENTER, textColor=colors.HexColor('#6b7280'),
                                              spaceAfter=6)
        header_style = ParagraphStyle('CRHeader', parent=styles['Normal'],
                                      fontName=PDF_FONT_BOLD, fontSize=8,
                                      alignment=TA_CENTER, textColor=colors.white)
        cell_style = ParagraphStyle('CRCell', parent=styles['Normal'],
                                    fontName=PDF_FONT, fontSize=8, alignment=TA_LEFT)
        cell_center = ParagraphStyle('CRCellC', parent=styles['Normal'],
                                     fontName=PDF_FONT, fontSize=8, alignment=TA_CENTER)
        dest_style_exam = ParagraphStyle('CRDestExam', parent=styles['Normal'],
                                         fontName=PDF_FONT_BOLD, fontSize=8,
                                         alignment=TA_CENTER, textColor=colors.HexColor('#1d4ed8'))
        dest_style_study = ParagraphStyle('CRDestStudy', parent=styles['Normal'],
                                          fontName=PDF_FONT_BOLD, fontSize=8,
                                          alignment=TA_CENTER, textColor=colors.HexColor('#92400e'))
        dest_style_none = ParagraphStyle('CRDestNone', parent=styles['Normal'],
                                         fontName=PDF_FONT, fontSize=8,
                                         alignment=TA_CENTER, textColor=colors.HexColor('#dc2626'))

        elements = []
        sorted_classes = sorted(by_class.keys())

        for cls_idx, class_name in enumerate(sorted_classes):
            students = by_class[class_name]

            plan_info = plan['plan_name']
            if exam_date_str:
                plan_info += f"  |  {exam_date_str}"
            elements.append(Paragraph(plan_info, plan_title_style))
            elements.append(Paragraph(f"{class_name} SINIFI", class_title_style))
            elements.append(Paragraph(f"Toplam {len(students)} öğrenci", class_subtitle_style))

            table_data = [[
                Paragraph('No', header_style),
                Paragraph('Ad Soyad', header_style),
                Paragraph('Gideceği Yer', header_style),
            ]]

            row_styles = []
            for idx, s in enumerate(students, 1):
                row_num = idx  # header is row 0, data starts at row 1
                if s.get('room_name'):
                    room_type = s.get('room_type', '')
                    if room_type == 'exam':
                        dest_text = f"{s['room_name']} - SINAV SALONU"
                        dest_para = Paragraph(dest_text, dest_style_exam)
                        row_styles.append(('BACKGROUND', (2, row_num), (2, row_num), colors.HexColor('#dbeafe')))
                    else:
                        dest_text = f"{s['room_name']} - DERS SALONU"
                        dest_para = Paragraph(dest_text, dest_style_study)
                        row_styles.append(('BACKGROUND', (2, row_num), (2, row_num), colors.HexColor('#fef9c3')))
                else:
                    dest_para = Paragraph("Yerleştirilmedi", dest_style_none)
                    row_styles.append(('BACKGROUND', (2, row_num), (2, row_num), colors.HexColor('#fee2e2')))

                table_data.append([
                    Paragraph(str(idx), cell_center),
                    Paragraph(s['student_name'], cell_style),
                    dest_para,
                ])

            page_w = A4[0] - 2.4*cm
            col_widths = [1.2*cm, page_w * 0.55, page_w * 0.38]
            t = Table(table_data, colWidths=col_widths, repeatRows=1)
            base_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('ALIGN', (2, 0), (2, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTNAME', (0, 1), (-1, -1), PDF_FONT),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#d1d5db')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ]
            t.setStyle(TableStyle(base_style + row_styles))
            elements.append(t)

            if cls_idx < len(sorted_classes) - 1:
                elements.append(PageBreak())

        doc.build(elements)
        buffer.seek(0)

        plan_name_safe = plan['plan_name'].replace(' ', '_')
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'sinif_raporu_{plan_name_safe}.pdf'
        )
    except Exception as e:
        logger.error(f"Error generating class report PDF: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()
