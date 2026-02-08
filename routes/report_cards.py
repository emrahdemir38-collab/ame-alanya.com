"""
Karne Yönetimi Routes - Admin PDF Karne Okuma Sistemi
"""
import os
import json
import logging
from flask import Blueprint, request, jsonify, render_template, current_app, send_file
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
import psycopg2
from psycopg2.extras import RealDictCursor
from datetime import datetime
import threading
from queue import Queue
from io import BytesIO
from utils.fmt_parser import FMTReportCardParser
from utils.image_report_parser import ImageReportParser
from utils.csv_parser import CSVExcelParser
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

def create_pdf_header(styles, default_font='DejaVuSans'):
    """PDF raporları için logo ve okul adı içeren header elementi listesi döndürür"""
    header_elements = []
    
    # Logo ekle
    logo_path = "static/images/school_logo.png"
    try:
        if os.path.exists(logo_path):
            logo = RLImage(logo_path, width=0.8*inch, height=0.8*inch)
            header_elements.append(logo)
    except Exception as e:
        logger.warning(f"Logo yüklenemedi: {e}")
    
    # Okul adı ve sistem adı - DejaVuSans font kullan (Türkçe karakter desteği)
    header_style = ParagraphStyle(
        'SchoolHeader',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.HexColor('#1e3a5f'),
        alignment=TA_CENTER,
        fontName=default_font,
        leading=18
    )
    subtitle_style = ParagraphStyle(
        'SystemHeader',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#667eea'),
        alignment=TA_CENTER,
        fontName=default_font
    )
    
    # Türkçe karakterlerle tam okul adı
    header_elements.append(Paragraph("Ayşe Melahat Erkin Ortaokulu", header_style))
    header_elements.append(Paragraph("Öğrenci Takip Sistemi", subtitle_style))
    header_elements.append(Spacer(1, 15))
    
    return header_elements

logger = logging.getLogger(__name__)

# PDF için DejaVuSans fontunu kaydet (Türkçe karakter desteği)
def register_turkish_fonts():
    """DejaVuSans fontlarını kaydet - Türkçe karakter desteği için"""
    try:
        pdfmetrics.registerFont(TTFont('DejaVuSans', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
        pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
        return 'DejaVuSans', 'DejaVuSans-Bold'
    except Exception as e:
        logger.warning(f"DejaVuSans font yüklenemedi: {e}")
        return 'Helvetica', 'Helvetica-Bold'

# Uygulama başlarken fontları kaydet
PDF_FONT, PDF_FONT_BOLD = register_turkish_fonts()

report_cards_bp = Blueprint('report_cards', __name__, url_prefix='/admin/report-cards')

def parse_outcome_code(outcome):
    """Kazanım kodunu sıralama için parse et. Örn: 7.1.3.2 -> (7, 1, 3, 2)"""
    if not outcome:
        return (999, 999, 999, 999)
    
    import re
    numbers = re.findall(r'\d+', outcome)
    if not numbers:
        return (999, 999, 999, 999)
    
    result = []
    for n in numbers[:4]:
        try:
            result.append(int(n))
        except:
            result.append(999)
    
    while len(result) < 4:
        result.append(0)
    
    return tuple(result)

def get_db():
    return psycopg2.connect(os.environ.get('DATABASE_URL'))

ALLOWED_EXTENSIONS = {'pdf'}
UPLOAD_FOLDER = 'uploads/report_cards'

# Object Storage referansı (app.py'den import edilecek)
object_storage = None

def init_object_storage(storage_client):
    """Object Storage client'ı ayarla"""
    global object_storage
    object_storage = storage_client

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ==================== PDF KUYRUK YÖNETİCİSİ ====================

class PDFQueueManager:
    """Thread-safe PDF işleme kuyruğu"""
    _instance = None
    _lock = threading.Lock()
    
    def __new__(cls):
        if cls._instance is None:
            with cls._lock:
                if cls._instance is None:
                    cls._instance = super().__new__(cls)
                    cls._instance._initialized = False
        return cls._instance
    
    def __init__(self):
        if self._initialized:
            return
        self._queue = Queue()
        self._processing = False
        self._current_job = None
        self._worker_thread = None
        self._initialized = True
        self._start_worker()
    
    def _start_worker(self):
        """Arka plan worker thread'i başlat"""
        if self._worker_thread is None or not self._worker_thread.is_alive():
            self._worker_thread = threading.Thread(target=self._worker_loop, daemon=True)
            self._worker_thread.start()
            logger.info("PDF Queue Worker started")
    
    def _worker_loop(self):
        """Sürekli çalışan worker döngüsü"""
        while True:
            try:
                job = self._queue.get()  # Blocking call
                if job is None:
                    continue
                    
                self._current_job = job
                self._processing = True
                
                # İşi çalıştır
                try:
                    job['status'] = 'processing'
                    self._update_db_status(job['report_card_id'], 'processing')
                    
                    # PDF parse işlemi
                    process_pdf_job(job)
                    
                except Exception as e:
                    logger.error(f"Queue job error: {e}")
                    self._update_db_status(job['report_card_id'], 'failed', str(e))
                finally:
                    self._processing = False
                    self._current_job = None
                    self._queue.task_done()
                    
            except Exception as e:
                logger.error(f"Worker loop error: {e}")
    
    def _update_db_status(self, report_card_id, status, error=None):
        """Veritabanında durum güncelle"""
        try:
            conn = get_db()
            cur = conn.cursor()
            if error:
                cur.execute("""
                    UPDATE report_cards SET parse_status = %s, parse_error = %s WHERE id = %s
                """, (status, error, report_card_id))
            else:
                cur.execute("""
                    UPDATE report_cards SET parse_status = %s WHERE id = %s
                """, (status, report_card_id))
            conn.commit()
            cur.close()
            conn.close()
        except Exception as e:
            logger.error(f"DB status update error: {e}")
    
    def add_job(self, job):
        """Kuyruğa iş ekle"""
        self._queue.put(job)
        return self._queue.qsize()
    
    def get_queue_status(self):
        """Kuyruk durumunu döndür"""
        return {
            'queue_size': self._queue.qsize(),
            'is_processing': self._processing,
            'current_job': self._current_job['report_card_id'] if self._current_job else None
        }

# Global queue manager instance (lazy initialization)
pdf_queue = None

def get_pdf_queue():
    global pdf_queue
    if pdf_queue is None:
        pdf_queue = PDFQueueManager()
    return pdf_queue


@report_cards_bp.route('/')
@login_required
def report_cards_page():
    """Karne yönetimi ana sayfası"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    return render_template('admin_report_cards.html')


@report_cards_bp.route('/api/classes')
@login_required
def get_classes():
    """Sistemdeki sınıfları getir - karne verilerinden"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Önce karne verilerindeki sınıfları al
        cur.execute("""
            SELECT DISTINCT class_name 
            FROM report_card_students 
            WHERE class_name IS NOT NULL
            ORDER BY class_name
        """)
        classes = [row['class_name'] for row in cur.fetchall()]
        
        # Eğer karne verisi yoksa users tablosundan al
        if not classes:
            cur.execute("""
                SELECT DISTINCT class_name 
                FROM users 
                WHERE class_name IS NOT NULL AND role = 'student'
                ORDER BY class_name
            """)
            classes = [row['class_name'] for row in cur.fetchall()]
        
        return jsonify(classes)
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/list')
@login_required
def list_report_cards():
    """Yüklenen karneleri listele - report_card_exams tablosundan"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    class_name = request.args.get('class_name')
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        if class_name:
            # Sınıf seçiliyse - report_card_results üzerinden ara
            class_normalized = class_name.replace('/', '')
            slash_class = class_name[:1] + '/' + class_name[1:] if len(class_name) >= 2 and '/' not in class_name else class_name
            
            query = """
                SELECT DISTINCT e.id, e.exam_name, e.grade_level, e.exam_date, e.created_at,
                       'completed' as parse_status,
                       (SELECT COUNT(*) FROM report_card_results r WHERE r.exam_id = e.id) as student_count
                FROM report_card_exams e
                WHERE EXISTS (
                    SELECT 1 FROM report_card_results r 
                    WHERE r.exam_id = e.id 
                    AND (r.class_name = %s OR r.class_name = %s OR REPLACE(r.class_name, '/', '') = %s)
                )
                ORDER BY e.created_at DESC
            """
            params = [class_name, slash_class, class_normalized]
        else:
            query = """
                SELECT e.id, e.exam_name, e.grade_level, e.exam_date, e.created_at,
                       'completed' as parse_status,
                       (SELECT COUNT(*) FROM report_card_results r WHERE r.exam_id = e.id) as student_count
                FROM report_card_exams e
                ORDER BY e.created_at DESC
            """
            params = []
        
        cur.execute(query, params)
        cards = cur.fetchall()
        
        # datetime nesnelerini string'e çevir
        for card in cards:
            if card.get('exam_date'):
                card['exam_date'] = card['exam_date'].isoformat() if hasattr(card['exam_date'], 'isoformat') else str(card['exam_date'])
            if card.get('created_at'):
                card['created_at'] = card['created_at'].isoformat()
        
        return jsonify(cards)
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/upload', methods=['POST'])
@login_required
def upload_report_card():
    """PDF karne yükle ve parse et - Sadece admin"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    if 'file' not in request.files:
        return jsonify({"error": "Dosya bulunamadı"}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "Dosya seçilmedi"}), 400
    
    if not allowed_file(file.filename):
        return jsonify({"error": "Sadece PDF dosyaları kabul edilir"}), 400
    
    exam_name = request.form.get('exam_name', 'Bilinmeyen Sınav')
    publisher = request.form.get('publisher', '')
    class_name = request.form.get('class_name', '')
    
    # Dosyayı kaydet (hem local hem Object Storage)
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    filename = secure_filename(f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)
    
    # Object Storage'a da kaydet
    storage_path = None
    if object_storage and object_storage.enabled:
        try:
            storage_path = f"report_cards/{filename}"
            with open(filepath, 'rb') as f:
                pdf_data = f.read()
            object_storage.upload_from_bytes(pdf_data, storage_path)
            logger.info(f"PDF Object Storage'a kaydedildi: {storage_path}")
        except Exception as e:
            logger.warning(f"Object Storage kayıt hatası: {e}")
            storage_path = None
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Kayıt oluştur
        cur.execute("""
            INSERT INTO report_cards (exam_name, publisher, class_name, pdf_filename, pdf_storage_path, uploaded_by, parse_status)
            VALUES (%s, %s, %s, %s, %s, %s, 'processing')
            RETURNING id
        """, (exam_name, publisher, class_name, filename, storage_path, current_user.id))
        
        report_card_id = cur.fetchone()['id']
        conn.commit()
        
        # Arkaplanda parse et
        thread = threading.Thread(
            target=parse_pdf_background,
            args=(report_card_id, filepath)
        )
        thread.daemon = True
        thread.start()
        
        return jsonify({
            "success": True,
            "message": "Dosya yüklendi, işleniyor...",
            "report_card_id": report_card_id
        })
        
    except Exception as e:
        conn.rollback()
        logger.error(f"Upload error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/fmt-answer-key-template', methods=['GET'])
@login_required
def download_fmt_answer_key_template():
    """Cevap anahtarı Excel şablonu indir"""
    grade = request.args.get('grade', '8')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Cevap Anahtarı"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    if grade in ['5', '6']:
        subjects = [
            ('Türkçe', 15), ('Sosyal Bilgiler', 10), ('Din Kültürü', 10),
            ('İngilizce', 10), ('Matematik', 15), ('Fen Bilimleri', 15)
        ]
        total_q = 75
    elif grade == '7':
        subjects = [
            ('Türkçe', 20), ('Sosyal Bilgiler', 10), ('Din Kültürü', 10),
            ('İngilizce', 10), ('Matematik', 20), ('Fen Bilimleri', 20)
        ]
        total_q = 90
    else:
        subjects = [
            ('Türkçe', 20), ('İnkılap Tarihi', 10), ('Din Kültürü', 10),
            ('İngilizce', 10), ('Matematik', 20), ('Fen Bilimleri', 20)
        ]
        total_q = 90
    
    ws['A1'] = f"Cevap Anahtarı Şablonu - {grade}. Sınıf (Toplam {total_q} Soru)"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:G1')
    
    ws['A3'] = "Kitapçık Türü"
    ws['B3'] = "Ders"
    ws['C3'] = "Soru Sayısı"
    ws['D3'] = "Cevaplar (A, B, C veya D harfleri)"
    
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}3'].fill = header_fill
        ws[f'{col}3'].font = header_font
        ws[f'{col}3'].border = thin_border
    
    row = 4
    for booklet in ['A', 'B']:
        for subject_name, q_count in subjects:
            ws[f'A{row}'] = booklet
            ws[f'B{row}'] = subject_name
            ws[f'C{row}'] = q_count
            ws[f'D{row}'] = ''
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].border = thin_border
            row += 1
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 50
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'cevap_anahtari_sablonu_{grade}_sinif.xlsx'
    )


def parse_excel_answer_key(file, grade: str = '8') -> tuple:
    """Excel dosyasından cevap anahtarını oku - A ve B kitapçık için (eski format)"""
    wb = load_workbook(file)
    ws = wb.active
    
    answer_key_a = {}
    answer_key_b = {}
    
    is_grade_8 = grade == '8'
    
    subject_mapping = {
        'türkçe': 'turkce',
        'sosyal bilgiler': 'sosyal',
        'din kültürü': 'din',
        'ingilizce': 'ingilizce',
        'matematik': 'matematik',
        'fen bilimleri': 'fen',
        'inkılap tarihi': 'inkilap',
        'inkilap tarihi': 'inkilap'
    }
    
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or len(row) < 4:
            continue
        
        booklet = str(row[0]).upper().strip() if row[0] else ''
        subject = str(row[1]).lower().strip() if row[1] else ''
        answers = str(row[3]).upper().strip() if row[3] else ''
        
        if 'sosyal/ink' in subject or 'sosyal/İnk' in subject.lower():
            subject_key = 'inkilap' if is_grade_8 else 'sosyal'
        else:
            subject_key = subject_mapping.get(subject)
        
        if not subject_key:
            continue
        
        answers = ''.join(c for c in answers if c in 'ABCD')
        
        if booklet == 'A':
            answer_key_a[subject_key] = answers
        elif booklet == 'B':
            answer_key_b[subject_key] = answers
    
    return answer_key_a, answer_key_b


def parse_kazanimli_excel(file) -> dict:
    """
    Kazanımlı Excel cevap anahtarını oku (Hız Yayınları formatı)
    Format: Soru No | Kazanım | A Cevabı | B Soru No
    Her ders başlığı satırda: TÜRKÇE, SOSYAL BİLGİLER, vb.
    
    Returns:
        {
            'answer_key_a': {'turkce': 'BBABCCBDAC...', ...},
            'answer_key_b': {'turkce': 'DACBBBACCB...', ...},
            'kazanimlar': {
                'turkce': [{'soru': 1, 'kazanim': 'T.S.5.1.3 ...', 'a_cevap': 'B', 'b_soru': 9}, ...]
            },
            'grade': '5',
            'total_questions': 75
        }
    """
    wb = load_workbook(file)
    ws = wb.active
    
    subject_mapping_upper = {
        'TÜRKÇE': 'turkce',
        'SOSYAL BİLGİLER': 'sosyal',
        'DİN KÜLTÜRÜ VE AHLAK BİLGİSİ': 'din',
        'DİN KÜLTÜRÜ': 'din',
        'İNGİLİZCE': 'ingilizce',
        'INGILIZCE': 'ingilizce',
        'MATEMATİK': 'matematik',
        'FEN BİLİMLERİ': 'fen',
        'FEN BİLİMLERİ': 'fen',
        'İNKILAP TARİHİ VE ATATÜRKÇÜLÜK': 'inkilap',
        'T.C. İNKILAP TARİHİ VE ATATÜRKÇÜLÜK': 'inkilap',
        'İNKILAP TARİHİ': 'inkilap'
    }
    
    result = {
        'answer_key_a': {},
        'answer_key_b': {},
        'kazanimlar': {},
        'grade': None,
        'total_questions': 0
    }
    
    current_subject = None
    current_subject_key = None
    a_answers = []
    b_mapping = []
    kazanimlar = []
    
    sheet_title = ws.title.lower() if ws.title else ''
    for g in ['5', '6', '7', '8']:
        if g in sheet_title:
            result['grade'] = g
            break
    
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or row[0] is None:
            continue
        
        first_cell = str(row[0]).strip()
        first_cell_upper = first_cell.upper()
        
        subject_headers = [
            'TÜRKÇE', 'SOSYAL BİLGİLER', 'DİN KÜLTÜRÜ VE AHLAK BİLGİSİ', 
            'İNGİLİZCE', 'INGILIZCE', 'MATEMATİK', 'FEN BİLİMLERİ',
            'İNKILAP TARİHİ VE ATATÜRKÇÜLÜK', 'T.C. İNKILAP TARİHİ VE ATATÜRKÇÜLÜK',
            'INKILAP TARİHİ VE ATATÜRKÇÜLÜK'
        ]
        
        is_subject_header = first_cell_upper in subject_headers or 'İNGİLİZCE' in first_cell_upper or 'INGILIZCE' in first_cell_upper
        
        if is_subject_header:
            if current_subject_key and a_answers:
                result['answer_key_a'][current_subject_key] = ''.join(a_answers)
                
                b_sorted = sorted(b_mapping, key=lambda x: x[1])
                b_answers = [item[0] for item in b_sorted]
                result['answer_key_b'][current_subject_key] = ''.join(b_answers)
                
                result['kazanimlar'][current_subject_key] = kazanimlar
                result['total_questions'] += len(a_answers)
            
            current_subject = first_cell_upper
            current_subject_key = subject_mapping_upper.get(first_cell_upper)
            if not current_subject_key and ('İNGİLİZCE' in first_cell_upper or 'INGILIZCE' in first_cell_upper):
                current_subject_key = 'ingilizce'
            if not current_subject_key and 'İNKILAP' in first_cell_upper:
                current_subject_key = 'inkilap'
            a_answers = []
            b_mapping = []
            kazanimlar = []
            continue
        
        if current_subject_key and len(row) >= 4:
            try:
                soru_no = int(row[0]) if row[0] else None
                kazanim = str(row[1]).strip() if row[1] else ''
                a_cevap = str(row[2]).upper().strip() if row[2] else ''
                b_soru = int(row[3]) if row[3] else None
                
                if soru_no and a_cevap in ['A', 'B', 'C', 'D'] and b_soru:
                    a_answers.append(a_cevap)
                    b_mapping.append((a_cevap, b_soru))
                    kazanimlar.append({
                        'soru': soru_no,
                        'kazanim': kazanim,
                        'a_cevap': a_cevap,
                        'b_soru': b_soru
                    })
            except (ValueError, TypeError):
                continue
    
    if current_subject_key and a_answers:
        result['answer_key_a'][current_subject_key] = ''.join(a_answers)
        b_sorted = sorted(b_mapping, key=lambda x: x[1])
        b_answers = [item[0] for item in b_sorted]
        result['answer_key_b'][current_subject_key] = ''.join(b_answers)
        result['kazanimlar'][current_subject_key] = kazanimlar
        result['total_questions'] += len(a_answers)
    
    return result


@report_cards_bp.route('/api/upload-fmt-answer-key', methods=['POST'])
@login_required
def upload_fmt_answer_key():
    """Kazanımlı Excel cevap anahtarı yükle - FMT CA"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    if 'file' not in request.files:
        return jsonify({"error": "Excel dosyası bulunamadı"}), 400
    
    file = request.files['file']
    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"error": "Geçerli bir Excel dosyası seçin (.xlsx veya .xls)"}), 400
    
    exam_name = request.form.get('exam_name', 'FMT Sınav')
    publisher = request.form.get('publisher', '')
    
    try:
        parsed = parse_kazanimli_excel(file)
        
        if not parsed['answer_key_a']:
            return jsonify({"error": "Excel dosyasından cevap anahtarı okunamadı. Format doğru mu?"}), 400
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            INSERT INTO fmt_answer_keys (exam_name, publisher, grade, answer_key_a, answer_key_b, kazanimlar, total_questions, created_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (exam_name) DO UPDATE SET
                publisher = EXCLUDED.publisher,
                grade = EXCLUDED.grade,
                answer_key_a = EXCLUDED.answer_key_a,
                answer_key_b = EXCLUDED.answer_key_b,
                kazanimlar = EXCLUDED.kazanimlar,
                total_questions = EXCLUDED.total_questions,
                updated_at = CURRENT_TIMESTAMP
            RETURNING id
        """, (
            exam_name,
            publisher,
            parsed['grade'],
            psycopg2.extras.Json(parsed['answer_key_a']),
            psycopg2.extras.Json(parsed['answer_key_b']),
            psycopg2.extras.Json(parsed['kazanimlar']),
            parsed['total_questions'],
            current_user.id
        ))
        
        key_id = cur.fetchone()['id']
        conn.commit()
        
        cur.close()
        conn.close()
        
        subject_names = {
            'turkce': 'Türkçe', 'sosyal': 'Sosyal Bilgiler', 'din': 'Din Kültürü',
            'ingilizce': 'İngilizce', 'matematik': 'Matematik', 'fen': 'Fen Bilimleri',
            'inkilap': 'İnkılap Tarihi'
        }
        
        subjects_info = []
        for subj, answers in parsed['answer_key_a'].items():
            subjects_info.append(f"{subject_names.get(subj, subj)}: {len(answers)} soru")
        
        return jsonify({
            "success": True,
            "message": f"Cevap anahtarı başarıyla yüklendi!",
            "id": key_id,
            "exam_name": exam_name,
            "grade": parsed['grade'],
            "total_questions": parsed['total_questions'],
            "subjects": subjects_info,
            "has_kazanimlar": bool(parsed['kazanimlar'])
        })
        
    except Exception as e:
        logger.error(f"FMT CA upload error: {e}")
        return jsonify({"error": str(e)}), 500


@report_cards_bp.route('/api/fmt-answer-keys', methods=['GET'])
@login_required
def get_fmt_answer_keys():
    """Kayıtlı FMT cevap anahtarlarını listele"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("""
            SELECT id, exam_name, publisher, grade, total_questions, 
                   created_at, updated_at
            FROM fmt_answer_keys
            ORDER BY created_at DESC
        """)
        keys = cur.fetchall()
        
        return jsonify({"success": True, "answer_keys": keys})
    except Exception as e:
        logger.error(f"FMT answer keys list error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/upload-fmt', methods=['POST'])
@login_required
def upload_fmt_file():
    """FMT dosyası yükle ve parse et - Sadece admin"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    if 'files[]' not in request.files and 'file' not in request.files:
        return jsonify({"error": "Dosya bulunamadı"}), 400
    
    files = request.files.getlist('files[]') or [request.files.get('file')]
    if not files or all(f.filename == '' for f in files):
        return jsonify({"error": "Dosya seçilmedi"}), 400
    
    exam_name = request.form.get('exam_name', 'FMT Sınav')
    publisher = request.form.get('publisher', '')
    grade = str(request.form.get('grade', '8'))
    
    answer_key_a = {}
    answer_key_b = {}
    
    excel_file = request.files.get('answer_key_excel')
    kazanimlar = {}
    if excel_file and excel_file.filename:
        try:
            # Önce basit formatı dene
            answer_key_a, answer_key_b = parse_excel_answer_key(excel_file, grade)
            logger.info(f"Basit Excel formatı denendi: A={len(answer_key_a)} ders, B={len(answer_key_b)} ders")
            
            # Eğer boşsa, kazanımlı formatı dene
            if not answer_key_a or len(answer_key_a) == 0:
                excel_file.seek(0)  # Dosyayı başa sar
                kazanimli_result = parse_kazanimli_excel(excel_file)
                if kazanimli_result and kazanimli_result.get('answer_key_a'):
                    answer_key_a = kazanimli_result['answer_key_a']
                    answer_key_b = kazanimli_result.get('answer_key_b', {})
                    kazanimlar = kazanimli_result.get('kazanimlar', {})
                    grade = kazanimli_result.get('grade', grade)
                    logger.info(f"Kazanımlı Excel formatı okundu: A={len(answer_key_a)} ders, sınıf={grade}")
        except Exception as e:
            logger.error(f"Excel okuma hatası: {e}")
            import traceback
            logger.error(traceback.format_exc())
    
    if not answer_key_a:
        answer_key_a = {
            'turkce': request.form.get('answer_turkce_a', request.form.get('answer_turkce', '')),
            'ingilizce': request.form.get('answer_ingilizce_a', request.form.get('answer_ingilizce', '')),
            'sosyal': request.form.get('answer_sosyal_a', request.form.get('answer_sosyal', '')),
            'inkilap': request.form.get('answer_inkilap_a', request.form.get('answer_inkilap', '')),
            'matematik': request.form.get('answer_matematik_a', request.form.get('answer_matematik', '')),
            'din': request.form.get('answer_din_a', request.form.get('answer_din', '')),
            'fen': request.form.get('answer_fen_a', request.form.get('answer_fen', ''))
        }
        answer_key_b = {
            'turkce': request.form.get('answer_turkce_b', ''),
            'ingilizce': request.form.get('answer_ingilizce_b', ''),
            'sosyal': request.form.get('answer_sosyal_b', ''),
            'inkilap': request.form.get('answer_inkilap_b', ''),
            'matematik': request.form.get('answer_matematik_b', ''),
            'din': request.form.get('answer_din_b', ''),
            'fen': request.form.get('answer_fen_b', '')
        }
    
    has_answer_key_a = any(v.strip() for v in answer_key_a.values())
    has_answer_key_b = any(v.strip() for v in answer_key_b.values())
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("""
            INSERT INTO report_cards (exam_name, publisher, class_name, pdf_filename, uploaded_by, parse_status)
            VALUES (%s, %s, %s, %s, %s, 'processing')
            RETURNING id
        """, (exam_name, publisher, 'FMT', 'fmt_upload.txt', current_user.id))
        
        report_card_id = cur.fetchone()['id']
        conn.commit()
        
        # Cevap anahtarını veritabanına kaydet (eğer varsa)
        if has_answer_key_a:
            try:
                import json
                cur.execute("""
                    INSERT INTO fmt_answer_keys (exam_name, publisher, grade, answer_key_a, answer_key_b, kazanimlar, total_questions, created_by)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (exam_name) DO UPDATE SET
                        publisher = EXCLUDED.publisher,
                        grade = EXCLUDED.grade,
                        answer_key_a = EXCLUDED.answer_key_a,
                        answer_key_b = EXCLUDED.answer_key_b,
                        kazanimlar = EXCLUDED.kazanimlar,
                        total_questions = EXCLUDED.total_questions,
                        updated_at = CURRENT_TIMESTAMP
                """, (
                    exam_name,
                    publisher,
                    grade,
                    json.dumps(answer_key_a, ensure_ascii=False),
                    json.dumps(answer_key_b, ensure_ascii=False) if answer_key_b else '{}',
                    json.dumps(kazanimlar, ensure_ascii=False) if kazanimlar else '{}',
                    75,
                    current_user.id
                ))
                conn.commit()
                logger.info(f"Cevap anahtarı veritabanına kaydedildi: {exam_name}, sınıf={grade}")
            except Exception as e:
                logger.error(f"Cevap anahtarı kaydetme hatası: {e}")
        
        all_students = []
        for file in files:
            if file and file.filename and not file.filename.endswith('.xlsx'):
                content = file.read()
                parser = FMTReportCardParser(
                    file_content=content,
                    answer_key=answer_key_a if has_answer_key_a else None,
                    answer_key_b=answer_key_b if has_answer_key_b else None
                )
                students = parser.parse()
                all_students.extend(students)
        
        thread = threading.Thread(
            target=process_fmt_students,
            args=(report_card_id, all_students, exam_name)
        )
        thread.daemon = False
        thread.start()
        
        return jsonify({
            "success": True,
            "message": f"FMT dosyaları yüklendi, {len(all_students)} öğrenci işleniyor...",
            "report_card_id": report_card_id,
            "student_count": len(all_students),
            "has_answer_key_a": has_answer_key_a,
            "has_answer_key_b": has_answer_key_b
        })
        
    except Exception as e:
        conn.rollback()
        logger.error(f"FMT upload error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


def detect_booklet_type(student_answers: dict, answer_key_a: dict, answer_key_b: dict) -> str:
    """
    Öğrenci cevaplarını her iki kitapçık cevap anahtarıyla karşılaştırarak
    en uygun kitapçık türünü otomatik tespit et.
    
    Returns: 'A' veya 'B'
    """
    score_a = 0
    score_b = 0
    
    for subj, answers in student_answers.items():
        if not answers:
            continue
        
        key_a = answer_key_a.get(subj, '') or ''
        key_b = answer_key_b.get(subj, '') or ''
        
        for i, char in enumerate(answers):
            if not char or char in [' ', '-', '*']:
                continue
            student_ans = char.upper()
            
            if i < len(key_a) and key_a[i].upper() == student_ans:
                score_a += 1
            if i < len(key_b) and key_b[i].upper() == student_ans:
                score_b += 1
    
    return 'B' if score_b > score_a else 'A'


def process_fmt_students(report_card_id: int, students: list, exam_name: str):
    """FMT öğrencilerini veritabanına kaydet ve cevap anahtarı ile eşleştir"""
    import psycopg2
    from psycopg2.extras import RealDictCursor
    import traceback
    
    logger.info(f"FMT grading thread başladı: {len(students)} öğrenci, report_card_id={report_card_id}")
    
    db_url = os.environ.get('DATABASE_URL')
    if not db_url:
        logger.error("DATABASE_URL bulunamadı!")
        return
        
    try:
        conn = psycopg2.connect(db_url)
        cur = conn.cursor(cursor_factory=RealDictCursor)
    except Exception as e:
        logger.error(f"Veritabanı bağlantı hatası: {e}")
        return
    
    try:
        saved_count = 0
        skipped_count = 0
        
        # Sınıf seviyesini belirle (ilk öğrenciden)
        grade = 5
        if students:
            grade = students[0].get('grade', 5)
        
        # Cevap anahtarını bul
        cur.execute("""
            SELECT id, answer_key_a, answer_key_b, kazanimlar 
            FROM fmt_answer_keys 
            WHERE grade = %s ORDER BY created_at DESC LIMIT 1
        """, (str(grade),))
        answer_key_row = cur.fetchone()
        
        answer_key_a = {}
        answer_key_b = {}
        kazanimlar = {}
        
        if answer_key_row:
            answer_key_a = answer_key_row.get('answer_key_a', {}) or {}
            answer_key_b = answer_key_row.get('answer_key_b', {}) or {}
            kazanimlar = answer_key_row.get('kazanimlar', {}) or {}
            logger.info(f"Cevap anahtarı bulundu: {grade}. sınıf")
        else:
            logger.warning(f"Cevap anahtarı bulunamadı: {grade}. sınıf")
        
        logger.info(f"Öğrenci işleme başlıyor: {len(students)} öğrenci")
        for idx, student in enumerate(students):
            try:
                if idx % 20 == 0:
                    logger.info(f"İşlenen: {idx}/{len(students)}")
                student_no = student.get('student_no', '')
                student_name = student.get('student_name', '')
                class_name = student.get('class_name', '')
                booklet_type = student.get('booklet_type', 'A')
                
                if not student_no:
                    student_no = f"FMT_{student_name[:10]}_{saved_count}"
                
                # Kullanıcı ID'sini, isim ve şube bilgisini veritabanından al
                # ÖNEMLİ: Önce student_no + sınıf ile eşleştir (aynı numara farklı sınıflarda olabilir)
                if idx == 0:
                    logger.info(f"İlk öğrenci: {student_no}, {student_name}, sınıf: {class_name}")
                
                user_row = None
                # CSV'den gelen sınıf bilgisi varsa önce sınıf + numara ile ara
                if class_name and student_no:
                    # Sınıf formatını normalize et: "7B" -> "7/B" veya "7/B" -> "7/B"
                    csv_class_normalized = class_name.replace('/', '')  # "7/B" -> "7B"
                    cur.execute("""
                        SELECT id, full_name, class_name FROM users 
                        WHERE student_no = %s 
                        AND (REPLACE(class_name, '/', '') = %s OR class_name = %s)
                        LIMIT 1
                    """, (student_no, csv_class_normalized, class_name))
                    user_row = cur.fetchone()
                
                # Sınıf eşleşmesi bulunamazsa sadece student_no ile ara
                if not user_row and student_no:
                    cur.execute("SELECT id, full_name, class_name FROM users WHERE student_no = %s LIMIT 1", (student_no,))
                    user_row = cur.fetchone()
                
                user_id = user_row['id'] if user_row else None
                if idx == 0:
                    logger.info(f"DB sorgu tamamlandı: user_id={user_id}")
                
                # Veritabanından TAM bilgileri al (FMT'deki eksik/hatalı bilgileri düzelt)
                if user_row:
                    # İsim veritabanından al (daha doğru)
                    db_name = user_row.get('full_name', '')
                    if db_name and len(db_name) > len(student_name):
                        student_name = db_name
                    
                    # Sınıf bilgisini veritabanından al
                    db_class = user_row.get('class_name', '')
                    if db_class:
                        # Sınıf formatını düzelt - sadece "/" yoksa dönüştür (5A -> 5/A)
                        if '/' not in db_class and len(db_class) == 2 and db_class[0].isdigit() and db_class[1].isalpha():
                            class_name = f"{db_class[0]}/{db_class[1]}"
                        else:
                            class_name = db_class
                        logger.debug(f"Veritabanından bilgi alındı: {student_no} -> {student_name}, {class_name}")
                
                # Ham cevapları hazırla
                raw_answers = {}
                subjects_data = student.get('subjects', {})
                if idx == 0:
                    logger.info(f"Subjects: {list(subjects_data.keys())}")
                for subj, subj_data in subjects_data.items():
                    # FMT parser 'student_answers' kullanıyor
                    raw_answers[subj] = subj_data.get('student_answers', '') or subj_data.get('answers', '')
                if idx == 0:
                    logger.info(f"Raw answers hazırlandı")
                
                # Kitapçık türü belirsizse veya 'A' default ise otomatik tespit et
                if booklet_type not in ['A', 'B'] or student.get('booklet_type') is None:
                    detected_type = detect_booklet_type(raw_answers, answer_key_a, answer_key_b)
                    if idx == 0:
                        logger.info(f"Kitapçık otomatik tespit: {detected_type}")
                    booklet_type = detected_type
                else:
                    # Explicit olarak A/B belirtilmiş olsa bile doğruluk kontrolü yap
                    detected_type = detect_booklet_type(raw_answers, answer_key_a, answer_key_b)
                    if detected_type != booklet_type:
                        logger.debug(f"Öğrenci {student_no}: Belirtilen {booklet_type}, tespit edilen {detected_type} - tespit edilen kullanılıyor")
                        booklet_type = detected_type
                
                # Cevap anahtarı ile eşleştirme yap
                total_correct = 0
                total_wrong = 0
                total_empty = 0
                subject_results = {}
                all_answers = []  # Hata karnesi için
                
                for subj, subj_data in subjects_data.items():
                    # FMT parser 'student_answers' kullanıyor
                    student_answers = subj_data.get('student_answers', '') or subj_data.get('answers', '')
                    subj_correct = 0
                    subj_wrong = 0
                    subj_empty = 0
                    
                    # Cevap anahtarı string'ini al (kitapçık tipine göre)
                    if booklet_type == 'A':
                        answer_string = answer_key_a.get(subj, '') or ''
                    else:
                        answer_string = answer_key_b.get(subj, '') or ''
                    
                    # Kazanım listesini al
                    subj_kazanimlar = kazanimlar.get(subj, []) or []
                    
                    for i, char in enumerate(student_answers):
                        q_num = i + 1
                        # _ karakteri boşluk anlamına geliyor (parse sırasında boşluklar _ ile değiştirildi)
                        if char == '_' or not char or not char.strip():
                            student_ans = ''
                        else:
                            student_ans = char.upper()
                        
                        # Cevap anahtarından doğru cevabı bul
                        correct_ans = ''
                        kazanim = ''
                        
                        # A kitapçığı için doğrudan index, B için mapping kullan
                        if booklet_type == 'A':
                            if i < len(answer_string):
                                correct_ans = answer_string[i].upper()
                            # Kazanım ara
                            for kz in subj_kazanimlar:
                                if kz.get('soru') == q_num:
                                    kazanim = kz.get('kazanim', '')
                                    break
                        else:
                            # B kitapçığı - b_soru ile eşle
                            for kz in subj_kazanimlar:
                                if kz.get('b_soru') == q_num:
                                    correct_ans = kz.get('a_cevap', '').upper()
                                    kazanim = kz.get('kazanim', '')
                                    break
                        
                        # Değerlendirme - _ karakteri de boşluk olarak kabul edilir
                        is_blank = (student_ans == '' or student_ans == ' ' or student_ans == '*' or student_ans == '-' or student_ans == '_')
                        is_correct = False
                        
                        if is_blank:
                            subj_empty += 1
                        elif correct_ans and student_ans == correct_ans:
                            is_correct = True
                            subj_correct += 1
                        elif correct_ans:
                            subj_wrong += 1
                        else:
                            # Cevap anahtarı yoksa boş say
                            subj_empty += 1
                        
                        # Hata karnesi için cevap detayı
                        all_answers.append({
                            'subject': subj,
                            'question_number': q_num,
                            'correct_answer': correct_ans,
                            'student_answer': student_ans,
                            'is_correct': is_correct,
                            'is_blank': is_blank,
                            'outcome_code': kazanim,
                            'outcome_text': kazanim
                        })
                    
                    subj_net = subj_correct - (subj_wrong / 4)
                    subject_results[subj] = {
                        'correct': subj_correct,
                        'wrong': subj_wrong,
                        'empty': subj_empty,
                        'net': round(subj_net, 2),
                        'question_count': len(student_answers)
                    }
                    
                    total_correct += subj_correct
                    total_wrong += subj_wrong
                    total_empty += subj_empty
                
                total_net = total_correct - (total_wrong / 4)
                total_questions = total_correct + total_wrong + total_empty
                success_rate = round((total_correct / total_questions * 100) if total_questions > 0 else 0, 2)
                
                results_json = {
                    'student_name': student_name,
                    'class_name': class_name,
                    'student_no': student_no,
                    'grade': grade,
                    'total_questions': total_questions,
                    'success_rate': success_rate,
                    'subjects': subject_results
                }
                
                # optical_student_results tablosuna kaydet
                cur.execute("""
                    INSERT INTO optical_student_results (
                        optical_exam_id, student_id, booklet_type, raw_answers, results,
                        total_correct, total_wrong, total_empty, total_net, total_score
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                """, (
                    report_card_id,
                    user_id,
                    booklet_type,
                    json.dumps(raw_answers),
                    json.dumps(results_json),
                    total_correct,
                    total_wrong,
                    total_empty,
                    round(total_net, 2),
                    0
                ))
                optical_result_id = cur.fetchone()['id']
                
                # report_card_students tablosuna da kaydet (hata karnesi için)
                cur.execute("""
                    INSERT INTO report_card_students (
                        report_card_id, user_id, student_name, student_no, class_name,
                        total_questions, total_correct, total_wrong, total_blank, total_net, success_rate
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                """, (
                    report_card_id,
                    user_id,
                    student_name,
                    student_no,
                    class_name,
                    total_questions,
                    total_correct,
                    total_wrong,
                    total_empty,
                    round(total_net, 2),
                    success_rate
                ))
                student_result_id = cur.fetchone()['id']
                
                # report_card_answers tablosuna kaydet (hata karnesi için) - batch INSERT
                if all_answers:
                    from psycopg2.extras import execute_values
                    answer_data = [(
                        student_result_id,
                        ans['subject'],
                        ans['question_number'],
                        ans['correct_answer'],
                        ans['student_answer'],
                        ans['is_correct'],
                        ans['is_blank'],
                        ans['outcome_code'],
                        ans['outcome_text']
                    ) for ans in all_answers]
                    execute_values(cur, """
                        INSERT INTO report_card_answers (
                            student_result_id, subject, question_number,
                            correct_answer, student_answer, is_correct, is_blank,
                            outcome_code, outcome_text
                        ) VALUES %s
                    """, answer_data)
                
                saved_count += 1
                conn.commit()
                
            except Exception as e:
                logger.error(f"FMT öğrenci kayıt hatası: {e}")
                conn.rollback()
                continue
        
        cur.execute("""
            UPDATE report_cards 
            SET parse_status = 'completed', student_count = %s
            WHERE id = %s
        """, (saved_count, report_card_id))
        conn.commit()
        
        logger.info(f"FMT işleme tamamlandı: {saved_count} kaydedildi, {skipped_count} atlandı")
        
    except Exception as e:
        logger.error(f"FMT process error: {e}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        try:
            cur.execute("""
                UPDATE report_cards SET parse_status = 'failed' WHERE id = %s
            """, (report_card_id,))
            conn.commit()
        except:
            pass
    finally:
        try:
            cur.close()
            conn.close()
        except:
            pass


@report_cards_bp.route('/api/reprocess/<int:report_card_id>', methods=['POST'])
@login_required
def reprocess_report_card(report_card_id):
    """Tek bir PDF'i yeniden işle - Sadece admin"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # PDF bilgilerini al
        cur.execute("""
            SELECT id, exam_name, pdf_filename, pdf_storage_path 
            FROM report_cards WHERE id = %s
        """, (report_card_id,))
        card = cur.fetchone()
        
        if not card:
            return jsonify({"error": "Sınav bulunamadı"}), 404
        
        # PDF dosyasını bul
        filepath = None
        temp_file = None
        
        # Önce local dosyayı kontrol et
        local_path = os.path.join(UPLOAD_FOLDER, card['pdf_filename']) if card['pdf_filename'] else None
        if local_path and os.path.exists(local_path):
            filepath = local_path
        # Object Storage'dan indir
        elif card['pdf_storage_path'] and object_storage and object_storage.enabled:
            try:
                import tempfile
                pdf_data, _ = object_storage.download_as_bytes(card['pdf_storage_path'])
                temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
                temp_file.write(pdf_data)
                temp_file.close()
                filepath = temp_file.name
                logger.info(f"PDF Object Storage'dan indirildi: {card['pdf_storage_path']}")
            except Exception as e:
                logger.error(f"Object Storage'dan indirme hatası: {e}")
                return jsonify({"error": f"PDF dosyası indirilemedi: {e}"}), 500
        
        if not filepath:
            return jsonify({"error": "PDF dosyası bulunamadı. Lütfen yeniden yükleyin."}), 404
        
        # Eski cevapları sil
        cur.execute("""
            DELETE FROM report_card_answers 
            WHERE student_result_id IN (
                SELECT id FROM report_card_students WHERE report_card_id = %s
            )
        """, (report_card_id,))
        
        # Eski öğrenci kayıtlarını sil
        cur.execute("DELETE FROM report_card_students WHERE report_card_id = %s", (report_card_id,))
        
        # Durumu güncelle
        cur.execute("""
            UPDATE report_cards 
            SET parse_status = 'processing', parse_error = NULL, 
                processed_student_count = 0, expected_student_count = 0
            WHERE id = %s
        """, (report_card_id,))
        conn.commit()
        
        # Thread-based parse with Flask app context
        import threading
        from flask import current_app
        app = current_app._get_current_object()
        
        def run_parse_in_thread(flask_app, rc_id, fpath, temp_path, db_url):
            """Thread içinde parse işlemi - doğrudan psycopg2 bağlantısı ile"""
            import traceback
            import psycopg2
            from psycopg2.extras import RealDictCursor
            
            conn2 = None
            cur2 = None
            try:
                # Thread içinde yeni veritabanı bağlantısı aç
                conn2 = psycopg2.connect(db_url)
                cur2 = conn2.cursor(cursor_factory=RealDictCursor)
                
                cur2.execute("""
                    UPDATE report_cards SET parse_error = 'Thread başladı...'
                    WHERE id = %s
                """, (rc_id,))
                conn2.commit()
                
                from utils.pdf_report_parser import parse_report_card_pdf
                
                cur2.execute("""
                    UPDATE report_cards SET parse_error = 'PDF parse ediliyor...'
                    WHERE id = %s
                """, (rc_id,))
                conn2.commit()
                
                students, error = parse_report_card_pdf(fpath)
                
                if error:
                    cur2.execute("""
                        UPDATE report_cards 
                        SET parse_status = 'failed', parse_error = %s
                        WHERE id = %s
                    """, (f"Parse hatası: {error}", rc_id))
                    conn2.commit()
                    return
                
                cur2.execute("""
                    UPDATE report_cards 
                    SET expected_student_count = %s, processed_student_count = 0,
                        parse_error = %s
                    WHERE id = %s
                """, (len(students), f"{len(students)} öğrenci bulundu, kaydediliyor...", rc_id))
                conn2.commit()
                
                cur2.execute("SELECT class_name, exam_name FROM report_cards WHERE id = %s", (rc_id,))
                report_info = cur2.fetchone()
                report_class = report_info['class_name']
                
                is_grade_level_upload = report_class and report_class.startswith('ALL_')
                grade_level = report_class.split('_')[1] if is_grade_level_upload else None
                
                inserted_count = 0
                updated_count = 0
                
                for idx, student_data in enumerate(students):
                    try:
                        student_name = student_data.get('student_name', '')
                        pdf_class = student_data.get('class_name')
                        student_no = student_data.get('student_no')
                        
                        # İlk öğrenci için debug log
                        if idx == 0:
                            cur2.execute("""
                                UPDATE report_cards SET parse_error = %s WHERE id = %s
                            """, (f"İlk öğrenci işleniyor: {student_name[:30]}...", rc_id))
                            conn2.commit()
                        
                        target_class = pdf_class
                        if not target_class and is_grade_level_upload:
                            target_class = f"{grade_level}/?"
                        elif not target_class:
                            target_class = report_class
                        
                        existing_student = None
                        if student_no:
                            cur2.execute("""
                                SELECT id FROM report_card_students 
                                WHERE report_card_id = %s AND student_no = %s
                            """, (rc_id, student_no))
                            existing_student = cur2.fetchone()
                        
                        if existing_student:
                            student_result_id = existing_student['id']
                            cur2.execute("DELETE FROM report_card_answers WHERE student_result_id = %s", (student_result_id,))
                            updated_count += 1
                        else:
                            cur2.execute("""
                                INSERT INTO report_card_students 
                                (report_card_id, student_name, class_name, student_no, grade,
                                 lgs_score, percentile, total_questions, total_correct, 
                                 total_wrong, total_blank, total_net, success_rate)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                                RETURNING id
                            """, (
                                rc_id, student_name, target_class, student_no,
                                student_data.get('grade', 8),
                                student_data.get('lgs_score'),
                                student_data.get('percentile'),
                                student_data.get('total_questions'),
                                student_data.get('total_correct'),
                                student_data.get('total_wrong'),
                                student_data.get('total_blank'),
                                student_data.get('total_net'),
                                student_data.get('success_rate')
                            ))
                            student_result_id = cur2.fetchone()['id']
                            inserted_count += 1
                        
                        for answer in student_data.get('answers', []):
                            cur2.execute("""
                                INSERT INTO report_card_answers
                                (student_result_id, subject, question_number, correct_answer,
                                 student_answer, is_correct, is_blank, outcome_code, outcome_description)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                            """, (
                                student_result_id, answer.get('subject'), answer.get('question_number'),
                                answer.get('correct_answer'), answer.get('student_answer', ''),
                                answer.get('is_correct', False), answer.get('is_blank', False),
                                answer.get('outcome_code'), answer.get('outcome_description')
                            ))
                        
                        conn2.commit()
                        
                        if (idx + 1) % 10 == 0:
                            cur2.execute("""
                                UPDATE report_cards SET processed_student_count = %s,
                                    parse_error = %s WHERE id = %s
                            """, (idx + 1, f"{idx + 1}/{len(students)} öğrenci işlendi", rc_id))
                            conn2.commit()
                    except Exception as student_err:
                        logger.error(f"Öğrenci kayıt hatası: {student_err}")
                        try:
                            cur2.execute("""
                                UPDATE report_cards SET parse_error = %s WHERE id = %s
                            """, (f"Öğrenci {idx+1} hatası: {str(student_err)[:200]}", rc_id))
                            conn2.commit()
                        except:
                            pass
                        continue
                
                cur2.execute("""
                    UPDATE report_cards 
                    SET parse_status = 'completed', 
                        student_count = %s,
                        processed_student_count = %s,
                        parse_error = %s
                    WHERE id = %s
                """, (inserted_count + updated_count, inserted_count + updated_count, 
                      f"{inserted_count} yeni, {updated_count} güncellendi", rc_id))
                conn2.commit()
                
            except Exception as e:
                logger.error(f"Thread parse hatası: {e}")
                traceback.print_exc()
                if conn2 and cur2:
                    try:
                        cur2.execute("""
                            UPDATE report_cards 
                            SET parse_status = 'failed', parse_error = %s
                            WHERE id = %s
                        """, (f"Thread hatası: {str(e)[:300]}", rc_id))
                        conn2.commit()
                    except:
                        pass
            finally:
                if temp_path:
                    try:
                        import os
                        os.unlink(temp_path)
                    except:
                        pass
                if cur2:
                    cur2.close()
                if conn2:
                    conn2.close()
        
        # Thread başlat - db_url'i şimdi al
        temp_path = temp_file.name if temp_file else None
        db_url = os.environ.get('DATABASE_URL')
        thread = threading.Thread(target=run_parse_in_thread, args=(app, report_card_id, filepath, temp_path, db_url))
        thread.daemon = False  # daemon=False ensures thread completes
        thread.start()
        
        return jsonify({
            "success": True,
            "message": f"'{card['exam_name']}' işleniyor... Sayfayı yenileyerek ilerlemeyi takip edin."
        })
        
    except Exception as e:
        conn.rollback()
        logger.error(f"Reprocess error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/migrate-local-pdfs', methods=['POST'])
@login_required
def migrate_local_pdfs():
    """Yerel PDF dosyalarını Object Storage'a yükle ve veritabanına ekle"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    if not os.path.exists(UPLOAD_FOLDER):
        return jsonify({"error": "PDF klasörü bulunamadı"}), 404
    
    pdf_files = [f for f in os.listdir(UPLOAD_FOLDER) if f.endswith('.pdf')]
    
    if not pdf_files:
        return jsonify({"error": "Yerel PDF dosyası bulunamadı"}), 404
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    migrated = 0
    errors = []
    
    try:
        for pdf_file in pdf_files:
            try:
                filepath = os.path.join(UPLOAD_FOLDER, pdf_file)
                
                cur.execute("SELECT id FROM report_cards WHERE pdf_filename = %s", (pdf_file,))
                if cur.fetchone():
                    continue
                
                storage_path = None
                if object_storage and object_storage.enabled:
                    try:
                        storage_path = f"report_cards/{pdf_file}"
                        with open(filepath, 'rb') as f:
                            pdf_data = f.read()
                        object_storage.upload_from_bytes(pdf_data, storage_path)
                        logger.info(f"PDF Object Storage'a migre edildi: {storage_path}")
                    except Exception as e:
                        logger.warning(f"Object Storage migrasyon hatası: {e}")
                        storage_path = None
                
                exam_name = "Migre Edilmiş Sınav"
                if "TG" in pdf_file:
                    exam_name = "TÜRKİYE GENELİ"
                
                cur.execute("""
                    INSERT INTO report_cards (exam_name, publisher, class_name, pdf_filename, pdf_storage_path, uploaded_by, parse_status)
                    VALUES (%s, %s, %s, %s, %s, %s, 'pending')
                    RETURNING id
                """, (exam_name, '', '', pdf_file, storage_path, current_user.id))
                
                report_card_id = cur.fetchone()['id']
                conn.commit()
                
                thread = threading.Thread(
                    target=parse_pdf_background,
                    args=(report_card_id, filepath)
                )
                thread.daemon = True
                thread.start()
                
                migrated += 1
                
            except Exception as e:
                errors.append(f"{pdf_file}: {str(e)}")
                logger.error(f"PDF migrasyon hatası: {pdf_file} - {e}")
        
        return jsonify({
            "success": True,
            "message": f"{migrated} PDF dosyası migre edildi ve işleniyor...",
            "migrated": migrated,
            "total_files": len(pdf_files),
            "errors": errors
        })
        
    except Exception as e:
        conn.rollback()
        logger.error(f"Migrate error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/reprocess-all', methods=['POST'])
@login_required
def reprocess_all_report_cards():
    """Tüm PDF'leri yeniden işle - Sadece admin"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Tüm karneleri al
        cur.execute("""
            SELECT id, exam_name, pdf_filename, pdf_storage_path 
            FROM report_cards 
            WHERE pdf_filename IS NOT NULL OR pdf_storage_path IS NOT NULL
        """)
        cards = cur.fetchall()
        
        if not cards:
            return jsonify({"error": "İşlenecek PDF bulunamadı"}), 404
        
        processed = 0
        errors = []
        
        for card in cards:
            try:
                # PDF dosyasını bul
                filepath = None
                
                local_path = os.path.join(UPLOAD_FOLDER, card['pdf_filename']) if card['pdf_filename'] else None
                if local_path and os.path.exists(local_path):
                    filepath = local_path
                elif card['pdf_storage_path'] and object_storage and object_storage.enabled:
                    try:
                        import tempfile
                        pdf_data, _ = object_storage.download_as_bytes(card['pdf_storage_path'])
                        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
                        temp_file.write(pdf_data)
                        temp_file.close()
                        filepath = temp_file.name
                    except:
                        pass
                
                if not filepath:
                    errors.append(f"{card['exam_name']}: PDF bulunamadı")
                    continue
                
                # Eski verileri sil
                cur.execute("""
                    DELETE FROM report_card_answers 
                    WHERE student_result_id IN (
                        SELECT id FROM report_card_students WHERE report_card_id = %s
                    )
                """, (card['id'],))
                cur.execute("DELETE FROM report_card_students WHERE report_card_id = %s", (card['id'],))
                cur.execute("""
                    UPDATE report_cards 
                    SET parse_status = 'processing', parse_error = NULL,
                        processed_student_count = 0, expected_student_count = 0
                    WHERE id = %s
                """, (card['id'],))
                conn.commit()
                
                # Arkaplanda parse et
                thread = threading.Thread(
                    target=parse_pdf_background,
                    args=(card['id'], filepath)
                )
                thread.daemon = True
                thread.start()
                
                processed += 1
                
            except Exception as e:
                errors.append(f"{card['exam_name']}: {str(e)}")
        
        return jsonify({
            "success": True,
            "message": f"{processed} PDF yeniden işleniyor...",
            "processed": processed,
            "errors": errors
        })
        
    except Exception as e:
        logger.error(f"Reprocess all error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


def parse_pdf_background(report_card_id: int, filepath: str):
    """Arkaplanda PDF parse et"""
    import sys
    import traceback
    import tempfile
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    
    conn = None
    cur = None
    temp_pdf_path = None
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Başladığını kaydet
        cur.execute("""
            UPDATE report_cards 
            SET parse_error = 'PARSE BAŞLADI: ' || NOW()::text
            WHERE id = %s
        """, (report_card_id,))
        conn.commit()
        
        # PDF bilgilerini al (Object Storage için)
        cur.execute("""
            SELECT pdf_filename, pdf_storage_path FROM report_cards WHERE id = %s
        """, (report_card_id,))
        card_info = cur.fetchone()
        
        # Dosya kontrolü - önce verilen path'i dene
        actual_filepath = filepath
        if not os.path.exists(actual_filepath):
            # Local path'i dene
            local_path = os.path.join(UPLOAD_FOLDER, card_info['pdf_filename']) if card_info['pdf_filename'] else None
            if local_path and os.path.exists(local_path):
                actual_filepath = local_path
            # Object Storage'dan indir
            elif card_info['pdf_storage_path'] and object_storage and object_storage.enabled:
                try:
                    cur.execute("""
                        UPDATE report_cards SET parse_error = 'Object Storage indir...'
                        WHERE id = %s
                    """, (report_card_id,))
                    conn.commit()
                    
                    pdf_data, _ = object_storage.download_as_bytes(card_info['pdf_storage_path'])
                    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
                    temp_file.write(pdf_data)
                    temp_file.close()
                    actual_filepath = temp_file.name
                    temp_pdf_path = actual_filepath
                    logger.info(f"PDF Object Storage'dan thread'de indirildi: {card_info['pdf_storage_path']}")
                except Exception as storage_err:
                    cur.execute("""
                        UPDATE report_cards 
                        SET parse_status = 'failed', parse_error = %s
                        WHERE id = %s
                    """, (f"STORAGE HATASI: {str(storage_err)}", report_card_id))
                    conn.commit()
                    return
            else:
                cur.execute("""
                    UPDATE report_cards 
                    SET parse_status = 'failed', parse_error = %s
                    WHERE id = %s
                """, (f"DOSYA BULUNAMADI: {filepath}", report_card_id))
                conn.commit()
                return
        
        # Parser'ı import et
        try:
            from utils.pdf_report_parser import parse_report_card_pdf
        except ImportError as ie:
            cur.execute("""
                UPDATE report_cards 
                SET parse_status = 'failed', parse_error = %s
                WHERE id = %s
            """, (f"IMPORT HATASI: {str(ie)}", report_card_id))
            conn.commit()
            return
        
        # PDF'i parse et
        cur.execute("""
            UPDATE report_cards SET parse_error = 'PDF parse ediliyor...'
            WHERE id = %s
        """, (report_card_id,))
        conn.commit()
        
        students, error = parse_report_card_pdf(actual_filepath)
        
        if error:
            cur.execute("""
                UPDATE report_cards 
                SET parse_status = 'failed', parse_error = %s
                WHERE id = %s
            """, (error, report_card_id))
            conn.commit()
            return
        
        # Beklenen öğrenci sayısını hemen kaydet (ilerleme için)
        cur.execute("""
            UPDATE report_cards 
            SET expected_student_count = %s, processed_student_count = 0
            WHERE id = %s
        """, (len(students), report_card_id))
        conn.commit()
        logger.info(f"PDF parsed: {len(students)} students found")
        
        # Karne bilgilerinden sınıf ve sınav bilgisini al
        cur.execute("SELECT class_name, exam_name FROM report_cards WHERE id = %s", (report_card_id,))
        report_info = cur.fetchone()
        report_class = report_info['class_name']
        exam_name = report_info['exam_name']
        
        # "ALL_X" formatı kontrolü - tüm sınıf seviyesi için tek PDF
        is_grade_level_upload = report_class and report_class.startswith('ALL_')
        grade_level = None
        if is_grade_level_upload:
            grade_level = report_class.split('_')[1]  # "ALL_7" -> "7"
        
        student_count = 0
        updated_count = 0
        inserted_count = 0
        skipped_count = 0
        
        for student_data in students:
            # Mevcut öğrenciyi bul
            student_name = student_data.get('student_name', '')
            pdf_class = student_data.get('class_name')  # PDF'den gelen sınıf (örn: "8/A", "7/B")
            
            # PDF'de sınıf yoksa User tablosundan al (student_no veya ad soyad ile)
            student_no = student_data.get('student_no')
            if not pdf_class and student_no:
                # Önce student_no ile ara
                cur.execute("""
                    SELECT class_name FROM users 
                    WHERE role = 'student' AND student_no = %s AND class_name IS NOT NULL
                    LIMIT 1
                """, (student_no,))
                user_result = cur.fetchone()
                if user_result and user_result['class_name']:
                    pdf_class = user_result['class_name']
                    logger.info(f"Sınıf bilgisi User tablosundan alındı (student_no={student_no}): {pdf_class}")
            
            # Hala bulunamadıysa ad soyad ile dene
            if not pdf_class and student_name:
                cur.execute("""
                    SELECT class_name FROM users 
                    WHERE role = 'student' 
                    AND UPPER(full_name) = UPPER(%s)
                    AND class_name IS NOT NULL
                    LIMIT 1
                """, (student_name,))
                user_result = cur.fetchone()
                if user_result and user_result['class_name']:
                    pdf_class = user_result['class_name']
                    logger.info(f"Sınıf bilgisi User tablosundan alındı (ad={student_name}): {pdf_class}")
            
            # Tüm sınıf seviyesi yüklemesinde, PDF'deki sınıfı kullan
            if is_grade_level_upload:
                if pdf_class:
                    # PDF'den gelen sınıf bilgisi var, doğru seviye mi kontrol et
                    # "8/A" -> "8", "7/B" -> "7"
                    pdf_grade = pdf_class.split('/')[0] if '/' in pdf_class else pdf_class[0]
                    if pdf_grade != grade_level:
                        # Bu öğrenci farklı sınıf seviyesinde, atla
                        skipped_count += 1
                        continue
                    student_class = pdf_class
                else:
                    # PDF'den sınıf bilgisi yok, varsayılan olarak seviye + A şubesi
                    student_class = f"{grade_level}/A"
            else:
                # Tek sınıf yüklemesi - eski mantık
                student_class = pdf_class or report_class
            lgs_score = student_data.get('lgs_score')
            
            user_id = None
            
            # 1. Önce öğrenci numarasıyla ara (en güvenilir)
            if student_no:
                # 1a. student_no alanında ara
                cur.execute("""
                    SELECT id FROM users 
                    WHERE role = 'student' AND student_no = %s
                    LIMIT 1
                """, (student_no,))
                result = cur.fetchone()
                if result:
                    user_id = result['id']
                
                # 1b. Bulunamadıysa username'de ara (bazı sistemlerde öğrenci no username olarak kayıtlı)
                if not user_id:
                    cur.execute("""
                        SELECT id FROM users 
                        WHERE role = 'student' AND username = %s
                        LIMIT 1
                    """, (student_no,))
                    result = cur.fetchone()
                    if result:
                        user_id = result['id']
            
            # 2. Öğrenci no ile bulunamadıysa, isimle ara
            if not user_id and student_name:
                cur.execute("""
                    SELECT id FROM users 
                    WHERE role = 'student' 
                    AND (
                        UPPER(full_name) = UPPER(%s)
                        OR UPPER(full_name) LIKE UPPER(%s)
                    )
                    AND (class_name = %s OR %s IS NULL)
                    LIMIT 1
                """, (student_name, f"%{student_name}%", student_class, student_class))
                
                result = cur.fetchone()
                if result:
                    user_id = result['id']
            
            # Mükerrer kontrol - öncelik: student_no > (isim + LGS puanı)
            existing_student_id = None
            
            # 1. Önce student_no ile kontrol (aynı sınavda aynı numara varsa)
            if student_no:
                cur.execute("""
                    SELECT rcs.id FROM report_card_students rcs
                    WHERE rcs.report_card_id = %s 
                    AND TRIM(rcs.student_no) = TRIM(%s)
                    LIMIT 1
                """, (report_card_id, student_no))
                existing = cur.fetchone()
                if existing:
                    existing_student_id = existing['id']
            
            # 2. student_no ile bulunamadıysa isim + LGS puanı ile kontrol
            if not existing_student_id and student_name and lgs_score is not None:
                cur.execute("""
                    SELECT rcs.id FROM report_card_students rcs
                    JOIN report_cards rc ON rcs.report_card_id = rc.id
                    WHERE rc.exam_name = %s 
                    AND UPPER(rcs.student_name) = UPPER(%s)
                    AND rcs.lgs_score = %s
                    LIMIT 1
                """, (exam_name, student_name, lgs_score))
                existing = cur.fetchone()
                if existing:
                    existing_student_id = existing['id']
            
            # Öğrencinin sınıf seviyesini al (6/A -> 6)
            student_grade = student_data.get('grade', 8)
            
            if existing_student_id:
                # Mevcut kaydı güncelle
                cur.execute("""
                    UPDATE report_card_students SET
                        user_id = %s, student_no = %s, class_name = %s, grade_level = %s,
                        percentile = %s, total_questions = %s, total_correct = %s, 
                        total_wrong = %s, total_blank = %s, total_net = %s, success_rate = %s
                    WHERE id = %s
                """, (
                    user_id, student_data.get('student_no'), student_class, student_grade,
                    student_data.get('percentile'), student_data.get('total_questions'), 
                    student_data.get('total_correct'), student_data.get('total_wrong'), 
                    student_data.get('total_blank'), student_data.get('total_net'), 
                    student_data.get('success_rate'), existing_student_id
                ))
                student_result_id = existing_student_id
                updated_count += 1
                
                # Eski ders ve cevap kayıtlarını sil
                cur.execute("DELETE FROM report_card_subjects WHERE student_result_id = %s", (existing_student_id,))
                cur.execute("DELETE FROM report_card_answers WHERE student_result_id = %s", (existing_student_id,))
            else:
                # Yeni kayıt oluştur
                cur.execute("""
                    INSERT INTO report_card_students (
                        report_card_id, user_id, student_name, student_no, class_name, grade_level,
                        lgs_score, percentile, total_questions, total_correct, 
                        total_wrong, total_blank, total_net, success_rate
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                """, (
                    report_card_id, user_id, student_name,
                    student_data.get('student_no'), student_class, student_grade,
                    student_data.get('lgs_score'), student_data.get('percentile'),
                    student_data.get('total_questions'), student_data.get('total_correct'),
                    student_data.get('total_wrong'), student_data.get('total_blank'),
                    student_data.get('total_net'), student_data.get('success_rate')
                ))
                student_result_id = cur.fetchone()['id']
                inserted_count += 1
            
            student_count += 1
            
            # Ders sonuçları
            for subject, subj_data in student_data.get('subjects', {}).items():
                cur.execute("""
                    INSERT INTO report_card_subjects (
                        student_result_id, subject, question_count, correct_count,
                        wrong_count, blank_count, net_score, success_rate,
                        correct_answers, student_answers
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    student_result_id, subject,
                    subj_data.get('question_count'), subj_data.get('correct_count'),
                    subj_data.get('wrong_count'), subj_data.get('blank_count'),
                    subj_data.get('net_score'), subj_data.get('success_rate'),
                    subj_data.get('correct_answers'), subj_data.get('student_answers')
                ))
            
            # Cevaplar
            for answer in student_data.get('answers', []):
                cur.execute("""
                    INSERT INTO report_card_answers (
                        student_result_id, subject, question_number,
                        correct_answer, student_answer, is_correct, is_blank,
                        outcome_code, outcome_text
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    student_result_id, answer.get('subject'),
                    answer.get('question_number'), answer.get('correct_answer'),
                    answer.get('student_answer'), answer.get('is_correct'),
                    answer.get('is_blank'), answer.get('outcome_code'),
                    answer.get('outcome_text')
                ))
            
            # Her 10 öğrencide bir commit yap (connection timeout önlemek için)
            if student_count % 10 == 0:
                cur.execute("""
                    UPDATE report_cards SET processed_student_count = %s WHERE id = %s
                """, (student_count, report_card_id))
                conn.commit()
                logger.info(f"Progress: {student_count} öğrenci kaydedildi")
        
        # SON KALAN ÖĞRENCİLER İÇİN COMMIT (10'un katı olmayan sayılar için kritik!)
        # Örn: 104 öğrenci varsa, son 4 öğrenci burada commit edilir
        if student_count % 10 != 0:
            cur.execute("""
                UPDATE report_cards SET processed_student_count = %s WHERE id = %s
            """, (student_count, report_card_id))
            conn.commit()
            logger.info(f"Final batch: {student_count} öğrenci kaydedildi (son {student_count % 10} öğrenci)")
        
        # Başarılı - istatistikleri de kaydet
        if is_grade_level_upload and skipped_count > 0:
            parse_stats = f"{inserted_count} yeni, {updated_count} güncellendi, {skipped_count} farklı sınıf atlandı"
        else:
            parse_stats = f"{inserted_count} yeni, {updated_count} güncellendi"
        cur.execute("""
            UPDATE report_cards 
            SET parse_status = 'completed', 
                student_count = %s,
                parse_error = %s,
                parsed_at = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (student_count, parse_stats, report_card_id))
        
        conn.commit()
        logger.info(f"PDF parse completed: {student_count} students ({inserted_count} new, {updated_count} updated, {skipped_count} skipped)")
        
    except Exception as e:
        error_detail = f"{str(e)}\n{traceback.format_exc()}"
        logger.error(f"Background parse error: {error_detail}")
        
        try:
            if conn:
                conn.rollback()
            if cur and conn:
                cur.execute("""
                    UPDATE report_cards 
                    SET parse_status = 'failed', parse_error = %s
                    WHERE id = %s
                """, (error_detail[:500], report_card_id))
                conn.commit()
        except Exception as db_err:
            logger.error(f"DB update error: {db_err}")
    finally:
        if temp_pdf_path and os.path.exists(temp_pdf_path):
            try:
                os.unlink(temp_pdf_path)
            except:
                pass
        if cur:
            cur.close()
        if conn:
            conn.close()


@report_cards_bp.route('/api/status/<int:report_card_id>')
@login_required
def get_parse_status(report_card_id):
    """Parse durumunu kontrol et"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("""
            SELECT parse_status, student_count, parse_error,
                   expected_student_count, processed_student_count
            FROM report_cards WHERE id = %s
        """, (report_card_id,))
        
        result = cur.fetchone()
        if not result:
            return jsonify({"error": "Karne bulunamadı"}), 404
        
        return jsonify(result)
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/students/<int:report_card_id>')
@login_required
def get_report_card_students(report_card_id):
    """Karnedeki öğrencileri listele"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("""
            SELECT rcs.*, u.full_name as matched_user_name
            FROM report_card_students rcs
            LEFT JOIN users u ON rcs.user_id = u.id
            WHERE rcs.report_card_id = %s
            ORDER BY rcs.student_name
        """, (report_card_id,))
        
        students = cur.fetchall()
        
        # Her öğrenci için ders sonuçlarını getir
        for student in students:
            cur.execute("""
                SELECT * FROM report_card_subjects
                WHERE student_result_id = %s
            """, (student['id'],))
            student['subjects'] = cur.fetchall()
            
            # Decimal tiplerini float'a çevir
            for key in ['lgs_score', 'percentile', 'total_net', 'success_rate']:
                if student.get(key):
                    student[key] = float(student[key])
            
            for subj in student['subjects']:
                for key in ['net_score', 'success_rate']:
                    if subj.get(key):
                        subj[key] = float(subj[key])
        
        return jsonify(students)
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/students-by-exams', methods=['POST'])
@login_required
def get_students_by_exams():
    """Birden fazla sınav için öğrenci listesi getir"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    data = request.get_json()
    report_card_ids = data.get('report_card_ids', [])
    class_filter = data.get('class_name', '')  # Sınıf filtresi
    
    import logging
    logging.info(f"students-by-exams called: report_card_ids={report_card_ids}, class_filter='{class_filter}'")
    
    if not report_card_ids:
        return jsonify([])
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Seçilen sınavlardaki öğrencileri getir - sınıf filtresi uygula
        if class_filter:
            # Sınıf filtresi varsa, sadece o sınıftaki öğrencileri getir
            # 5D -> 5/D veya 5/D formatlarını destekle
            class_variants = [class_filter]
            if '/' not in class_filter and len(class_filter) == 2:
                class_variants.append(f"{class_filter[0]}/{class_filter[1]}")
            elif '/' in class_filter:
                class_variants.append(class_filter.replace('/', ''))
            
            cur.execute("""
                WITH student_keys AS (
                    SELECT 
                        rcs.user_id, 
                        rcs.student_name, 
                        rcs.class_name,
                        rcs.report_card_id,
                        COALESCE(rcs.user_id::text, UPPER(rcs.student_name)) as student_key
                    FROM report_card_students rcs
                    WHERE rcs.report_card_id = ANY(%s)
                    AND rcs.class_name = ANY(%s)
                )
                SELECT DISTINCT ON (student_key)
                    sk.user_id, sk.student_name, sk.class_name,
                    u.full_name as matched_user_name,
                    (SELECT COUNT(DISTINCT report_card_id) FROM student_keys WHERE student_key = sk.student_key) as exam_count
                FROM student_keys sk
                LEFT JOIN users u ON sk.user_id = u.id
                ORDER BY student_key, sk.student_name
            """, (report_card_ids, class_variants))
        else:
            cur.execute("""
                WITH student_keys AS (
                    SELECT 
                        rcs.user_id, 
                        rcs.student_name, 
                        rcs.class_name,
                        rcs.report_card_id,
                        COALESCE(rcs.user_id::text, UPPER(rcs.student_name)) as student_key
                    FROM report_card_students rcs
                    WHERE rcs.report_card_id = ANY(%s)
                )
                SELECT DISTINCT ON (student_key)
                    sk.user_id, sk.student_name, sk.class_name,
                    u.full_name as matched_user_name,
                    (SELECT COUNT(DISTINCT report_card_id) FROM student_keys WHERE student_key = sk.student_key) as exam_count
                FROM student_keys sk
                LEFT JOIN users u ON sk.user_id = u.id
                ORDER BY student_key, sk.student_name
            """, (report_card_ids,))
        
        students = cur.fetchall()
        total_exams = len(report_card_ids)
        
        # Her öğrenci için tüm sınavlarda olup olmadığını işaretle
        for student in students:
            student['in_all_exams'] = int(student['exam_count']) >= total_exams
        
        return jsonify(students)
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/delete/<int:report_card_id>', methods=['DELETE'])
@login_required
def delete_report_card(report_card_id):
    """Karneyi sil - Sadece admin"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor()
    
    try:
        # Önce dosyayı bul
        cur.execute("SELECT pdf_filename FROM report_cards WHERE id = %s", (report_card_id,))
        result = cur.fetchone()
        
        if result and result[0]:
            filepath = os.path.join(UPLOAD_FOLDER, result[0])
            if os.path.exists(filepath):
                os.remove(filepath)
        
        # Veritabanından sil (cascade ile ilişkili kayıtlar da silinir)
        cur.execute("DELETE FROM report_cards WHERE id = %s", (report_card_id,))
        conn.commit()
        
        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/error-report', methods=['POST'])
@login_required
def generate_error_report():
    """Hata karnesi oluştur"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    data = request.get_json()
    report_card_ids = data.get('report_card_ids', [])
    student_user_id = data.get('student_user_id')
    
    if not report_card_ids:
        return jsonify({"error": "En az bir sınav seçin"}), 400
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Öğrencinin yanlış yaptığı soruları getir - SQL injection koruması için ANY() kullan
        if student_user_id:
            cur.execute("""
                SELECT 
                    rca.subject, rca.question_number, rca.correct_answer, 
                    rca.student_answer, rca.outcome_code, rca.outcome_text,
                    rc.exam_name
                FROM report_card_answers rca
                JOIN report_card_students rcs ON rca.student_result_id = rcs.id
                JOIN report_cards rc ON rcs.report_card_id = rc.id
                WHERE rc.id = ANY(%s)
                AND (rca.is_correct = FALSE OR rca.is_blank = TRUE)
                AND rcs.user_id = %s
                ORDER BY rca.subject, rc.exam_name, rca.question_number
            """, (report_card_ids, student_user_id))
        else:
            cur.execute("""
                SELECT 
                    rca.subject, rca.question_number, rca.correct_answer, 
                    rca.student_answer, rca.outcome_code, rca.outcome_text,
                    rc.exam_name
                FROM report_card_answers rca
                JOIN report_card_students rcs ON rca.student_result_id = rcs.id
                JOIN report_cards rc ON rcs.report_card_id = rc.id
                WHERE rc.id = ANY(%s)
                AND (rca.is_correct = FALSE OR rca.is_blank = TRUE)
                ORDER BY rca.subject, rc.exam_name, rca.question_number
            """, (report_card_ids,))
        errors = cur.fetchall()
        
        # Derslere göre grupla
        grouped = {}
        for error in errors:
            subject = error['subject']
            if subject not in grouped:
                grouped[subject] = []
            grouped[subject].append(error)
        
        return jsonify({
            "total_errors": len(errors),
            "by_subject": grouped
        })
        
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/error-report-pdf', methods=['GET', 'POST'])
@login_required
def generate_error_report_pdf():
    """Hata karnesi PDF oluştur"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    # GET veya POST parametreleri al
    if request.method == 'GET':
        ids_str = request.args.get('report_card_ids', '')
        report_card_ids = [int(x) for x in ids_str.split(',') if x.strip().isdigit()]
        student_user_id = request.args.get('student_user_id')
        if student_user_id and student_user_id.isdigit():
            student_user_id = int(student_user_id)
        else:
            student_user_id = None
        student_name_filter = request.args.get('student_name_filter')
        student_name = request.args.get('student_name', 'Öğrenci')
    elif request.is_json:
        data = request.get_json()
        report_card_ids = data.get('report_card_ids', [])
        student_user_id = data.get('student_user_id')
        student_name_filter = data.get('student_name_filter')
        student_name = data.get('student_name', 'Öğrenci')
    else:
        ids_str = request.form.get('report_card_ids', '')
        report_card_ids = [int(x) for x in ids_str.split(',') if x.strip().isdigit()]
        student_user_id = request.form.get('student_user_id')
        if student_user_id and student_user_id.isdigit():
            student_user_id = int(student_user_id)
        else:
            student_user_id = None
        student_name_filter = request.form.get('student_name_filter')
        student_name = request.form.get('student_name', 'Öğrenci')
    
    if not report_card_ids:
        return jsonify({"error": "En az bir sınav seçin"}), 400
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Sınav isimlerini al - SQL injection koruması için tuple kullan
        cur.execute("SELECT id, exam_name FROM report_cards WHERE id = ANY(%s)", (report_card_ids,))
        exams = {row['id']: row['exam_name'] for row in cur.fetchall()}
        
        # Öğrenci bilgilerini al
        class_name = ''
        if student_user_id:
            cur.execute("SELECT full_name, class_name FROM users WHERE id = %s", (student_user_id,))
            user_info = cur.fetchone()
            if user_info:
                student_name = user_info['full_name']
                class_name = user_info.get('class_name', '')
        elif student_name_filter:
            student_name = student_name_filter
            # Öğrenci isminden sınıf bilgisini al
            cur.execute("""
                SELECT class_name FROM report_card_students 
                WHERE student_name = %s AND report_card_id = ANY(%s)
                LIMIT 1
            """, (student_name_filter, report_card_ids))
            class_info = cur.fetchone()
            if class_info:
                class_name = class_info.get('class_name', '')
        
        # Yanlış soruları getir - SQL injection koruması için ANY() kullan
        if student_user_id:
            cur.execute("""
                SELECT 
                    rca.subject, rca.question_number, rca.correct_answer, 
                    rca.student_answer, rca.outcome_code, rca.outcome_text,
                    rca.is_blank, rc.exam_name, rc.id as report_card_id
                FROM report_card_answers rca
                JOIN report_card_students rcs ON rca.student_result_id = rcs.id
                JOIN report_cards rc ON rcs.report_card_id = rc.id
                WHERE rc.id = ANY(%s)
                AND (rca.is_correct = FALSE OR rca.is_blank = TRUE)
                AND rcs.user_id = %s
                ORDER BY rca.subject, rc.exam_name, rca.question_number
            """, (report_card_ids, student_user_id))
        elif student_name_filter:
            cur.execute("""
                SELECT 
                    rca.subject, rca.question_number, rca.correct_answer, 
                    rca.student_answer, rca.outcome_code, rca.outcome_text,
                    rca.is_blank, rc.exam_name, rc.id as report_card_id
                FROM report_card_answers rca
                JOIN report_card_students rcs ON rca.student_result_id = rcs.id
                JOIN report_cards rc ON rcs.report_card_id = rc.id
                WHERE rc.id = ANY(%s)
                AND (rca.is_correct = FALSE OR rca.is_blank = TRUE)
                AND UPPER(rcs.student_name) = UPPER(%s)
                ORDER BY rca.subject, rc.exam_name, rca.question_number
            """, (report_card_ids, student_name_filter))
        else:
            cur.execute("""
                SELECT 
                    rca.subject, rca.question_number, rca.correct_answer, 
                    rca.student_answer, rca.outcome_code, rca.outcome_text,
                    rca.is_blank, rc.exam_name, rc.id as report_card_id
                FROM report_card_answers rca
                JOIN report_card_students rcs ON rca.student_result_id = rcs.id
                JOIN report_cards rc ON rcs.report_card_id = rc.id
                WHERE rc.id = ANY(%s)
                AND (rca.is_correct = FALSE OR rca.is_blank = TRUE)
                ORDER BY rca.subject, rc.exam_name, rca.question_number
            """, (report_card_ids,))
        errors = cur.fetchall()
        
        # PDF oluştur
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        
        # Türkçe karakter desteği için font
        try:
            pdfmetrics.registerFont(TTFont('DejaVuSans', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
            pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
            default_font = 'DejaVuSans'
        except:
            default_font = 'Helvetica'
        
        # Başlık stili
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontName=default_font,
            fontSize=18,
            alignment=TA_CENTER,
            spaceAfter=20
        )
        
        subject_style = ParagraphStyle(
            'Subject',
            parent=styles['Heading2'],
            fontName=default_font,
            fontSize=14,
            textColor=colors.HexColor('#2563eb'),
            spaceAfter=10,
            spaceBefore=15
        )
        
        normal_style = ParagraphStyle(
            'Normal',
            parent=styles['Normal'],
            fontName=default_font,
            fontSize=10
        )
        
        # Logo ve okul adı header'ı
        header_elements = create_pdf_header(styles, default_font)
        elements.extend(header_elements)
        
        # Başlık
        elements.append(Paragraph(f"HATA KARNESİ", title_style))
        elements.append(Paragraph(f"{student_name} {f'- {class_name}' if class_name else ''}", normal_style))
        elements.append(Paragraph(f"Seçilen Sınavlar: {', '.join(exams.values())}", normal_style))
        elements.append(Spacer(1, 20))
        
        # Ders isim çevirisi
        subject_names = {
            'turkce': 'Türkçe',
            'matematik': 'Matematik',
            'fen': 'Fen Bilimleri',
            'inkilap': 'İnkılap Tarihi',
            'din': 'Din Kültürü',
            'ingilizce': 'İngilizce',
            'sosyal': 'Sosyal Bilgiler'
        }
        
        # Derslere göre grupla
        grouped = {}
        for error in errors:
            subject = error['subject']
            if subject not in grouped:
                grouped[subject] = []
            grouped[subject].append(error)
        
        # Esnek hücre stilleri
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontName=default_font,
            fontSize=7,
            leading=9,
            wordWrap='CJK'
        )
        
        # Header cell style
        header_cell_style = ParagraphStyle(
            'HeaderCellStyle',
            parent=styles['Normal'],
            fontName=default_font,
            fontSize=9,
            textColor=colors.white,
            alignment=TA_CENTER
        )
        
        if not grouped:
            elements.append(Paragraph("Bu öğrenci seçilen sınavlarda yanlış yapmamış!", normal_style))
        else:
            # Her ders için tablo oluştur
            for subject, subject_errors in grouped.items():
                subject_display = subject_names.get(subject, subject)
                elements.append(Paragraph(f"{subject_display}", subject_style))
                
                # Tablo verisi - tüm hücreler Paragraph olmalı (Türkçe karakter desteği için)
                table_data = [[
                    Paragraph('Soru', header_cell_style),
                    Paragraph('Doğru', header_cell_style),
                    Paragraph('Öğrenci', header_cell_style),
                    Paragraph('Durum', header_cell_style),
                    Paragraph('Kazanım', header_cell_style)
                ]]
                
                for err in subject_errors:
                    status = 'Boş' if err.get('is_blank') else 'Yanlış'
                    outcome_text = err.get('outcome_text') or '-'
                    outcome_para = Paragraph(outcome_text, cell_style)
                    
                    table_data.append([
                        Paragraph(str(err.get('question_number', '')), cell_style),
                        Paragraph(err.get('correct_answer') or '-', cell_style),
                        Paragraph(err.get('student_answer') or '-', cell_style),
                        Paragraph(status, cell_style),
                        outcome_para
                    ])
                
                table = Table(table_data, colWidths=[40, 40, 50, 45, None])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                    ('ALIGN', (0, 0), (3, -1), 'CENTER'),
                    ('ALIGN', (4, 1), (4, -1), 'LEFT'),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('TOPPADDING', (0, 1), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
                ]))
                
                elements.append(table)
                elements.append(Spacer(1, 15))
        
        # Özet
        elements.append(Spacer(1, 20))
        elements.append(Paragraph(f"Toplam Hata: {len(errors)} soru", normal_style))
        
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"hata_karnesi_{student_name.replace(' ', '_')}.pdf"
        )
        
    except Exception as e:
        logger.error(f"Error report PDF error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


# ============ ÖĞRENCİ SELF-SERVICE ENDPOINTS ============

@report_cards_bp.route('/student/my-reports')
@login_required
def student_my_reports():
    """Öğrencinin kendi karnelerini listele"""
    if current_user.role != 'student':
        return jsonify({"error": "Sadece öğrenciler erişebilir"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # user_id ile eşleşen VEYA isim benzerliği olan kayıtları getir
        cur.execute("""
            SELECT rc.id, rc.exam_name, rc.class_name, rc.upload_date,
                   rcs.total_net, rcs.lgs_score, rcs.percentile,
                   rcs.total_correct, rcs.total_wrong, rcs.total_blank
            FROM report_card_students rcs
            JOIN report_cards rc ON rcs.report_card_id = rc.id
            WHERE rcs.user_id = %s 
            OR (rcs.user_id IS NULL AND UPPER(rcs.student_name) = UPPER(%s))
            ORDER BY rc.upload_date DESC
        """, (current_user.id, current_user.full_name))
        
        reports = cur.fetchall()
        
        # Decimal tiplerini float'a çevir
        for r in reports:
            for key in ['total_net', 'lgs_score', 'percentile']:
                if r.get(key):
                    r[key] = float(r[key])
            if r.get('upload_date'):
                r['upload_date'] = r['upload_date'].isoformat()
        
        return jsonify(reports)
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/student/error-report-pdf', methods=['POST'])
@login_required
def student_error_report_pdf():
    """Öğrenci kendi hata karnesini PDF olarak indir"""
    if current_user.role != 'student':
        return jsonify({"error": "Sadece öğrenciler erişebilir"}), 403
    
    if request.is_json:
        data = request.get_json()
        report_card_ids = data.get('report_card_ids', [])
    else:
        ids_str = request.form.get('report_card_ids', '')
        report_card_ids = [int(x) for x in ids_str.split(',') if x.strip().isdigit()]
    
    if not report_card_ids:
        return jsonify({"error": "En az bir sınav seçin"}), 400
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Önce bu öğrencinin bu sınavlarda kaydı var mı kontrol et (user_id veya isim ile)
        cur.execute("""
            SELECT DISTINCT rc.id FROM report_cards rc
            JOIN report_card_students rcs ON rcs.report_card_id = rc.id
            WHERE rc.id = ANY(%s) 
            AND (rcs.user_id = %s OR (rcs.user_id IS NULL AND UPPER(rcs.student_name) = UPPER(%s)))
        """, (report_card_ids, current_user.id, current_user.full_name))
        
        valid_ids = [r['id'] for r in cur.fetchall()]
        if not valid_ids:
            return jsonify({"error": "Bu sınavlarda kaydınız bulunamadı"}), 404
        
        # Sınav isimlerini al
        cur.execute("SELECT id, exam_name FROM report_cards WHERE id = ANY(%s)", (valid_ids,))
        exams = {row['id']: row['exam_name'] for row in cur.fetchall()}
        
        # Yanlış soruları getir (user_id veya isim ile)
        cur.execute("""
            SELECT 
                rca.subject, rca.question_number, rca.correct_answer, 
                rca.student_answer, rca.outcome_code, rca.outcome_text,
                rca.is_blank, rc.exam_name, rc.id as report_card_id
            FROM report_card_answers rca
            JOIN report_card_students rcs ON rca.student_result_id = rcs.id
            JOIN report_cards rc ON rcs.report_card_id = rc.id
            WHERE rc.id = ANY(%s)
            AND (rcs.user_id = %s OR (rcs.user_id IS NULL AND UPPER(rcs.student_name) = UPPER(%s)))
            AND (rca.is_correct = FALSE OR rca.is_blank = TRUE)
            ORDER BY rca.subject, rc.exam_name, rca.question_number
        """, (valid_ids, current_user.id, current_user.full_name))
        errors = cur.fetchall()
        
        # PDF oluştur (aynı format)
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        elements = []
        
        default_font = 'Helvetica'
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle('Title', fontName=default_font, fontSize=16, alignment=1, spaceAfter=20)
        subject_style = ParagraphStyle('Subject', fontName=default_font, fontSize=12, textColor=colors.HexColor('#3b82f6'), spaceBefore=15, spaceAfter=10)
        normal_style = ParagraphStyle('Normal', fontName=default_font, fontSize=10)
        
        # Logo ve okul adı header'ı
        header_elements = create_pdf_header(styles, default_font)
        elements.extend(header_elements)
        
        elements.append(Paragraph("HATA KARNESİ", title_style))
        elements.append(Paragraph(f"{current_user.full_name}", normal_style))
        elements.append(Paragraph(f"Seçilen Sınavlar: {', '.join(exams.values())}", normal_style))
        elements.append(Spacer(1, 20))
        
        subject_names = {
            'turkce': 'Türkçe', 'matematik': 'Matematik', 'fen': 'Fen Bilimleri',
            'inkilap': 'İnkılap Tarihi', 'din': 'Din Kültürü', 'ingilizce': 'İngilizce',
            'sosyal': 'Sosyal Bilgiler'
        }
        
        grouped = {}
        for error in errors:
            subject = error['subject']
            if subject not in grouped:
                grouped[subject] = []
            grouped[subject].append(error)
        
        # Esnek hücre stilleri (öğrenci)
        cell_style = ParagraphStyle(
            'CellStyleStudent',
            parent=styles['Normal'],
            fontName=default_font,
            fontSize=7,
            leading=9,
            wordWrap='CJK'
        )
        
        # Header cell style (öğrenci)
        header_cell_style = ParagraphStyle(
            'HeaderCellStyleStudent',
            parent=styles['Normal'],
            fontName=default_font,
            fontSize=9,
            textColor=colors.white,
            alignment=TA_CENTER
        )
        
        if not grouped:
            elements.append(Paragraph("Tebrikler! Seçilen sınavlarda yanlış yapmamışsın!", normal_style))
        else:
            for subject, subject_errors in grouped.items():
                subject_display = subject_names.get(subject, subject)
                elements.append(Paragraph(f"{subject_display}", subject_style))
                
                # Tablo verisi - tüm hücreler Paragraph olmalı
                table_data = [[
                    Paragraph('Soru', header_cell_style),
                    Paragraph('Doğru', header_cell_style),
                    Paragraph('Öğrenci', header_cell_style),
                    Paragraph('Durum', header_cell_style),
                    Paragraph('Kazanım', header_cell_style)
                ]]
                
                for err in subject_errors:
                    status = 'Boş' if err.get('is_blank') else 'Yanlış'
                    outcome_text = err.get('outcome_text') or '-'
                    outcome_para = Paragraph(outcome_text, cell_style)
                    
                    table_data.append([
                        Paragraph(str(err.get('question_number', '')), cell_style),
                        Paragraph(err.get('correct_answer') or '-', cell_style),
                        Paragraph(err.get('student_answer') or '-', cell_style),
                        Paragraph(status, cell_style),
                        outcome_para
                    ])
                
                table = Table(table_data, colWidths=[40, 40, 50, 45, None])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                    ('ALIGN', (0, 0), (3, -1), 'CENTER'),
                    ('ALIGN', (4, 1), (4, -1), 'LEFT'),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('TOPPADDING', (0, 1), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
                ]))
                
                elements.append(table)
                elements.append(Spacer(1, 15))
        
        elements.append(Spacer(1, 20))
        elements.append(Paragraph(f"Toplam Hata: {len(errors)} soru", normal_style))
        
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"hata_karnesi_{current_user.full_name.replace(' ', '_')}.pdf"
        )
        
    except Exception as e:
        logger.error(f"Student error report PDF error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/comparison-error-report', methods=['POST'])
@login_required
def get_comparison_error_report():
    """Karşılaştırmalı hata karnesi - TG1 vs TG2 gelişim takibi"""
    data = request.get_json()
    report_card_ids = data.get('report_card_ids', [])
    
    if current_user.role == 'student':
        student_user_id = current_user.id
        student_name_filter = None
    else:
        student_user_id = data.get('student_user_id')
        student_name_filter = data.get('student_name_filter')
    
    if len(report_card_ids) < 2:
        return jsonify({"error": "Karşılaştırma için en az 2 sınav seçin"}), 400
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("""
            SELECT id, exam_name, upload_date,
                   COALESCE(NULLIF(regexp_replace(exam_name, '[^0-9]', '', 'g'), '')::int, 0) as exam_num
            FROM report_cards 
            WHERE id = ANY(%s) 
            ORDER BY exam_num ASC, exam_name ASC
        """, (report_card_ids,))
        exams = cur.fetchall()
        
        if len(exams) < 2:
            return jsonify({"error": "Seçilen sınavlar bulunamadı"}), 404
        
        first_exam = exams[0]
        last_exam = exams[-1]
        
        base_query = """
            SELECT 
                rca.subject, rca.question_number, rca.correct_answer, 
                rca.student_answer, rca.is_correct, rca.is_blank,
                rca.outcome_code, rca.outcome_text,
                rc.id as report_card_id, rc.exam_name
            FROM report_card_answers rca
            JOIN report_card_students rcs ON rca.student_result_id = rcs.id
            JOIN report_cards rc ON rcs.report_card_id = rc.id
            WHERE rc.id = %s
        """
        
        if student_user_id:
            base_query += " AND rcs.user_id = %s"
            cur.execute(base_query, (first_exam['id'], student_user_id))
            first_answers = cur.fetchall()
            cur.execute(base_query, (last_exam['id'], student_user_id))
            last_answers = cur.fetchall()
        elif student_name_filter:
            base_query += " AND UPPER(rcs.student_name) = UPPER(%s)"
            cur.execute(base_query, (first_exam['id'], student_name_filter))
            first_answers = cur.fetchall()
            cur.execute(base_query, (last_exam['id'], student_name_filter))
            last_answers = cur.fetchall()
        else:
            return jsonify({"error": "Öğrenci seçimi gerekli"}), 400
        
        first_map = {}
        for ans in first_answers:
            key = ans.get('outcome_code') or f"{ans['subject']}_{ans['question_number']}"
            first_map[key] = ans
        
        last_map = {}
        for ans in last_answers:
            key = ans.get('outcome_code') or f"{ans['subject']}_{ans['question_number']}"
            last_map[key] = ans
        
        comparison = []
        all_keys = set(first_map.keys()) | set(last_map.keys())
        
        for key in all_keys:
            first = first_map.get(key)
            last = last_map.get(key)
            
            if first and last:
                first_correct = first.get('is_correct', False)
                last_correct = last.get('is_correct', False)
                
                if not first_correct and last_correct:
                    status = 'improved'
                elif first_correct and not last_correct:
                    status = 'regressed'
                elif not first_correct and not last_correct:
                    status = 'still_wrong'
                else:
                    status = 'still_correct'
                
                comparison.append({
                    'subject': first['subject'],
                    'question_number': first['question_number'],
                    'outcome_code': first.get('outcome_code') or last.get('outcome_code'),
                    'outcome_text': first.get('outcome_text') or last.get('outcome_text'),
                    'first_exam': first_exam['exam_name'],
                    'first_correct': first_correct,
                    'first_answer': first.get('student_answer'),
                    'last_exam': last_exam['exam_name'],
                    'last_correct': last_correct,
                    'last_answer': last.get('student_answer'),
                    'correct_answer': first.get('correct_answer') or last.get('correct_answer'),
                    'status': status
                })
        
        stats = {
            'improved': len([c for c in comparison if c['status'] == 'improved']),
            'regressed': len([c for c in comparison if c['status'] == 'regressed']),
            'still_wrong': len([c for c in comparison if c['status'] == 'still_wrong']),
            'still_correct': len([c for c in comparison if c['status'] == 'still_correct']),
            'total': len(comparison)
        }
        
        by_subject = {}
        for item in comparison:
            subj = item['subject']
            if subj not in by_subject:
                by_subject[subj] = []
            if item['status'] != 'still_correct':
                by_subject[subj].append(item)
        
        return jsonify({
            "first_exam": first_exam['exam_name'],
            "last_exam": last_exam['exam_name'],
            "stats": stats,
            "by_subject": by_subject,
            "comparison": [c for c in comparison if c['status'] != 'still_correct']
        })
        
    except Exception as e:
        logger.error(f"Comparison error report error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


# ==================== TOPLU GELİŞİM RAPORU ====================

@report_cards_bp.route('/api/multi-exam-progress', methods=['POST'])
@login_required
def get_multi_exam_progress():
    """Tüm sınavları karşılaştıran toplu gelişim raporu - yeni tablolardan"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    data = request.get_json() if request.is_json else request.form
    class_name = data.get('class_name') or data.get('class_id')
    
    if not class_name:
        return jsonify({"error": "Sınıf seçilmedi"}), 400
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        class_normalized = class_name.replace('/', '')
        
        if current_user.role == 'teacher':
            cur.execute("SELECT class_name FROM teacher_classes WHERE teacher_id = %s", (current_user.id,))
            allowed_classes = [row['class_name'] for row in cur.fetchall()]
            allowed_normalized = [c.replace('/', '') for c in allowed_classes]
            if class_name not in allowed_classes and class_normalized not in allowed_normalized:
                return jsonify({"error": "Bu sınıfa erişim yetkiniz yok"}), 403
        
        cur.execute("""
            SELECT DISTINCT e.id, e.exam_name, e.upload_date
            FROM report_card_exams e
            JOIN report_card_results r ON r.exam_id = e.id
            WHERE r.class_name = %s OR REPLACE(r.class_name, '/', '') = %s
            ORDER BY e.upload_date ASC, e.exam_name ASC
        """, (class_name, class_normalized))
        exams = cur.fetchall()
        
        if len(exams) < 1:
            return jsonify({"error": "Bu sınıfa ait karne bulunamadı"}), 404
        
        exam_names = [e['exam_name'] for e in exams]
        
        cur.execute("""
            SELECT r.student_name, r.student_no, e.exam_name, r.totals
            FROM report_card_results r
            JOIN report_card_exams e ON r.exam_id = e.id
            WHERE r.class_name = %s OR REPLACE(r.class_name, '/', '') = %s
            ORDER BY r.student_name, e.upload_date ASC
        """, (class_name, class_normalized))
        results = cur.fetchall()
        
        students = {}
        for row in results:
            student_key = row['student_no'].strip() if row['student_no'] else row['student_name']
            
            if student_key not in students:
                students[student_key] = {
                    'student_name': row['student_name'],
                    'student_number': row['student_no'],
                    'exams': {}
                }
            
            students[student_key]['student_name'] = row['student_name']
            
            totals = row.get('totals') or {}
            if isinstance(totals, str):
                try:
                    totals = json.loads(totals)
                except:
                    totals = {}
            
            total_correct = totals.get('correct_count', totals.get('correct', 0)) or 0
            total_wrong = totals.get('wrong_count', totals.get('wrong', 0)) or 0
            total_net = totals.get('net_score', totals.get('net', 0)) or 0
            
            students[student_key]['exams'][row['exam_name']] = {
                'total_correct': total_correct,
                'total_wrong': total_wrong,
                'total_net': round(float(total_net), 2)
            }
        
        for student_data in students.values():
            exams_data = student_data['exams']
            available_exams = [e for e in exam_names if e in exams_data]
            if len(available_exams) >= 2:
                first = exams_data[available_exams[0]]
                last = exams_data[available_exams[-1]]
                student_data['net_change'] = round(last['total_net'] - first['total_net'], 2)
            else:
                student_data['net_change'] = 0
        
        return jsonify({
            "exam_names": exam_names,
            "students": list(students.values())
        })
        
    except Exception as e:
        logger.error(f"Multi-exam progress error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


# ==================== ÖĞRENCİ KİŞİSEL GELİŞİM ====================

@report_cards_bp.route('/api/student-progress', methods=['GET'])
@login_required
def get_student_progress():
    """Öğrencinin kendi sınavlarındaki gelişimi"""
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Öğrencinin tüm sınavlarını kronolojik sırada al
        cur.execute("""
            SELECT rc.exam_name, rcs.total_correct, rcs.total_wrong, rcs.total_blank,
                   rcs.total_net, rcs.success_rate, rcs.lgs_score, rcs.percentile,
                   COALESCE(NULLIF(regexp_replace(rc.exam_name, '[^0-9]', '', 'g'), '')::int, 0) as exam_num
            FROM report_card_students rcs
            JOIN report_cards rc ON rcs.report_card_id = rc.id
            WHERE rcs.user_id = %s
            ORDER BY exam_num ASC, rc.exam_name ASC
        """, (current_user.id,))
        exams = cur.fetchall()
        
        if len(exams) == 0:
            return jsonify({"error": "Henüz sınav verisi bulunamadı", "exams": []})
        
        exam_list = []
        for exam in exams:
            exam_list.append({
                'exam_name': exam['exam_name'],
                'total_correct': exam['total_correct'],
                'total_wrong': exam['total_wrong'],
                'total_blank': exam['total_blank'],
                'total_net': float(exam['total_net']) if exam['total_net'] else 0,
                'success_rate': float(exam['success_rate']) if exam['success_rate'] else 0,
                'lgs_score': float(exam['lgs_score']) if exam['lgs_score'] else None,
                'percentile': float(exam['percentile']) if exam['percentile'] else None
            })
        
        # Gelişim hesapla
        if len(exam_list) >= 2:
            first = exam_list[0]
            last = exam_list[-1]
            net_change = round(last['total_net'] - first['total_net'], 2)
            success_change = round(last['success_rate'] - first['success_rate'], 2)
        else:
            net_change = 0
            success_change = 0
        
        return jsonify({
            "exams": exam_list,
            "net_change": net_change,
            "success_change": success_change
        })
        
    except Exception as e:
        logger.error(f"Student progress error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


# ==================== ÖĞRENCİ KAZANIM ANALİZİ ====================

@report_cards_bp.route('/api/student-outcomes', methods=['GET'])
@login_required
def get_student_outcomes():
    """Öğrencinin kendi kazanım bazlı analizi"""
    if current_user.role != 'student':
        return jsonify({"error": "Sadece öğrenciler erişebilir"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("""
            SELECT rca.outcome_code, rca.subject, rca.outcome_text as description,
                   COUNT(*) as total_questions,
                   SUM(CASE WHEN rca.is_correct THEN 1 ELSE 0 END) as total_correct,
                   SUM(CASE WHEN NOT rca.is_correct AND NOT rca.is_blank THEN 1 ELSE 0 END) as total_wrong,
                   SUM(CASE WHEN rca.is_blank THEN 1 ELSE 0 END) as total_blank
            FROM report_card_answers rca
            JOIN report_card_students rcs ON rca.student_result_id = rcs.id
            WHERE rcs.user_id = %s AND rca.outcome_code IS NOT NULL
            GROUP BY rca.outcome_code, rca.subject, rca.outcome_text
            ORDER BY rca.subject, rca.outcome_code
        """, (current_user.id,))
        outcomes = cur.fetchall()
        
        by_subject = {}
        for outcome in outcomes:
            subject = outcome['subject']
            if subject not in by_subject:
                by_subject[subject] = []
            
            total_q = outcome['total_questions']
            total_c = outcome['total_correct']
            success_rate = round((total_c / total_q * 100), 1) if total_q > 0 else 0
            
            by_subject[subject].append({
                'outcome_code': outcome['outcome_code'],
                'outcome_text': outcome['description'],
                'total_questions': total_q,
                'total_correct': total_c,
                'success_rate': success_rate
            })
        
        return jsonify({"subjects": by_subject})
        
    except Exception as e:
        logger.error(f"Student outcomes error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/student-priority-topics', methods=['GET'])
@login_required
def get_student_priority_topics():
    """Öğrencinin en çok hata yaptığı konular - öncelikli sıralama"""
    if current_user.role != 'student':
        return jsonify({"error": "Sadece öğrenciler erişebilir"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("""
            SELECT 
                rca.subject,
                rca.outcome_code,
                COALESCE(lo.outcome_text, rca.outcome_text, rca.outcome_code) as outcome_text,
                COUNT(*) as total_questions,
                SUM(CASE WHEN rca.is_correct THEN 1 ELSE 0 END) as correct_count,
                SUM(CASE WHEN NOT rca.is_correct AND NOT rca.is_blank THEN 1 ELSE 0 END) as wrong_count,
                SUM(CASE WHEN rca.is_blank THEN 1 ELSE 0 END) as blank_count
            FROM report_card_answers rca
            JOIN report_card_students rcs ON rca.student_result_id = rcs.id
            LEFT JOIN learning_outcomes lo ON UPPER(TRIM(rca.outcome_code)) = UPPER(TRIM(lo.outcome_code))
            WHERE rcs.user_id = %s AND rca.outcome_code IS NOT NULL
            GROUP BY rca.subject, rca.outcome_code, lo.outcome_text, rca.outcome_text
            ORDER BY wrong_count DESC, blank_count DESC
        """, (current_user.id,))
        topics = cur.fetchall()
        
        subject_names = {
            'turkce': 'Türkçe', 'matematik': 'Matematik', 'fen': 'Fen Bilimleri',
            'inkilap': 'İnkılap Tarihi', 'din': 'Din Kültürü', 'ingilizce': 'İngilizce',
            'sosyal': 'Sosyal Bilgiler'
        }
        
        by_subject = {}
        priority_list = []
        
        for topic in topics:
            subject = topic['subject']
            error_count = (topic['wrong_count'] or 0) + (topic['blank_count'] or 0)
            total = topic['total_questions'] or 1
            error_rate = round((error_count / total) * 100, 1)
            
            if subject not in by_subject:
                by_subject[subject] = {
                    'name': subject_names.get(subject, subject),
                    'total_errors': 0,
                    'topics': []
                }
            
            by_subject[subject]['total_errors'] += error_count
            by_subject[subject]['topics'].append({
                'outcome_code': topic['outcome_code'],
                'outcome_text': topic['outcome_text'] or topic['outcome_code'],
                'wrong_count': topic['wrong_count'] or 0,
                'blank_count': topic['blank_count'] or 0,
                'total_errors': error_count,
                'total_questions': total,
                'error_rate': error_rate
            })
            
            if error_count > 0:
                priority_list.append({
                    'subject': subject,
                    'subject_name': subject_names.get(subject, subject),
                    'outcome_code': topic['outcome_code'],
                    'outcome_text': topic['outcome_text'] or topic['outcome_code'],
                    'total_errors': error_count,
                    'error_rate': error_rate
                })
        
        for subj in by_subject.values():
            subj['topics'].sort(key=lambda x: x['total_errors'], reverse=True)
        
        priority_list.sort(key=lambda x: x['total_errors'], reverse=True)
        
        return jsonify({
            "by_subject": by_subject,
            "priority_list": priority_list[:20],
            "total_error_topics": len([t for t in topics if (t['wrong_count'] or 0) + (t['blank_count'] or 0) > 0])
        })
        
    except Exception as e:
        logger.error(f"Student priority topics error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/student/priority-topics-pdf', methods=['GET'])
@login_required
def student_priority_topics_pdf():
    """Öğrenci öncelikli konular PDF"""
    if current_user.role != 'student':
        return jsonify({"error": "Sadece öğrenciler erişebilir"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("""
            SELECT 
                rca.subject,
                rca.outcome_code,
                COALESCE(lo.outcome_text, rca.outcome_text, rca.outcome_code) as outcome_text,
                SUM(CASE WHEN NOT rca.is_correct AND NOT rca.is_blank THEN 1 ELSE 0 END) as wrong_count,
                SUM(CASE WHEN rca.is_blank THEN 1 ELSE 0 END) as blank_count
            FROM report_card_answers rca
            JOIN report_card_students rcs ON rca.student_result_id = rcs.id
            LEFT JOIN learning_outcomes lo ON UPPER(TRIM(rca.outcome_code)) = UPPER(TRIM(lo.outcome_code))
            WHERE rcs.user_id = %s AND rca.outcome_code IS NOT NULL
            GROUP BY rca.subject, rca.outcome_code, lo.outcome_text, rca.outcome_text
            ORDER BY (SUM(CASE WHEN NOT rca.is_correct AND NOT rca.is_blank THEN 1 ELSE 0 END) + 
                     SUM(CASE WHEN rca.is_blank THEN 1 ELSE 0 END)) DESC
        """, (current_user.id,))
        topics = cur.fetchall()
        
        cur.execute("SELECT student_no, student_name FROM report_card_students WHERE user_id = %s LIMIT 1", 
                   (current_user.id,))
        student_info = cur.fetchone()
        
        subject_names = {
            'turkce': 'Türkçe', 'matematik': 'Matematik', 'fen': 'Fen Bilimleri',
            'inkilap': 'İnkılap Tarihi', 'din': 'Din Kültürü', 'ingilizce': 'İngilizce',
            'sosyal': 'Sosyal Bilgiler'
        }
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontName=PDF_FONT_BOLD, 
                                      fontSize=16, alignment=1, spaceAfter=20, textColor=colors.HexColor('#1f2937'))
        header_style = ParagraphStyle('Header', parent=styles['Heading2'], fontName=PDF_FONT_BOLD,
                                       fontSize=12, spaceAfter=10, textColor=colors.HexColor('#dc2626'))
        normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontName=PDF_FONT, fontSize=10)
        
        logo_path = 'static/images/school_logo.png'
        if os.path.exists(logo_path):
            logo = RLImage(logo_path, width=60, height=60)
            elements.append(logo)
        
        elements.append(Paragraph("Öncelikli Konular Raporu", title_style))
        if student_info:
            elements.append(Paragraph(f"Öğrenci: {student_info['student_name']} ({student_info['student_no']})", normal_style))
        elements.append(Paragraph(f"Tarih: {datetime.now().strftime('%d.%m.%Y')}", normal_style))
        elements.append(Spacer(1, 20))
        
        priority_topics = [t for t in topics if (t['wrong_count'] or 0) + (t['blank_count'] or 0) > 0][:15]
        
        if priority_topics:
            elements.append(Paragraph("En Çok Hata Yapılan Konular", header_style))
            
            data = [['Sıra', 'Ders', 'Kazanım', 'Yanlış', 'Boş', 'Toplam']]
            for i, topic in enumerate(priority_topics, 1):
                total = (topic['wrong_count'] or 0) + (topic['blank_count'] or 0)
                data.append([
                    str(i),
                    subject_names.get(topic['subject'], topic['subject']),
                    Paragraph(topic['outcome_text'][:60] + '...' if len(topic['outcome_text']) > 60 else topic['outcome_text'], normal_style),
                    str(topic['wrong_count'] or 0),
                    str(topic['blank_count'] or 0),
                    str(total)
                ])
            
            table = Table(data, colWidths=[30, 70, 250, 45, 45, 50])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc2626')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD),
                ('FONTNAME', (0, 1), (-1, -1), PDF_FONT),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (2, 1), (2, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fef2f2')])
            ]))
            elements.append(table)
        else:
            elements.append(Paragraph("Henüz yeterli veri bulunmamaktadır.", normal_style))
        
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'oncelikli_konular_{datetime.now().strftime("%Y%m%d")}.pdf'
        )
        
    except Exception as e:
        logger.error(f"Priority topics PDF error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/student/outcome-analysis-pdf', methods=['GET'])
@login_required
def student_outcome_analysis_pdf():
    """Öğrenci kazanım analizi PDF"""
    if current_user.role != 'student':
        return jsonify({"error": "Sadece öğrenciler erişebilir"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("""
            SELECT 
                rca.subject,
                rca.outcome_code,
                COALESCE(lo.outcome_text, rca.outcome_text, rca.outcome_code) as outcome_text,
                COUNT(*) as total_questions,
                SUM(CASE WHEN rca.is_correct THEN 1 ELSE 0 END) as correct_count
            FROM report_card_answers rca
            JOIN report_card_students rcs ON rca.student_result_id = rcs.id
            LEFT JOIN learning_outcomes lo ON UPPER(TRIM(rca.outcome_code)) = UPPER(TRIM(lo.outcome_code))
            WHERE rcs.user_id = %s AND rca.outcome_code IS NOT NULL
            GROUP BY rca.subject, rca.outcome_code, lo.outcome_text, rca.outcome_text
            ORDER BY rca.subject, rca.outcome_code
        """, (current_user.id,))
        outcomes = cur.fetchall()
        
        cur.execute("SELECT student_no, student_name FROM report_card_students WHERE user_id = %s LIMIT 1", 
                   (current_user.id,))
        student_info = cur.fetchone()
        
        subject_names = {
            'turkce': 'Türkçe', 'matematik': 'Matematik', 'fen': 'Fen Bilimleri',
            'inkilap': 'İnkılap Tarihi', 'din': 'Din Kültürü', 'ingilizce': 'İngilizce',
            'sosyal': 'Sosyal Bilgiler'
        }
        
        by_subject = {}
        for outcome in outcomes:
            subject = outcome['subject']
            if subject not in by_subject:
                by_subject[subject] = []
            total = outcome['total_questions'] or 1
            correct = outcome['correct_count'] or 0
            rate = round((correct / total) * 100, 1)
            by_subject[subject].append({
                'code': outcome['outcome_code'],
                'text': outcome['outcome_text'],
                'rate': rate
            })
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontName=PDF_FONT_BOLD, 
                                      fontSize=16, alignment=1, spaceAfter=20, textColor=colors.HexColor('#1f2937'))
        header_style = ParagraphStyle('Header', parent=styles['Heading2'], fontName=PDF_FONT_BOLD,
                                       fontSize=12, spaceAfter=10, textColor=colors.HexColor('#3b82f6'))
        normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontName=PDF_FONT, fontSize=10)
        
        logo_path = 'static/images/school_logo.png'
        if os.path.exists(logo_path):
            logo = RLImage(logo_path, width=60, height=60)
            elements.append(logo)
        
        elements.append(Paragraph("Kazanım Analizi Raporu", title_style))
        if student_info:
            elements.append(Paragraph(f"Öğrenci: {student_info['student_name']} ({student_info['student_no']})", normal_style))
        elements.append(Paragraph(f"Tarih: {datetime.now().strftime('%d.%m.%Y')}", normal_style))
        elements.append(Spacer(1, 20))
        
        for subject, items in by_subject.items():
            elements.append(Paragraph(subject_names.get(subject, subject), header_style))
            
            weak = [i for i in items if i['rate'] < 50]
            strong = [i for i in items if i['rate'] >= 80]
            
            if weak:
                elements.append(Paragraph("Geliştirilmesi Gereken:", normal_style))
                data = [['Kazanım', 'Başarı']]
                for item in weak[:10]:
                    text = item['text'][:70] + '...' if len(item['text']) > 70 else item['text']
                    data.append([Paragraph(text, normal_style), f"%{item['rate']}"])
                
                table = Table(data, colWidths=[400, 60])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fef2f2')),
                    ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD),
                    ('FONTNAME', (0, 1), (-1, -1), PDF_FONT),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                elements.append(table)
                elements.append(Spacer(1, 10))
            
            if strong:
                elements.append(Paragraph("Güçlü Alanlar:", normal_style))
                data = [['Kazanım', 'Başarı']]
                for item in strong[:10]:
                    text = item['text'][:70] + '...' if len(item['text']) > 70 else item['text']
                    data.append([Paragraph(text, normal_style), f"%{item['rate']}"])
                
                table = Table(data, colWidths=[400, 60])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f0fdf4')),
                    ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD),
                    ('FONTNAME', (0, 1), (-1, -1), PDF_FONT),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                elements.append(table)
            
            elements.append(Spacer(1, 15))
        
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'kazanim_analizi_{datetime.now().strftime("%Y%m%d")}.pdf'
        )
        
    except Exception as e:
        logger.error(f"Outcome analysis PDF error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


# ==================== ÖĞRETMEN - ÖĞRENCİ BAZLI KAZANIM ANALİZİ ====================

@report_cards_bp.route('/api/class-students', methods=['GET'])
@login_required
def get_class_students():
    """Sınıftaki öğrenci listesini getir - seçilen sınavlara göre filtrele"""
    logger.info(f"Class students API called by {current_user.username}, role: {current_user.role}")
    
    if current_user.role not in ['admin', 'teacher']:
        logger.warning(f"Unauthorized access attempt by {current_user.username}")
        return jsonify({"error": "Yetkisiz erişim", "students": []}), 403
    
    class_name = request.args.get('class_name', '')
    exam_ids_str = request.args.get('exam_ids', '')
    logger.info(f"Requested class_name: {class_name}, exam_ids: {exam_ids_str}")
    
    if not class_name:
        return jsonify({"students": []})
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        if exam_ids_str:
            exam_ids = [int(x) for x in exam_ids_str.split(',') if x.strip().isdigit()]
            if exam_ids:
                cur.execute("""
                    SELECT DISTINCT r.student_no, r.student_name
                    FROM report_card_results r
                    WHERE r.class_name = %s AND r.exam_id = ANY(%s)
                    ORDER BY r.student_name
                """, (class_name, exam_ids))
            else:
                cur.execute("""
                    SELECT DISTINCT student_no, student_name
                    FROM report_card_results
                    WHERE class_name = %s
                    ORDER BY student_name
                """, (class_name,))
        else:
            cur.execute("""
                SELECT DISTINCT student_no, student_name
                FROM report_card_results
                WHERE class_name = %s
                ORDER BY student_name
            """, (class_name,))
        
        raw_students = cur.fetchall()
        
        # Aynı numaralı öğrencileri birleştir - en uzun ismi seç
        seen = {}
        for s in raw_students:
            normalized_no = s['student_no'].lstrip('0') or '0'
            name = s['student_name']
            
            if normalized_no not in seen:
                seen[normalized_no] = {'student_no': s['student_no'], 'student_name': name}
            else:
                # Aynı numara varsa, daha uzun ismi seç (eksik harf sorunu için)
                if len(name) > len(seen[normalized_no]['student_name']):
                    seen[normalized_no] = {'student_no': s['student_no'], 'student_name': name}
        
        students = sorted(seen.values(), key=lambda x: x['student_name'])
        logger.info(f"Found {len(students)} unique students for class {class_name}")
        return jsonify({"students": students})
    except Exception as e:
        logger.error(f"Class students error: {e}")
        return jsonify({"students": []})
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/student-outcome-analysis', methods=['POST'])
@login_required
def get_student_outcome_analysis():
    """Öğretmen için öğrenci bazlı kazanım analizi"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    data = request.get_json()
    class_name = data.get('class_name')
    student_no = data.get('student_no')
    
    if not class_name or not student_no:
        return jsonify({"error": "Sınıf ve öğrenci numarası gerekli"}), 400
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Kazanım analizi
        cur.execute("""
            SELECT 
                rca.subject,
                rca.outcome_code,
                COALESCE(lo.outcome_text, rca.outcome_text, rca.outcome_code) as outcome_text,
                COUNT(*) as total_questions,
                SUM(CASE WHEN rca.is_correct THEN 1 ELSE 0 END) as correct_count,
                SUM(CASE WHEN NOT rca.is_correct AND NOT rca.is_blank THEN 1 ELSE 0 END) as wrong_count,
                SUM(CASE WHEN rca.is_blank THEN 1 ELSE 0 END) as blank_count
            FROM report_card_answers rca
            JOIN report_card_students rcs ON rca.student_result_id = rcs.id
            LEFT JOIN learning_outcomes lo ON UPPER(TRIM(rca.outcome_code)) = UPPER(TRIM(lo.outcome_code))
            WHERE rcs.class_name = %s AND rcs.student_no = %s AND rca.outcome_code IS NOT NULL
            GROUP BY rca.subject, rca.outcome_code, lo.outcome_text, rca.outcome_text
            ORDER BY rca.subject, rca.outcome_code
        """, (class_name, student_no))
        outcomes = cur.fetchall()
        
        subject_names = {
            'turkce': 'Türkçe', 'matematik': 'Matematik', 'fen': 'Fen Bilimleri',
            'inkilap': 'İnkılap Tarihi', 'din': 'Din Kültürü', 'ingilizce': 'İngilizce',
            'sosyal': 'Sosyal Bilgiler'
        }
        
        by_subject = {}
        priority_topics = []
        
        for outcome in outcomes:
            subject = outcome['subject']
            total = outcome['total_questions'] or 1
            correct = outcome['correct_count'] or 0
            wrong = outcome['wrong_count'] or 0
            blank = outcome['blank_count'] or 0
            success_rate = round((correct / total) * 100, 1)
            error_count = wrong + blank
            
            if subject not in by_subject:
                by_subject[subject] = []
            
            by_subject[subject].append({
                'outcome_code': outcome['outcome_code'],
                'outcome_text': outcome['outcome_text'] or outcome['outcome_code'],
                'total_questions': total,
                'correct_count': correct,
                'wrong_count': wrong,
                'blank_count': blank,
                'success_rate': success_rate
            })
            
            if error_count > 0:
                priority_topics.append({
                    'subject': subject,
                    'outcome_code': outcome['outcome_code'],
                    'outcome_text': outcome['outcome_text'] or outcome['outcome_code'],
                    'total_errors': error_count,
                    'wrong_count': wrong,
                    'blank_count': blank
                })
        
        priority_topics.sort(key=lambda x: x['total_errors'], reverse=True)
        
        return jsonify({
            "by_subject": by_subject,
            "priority_topics": priority_topics[:15]
        })
        
    except Exception as e:
        logger.error(f"Student outcome analysis error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/student-outcome-pdf', methods=['GET'])
@login_required
def get_student_outcome_pdf():
    """Öğretmen için öğrenci bazlı kazanım PDF"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    class_name = request.args.get('class_name')
    student_no = request.args.get('student_no')
    
    if not class_name or not student_no:
        return jsonify({"error": "Sınıf ve öğrenci numarası gerekli"}), 400
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("SELECT student_name FROM report_card_students WHERE class_name = %s AND student_no = %s LIMIT 1",
                   (class_name, student_no))
        student = cur.fetchone()
        student_name = student['student_name'] if student else student_no
        
        cur.execute("""
            SELECT 
                rca.subject,
                rca.outcome_code,
                COALESCE(lo.outcome_text, rca.outcome_text, rca.outcome_code) as outcome_text,
                COUNT(*) as total_questions,
                SUM(CASE WHEN rca.is_correct THEN 1 ELSE 0 END) as correct_count,
                SUM(CASE WHEN NOT rca.is_correct AND NOT rca.is_blank THEN 1 ELSE 0 END) as wrong_count,
                SUM(CASE WHEN rca.is_blank THEN 1 ELSE 0 END) as blank_count
            FROM report_card_answers rca
            JOIN report_card_students rcs ON rca.student_result_id = rcs.id
            LEFT JOIN learning_outcomes lo ON UPPER(TRIM(rca.outcome_code)) = UPPER(TRIM(lo.outcome_code))
            WHERE rcs.class_name = %s AND rcs.student_no = %s AND rca.outcome_code IS NOT NULL
            GROUP BY rca.subject, rca.outcome_code, lo.outcome_text, rca.outcome_text
            ORDER BY (SUM(CASE WHEN NOT rca.is_correct AND NOT rca.is_blank THEN 1 ELSE 0 END) + 
                     SUM(CASE WHEN rca.is_blank THEN 1 ELSE 0 END)) DESC
        """, (class_name, student_no))
        outcomes = cur.fetchall()
        
        subject_names = {
            'turkce': 'Türkçe', 'matematik': 'Matematik', 'fen': 'Fen Bilimleri',
            'inkilap': 'İnkılap Tarihi', 'din': 'Din Kültürü', 'ingilizce': 'İngilizce',
            'sosyal': 'Sosyal Bilgiler'
        }
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontName=PDF_FONT_BOLD, 
                                      fontSize=16, alignment=1, spaceAfter=20, textColor=colors.HexColor('#1f2937'))
        header_style = ParagraphStyle('Header', parent=styles['Heading2'], fontName=PDF_FONT_BOLD,
                                       fontSize=12, spaceAfter=10, textColor=colors.HexColor('#8b5cf6'))
        priority_style = ParagraphStyle('Priority', parent=styles['Heading2'], fontName=PDF_FONT_BOLD,
                                       fontSize=12, spaceAfter=10, textColor=colors.HexColor('#dc2626'))
        normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontName=PDF_FONT, fontSize=10)
        
        logo_path = 'static/images/school_logo.png'
        if os.path.exists(logo_path):
            logo = RLImage(logo_path, width=60, height=60)
            elements.append(logo)
        
        elements.append(Paragraph("Öğrenci Kazanım Analizi", title_style))
        elements.append(Paragraph(f"Öğrenci: {student_name} ({student_no}) - Sınıf: {class_name}", normal_style))
        elements.append(Paragraph(f"Tarih: {datetime.now().strftime('%d.%m.%Y')}", normal_style))
        elements.append(Spacer(1, 20))
        
        priority_topics = [o for o in outcomes if (o['wrong_count'] or 0) + (o['blank_count'] or 0) > 0][:10]
        
        if priority_topics:
            elements.append(Paragraph("Öncelikli Konular (Telafi Gerekli)", priority_style))
            
            cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontName=PDF_FONT, fontSize=8, leading=10)
            data = [['Sıra', 'Ders', 'Kazanım', 'Yanlış', 'Boş', 'Toplam']]
            for i, topic in enumerate(priority_topics, 1):
                total_err = (topic['wrong_count'] or 0) + (topic['blank_count'] or 0)
                text = topic['outcome_text'] or topic.get('outcome_code', '')
                data.append([
                    str(i),
                    subject_names.get(topic['subject'], topic['subject']),
                    Paragraph(text, cell_style),
                    str(topic['wrong_count'] or 0),
                    str(topic['blank_count'] or 0),
                    str(total_err)
                ])
            
            table = Table(data, colWidths=[30, 60, 260, 40, 40, 50])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc2626')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD),
                ('FONTNAME', (0, 1), (-1, -1), PDF_FONT),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (2, 1), (2, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fef2f2')])
            ]))
            elements.append(table)
            elements.append(Spacer(1, 20))
        
        # Kazanım detayları ders bazlı
        elements.append(Paragraph("Kazanım Detayları", header_style))
        
        by_subject = {}
        for o in outcomes:
            subj = o['subject']
            if subj not in by_subject:
                by_subject[subj] = []
            total = o['total_questions'] or 1
            correct = o['correct_count'] or 0
            rate = round((correct / total) * 100, 1)
            by_subject[subj].append({
                'text': o['outcome_text'],
                'rate': rate
            })
        
        for subject, items in by_subject.items():
            elements.append(Paragraph(subject_names.get(subject, subject), normal_style))
            
            weak = [i for i in items if i['rate'] < 50]
            strong = [i for i in items if i['rate'] >= 80]
            
            if weak:
                cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontName=PDF_FONT, fontSize=8, leading=10)
                data = [['Geliştirilmesi Gereken Kazanımlar', 'Başarı']]
                for item in weak[:8]:
                    text = item['text'] or ''
                    data.append([Paragraph(text, cell_style), f"%{item['rate']}"])
                
                table = Table(data, colWidths=[400, 60])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fef2f2')),
                    ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD),
                    ('FONTNAME', (0, 1), (-1, -1), PDF_FONT),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                elements.append(table)
            
            elements.append(Spacer(1, 10))
        
        doc.build(elements)
        buffer.seek(0)
        
        safe_name = student_name.replace(' ', '_')[:20]
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'kazanim_{safe_name}_{datetime.now().strftime("%Y%m%d")}.pdf'
        )
        
    except Exception as e:
        logger.error(f"Student outcome PDF error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


# ==================== KAZANIM BAZLI ANALİTİK ====================

@report_cards_bp.route('/api/outcome-analytics', methods=['POST'])
@login_required
def get_outcome_analytics():
    """Kazanım bazlı toplam soru ve doğru analizi"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    data = request.get_json() if request.is_json else request.form
    class_name = data.get('class_id') if request.is_json else request.form.get('class_id')
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Sınıfa ait tüm karnelerdeki kazanım bazlı verileri topla
        cur.execute("""
            SELECT rca.outcome_code, rca.subject, rca.outcome_text as description,
                   COUNT(*) as total_questions,
                   SUM(CASE WHEN rca.is_correct THEN 1 ELSE 0 END) as total_correct,
                   SUM(CASE WHEN NOT rca.is_correct AND NOT rca.is_blank THEN 1 ELSE 0 END) as total_wrong,
                   SUM(CASE WHEN rca.is_blank THEN 1 ELSE 0 END) as total_blank,
                   COUNT(DISTINCT rc.id) as exam_count,
                   COUNT(DISTINCT rcs.id) as student_attempts
            FROM report_card_answers rca
            JOIN report_card_students rcs ON rca.student_result_id = rcs.id
            JOIN report_cards rc ON rcs.report_card_id = rc.id
            WHERE rcs.class_name IN (%s, %s, 'ALL_' || SUBSTRING(%s, 1, 1)) AND rca.outcome_code IS NOT NULL
            GROUP BY rca.outcome_code, rca.subject, rca.outcome_text
            ORDER BY rca.subject, rca.outcome_code
        """, (class_name, class_name[:1] + '/' + class_name[1:] if len(class_name) >= 2 else class_name, class_name))
        outcomes = cur.fetchall()
        
        # Ders bazlı grupla
        by_subject = {}
        subject_totals = {}
        
        for outcome in outcomes:
            subject = outcome['subject']
            if subject not in by_subject:
                by_subject[subject] = []
                subject_totals[subject] = {'total_questions': 0, 'total_correct': 0}
            
            total_q = outcome['total_questions']
            total_c = outcome['total_correct']
            accuracy = round((total_c / total_q * 100), 1) if total_q > 0 else 0
            
            by_subject[subject].append({
                'outcome_code': outcome['outcome_code'],
                'description': outcome['description'],
                'total_questions': total_q,
                'total_correct': total_c,
                'total_wrong': outcome['total_wrong'],
                'total_blank': outcome['total_blank'],
                'accuracy': accuracy,
                'exam_count': outcome['exam_count'],
                'student_attempts': outcome['student_attempts']
            })
            
            subject_totals[subject]['total_questions'] += total_q
            subject_totals[subject]['total_correct'] += total_c
        
        # Ders ortalamaları hesapla
        for subject, totals in subject_totals.items():
            if totals['total_questions'] > 0:
                totals['average_accuracy'] = round(
                    (totals['total_correct'] / totals['total_questions'] * 100), 1
                )
            else:
                totals['average_accuracy'] = 0
        
        return jsonify({
            "by_subject": by_subject,
            "subject_totals": subject_totals,
            "total_outcomes": len(outcomes)
        })
        
    except Exception as e:
        logger.error(f"Outcome analytics error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/class-outcome-pdf', methods=['GET'])
@login_required
def get_class_outcome_pdf():
    """Sınıf bazlı kazanım analizi PDF - Çoklu sınav destekli (report_card_results)"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    exam_ids_str = request.args.get('exam_ids', '')
    class_name = request.args.get('class_name', '')
    
    if not exam_ids_str:
        return jsonify({"error": "Sınav ID'leri gerekli"}), 400
    
    exam_ids = [int(x) for x in exam_ids_str.split(',') if x.strip().isdigit()]
    if not exam_ids:
        return jsonify({"error": "Geçerli sınav ID'si yok"}), 400
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        placeholders = ','.join(['%s'] * len(exam_ids))
        query = f"""
            SELECT r.subjects, r.class_name, e.exam_name
            FROM report_card_results r
            JOIN report_card_exams e ON r.exam_id = e.id
            WHERE r.exam_id IN ({placeholders})
        """
        params = list(exam_ids)
        
        if class_name:
            class_normalized = class_name.replace('/', '')
            query += " AND (r.class_name = %s OR REPLACE(r.class_name, '/', '') = %s)"
            params.extend([class_name, class_normalized])
        
        cur.execute(query, params)
        results = cur.fetchall()
        
        if not results:
            return jsonify({"error": "Veri bulunamadı"}), 404
        
        outcome_stats = {}
        
        for result in results:
            subjects = result.get('subjects') or {}
            if isinstance(subjects, str):
                subjects = json.loads(subjects)
            
            for subj_key, subj_data in subjects.items():
                subject_label = subj_data.get('subject_label', subj_key)
                if subject_label not in outcome_stats:
                    outcome_stats[subject_label] = {}
                
                for ans in subj_data.get('answers', []):
                    outcome = ans.get('outcome', '') or '-'
                    if outcome not in outcome_stats[subject_label]:
                        outcome_stats[subject_label][outcome] = {'correct': 0, 'wrong': 0, 'blank': 0, 'total': 0}
                    
                    outcome_stats[subject_label][outcome]['total'] += 1
                    if ans.get('status') == 'correct':
                        outcome_stats[subject_label][outcome]['correct'] += 1
                    elif ans.get('status') == 'wrong':
                        outcome_stats[subject_label][outcome]['wrong'] += 1
                    else:
                        outcome_stats[subject_label][outcome]['blank'] += 1
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30, bottomMargin=30, leftMargin=30, rightMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        
        elements.extend(create_pdf_header(styles))
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontName=PDF_FONT, fontSize=16, alignment=TA_CENTER, spaceAfter=15)
        subject_style = ParagraphStyle('Subject', parent=styles['Heading2'], fontName=PDF_FONT, fontSize=12, textColor=colors.HexColor('#6366f1'), spaceAfter=8, spaceBefore=12)
        normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontName=PDF_FONT, fontSize=10)
        cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontName=PDF_FONT, fontSize=7, leading=9, wordWrap='CJK')
        
        elements.append(Paragraph("SINIF KAZANIM ANALİZİ", title_style))
        elements.append(Paragraph(f"Sınıf: {class_name or 'Tüm Sınıflar'}", normal_style))
        elements.append(Paragraph(f"Tarih: {datetime.now().strftime('%d.%m.%Y')}", normal_style))
        elements.append(Spacer(1, 15))
        
        for subject, outcomes in outcome_stats.items():
            elements.append(Paragraph(subject, subject_style))
            
            table_data = [['Kazanım', 'Doğru', 'Yanlış', 'Boş', 'Başarı %']]
            for outcome, stats in outcomes.items():
                total = stats['total']
                success = round((stats['correct'] / total * 100), 1) if total > 0 else 0
                table_data.append([
                    Paragraph(outcome, cell_style),
                    str(stats['correct']),
                    str(stats['wrong']),
                    str(stats['blank']),
                    f"%{success}"
                ])
            
            t = Table(table_data, colWidths=[300, 45, 45, 45, 50])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e0e7ff')),
                ('FONTNAME', (0, 0), (-1, -1), PDF_FONT),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 10))
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f'sinif_kazanim_{class_name.replace("/", "_") if class_name else "tum"}_{datetime.now().strftime("%Y%m%d")}.pdf'
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)
        
    except Exception as e:
        logger.error(f"Class outcome PDF error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


# ==================== KAZANIM YÖNETİMİ ====================

@report_cards_bp.route('/api/outcomes/template')
@login_required
def download_outcomes_template():
    """Kazanım yükleme şablonu indir"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Kazanımlar"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ["Sınıf Seviyesi", "Ders", "Kazanım Kodu", "Kazanım Açıklaması"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    sample_data = [
        [8, "turkce", "T.8.1.3.", "Dinlediklerini/izlediklerini özetler."],
        [8, "turkce", "T.8.3.5.1.", "Bağlamdan yararlanarak bilmediği kelimelerin anlamını tahmin eder."],
        [8, "matematik", "M.8.1.1.2.", "İki doğal sayının ebob ve ekok'unu hesaplar."],
        [8, "fen", "F.8.1.1.1.", "Mevsimlerin oluşumunu açıklar."],
        [8, "inkilap", "İTA.8.1.1.", "Avrupa'daki gelişmelerin Osmanlı Devleti'ne etkilerini analiz eder."],
        [8, "din", "D.8.1.1.", "Kader ve kaza inancını açıklar."],
        [8, "ingilizce", "E.8.1.2.", "Making simple inquiries."],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 60
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='kazanim_sablonu.xlsx'
    )


@report_cards_bp.route('/api/outcomes/upload', methods=['POST'])
@login_required
def upload_outcomes():
    """Kazanım Excel dosyası yükle"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    if 'file' not in request.files:
        return jsonify({"error": "Dosya seçilmedi"}), 400
    
    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"error": "Sadece Excel dosyası (.xlsx) yükleyebilirsiniz"}), 400
    
    from openpyxl import load_workbook
    
    try:
        wb = load_workbook(file)
        ws = wb.active
        
        conn = get_db()
        cur = conn.cursor()
        
        subject_map = {
            # Türkçe
            'türkçe': 'turkce', 'turkce': 'turkce', 'Türkçe': 'turkce', 'TÜRKÇE': 'turkce',
            # Matematik
            'matematik': 'matematik', 'Matematik': 'matematik', 'MATEMATIK': 'matematik', 'mat': 'matematik',
            # Fen Bilimleri
            'fen': 'fen', 'Fen': 'fen', 'FEN': 'fen',
            'fen bilimleri': 'fen', 'Fen Bilimleri': 'fen', 'FEN BİLİMLERİ': 'fen', 'fen bilgisi': 'fen',
            # İnkılap Tarihi
            'inkılap': 'inkilap', 'inkilap': 'inkilap', 'İnkılap': 'inkilap', 'INKILAP': 'inkilap',
            'inkılap tarihi': 'inkilap', 'inkilap tarihi': 'inkilap', 'İnkılap Tarihi': 'inkilap',
            't.c. inkılap tarihi': 'inkilap', 'tc inkılap tarihi': 'inkilap',
            # Din Kültürü
            'din': 'din', 'Din': 'din', 'DIN': 'din',
            'din kültürü': 'din', 'Din Kültürü': 'din', 'DİN KÜLTÜRÜ': 'din',
            'din kültürü ve ahlak bilgisi': 'din', 'dkab': 'din',
            # İngilizce
            'ingilizce': 'ingilizce', 'İngilizce': 'ingilizce', 'INGILIZCE': 'ingilizce', 'İNGİLİZCE': 'ingilizce',
            'yabancı dil': 'ingilizce', 'yabanci dil': 'ingilizce',
            # Sosyal Bilgiler (5-6-7. sınıflar için)
            'sosyal bilgiler': 'sosyal', 'Sosyal Bilgiler': 'sosyal', 'SOSYAL BİLGİLER': 'sosyal', 'sosyal': 'sosyal'
        }
        
        inserted = 0
        updated = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row or len(row) < 4:
                continue
            
            grade_val = row[0]
            if grade_val is None or (isinstance(grade_val, str) and not grade_val.strip()):
                continue
            
            try:
                grade_level = int(grade_val) if grade_val else 8
                subject_raw = str(row[1]).lower().strip() if row[1] else ''
                subject = subject_map.get(subject_raw, subject_raw)
                outcome_code = str(row[2]).strip() if row[2] else ''
                outcome_text = str(row[3]).strip() if row[3] else ''
                
                if not outcome_code or not outcome_text:
                    errors.append(f"Satır {row_idx}: Kazanım kodu veya açıklaması eksik")
                    continue
                
                cur.execute("""
                    INSERT INTO learning_outcomes (grade_level, subject, outcome_code, outcome_text)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (outcome_code) 
                    DO UPDATE SET 
                        grade_level = EXCLUDED.grade_level,
                        subject = EXCLUDED.subject,
                        outcome_text = EXCLUDED.outcome_text
                    RETURNING (xmax = 0) as inserted
                """, (grade_level, subject, outcome_code, outcome_text))
                
                result = cur.fetchone()
                if result and result[0]:
                    inserted += 1
                else:
                    updated += 1
                    
            except Exception as e:
                errors.append(f"Satır {row_idx}: {str(e)}")
        
        conn.commit()
        
        return jsonify({
            "success": True,
            "inserted": inserted,
            "updated": updated,
            "errors": errors[:10]
        })
        
    except Exception as e:
        logger.error(f"Outcomes upload error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if 'cur' in locals():
            cur.close()
        if 'conn' in locals():
            conn.close()


@report_cards_bp.route('/api/outcomes')
@login_required
def get_outcomes():
    """Kayıtlı kazanımları getir"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("""
            SELECT id, grade_level, subject, outcome_code, outcome_text
            FROM learning_outcomes
            ORDER BY grade_level, subject, outcome_code
        """)
        outcomes = cur.fetchall()
        
        by_subject = {}
        for o in outcomes:
            subj = o['subject']
            if subj not in by_subject:
                by_subject[subj] = []
            by_subject[subj].append(o)
        
        return jsonify({
            "total": len(outcomes),
            "by_subject": by_subject,
            "outcomes": outcomes
        })
        
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/outcomes/subjects', methods=['GET'])
@login_required
def get_outcome_subjects():
    """Mevcut ders adlarını ve sayılarını getir"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("""
            SELECT subject, COUNT(*) as count 
            FROM learning_outcomes 
            GROUP BY subject 
            ORDER BY count DESC
        """)
        subjects = cur.fetchall()
        return jsonify({"subjects": subjects})
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/outcomes/merge-subjects', methods=['POST'])
@login_required
def merge_outcome_subjects():
    """Seçilen ders adlarını birleştir"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    data = request.get_json()
    source_subjects = data.get('sources', [])  # Birleştirilecek kaynak ders adları
    target_subject = data.get('target', '')     # Hedef ders adı
    
    if not source_subjects or not target_subject:
        return jsonify({"error": "Kaynak ve hedef ders adları gerekli"}), 400
    
    conn = get_db()
    cur = conn.cursor()
    
    try:
        total_merged = 0
        
        for source in source_subjects:
            if source != target_subject:
                cur.execute("""
                    UPDATE learning_outcomes 
                    SET subject = %s 
                    WHERE subject = %s
                """, (target_subject, source))
                merged = cur.rowcount
                total_merged += merged
                logger.info(f"Merged {merged} outcomes: {source} -> {target_subject}")
        
        # Tekrar eden outcome_code'ları temizle
        cur.execute("""
            DELETE FROM learning_outcomes a
            USING learning_outcomes b
            WHERE a.id < b.id 
            AND UPPER(TRIM(a.outcome_code)) = UPPER(TRIM(b.outcome_code))
        """)
        duplicates_removed = cur.rowcount
        
        conn.commit()
        
        # Güncel durumu getir
        cur.execute("""
            SELECT subject, COUNT(*) as count 
            FROM learning_outcomes 
            GROUP BY subject 
            ORDER BY count DESC
        """)
        
        return jsonify({
            "success": True,
            "merged": total_merged,
            "duplicates_removed": duplicates_removed,
            "subjects": [{"subject": r[0], "count": r[1]} for r in cur.fetchall()]
        })
        
    except Exception as e:
        conn.rollback()
        logger.error(f"Merge subjects error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/outcomes/clear', methods=['DELETE'])
@login_required
def clear_outcomes():
    """Tüm kazanımları veya seçili dersleri sil"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor()
    
    try:
        # Query parametrelerinden silinecek dersleri al
        subjects = request.args.getlist('subjects')
        
        if subjects:
            # Seçili dersleri sil
            placeholders = ','.join(['%s'] * len(subjects))
            cur.execute(f"DELETE FROM learning_outcomes WHERE subject IN ({placeholders})", subjects)
        else:
            # Tüm kazanımları sil
            cur.execute("DELETE FROM learning_outcomes")
        
        deleted = cur.rowcount
        conn.commit()
        return jsonify({"success": True, "deleted": deleted})
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/outcomes/sync', methods=['POST'])
@login_required
def sync_outcome_texts():
    """Yüklenen kazanımlarla mevcut sınav verilerindeki kazanım isimlerini güncelle"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # 0. Önce sınıf seviyesine göre ders adlarını düzelt
        # 5, 6, 7. sınıf için inkilap -> sosyal
        cur.execute("""
            UPDATE report_card_answers rca
            SET subject = 'sosyal'
            FROM report_card_students rcs
            WHERE rca.student_result_id = rcs.id
            AND LEFT(rcs.class_name, 1) IN ('5', '6', '7')
            AND LOWER(rca.subject) = 'inkilap'
        """)
        subject_fixed = cur.rowcount
        
        # learning_outcomes tablosunda da düzelt
        cur.execute("""
            UPDATE learning_outcomes
            SET subject = 'sosyal'
            WHERE grade_level IN (5, 6, 7)
            AND LOWER(subject) IN ('inkilap', 'inkılap tarihi', 'inkilap tarihi', 'İnkılap Tarihi')
        """)
        
        # 1. Önce tam eşleşme dene
        cur.execute("""
            UPDATE report_card_answers rca
            SET outcome_text = lo.outcome_text
            FROM learning_outcomes lo
            WHERE UPPER(TRIM(rca.outcome_code)) = UPPER(TRIM(lo.outcome_code))
            AND (rca.outcome_text IS NULL OR rca.outcome_text = '' OR rca.outcome_text != lo.outcome_text)
        """)
        updated_count = cur.rowcount
        
        # 2. Nokta farkı olan eşleşmeleri dene (T.8.1.1 vs T.8.1.1.)
        cur.execute("""
            UPDATE report_card_answers rca
            SET outcome_text = lo.outcome_text
            FROM learning_outcomes lo
            WHERE UPPER(RTRIM(TRIM(rca.outcome_code), '.')) = UPPER(RTRIM(TRIM(lo.outcome_code), '.'))
            AND (rca.outcome_text IS NULL OR rca.outcome_text = '')
        """)
        updated_count += cur.rowcount
        
        cur.execute("SELECT COUNT(*) as total FROM learning_outcomes")
        total_outcomes = cur.fetchone()['total']
        
        # Eşleşme kontrolü - her iki format için
        cur.execute("""
            SELECT COUNT(DISTINCT rca.outcome_code) as matched
            FROM report_card_answers rca
            JOIN learning_outcomes lo ON 
                UPPER(RTRIM(TRIM(rca.outcome_code), '.')) = UPPER(RTRIM(TRIM(lo.outcome_code), '.'))
        """)
        matched = cur.fetchone()['matched']
        
        cur.execute("SELECT COUNT(DISTINCT outcome_code) as total FROM report_card_answers WHERE outcome_code IS NOT NULL AND TRIM(outcome_code) != ''")
        total_in_exams = cur.fetchone()['total']
        
        # Eşleşmeyen kazanımları PDF verilerinden al ve learning_outcomes'a ekle
        cur.execute("""
            SELECT DISTINCT 
                rca.outcome_code,
                rca.subject,
                rca.outcome_text
            FROM report_card_answers rca
            WHERE rca.outcome_code IS NOT NULL 
            AND TRIM(rca.outcome_code) != ''
            AND NOT EXISTS (
                SELECT 1 FROM learning_outcomes lo 
                WHERE UPPER(RTRIM(TRIM(rca.outcome_code), '.')) = UPPER(RTRIM(TRIM(lo.outcome_code), '.'))
            )
        """)
        unmatched = cur.fetchall()
        
        # Eksik kazanımları otomatik ekle
        auto_added = 0
        for row in unmatched:
            outcome_code = row['outcome_code'].strip() if row['outcome_code'] else None
            subject = row['subject'] if row['subject'] else 'bilinmiyor'
            outcome_text = row['outcome_text'] if row['outcome_text'] else outcome_code
            
            if not outcome_code:
                continue
            
            # Kazanım kodundan sınıf seviyesini tahmin et (T.8.1.1 -> 8, M.7.2.3 -> 7)
            grade_level = None
            import re
            grade_match = re.search(r'\.(\d)\.\d', outcome_code)
            if grade_match:
                grade_level = int(grade_match.group(1))
            
            try:
                cur.execute("""
                    INSERT INTO learning_outcomes (outcome_code, outcome_text, subject, grade_level)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (outcome_code) DO NOTHING
                """, (outcome_code, outcome_text, subject, grade_level))
                if cur.rowcount > 0:
                    auto_added += 1
            except Exception as e:
                logger.warning(f"Auto-add outcome failed for {outcome_code}: {e}")
        
        # Tekrar eşleşme sayısını hesapla
        cur.execute("SELECT COUNT(*) as total FROM learning_outcomes")
        total_outcomes = cur.fetchone()['total']
        
        cur.execute("""
            SELECT COUNT(DISTINCT rca.outcome_code) as matched
            FROM report_card_answers rca
            JOIN learning_outcomes lo ON 
                UPPER(RTRIM(TRIM(rca.outcome_code), '.')) = UPPER(RTRIM(TRIM(lo.outcome_code), '.'))
        """)
        matched = cur.fetchone()['matched']
        
        # Hala eşleşmeyen varsa örneklerini getir
        cur.execute("""
            SELECT DISTINCT rca.outcome_code 
            FROM report_card_answers rca
            WHERE rca.outcome_code IS NOT NULL 
            AND TRIM(rca.outcome_code) != ''
            AND NOT EXISTS (
                SELECT 1 FROM learning_outcomes lo 
                WHERE UPPER(RTRIM(TRIM(rca.outcome_code), '.')) = UPPER(RTRIM(TRIM(lo.outcome_code), '.'))
            )
            LIMIT 10
        """)
        unmatched_samples = [r['outcome_code'] for r in cur.fetchall()]
        
        conn.commit()
        
        return jsonify({
            "success": True,
            "updated_records": updated_count,
            "auto_added_outcomes": auto_added,
            "total_outcomes_in_db": total_outcomes,
            "matched_outcomes": matched,
            "total_outcomes_in_exams": total_in_exams,
            "unmatched_samples": unmatched_samples,
            "message": f"{updated_count} kayıt güncellendi. {auto_added} yeni kazanım eklendi. {matched}/{total_in_exams} kazanım eşleştirildi."
        })
        
    except Exception as e:
        conn.rollback()
        logger.error(f"Outcome sync error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


# ==================== ÇIFT KAYIT YÖNETİMİ ====================

@report_cards_bp.route('/api/student-duplicates', methods=['GET'])
@login_required
def get_student_duplicates():
    """Aynı öğrenci numarasına sahip farklı isimli kayıtları tespit et"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("""
            SELECT 
                LTRIM(student_no, '0') as student_no_normalized,
                array_agg(DISTINCT student_no) as no_variants,
                array_agg(DISTINCT student_name ORDER BY student_name) as name_variants,
                COUNT(DISTINCT student_name) as variant_count,
                COUNT(*) as total_records,
                array_agg(DISTINCT class_name) as classes
            FROM report_card_results 
            WHERE student_no IS NOT NULL AND TRIM(student_no) != ''
            GROUP BY LTRIM(student_no, '0')
            HAVING COUNT(DISTINCT student_no) > 1 OR COUNT(DISTINCT student_name) > 1
            ORDER BY LTRIM(student_no, '0')
        """)
        duplicates = cur.fetchall()
        
        return jsonify({
            "duplicates": duplicates,
            "total_duplicates": len(duplicates)
        })
        
    except Exception as e:
        logger.error(f"Duplicate detection error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/student-duplicates/merge', methods=['POST'])
@login_required
def merge_student_duplicates():
    """Aynı öğrenci numarasına sahip kayıtları tek bir isim ve numara altında birleştir"""
    if current_user.role != 'admin':
        return jsonify({"error": "Sadece admin bu işlemi yapabilir"}), 403
    
    data = request.get_json()
    student_no = data.get('student_no')
    canonical_name = data.get('canonical_name')
    canonical_no = data.get('canonical_no', student_no)
    
    if not student_no or not canonical_name:
        return jsonify({"error": "Öğrenci numarası ve isim gerekli"}), 400
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        student_no_normalized = student_no.lstrip('0') if student_no else ''
        
        cur.execute("""
            SELECT id, student_no, student_name, class_name
            FROM report_card_results 
            WHERE LTRIM(student_no, '0') = %s
        """, (student_no_normalized,))
        records = cur.fetchall()
        
        if len(records) == 0:
            return jsonify({"error": "Bu numaraya ait kayıt bulunamadı"}), 404
        
        cur.execute("""
            UPDATE report_card_results 
            SET student_name = %s, student_no = %s
            WHERE LTRIM(student_no, '0') = %s
        """, (canonical_name, canonical_no, student_no_normalized))
        updated_count = cur.rowcount
        
        conn.commit()
        
        return jsonify({
            "success": True,
            "student_no": canonical_no,
            "canonical_name": canonical_name,
            "updated_records": updated_count,
            "message": f"{updated_count} kayıt '{canonical_name}' ({canonical_no}) olarak güncellendi"
        })
        
    except Exception as e:
        conn.rollback()
        logger.error(f"Merge duplicates error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/student-duplicates/merge-all', methods=['POST'])
@login_required
def merge_all_student_duplicates():
    """Tüm çift kayıtları normalize et - baştaki sıfırları temizle ve en son ismi kullan"""
    if current_user.role != 'admin':
        return jsonify({"error": "Sadece admin bu işlemi yapabilir"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("""
            UPDATE report_card_results
            SET student_no = LTRIM(student_no, '0')
            WHERE student_no ~ '^0+[0-9]'
        """)
        no_updated = cur.rowcount
        
        cur.execute("""
            WITH latest_names AS (
                SELECT DISTINCT ON (LTRIM(student_no, '0'))
                    LTRIM(student_no, '0') as student_no_normalized,
                    student_name
                FROM report_card_results
                WHERE student_no IS NOT NULL AND TRIM(student_no) != ''
                ORDER BY LTRIM(student_no, '0'), id DESC
            )
            UPDATE report_card_results r
            SET student_name = ln.student_name
            FROM latest_names ln
            WHERE LTRIM(r.student_no, '0') = ln.student_no_normalized
            AND r.student_name != ln.student_name
        """)
        name_updated = cur.rowcount
        
        conn.commit()
        
        return jsonify({
            "success": True,
            "no_updated": no_updated,
            "name_updated": name_updated,
            "message": f"{no_updated} numara normalize edildi, {name_updated} isim güncellendi"
        })
        
    except Exception as e:
        conn.rollback()
        logger.error(f"Merge all duplicates error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


# ==================== KAYIT EŞLEŞTİRME ====================

@report_cards_bp.route('/api/match-students', methods=['POST'])
@login_required
def match_students():
    """Excel'den gelen öğrenci isimlerini sistemdeki kayıtlarla eşleştir"""
    if current_user.role != 'admin':
        return jsonify({"error": "Sadece admin bu işlemi yapabilir"}), 403
    
    if 'file' not in request.files:
        return jsonify({"error": "Dosya yüklenmedi"}), 400
    
    file = request.files['file']
    if not file.filename:
        return jsonify({"error": "Dosya seçilmedi"}), 400
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"error": "Sadece .xlsx veya .xls dosyası desteklenir"}), 400
    
    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active
        
        input_names = []
        name_column_idx = None
        
        # İlk satırdan başlayarak "ADI", "AD", "İSİM", "SOYADI" gibi başlık ara
        header_keywords = ['ADI', 'AD', 'İSİM', 'ISIM', 'AD SOYAD', 'ADI SOYADI', 'ÖĞRENCİ', 'OGRENCI', 'İSİM SOYİSİM']
        start_row = 1
        
        # İlk 5 satıra bak, başlık bul
        for row_idx in range(1, 6):
            row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0] if row_idx <= ws.max_row else None
            if row:
                for col_idx, cell in enumerate(row):
                    if cell:
                        cell_upper = str(cell).upper().strip()
                        for keyword in header_keywords:
                            if keyword in cell_upper:
                                name_column_idx = col_idx
                                start_row = row_idx + 1
                                break
                    if name_column_idx is not None:
                        break
            if name_column_idx is not None:
                break
        
        # Başlık bulunamadıysa ilk sütunu kullan, 2. satırdan başla
        if name_column_idx is None:
            name_column_idx = 0
            start_row = 2
        
        # İsimleri oku
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            if row and len(row) > name_column_idx:
                cell_value = row[name_column_idx]
                if cell_value:
                    name = str(cell_value).strip()
                    # Sadece sayı veya boş değilse ekle
                    if name and not name.replace('.', '').replace(',', '').isdigit():
                        input_names.append(name)
        
        if not input_names:
            return jsonify({"error": f"Excel'de isim bulunamadı. Dosyanın ilk sütununda (veya 'ADI', 'AD SOYAD' başlıklı sütunda) isimler olmalı."}), 400
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT DISTINCT student_name, student_no, class_name
            FROM report_card_results
            WHERE student_name IS NOT NULL AND TRIM(student_name) != ''
        """)
        db_students = cur.fetchall()
        
        def normalize_name(name):
            import unicodedata
            name = name.upper().strip()
            name = unicodedata.normalize('NFKD', name)
            tr_map = {'İ': 'I', 'Ş': 'S', 'Ğ': 'G', 'Ü': 'U', 'Ö': 'O', 'Ç': 'C'}
            for tr_char, en_char in tr_map.items():
                name = name.replace(tr_char, en_char)
            return ' '.join(name.split())
        
        def similarity(a, b):
            a_norm = normalize_name(a)
            b_norm = normalize_name(b)
            
            if a_norm == b_norm:
                return 1.0
            
            a_words = set(a_norm.split())
            b_words = set(b_norm.split())
            
            if not a_words or not b_words:
                return 0.0
            
            common = len(a_words & b_words)
            total = len(a_words | b_words)
            
            return common / total if total > 0 else 0.0
        
        results = []
        matched_count = 0
        partial_count = 0
        not_found_count = 0
        
        for input_name in input_names:
            best_match = None
            best_score = 0.0
            
            for student in db_students:
                score = similarity(input_name, student['student_name'])
                if score > best_score:
                    best_score = score
                    best_match = student
            
            if best_score >= 0.9:
                results.append({
                    'input_name': input_name,
                    'matched_name': best_match['student_name'],
                    'student_no': best_match['student_no'],
                    'class_name': best_match['class_name'],
                    'status': 'matched',
                    'score': best_score
                })
                matched_count += 1
            elif best_score >= 0.5:
                results.append({
                    'input_name': input_name,
                    'matched_name': best_match['student_name'],
                    'student_no': best_match['student_no'],
                    'class_name': best_match['class_name'],
                    'status': 'partial',
                    'score': best_score
                })
                partial_count += 1
            else:
                results.append({
                    'input_name': input_name,
                    'matched_name': None,
                    'student_no': None,
                    'class_name': None,
                    'status': 'not_found',
                    'score': 0
                })
                not_found_count += 1
        
        return jsonify({
            'results': results,
            'total': len(results),
            'matched': matched_count,
            'partial': partial_count,
            'not_found': not_found_count
        })
        
    except Exception as e:
        logger.error(f"Match students error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if 'cur' in dir():
            cur.close()
        if 'conn' in dir():
            conn.close()


@report_cards_bp.route('/api/match-students/download', methods=['POST'])
@login_required
def download_matching_results():
    """Eşleştirme sonuçlarını Excel olarak indir"""
    if current_user.role != 'admin':
        return jsonify({"error": "Sadece admin bu işlemi yapabilir"}), 403
    
    data = request.get_json()
    results = data.get('results', [])
    
    if not results:
        return jsonify({"error": "İndirilecek veri yok"}), 400
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Eşleştirme Sonuçları"
        
        headers = ['Sıra', 'Girilen İsim', 'Doğru İsim', 'Sınıf', 'Okul No', 'Durum']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        for idx, item in enumerate(results, 1):
            ws.cell(row=idx+1, column=1, value=idx)
            ws.cell(row=idx+1, column=2, value=item.get('input_name', ''))
            ws.cell(row=idx+1, column=3, value=item.get('matched_name', ''))
            ws.cell(row=idx+1, column=4, value=item.get('class_name', ''))
            ws.cell(row=idx+1, column=5, value=item.get('student_no', ''))
            
            status = item.get('status', '')
            if status == 'matched':
                status_text = 'Tam Eşleşme'
            elif status == 'partial':
                status_text = 'Benzer'
            else:
                status_text = 'Bulunamadı'
            ws.cell(row=idx+1, column=6, value=status_text)
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'eslestirme_sonuclari_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        )
        
    except Exception as e:
        logger.error(f"Download matching results error: {e}")
        return jsonify({"error": str(e)}), 500


# ==================== EKSİK KAYIT TESPİT VE TAMAMLAMA ====================

@report_cards_bp.route('/api/incomplete-exams', methods=['GET'])
@login_required
def get_incomplete_exams():
    """Eksik öğrenci kaydı olan sınavları listele"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # expected_student_count vs student_count karşılaştırması
        cur.execute("""
            SELECT 
                id, exam_name, class_name, grade_level,
                expected_student_count, student_count, processed_student_count,
                pdf_filename, parse_status, created_at,
                COALESCE(expected_student_count, 0) - COALESCE(student_count, 0) as missing_count
            FROM report_cards
            WHERE parse_status = 'completed'
            AND expected_student_count IS NOT NULL
            AND expected_student_count > COALESCE(student_count, 0)
            ORDER BY created_at DESC
        """)
        incomplete = cur.fetchall()
        
        return jsonify({
            "incomplete_exams": incomplete,
            "total": len(incomplete)
        })
        
    except Exception as e:
        logger.error(f"Incomplete exams error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/complete-exam/<int:report_card_id>', methods=['POST'])
@login_required
def complete_exam_students(report_card_id):
    """Mevcut sınava PDF yeniden yükleyerek eksik öğrencileri tamamla"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    if 'pdf_file' not in request.files:
        return jsonify({"error": "PDF dosyası gerekli"}), 400
    
    pdf_file = request.files['pdf_file']
    if pdf_file.filename == '':
        return jsonify({"error": "Dosya seçilmedi"}), 400
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Mevcut sınavı kontrol et
        cur.execute("SELECT * FROM report_cards WHERE id = %s", (report_card_id,))
        exam = cur.fetchone()
        
        if not exam:
            return jsonify({"error": "Sınav bulunamadı"}), 404
        
        # Mevcut öğrenci sayısını al
        cur.execute("""
            SELECT COUNT(*) as current_count FROM report_card_students WHERE report_card_id = %s
        """, (report_card_id,))
        current_count = cur.fetchone()['current_count']
        
        # PDF'i parse et
        pdf_content = pdf_file.read()
        parsed_data = parse_karne_pdf(pdf_content)
        
        if not parsed_data or 'students' not in parsed_data:
            return jsonify({"error": "PDF parse edilemedi"}), 400
        
        students = parsed_data['students']
        added_count = 0
        updated_count = 0
        skipped_count = 0
        
        for student_data in students:
            student_no = student_data.get('student_no', '').strip()
            student_name = student_data.get('student_name', '').strip()
            
            if not student_name:
                continue
            
            # Öğrenci numarasına göre kontrol et
            if student_no:
                cur.execute("""
                    SELECT id FROM report_card_students 
                    WHERE report_card_id = %s AND TRIM(student_no) = TRIM(%s)
                """, (report_card_id, student_no))
            else:
                cur.execute("""
                    SELECT id FROM report_card_students 
                    WHERE report_card_id = %s AND TRIM(student_name) = TRIM(%s)
                """, (report_card_id, student_name))
            
            existing = cur.fetchone()
            
            if existing:
                skipped_count += 1
                continue
            
            # Yeni öğrenci ekle
            cur.execute("""
                INSERT INTO report_card_students (
                    report_card_id, student_name, student_no, class_name,
                    total_correct, total_wrong, total_blank, total_net,
                    success_rate, lgs_score, percentile, class_rank, school_rank
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                report_card_id, student_name, student_no,
                student_data.get('class_name', exam['class_name']),
                student_data.get('total_correct'), student_data.get('total_wrong'),
                student_data.get('total_blank'), student_data.get('total_net'),
                student_data.get('success_rate'), student_data.get('lgs_score'),
                student_data.get('percentile'), student_data.get('class_rank'),
                student_data.get('school_rank')
            ))
            student_result_id = cur.fetchone()['id']
            added_count += 1
            
            # Ders sonuçları
            for subject, subj_data in student_data.get('subjects', {}).items():
                cur.execute("""
                    INSERT INTO report_card_subjects (
                        student_result_id, subject, question_count, correct_count,
                        wrong_count, blank_count, net_score, success_rate,
                        correct_answers, student_answers
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    student_result_id, subject,
                    subj_data.get('question_count'), subj_data.get('correct_count'),
                    subj_data.get('wrong_count'), subj_data.get('blank_count'),
                    subj_data.get('net_score'), subj_data.get('success_rate'),
                    subj_data.get('correct_answers'), subj_data.get('student_answers')
                ))
            
            # Cevaplar
            for answer in student_data.get('answers', []):
                cur.execute("""
                    INSERT INTO report_card_answers (
                        student_result_id, subject, question_number,
                        correct_answer, student_answer, is_correct, is_blank,
                        outcome_code, outcome_text
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    student_result_id, answer.get('subject'),
                    answer.get('question_number'), answer.get('correct_answer'),
                    answer.get('student_answer'), answer.get('is_correct'),
                    answer.get('is_blank'), answer.get('outcome_code'),
                    answer.get('outcome_text')
                ))
        
        # Öğrenci sayısını güncelle
        cur.execute("""
            UPDATE report_cards 
            SET student_count = student_count + %s,
                parse_error = COALESCE(parse_error, '') || ' | Tamamlama: +' || %s || ' öğrenci'
            WHERE id = %s
        """, (added_count, added_count, report_card_id))
        
        conn.commit()
        
        return jsonify({
            "success": True,
            "added": added_count,
            "skipped": skipped_count,
            "message": f"{added_count} yeni öğrenci eklendi, {skipped_count} zaten mevcuttu"
        })
        
    except Exception as e:
        conn.rollback()
        logger.error(f"Reprocess error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/debug')
@login_required
def admin_debug_page():
    """Production debug/log sayfası - Veritabanı durumu ve hatalar"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        stats = {}
        
        cur.execute("SELECT COUNT(*) as total FROM report_cards")
        stats['total_exams'] = cur.fetchone()['total']
        
        cur.execute("SELECT parse_status, COUNT(*) as count FROM report_cards GROUP BY parse_status")
        stats['status_counts'] = {r['parse_status']: r['count'] for r in cur.fetchall()}
        
        cur.execute("SELECT COUNT(*) as total FROM report_card_students")
        stats['total_students'] = cur.fetchone()['total']
        
        cur.execute("SELECT COUNT(*) as total FROM report_card_answers")
        stats['total_answers'] = cur.fetchone()['total']
        
        cur.execute("""
            SELECT id, exam_name, class_name, parse_status, parse_error, 
                   student_count, expected_student_count, processed_student_count,
                   pdf_filename, pdf_storage_path, created_at
            FROM report_cards 
            ORDER BY created_at DESC 
            LIMIT 20
        """)
        recent_exams_raw = cur.fetchall()
        
        recent_exams = []
        for exam in recent_exams_raw:
            exam_dict = dict(exam)
            local_path = os.path.join(UPLOAD_FOLDER, exam['pdf_filename']) if exam['pdf_filename'] else None
            exam_dict['local_exists'] = os.path.exists(local_path) if local_path else False
            exam_dict['storage_exists'] = bool(exam['pdf_storage_path'])
            recent_exams.append(exam_dict)
        
        cur.execute("""
            SELECT id, exam_name, parse_error, created_at 
            FROM report_cards 
            WHERE parse_error IS NOT NULL AND parse_error != ''
            ORDER BY created_at DESC 
            LIMIT 20
        """)
        errors = cur.fetchall()
        
        sample_outcomes = []
        cur.execute("""
            SELECT rca.id, rca.subject, rca.question_number, 
                   rca.correct_answer, rca.student_answer, rca.is_correct,
                   rca.outcome_code, rcs.student_name, rc.exam_name
            FROM report_card_answers rca
            JOIN report_card_students rcs ON rca.student_result_id = rcs.id
            JOIN report_cards rc ON rcs.report_card_id = rc.id
            ORDER BY rca.id DESC
            LIMIT 50
        """)
        sample_outcomes = cur.fetchall()
        
        upload_folder_exists = os.path.exists(UPLOAD_FOLDER)
        upload_folder_files = os.listdir(UPLOAD_FOLDER) if upload_folder_exists else []
        
        # Object Storage PDF sayısı
        storage_pdf_count = 0
        storage_pdf_files = []
        if object_storage and object_storage.enabled:
            try:
                storage_objects = list(object_storage.client.list(prefix='report_cards/'))
                storage_pdf_files = [o.name for o in storage_objects if o.name.endswith('.pdf')]
                storage_pdf_count = len(storage_pdf_files)
            except Exception as e:
                logger.warning(f"Object Storage liste hatası: {e}")
        
        return render_template('admin_debug.html',
                             stats=stats,
                             recent_exams=recent_exams,
                             errors=errors,
                             sample_outcomes=sample_outcomes,
                             upload_folder_exists=upload_folder_exists,
                             upload_folder_files=upload_folder_files[:20],
                             storage_pdf_count=storage_pdf_count,
                             storage_pdf_files=storage_pdf_files[:20])
        
    except Exception as e:
        logger.error(f"Debug page error: {e}")
        return f"Hata: {str(e)}", 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/fix-student-numbers', methods=['GET', 'POST'])
@login_required
def fix_student_numbers():
    """Öğrenci numaralarındaki baştaki sıfırları temizle"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("""
            UPDATE report_card_results 
            SET student_no = LTRIM(student_no, '0')
            WHERE student_no LIKE '0%' AND student_no != '0'
        """)
        updated_results = cur.rowcount
        
        cur.execute("""
            UPDATE report_card_students 
            SET student_no = LTRIM(student_no, '0')
            WHERE student_no LIKE '0%' AND student_no != '0'
        """)
        updated_students = cur.rowcount
        
        conn.commit()
        
        return jsonify({
            "success": True,
            "message": f"{updated_results} sonuç ve {updated_students} öğrenci kaydı güncellendi"
        })
    except Exception as e:
        conn.rollback()
        logger.error(f"Öğrenci numarası düzeltme hatası: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/debug-stats')
@login_required
def get_debug_stats():
    """JSON formatında debug istatistikleri"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("""
            SELECT 
                (SELECT COUNT(*) FROM report_cards) as total_exams,
                (SELECT COUNT(*) FROM report_cards WHERE parse_status = 'completed') as completed,
                (SELECT COUNT(*) FROM report_cards WHERE parse_status = 'processing') as processing,
                (SELECT COUNT(*) FROM report_cards WHERE parse_status = 'failed') as failed,
                (SELECT COUNT(*) FROM report_card_students) as total_students,
                (SELECT COUNT(*) FROM report_card_answers) as total_answers,
                (SELECT COUNT(*) FROM report_card_answers WHERE is_correct = true) as correct_answers,
                (SELECT COUNT(*) FROM report_card_answers WHERE is_correct = false) as wrong_answers
        """)
        stats = cur.fetchone()
        
        cur.execute("""
            SELECT id, exam_name, class_name, parse_status, parse_error,
                   student_count, expected_student_count, processed_student_count
            FROM report_cards 
            ORDER BY created_at DESC 
            LIMIT 10
        """)
        recent = cur.fetchall()
        
        return jsonify({
            "stats": stats,
            "recent_exams": recent,
            "timestamp": datetime.now().isoformat()
        })
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/migrate-fmt-to-report', methods=['POST'])
@login_required
def migrate_fmt_to_report():
    """Mevcut FMT/optical verilerini report_card tablolarına aktar"""
    if current_user.role not in ['admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    data = request.get_json() or {}
    report_card_id = data.get('report_card_id')
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # FMT sınavlarını bul
        if report_card_id:
            cur.execute("SELECT id, exam_name FROM report_cards WHERE id = %s", (report_card_id,))
        else:
            cur.execute("SELECT id, exam_name FROM report_cards WHERE class_name = 'FMT'")
        
        exams = cur.fetchall()
        
        total_migrated = 0
        total_answers = 0
        
        for exam in exams:
            exam_id = exam['id']
            
            # Optical sonuçları al
            cur.execute("""
                SELECT osr.id, osr.student_id, osr.booklet_type, osr.raw_answers, osr.results,
                       osr.total_correct, osr.total_wrong, osr.total_empty, osr.total_net
                FROM optical_student_results osr
                WHERE osr.optical_exam_id = %s
            """, (exam_id,))
            
            optical_results = cur.fetchall()
            
            # Cevap anahtarını bul
            results_sample = optical_results[0]['results'] if optical_results else {}
            grade = results_sample.get('grade', 5) if results_sample else 5
            
            cur.execute("""
                SELECT answer_key_a, answer_key_b, kazanimlar 
                FROM fmt_answer_keys 
                WHERE grade = %s ORDER BY created_at DESC LIMIT 1
            """, (str(grade),))
            ak_row = cur.fetchone()
            
            answer_key_a = ak_row.get('answer_key_a', {}) if ak_row else {}
            answer_key_b = ak_row.get('answer_key_b', {}) if ak_row else {}
            kazanimlar = ak_row.get('kazanimlar', {}) if ak_row else {}
            
            for osr in optical_results:
                results = osr['results'] or {}
                student_name = results.get('student_name', '')
                student_no = results.get('student_no', '')
                class_name = results.get('class_name', '')
                booklet_type = osr['booklet_type'] or 'A'
                
                # Cevapları results->subjects->student_answers'dan al
                subjects_data = results.get('subjects', {})
                raw_answers = {}
                for subj, subj_info in subjects_data.items():
                    if isinstance(subj_info, dict):
                        raw_answers[subj] = subj_info.get('student_answers', '')
                
                # Zaten aktarılmış mı kontrol et
                cur.execute("""
                    SELECT id FROM report_card_students 
                    WHERE report_card_id = %s AND student_no = %s
                """, (exam_id, student_no))
                
                if cur.fetchone():
                    continue  # Zaten var
                
                # report_card_students'a ekle
                cur.execute("""
                    INSERT INTO report_card_students (
                        report_card_id, user_id, student_name, student_no, class_name,
                        total_questions, total_correct, total_wrong, total_blank, total_net, success_rate
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                """, (
                    exam_id,
                    osr['student_id'],
                    student_name,
                    student_no,
                    class_name,
                    osr['total_correct'] + osr['total_wrong'] + osr['total_empty'],
                    osr['total_correct'],
                    osr['total_wrong'],
                    osr['total_empty'],
                    osr['total_net'],
                    round((osr['total_correct'] / max(1, osr['total_correct'] + osr['total_wrong'] + osr['total_empty'])) * 100, 2)
                ))
                student_result_id = cur.fetchone()['id']
                total_migrated += 1
                
                # Cevapları işle ve report_card_answers'a ekle
                for subj, answers_str in raw_answers.items():
                    if not answers_str:
                        continue
                    
                    # Cevap anahtarı
                    if booklet_type == 'A':
                        answer_string = answer_key_a.get(subj, '') or ''
                    else:
                        answer_string = answer_key_b.get(subj, '') or ''
                    
                    subj_kazanimlar = kazanimlar.get(subj, []) or []
                    
                    for i, char in enumerate(answers_str):
                        q_num = i + 1
                        student_ans = char.upper() if char and char.strip() else ''
                        
                        correct_ans = ''
                        kazanim = ''
                        
                        if booklet_type == 'A':
                            if i < len(answer_string):
                                correct_ans = answer_string[i].upper()
                            for kz in subj_kazanimlar:
                                if kz.get('soru') == q_num:
                                    kazanim = kz.get('kazanim', '')
                                    break
                        else:
                            for kz in subj_kazanimlar:
                                if kz.get('b_soru') == q_num:
                                    correct_ans = kz.get('a_cevap', '').upper()
                                    kazanim = kz.get('kazanim', '')
                                    break
                        
                        is_blank = (student_ans == '' or student_ans == ' ' or student_ans == '*')
                        is_correct = (not is_blank and correct_ans and student_ans == correct_ans)
                        
                        cur.execute("""
                            INSERT INTO report_card_answers (
                                student_result_id, subject, question_number,
                                correct_answer, student_answer, is_correct, is_blank,
                                outcome_code, outcome_text
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            student_result_id,
                            subj,
                            q_num,
                            correct_ans,
                            student_ans,
                            is_correct,
                            is_blank,
                            kazanim,
                            kazanim
                        ))
                        total_answers += 1
            
            conn.commit()
        
        return jsonify({
            "success": True,
            "migrated_students": total_migrated,
            "migrated_answers": total_answers,
            "message": f"{total_migrated} öğrenci ve {total_answers} cevap aktarıldı"
        })
        
    except Exception as e:
        conn.rollback()
        logger.error(f"FMT migration error: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


# ==================== GÖRSEL YÜKLEME SİSTEMİ ====================

@report_cards_bp.route('/image-upload')
@login_required
def image_upload_page():
    """Görsel yükleme sayfası"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    return render_template('admin_image_upload.html')


@report_cards_bp.route('/api/upload-image', methods=['POST'])
@login_required
def upload_image_report():
    """Görsel veya PDF dosyasından sınav sonuçlarını parse et (Gemini AI)"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    if 'images' not in request.files:
        return jsonify({"error": "Dosya bulunamadı"}), 400
    
    files = request.files.getlist('images')
    if not files or files[0].filename == '':
        return jsonify({"error": "Dosya seçilmedi"}), 400
    
    exam_name = request.form.get('exam_name', 'Gemini Okuma')
    
    all_students = []
    errors = []
    parser = ImageReportParser()
    
    for file in files:
        if file and file.filename:
            filename = file.filename.lower()
            file_data = file.read()
            
            try:
                logger.info(f"Dosya işleniyor: {filename}, boyut: {len(file_data)} bytes")
                if filename.endswith('.pdf'):
                    logger.info(f"PDF işleniyor: {filename}")
                    result = parser.parse_pdf(file_data, first_page_only=True)
                    logger.info(f"PDF sonuç: {result}")
                elif filename.endswith(('.png', '.jpg', '.jpeg', '.webp')):
                    mime_type = 'image/png' if filename.endswith('.png') else 'image/jpeg'
                    if filename.endswith('.webp'):
                        mime_type = 'image/webp'
                    result = parser.parse_image(file_data, mime_type)
                else:
                    errors.append(f"{file.filename}: Desteklenmeyen format")
                    continue
                
                if result.get('success'):
                    all_students.extend(result.get('students', []))
                else:
                    errors.append(f"{file.filename}: {result.get('error', 'Bilinmeyen hata')}")
            except Exception as e:
                errors.append(f"{file.filename}: {str(e)}")
    
    if not all_students and errors:
        return jsonify({"error": "; ".join(errors)}), 400
    
    return jsonify({
        "success": True,
        "students": all_students,
        "student_count": len(all_students),
        "exam_name": exam_name,
        "errors": errors
    })


@report_cards_bp.route('/api/save-image-results', methods=['POST'])
@login_required
def save_image_results():
    """Parse edilen CSV/Excel sonuçlarını veritabanına kaydet"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    data = request.json
    if not data:
        return jsonify({"error": "Veri bulunamadı"}), 400
    
    students = data.get('students', [])
    exam_name = data.get('exam_name', 'CSV Yükleme')
    
    if not students:
        return jsonify({"error": "Öğrenci verisi bulunamadı"}), 400
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        saved_count = 0
        errors = []
        
        import re
        grade_match = re.match(r'^(\d+)', exam_name)
        exam_grade_level = int(grade_match.group(1)) if grade_match else None
        
        cur.execute("""
            INSERT INTO report_card_exams (exam_name, exam_date, upload_date, source_type, created_by, grade_level)
            VALUES (%s, %s, NOW(), 'csv', %s, %s)
            RETURNING id
        """, (exam_name, datetime.now().strftime('%Y-%m-%d'), current_user.id, exam_grade_level))
        exam_id = cur.fetchone()['id']
        conn.commit()
        
        for student in students:
            try:
                student_name = student.get('name', '').strip()
                class_name = student.get('class_name', '').strip()
                grade = student.get('grade', 5)
                student_no = str(student.get('student_no', '')).lstrip('0') or '0'
                booklet_type = student.get('booklet_type', 'A')
                
                if not student_name and not student_no:
                    errors.append(f"Öğrenci adı veya numarası boş")
                    continue
                
                user_id = None
                if student_no:
                    cur.execute("""
                        SELECT id FROM users WHERE role = 'student' AND username = %s
                    """, (student_no,))
                    user = cur.fetchone()
                    if user:
                        user_id = user['id']
                
                totals = student.get('totals', {})
                subjects = student.get('subjects', {})
                
                cur.execute("""
                    INSERT INTO report_card_results (
                        exam_id, student_id, student_no, student_name, class_name,
                        grade_level, booklet_type, subjects, totals
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                """, (
                    exam_id, user_id, student_no, student_name, class_name,
                    grade, booklet_type, json.dumps(subjects), json.dumps(totals)
                ))
                
                conn.commit()
                saved_count += 1
                
            except Exception as e:
                conn.rollback()
                logger.error(f"Öğrenci kayıt hatası: {e}")
                errors.append(f"{student_name}: {str(e)}")
        
        return jsonify({
            "success": True,
            "saved_count": saved_count,
            "exam_id": exam_id,
            "errors": errors if errors else None
        })
        
    except Exception as e:
        conn.rollback()
        logger.error(f"Kayıt hatası: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/save-image-results-old', methods=['POST'])
@login_required
def save_image_results_old():
    """Eski format - Parse edilen görsel sonuçlarını veritabanına kaydet"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    data = request.json
    if not data:
        return jsonify({"error": "Veri bulunamadı"}), 400
    
    students = data.get('students', [])
    exam_name = data.get('exam_name', 'Görsel Yükleme')
    
    if not students:
        return jsonify({"error": "Öğrenci verisi bulunamadı"}), 400
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        saved_count = 0
        errors = []
        
        cur.execute("""
            INSERT INTO report_cards (exam_name, class_name, upload_date, parse_status)
            VALUES (%s, '', NOW(), 'completed')
            RETURNING id
        """, (exam_name,))
        report_card_id = cur.fetchone()['id']
        conn.commit()
        
        for student in students:
            try:
                student_name = student.get('name', '').strip()
                class_name = student.get('class_name', '').strip()
                grade = student.get('grade', 8)
                
                if not student_name or not class_name:
                    errors.append(f"Ogrenci adi veya sinif bos: {student_name}")
                    continue
                
                totals = student.get('totals', {})
                
                cur.execute("""
                    INSERT INTO report_card_students (
                        report_card_id, student_name, class_name, grade_level, student_no,
                        total_correct, total_wrong, total_blank, total_net, success_rate
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                """, (
                    report_card_id, student_name, class_name, grade, student.get('student_no', ''),
                    totals.get('correct_count', 0),
                    totals.get('wrong_count', 0),
                    totals.get('blank_count', 0),
                    totals.get('net_score', 0),
                    totals.get('success_rate', 0)
                ))
                
                conn.commit()
                saved_count += 1
                
            except Exception as e:
                conn.rollback()
                logger.error(f"Öğrenci kayıt hatası: {e}")
                errors.append(f"{student_name}: {str(e)}")
        
        return jsonify({
            "success": True,
            "saved_count": saved_count,
            "errors": errors if errors else None
        })
        
    except Exception as e:
        conn.rollback()
        logger.error(f"Kayıt hatası: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/save-image-results-legacy', methods=['POST'])
@login_required
def save_image_results_legacy():
    """Legacy format için"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    data = request.json
    if not data:
        return jsonify({"error": "Veri bulunamadı"}), 400
    
    students = data.get('students', [])
    exam_name = data.get('exam_name', 'Görsel Yükleme')
    
    if not students:
        return jsonify({"error": "Öğrenci verisi bulunamadı"}), 400
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        saved_count = 0
        errors = []
        
        for student in students:
            try:
                student_name = student.get('name', '').strip()
                class_name = student.get('class_name', '').strip()
                grade = student.get('grade', 8)
                
                if not student_name or not class_name:
                    errors.append(f"Ogrenci adi veya sinif bos: {student_name}")
                    continue
                
                cur.execute("""
                    SELECT id FROM report_card_students 
                    WHERE student_name = %s AND class_name = %s
                """, (student_name, class_name))
                existing = cur.fetchone()
                
                totals = student.get('totals', {})
                
                if existing:
                    cur.execute("""
                        UPDATE report_card_students SET
                            total_correct = %s, total_wrong = %s, total_blank = %s,
                            total_net = %s, success_rate = %s
                        WHERE id = %s
                    """, (
                        totals.get('correct_count', 0),
                        totals.get('wrong_count', 0),
                        totals.get('blank_count', 0),
                        totals.get('net_score', 0),
                        totals.get('success_rate', 0),
                        existing['id']
                    ))
                else:
                    cur.execute("""
                        INSERT INTO report_card_students (
                            student_name, class_name, grade_level, student_no,
                            total_correct, total_wrong, total_blank, total_net, success_rate
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        student_name, class_name, grade, student.get('student_no', ''),
                        totals.get('correct_count', 0),
                        totals.get('wrong_count', 0),
                        totals.get('blank_count', 0),
                        totals.get('net_score', 0),
                        totals.get('success_rate', 0)
                    ))
                
                conn.commit()
                saved_count += 1
                
            except Exception as e:
                conn.rollback()
                logger.error(f"Öğrenci kayıt hatası: {e}")
                errors.append(f"{student_name}: {str(e)}")
        
        return jsonify({
            "success": True,
            "saved_count": saved_count,
            "errors": errors if errors else None
        })
        
    except Exception as e:
        conn.rollback()
        logger.error(f"Kayıt hatası: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/save-image-results-detailed', methods=['POST'])
@login_required
def save_image_results_detailed():
    """Detaylı kayıt format - UNUSED"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    data = request.json
    if not data:
        return jsonify({"error": "Veri bulunamadı"}), 400
    
    students = data.get('students', [])
    exam_name = data.get('exam_name', 'Görsel Yükleme')
    
    if not students:
        return jsonify({"error": "Öğrenci verisi bulunamadı"}), 400
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        saved_count = 0
        errors = []
        
        import re
        grade_match = re.match(r'^(\d+)', exam_name)
        exam_grade_level = int(grade_match.group(1)) if grade_match else None
        
        cur.execute("""
            INSERT INTO report_card_exams (exam_name, exam_date, upload_date, source_type, grade_level)
            VALUES (%s, %s, NOW(), 'image', %s)
            RETURNING id
        """, (exam_name, datetime.now().strftime('%Y-%m-%d'), exam_grade_level))
        exam_id = cur.fetchone()['id']
        
        for student in students:
            try:
                student_name = student.get('name', '').strip()
                class_name = student.get('class_name', '').strip()
                grade = student.get('grade', 8)
                
                if not student_name or not class_name:
                    errors.append(f"Ogrenci adi veya sinif bos: {student_name}")
                    continue
                
                cur.execute("""
                    SELECT id FROM report_card_students 
                    WHERE student_name = %s AND class_name = %s
                """, (student_name, class_name))
                existing = cur.fetchone()
                
                if existing:
                    student_id = existing['id']
                else:
                    cur.execute("""
                        INSERT INTO report_card_students (student_name, class_name, grade_level, student_no)
                        VALUES (%s, %s, %s, %s)
                        RETURNING id
                    """, (student_name, class_name, grade, student.get('student_no', '')))
                    student_id = cur.fetchone()['id']
                
                totals = student.get('totals', {})
                cur.execute("""
                    INSERT INTO report_card_results (
                        student_id, exam_id, total_questions, total_correct,
                        total_wrong, total_blank, total_net, total_score, success_rate
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    RETURNING id
                """, (
                    student_id, exam_id,
                    totals.get('question_count', 0),
                    totals.get('correct_count', 0),
                    totals.get('wrong_count', 0),
                    totals.get('blank_count', 0),
                    totals.get('net_score', 0),
                    totals.get('total_score', 0),
                    totals.get('success_rate', 0)
                ))
                result_id = cur.fetchone()['id']
                
                for subject_key, subject_data in student.get('subjects', {}).items():
                    cur.execute("""
                        INSERT INTO report_card_subjects (
                            result_id, subject_name, question_count, correct_count,
                            wrong_count, blank_count, net_score, success_rate
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        result_id,
                        subject_key,
                        subject_data.get('question_count', 0),
                        subject_data.get('correct_count', 0),
                        subject_data.get('wrong_count', 0),
                        subject_data.get('blank_count', 0),
                        subject_data.get('net_score', 0),
                        subject_data.get('success_rate', 0)
                    ))
                
                for answer in student.get('answers', []):
                    cur.execute("""
                        INSERT INTO report_card_answers (
                            student_result_id, subject, question_number,
                            correct_answer, student_answer, is_correct, is_blank
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """, (
                        result_id,
                        answer.get('subject', ''),
                        answer.get('question_number', 0),
                        answer.get('correct_answer', ''),
                        answer.get('student_answer', ''),
                        answer.get('is_correct', False),
                        answer.get('is_blank', False)
                    ))
                
                saved_count += 1
                
            except Exception as e:
                errors.append(f"{student.get('name', 'Bilinmeyen')}: {str(e)}")
                logger.error(f"Öğrenci kayıt hatası: {e}")
        
        conn.commit()
        
        return jsonify({
            "success": True,
            "saved_count": saved_count,
            "errors": errors,
            "message": f"{saved_count} öğrenci kaydedildi"
        })
        
    except Exception as e:
        conn.rollback()
        logger.error(f"Kayıt hatası: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/upload-csv', methods=['POST'])
@login_required
def upload_csv_report():
    """CSV/Excel dosyasından sınav sonuçlarını parse et"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    if 'file' not in request.files:
        return jsonify({"error": "Dosya bulunamadı"}), 400
    
    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({"error": "Dosya seçilmedi"}), 400
    
    filename = file.filename.lower()
    file_data = file.read()
    
    fetch_student_names = request.form.get('fetch_student_names', '0') == '1'
    grade_level = request.form.get('grade_level', '8')
    
    try:
        grade_level = int(grade_level)
    except:
        grade_level = 8
    
    answer_key_a = {}
    answer_key_b = {}
    
    if 'answer_key' in request.files:
        ak_file = request.files['answer_key']
        if ak_file.filename:
            answer_key_a, answer_key_b = parse_answer_key_with_outcomes(ak_file, grade_level)
    
    parser = CSVExcelParser(answer_key={'A': answer_key_a, 'B': answer_key_b})
    
    try:
        if filename.endswith('.csv') or filename.endswith('.txt'):
            result = parser.parse_csv(file_data)
        elif filename.endswith(('.xlsx', '.xls')):
            result = parser.parse_excel(file_data)
        else:
            return jsonify({"error": "Desteklenmeyen format. CSV, TXT veya Excel dosyası yükleyin."}), 400
        
        if result.get('success'):
            students = result.get('students', [])
            
            if fetch_student_names:
                students = fetch_student_names_from_db(students)
            
            return jsonify({
                "success": True,
                "students": students,
                "student_count": len(students),
                "errors": result.get('errors')
            })
        else:
            return jsonify({"error": result.get('error', 'Bilinmeyen hata')}), 400
            
    except Exception as e:
        logger.error(f"CSV/Excel parse hatası: {e}")
        return jsonify({"error": str(e)}), 500


def parse_answer_key_with_outcomes(file, grade_level=8):
    """Kazanımlı Excel cevap anahtarını parse et - tek dosyadan A ve B kitapçık
    Format: Her ders başlığı altında satırlar
    Sütunlar: Soru No | Kazanım | A (cevap) | B | C | D | G (B kitapçık karşılığı)
    
    Sınıf seviyesine göre soru sayıları:
    - 5-6. sınıf: Türkçe 15, Mat 10, Fen 10, Sosyal 10, İng 15, Din 15
    - 7-8. sınıf: Türkçe 20, Mat 10, Fen 10, Sosyal/İnkılap 10, İng 20, Din 20
    - 8. sınıf: Sosyal Bilgiler yerine İnkılap Tarihi
    
    Returns: (answer_key_a, answer_key_b)
    """
    if grade_level in [5, 6]:
        SUBJECT_QUESTION_COUNTS = {
            'turkce': 15, 'matematik': 15, 'fen': 15,
            'sosyal': 10, 'ingilizce': 10, 'din': 10
        }
    else:
        SUBJECT_QUESTION_COUNTS = {
            'turkce': 20, 'matematik': 20, 'fen': 20,
            'sosyal': 10, 'inkilap': 10, 'ingilizce': 10, 'din': 10
        }
    
    logger.info(f"Sınıf {grade_level} için soru dağılımı: {SUBJECT_QUESTION_COUNTS}")
    
    try:
        import pandas as pd
        df = pd.read_excel(file, header=None)
        
        answer_key_a = {}
        answer_key_b = {}
        
        # Pattern'ler normalize edilmiş (ı->i, i̇->i) - first_cell de normalize edildi
        subject_patterns = [
            ('turkce', 'turkce'),
            ('din kulturu ve ahlak bilgisi', 'din'),
            ('din kulturu ve ahlak', 'din'),
            ('din kulturu', 'din'),
            ('sosyal bilgiler', 'sosyal'),
            # İnkılap Tarihi - normalize edilmiş
            ('t.c. inkilap tarihi ve ataturkculuk', 'inkilap'),
            ('t.c. inkilap tarihi ve ataturkcu', 'inkilap'),
            ('inkilap tarihi ve ataturkculuk', 'inkilap'),
            ('inkilap tarihi ve ataturkcu', 'inkilap'),
            ('inkilap tarihi', 'inkilap'),
            ('t.c. inkilap', 'inkilap'),
            ('inkilap', 'inkilap'),
            ('ataturkculuk', 'inkilap'),
            ('ataturkcu', 'inkilap'),
            # İngilizce
            ('ingilizce', 'ingilizce'),
            ('english', 'ingilizce'),
            ('yabanci dil', 'ingilizce'),
            # Matematik
            ('matematik', 'matematik'),
            # Fen
            ('fen bilimleri', 'fen'),
            ('fen', 'fen'),
            # Sosyal
            ('sosyal', 'sosyal'),
        ]
        
        current_subject = None
        questions_a = []
        
        def normalize_turkish(text):
            """Türkçe karakterleri normalize et - büyük/küçük harf sorunlarını çöz"""
            text = text.lower()
            # Combining dot above karakterini kaldır (İ -> i̇ sorununu çözer)
            text = text.replace('i̇', 'i')
            # Diğer Türkçe karakter varyasyonlarını normalize et
            text = text.replace('ı', 'i')  # noktasız ı -> i
            text = text.replace('ü', 'u')
            text = text.replace('ö', 'o')
            text = text.replace('ş', 's')
            text = text.replace('ğ', 'g')
            text = text.replace('ç', 'c')
            return text
        
        for idx, row in df.iterrows():
            first_cell = normalize_turkish(str(row.iloc[0]).strip()) if pd.notna(row.iloc[0]) else ''
            second_cell = normalize_turkish(str(row.iloc[1]).strip()) if len(row) > 1 and pd.notna(row.iloc[1]) else ''
            
            detected_subject = None
            
            try:
                int(first_cell)
                is_question_row = True
            except:
                is_question_row = False
            
            if not is_question_row:
                for pattern, subject in subject_patterns:
                    if pattern in first_cell:
                        detected_subject = subject
                        logger.info(f"Ders başlığı bulundu satır {idx}: '{first_cell}' -> {subject}")
                        break
            
            if detected_subject:
                if current_subject and questions_a:
                    answer_key_a[current_subject] = {
                        'answers': ''.join([q.get('answer', '') for q in questions_a]),
                        'questions': questions_a
                    }
                    questions_b = build_b_booklet_questions(questions_a)
                    if questions_b:
                        answer_key_b[current_subject] = {
                            'answers': ''.join([q.get('answer', '') for q in questions_b]),
                            'questions': questions_b
                        }
                current_subject = detected_subject
                questions_a = []
                continue
            
            if 'kazanım' in second_cell or 'kazanim' in second_cell:
                continue
            
            if current_subject:
                try:
                    q_num = int(row.iloc[0]) if pd.notna(row.iloc[0]) else None
                    if q_num is None:
                        continue
                    
                    kazanim = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else ''
                    
                    answer = ''
                    b_karsiligi = None
                    
                    if len(row) > 2 and pd.notna(row.iloc[2]):
                        val = str(row.iloc[2]).strip().upper()
                        if val in ['A', 'B', 'C', 'D']:
                            answer = val
                    
                    b_col_idx = 3
                    if len(row) > b_col_idx and pd.notna(row.iloc[b_col_idx]):
                        try:
                            b_karsiligi = int(row.iloc[b_col_idx])
                        except:
                            pass
                    
                    if q_num and (answer or kazanim):
                        questions_a.append({
                            'question_number': q_num,
                            'answer': answer,
                            'outcome': kazanim,
                            'b_equivalent': b_karsiligi
                        })
                        
                except Exception as e:
                    logger.debug(f"Satır parse hatası: {e}")
                    continue
        
        if current_subject and questions_a:
            answer_key_a[current_subject] = {
                'answers': ''.join([q.get('answer', '') for q in questions_a]),
                'questions': questions_a
            }
            questions_b = build_b_booklet_questions(questions_a)
            if questions_b:
                answer_key_b[current_subject] = {
                    'answers': ''.join([q.get('answer', '') for q in questions_b]),
                    'questions': questions_b
                }
        
        logger.info(f"Cevap anahtarı A parse edildi: {list(answer_key_a.keys())}")
        for subj, data in answer_key_a.items():
            logger.info(f"  A-{subj}: {len(data.get('questions', []))} soru, cevaplar: {data.get('answers', '')[:20]}...")
        
        logger.info(f"Cevap anahtarı B parse edildi: {list(answer_key_b.keys())}")
        for subj, data in answer_key_b.items():
            logger.info(f"  B-{subj}: {len(data.get('questions', []))} soru, cevaplar: {data.get('answers', '')[:20]}...")
        
        return answer_key_a, answer_key_b
        
    except Exception as e:
        logger.error(f"Cevap anahtarı parse hatası: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return {}, {}


def build_b_booklet_questions(questions_a):
    """A kitapçık sorularından B kitapçık cevap anahtarını oluştur
    B kitapçık soru sırası G sütunundaki değerlere göre belirlenir
    """
    b_mapping = {}
    for q in questions_a:
        b_equiv = q.get('b_equivalent')
        if b_equiv:
            b_mapping[b_equiv] = q
    
    if not b_mapping:
        return []
    
    questions_b = []
    for b_num in sorted(b_mapping.keys()):
        q = b_mapping[b_num]
        questions_b.append({
            'question_number': b_num,
            'answer': q.get('answer', ''),
            'outcome': q.get('outcome', ''),
            'a_equivalent': q.get('question_number')
        })
    
    return questions_b


def fetch_student_names_from_db(students):
    """Öğrenci numaralarından isimleri veritabanından çek"""
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        for student in students:
            student_no = student.get('student_no', '')
            csv_class = student.get('class_name', '')  # CSV'den gelen sınıf bilgisi
            
            if student_no:
                result = None
                
                # Önce student_no + sınıf ile eşleştir (aynı numara farklı sınıflarda olabilir)
                if csv_class:
                    csv_class_normalized = csv_class.replace('/', '')  # "7/B" -> "7B"
                    cur.execute("""
                        SELECT full_name, class_name FROM users 
                        WHERE role = 'student' AND student_no = %s
                        AND (REPLACE(class_name, '/', '') = %s OR class_name = %s)
                        LIMIT 1
                    """, (student_no, csv_class_normalized, csv_class))
                    result = cur.fetchone()
                
                # Sınıf eşleşmesi bulunamazsa sadece student_no ile ara
                if not result:
                    cur.execute("""
                        SELECT full_name, class_name FROM users 
                        WHERE role = 'student' AND student_no = %s
                        LIMIT 1
                    """, (student_no,))
                    result = cur.fetchone()
                
                if result:
                    student['name'] = result['full_name']
                    if result['class_name']:
                        student['class_name'] = result['class_name']
        
        cur.close()
        conn.close()
        
    except Exception as e:
        logger.error(f"Öğrenci ismi çekme hatası: {e}")
    
    return students


# ==================== YENİ TABLO YAPISINA UYGUN ENDPOINT'LER ====================

@report_cards_bp.route('/api/exams-list', methods=['GET'])
@login_required
def get_exams_list():
    """Yeni tablolardan sınav listesi - sınıf filtresi destekli"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    class_filter = request.args.get('class_name', '')
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        if class_filter:
            class_normalized = class_filter.replace('/', '')
            cur.execute("""
                SELECT DISTINCT e.id, e.exam_name, e.exam_date, e.source_type,
                       COUNT(r.id) as student_count,
                       e.created_at, e.grade_level,
                       STRING_AGG(DISTINCT r.class_name, ', ' ORDER BY r.class_name) as class_name
                FROM report_card_exams e
                JOIN report_card_results r ON e.id = r.exam_id
                WHERE r.class_name = %s OR REPLACE(r.class_name, '/', '') = %s
                GROUP BY e.id
                ORDER BY e.created_at DESC
            """, (class_filter, class_normalized))
        else:
            cur.execute("""
                SELECT e.id, e.exam_name, e.exam_date, e.source_type,
                       COUNT(r.id) as student_count,
                       e.created_at, e.grade_level,
                       STRING_AGG(DISTINCT r.class_name, ', ' ORDER BY r.class_name) as class_name
                FROM report_card_exams e
                LEFT JOIN report_card_results r ON e.id = r.exam_id
                GROUP BY e.id
                ORDER BY e.created_at DESC
            """)
        exams = cur.fetchall()
        
        return jsonify({"exams": exams})
        
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/delete-exam/<int:exam_id>', methods=['DELETE'])
@login_required
def delete_exam(exam_id):
    """Sınavı ve sonuçlarını sil"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor()
    
    try:
        cur.execute("DELETE FROM report_card_results WHERE exam_id = %s", (exam_id,))
        deleted_results = cur.rowcount
        
        cur.execute("DELETE FROM report_card_exams WHERE id = %s", (exam_id,))
        deleted_exam = cur.rowcount
        
        conn.commit()
        
        return jsonify({
            "success": True,
            "message": f"Sınav ve {deleted_results} öğrenci sonucu silindi"
        })
        
    except Exception as e:
        conn.rollback()
        logger.error(f"Sınav silme hatası: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/exam-students/<int:exam_id>', methods=['GET'])
@login_required
def get_exam_students(exam_id):
    """Belirli sınavın öğrenci listesi"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    class_name = request.args.get('class_name', '')
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Öğretmen için sınıf kontrolü
        allowed_classes = None
        allowed_classes_normalized = None
        if current_user.role == 'teacher':
            cur.execute("SELECT class_name FROM teacher_classes WHERE teacher_id = %s", (current_user.id,))
            allowed_classes = [row['class_name'] for row in cur.fetchall()]
            if allowed_classes:
                allowed_classes_normalized = [c.replace('/', '') for c in allowed_classes]
        
        query = """
            SELECT r.id, r.student_no, r.student_name, r.class_name,
                   r.grade_level, r.booklet_type, r.totals
            FROM report_card_results r
            WHERE r.exam_id = %s
        """
        params = [exam_id]
        
        if class_name:
            class_name_normalized = class_name.replace('/', '')
            query += " AND (r.class_name = %s OR REPLACE(r.class_name, '/', '') = %s)"
            params.extend([class_name, class_name_normalized])
        elif allowed_classes and allowed_classes_normalized:
            query += " AND (r.class_name = ANY(%s) OR REPLACE(r.class_name, '/', '') = ANY(%s))"
            params.extend([allowed_classes, allowed_classes_normalized])
        
        query += " ORDER BY r.class_name, r.student_name"
        
        cur.execute(query, params)
        students = cur.fetchall()
        
        for s in students:
            totals = s.get('totals') or {}
            if isinstance(totals, str):
                try:
                    totals = json.loads(totals)
                except:
                    totals = {}
            if not isinstance(totals, dict):
                totals = {}
            s['total_correct'] = totals.get('correct_count', totals.get('correct', 0)) or 0
            s['total_wrong'] = totals.get('wrong_count', totals.get('wrong', 0)) or 0
            s['total_blank'] = totals.get('blank_count', totals.get('blank', 0)) or 0
            net_val = totals.get('net_score', totals.get('net', 0))
            s['total_net'] = float(net_val) if net_val is not None else 0
            if 'totals' in s:
                del s['totals']
        
        return jsonify({"students": students})
        
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/student-error-report/<int:result_id>', methods=['GET'])
@login_required
def get_student_error_report(result_id):
    """Tek öğrenci hata karnesi - yeni tablolardan"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("""
            SELECT r.*, e.exam_name
            FROM report_card_results r
            JOIN report_card_exams e ON r.exam_id = e.id
            WHERE r.id = %s
        """, (result_id,))
        result = cur.fetchone()
        
        if not result:
            return jsonify({"error": "Sonuç bulunamadı"}), 404
        
        subjects = result.get('subjects') or {}
        if isinstance(subjects, str):
            try:
                subjects = json.loads(subjects)
            except:
                subjects = {}
        if not isinstance(subjects, dict):
            subjects = {}
        errors_by_subject = {}
        
        for subj_key, subj_data in subjects.items():
            wrong_answers = []
            for ans in subj_data.get('answers', []):
                if ans.get('status') in ['wrong', 'blank']:
                    wrong_answers.append({
                        'question_number': ans.get('question_number'),
                        'correct_answer': ans.get('correct_answer'),
                        'student_answer': ans.get('student_answer', ''),
                        'outcome': ans.get('outcome', ''),
                        'status': ans.get('status')
                    })
            
            if wrong_answers:
                wrong_answers.sort(key=lambda x: parse_outcome_code(x.get('outcome', '')))
                errors_by_subject[subj_key] = {
                    'subject_label': subj_data.get('subject_label', subj_key),
                    'errors': wrong_answers,
                    'correct_count': subj_data.get('correct_count', 0),
                    'wrong_count': subj_data.get('wrong_count', 0),
                    'blank_count': subj_data.get('blank_count', 0)
                }
        
        totals = result.get('totals') or {}
        if isinstance(totals, str):
            try:
                totals = json.loads(totals)
            except:
                totals = {}
        if not isinstance(totals, dict):
            totals = {}
        
        return jsonify({
            "student_name": result['student_name'],
            "student_no": result['student_no'],
            "class_name": result['class_name'],
            "exam_name": result['exam_name'],
            "errors_by_subject": errors_by_subject,
            "totals": totals
        })
        
    except Exception as e:
        logger.error(f"Hata karnesi hatası: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()




@report_cards_bp.route('/api/class-outcome-report/<int:exam_id>', methods=['GET'])
@login_required
def get_class_outcome_report(exam_id):
    """Sınıf bazlı kazanım raporu"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    class_name = request.args.get('class_name', '')
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        query = """
            SELECT r.subjects, r.student_name
            FROM report_card_results r
            WHERE r.exam_id = %s
        """
        params = [exam_id]
        
        if class_name:
            # Sınıf formatını normalize et (8/A -> 8A veya tersi)
            class_name_normalized = class_name.replace('/', '')
            query += " AND (r.class_name = %s OR REPLACE(r.class_name, '/', '') = %s)"
            params.extend([class_name, class_name_normalized])
        
        cur.execute(query, params)
        results = cur.fetchall()
        
        outcome_stats = {}
        
        for result in results:
            subjects = result.get('subjects', {})
            if isinstance(subjects, str):
                subjects = json.loads(subjects)
            for subj_key, subj_data in subjects.items():
                if subj_key not in outcome_stats:
                    outcome_stats[subj_key] = {}
                
                for ans in subj_data.get('answers', []):
                    outcome = ans.get('outcome', '')
                    if not outcome:
                        continue
                    
                    if outcome not in outcome_stats[subj_key]:
                        outcome_stats[subj_key][outcome] = {'correct': 0, 'wrong': 0, 'blank': 0, 'total': 0}
                    
                    outcome_stats[subj_key][outcome]['total'] += 1
                    if ans.get('status') == 'correct':
                        outcome_stats[subj_key][outcome]['correct'] += 1
                    elif ans.get('status') == 'wrong':
                        outcome_stats[subj_key][outcome]['wrong'] += 1
                    else:
                        outcome_stats[subj_key][outcome]['blank'] += 1
        
        for subj_key in outcome_stats:
            for outcome in outcome_stats[subj_key]:
                stats = outcome_stats[subj_key][outcome]
                if stats['total'] > 0:
                    stats['success_rate'] = round((stats['correct'] / stats['total']) * 100, 1)
                else:
                    stats['success_rate'] = 0
        
        return jsonify({
            "exam_id": exam_id,
            "class_name": class_name,
            "student_count": len(results),
            "outcome_stats": outcome_stats
        })
        
    except Exception as e:
        logger.error(f"Sınıf kazanım raporu hatası: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/student-outcome-report/<int:result_id>', methods=['GET'])
@login_required
def get_student_outcome_report(result_id):
    """Öğrenci kazanım raporu"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("""
            SELECT r.*, e.exam_name
            FROM report_card_results r
            JOIN report_card_exams e ON r.exam_id = e.id
            WHERE r.id = %s
        """, (result_id,))
        result = cur.fetchone()
        
        if not result:
            return jsonify({"error": "Sonuç bulunamadı"}), 404
        
        subjects = result.get('subjects', {})
        if isinstance(subjects, str):
            subjects = json.loads(subjects)
        outcome_analysis = {}
        
        for subj_key, subj_data in subjects.items():
            outcome_analysis[subj_key] = {
                'subject_label': subj_data.get('subject_label', subj_key),
                'outcomes': {}
            }
            
            for ans in subj_data.get('answers', []):
                outcome = ans.get('outcome', '')
                if not outcome:
                    continue
                
                if outcome not in outcome_analysis[subj_key]['outcomes']:
                    outcome_analysis[subj_key]['outcomes'][outcome] = {
                        'questions': [],
                        'correct': 0,
                        'wrong': 0,
                        'blank': 0
                    }
                
                outcome_analysis[subj_key]['outcomes'][outcome]['questions'].append({
                    'question_number': ans.get('question_number'),
                    'status': ans.get('status')
                })
                
                if ans.get('status') == 'correct':
                    outcome_analysis[subj_key]['outcomes'][outcome]['correct'] += 1
                elif ans.get('status') == 'wrong':
                    outcome_analysis[subj_key]['outcomes'][outcome]['wrong'] += 1
                else:
                    outcome_analysis[subj_key]['outcomes'][outcome]['blank'] += 1
        
        return jsonify({
            "student_name": result['student_name'],
            "student_no": result['student_no'],
            "class_name": result['class_name'],
            "exam_name": result['exam_name'],
            "outcome_analysis": outcome_analysis
        })
        
    except Exception as e:
        logger.error(f"Öğrenci kazanım raporu hatası: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/student-outcome-pdf-new/<int:result_id>', methods=['GET'])
@login_required
def get_student_outcome_pdf_new(result_id):
    """Ogrenci kazanim raporu PDF"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erisim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("""
            SELECT r.*, e.exam_name
            FROM report_card_results r
            JOIN report_card_exams e ON r.exam_id = e.id
            WHERE r.id = %s
        """, (result_id,))
        result = cur.fetchone()
        
        if not result:
            return jsonify({"error": "Sonuc bulunamadi"}), 404
        
        subjects = result.get('subjects', {})
        if isinstance(subjects, str):
            subjects = json.loads(subjects)
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        
        elements.extend(create_pdf_header(styles))
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontName=PDF_FONT, fontSize=16, alignment=TA_CENTER, spaceAfter=15)
        normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontName=PDF_FONT, fontSize=10)
        subject_style = ParagraphStyle('Subject', parent=styles['Heading2'], fontName=PDF_FONT, fontSize=12, textColor=colors.HexColor('#2563eb'), spaceAfter=8, spaceBefore=12)
        
        elements.append(Paragraph("OGRENCI KAZANIM RAPORU", title_style))
        elements.append(Paragraph(f"Ogrenci: {result['student_name']}", normal_style))
        elements.append(Paragraph(f"Sinif: {result['class_name']}", normal_style))
        elements.append(Paragraph(f"Sinav: {result['exam_name']}", normal_style))
        elements.append(Spacer(1, 15))
        
        for subj_key, subj_data in subjects.items():
            subject_label = subj_data.get('subject_label', subj_key)
            elements.append(Paragraph(subject_label, subject_style))
            
            outcome_stats = {}
            for ans in subj_data.get('answers', []):
                outcome = ans.get('outcome', '')
                if not outcome:
                    continue
                if outcome not in outcome_stats:
                    outcome_stats[outcome] = {'correct': 0, 'wrong': 0, 'blank': 0}
                if ans.get('status') == 'correct':
                    outcome_stats[outcome]['correct'] += 1
                elif ans.get('status') == 'wrong':
                    outcome_stats[outcome]['wrong'] += 1
                else:
                    outcome_stats[outcome]['blank'] += 1
            
            table_data = [['Kazanim', 'D', 'Y', 'B', '%']]
            for outcome, stats in outcome_stats.items():
                total = stats['correct'] + stats['wrong'] + stats['blank']
                rate = round((stats['correct'] / total) * 100) if total > 0 else 0
                table_data.append([outcome[:60], str(stats['correct']), str(stats['wrong']), str(stats['blank']), f'{rate}%'])
            
            if len(table_data) > 1:
                t = Table(table_data, colWidths=[300, 40, 40, 40, 50])
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f3f4f6')),
                    ('FONTNAME', (0, 0), (-1, -1), PDF_FONT),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
                    ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 4),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ]))
                elements.append(t)
        
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=f'kazanim_raporu_{result["student_name"]}.pdf')
        
    except Exception as e:
        logger.error(f"PDF olusturma hatasi: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/class-outcome-pdf-new/<int:exam_id>', methods=['GET'])
@login_required
def get_class_outcome_pdf_new(exam_id):
    """Sinif kazanim raporu PDF"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erisim"}), 403
    
    class_name = request.args.get('class_name', '')
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("SELECT exam_name FROM report_card_exams WHERE id = %s", (exam_id,))
        exam = cur.fetchone()
        if not exam:
            return jsonify({"error": "Sinav bulunamadi"}), 404
        
        query = "SELECT * FROM report_card_results WHERE exam_id = %s"
        params = [exam_id]
        if class_name:
            # Sınıf formatını normalize et (8/A -> 8A veya tersi)
            class_name_normalized = class_name.replace('/', '')
            query += " AND (class_name = %s OR REPLACE(class_name, '/', '') = %s)"
            params.extend([class_name, class_name_normalized])
        
        cur.execute(query, params)
        results = cur.fetchall()
        
        outcome_totals = {}
        
        for result in results:
            subjects = result.get('subjects', {})
            if isinstance(subjects, str):
                subjects = json.loads(subjects)
            
            for subj_key, subj_data in subjects.items():
                subject_label = subj_data.get('subject_label', subj_key)
                if subject_label not in outcome_totals:
                    outcome_totals[subject_label] = {}
                
                for ans in subj_data.get('answers', []):
                    outcome = ans.get('outcome', '')
                    if not outcome:
                        continue
                    if outcome not in outcome_totals[subject_label]:
                        outcome_totals[subject_label][outcome] = {'correct': 0, 'wrong': 0, 'blank': 0}
                    if ans.get('status') == 'correct':
                        outcome_totals[subject_label][outcome]['correct'] += 1
                    elif ans.get('status') == 'wrong':
                        outcome_totals[subject_label][outcome]['wrong'] += 1
                    else:
                        outcome_totals[subject_label][outcome]['blank'] += 1
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        
        elements.extend(create_pdf_header(styles))
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontName=PDF_FONT, fontSize=16, alignment=TA_CENTER, spaceAfter=15)
        normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontName=PDF_FONT, fontSize=10)
        subject_style = ParagraphStyle('Subject', parent=styles['Heading2'], fontName=PDF_FONT, fontSize=12, textColor=colors.HexColor('#2563eb'), spaceAfter=8, spaceBefore=12)
        
        elements.append(Paragraph("SINIF KAZANIM RAPORU", title_style))
        elements.append(Paragraph(f"Sinav: {exam['exam_name']}", normal_style))
        if class_name:
            elements.append(Paragraph(f"Sinif: {class_name}", normal_style))
        elements.append(Paragraph(f"Toplam Ogrenci: {len(results)}", normal_style))
        elements.append(Spacer(1, 15))
        
        for subject_label, outcomes in outcome_totals.items():
            elements.append(Paragraph(subject_label, subject_style))
            
            table_data = [['Kazanim', 'D', 'Y', 'B', '%']]
            for outcome, stats in outcomes.items():
                total = stats['correct'] + stats['wrong'] + stats['blank']
                rate = round((stats['correct'] / total) * 100) if total > 0 else 0
                table_data.append([outcome[:60], str(stats['correct']), str(stats['wrong']), str(stats['blank']), f'{rate}%'])
            
            if len(table_data) > 1:
                t = Table(table_data, colWidths=[300, 40, 40, 40, 50])
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f3f4f6')),
                    ('FONTNAME', (0, 0), (-1, -1), PDF_FONT),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
                    ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 4),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ]))
                elements.append(t)
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f'sinif_kazanim_{class_name or "tum"}.pdf'
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)
        
    except Exception as e:
        logger.error(f"PDF olusturma hatasi: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


# ==================== TXT DÖNÜŞTÜRÜCÜ ====================

@report_cards_bp.route('/txt-converter')
@login_required
def txt_converter_page():
    """TXT to CSV dönüştürücü sayfası"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    return render_template('admin_txt_converter.html')


@report_cards_bp.route('/api/convert-txt', methods=['POST'])
@login_required
def convert_txt_to_csv():
    """TXT dosyasını CSV'ye dönüştür"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    if 'txt_file' not in request.files:
        return jsonify({"error": "TXT dosyası bulunamadı"}), 400
    
    txt_file = request.files['txt_file']
    grade = request.form.get('grade', '5')
    
    if txt_file.filename == '':
        return jsonify({"error": "Dosya seçilmedi"}), 400
    
    try:
        # TXT dosyasını oku - farklı encoding'leri dene
        raw_data = txt_file.read()
        content = None
        encodings = ['utf-8', 'windows-1254', 'iso-8859-9', 'cp1254', 'latin-1']
        
        for encoding in encodings:
            try:
                content = raw_data.decode(encoding)
                logger.info(f"TXT dosyası {encoding} encoding ile okundu")
                break
            except (UnicodeDecodeError, LookupError):
                continue
        
        if content is None:
            return jsonify({"error": "Dosya encoding'i okunamadı"}), 400
            
        lines = content.splitlines()
        
        processed_data = []
        
        for line in lines:
            if len(line) < 50:
                continue
            
            # TXT formatı - debug ile analiz
            # Kurum(8) + Boşluk(11) + ÖğrNo(5) + Oturum(1) + Kitapçık(1) + AdSoyad(20) + Sınıf(3) + Cevaplar
            kurum_kodu = line[0:8].strip()
            ogrenci_no = line[19:24].strip()  # 5 hane
            oturum = line[24:25].strip()      # 1 veya 2
            kitapcik = line[25:26].strip()    # A veya B
            ad_soyad = line[26:46].strip()    # 20 karakter
            
            # Sınıf alanı: 46-50 arası (4 karakter) - "7A", "7 B", " 7D" gibi formatlar olabilir
            sinif_raw = line[46:50]
            # Sınıf formatını düzelt: [5-8] ve [A-D] arasında boşluk olabilir
            import re
            # Önce "7 B" veya "7B" formatını ara (rakam + opsiyonel boşluk + harf)
            sinif_match = re.search(r'([5-8])\s*([A-D])', sinif_raw)
            if sinif_match:
                # Boşluksuz formata dönüştür: "7B"
                sinif = sinif_match.group(1) + sinif_match.group(2)
            else:
                # Sadece sınıf numarası varsa (7, 8 gibi)
                sinif_num_match = re.search(r'[5-8]', sinif_raw)
                if sinif_num_match:
                    sinif = sinif_num_match.group()
                else:
                    sinif = sinif_raw.strip()
            
            # Cevaplar 50. karakterden başlıyor (sınıf alanı dahil 50 karakter meta)
            cevaplar_raw = line[50:]
            
            # Debug log
            if len(processed_data) < 3:
                current_app.logger.info(f"DEBUG TXT satır: oturum={oturum}, sinif={sinif}, cevaplar_ilk60=[{cevaplar_raw[:60]}]")
            
            # Sınıfa göre soru sayılarını ve alan genişliklerini belirle
            # TXT formatı: Her cevap bloğu sonrası padding var
            if grade in ['5', '6']:
                # Oturum 2: Türkçe(15)+pad(5), Sosyal(10)+pad(10), Din(10)+pad(10), İng(10)
                # Oturum 1: Mat(15)+pad(5), Fen(15)
                turkce_len = 15
                turkce_pad = 5
                sosyal_len = 10
                sosyal_pad = 10
                din_len = 10
                din_pad = 10
                ing_len = 10
                mat_len = 15
                mat_pad = 5
                fen_len = 15
            else:  # 7, 8
                turkce_len = 20
                turkce_pad = 0
                sosyal_len = 10
                sosyal_pad = 10
                din_len = 10
                din_pad = 10
                ing_len = 10
                mat_len = 20
                mat_pad = 0
                fen_len = 20
            
            if oturum == "2":
                # Sözel oturum: Türkçe, Sosyal, Din, İngilizce
                pos = 0
                turkce = cevaplar_raw[pos:pos+turkce_len]
                pos += turkce_len + turkce_pad
                sosyal = cevaplar_raw[pos:pos+sosyal_len]
                pos += sosyal_len + sosyal_pad
                din = cevaplar_raw[pos:pos+din_len]
                pos += din_len + din_pad
                ing = cevaplar_raw[pos:pos+ing_len]
                mat = ""
                fen = ""
            else:
                # Sayısal oturum: Mat ve Fen
                if grade in ['5', '6']:
                    # 5-6. sınıf: Mat 80, Fen 100
                    mat = cevaplar_raw[80:80+mat_len]
                    fen = cevaplar_raw[100:100+fen_len]
                else:
                    # 7-8. sınıf: Mat 80. pozisyondan (20 karakter), Fen 100. pozisyondan (20 karakter)
                    mat = cevaplar_raw[80:80+mat_len]
                    fen = cevaplar_raw[100:100+fen_len]
                turkce = ""
                sosyal = ""
                din = ""
                ing = ""
            
            # Oturum 2: Türkçe/Sosyal/Din/İng dolu, Mat/Fen boş
            # Oturum 1: Türkçe/Sosyal/Din/İng boş, Mat/Fen dolu
            row = [
                kurum_kodu,
                ogrenci_no,
                oturum,
                kitapcik,
                ad_soyad,
                sinif,
                turkce if oturum == "2" else "",
                sosyal if oturum == "2" else "",
                din if oturum == "2" else "",
                ing if oturum == "2" else "",
                mat if oturum == "1" else "",
                fen if oturum == "1" else ""
            ]
            processed_data.append(row)
        
        if not processed_data:
            return jsonify({"error": "İşlenecek veri bulunamadı"}), 400
        
        # Önce Sözel (2), sonra Sayısal (1) gelecek şekilde sırala
        processed_data.sort(key=lambda x: x[2], reverse=True)
        
        # CSV oluştur - başlık satırı olmadan
        import io
        import csv
        
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')
        writer.writerows(processed_data)
        
        csv_content = output.getvalue()
        
        # BytesIO ile dosya oluştur
        buffer = BytesIO()
        buffer.write(csv_content.encode('utf-8-sig'))
        buffer.seek(0)
        
        filename = f'{grade}_Sinif_Sonuc.csv'
        return send_file(buffer, mimetype='text/csv', as_attachment=True, download_name=filename)
        
    except Exception as e:
        logger.error(f"TXT dönüştürme hatası: {e}")
        return jsonify({"error": str(e)}), 500


# ==================== ÖĞRENCİ/ÖĞRETMEN KARNE ANALİZİ ====================

@report_cards_bp.route('/student-analysis')
@login_required
def student_analysis_page():
    """Öğrenci karne analizi sayfası"""
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    return render_template('student_karne_analysis.html')


@report_cards_bp.route('/teacher-analysis')
@login_required
def teacher_analysis_page():
    """Öğretmen karne analizi sayfası"""
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("SELECT class_name FROM teacher_classes WHERE teacher_id = %s ORDER BY class_name", (current_user.id,))
        teacher_classes = [row['class_name'] for row in cur.fetchall()]
    except:
        teacher_classes = []
    finally:
        cur.close()
        conn.close()
    
    return render_template('teacher_karne_analysis.html', teacher_classes=teacher_classes)


@report_cards_bp.route('/api/student-my-exams')
@login_required
def student_my_exams():
    """Öğrencinin CSV sınavlarını getir"""
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Kullanıcı bilgilerini veritabanından al (student_no dahil)
        cur.execute("SELECT student_no, class_name, full_name FROM users WHERE id = %s", (current_user.id,))
        user_data = cur.fetchone()
        
        student_class = user_data.get('class_name', '') if user_data else ''
        student_no = user_data.get('student_no', '') if user_data else ''
        student_name = user_data.get('full_name', '') if user_data else ''
        
        # Sınıf grade seviyesini çıkar (örn: "6/C" -> "6", "7A" -> "7")
        grade_level = ''
        if student_class:
            grade_level = student_class[0] if student_class[0].isdigit() else ''
        
        # student_no'yu farklı formatlarda ara (başında sıfır olabilir)
        student_no_padded = student_no.zfill(5) if student_no else ''
        student_no_int = student_no.lstrip('0') if student_no else ''
        
        query = """
            SELECT e.id, e.exam_name, e.grade_level, e.upload_date,
                   r.id as result_id, r.student_name, r.class_name, r.totals
            FROM report_card_exams e
            JOIN report_card_results r ON r.exam_id = e.id
            WHERE (
                UPPER(r.student_name) = UPPER(%s)
                OR r.student_no = %s
                OR r.student_no = %s
                OR LTRIM(r.student_no, '0') = %s
            )
            ORDER BY e.upload_date DESC
        """
        cur.execute(query, (student_name, student_no, student_no_padded, student_no_int))
        exams = cur.fetchall()
        
        for e in exams:
            if e.get('upload_date'):
                e['upload_date'] = e['upload_date'].isoformat()
            # totals JSON'dan değerleri çıkar
            totals = e.get('totals', {})
            if isinstance(totals, str):
                totals = json.loads(totals)
            e['total_correct'] = totals.get('correct_count', totals.get('correct', 0))
            e['total_wrong'] = totals.get('wrong_count', totals.get('wrong', 0))
            e['total_blank'] = totals.get('blank_count', totals.get('blank', 0))
            net_val = totals.get('net_score', totals.get('net', 0))
            e['total_net'] = float(net_val) if net_val is not None else 0
            del e['totals']
        
        return jsonify(exams)
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/teacher-class-exams')
@login_required
def teacher_class_exams():
    """Öğretmenin sınıflarına ait CSV sınavlarını getir"""
    if current_user.role not in ['teacher', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        if current_user.role == 'admin':
            cur.execute("SELECT DISTINCT class_name FROM report_card_results ORDER BY class_name")
            teacher_classes = [row['class_name'] for row in cur.fetchall()]
            
            query = """
                SELECT DISTINCT e.id, e.exam_name, e.grade_level, e.upload_date,
                       COUNT(DISTINCT r.id) as student_count
                FROM report_card_exams e
                JOIN report_card_results r ON r.exam_id = e.id
                GROUP BY e.id, e.exam_name, e.grade_level, e.upload_date
                ORDER BY e.upload_date DESC
            """
            cur.execute(query)
        else:
            cur.execute("""
                SELECT class_name FROM teacher_classes WHERE teacher_id = %s
            """, (current_user.id,))
            teacher_classes = [row['class_name'] for row in cur.fetchall()]
            
            if not teacher_classes:
                return jsonify({"exams": [], "classes": []})
            
            query = """
                SELECT DISTINCT e.id, e.exam_name, e.grade_level, e.upload_date,
                       COUNT(DISTINCT r.id) as student_count
                FROM report_card_exams e
                JOIN report_card_results r ON r.exam_id = e.id
                WHERE r.class_name = ANY(%s) OR REPLACE(r.class_name, '/', '') = ANY(%s)
                GROUP BY e.id, e.exam_name, e.grade_level, e.upload_date
                ORDER BY e.upload_date DESC
            """
            classes_normalized = [c.replace('/', '') for c in teacher_classes]
            cur.execute(query, (teacher_classes, classes_normalized))
        
        exams = cur.fetchall()
        
        for e in exams:
            if e.get('upload_date'):
                e['upload_date'] = e['upload_date'].isoformat()
        
        return jsonify({"exams": exams, "classes": teacher_classes})
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/student-outcome-analysis/<int:result_id>')
@login_required
def student_outcome_analysis(result_id):
    """Öğrenci kazanım analizi"""
    if current_user.role not in ['student', 'teacher', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("""
            SELECT r.*, e.exam_name
            FROM report_card_results r
            JOIN report_card_exams e ON r.exam_id = e.id
            WHERE r.id = %s
        """, (result_id,))
        result = cur.fetchone()
        
        if not result:
            return jsonify({"error": "Sonuç bulunamadı"}), 404
        
        # Yetkilendirme kontrolü
        if current_user.role == 'student':
            cur.execute("SELECT student_no, class_name, full_name FROM users WHERE id = %s", (current_user.id,))
            user_data = cur.fetchone()
            student_no = user_data.get('student_no', '') if user_data else ''
            student_class = user_data.get('class_name', '') if user_data else ''
            student_name = user_data.get('full_name', '') if user_data else ''
            
            student_no_padded = student_no.zfill(5) if student_no else ''
            student_no_int = student_no.lstrip('0') if student_no else ''
            result_student_no = result.get('student_no', '') or ''
            result_student_no_int = result_student_no.lstrip('0') if result_student_no else ''
            
            class_normalized = student_class.replace('/', '')
            result_class_normalized = (result.get('class_name') or '').replace('/', '')
            
            student_no_match = (
                result_student_no == student_no or
                result_student_no == student_no_padded or
                result_student_no_int == student_no_int
            )
            class_match = (result.get('class_name') == student_class or result_class_normalized == class_normalized)
            name_match = result.get('student_name', '').upper() == student_name.upper()
            
            is_owner = (student_no_match and class_match) or name_match
            if not is_owner:
                return jsonify({"error": "Bu veriye erişim yetkiniz yok"}), 403
        
        elif current_user.role == 'teacher':
            cur.execute("SELECT class_name FROM teacher_classes WHERE teacher_id = %s", (current_user.id,))
            teacher_classes = [row['class_name'] for row in cur.fetchall()]
            teacher_classes_normalized = [c.replace('/', '') for c in teacher_classes]
            result_class = result.get('class_name', '')
            result_class_normalized = result_class.replace('/', '')
            
            if result_class not in teacher_classes and result_class_normalized not in teacher_classes_normalized:
                return jsonify({"error": "Bu sınıfa erişim yetkiniz yok"}), 403
        
        subjects = result.get('subjects', {})
        if isinstance(subjects, str):
            subjects = json.loads(subjects)
        
        outcome_data = {}
        for subj_key, subj_data in subjects.items():
            subject_label = subj_data.get('subject_label', subj_key)
            outcomes = {}
            
            for ans in subj_data.get('answers', []):
                outcome = ans.get('outcome', '')
                if not outcome:
                    continue
                
                if outcome not in outcomes:
                    outcomes[outcome] = {'correct': 0, 'wrong': 0, 'blank': 0, 'total': 0}
                
                outcomes[outcome]['total'] += 1
                if ans.get('status') == 'correct':
                    outcomes[outcome]['correct'] += 1
                elif ans.get('status') == 'wrong':
                    outcomes[outcome]['wrong'] += 1
                else:
                    outcomes[outcome]['blank'] += 1
            
            for oc in outcomes.values():
                total = oc['total']
                oc['success_rate'] = round((oc['correct'] / total * 100), 1) if total > 0 else 0
            
            if outcomes:
                sorted_outcomes = dict(sorted(outcomes.items(), key=lambda x: parse_outcome_code(x[0])))
                outcome_data[subject_label] = sorted_outcomes
        
        return jsonify({
            "student_name": result['student_name'],
            "class_name": result['class_name'],
            "exam_name": result['exam_name'],
            "outcome_data": outcome_data
        })
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/student-error-details/<int:result_id>')
@login_required
def student_error_details(result_id):
    """Öğrenci hata karnesi detayları"""
    if current_user.role not in ['student', 'teacher', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("""
            SELECT r.*, e.exam_name
            FROM report_card_results r
            JOIN report_card_exams e ON r.exam_id = e.id
            WHERE r.id = %s
        """, (result_id,))
        result = cur.fetchone()
        
        if not result:
            return jsonify({"error": "Sonuç bulunamadı"}), 404
        
        # Yetkilendirme kontrolü
        if current_user.role == 'student':
            cur.execute("SELECT student_no, class_name, full_name FROM users WHERE id = %s", (current_user.id,))
            user_data = cur.fetchone()
            student_no = user_data.get('student_no', '') if user_data else ''
            student_class = user_data.get('class_name', '') if user_data else ''
            student_name = user_data.get('full_name', '') if user_data else ''
            
            student_no_padded = student_no.zfill(5) if student_no else ''
            student_no_int = student_no.lstrip('0') if student_no else ''
            result_student_no = result.get('student_no', '') or ''
            result_student_no_int = result_student_no.lstrip('0') if result_student_no else ''
            
            class_normalized = student_class.replace('/', '')
            result_class_normalized = (result.get('class_name') or '').replace('/', '')
            
            student_no_match = (
                result_student_no == student_no or
                result_student_no == student_no_padded or
                result_student_no_int == student_no_int
            )
            class_match = (result.get('class_name') == student_class or result_class_normalized == class_normalized)
            name_match = result.get('student_name', '').upper() == student_name.upper()
            
            is_owner = (student_no_match and class_match) or name_match
            if not is_owner:
                return jsonify({"error": "Bu veriye erişim yetkiniz yok"}), 403
        
        elif current_user.role == 'teacher':
            cur.execute("SELECT class_name FROM teacher_classes WHERE teacher_id = %s", (current_user.id,))
            teacher_classes = [row['class_name'] for row in cur.fetchall()]
            teacher_classes_normalized = [c.replace('/', '') for c in teacher_classes]
            result_class = result.get('class_name', '')
            result_class_normalized = result_class.replace('/', '')
            
            if result_class not in teacher_classes and result_class_normalized not in teacher_classes_normalized:
                return jsonify({"error": "Bu sınıfa erişim yetkiniz yok"}), 403
        
        subjects = result.get('subjects', {})
        if isinstance(subjects, str):
            subjects = json.loads(subjects)
        
        error_data = {}
        calc_correct = 0
        calc_wrong = 0
        calc_blank = 0
        
        for subj_key, subj_data in subjects.items():
            subject_label = subj_data.get('subject_label', subj_key)
            errors = []
            
            for ans in subj_data.get('answers', []):
                status = ans.get('status')
                if status == 'correct':
                    calc_correct += 1
                elif status == 'wrong':
                    calc_wrong += 1
                    errors.append({
                        'question_no': ans.get('question_no'),
                        'student_answer': ans.get('student_answer', '-'),
                        'correct_answer': ans.get('correct_answer', '-'),
                        'outcome': ans.get('outcome', ''),
                        'status': status
                    })
                elif status == 'blank':
                    calc_blank += 1
                    errors.append({
                        'question_no': ans.get('question_no'),
                        'student_answer': ans.get('student_answer', '-'),
                        'correct_answer': ans.get('correct_answer', '-'),
                        'outcome': ans.get('outcome', ''),
                        'status': status
                    })
            
            if errors:
                error_data[subject_label] = errors
        
        total_correct = result.get('total_correct') or calc_correct
        total_wrong = result.get('total_wrong') or calc_wrong
        total_blank = result.get('total_blank') or calc_blank
        
        return jsonify({
            "student_name": result['student_name'],
            "class_name": result['class_name'],
            "exam_name": result['exam_name'],
            "total_correct": total_correct,
            "total_wrong": total_wrong,
            "total_blank": total_blank,
            "error_data": error_data
        })
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/student-error-pdf/<int:result_id>')
@login_required
def student_error_pdf_download(result_id):
    """Öğrenci hata karnesi PDF"""
    if current_user.role not in ['student', 'teacher', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("""
            SELECT r.*, e.exam_name
            FROM report_card_results r
            JOIN report_card_exams e ON r.exam_id = e.id
            WHERE r.id = %s
        """, (result_id,))
        result = cur.fetchone()
        
        if not result:
            return jsonify({"error": "Sonuç bulunamadı"}), 404
        
        # Yetkilendirme kontrolü
        if current_user.role == 'student':
            cur.execute("SELECT student_no, class_name, full_name FROM users WHERE id = %s", (current_user.id,))
            user_data = cur.fetchone()
            student_no = user_data.get('student_no', '') if user_data else ''
            student_class = user_data.get('class_name', '') if user_data else ''
            student_name = user_data.get('full_name', '') if user_data else ''
            
            student_no_padded = student_no.zfill(5) if student_no else ''
            student_no_int = student_no.lstrip('0') if student_no else ''
            result_student_no = result.get('student_no', '') or ''
            result_student_no_int = result_student_no.lstrip('0') if result_student_no else ''
            
            class_normalized = student_class.replace('/', '')
            result_class_normalized = (result.get('class_name') or '').replace('/', '')
            
            student_no_match = (
                result_student_no == student_no or
                result_student_no == student_no_padded or
                result_student_no_int == student_no_int
            )
            class_match = (result.get('class_name') == student_class or result_class_normalized == class_normalized)
            name_match = result.get('student_name', '').upper() == student_name.upper()
            
            is_owner = (student_no_match and class_match) or name_match
            if not is_owner:
                return jsonify({"error": "Bu veriye erişim yetkiniz yok"}), 403
        
        elif current_user.role == 'teacher':
            cur.execute("SELECT class_name FROM teacher_classes WHERE teacher_id = %s", (current_user.id,))
            teacher_classes = [row['class_name'] for row in cur.fetchall()]
            teacher_classes_normalized = [c.replace('/', '') for c in teacher_classes]
            result_class = result.get('class_name', '')
            result_class_normalized = result_class.replace('/', '')
            
            if result_class not in teacher_classes and result_class_normalized not in teacher_classes_normalized:
                return jsonify({"error": "Bu sınıfa erişim yetkiniz yok"}), 403
        
        subjects = result.get('subjects', {})
        if isinstance(subjects, str):
            subjects = json.loads(subjects)
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        
        elements.extend(create_pdf_header(styles))
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontName=PDF_FONT, fontSize=16, alignment=TA_CENTER, spaceAfter=15)
        normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontName=PDF_FONT, fontSize=10)
        subject_style = ParagraphStyle('Subject', parent=styles['Heading2'], fontName=PDF_FONT, fontSize=12, textColor=colors.HexColor('#dc2626'), spaceAfter=8, spaceBefore=12)
        
        elements.append(Paragraph("HATA KARNESİ", title_style))
        elements.append(Paragraph(f"Öğrenci: {result['student_name']}", normal_style))
        elements.append(Paragraph(f"Sınıf: {result['class_name']}", normal_style))
        elements.append(Paragraph(f"Sınav: {result['exam_name']}", normal_style))
        elements.append(Spacer(1, 15))
        
        for subj_key, subj_data in subjects.items():
            subject_label = subj_data.get('subject_label', subj_key)
            errors = []
            
            for ans in subj_data.get('answers', []):
                if ans.get('status') in ['wrong', 'blank']:
                    errors.append(ans)
            
            if errors:
                elements.append(Paragraph(subject_label, subject_style))
                
                outcome_style = ParagraphStyle('OutcomeCell', parent=styles['Normal'], fontName=PDF_FONT, fontSize=7, leading=9, wordWrap='CJK')
                
                table_data = [['Soru', 'Cevabınız', 'Doğru', 'Kazanım']]
                for err in errors:
                    q_no = str(err.get('question_no', '-'))
                    student_ans = err.get('student_answer', '-')
                    correct_ans = err.get('correct_answer', '-')
                    outcome_text = err.get('outcome', '') or '-'
                    outcome_para = Paragraph(outcome_text, outcome_style)
                    table_data.append([q_no, student_ans, correct_ans, outcome_para])
                
                t = Table(table_data, colWidths=[35, 50, 50, 335])
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fee2e2')),
                    ('FONTNAME', (0, 0), (-1, -1), PDF_FONT),
                    ('FONTSIZE', (0, 0), (-1, 0), 8),
                    ('FONTSIZE', (0, 1), (2, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#fca5a5')),
                    ('ALIGN', (0, 0), (2, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ]))
                elements.append(t)
                elements.append(Spacer(1, 10))
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f'hata_karnesi_{result["student_name"].replace(" ", "_")}.pdf'
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)
        
    except Exception as e:
        logger.error(f"PDF oluşturma hatası: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/student-outcome-pdf/<int:result_id>')
@login_required
def student_outcome_pdf_download(result_id):
    """Öğrenci kazanım analizi PDF"""
    if current_user.role not in ['student', 'teacher', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("""
            SELECT r.*, e.exam_name
            FROM report_card_results r
            JOIN report_card_exams e ON r.exam_id = e.id
            WHERE r.id = %s
        """, (result_id,))
        result = cur.fetchone()
        
        if not result:
            return jsonify({"error": "Sonuç bulunamadı"}), 404
        
        if current_user.role == 'student':
            cur.execute("SELECT student_no, class_name, full_name FROM users WHERE id = %s", (current_user.id,))
            user_data = cur.fetchone()
            student_no = user_data.get('student_no', '') if user_data else ''
            student_class = user_data.get('class_name', '') if user_data else ''
            student_name = user_data.get('full_name', '') if user_data else ''
            
            student_no_padded = student_no.zfill(5) if student_no else ''
            student_no_int = student_no.lstrip('0') if student_no else ''
            result_student_no = result.get('student_no', '') or ''
            result_student_no_int = result_student_no.lstrip('0') if result_student_no else ''
            
            class_normalized = student_class.replace('/', '')
            result_class_normalized = (result.get('class_name') or '').replace('/', '')
            
            student_no_match = (
                result_student_no == student_no or
                result_student_no == student_no_padded or
                result_student_no_int == student_no_int
            )
            class_match = (result.get('class_name') == student_class or result_class_normalized == class_normalized)
            name_match = result.get('student_name', '').upper() == student_name.upper()
            
            is_owner = (student_no_match and class_match) or name_match
            if not is_owner:
                return jsonify({"error": "Bu veriye erişim yetkiniz yok"}), 403
        
        elif current_user.role == 'teacher':
            cur.execute("SELECT class_name FROM teacher_classes WHERE teacher_id = %s", (current_user.id,))
            teacher_classes = [row['class_name'] for row in cur.fetchall()]
            teacher_classes_normalized = [c.replace('/', '') for c in teacher_classes]
            result_class = result.get('class_name', '')
            result_class_normalized = result_class.replace('/', '')
            
            if result_class not in teacher_classes and result_class_normalized not in teacher_classes_normalized:
                return jsonify({"error": "Bu sınıfa erişim yetkiniz yok"}), 403
        
        subjects = result.get('subjects', {})
        if isinstance(subjects, str):
            subjects = json.loads(subjects)
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        
        elements.extend(create_pdf_header(styles))
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontName=PDF_FONT, fontSize=16, alignment=TA_CENTER, spaceAfter=15)
        normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontName=PDF_FONT, fontSize=10)
        subject_style = ParagraphStyle('Subject', parent=styles['Heading2'], fontName=PDF_FONT, fontSize=12, textColor=colors.HexColor('#2563eb'), spaceAfter=8, spaceBefore=12)
        
        elements.append(Paragraph("KAZANIM ANALİZİ RAPORU", title_style))
        elements.append(Paragraph(f"Öğrenci: {result['student_name']}", normal_style))
        elements.append(Paragraph(f"Sınıf: {result['class_name']}", normal_style))
        elements.append(Paragraph(f"Sınav: {result['exam_name']}", normal_style))
        elements.append(Spacer(1, 15))
        
        for subj_key, subj_data in subjects.items():
            subject_label = subj_data.get('subject_label', subj_key)
            
            outcome_stats = {}
            for ans in subj_data.get('answers', []):
                outcome = ans.get('outcome', '')
                if not outcome:
                    continue
                if outcome not in outcome_stats:
                    outcome_stats[outcome] = {'correct': 0, 'wrong': 0, 'blank': 0, 'total': 0}
                outcome_stats[outcome]['total'] += 1
                if ans.get('status') == 'correct':
                    outcome_stats[outcome]['correct'] += 1
                elif ans.get('status') == 'wrong':
                    outcome_stats[outcome]['wrong'] += 1
                else:
                    outcome_stats[outcome]['blank'] += 1
            
            if outcome_stats:
                elements.append(Paragraph(subject_label, subject_style))
                
                table_data = [['Kazanım', 'D', 'Y', 'B', '%']]
                for outcome, stats in outcome_stats.items():
                    total = stats['total']
                    rate = round((stats['correct'] / total) * 100) if total > 0 else 0
                    table_data.append([outcome[:60], str(stats['correct']), str(stats['wrong']), str(stats['blank']), f'%{rate}'])
                
                t = Table(table_data, colWidths=[300, 40, 40, 40, 50])
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dbeafe')),
                    ('FONTNAME', (0, 0), (-1, -1), PDF_FONT),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#93c5fd')),
                    ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 4),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ]))
                elements.append(t)
                elements.append(Spacer(1, 10))
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f'kazanim_analizi_{result["student_name"].replace(" ", "_")}.pdf'
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)
        
    except Exception as e:
        logger.error(f"PDF oluşturma hatası: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/class-outcome-analysis/<int:exam_id>')
@report_cards_bp.route('/api/class-outcome-analysis-multi')
@login_required
def class_outcome_analysis(exam_id=None):
    """Sınıf bazlı kazanım analizi - tek veya çoklu sınav"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    class_name = request.args.get('class_name', '')
    exam_ids_str = request.args.get('exam_ids', '')
    
    if exam_ids_str:
        exam_ids = [int(x) for x in exam_ids_str.split(',') if x.strip().isdigit()]
    elif exam_id:
        exam_ids = [exam_id]
    else:
        return jsonify({"error": "Sınav seçilmedi"}), 400
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        if current_user.role == 'teacher':
            cur.execute("SELECT class_name FROM teacher_classes WHERE teacher_id = %s", (current_user.id,))
            allowed_classes = [row['class_name'] for row in cur.fetchall()]
            allowed_classes_normalized = [c.replace('/', '') for c in allowed_classes]
            
            if class_name:
                class_normalized = class_name.replace('/', '')
                if class_name not in allowed_classes and class_normalized not in allowed_classes_normalized:
                    return jsonify({"error": "Bu sınıfa erişim yetkiniz yok"}), 403
        
        placeholders = ','.join(['%s'] * len(exam_ids))
        query = f"""
            SELECT r.subjects 
            FROM report_card_results r
            WHERE r.exam_id IN ({placeholders})
        """
        params = exam_ids
        
        if class_name:
            class_normalized = class_name.replace('/', '')
            query += " AND (r.class_name = %s OR REPLACE(r.class_name, '/', '') = %s)"
            params.extend([class_name, class_normalized])
        elif current_user.role == 'teacher' and allowed_classes:
            query += " AND (r.class_name = ANY(%s) OR REPLACE(r.class_name, '/', '') = ANY(%s))"
            params.extend([allowed_classes, allowed_classes_normalized])
        
        cur.execute(query, params)
        results = cur.fetchall()
        
        outcome_stats = {}
        
        for result in results:
            subjects = result.get('subjects') or {}
            if isinstance(subjects, str):
                try:
                    subjects = json.loads(subjects)
                except:
                    continue
            
            for subj_key, subj_data in subjects.items():
                subject_label = subj_data.get('subject_label', subj_key)
                
                if subject_label not in outcome_stats:
                    outcome_stats[subject_label] = {}
                
                for ans in subj_data.get('answers', []):
                    outcome = ans.get('outcome', '')
                    if not outcome:
                        continue
                    
                    if outcome not in outcome_stats[subject_label]:
                        outcome_stats[subject_label][outcome] = {'correct': 0, 'wrong': 0, 'blank': 0, 'total': 0}
                    
                    outcome_stats[subject_label][outcome]['total'] += 1
                    if ans.get('status') == 'correct':
                        outcome_stats[subject_label][outcome]['correct'] += 1
                    elif ans.get('status') == 'wrong':
                        outcome_stats[subject_label][outcome]['wrong'] += 1
                    else:
                        outcome_stats[subject_label][outcome]['blank'] += 1
        
        result_data = {}
        for subject, outcomes in outcome_stats.items():
            result_data[subject] = []
            for outcome, stats in outcomes.items():
                result_data[subject].append({
                    'outcome': outcome,
                    'correct': stats['correct'],
                    'wrong': stats['wrong'],
                    'blank': stats['blank'],
                    'total': stats['total']
                })
            result_data[subject].sort(key=lambda x: parse_outcome_code(x['outcome']))
        
        return jsonify({"subjects": result_data, "student_count": len(results)})
        
    except Exception as e:
        logger.error(f"Sınıf kazanım analizi hatası: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/student-exams')
@login_required
def get_student_exams():
    """Öğrencinin katıldığı sınavları listele"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    student_no = request.args.get('student_no', '')
    class_name = request.args.get('class_name', '')
    
    if not student_no:
        return jsonify({"error": "Öğrenci numarası gerekli"}), 400
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        class_normalized = class_name.replace('/', '') if class_name else ''
        cur.execute("""
            SELECT r.id as result_id, r.exam_id, e.exam_name
            FROM report_card_results r
            JOIN report_card_exams e ON r.exam_id = e.id
            WHERE r.student_no = %s 
            AND (r.class_name = %s OR REPLACE(r.class_name, '/', '') = %s OR %s = '')
            ORDER BY e.created_at DESC
        """, (student_no, class_name, class_normalized, class_name))
        exams = cur.fetchall()
        
        return jsonify({"exams": exams})
        
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/student-multi-outcome-by-result')
@login_required
def student_multi_outcome():
    """Öğrenci çoklu sınav kazanım analizi (result_ids ile - eski versiyon)"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    result_ids_str = request.args.get('result_ids', '')
    if not result_ids_str:
        return jsonify({"error": "Sonuç ID'leri gerekli"}), 400
    
    result_ids = [int(x) for x in result_ids_str.split(',') if x.strip().isdigit()]
    if not result_ids:
        return jsonify({"error": "Geçerli sonuç ID'si yok"}), 400
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        placeholders = ','.join(['%s'] * len(result_ids))
        cur.execute(f"""
            SELECT r.subjects, e.exam_name
            FROM report_card_results r
            JOIN report_card_exams e ON r.exam_id = e.id
            WHERE r.id IN ({placeholders})
        """, result_ids)
        results = cur.fetchall()
        
        outcome_data = {}
        
        for result in results:
            subjects = result.get('subjects') or {}
            if isinstance(subjects, str):
                try:
                    subjects = json.loads(subjects)
                except:
                    continue
            
            for subj_key, subj_data in subjects.items():
                subject_label = subj_data.get('subject_label', subj_key)
                
                if subject_label not in outcome_data:
                    outcome_data[subject_label] = {}
                
                for ans in subj_data.get('answers', []):
                    outcome = ans.get('outcome', '')
                    if not outcome:
                        continue
                    
                    if outcome not in outcome_data[subject_label]:
                        outcome_data[subject_label][outcome] = {'correct': 0, 'wrong': 0, 'blank': 0, 'total': 0}
                    
                    outcome_data[subject_label][outcome]['total'] += 1
                    if ans.get('status') == 'correct':
                        outcome_data[subject_label][outcome]['correct'] += 1
                    elif ans.get('status') == 'wrong':
                        outcome_data[subject_label][outcome]['wrong'] += 1
                    else:
                        outcome_data[subject_label][outcome]['blank'] += 1
        
        for subject in outcome_data:
            for outcome in outcome_data[subject]:
                stats = outcome_data[subject][outcome]
                total = stats['total']
                stats['success_rate'] = round((stats['correct'] / total * 100), 1) if total > 0 else 0
        
        return jsonify({"outcome_data": outcome_data})
        
    except Exception as e:
        logger.error(f"Öğrenci çoklu kazanım hatası: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/student-multi-error-by-result')
@login_required
def student_multi_error():
    """Öğrenci çoklu sınav hata karnesi (result_ids ile - eski versiyon)"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    result_ids_str = request.args.get('result_ids', '')
    if not result_ids_str:
        return jsonify({"error": "Sonuç ID'leri gerekli"}), 400
    
    result_ids = [int(x) for x in result_ids_str.split(',') if x.strip().isdigit()]
    if not result_ids:
        return jsonify({"error": "Geçerli sonuç ID'si yok"}), 400
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        placeholders = ','.join(['%s'] * len(result_ids))
        cur.execute(f"""
            SELECT r.subjects, e.exam_name
            FROM report_card_results r
            JOIN report_card_exams e ON r.exam_id = e.id
            WHERE r.id IN ({placeholders})
        """, result_ids)
        results = cur.fetchall()
        
        errors_by_subject = {}
        total_correct = 0
        total_wrong = 0
        total_blank = 0
        
        for result in results:
            exam_name = result.get('exam_name', '')
            subjects = result.get('subjects') or {}
            if isinstance(subjects, str):
                try:
                    subjects = json.loads(subjects)
                except:
                    continue
            
            for subj_key, subj_data in subjects.items():
                subject_label = subj_data.get('subject_label', subj_key)
                
                if subject_label not in errors_by_subject:
                    errors_by_subject[subject_label] = {
                        'subject_label': subject_label,
                        'errors': []
                    }
                
                for ans in subj_data.get('answers', []):
                    status = ans.get('status', '')
                    if status == 'correct':
                        total_correct += 1
                    elif status == 'wrong':
                        total_wrong += 1
                        errors_by_subject[subject_label]['errors'].append({
                            'question_number': ans.get('question_number'),
                            'exam_name': exam_name,
                            'correct_answer': ans.get('correct_answer'),
                            'student_answer': ans.get('student_answer', ''),
                            'outcome': ans.get('outcome', ''),
                            'status': 'wrong'
                        })
                    else:
                        total_blank += 1
                        errors_by_subject[subject_label]['errors'].append({
                            'question_number': ans.get('question_number'),
                            'exam_name': exam_name,
                            'correct_answer': ans.get('correct_answer'),
                            'student_answer': ans.get('student_answer', ''),
                            'outcome': ans.get('outcome', ''),
                            'status': 'blank'
                        })
        
        errors_by_subject = {k: v for k, v in errors_by_subject.items() if v['errors']}
        
        for subj in errors_by_subject.values():
            subj['errors'].sort(key=lambda x: parse_outcome_code(x.get('outcome', '')))
        
        net = total_correct - (total_wrong / 4)
        
        return jsonify({
            "errors_by_subject": errors_by_subject,
            "totals": {
                "correct": total_correct,
                "wrong": total_wrong,
                "blank": total_blank,
                "net": net
            }
        })
        
    except Exception as e:
        logger.error(f"Öğrenci çoklu hata karnesi hatası: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


# ==================== ÖĞRENCİ ÇOKLU SINAV ANALİZİ (student_no + exam_ids) ====================

@report_cards_bp.route('/api/student-multi-outcome')
@login_required
def student_multi_outcome_by_no():
    """Öğrenci çoklu sınav kazanım analizi (student_no ve exam_ids ile)"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    student_no = request.args.get('student_no', '')
    exam_ids_str = request.args.get('exam_ids', '')
    
    if not student_no:
        return jsonify({"error": "Öğrenci numarası gerekli"}), 400
    if not exam_ids_str:
        return jsonify({"error": "Sınav ID'leri gerekli"}), 400
    
    exam_ids = [int(x) for x in exam_ids_str.split(',') if x.strip().isdigit()]
    if not exam_ids:
        return jsonify({"error": "Geçerli sınav ID'si yok"}), 400
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        placeholders = ','.join(['%s'] * len(exam_ids))
        cur.execute(f"""
            SELECT r.subjects, r.student_name, e.exam_name
            FROM report_card_results r
            JOIN report_card_exams e ON r.exam_id = e.id
            WHERE r.student_no = %s AND r.exam_id IN ({placeholders})
        """, [student_no] + exam_ids)
        results = cur.fetchall()
        
        if not results:
            return jsonify({"error": "Veri bulunamadı"}), 404
        
        student_name = results[0].get('student_name', student_no)
        outcome_data = {}
        
        for result in results:
            subjects = result.get('subjects') or {}
            if isinstance(subjects, str):
                subjects = json.loads(subjects)
            
            for subj_key, subj_data in subjects.items():
                subject_label = subj_data.get('subject_label', subj_key)
                if subject_label not in outcome_data:
                    outcome_data[subject_label] = {}
                
                for ans in subj_data.get('answers', []):
                    outcome = ans.get('outcome', '') or '-'
                    if outcome not in outcome_data[subject_label]:
                        outcome_data[subject_label][outcome] = {'correct': 0, 'wrong': 0, 'blank': 0, 'total': 0}
                    
                    outcome_data[subject_label][outcome]['total'] += 1
                    if ans.get('status') == 'correct':
                        outcome_data[subject_label][outcome]['correct'] += 1
                    elif ans.get('status') == 'wrong':
                        outcome_data[subject_label][outcome]['wrong'] += 1
                    else:
                        outcome_data[subject_label][outcome]['blank'] += 1
        
        sorted_outcome_data = {}
        for subj, outcomes in outcome_data.items():
            sorted_outcome_data[subj] = dict(sorted(outcomes.items(), key=lambda x: parse_outcome_code(x[0])))
        
        return jsonify({"student_name": student_name, "outcomes": sorted_outcome_data})
        
    except Exception as e:
        logger.error(f"Öğrenci kazanım analizi hatası: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/student-multi-error')
@login_required
def student_multi_error_by_no():
    """Öğrenci çoklu sınav hata karnesi (student_no ve exam_ids ile)"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    student_no = request.args.get('student_no', '')
    exam_ids_str = request.args.get('exam_ids', '')
    
    if not student_no:
        return jsonify({"error": "Öğrenci numarası gerekli"}), 400
    if not exam_ids_str:
        return jsonify({"error": "Sınav ID'leri gerekli"}), 400
    
    exam_ids = [int(x) for x in exam_ids_str.split(',') if x.strip().isdigit()]
    if not exam_ids:
        return jsonify({"error": "Geçerli sınav ID'si yok"}), 400
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        placeholders = ','.join(['%s'] * len(exam_ids))
        cur.execute(f"""
            SELECT r.subjects, r.student_name, e.exam_name
            FROM report_card_results r
            JOIN report_card_exams e ON r.exam_id = e.id
            WHERE r.student_no = %s AND r.exam_id IN ({placeholders})
        """, [student_no] + exam_ids)
        results = cur.fetchall()
        
        if not results:
            return jsonify({"error": "Veri bulunamadı"}), 404
        
        student_name = results[0].get('student_name', student_no)
        error_list = {}
        total_correct = 0
        total_wrong = 0
        total_blank = 0
        
        for result in results:
            exam_name = result.get('exam_name', '')
            subjects = result.get('subjects') or {}
            if isinstance(subjects, str):
                subjects = json.loads(subjects)
            
            for subj_key, subj_data in subjects.items():
                subject_label = subj_data.get('subject_label', subj_key)
                if subject_label not in error_list:
                    error_list[subject_label] = []
                
                for ans in subj_data.get('answers', []):
                    q_no = ans.get('question_number', '-')
                    outcome = ans.get('outcome', '') or '-'
                    status = ans.get('status', '')
                    
                    if status == 'correct':
                        total_correct += 1
                    elif status == 'wrong':
                        total_wrong += 1
                        error_list[subject_label].append({
                            'question_number': q_no,
                            'exam_name': exam_name,
                            'outcome': outcome,
                            'status': 'wrong',
                            'student_answer': ans.get('student_answer', '-'),
                            'correct_answer': ans.get('correct_answer', '-')
                        })
                    else:
                        total_blank += 1
                        error_list[subject_label].append({
                            'question_number': q_no,
                            'exam_name': exam_name,
                            'outcome': outcome,
                            'status': 'blank',
                            'student_answer': '-',
                            'correct_answer': ans.get('correct_answer', '-')
                        })
        
        error_list = {k: v for k, v in error_list.items() if v}
        
        for subj in error_list.values():
            subj.sort(key=lambda x: parse_outcome_code(x.get('outcome', '')))
        
        return jsonify({
            "student_name": student_name,
            "errors": error_list,
            "summary": {"correct": total_correct, "wrong": total_wrong, "blank": total_blank}
        })
        
    except Exception as e:
        logger.error(f"Öğrenci hata karnesi hatası: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/student-multi-outcome-pdf')
@login_required
def student_multi_outcome_pdf():
    """Öğrenci çoklu sınav kazanım analizi PDF"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    student_no = request.args.get('student_no', '')
    exam_ids_str = request.args.get('exam_ids', '')
    
    if not student_no or not exam_ids_str:
        return jsonify({"error": "Parametreler eksik"}), 400
    
    exam_ids = [int(x) for x in exam_ids_str.split(',') if x.strip().isdigit()]
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        placeholders = ','.join(['%s'] * len(exam_ids))
        cur.execute(f"""
            SELECT r.subjects, r.student_name, r.class_name, e.exam_name
            FROM report_card_results r
            JOIN report_card_exams e ON r.exam_id = e.id
            WHERE r.student_no = %s AND r.exam_id IN ({placeholders})
        """, [student_no] + exam_ids)
        results = cur.fetchall()
        
        if not results:
            return jsonify({"error": "Veri bulunamadı"}), 404
        
        student_name = results[0].get('student_name', student_no)
        class_name = results[0].get('class_name', '')
        
        outcome_data = {}
        for result in results:
            subjects = result.get('subjects') or {}
            if isinstance(subjects, str):
                subjects = json.loads(subjects)
            
            for subj_key, subj_data in subjects.items():
                subject_label = subj_data.get('subject_label', subj_key)
                if subject_label not in outcome_data:
                    outcome_data[subject_label] = {}
                
                for ans in subj_data.get('answers', []):
                    outcome = ans.get('outcome', '') or '-'
                    if outcome not in outcome_data[subject_label]:
                        outcome_data[subject_label][outcome] = {'correct': 0, 'wrong': 0, 'blank': 0, 'total': 0}
                    
                    outcome_data[subject_label][outcome]['total'] += 1
                    if ans.get('status') == 'correct':
                        outcome_data[subject_label][outcome]['correct'] += 1
                    elif ans.get('status') == 'wrong':
                        outcome_data[subject_label][outcome]['wrong'] += 1
                    else:
                        outcome_data[subject_label][outcome]['blank'] += 1
        
        sorted_outcome_data = {}
        for subj, outcomes in outcome_data.items():
            sorted_outcome_data[subj] = dict(sorted(outcomes.items(), key=lambda x: parse_outcome_code(x[0])))
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30, bottomMargin=30, leftMargin=30, rightMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        
        elements.extend(create_pdf_header(styles))
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontName=PDF_FONT, fontSize=14, alignment=TA_CENTER, spaceAfter=12)
        subject_style = ParagraphStyle('Subject', parent=styles['Heading2'], fontName=PDF_FONT, fontSize=11, textColor=colors.HexColor('#6366f1'), spaceAfter=6)
        normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontName=PDF_FONT, fontSize=9)
        cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontName=PDF_FONT, fontSize=7, leading=9, wordWrap='CJK')
        
        elements.append(Paragraph("ÖĞRENCİ KAZANIM ANALİZİ", title_style))
        elements.append(Paragraph(f"Öğrenci: {student_name} | Sınıf: {class_name}", normal_style))
        elements.append(Paragraph(f"Tarih: {datetime.now().strftime('%d.%m.%Y')}", normal_style))
        elements.append(Spacer(1, 12))
        
        for subject, outcomes in sorted_outcome_data.items():
            elements.append(Paragraph(subject, subject_style))
            
            table_data = [['Kazanım', 'Doğru', 'Yanlış', 'Boş', 'Başarı %']]
            for outcome, stats in outcomes.items():
                total = stats['total'] or 1
                success = round((stats['correct'] / total * 100), 1)
                table_data.append([
                    Paragraph(outcome, cell_style),
                    str(stats['correct']),
                    str(stats['wrong']),
                    str(stats['blank']),
                    f"%{success}"
                ])
            
            t = Table(table_data, colWidths=[300, 40, 40, 40, 50])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e0e7ff')),
                ('FONTNAME', (0, 0), (-1, -1), PDF_FONT),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 8))
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f'ogrenci_kazanim_{student_no}_{datetime.now().strftime("%Y%m%d")}.pdf'
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)
        
    except Exception as e:
        logger.error(f"Öğrenci kazanım PDF hatası: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/student-multi-error-pdf')
@login_required
def student_multi_error_pdf():
    """Öğrenci çoklu sınav hata karnesi PDF"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    student_no = request.args.get('student_no', '')
    exam_ids_str = request.args.get('exam_ids', '')
    
    if not student_no or not exam_ids_str:
        return jsonify({"error": "Parametreler eksik"}), 400
    
    exam_ids = [int(x) for x in exam_ids_str.split(',') if x.strip().isdigit()]
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        placeholders = ','.join(['%s'] * len(exam_ids))
        cur.execute(f"""
            SELECT r.subjects, r.student_name, r.class_name, e.exam_name
            FROM report_card_results r
            JOIN report_card_exams e ON r.exam_id = e.id
            WHERE r.student_no = %s AND r.exam_id IN ({placeholders})
        """, [student_no] + exam_ids)
        results = cur.fetchall()
        
        if not results:
            return jsonify({"error": "Veri bulunamadı"}), 404
        
        student_name = results[0].get('student_name', student_no)
        class_name = results[0].get('class_name', '')
        
        error_list = {}
        total_correct = 0
        total_wrong = 0
        total_blank = 0
        
        for result in results:
            exam_name = result.get('exam_name', '')
            subjects = result.get('subjects') or {}
            if isinstance(subjects, str):
                subjects = json.loads(subjects)
            
            for subj_key, subj_data in subjects.items():
                subject_label = subj_data.get('subject_label', subj_key)
                if subject_label not in error_list:
                    error_list[subject_label] = []
                
                for ans in subj_data.get('answers', []):
                    q_no = ans.get('question_number', '-')
                    outcome = ans.get('outcome', '') or '-'
                    status = ans.get('status', '')
                    
                    if status == 'correct':
                        total_correct += 1
                    elif status == 'wrong':
                        total_wrong += 1
                        error_list[subject_label].append({
                            'exam_name': exam_name,
                            'question_number': q_no,
                            'outcome': outcome,
                            'status': 'Hatalı'
                        })
                    else:
                        total_blank += 1
                        error_list[subject_label].append({
                            'exam_name': exam_name,
                            'question_number': q_no,
                            'outcome': outcome,
                            'status': 'Boş'
                        })
        
        error_list = {k: v for k, v in error_list.items() if v}
        
        for subj_err_list in error_list.values():
            subj_err_list.sort(key=lambda x: parse_outcome_code(x.get('outcome', '')))
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30, bottomMargin=30, leftMargin=30, rightMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        
        elements.extend(create_pdf_header(styles))
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontName=PDF_FONT, fontSize=14, alignment=TA_CENTER, spaceAfter=12)
        subject_style = ParagraphStyle('Subject', parent=styles['Heading2'], fontName=PDF_FONT, fontSize=11, textColor=colors.HexColor('#dc2626'), spaceAfter=6)
        normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontName=PDF_FONT, fontSize=9)
        cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontName=PDF_FONT, fontSize=7, leading=9, wordWrap='CJK')
        
        elements.append(Paragraph("HATA KARNESİ", title_style))
        elements.append(Paragraph(f"Öğrenci: {student_name} | Sınıf: {class_name}", normal_style))
        elements.append(Paragraph(f"Toplam: {total_correct} Doğru / {total_wrong} Hatalı / {total_blank} Boş", normal_style))
        elements.append(Paragraph(f"Tarih: {datetime.now().strftime('%d.%m.%Y')}", normal_style))
        elements.append(Spacer(1, 12))
        
        for subject, errors in error_list.items():
            elements.append(Paragraph(f"{subject} ({len(errors)} hata)", subject_style))
            
            table_data = [['Sınav', 'Soru', 'Kazanım', 'Durum']]
            for err in errors:
                table_data.append([
                    Paragraph(err['exam_name'], cell_style),
                    f"{err['question_number']}. Soru",
                    Paragraph(err['outcome'], cell_style),
                    err['status']
                ])
            
            t = Table(table_data, colWidths=[120, 50, 260, 50])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fee2e2')),
                ('FONTNAME', (0, 0), (-1, -1), PDF_FONT),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#fca5a5')),
                ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                ('ALIGN', (3, 0), (3, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 8))
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f'ogrenci_hata_{student_no}_{datetime.now().strftime("%Y%m%d")}.pdf'
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)
        
    except Exception as e:
        logger.error(f"Öğrenci hata PDF hatası: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/class-error-analysis-multi')
@login_required
def class_error_analysis_multi():
    """Sınıf bazlı hata analizi - çoklu sınav desteği"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    class_name = request.args.get('class_name', '')
    exam_ids_str = request.args.get('exam_ids', '')
    
    if not exam_ids_str:
        return jsonify({"error": "Sınav seçilmedi"}), 400
    
    exam_ids = [int(x) for x in exam_ids_str.split(',') if x.strip().isdigit()]
    if not exam_ids:
        return jsonify({"error": "Geçerli sınav seçilmedi"}), 400
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        if current_user.role == 'teacher':
            cur.execute("SELECT class_name FROM teacher_classes WHERE teacher_id = %s", (current_user.id,))
            allowed_classes = [row['class_name'] for row in cur.fetchall()]
            allowed_classes_normalized = [c.replace('/', '') for c in allowed_classes]
            
            if class_name:
                class_normalized = class_name.replace('/', '')
                if class_name not in allowed_classes and class_normalized not in allowed_classes_normalized:
                    return jsonify({"error": "Bu sınıfa erişim yetkiniz yok"}), 403
        
        placeholders = ','.join(['%s'] * len(exam_ids))
        query = f"""
            SELECT r.subjects, r.class_name, e.exam_name
            FROM report_card_results r
            JOIN report_card_exams e ON r.exam_id = e.id
            WHERE r.exam_id IN ({placeholders})
        """
        params = list(exam_ids)
        
        if class_name:
            class_normalized = class_name.replace('/', '')
            query += " AND (r.class_name = %s OR REPLACE(r.class_name, '/', '') = %s)"
            params.extend([class_name, class_normalized])
        elif current_user.role == 'teacher' and allowed_classes:
            query += " AND (r.class_name = ANY(%s) OR REPLACE(r.class_name, '/', '') = ANY(%s))"
            params.extend([allowed_classes, allowed_classes_normalized])
        
        cur.execute(query, params)
        results = cur.fetchall()
        
        error_stats = {}
        total_correct = 0
        total_wrong = 0
        total_blank = 0
        
        for result in results:
            exam_name = result.get('exam_name', '')
            subjects = result.get('subjects') or {}
            if isinstance(subjects, str):
                try:
                    subjects = json.loads(subjects)
                except:
                    continue
            
            for subj_key, subj_data in subjects.items():
                subject_label = subj_data.get('subject_label', subj_key)
                
                if subject_label not in error_stats:
                    error_stats[subject_label] = {}
                
                for ans in subj_data.get('answers', []):
                    q_num = ans.get('question_number', 0)
                    outcome = ans.get('outcome', '')
                    status = ans.get('status', '')
                    
                    key = f"{exam_name}|{q_num}"
                    
                    if key not in error_stats[subject_label]:
                        error_stats[subject_label][key] = {
                            'exam_name': exam_name,
                            'question_number': q_num,
                            'outcome': outcome,
                            'correct': 0,
                            'wrong': 0,
                            'blank': 0,
                            'total': 0
                        }
                    
                    error_stats[subject_label][key]['total'] += 1
                    if status == 'correct':
                        error_stats[subject_label][key]['correct'] += 1
                        total_correct += 1
                    elif status == 'wrong':
                        error_stats[subject_label][key]['wrong'] += 1
                        total_wrong += 1
                    else:
                        error_stats[subject_label][key]['blank'] += 1
                        total_blank += 1
        
        result_data = {}
        for subject, questions in error_stats.items():
            result_data[subject] = []
            for key, stats in sorted(questions.items(), key=lambda x: (parse_outcome_code(x[1]['outcome']), x[1]['exam_name'], x[1]['question_number'])):
                result_data[subject].append({
                    'exam_name': stats['exam_name'],
                    'question_number': stats['question_number'],
                    'outcome': stats['outcome'],
                    'correct': stats['correct'],
                    'wrong': stats['wrong'],
                    'blank': stats['blank'],
                    'total': stats['total']
                })
        
        return jsonify({
            "subjects": result_data,
            "student_count": len(results),
            "totals": {
                "correct": total_correct,
                "wrong": total_wrong,
                "blank": total_blank
            }
        })
        
    except Exception as e:
        logger.error(f"Sınıf hata analizi hatası: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/class-error-pdf')
@login_required
def class_error_pdf_multi():
    """Sınıf bazlı hata analizi PDF - Çoklu sınav destekli"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    exam_ids_str = request.args.get('exam_ids', '')
    class_name = request.args.get('class_name', '')
    
    if not exam_ids_str:
        return jsonify({"error": "Sınav ID'leri gerekli"}), 400
    
    exam_ids = [int(x) for x in exam_ids_str.split(',') if x.strip().isdigit()]
    if not exam_ids:
        return jsonify({"error": "Geçerli sınav ID'si yok"}), 400
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        placeholders = ','.join(['%s'] * len(exam_ids))
        query = f"""
            SELECT r.subjects, r.class_name, e.exam_name
            FROM report_card_results r
            JOIN report_card_exams e ON r.exam_id = e.id
            WHERE r.exam_id IN ({placeholders})
        """
        params = list(exam_ids)
        
        if class_name:
            class_normalized = class_name.replace('/', '')
            query += " AND (r.class_name = %s OR REPLACE(r.class_name, '/', '') = %s)"
            params.extend([class_name, class_normalized])
        
        cur.execute(query, params)
        results = cur.fetchall()
        
        if not results:
            return jsonify({"error": "Veri bulunamadı"}), 404
        
        error_stats = {}
        total_correct = 0
        total_wrong = 0
        total_blank = 0
        student_count = len(results)
        
        for result in results:
            subjects = result.get('subjects') or {}
            if isinstance(subjects, str):
                subjects = json.loads(subjects)
            
            for subj_key, subj_data in subjects.items():
                subject_label = subj_data.get('subject_label', subj_key)
                if subject_label not in error_stats:
                    error_stats[subject_label] = {}
                
                for ans in subj_data.get('answers', []):
                    q_no = ans.get('question_number', '-')
                    outcome = ans.get('outcome', '') or '-'
                    key = f"{q_no}_{outcome}"
                    
                    if key not in error_stats[subject_label]:
                        error_stats[subject_label][key] = {'question_number': q_no, 'outcome': outcome, 'correct': 0, 'wrong': 0, 'blank': 0, 'total': 0}
                    
                    error_stats[subject_label][key]['total'] += 1
                    if ans.get('status') == 'correct':
                        error_stats[subject_label][key]['correct'] += 1
                        total_correct += 1
                    elif ans.get('status') == 'wrong':
                        error_stats[subject_label][key]['wrong'] += 1
                        total_wrong += 1
                    else:
                        error_stats[subject_label][key]['blank'] += 1
                        total_blank += 1
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30, bottomMargin=30, leftMargin=30, rightMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        
        elements.extend(create_pdf_header(styles))
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontName=PDF_FONT, fontSize=16, alignment=TA_CENTER, spaceAfter=15)
        subject_style = ParagraphStyle('Subject', parent=styles['Heading2'], fontName=PDF_FONT, fontSize=12, textColor=colors.HexColor('#dc2626'), spaceAfter=8, spaceBefore=12)
        normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontName=PDF_FONT, fontSize=10)
        cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontName=PDF_FONT, fontSize=7, leading=9, wordWrap='CJK')
        
        elements.append(Paragraph("SINIF HATA ANALİZİ", title_style))
        elements.append(Paragraph(f"Sınıf: {class_name or 'Tüm Sınıflar'} | Öğrenci: {student_count}", normal_style))
        elements.append(Paragraph(f"Toplam: {total_correct} Doğru / {total_wrong} Yanlış / {total_blank} Boş", normal_style))
        elements.append(Paragraph(f"Tarih: {datetime.now().strftime('%d.%m.%Y')}", normal_style))
        elements.append(Spacer(1, 15))
        
        for subject, questions in error_stats.items():
            elements.append(Paragraph(subject, subject_style))
            
            sorted_questions = sorted(questions.values(), key=lambda x: x['wrong'] + x['blank'], reverse=True)
            
            table_data = [['S.No', 'Kazanım', 'Yanlış', 'Boş', 'Hata %']]
            for q in sorted_questions[:20]:
                total = q['total']
                error_rate = round(((q['wrong'] + q['blank']) / total * 100), 1) if total > 0 else 0
                table_data.append([
                    str(q['question_number']),
                    Paragraph(q['outcome'], cell_style),
                    str(q['wrong']),
                    str(q['blank']),
                    f"%{error_rate}"
                ])
            
            t = Table(table_data, colWidths=[40, 305, 45, 45, 50])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fee2e2')),
                ('FONTNAME', (0, 0), (-1, -1), PDF_FONT),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#fca5a5')),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 10))
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f'sinif_hata_{class_name.replace("/", "_") if class_name else "tum"}_{datetime.now().strftime("%Y%m%d")}.pdf'
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)
        
    except Exception as e:
        logger.error(f"Sınıf hata PDF hatası: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


# ==================== ÖĞRENCİ ÇOKLU SINAV API'LERİ ====================

@report_cards_bp.route('/api/my-multi-outcome')
@login_required
def my_multi_outcome():
    """Öğrenci kendi çoklu sınav kazanım analizi"""
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    result_ids_str = request.args.get('result_ids', '')
    if not result_ids_str:
        return jsonify({"error": "Sonuç ID'leri gerekli"}), 400
    
    result_ids = [int(x) for x in result_ids_str.split(',') if x.strip().isdigit()]
    if not result_ids:
        return jsonify({"error": "Geçerli sonuç ID'si yok"}), 400
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("SELECT student_no, class_name FROM users WHERE id = %s", (current_user.id,))
        user_data = cur.fetchone()
        student_no = user_data.get('student_no', '') if user_data else ''
        
        placeholders = ','.join(['%s'] * len(result_ids))
        cur.execute(f"""
            SELECT r.subjects, e.exam_name
            FROM report_card_results r
            JOIN report_card_exams e ON r.exam_id = e.id
            WHERE r.id IN ({placeholders}) AND r.student_no = %s
        """, result_ids + [student_no])
        results = cur.fetchall()
        
        if not results:
            return jsonify({"error": "Sonuç bulunamadı veya erişim yetkiniz yok"}), 404
        
        outcome_data = {}
        
        for result in results:
            subjects = result.get('subjects') or {}
            if isinstance(subjects, str):
                try:
                    subjects = json.loads(subjects)
                except:
                    continue
            
            for subj_key, subj_data in subjects.items():
                subject_label = subj_data.get('subject_label', subj_key)
                
                if subject_label not in outcome_data:
                    outcome_data[subject_label] = {}
                
                for ans in subj_data.get('answers', []):
                    outcome = ans.get('outcome', '')
                    if not outcome:
                        continue
                    
                    if outcome not in outcome_data[subject_label]:
                        outcome_data[subject_label][outcome] = {'correct': 0, 'wrong': 0, 'blank': 0, 'total': 0}
                    
                    outcome_data[subject_label][outcome]['total'] += 1
                    if ans.get('status') == 'correct':
                        outcome_data[subject_label][outcome]['correct'] += 1
                    elif ans.get('status') == 'wrong':
                        outcome_data[subject_label][outcome]['wrong'] += 1
                    else:
                        outcome_data[subject_label][outcome]['blank'] += 1
        
        for subject in outcome_data:
            for outcome in outcome_data[subject]:
                stats = outcome_data[subject][outcome]
                total = stats['total']
                stats['success_rate'] = round((stats['correct'] / total * 100), 1) if total > 0 else 0
        
        return jsonify({"outcome_data": outcome_data})
        
    except Exception as e:
        logger.error(f"Öğrenci çoklu kazanım hatası: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/my-multi-error')
@login_required
def my_multi_error():
    """Öğrenci kendi çoklu sınav hata karnesi"""
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    result_ids_str = request.args.get('result_ids', '')
    if not result_ids_str:
        return jsonify({"error": "Sonuç ID'leri gerekli"}), 400
    
    result_ids = [int(x) for x in result_ids_str.split(',') if x.strip().isdigit()]
    if not result_ids:
        return jsonify({"error": "Geçerli sonuç ID'si yok"}), 400
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("SELECT student_no FROM users WHERE id = %s", (current_user.id,))
        user_data = cur.fetchone()
        student_no = user_data.get('student_no', '') if user_data else ''
        
        placeholders = ','.join(['%s'] * len(result_ids))
        cur.execute(f"""
            SELECT r.subjects, e.exam_name
            FROM report_card_results r
            JOIN report_card_exams e ON r.exam_id = e.id
            WHERE r.id IN ({placeholders}) AND r.student_no = %s
        """, result_ids + [student_no])
        results = cur.fetchall()
        
        if not results:
            return jsonify({"error": "Sonuç bulunamadı veya erişim yetkiniz yok"}), 404
        
        errors_by_subject = {}
        total_correct = 0
        total_wrong = 0
        total_blank = 0
        
        for result in results:
            exam_name = result.get('exam_name', '')
            subjects = result.get('subjects') or {}
            if isinstance(subjects, str):
                try:
                    subjects = json.loads(subjects)
                except:
                    continue
            
            for subj_key, subj_data in subjects.items():
                subject_label = subj_data.get('subject_label', subj_key)
                
                if subject_label not in errors_by_subject:
                    errors_by_subject[subject_label] = {'subject_label': subject_label, 'errors': []}
                
                for ans in subj_data.get('answers', []):
                    status = ans.get('status', '')
                    if status == 'correct':
                        total_correct += 1
                    elif status == 'wrong':
                        total_wrong += 1
                        errors_by_subject[subject_label]['errors'].append({
                            'question_number': ans.get('question_number'),
                            'exam_name': exam_name,
                            'correct_answer': ans.get('correct_answer'),
                            'student_answer': ans.get('student_answer', ''),
                            'status': 'wrong'
                        })
                    else:
                        total_blank += 1
                        errors_by_subject[subject_label]['errors'].append({
                            'question_number': ans.get('question_number'),
                            'exam_name': exam_name,
                            'correct_answer': ans.get('correct_answer'),
                            'student_answer': ans.get('student_answer', ''),
                            'status': 'blank'
                        })
        
        errors_by_subject = {k: v for k, v in errors_by_subject.items() if v['errors']}
        net = total_correct - (total_wrong / 4)
        
        return jsonify({
            "errors_by_subject": errors_by_subject,
            "totals": {"correct": total_correct, "wrong": total_wrong, "blank": total_blank, "net": net}
        })
        
    except Exception as e:
        logger.error(f"Öğrenci çoklu hata karnesi hatası: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/my-multi-outcome-pdf')
@login_required
def my_multi_outcome_pdf():
    """Öğrenci çoklu sınav kazanım PDF"""
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    result_ids_str = request.args.get('result_ids', '')
    if not result_ids_str:
        return jsonify({"error": "Sonuç ID'leri gerekli"}), 400
    
    result_ids = [int(x) for x in result_ids_str.split(',') if x.strip().isdigit()]
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("SELECT student_no FROM users WHERE id = %s", (current_user.id,))
        user_data = cur.fetchone()
        student_no = user_data.get('student_no', '') if user_data else ''
        
        placeholders = ','.join(['%s'] * len(result_ids))
        cur.execute(f"""
            SELECT r.subjects, r.student_name, r.class_name, e.exam_name
            FROM report_card_results r
            JOIN report_card_exams e ON r.exam_id = e.id
            WHERE r.id IN ({placeholders}) AND r.student_no = %s
        """, result_ids + [student_no])
        results = cur.fetchall()
        
        if not results:
            return jsonify({"error": "Veri bulunamadı"}), 404
        
        student_name = results[0].get('student_name', student_no)
        class_name = results[0].get('class_name', '')
        
        outcome_data = {}
        for result in results:
            subjects = result.get('subjects') or {}
            if isinstance(subjects, str):
                subjects = json.loads(subjects)
            
            for subj_key, subj_data in subjects.items():
                subject_label = subj_data.get('subject_label', subj_key)
                if subject_label not in outcome_data:
                    outcome_data[subject_label] = {}
                
                for ans in subj_data.get('answers', []):
                    outcome = ans.get('outcome', '') or '-'
                    if outcome not in outcome_data[subject_label]:
                        outcome_data[subject_label][outcome] = {'correct': 0, 'wrong': 0, 'blank': 0, 'total': 0}
                    
                    outcome_data[subject_label][outcome]['total'] += 1
                    if ans.get('status') == 'correct':
                        outcome_data[subject_label][outcome]['correct'] += 1
                    elif ans.get('status') == 'wrong':
                        outcome_data[subject_label][outcome]['wrong'] += 1
                    else:
                        outcome_data[subject_label][outcome]['blank'] += 1
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30, bottomMargin=30, leftMargin=30, rightMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        
        elements.extend(create_pdf_header(styles))
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontName=PDF_FONT, fontSize=14, alignment=TA_CENTER, spaceAfter=12)
        subject_style = ParagraphStyle('Subject', parent=styles['Heading2'], fontName=PDF_FONT, fontSize=11, textColor=colors.HexColor('#6366f1'), spaceAfter=6)
        normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontName=PDF_FONT, fontSize=9)
        cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontName=PDF_FONT, fontSize=7, leading=9, wordWrap='CJK')
        
        elements.append(Paragraph("KAZANIM ANALİZİ", title_style))
        elements.append(Paragraph(f"Öğrenci: {student_name} | Sınıf: {class_name}", normal_style))
        elements.append(Paragraph(f"Tarih: {datetime.now().strftime('%d.%m.%Y')}", normal_style))
        elements.append(Spacer(1, 12))
        
        for subject, outcomes in outcome_data.items():
            elements.append(Paragraph(subject, subject_style))
            
            table_data = [['Kazanım', 'Doğru', 'Hatalı', 'Boş', 'Başarı %']]
            for outcome, stats in outcomes.items():
                total = stats['total'] or 1
                success = round((stats['correct'] / total * 100), 1)
                table_data.append([
                    Paragraph(outcome, cell_style),
                    str(stats['correct']),
                    str(stats['wrong']),
                    str(stats['blank']),
                    f"%{success}"
                ])
            
            t = Table(table_data, colWidths=[300, 40, 40, 40, 50])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e0e7ff')),
                ('FONTNAME', (0, 0), (-1, -1), PDF_FONT),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 8))
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f'kazanim_analizi_{datetime.now().strftime("%Y%m%d")}.pdf'
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)
        
    except Exception as e:
        logger.error(f"Öğrenci kazanım PDF hatası: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/my-multi-error-pdf')
@login_required
def my_multi_error_pdf():
    """Öğrenci çoklu sınav hata karnesi PDF"""
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    result_ids_str = request.args.get('result_ids', '')
    if not result_ids_str:
        return jsonify({"error": "Sonuç ID'leri gerekli"}), 400
    
    result_ids = [int(x) for x in result_ids_str.split(',') if x.strip().isdigit()]
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("SELECT student_no FROM users WHERE id = %s", (current_user.id,))
        user_data = cur.fetchone()
        student_no = user_data.get('student_no', '') if user_data else ''
        
        placeholders = ','.join(['%s'] * len(result_ids))
        cur.execute(f"""
            SELECT r.subjects, r.student_name, r.class_name, e.exam_name
            FROM report_card_results r
            JOIN report_card_exams e ON r.exam_id = e.id
            WHERE r.id IN ({placeholders}) AND r.student_no = %s
        """, result_ids + [student_no])
        results = cur.fetchall()
        
        if not results:
            return jsonify({"error": "Veri bulunamadı"}), 404
        
        student_name = results[0].get('student_name', student_no)
        class_name = results[0].get('class_name', '')
        
        error_list = {}
        total_correct = 0
        total_wrong = 0
        total_blank = 0
        
        for result in results:
            exam_name = result.get('exam_name', '')
            subjects = result.get('subjects') or {}
            if isinstance(subjects, str):
                subjects = json.loads(subjects)
            
            for subj_key, subj_data in subjects.items():
                subject_label = subj_data.get('subject_label', subj_key)
                if subject_label not in error_list:
                    error_list[subject_label] = []
                
                for ans in subj_data.get('answers', []):
                    q_no = ans.get('question_number', '-')
                    outcome = ans.get('outcome', '') or '-'
                    status = ans.get('status', '')
                    
                    if status == 'correct':
                        total_correct += 1
                    elif status == 'wrong':
                        total_wrong += 1
                        error_list[subject_label].append({
                            'exam_name': exam_name,
                            'question_number': q_no,
                            'outcome': outcome,
                            'status': 'Hatalı'
                        })
                    else:
                        total_blank += 1
                        error_list[subject_label].append({
                            'exam_name': exam_name,
                            'question_number': q_no,
                            'outcome': outcome,
                            'status': 'Boş'
                        })
        
        error_list = {k: v for k, v in error_list.items() if v}
        
        for subj_err_list in error_list.values():
            subj_err_list.sort(key=lambda x: parse_outcome_code(x.get('outcome', '')))
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30, bottomMargin=30, leftMargin=30, rightMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        
        elements.extend(create_pdf_header(styles))
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontName=PDF_FONT, fontSize=14, alignment=TA_CENTER, spaceAfter=12)
        subject_style = ParagraphStyle('Subject', parent=styles['Heading2'], fontName=PDF_FONT, fontSize=11, textColor=colors.HexColor('#dc2626'), spaceAfter=6)
        normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontName=PDF_FONT, fontSize=9)
        cell_style = ParagraphStyle('Cell', parent=styles['Normal'], fontName=PDF_FONT, fontSize=7, leading=9, wordWrap='CJK')
        
        elements.append(Paragraph("HATA KARNESİ", title_style))
        elements.append(Paragraph(f"Öğrenci: {student_name} | Sınıf: {class_name}", normal_style))
        elements.append(Paragraph(f"Toplam: {total_correct} Doğru / {total_wrong} Hatalı / {total_blank} Boş", normal_style))
        elements.append(Paragraph(f"Tarih: {datetime.now().strftime('%d.%m.%Y')}", normal_style))
        elements.append(Spacer(1, 12))
        
        for subject, errors in error_list.items():
            elements.append(Paragraph(f"{subject} ({len(errors)} hata)", subject_style))
            
            table_data = [['Sınav', 'Soru', 'Kazanım', 'Durum']]
            for err in errors:
                table_data.append([
                    Paragraph(err['exam_name'], cell_style),
                    f"{err['question_number']}. Soru",
                    Paragraph(err['outcome'], cell_style),
                    err['status']
                ])
            
            t = Table(table_data, colWidths=[120, 50, 260, 50])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fee2e2')),
                ('FONTNAME', (0, 0), (-1, -1), PDF_FONT),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#fca5a5')),
                ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                ('ALIGN', (3, 0), (3, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 8))
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f'hata_karnesi_{datetime.now().strftime("%Y%m%d")}.pdf'
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)
        
    except Exception as e:
        logger.error(f"Öğrenci hata PDF hatası: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@report_cards_bp.route('/api/my-progress')
@login_required
def my_progress():
    """Öğrenci gelişim raporu"""
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("SELECT student_no FROM users WHERE id = %s", (current_user.id,))
        user_data = cur.fetchone()
        student_no = user_data.get('student_no', '') if user_data else ''
        
        cur.execute("""
            SELECT r.id as result_id, e.exam_name, r.totals, e.created_at
            FROM report_card_results r
            JOIN report_card_exams e ON r.exam_id = e.id
            WHERE r.student_no = %s
            ORDER BY e.created_at ASC
        """, (student_no,))
        raw_exams = cur.fetchall()
        
        exams = []
        for exam in raw_exams:
            totals = exam.get('totals') or {}
            if isinstance(totals, str):
                import json
                totals = json.loads(totals)
            
            total_correct = totals.get('correct_count', totals.get('correct', 0)) or 0
            total_wrong = totals.get('wrong_count', totals.get('wrong', 0)) or 0
            total_blank = totals.get('blank_count', totals.get('empty', totals.get('blank', 0))) or 0
            total_net = totals.get('net', total_correct - total_wrong / 4)
            
            exams.append({
                'result_id': exam['result_id'],
                'exam_name': exam['exam_name'],
                'total_correct': total_correct,
                'total_wrong': total_wrong,
                'total_blank': total_blank,
                'total_net': round(total_net, 2) if total_net else 0,
                'created_at': exam['created_at'].isoformat() if exam.get('created_at') else None
            })
        
        return jsonify({"exams": exams})
        
    except Exception as e:
        logger.error(f"Öğrenci gelişim raporu hatası: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()
