import os
import io
import logging
import requests
from datetime import datetime
from flask import Blueprint, request, jsonify, render_template, send_file, session as flask_session, redirect
from flask_login import login_required, current_user
import psycopg2
from psycopg2.extras import RealDictCursor
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from bs4 import BeautifulSoup

logger = logging.getLogger(__name__)

lgs_results_bp = Blueprint('lgs_results', __name__)

ALLOWED_MEB_DOMAINS = [
    'sonuc.meb.gov.tr',
    'www.meb.gov.tr',
    'meb.gov.tr',
]

def is_allowed_url(url):
    try:
        from urllib.parse import urlparse
        parsed = urlparse(url)
        return parsed.hostname in ALLOWED_MEB_DOMAINS and parsed.scheme == 'https'
    except:
        return False

def get_db():
    return psycopg2.connect(os.environ.get('DATABASE_URL'))

def load_cookies_to_session(s):
    cookies_data = flask_session.get('meb_cookies', [])
    if isinstance(cookies_data, list):
        for c in cookies_data:
            s.cookies.set(c['name'], c['value'], domain=c.get('domain', ''), path=c.get('path', '/'))
    elif isinstance(cookies_data, dict):
        for k, v in cookies_data.items():
            s.cookies.set(k, v)

def save_cookies_from_session(s):
    cookies_list = []
    for cookie in s.cookies:
        cookies_list.append({
            'name': cookie.name,
            'value': cookie.value,
            'domain': cookie.domain,
            'path': cookie.path
        })
    flask_session['meb_cookies'] = cookies_list

def init_lgs_tables():
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS lgs_results (
                id SERIAL PRIMARY KEY,
                student_name VARCHAR(200),
                tc_no VARCHAR(20),
                school_no VARCHAR(20),
                birth_date DATE,
                exam_year INTEGER DEFAULT 2024,
                sonuc_tipi VARCHAR(20) DEFAULT 'sinav',
                sinav_puan DECIMAL(8,4),
                genel_yuzdelik VARCHAR(50),
                il_yuzdelik VARCHAR(50),
                turkce_dogru INTEGER DEFAULT 0,
                turkce_yanlis INTEGER DEFAULT 0,
                turkce_bos INTEGER DEFAULT 0,
                matematik_dogru INTEGER DEFAULT 0,
                matematik_yanlis INTEGER DEFAULT 0,
                matematik_bos INTEGER DEFAULT 0,
                fen_dogru INTEGER DEFAULT 0,
                fen_yanlis INTEGER DEFAULT 0,
                fen_bos INTEGER DEFAULT 0,
                inkilap_dogru INTEGER DEFAULT 0,
                inkilap_yanlis INTEGER DEFAULT 0,
                inkilap_bos INTEGER DEFAULT 0,
                din_dogru INTEGER DEFAULT 0,
                din_yanlis INTEGER DEFAULT 0,
                din_bos INTEGER DEFAULT 0,
                yabanci_dil_dogru INTEGER DEFAULT 0,
                yabanci_dil_yanlis INTEGER DEFAULT 0,
                yabanci_dil_bos INTEGER DEFAULT 0,
                yerlestigi_okul VARCHAR(500),
                baba_adi VARCHAR(200),
                yerlestirme_turu VARCHAR(200),
                okul_turu VARCHAR(200),
                bolum_alan VARCHAR(300),
                sonuc_text VARCHAR(500),
                pansiyon VARCHAR(200),
                sinif VARCHAR(10),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(tc_no, exam_year)
            )
        """)
        conn.commit()
        logger.info("LGS results table created/verified")
    except Exception as e:
        logger.error(f"Error creating LGS tables: {e}")
        if conn:
            conn.rollback()
    finally:
        if conn:
            conn.close()


@lgs_results_bp.route('/admin/lgs-results')
@login_required
def lgs_results_page():
    if current_user.role != 'admin':
        return redirect('/')
    return render_template('admin_lgs_results.html')


@lgs_results_bp.route('/admin/lgs-results/template')
@login_required
def download_lgs_template():
    if current_user.role != 'admin':
        return jsonify({'error': 'Yetkisiz'}), 403

    wb = Workbook()
    ws = wb.active
    ws.title = "LGS Öğrenci Listesi"

    headers = ['TC Kimlik No', 'Okul No', 'Doğum Tarihi (GG.AA.YYYY)', 'Sınıf']
    header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12

    example_data = [
        ['12345678901', '65', '15.03.2012', '8A'],
        ['98765432101', '65', '22.07.2012', '8B'],
    ]
    example_font = Font(name='Calibri', size=11, color='888888', italic=True)
    for row_idx, row_data in enumerate(example_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = example_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    ws.cell(row=4, column=1, value="NOT: Örnek verileri silip gerçek öğrenci bilgilerini giriniz.").font = Font(color='FF0000', bold=True, size=10)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='LGS_Ogrenci_Sablonu.xlsx'
    )


@lgs_results_bp.route('/admin/lgs-results/upload-excel', methods=['POST'])
@login_required
def upload_lgs_excel():
    if current_user.role != 'admin':
        return jsonify({'error': 'Yetkisiz'}), 403

    if 'file' not in request.files:
        return jsonify({'error': 'Dosya seçilmedi'}), 400

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls', '.xlsm')):
        return jsonify({'error': 'Geçersiz dosya formatı. Excel dosyası yükleyin.'}), 400

    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active

        students = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            tc = str(row[0].value).strip() if row[0].value else None
            okul_no = str(row[1].value).strip() if row[1].value else None
            dogum = row[2].value if len(row) > 2 and row[2].value else None
            sinif = str(row[3].value).strip() if len(row) > 3 and row[3].value else ''

            if not tc or tc == 'None' or not tc.isdigit():
                continue
            if len(tc) != 11:
                continue

            dogum_str = ''
            if dogum:
                if isinstance(dogum, datetime):
                    dogum_str = dogum.strftime('%d.%m.%Y')
                else:
                    dogum_str = str(dogum).strip()

            if dogum_str:
                import re
                parts = dogum_str.split('.')
                if len(parts) == 3:
                    g, a, y = parts
                    if len(y) > 4:
                        y = y[:4]
                    if len(g) == 1:
                        g = '0' + g
                    if len(a) == 1:
                        a = '0' + a
                    dogum_str = f"{g}.{a}.{y}"

            students.append({
                'tc': tc,
                'okul_no': okul_no or '',
                'dogum_tarihi': dogum_str,
                'sinif': sinif
            })

        if not students:
            return jsonify({'error': 'Excel dosyasında geçerli öğrenci bulunamadı'}), 400

        return jsonify({
            'success': True,
            'students': students,
            'count': len(students)
        })

    except Exception as e:
        logger.error(f"Excel upload error: {e}")
        return jsonify({'error': f'Excel okuma hatası: {str(e)}'}), 500


@lgs_results_bp.route('/admin/lgs-results/get-captcha', methods=['POST'])
@login_required
def get_meb_captcha():
    if current_user.role != 'admin':
        return jsonify({'error': 'Yetkisiz'}), 403

    data = request.get_json()
    sinav_url = data.get('sinav_url', 'https://sonuc.meb.gov.tr/')

    if not is_allowed_url(sinav_url):
        return jsonify({'error': 'Geçersiz URL. Sadece MEB alan adları desteklenir (sonuc.meb.gov.tr, www.meb.gov.tr)'}), 400

    try:
        s = requests.Session()
        resp = s.get(sinav_url, timeout=10, allow_redirects=True)

        if resp.status_code != 200:
            return jsonify({'error': f'MEB sayfası açılamadı (HTTP {resp.status_code})'}), 500

        soup = BeautifulSoup(resp.text, 'html.parser')

        sinav_id_input = soup.find('input', {'name': 'SINAV_ID'})
        sinav_id = sinav_id_input['value'] if sinav_id_input else ''

        captcha_img = soup.find('img', {'id': 'image'})
        if not captcha_img:
            captcha_img = soup.find('img', {'id': 'capcha'})
        if not captcha_img:
            captcha_img = soup.find('img', {'name': 'capcha'})
        if not captcha_img:
            for img in soup.find_all('img'):
                src = img.get('src', '')
                if 'captcha' in src.lower():
                    captcha_img = img
                    break

        gk_input = soup.find('input', {'name': 'GUVENLIKKODU'})
        has_captcha = captcha_img is not None or gk_input is not None

        captcha_base64 = None
        if captcha_img:
            captcha_src = captcha_img.get('src', '')
            if captcha_src:
                if captcha_src.startswith('/'):
                    base_url = '/'.join(resp.url.rstrip('/').split('/')[:3])
                    captcha_url = base_url + captcha_src
                else:
                    base_url = '/'.join(resp.url.rstrip('/').split('/')[:-1])
                    captcha_url = base_url + '/' + captcha_src

                captcha_resp = s.get(captcha_url, timeout=10)
                if captcha_resp.status_code == 200:
                    import base64
                    captcha_base64 = base64.b64encode(captcha_resp.content).decode('utf-8')

        save_cookies_from_session(s)
        flask_session['meb_sinav_id'] = sinav_id
        flask_session['meb_sinav_url'] = resp.url

        form_action = 'sonuc.php'
        form_tag = soup.find('form')
        if form_tag and form_tag.get('action'):
            form_action = form_tag['action']

        flask_session['meb_form_action'] = form_action

        field_tc = 'ADAY_NO'
        field_okul = 'GUVENLIKNUMARASI'
        tc_input = soup.find('input', {'name': 'ADAY_NO'})
        if not tc_input:
            tc_input = soup.find('input', {'name': 'TCNO'})
            if tc_input:
                field_tc = 'TCNO'
        okul_input = soup.find('input', {'name': 'GUVENLIKNUMARASI'})
        if not okul_input:
            okul_input = soup.find('input', {'name': 'OKULNO'})
            if okul_input:
                field_okul = 'OKULNO'

        flask_session['meb_field_tc'] = field_tc
        flask_session['meb_field_okul'] = field_okul

        return jsonify({
            'success': True,
            'has_captcha': has_captcha,
            'captcha_image': captcha_base64,
            'sinav_id': sinav_id,
            'field_tc': field_tc,
            'field_okul': field_okul
        })

    except requests.exceptions.Timeout:
        return jsonify({'error': 'MEB sunucusuna bağlanılamadı (zaman aşımı)'}), 500
    except Exception as e:
        logger.error(f"Captcha fetch error: {e}")
        return jsonify({'error': f'Bağlantı hatası: {str(e)}'}), 500


@lgs_results_bp.route('/admin/lgs-results/fetch-single', methods=['POST'])
@login_required
def fetch_single_result():
    if current_user.role != 'admin':
        return jsonify({'error': 'Yetkisiz'}), 403

    data = request.get_json()
    tc = data.get('tc', '')
    okul_no = data.get('okul_no', '')
    gun = data.get('gun', '')
    ay = data.get('ay', '')
    yil = data.get('yil', '')
    captcha_code = data.get('captcha_code', '')
    sinif = data.get('sinif', '')
    exam_year = data.get('exam_year', 2024)
    sonuc_tipi = data.get('sonuc_tipi', 'sinav')

    sinav_url = flask_session.get('meb_sinav_url', 'https://sonuc.meb.gov.tr/')
    sinav_id = flask_session.get('meb_sinav_id', '')
    form_action = flask_session.get('meb_form_action', 'sonuc.php')
    field_tc = flask_session.get('meb_field_tc', 'ADAY_NO')
    field_okul = flask_session.get('meb_field_okul', 'GUVENLIKNUMARASI')

    try:
        s = requests.Session()
        load_cookies_to_session(s)

        if form_action.startswith('http'):
            post_url = form_action
        else:
            base_url = '/'.join(sinav_url.rstrip('/').split('/')[:-1])
            post_url = base_url + '/' + form_action.lstrip('/')

        form_data = {
            field_tc: tc,
            field_okul: okul_no,
            'GUN': gun,
            'AY': ay,
            'YIL': yil,
            'SINAV_ID': sinav_id,
            'Submit': 'Tamam'
        }

        if captcha_code:
            form_data['GUVENLIKKODU'] = captcha_code

        headers = {
            'Referer': sinav_url,
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Content-Type': 'application/x-www-form-urlencoded'
        }

        resp = s.post(post_url, data=form_data, headers=headers, timeout=15, allow_redirects=True)

        if resp.status_code != 200:
            return jsonify({'error': f'MEB yanıt vermedi (HTTP {resp.status_code})'}), 500

        save_cookies_from_session(s)

        soup = BeautifulSoup(resp.text, 'html.parser')

        tables = soup.find_all('table')
        if len(tables) < 3:
            error_text = soup.get_text()
            logger.warning(f"MEB response text (no tables): {error_text[:500]}")

            meb_error = None
            error_div = soup.find('div', {'id': 'hata'})
            if error_div and error_div.get_text(strip=True):
                meb_error = error_div.get_text(strip=True)
            if not meb_error:
                for line in error_text.split('\n'):
                    line = line.strip()
                    if line and ('yanlış' in line.lower() or 'hatalı' in line.lower() or 'bulunamadı' in line.lower()):
                        meb_error = line
                        break

            if meb_error:
                is_captcha_error = 'güvenlik kod' in meb_error.lower() or 'captcha' in meb_error.lower()
                return jsonify({
                    'error': meb_error,
                    'is_captcha_error': is_captcha_error
                }), 400

            if 'bulunamadı' in error_text.lower():
                return jsonify({'error': 'Sonuç bulunamadı. Bilgileri kontrol edin.'}), 404
            return jsonify({'error': 'MEB sonuç sayfası beklenmeyen formatta. Güvenlik kodu yanlış olabilir.', 'is_captcha_error': True}), 400

        result = parse_meb_tables(tables, sonuc_tipi)
        result['tc'] = tc
        result['sinif'] = sinif
        result['exam_year'] = exam_year

        save_lgs_result(result, tc, okul_no, gun, ay, yil, sinif, exam_year)

        return jsonify({
            'success': True,
            'result': result
        })

    except requests.exceptions.Timeout:
        return jsonify({'error': 'MEB sunucusu zaman aşımına uğradı'}), 500
    except Exception as e:
        logger.error(f"Fetch single result error: {e}")
        return jsonify({'error': f'Sonuç çekme hatası: {str(e)}'}), 500


@lgs_results_bp.route('/admin/lgs-results/fetch-batch', methods=['POST'])
@login_required
def fetch_batch_results():
    if current_user.role != 'admin':
        return jsonify({'error': 'Yetkisiz'}), 403

    data = request.get_json()
    students = data.get('students', [])
    exam_year = data.get('exam_year', 2024)

    sinav_url = flask_session.get('meb_sinav_url', 'https://sonuc.meb.gov.tr/')
    sinav_id = flask_session.get('meb_sinav_id', '')
    form_action = flask_session.get('meb_form_action', 'sonuc.php')
    field_tc = flask_session.get('meb_field_tc', 'ADAY_NO')
    field_okul = flask_session.get('meb_field_okul', 'GUVENLIKNUMARASI')

    results = []
    errors = []

    s = requests.Session()
    load_cookies_to_session(s)

    if form_action.startswith('http'):
        post_url = form_action
    else:
        base_url = '/'.join(sinav_url.rstrip('/').split('/')[:-1])
        post_url = base_url + '/' + form_action.lstrip('/')

    headers = {
        'Referer': sinav_url,
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Content-Type': 'application/x-www-form-urlencoded'
    }

    for idx, student in enumerate(students):
        tc = student.get('tc', '')
        okul_no = student.get('okul_no', '')
        dogum = student.get('dogum_tarihi', '')
        sinif = student.get('sinif', '')

        parts = dogum.split('.')
        if len(parts) == 3:
            gun, ay, yil = parts
        else:
            errors.append({'tc': tc, 'error': 'Geçersiz doğum tarihi formatı'})
            continue

        try:
            if idx > 0:
                resp_page = s.get(sinav_url, timeout=10, allow_redirects=True)
                soup_page = BeautifulSoup(resp_page.text, 'html.parser')
                sinav_id_input = soup_page.find('input', {'name': 'SINAV_ID'})
                if sinav_id_input:
                    sinav_id = sinav_id_input['value']

            form_data = {
                field_tc: tc,
                field_okul: okul_no,
                'GUN': gun,
                'AY': ay,
                'YIL': yil,
                'SINAV_ID': sinav_id,
                'Submit': 'Tamam'
            }

            resp = s.post(post_url, data=form_data, headers=headers, timeout=15, allow_redirects=True)

            if resp.status_code != 200:
                errors.append({'tc': tc, 'error': f'HTTP {resp.status_code}'})
                continue

            soup = BeautifulSoup(resp.text, 'html.parser')
            tables = soup.find_all('table')

            if len(tables) < 3:
                page_text = soup.get_text()
                if 'güvenlik' in page_text.lower() or 'captcha' in page_text.lower():
                    return jsonify({
                        'error': 'Güvenlik kodu gerekli. Lütfen güvenlik kodlu moda geçin.',
                        'captcha_required': True,
                        'processed': len(results),
                        'results': results,
                        'errors': errors
                    }), 403

                errors.append({'tc': tc, 'error': 'Sonuç bulunamadı'})
                continue

            result = parse_meb_tables(tables)
            result['tc'] = tc
            result['sinif'] = sinif
            result['exam_year'] = exam_year

            save_lgs_result(result, tc, okul_no, gun, ay, yil, sinif, exam_year)
            results.append(result)

        except requests.exceptions.Timeout:
            errors.append({'tc': tc, 'error': 'Zaman aşımı'})
        except Exception as e:
            errors.append({'tc': tc, 'error': str(e)})

    save_cookies_from_session(s)

    return jsonify({
        'success': True,
        'results': results,
        'errors': errors,
        'total': len(students),
        'success_count': len(results),
        'error_count': len(errors)
    })


def parse_meb_tables(tables, sonuc_tipi='sinav'):
    result = {'sonuc_tipi': sonuc_tipi}
    try:
        logger.info(f"MEB parse ({sonuc_tipi}): {len(tables)} tables found")
        for ti, t in enumerate(tables):
            rows = t.find_all('tr')
            logger.info(f"  Table[{ti}]: {len(rows)} rows")
            for ri, r in enumerate(rows):
                cells = r.find_all(['td', 'th'])
                cell_texts = [c.get_text(strip=True) for c in cells]
                logger.info(f"    Row[{ri}]: {cell_texts}")

        all_labels = {}
        for table in tables:
            rows = table.find_all('tr')
            for row in rows:
                cells = row.find_all(['td', 'th'])
                if len(cells) >= 2:
                    label = cells[0].get_text(strip=True)
                    value = cells[1].get_text(strip=True)
                    all_labels[label.lower()] = value

        for label, value in all_labels.items():
            if 'adı soyadı' in label or ('adı' in label and 'soyadı' in label):
                result['ad_soyad'] = value
            elif label.strip() == 'adı soyadı' or label.strip() == 'ad soyad':
                result['ad_soyad'] = value
            elif 'baba' in label:
                result['baba_adi'] = value
            elif 'doğum' in label:
                pass

        sinav_puan = '0'
        for table in tables:
            text = table.get_text()
            if 'Merkezi Sınav Puanı' in text or 'Sınav Puanı' in text or 'Ağırlıklı Puan' in text:
                rows = table.find_all('tr')
                for row in rows:
                    cells = row.find_all(['td', 'th'])
                    for cell in cells:
                        ct = cell.get_text(strip=True)
                        if ct and ct.replace(',', '').replace('.', '').replace('-', '').isdigit() and len(ct) > 2:
                            sinav_puan = ct
                            break
                    if sinav_puan != '0':
                        break

        if sinav_puan == '0':
            for label, value in all_labels.items():
                if 'puan' in label and value.replace(',', '').replace('.', '').replace('-', '').isdigit():
                    sinav_puan = value
                    break

        puan_parts = sinav_puan.split(',')
        if len(puan_parts) >= 2:
            result['sinav_puan'] = puan_parts[0] + ',' + puan_parts[1][:4]
        else:
            result['sinav_puan'] = sinav_puan

        if sonuc_tipi == 'yerlestirme':
            for label, value in all_labels.items():
                if 'yerleşme türü' in label or 'yerleştirme türü' in label:
                    result['yerlestirme_turu'] = value
                elif 'yerleştirildiğiniz okul' in label or 'okul adı' in label:
                    result['yerlestigi_okul'] = value
                elif 'okul türü' in label:
                    result['okul_turu'] = value
                elif 'bölüm' in label or 'alan' in label:
                    result['bolum_alan'] = value
                elif label.strip() == 'sonuç' or label.strip() == 'sonuc':
                    result['sonuc_text'] = value
                elif 'pansiyon' in label:
                    result['pansiyon'] = value

        elif sonuc_tipi == 'sinav':
            for label, value in all_labels.items():
                if 'genel' in label and 'yüzdelik' in label:
                    result['genel_yuzdelik'] = value
                elif 'il' in label and 'yüzdelik' in label:
                    result['il_yuzdelik'] = value

            dersler_map = {
                'türkçe': 'turkce',
                'matematik': 'matematik',
                'fen': 'fen',
                'ink': 'inkilap',
                'atatürk': 'inkilap',
                'din': 'din',
                'yabancı': 'yabanci_dil',
                'ingilizce': 'yabanci_dil',
            }

            for ders_key in ['turkce', 'matematik', 'fen', 'inkilap', 'din', 'yabanci_dil']:
                result[f'{ders_key}_dogru'] = 0
                result[f'{ders_key}_yanlis'] = 0
                result[f'{ders_key}_bos'] = 0

            for table in tables:
                rows = table.find_all('tr')
                for row in rows:
                    cells = row.find_all(['td', 'th'])
                    if len(cells) >= 4:
                        label = cells[0].get_text(strip=True).lower()
                        for ders_label, ders_key in dersler_map.items():
                            if ders_label in label:
                                result[f'{ders_key}_dogru'] = safe_int(cells[1].get_text(strip=True))
                                result[f'{ders_key}_yanlis'] = safe_int(cells[2].get_text(strip=True))
                                result[f'{ders_key}_bos'] = safe_int(cells[3].get_text(strip=True))
                                break

        if 'ad_soyad' not in result:
            result['ad_soyad'] = ''

    except Exception as e:
        logger.error(f"Parse MEB tables error: {e}")

    return result


def safe_int(val):
    try:
        return int(str(val).strip())
    except:
        return 0


def save_lgs_result(result, tc, okul_no, gun, ay, yil, sinif, exam_year):
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()

        birth_date = f"{yil}-{ay}-{gun}"
        puan_str = result.get('sinav_puan', '0').replace(',', '.')
        try:
            puan_val = float(puan_str)
        except:
            puan_val = 0

        sonuc_tipi = result.get('sonuc_tipi', 'sinav')

        cur.execute("""
            INSERT INTO lgs_results (
                student_name, tc_no, school_no, birth_date, exam_year, sonuc_tipi,
                sinav_puan, genel_yuzdelik, il_yuzdelik,
                turkce_dogru, turkce_yanlis, turkce_bos,
                matematik_dogru, matematik_yanlis, matematik_bos,
                fen_dogru, fen_yanlis, fen_bos,
                inkilap_dogru, inkilap_yanlis, inkilap_bos,
                din_dogru, din_yanlis, din_bos,
                yabanci_dil_dogru, yabanci_dil_yanlis, yabanci_dil_bos,
                yerlestigi_okul, baba_adi, yerlestirme_turu, okul_turu, bolum_alan, sonuc_text, pansiyon,
                sinif, updated_at
            ) VALUES (
                %s, %s, %s, %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s,
                %s, %s, %s, %s, %s, %s, %s,
                %s, CURRENT_TIMESTAMP
            )
            ON CONFLICT (tc_no, exam_year) DO UPDATE SET
                student_name = EXCLUDED.student_name,
                school_no = EXCLUDED.school_no,
                birth_date = EXCLUDED.birth_date,
                sonuc_tipi = EXCLUDED.sonuc_tipi,
                sinav_puan = EXCLUDED.sinav_puan,
                genel_yuzdelik = EXCLUDED.genel_yuzdelik,
                il_yuzdelik = EXCLUDED.il_yuzdelik,
                turkce_dogru = EXCLUDED.turkce_dogru,
                turkce_yanlis = EXCLUDED.turkce_yanlis,
                turkce_bos = EXCLUDED.turkce_bos,
                matematik_dogru = EXCLUDED.matematik_dogru,
                matematik_yanlis = EXCLUDED.matematik_yanlis,
                matematik_bos = EXCLUDED.matematik_bos,
                fen_dogru = EXCLUDED.fen_dogru,
                fen_yanlis = EXCLUDED.fen_yanlis,
                fen_bos = EXCLUDED.fen_bos,
                inkilap_dogru = EXCLUDED.inkilap_dogru,
                inkilap_yanlis = EXCLUDED.inkilap_yanlis,
                inkilap_bos = EXCLUDED.inkilap_bos,
                din_dogru = EXCLUDED.din_dogru,
                din_yanlis = EXCLUDED.din_yanlis,
                din_bos = EXCLUDED.din_bos,
                yabanci_dil_dogru = EXCLUDED.yabanci_dil_dogru,
                yabanci_dil_yanlis = EXCLUDED.yabanci_dil_yanlis,
                yabanci_dil_bos = EXCLUDED.yabanci_dil_bos,
                yerlestigi_okul = EXCLUDED.yerlestigi_okul,
                baba_adi = EXCLUDED.baba_adi,
                yerlestirme_turu = EXCLUDED.yerlestirme_turu,
                okul_turu = EXCLUDED.okul_turu,
                bolum_alan = EXCLUDED.bolum_alan,
                sonuc_text = EXCLUDED.sonuc_text,
                pansiyon = EXCLUDED.pansiyon,
                sinif = EXCLUDED.sinif,
                updated_at = CURRENT_TIMESTAMP
        """, (
            result.get('ad_soyad', ''),
            tc, okul_no, birth_date, exam_year, sonuc_tipi,
            puan_val,
            result.get('genel_yuzdelik', ''),
            result.get('il_yuzdelik', ''),
            result.get('turkce_dogru', 0), result.get('turkce_yanlis', 0), result.get('turkce_bos', 0),
            result.get('matematik_dogru', 0), result.get('matematik_yanlis', 0), result.get('matematik_bos', 0),
            result.get('fen_dogru', 0), result.get('fen_yanlis', 0), result.get('fen_bos', 0),
            result.get('inkilap_dogru', 0), result.get('inkilap_yanlis', 0), result.get('inkilap_bos', 0),
            result.get('din_dogru', 0), result.get('din_yanlis', 0), result.get('din_bos', 0),
            result.get('yabanci_dil_dogru', 0), result.get('yabanci_dil_yanlis', 0), result.get('yabanci_dil_bos', 0),
            result.get('yerlestigi_okul', ''),
            result.get('baba_adi', ''),
            result.get('yerlestirme_turu', ''),
            result.get('okul_turu', ''),
            result.get('bolum_alan', ''),
            result.get('sonuc_text', ''),
            result.get('pansiyon', ''),
            sinif
        ))
        conn.commit()
    except Exception as e:
        logger.error(f"Save LGS result error: {e}")
        if conn:
            conn.rollback()
    finally:
        if conn:
            conn.close()


@lgs_results_bp.route('/admin/lgs-results/list')
@login_required
def list_lgs_results():
    if current_user.role != 'admin':
        return jsonify({'error': 'Yetkisiz'}), 403

    exam_year = request.args.get('year', type=int)
    tipi = request.args.get('tipi', '')

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        conditions = []
        params = []
        if exam_year:
            conditions.append("exam_year = %s")
            params.append(exam_year)
        if tipi:
            conditions.append("COALESCE(sonuc_tipi, 'sinav') = %s")
            params.append(tipi)

        where_clause = " WHERE " + " AND ".join(conditions) if conditions else ""
        cur.execute(f"SELECT * FROM lgs_results{where_clause} ORDER BY exam_year DESC, sinav_puan DESC", params)

        results = cur.fetchall()

        for r in results:
            for key, val in r.items():
                if isinstance(val, datetime):
                    r[key] = val.strftime('%d.%m.%Y %H:%M')
                elif hasattr(val, '__float__'):
                    r[key] = float(val)

        return jsonify({'success': True, 'results': results})

    except Exception as e:
        logger.error(f"List LGS results error: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            conn.close()


@lgs_results_bp.route('/admin/lgs-results/delete/<int:result_id>', methods=['DELETE'])
@login_required
def delete_lgs_result(result_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'Yetkisiz'}), 403

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM lgs_results WHERE id = %s", (result_id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Delete LGS result error: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            conn.close()


@lgs_results_bp.route('/admin/lgs-results/export')
@login_required
def export_lgs_results():
    if current_user.role != 'admin':
        return jsonify({'error': 'Yetkisiz'}), 403

    exam_year = request.args.get('year', type=int)

    conn = None
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)

        if exam_year:
            cur.execute("SELECT * FROM lgs_results WHERE exam_year = %s ORDER BY sinav_puan DESC", (exam_year,))
        else:
            cur.execute("SELECT * FROM lgs_results ORDER BY exam_year DESC, sinav_puan DESC")

        results = cur.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = f"LGS Sonuçları {exam_year or 'Tümü'}"

        headers = [
            'Sıra', 'Ad Soyad', 'TC', 'Sınıf', 'Sınav Puanı',
            'Genel Yüzdelik', 'İl Yüzdelik',
            'Türkçe D', 'Türkçe Y', 'Türkçe B',
            'Mat D', 'Mat Y', 'Mat B',
            'Fen D', 'Fen Y', 'Fen B',
            'İnk D', 'İnk Y', 'İnk B',
            'Din D', 'Din Y', 'Din B',
            'Yab D', 'Yab Y', 'Yab B',
            'Yerleştiği Okul', 'Yıl'
        ]

        header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border

        for idx, r in enumerate(results, 1):
            row_data = [
                idx,
                r.get('student_name', ''),
                r.get('tc_no', ''),
                r.get('sinif', ''),
                float(r['sinav_puan']) if r.get('sinav_puan') else 0,
                r.get('genel_yuzdelik', ''),
                r.get('il_yuzdelik', ''),
                r.get('turkce_dogru', 0), r.get('turkce_yanlis', 0), r.get('turkce_bos', 0),
                r.get('matematik_dogru', 0), r.get('matematik_yanlis', 0), r.get('matematik_bos', 0),
                r.get('fen_dogru', 0), r.get('fen_yanlis', 0), r.get('fen_bos', 0),
                r.get('inkilap_dogru', 0), r.get('inkilap_yanlis', 0), r.get('inkilap_bos', 0),
                r.get('din_dogru', 0), r.get('din_yanlis', 0), r.get('din_bos', 0),
                r.get('yabanci_dil_dogru', 0), r.get('yabanci_dil_yanlis', 0), r.get('yabanci_dil_bos', 0),
                r.get('yerlestigi_okul', ''),
                r.get('exam_year', '')
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=idx + 1, column=col, value=val)
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'LGS_Sonuclari_{exam_year or "Tumu"}.xlsx'
        )

    except Exception as e:
        logger.error(f"Export LGS results error: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            conn.close()


@lgs_results_bp.route('/admin/lgs-results/refresh-captcha', methods=['POST'])
@login_required
def refresh_captcha():
    if current_user.role != 'admin':
        return jsonify({'error': 'Yetkisiz'}), 403

    sinav_url = flask_session.get('meb_sinav_url', 'https://sonuc.meb.gov.tr/')

    try:
        import base64
        s = requests.Session()
        load_cookies_to_session(s)

        from urllib.parse import urlparse
        parsed = urlparse(sinav_url)
        base_url = f"{parsed.scheme}://{parsed.netloc}"
        captcha_url = base_url + '/sinavlar/sonuc/captcha.php'

        captcha_resp = s.get(captcha_url, timeout=10, headers={
            'Referer': sinav_url,
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })

        captcha_base64 = None
        if captcha_resp.status_code == 200 and len(captcha_resp.content) > 100:
            captcha_base64 = base64.b64encode(captcha_resp.content).decode('utf-8')

        save_cookies_from_session(s)

        return jsonify({
            'success': True,
            'captcha_image': captcha_base64
        })

    except Exception as e:
        logger.error(f"Refresh captcha error: {e}")
        return jsonify({'error': str(e)}), 500
