from flask import Flask, request, jsonify, send_from_directory, render_template, session, redirect, url_for, send_file, g, make_response
from flask_cors import CORS
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
import os
import pandas as pd
from datetime import datetime, timedelta, timezone
import uuid
import psycopg2
from psycopg2.extras import RealDictCursor
import json
from itsdangerous import URLSafeTimedSerializer, SignatureExpired, BadSignature
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import io
from io import BytesIO
import logging
from contextlib import contextmanager
import subprocess
import tempfile
import gzip
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak, Flowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

class RotatedParagraph(Flowable):
    """ReportLab'de döndürülmüş paragraf"""
    def __init__(self, text, style, angle=90, width_val=0.4*inch, height_val=2*inch):
        Flowable.__init__(self)
        self.text = text
        self.style = style
        self.angle = angle
        self._width = width_val
        self._height = height_val
    
    def width(self):
        return self._width
    
    def height(self):
        return self._height
    
    def draw(self):
        try:
            para = Paragraph(self.text, self.style)
            self.canv.saveState()
            self.canv.translate(self._width/2, self._height/2)
            self.canv.rotate(self.angle)
            para.drawOn(self.canv, -self._width/4, -self._height/4)
            self.canv.restoreState()
        except Exception as e:
            logger.error(f"RotatedParagraph draw error: {e}")
            # Fallback - normal metin
            Paragraph(self.text, self.style).drawOn(self.canv, 0, 0)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Image as RLImage
import matplotlib

# PDF rapor header'ı oluşturma - logo ve okul adı
def create_pdf_header(styles, logo_width=0.8*inch, logo_height=0.8*inch):
    """PDF raporları için logo ve okul adı içeren header elementi listesi döndürür"""
    header_elements = []
    
    # Logo ekle
    logo_path = "static/images/school_logo.png"
    try:
        import os
        if os.path.exists(logo_path):
            logo = RLImage(logo_path, width=logo_width, height=logo_height)
            header_elements.append(logo)
    except Exception as e:
        logger.warning(f"Logo could not be loaded: {e}")
    
    # Okul adı ve sistem adı
    header_style = ParagraphStyle(
        'SchoolHeader',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.HexColor('#1e3a5f'),
        alignment=TA_CENTER,
        fontName='DejaVuSans',
        leading=18
    )
    subtitle_style = ParagraphStyle(
        'SystemHeader',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#667eea'),
        alignment=TA_CENTER,
        fontName='DejaVuSans'
    )
    
    header_elements.append(Paragraph("Ayşe Melahat Erkin Ortaokulu", header_style))
    header_elements.append(Paragraph("Öğrenci Takip Sistemi", subtitle_style))
    header_elements.append(Spacer(1, 15))
    
    return header_elements

def create_pdf_footer_text():
    """PDF raporları için footer yazısı döndürür"""
    return "Ayşe Melahat Erkin Ortaokulu - Öğrenci Takip Sistemi"
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from google.cloud import storage
import requests
from replit.object_storage import Client
from PIL import Image as PILImage

# Logging yapılandırması
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

# Replit Object Storage yapılandırması
OBJECT_STORAGE_BUCKET_ID = os.environ.get('OBJECT_STORAGE_BUCKET_ID', '')

class ObjectStorageClient:
    """Replit Object Storage wrapper - Resmi SDK kullanır"""
    
    def __init__(self):
        self.bucket_id = OBJECT_STORAGE_BUCKET_ID
        self.enabled = False
        self.client = None
        self._initialize()
    
    def _initialize(self):
        """Object Storage kullanılabilir mi kontrol et"""
        try:
            if not self.bucket_id:
                logger.warning("OBJECT_STORAGE_BUCKET_ID tanımlı değil - Object Storage kullanılamaz")
                return
            
            # Replit Object Storage client oluştur
            try:
                self.client = Client()
                self.enabled = True
                logger.info(f"✅ Object Storage hazır: {self.bucket_id}")
            except Exception as test_error:
                logger.warning(f"⚠️ Object Storage başlatma hatası: {test_error}")
                
        except Exception as e:
            logger.error(f"❌ Object Storage başlatılamadı: {e}")
            self.enabled = False
    
    def upload_from_file(self, source_file, destination_path):
        """Dosyayı Object Storage'a yükle"""
        if not self.enabled or not self.client:
            raise Exception("Object Storage kullanılamıyor")
        
        try:
            # Dosyayı oku
            source_file.seek(0)
            file_data = source_file.read()
            
            # Replit SDK ile yükle
            self.client.upload_from_bytes(destination_path, file_data)
            
            logger.info(f"✅ Dosya Object Storage'a yüklendi: {destination_path}")
            return destination_path
        except Exception as e:
            logger.error(f"❌ Dosya yükleme hatası: {e}")
            raise
    
    def upload_from_bytes(self, data, destination_path, content_type=None):
        """Byte verisini Object Storage'a yükle"""
        if not self.enabled or not self.client:
            raise Exception("Object Storage kullanılamıyor")
        
        try:
            self.client.upload_from_bytes(destination_path, data)
            logger.info(f"✅ Veri Object Storage'a yüklendi: {destination_path}")
            return destination_path
        except Exception as e:
            logger.error(f"❌ Veri yükleme hatası: {e}")
            raise
    
    def download_as_bytes(self, source_path):
        """Object Storage'dan dosya indir"""
        if not self.enabled or not self.client:
            raise Exception("Object Storage kullanılamıyor")
        
        try:
            content = self.client.download_as_bytes(source_path)
            content_type = 'application/octet-stream'
            
            # Dosya uzantısına göre content type belirle
            if source_path.endswith('.pdf'):
                content_type = 'application/pdf'
            elif source_path.endswith(('.jpg', '.jpeg')):
                content_type = 'image/jpeg'
            elif source_path.endswith('.png'):
                content_type = 'image/png'
            
            return content, content_type
        except Exception as e:
            logger.error(f"❌ Dosya indirme hatası: {e}")
            raise FileNotFoundError(f"Dosya bulunamadı: {source_path}")
    
    def delete(self, file_path):
        """Object Storage'dan dosya sil"""
        if not self.enabled or not self.client:
            raise Exception("Object Storage kullanılamıyor")
        
        try:
            self.client.delete(file_path)
            logger.info(f"✅ Dosya silindi: {file_path}")
        except Exception as e:
            logger.error(f"❌ Dosya silme hatası: {e}")
            raise
    
    def list_files(self, prefix=None):
        """Object Storage'daki dosyaları listele"""
        if not self.enabled or not self.client:
            return []
        
        try:
            return list(self.client.list())
        except Exception as e:
            logger.warning(f"Dosya listeleme hatası: {e}")
            return []
    
    def is_available(self):
        """Object Storage kullanılabilir mi?"""
        return self.enabled

# Object Storage client
object_storage = ObjectStorageClient()


app = Flask(__name__)
app.secret_key = os.environ.get('SESSION_SECRET', os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production'))

# Token serializer for PDF downloads (APK session bypass)
pdf_token_serializer = URLSafeTimedSerializer(app.secret_key)

def generate_pdf_token(teacher_id, class_name, exam_number):
    """Öğretmen rapor PDF indirme için güvenli token oluştur"""
    data = {
        'teacher_id': teacher_id,
        'class_name': class_name,
        'exam_number': exam_number
    }
    return pdf_token_serializer.dumps(data, salt='pdf-download')

def verify_pdf_token(token, max_age=300):
    """Token doğrula (5 dakika geçerli)"""
    try:
        data = pdf_token_serializer.loads(token, salt='pdf-download', max_age=max_age)
        return data
    except (SignatureExpired, BadSignature):
        return None

# Session yapılandırması - GÜVENLİK AYARLARI
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # Lax: iframe ve proxy uyumlu
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SECURE'] = False  # Replit proxy için False
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=30)  # 30 dakika sonra oturum sona erer
app.config['SESSION_REFRESH_EACH_REQUEST'] = True  # Her istekte cookie yenilenir
app.config['SESSION_COOKIE_NAME'] = 'ameo_session'  # Özel session cookie adı
app.config['ONESIGNAL_APP_ID'] = os.environ.get('ONESIGNAL_APP_ID', '')
app.config['ONESIGNAL_API_KEY'] = os.environ.get('ONESIGNAL_API_KEY', '')

CORS(app, supports_credentials=True)

NOTIFICATION_ALLOWED_USERS = ['admin', '32260940130']

def can_send_notification(user):
    """Kullanıcının bildirim gönderme yetkisi var mı kontrol et"""
    if not user:
        return False
    if user.role == 'admin':
        return True
    if user.username in NOTIFICATION_ALLOWED_USERS:
        return True
    return False

def send_push_notification(title, message, url="https://ameo-alanya.com", target_classes=None, target_role=None):
    """
    OneSignal üzerinden push bildirim gönder (Yeni API - 2024)
    
    target_classes: Liste olarak sınıflar ["8A", "8B"] veya sınıf seviyesi ["8"] 
    target_role: "student", "teacher" veya None (herkese)
    """
    try:
        onesignal_app_id = app.config.get('ONESIGNAL_APP_ID')
        onesignal_api_key = app.config.get('ONESIGNAL_API_KEY')
        
        logger.info(f"🔔 Bildirim gönderiliyor: {title}")
        logger.info(f"📱 App ID mevcut: {bool(onesignal_app_id)}, API Key mevcut: {bool(onesignal_api_key)}")
        logger.info(f"🎯 Hedef sınıflar: {target_classes}, Hedef rol: {target_role}")
        
        if not onesignal_app_id or not onesignal_api_key:
            logger.warning("❌ OneSignal credentials not configured - App ID veya API Key eksik")
            return False
        
        headers = {
            "Authorization": f"Basic {onesignal_api_key}",
            "Content-Type": "application/json"
        }
        
        payload = {
            "app_id": onesignal_app_id,
            "target_channel": "push",
            "contents": {"tr": message, "en": message},
            "headings": {"tr": title, "en": title},
            "url": url
        }
        
        # Hedef sınıf veya rol varsa filtre kullan
        if target_classes or target_role:
            filters = []
            
            # Sınıf filtresi
            if target_classes:
                class_filters = []
                for i, cls in enumerate(target_classes):
                    if len(cls) == 1:  # Sadece sınıf seviyesi (örn: "8")
                        class_filters.append({"field": "tag", "key": "grade", "relation": "=", "value": cls})
                    else:  # Tam sınıf adı (örn: "8A")
                        class_filters.append({"field": "tag", "key": "class_name", "relation": "=", "value": cls})
                    if i < len(target_classes) - 1:
                        class_filters.append({"operator": "OR"})
                filters.extend(class_filters)
            
            # Rol filtresi
            if target_role:
                if filters:
                    filters.append({"operator": "AND"})
                filters.append({"field": "tag", "key": "role", "relation": "=", "value": target_role})
            
            payload["filters"] = filters
            logger.info(f"🔍 Filtre uygulandı: {filters}")
        else:
            # Hedef yoksa herkese gönder
            payload["included_segments"] = ["Total Subscriptions"]
        
        logger.info(f"📤 OneSignal API'ye istek gönderiliyor...")
        
        response = requests.post(
            "https://api.onesignal.com/notifications",
            headers=headers,
            json=payload
        )
        
        result = response.json()
        logger.info(f"📥 OneSignal yanıtı: Status={response.status_code}, Response={result}")
        
        if response.status_code == 200:
            recipients = result.get('recipients', 0)
            logger.info(f"✅ Push notification sent: {title} - Recipients: {recipients}")
            return True
        else:
            errors = result.get('errors', [])
            if 'All included players are not subscribed' in str(errors):
                logger.warning(f"⚠️ Bildirim gönderildi ama hedef kullanıcı yok: {title}")
                return True
            logger.error(f"❌ Push notification failed: {errors}")
            return False
    except Exception as e:
        logger.error(f"❌ Push notification error: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return False

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'user_login_page'  # Yetkisiz erişimde giriş sayfasına yönlendir
login_manager.login_message = 'Bu sayfaya erişmek için giriş yapmalısınız.'
login_manager.login_message_category = 'warning'
login_manager.session_protection = 'basic'  # Basic koruma - Replit proxy uyumlu

@login_manager.unauthorized_handler
def unauthorized_callback():
    """API çağrıları için JSON, sayfa çağrıları için redirect döndür"""
    if request.path.startswith('/api/') or request.headers.get('Accept', '').find('application/json') >= 0:
        return jsonify({"error": "Oturum süresi doldu. Lütfen tekrar giriş yapın.", "redirect": "/ameo_kullanıcı_giriş"}), 401
    return redirect(url_for('user_login_page'))

UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

# PDF için Türkçe karakter desteği - DejaVu fontlarını register et
try:
    pdfmetrics.registerFont(TTFont('DejaVuSans', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
    pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
    logger.info("✅ PDF Turkish fonts registered successfully")
except Exception as font_error:
    logger.warning(f"⚠️ PDF font registration failed: {font_error}")

ALLOWED_EXTENSIONS = {
    'pdf': ['pdf'],
    'image': ['jpg', 'jpeg', 'png', 'gif'],
    'excel': ['xlsx', 'xls'],
    'document': ['doc', 'docx', 'txt'],
    'all': ['pdf', 'jpg', 'jpeg', 'png', 'gif', 'xlsx', 'xls', 'doc', 'docx', 'txt']
}

def save_uploaded_file(file, category="general"):
    """Dosyayı Object Storage'a (varsa) veya yerel uploads/ dizinine kaydet"""
    if not file or not file.filename:
        raise ValueError("Geçersiz dosya")
    
    filename = secure_filename(file.filename)
    unique_filename = f"{uuid.uuid4().hex}_{filename}"
    
    # Object Storage varsa oraya kaydet
    if object_storage.is_available():
        storage_path = f"{category}/{unique_filename}"
        try:
            object_storage.upload_from_file(file, storage_path)
            logger.info(f"Dosya Object Storage'a kaydedildi: {storage_path}")
            return f"/storage/{storage_path}"  # storage:// prefix ile döndür
        except Exception as e:
            logger.error(f"Object Storage yükleme hatası: {e}")
            # Fallback: yerel kaydet
    
    # Yerel uploads/ dizinine kaydet (fallback veya Object Storage yoksa)
    category_dir = os.path.join(UPLOAD_DIR, category)
    os.makedirs(category_dir, exist_ok=True)
    filepath = os.path.join(category_dir, unique_filename)
    file.seek(0)  # File pointer'ı başa al
    file.save(filepath)
    logger.info(f"Dosya yerel dizine kaydedildi: {filepath}")
    return f"/{filepath}"  # /uploads/category/filename formatında döndür

def get_file_data(file_path):
    """Dosya verisini Object Storage'dan veya yerel dizinden al"""
    # Object Storage path kontrolü (/storage/ ile başlıyor mu?)
    if file_path.startswith("/storage/"):
        if not object_storage.is_available():
            raise FileNotFoundError("Object Storage kullanılamıyor")
        
        storage_path = file_path[len("/storage/"):]  # /storage/ prefix'ini kaldır
        try:
            data, content_type = object_storage.download_as_bytes(storage_path)
            return data, content_type or "application/octet-stream"
        except Exception as e:
            logger.error(f"Object Storage indirme hatası: {e}")
            raise FileNotFoundError(f"Dosya bulunamadı: {file_path}")
    
    # Yerel dosya
    if file_path.startswith("/"):
        file_path = file_path[1:]  # başındaki / işaretini kaldır
    
    full_path = file_path if os.path.isabs(file_path) else os.path.join(os.getcwd(), file_path)
    
    if not os.path.exists(full_path):
        raise FileNotFoundError(f"Dosya bulunamadı: {file_path}")
    
    # Dosya uzantısından content type belirle
    ext = os.path.splitext(file_path)[1].lower()
    content_type_map = {
        '.pdf': 'application/pdf',
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.png': 'image/png',
        '.gif': 'image/gif',
        '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        '.xls': 'application/vnd.ms-excel',
        '.doc': 'application/msword',
        '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        '.txt': 'text/plain'
    }
    content_type = content_type_map.get(ext, 'application/octet-stream')
    
    with open(full_path, 'rb') as f:
        data = f.read()
    
    return data, content_type

def get_db():
    """Veritabanı bağlantısı oluştur"""
    return psycopg2.connect(os.environ.get('DATABASE_URL'))


def allowed_file(filename, file_type='all'):
    if not filename or '.' not in filename:
        return False
    ext = filename.rsplit('.', 1)[1].lower()
    return ext in ALLOWED_EXTENSIONS.get(file_type, ALLOWED_EXTENSIONS['all'])

def remove_overlapping_classes(target_classes):
    """Çakışan sınıf seçimlerini temizle (örn: hem '7A' hem 'tüm_7' varsa, '7A'yı kaldır)"""
    classes_set = set(target_classes)
    
    if 'tüm_okul' in classes_set:
        return ['tüm_okul']
    
    for level in ['5', '6', '7', '8']:
        tum_key = f'tüm_{level}'
        if tum_key in classes_set:
            for branch in ['A', 'B', 'C', 'D', 'E']:
                classes_set.discard(f'{level}{branch}')
    
    return list(classes_set)

class User(UserMixin):
    def __init__(self, id, username, role, full_name, class_name=None):
        self.id = id
        self.username = username
        self.role = role
        self.full_name = full_name
        self.class_name = class_name

@login_manager.user_loader
def load_user(user_id):
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT * FROM users WHERE id = %s", (user_id,))
        user_data = cur.fetchone()
        
        if user_data:
            return User(
                user_data['id'], 
                user_data['username'], 
                user_data['role'], 
                user_data['full_name'],
                user_data.get('class_name')
            )
        return None
    except Exception as e:
        logger.error(f"Error loading user: {e}")
        return None
    finally:
        if conn:
            conn.close()

@app.route("/health")
def health_check():
    return "OK", 200

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/ameo_admin_giris")
def admin_login_page():
    return render_template("admin_login.html")

@app.route("/ameo_kullanıcı_giriş")
def user_login_page():
    return render_template("user_login.html")

@app.route("/veli-portal")
def parent_portal():
    return render_template("parent_portal.html")

@app.route("/dashboard")
@login_required
def dashboard():
    """Genel dashboard - kullanıcıyı rolüne göre yönlendirir"""
    role_routes = {
        'admin': '/admin/dashboard',
        'teacher': '/teacher/dashboard',
        'student': '/student/dashboard'
    }
    return redirect(role_routes.get(current_user.role, '/'))

@app.route("/admin/dashboard")
@login_required
def admin_dashboard():
    if current_user.role != 'admin':
        return redirect('/')
    return render_template("admin_dashboard.html")

@app.route("/admin/question-analysis")
@login_required
def admin_question_analysis():
    if current_user.role != 'admin':
        return redirect('/')
    return render_template("admin_question_analysis.html")

@app.route("/api/last-username")
def get_last_username():
    """Son giriş yapan kullanıcının adını döndür (APK için)"""
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Son giriş yapan kullanıcı
        cur.execute("""
            SELECT username FROM users 
            WHERE last_login_at IS NOT NULL 
            ORDER BY last_login_at DESC 
            LIMIT 1
        """)
        user = cur.fetchone()
        cur.close()
        conn.close()
        
        if user:
            return jsonify({"last_username": user['username']}), 200
        else:
            return jsonify({"last_username": None}), 200
    except Exception as e:
        logger.error(f"Get last username error: {str(e)}")
        return jsonify({"last_username": None}), 200

@app.route("/download/kodular-rehberi")
@login_required
def download_kodular_guide():
    """Kodular Rehberi - PDF/PNG indir (APK uyumlu)"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
        from reportlab.lib.enums import TA_LEFT, TA_CENTER
        from reportlab.lib.units import inch
        
        # Önce PDF oluştur
        pdf_buffer = BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
        story = []
        styles = getSampleStyleSheet()
        
        # Başlık stili
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1f2937'),
            alignment=TA_CENTER,
            fontName='DejaVuSans-Bold',
            spaceAfter=20
        )
        
        # Heading stili
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#1f2937'),
            fontName='DejaVuSans-Bold',
            spaceAfter=10,
            spaceBefore=10
        )
        
        # Normal text stili
        text_style = ParagraphStyle(
            'CustomText',
            parent=styles['Normal'],
            fontSize=10,
            fontName='DejaVuSans',
            spaceAfter=8,
            leading=14
        )
        
        # İçerik
        story.append(Paragraph("KODULAR REHBERİ - ÇOK DETAYLI TARIF", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # ÖN BİLGİ
        story.append(Paragraph("KODULAR ARAYÜZÜ NASIL ÇALIŞIR?", heading_style))
        story.append(Paragraph("Kodular'da 2 sekmesi var:<br/><b>1. Designer Sekmesi:</b> Telefon ekranını tasarladığın yer<br/><b>2. Blocks Sekmesi:</b> Butonların ne yapacağını yazdığın yer", text_style))
        story.append(Spacer(1, 0.15*inch))
        story.append(Paragraph("Blocks sekmesinde 3 alan var:<br/><b>SOL:</b> Tüm öğeler listelenir<br/><b> ORTADA:</b> Kod bloklarını birleştirdiğin boş alan<br/><b>SAĞ:</b> Blokları yapıştırdığında özelliklerini gösteriri", text_style))
        story.append(Spacer(1, 0.25*inch))
        
        # AŞAMA 1
        story.append(Paragraph("AŞAMA 1: TINYDB EKLEME", heading_style))
        story.append(Paragraph("<b>STEP 1:</b> Kodular.io'da Designer sekmesine git", text_style))
        story.append(Paragraph("<b>STEP 2:</b> Sağ tarafta aşağı kaydır → &quot;Storage&quot; bölümünü bul", text_style))
        story.append(Paragraph("<b>STEP 3:</b> &quot;TinyDB&quot;'yi bul (mor/mavi kutucuk) → orta alana sürükle", text_style))
        story.append(Paragraph("<b>STEP 4:</b> Bitti! TinyDB1 adında bir bileşen eklendi", text_style))
        story.append(Spacer(1, 0.2*inch))
        
        # AŞAMA 2 - KOMPLİKE
        story.append(Paragraph("AŞAMA 2: LOGIN BUTONUNA KOD EKLE (EN ÖNEMLİSİ)", heading_style))
        story.append(Paragraph("<b>STEP 1:</b> Kodular.io'da &quot;Blocks&quot; sekmesine tıkla", text_style))
        story.append(Spacer(1, 0.1*inch))
        
        story.append(Paragraph("<b>STEP 2:</b> Sol tarafta hangi öğeleri ara:", text_style))
        story.append(Paragraph("• Aşağı kaydır<br/>• &quot;LoginScreen&quot; veya &quot;MyLoginScreen&quot; isminde bir öğe ara (turuncu olacak)<br/>• Bulduğunda, yanındaki &quot;▶&quot; sembolüne tıkla (AÇILACAK)", text_style))
        story.append(Spacer(1, 0.1*inch))
        
        story.append(Paragraph("<b>STEP 3:</b> LoginScreen açıldı. İçinde ne var bak:", text_style))
        story.append(Paragraph("• &quot;LoginButton&quot; isminde bir öğe ara<br/>• Bulduğunda, yanındaki &quot;▶&quot; tıkla (AÇILACAK)", text_style))
        story.append(Spacer(1, 0.1*inch))
        
        story.append(Paragraph("<b>STEP 4:</b> LoginButton açıldı. Şunu göreceksin:", text_style))
        story.append(Paragraph("• &quot;when LoginButton.Click&quot; - BU ONU!<br/>• Bu satırı SOL TARAFTAN SAĞ TARAFTA BOŞALAN (ORTAYA) SÜRÜKLE", text_style))
        story.append(Spacer(1, 0.15*inch))
        
        story.append(Paragraph("<b>STEP 5:</b> Şimdi sol tarafta &quot;Control&quot; bölümü ara:", text_style))
        story.append(Paragraph("• Aşağı kaydır<br/>• &quot;Control&quot; isminde sarı bir kategori göreceksin<br/>• Yanındaki &quot;▶&quot; tıkla", text_style))
        story.append(Spacer(1, 0.1*inch))
        
        story.append(Paragraph("<b>STEP 6:</b> Control açıldı. İçinde blokları ara:", text_style))
        story.append(Paragraph("• &quot;if&quot; isminde mavi bir blok ara<br/>• Bunu LoginButton.Click'in ALTINA yapıştır", text_style))
        story.append(Spacer(1, 0.15*inch))
        
        story.append(Paragraph("<b>★ BURASI ÖNEMLİ ★ If bloğuna RememberCheckbox koyma:", heading_style))
        story.append(Paragraph("• If bloğu koydun, değil mi?<br/>• Sağ tarafta gördüğün kutulara bak<br/>• &quot;condition&quot; veya &quot;test&quot; isminde bir yer var - orada bir boşluk göreceksin<br/>• O boşluğa tıkla ve sol taraftan &quot;RememberCheckbox&quot;'ı ara<br/>• Bulduğunda, yanındaki &quot;▶&quot; tıkla → &quot;RememberCheckbox.Checked&quot; bloğunu bul<br/>• O bloğu o boşluğa SÜRÜKLE VE BIRAK", text_style))
        story.append(Spacer(1, 0.15*inch))
        
        story.append(Paragraph("<b>STEP 7:</b> if bloğunun altında &quot;then&quot; alanı var:", text_style))
        story.append(Paragraph("• Sol tarafta &quot;TinyDB1&quot; ara<br/>• Yanındaki &quot;▶&quot; tıkla<br/>• &quot;call TinyDB1.StoreValue&quot; bloğunu bul ve then alanına SÜRÜKLE<br/>• Sağ tarafta ayarları doldur: tag=&quot;login_data&quot;, valueKey=&quot;username&quot;, value=UsernameInput.Text", text_style))
        story.append(Spacer(1, 0.15*inch))
        
        story.append(Paragraph("<b>STEP 8:</b> Password için AYNISINI TEKRARLA:", text_style))
        story.append(Paragraph("• TinyDB1.StoreValue bloğunu bir daha then alanına ekle<br/>• Ama bu sefer: valueKey=&quot;password&quot;, value=PasswordInput.Text", text_style))
        story.append(Spacer(1, 0.25*inch))
        
        # AŞAMA 3
        story.append(Paragraph("AŞAMA 3: APP BAŞLANGICINDA OTOMATIK DOLDUR", heading_style))
        story.append(Paragraph("<b>STEP 1:</b> Sol tarafta &quot;Screen1&quot; ara ve yanındaki ▶ tıkla", text_style))
        story.append(Paragraph("<b>STEP 2:</b> &quot;when Screen1.Initialize&quot; bloğunu bul ve ortaya sürükle", text_style))
        story.append(Paragraph("<b>STEP 3:</b> Sol tarafta &quot;UsernameInput&quot; ara, ▶ tıkla, &quot;set UsernameInput.Text&quot; sürükle", text_style))
        story.append(Paragraph("<b>STEP 4:</b> Sağ tarafta &quot;to&quot; kısmına TinyDB1.GetValue bloğunu koy (tag: &quot;login_data&quot;, valueKey: &quot;username&quot;)", text_style))
        story.append(Paragraph("<b>STEP 5:</b> Aynısını PasswordInput için yap (valueKey: &quot;password&quot;)", text_style))
        story.append(Spacer(1, 0.2*inch))
        
        # AŞAMA 4
        story.append(Paragraph("AŞAMA 4: LOGOUT'TA TEMİZLE", heading_style))
        story.append(Paragraph("<b>STEP 1:</b> Sol tarafta &quot;LogoutButton&quot; ara ve ▶ tıkla", text_style))
        story.append(Paragraph("<b>STEP 2:</b> &quot;when LogoutButton.Click&quot; bloğunu ortaya sürükle", text_style))
        story.append(Paragraph("<b>STEP 3:</b> TinyDB1.DeleteValue bloğunu (2 kez) ekle - username ve password için", text_style))
        
        doc.build(story)
        pdf_buffer.seek(0)
        
        # Her zaman PNG'ye dönüştür - APK ve tarayıcı için (PNG herkes açabilir)
        try:
            from pdf2image import convert_from_bytes
            
            # PDF'yi PNG'ye dönüştür (1. sayfa)
            images = convert_from_bytes(pdf_buffer.getvalue(), first_page=1, last_page=1, dpi=100)
            
            if images:
                # Resmi PNG olarak kaydet
                img_buffer = BytesIO()
                images[0].save(img_buffer, format='PNG', optimize=True)
                img_buffer.seek(0)
                
                logger.info("✅ PDF başarıyla PNG'ye dönüştürüldü")
                
                return send_file(
                    img_buffer,
                    mimetype='image/png',
                    as_attachment=True,
                    download_name='Kodular_Rehberi.png'
                )
            else:
                logger.error("PNG dönüşümü başarısız - görüntü oluşturulamadı")
                raise Exception("Görüntü oluşturulamadı")
        
        except Exception as e:
            logger.error(f"PNG conversion failed: {str(e)}, PDF gönderiliyor...")
            # Fallback: PDF'yi direkt gönder
            return send_file(
                BytesIO(pdf_buffer.getvalue()),
                mimetype='application/pdf',
                as_attachment=True,
                download_name='Kodular_Rehberi.pdf'
            )
    
    except Exception as e:
        logger.error(f"Kodular guide error: {str(e)}")
        return jsonify({"error": f"Hata: {str(e)}"}), 500

@app.route("/admin/users")
@login_required
def admin_users_page():
    if current_user.role != 'admin':
        return redirect('/')
    return render_template("admin_users.html")

@app.route("/admin/classes-page")
@login_required
def admin_classes_page():
    if current_user.role != 'admin':
        return redirect('/')
    return render_template("admin_classes.html")

@app.route("/admin/files-page")
@login_required
def admin_files_page():
    if current_user.role != 'admin':
        return redirect('/')
    return render_template("admin_files.html")

@app.route("/admin/reports")
@login_required
def admin_reports_page():
    if current_user.role != 'admin':
        return redirect('/')
    return render_template("admin_reports.html")

@app.route("/teacher/dashboard")
@login_required
def teacher_dashboard():
    if current_user.role not in ['teacher', 'admin']:
        return redirect('/')
    return render_template("teacher_dashboard.html")

@app.route("/teacher/classes", methods=["GET"])
@login_required
def teacher_list_classes():
    """Öğretmenler için sınıf listesi - Sınav oluştururken kullanılır"""
    if current_user.role not in ['teacher', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT id, name, level, branch, type, is_active
            FROM classes
            WHERE is_active = TRUE
            ORDER BY 
                CASE 
                    WHEN type = 'standard' THEN 0
                    ELSE 1
                END,
                level, branch, name
        """)
        classes = cur.fetchall()
        
        cur.close()
        conn.close()
        
        return jsonify({"classes": [{
            "id": c['id'],
            "name": c['name'],
            "level": c['level'],
            "branch": c['branch'],
            "type": c['type'],
            "is_active": c['is_active']
        } for c in classes]})
    
    except Exception as e:
        logger.error(f"Teacher list classes error: {str(e)}")
        return jsonify({"error": f"Sınıflar listelenemedi: {str(e)}"}), 500

@app.route("/student/dashboard")
@login_required
def student_dashboard():
    if current_user.role not in ['student', 'admin']:
        return redirect('/')
    return render_template("student_dashboard.html")

@app.route("/parent/dashboard")
@login_required
def parent_dashboard():
    if current_user.role != 'parent':
        return redirect('/')
    return render_template("parent_dashboard.html")

@app.route("/admin/announcements")
@login_required
def admin_announcements_page():
    if current_user.role != 'admin':
        return redirect('/')
    return render_template("admin_announcements.html")

@app.route("/admin/settings")
@login_required
def admin_settings_page():
    if current_user.role != 'admin':
        return redirect('/')
    return render_template("admin_settings.html")

@app.route("/profile/settings")
@login_required
def profile_settings_page():
    """Kullanıcı profil ayarları sayfası - tüm roller için"""
    return render_template("profile_settings.html")

@app.route("/api/profile/change-password", methods=["POST"])
@login_required
def profile_change_password():
    """Kullanıcı şifresini değiştirir - Admin ve öğretmenler (öğrenciler değiştiremez)"""
    # Admin ve öğretmenler şifre değiştirebilir, öğrenciler değiştiremez
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Şifre değiştirme yetkiniz yok. Lütfen okul yönetimine başvurun."}), 403
    
    try:
        data = request.get_json()
        current_password = data.get('current_password')
        new_password = data.get('new_password')
        
        if not current_password or not new_password:
            return jsonify({"error": "Tüm alanlar gerekli"}), 400
        
        if len(new_password) < 6:
            return jsonify({"error": "Yeni şifre en az 6 karakter olmalıdır"}), 400
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Mevcut şifreyi kontrol et (sütun adı: password)
        cur.execute("SELECT password FROM users WHERE id = %s", (current_user.id,))
        user_data = cur.fetchone()
        
        if not user_data or not check_password_hash(user_data['password'], current_password):
            cur.close()
            conn.close()
            return jsonify({"error": "Mevcut şifre yanlış"}), 401
        
        # Yeni şifreyi hashle ve güncelle
        new_password_hash = generate_password_hash(new_password)
        cur.execute("""
            UPDATE users 
            SET password = %s 
            WHERE id = %s
        """, (new_password_hash, current_user.id))
        
        conn.commit()
        cur.close()
        conn.close()
        
        logger.info(f"✅ Şifre değiştirildi: {current_user.username} (role: {current_user.role})")
        return jsonify({"success": True, "message": "Şifreniz başarıyla güncellendi"})
    
    except Exception as e:
        logger.error(f"Change password error: {str(e)}")
        return jsonify({"error": "Şifre güncellenirken hata oluştu"}), 500

@app.route("/admin/teacher-assignments")
@login_required
def admin_teacher_assignments_page():
    if current_user.role != 'admin':
        return redirect('/')
    return render_template("admin_teacher_assignments.html")

@app.route("/teacher/exams")
@login_required
def teacher_exams_page():
    if current_user.role not in ['teacher', 'admin']:
        return redirect('/')
    return render_template("teacher_exams.html")

@app.route("/teacher/exams/<exam_id>/report")
@login_required
def teacher_exam_report_page(exam_id):
    if current_user.role not in ['teacher', 'admin']:
        return redirect('/')
    return render_template("teacher_exam_report.html", exam_id=exam_id)

@app.route("/teacher/assignments")
@login_required
def teacher_assignments_page():
    if current_user.role not in ['teacher', 'admin']:
        return redirect('/')
    return render_template("teacher_assignments.html")

@app.route("/teacher/announcements")
@login_required
def teacher_announcements_page():
    if current_user.role not in ['teacher', 'admin']:
        return redirect('/')
    return render_template("teacher_announcements.html")

@app.route("/teacher/questions")
@login_required
def teacher_questions_page():
    if current_user.role not in ['teacher', 'admin']:
        return redirect('/')
    return render_template("teacher_questions.html")

@app.route("/teacher/question-asks")
@login_required
def teacher_question_asks_page():
    if current_user.role != 'teacher':
        return redirect('/')
    return render_template("teacher_question_asks.html")

@app.route("/teacher/daily-tracking")
@login_required
def teacher_daily_tracking_page():
    if current_user.role not in ['teacher', 'admin']:
        return redirect('/')
    return render_template("teacher_daily_tracking.html")

@app.route("/teacher/exam-calendar")
@login_required
def teacher_exam_calendar_page():
    if current_user.role not in ['teacher', 'admin']:
        return redirect('/')
    return render_template("teacher_exam_calendar.html")

@app.route("/teacher/study-plan")
@login_required
def teacher_study_plan_page():
    if current_user.role not in ['teacher', 'admin']:
        return redirect('/')
    return render_template("teacher_study_plan.html")

@app.route("/teacher/question-analysis")
@login_required
def teacher_question_analysis_page():
    if current_user.role not in ['teacher', 'admin']:
        return redirect('/')
    return render_template("teacher_question_asks.html")

@app.route("/teacher/lesson-schedule")
@login_required
def teacher_lesson_schedule_page():
    if current_user.role != 'teacher':
        return redirect('/')
    return render_template("teacher_lesson_schedule.html")

@app.route("/teacher/report-cards")
@login_required
def teacher_report_cards_page():
    """Öğretmen karne analizi sayfası - sadece analiz yapabilir"""
    if current_user.role not in ['teacher', 'admin']:
        return redirect('/')
    return redirect('/admin/report-cards/teacher-analysis')

@app.route("/student/exams")
@login_required
def student_exams_page():
    if current_user.role not in ['student', 'admin']:
        return redirect('/')
    return render_template("student_exams.html")

@app.route("/student/exam/<exam_id>/results")
@login_required
def student_exam_result_page(exam_id):
    if current_user.role != 'student':
        return redirect('/')
    return render_template("student_exam_result.html")

@app.route("/student/assignments")
@login_required
def student_assignments_page():
    if current_user.role not in ['student', 'admin']:
        return redirect('/')
    return render_template("student_assignments.html")

@app.route("/student/announcements")
@login_required
def student_announcements_page():
    if current_user.role not in ['student', 'admin']:
        return redirect('/')
    return render_template("student_announcements.html")

@app.route("/student/ask-question-page")
@login_required
def student_ask_question_page():
    if current_user.role not in ['student', 'admin']:
        return redirect('/')
    return render_template("student_ask_question.html")

@app.route("/api/admin/login", methods=["POST"])
def admin_login():
    conn = None
    cur = None
    try:
        logger.info("Admin login request started")
        
        # Try multiple methods to get JSON data
        data = None
        try:
            data = request.get_json(force=True, silent=True)
        except:
            pass
        
        # Fallback: try to parse request.data directly
        if not data:
            try:
                data = json.loads(request.data.decode('utf-8'))
            except:
                pass
        
        logger.info(f"Request data: {data}")
        
        if not data:
            logger.error("Request body is empty")
            return jsonify({"error": "Veri gönderilemedi"}), 400
        
        username = data.get("username", "").strip()
        password = data.get("password", "").strip()
        
        logger.info(f"Attempting login for username: {username}")
        
        if not username or not password:
            return jsonify({"error": "Kullanıcı adı ve şifre gerekli"}), 400
        
        try:
            conn = get_db()
            logger.info("Database connection established")
        except Exception as db_err:
            logger.error(f"Database connection error: {str(db_err)}")
            return jsonify({"error": "Veritabanı bağlantı hatası"}), 500
        
        try:
            cur = conn.cursor(cursor_factory=RealDictCursor)
            cur.execute("SELECT * FROM users WHERE username = %s AND role = 'admin'", (username,))
            user_data = cur.fetchone()
            logger.info(f"User found: {user_data is not None}")
        except Exception as query_err:
            logger.error(f"Database query error: {str(query_err)}")
            cur.close()
            conn.close()
            return jsonify({"error": "Kullanıcı sorgusu hatası"}), 500
        
        if user_data and check_password_hash(user_data['password'], password):
            try:
                # Update last_login_at
                cur.execute("UPDATE users SET last_login_at = CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul' WHERE id = %s", (user_data['id'],))
                logger.info("last_login_at updated")
                
                # Create user_sessions record
                cur.execute("""
                    INSERT INTO user_sessions (user_id, login_at)
                    VALUES (%s, CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul')
                """, (user_data['id'],))
                logger.info("user_session record created")

                conn.commit()
                logger.info("Changes committed to database")
            except Exception as update_err:
                logger.error(f"Update error: {str(update_err)}", exc_info=True)
                conn.rollback()
                cur.close()
                conn.close()
                return jsonify({"error": "Giriş kaydı hatası"}), 500
            
            try:
                user = User(
                    user_data['id'], 
                    user_data['username'], 
                    user_data['role'], 
                    user_data['full_name'],
                    user_data.get('class_name')
                )
                login_user(user, remember=True)  # Oturum 30 dakika süresince devam eder
                session.permanent = True  # 30 dakika oturum süresi (PERMANENT_SESSION_LIFETIME)
                session['login_time'] = datetime.now().isoformat()
                logger.info(f"User {username} logged in successfully")
            except Exception as session_err:
                logger.error(f"Session error: {str(session_err)}")
                cur.close()
                conn.close()
                return jsonify({"error": "Oturum hatası"}), 500
            
            cur.close()
            conn.close()
            return jsonify({
                "success": True,
                "user": {
                    "id": user_data['id'],
                    "username": user_data['username'],
                    "role": user_data['role'],
                    "full_name": user_data['full_name']
                }
            }), 200
        
        cur.close()
        conn.close()
        logger.warning(f"Failed login attempt for username: {username}")
        return jsonify({"error": "Kullanıcı adı veya şifre hatalı"}), 401
    
    except Exception as e:
        logger.error(f"Admin login error: {str(e)}", exc_info=True)
        if cur:
            try:
                cur.close()
            except:
                pass
        if conn:
            try:
                conn.close()
            except:
                pass
        return jsonify({"error": "Giriş yapılırken hata oluştu"}), 500

@app.route("/api/user/login", methods=["POST"])
def user_login():
    conn = None
    cur = None
    try:
        logger.info("User login request started")
        
        # Try multiple methods to get JSON data
        data = None
        try:
            data = request.get_json(force=True, silent=True)
        except:
            pass
        
        # Fallback: try to parse request.data directly
        if not data:
            try:
                data = json.loads(request.data.decode('utf-8'))
            except:
                pass
        
        logger.info(f"Request data: {data}")
        
        if not data:
            logger.error("Request body is empty")
            return jsonify({"error": "Veri gönderilemedi"}), 400
        
        username = data.get("username", "").strip()
        password = data.get("password", "").strip()
        
        logger.info(f"Attempting login for username: {username}")
        
        if not username or not password:
            return jsonify({"error": "Kullanıcı adı ve şifre gerekli"}), 400
        
        try:
            conn = get_db()
            logger.info("Database connection established")
        except Exception as db_err:
            logger.error(f"Database connection error: {str(db_err)}")
            return jsonify({"error": "Veritabanı bağlantı hatası"}), 500
        
        try:
            cur = conn.cursor(cursor_factory=RealDictCursor)
            cur.execute("SELECT * FROM users WHERE username = %s AND role IN ('admin', 'teacher', 'student', 'parent')", (username,))
            user_data = cur.fetchone()
            logger.info(f"User found: {user_data is not None}")
        except Exception as query_err:
            logger.error(f"Database query error: {str(query_err)}")
            cur.close()
            conn.close()
            return jsonify({"error": "Kullanıcı sorgusu hatası"}), 500
        
        if user_data and check_password_hash(user_data['password'], password):
            try:
                # Update last_login_at
                cur.execute("UPDATE users SET last_login_at = CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul' WHERE id = %s", (user_data['id'],))
                logger.info("last_login_at updated")
                
                # Create user_sessions record
                cur.execute("""
                    INSERT INTO user_sessions (user_id, login_at)
                    VALUES (%s, CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul')
                """, (user_data['id'],))
                logger.info("user_session record created")

                conn.commit()
                logger.info("Changes committed to database")
            except Exception as update_err:
                logger.error(f"Update error: {str(update_err)}", exc_info=True)
                conn.rollback()
                cur.close()
                conn.close()
                return jsonify({"error": "Giriş kaydı hatası"}), 500
            
            try:
                user = User(
                    user_data['id'], 
                    user_data['username'], 
                    user_data['role'], 
                    user_data['full_name'],
                    user_data.get('class_name')
                )
                login_user(user, remember=True)  # Oturum 30 dakika süresince devam eder
                session.permanent = True  # 30 dakika oturum süresi (PERMANENT_SESSION_LIFETIME)
                session['login_time'] = datetime.now().isoformat()
                logger.info(f"User {username} logged in successfully")
            except Exception as session_err:
                logger.error(f"Session error: {str(session_err)}")
                cur.close()
                conn.close()
                return jsonify({"error": "Oturum hatası"}), 500
            
            cur.close()
            conn.close()
            return jsonify({
                "success": True,
                "user": {
                    "id": user_data['id'],
                    "username": user_data['username'],
                    "role": user_data['role'],
                    "full_name": user_data['full_name']
                }
            }), 200
        
        cur.close()
        conn.close()
        logger.warning(f"Failed login attempt for username: {username}")
        return jsonify({"error": "Kullanıcı adı veya şifre hatalı"}), 401
    
    except Exception as e:
        logger.error(f"User login error: {str(e)}", exc_info=True)
        if cur:
            try:
                cur.close()
            except:
                pass
        if conn:
            try:
                conn.close()
            except:
                pass
        return jsonify({"error": "Giriş yapılırken hata oluştu"}), 500

@app.route("/api/logout", methods=["POST"])
@login_required
def logout():
    try:
        user_id = current_user.id
        conn = get_db()
        cur = conn.cursor()
        
        # Update last_logout_at
        cur.execute("UPDATE users SET last_logout_at = CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul' WHERE id = %s", (user_id,))
        
        # Record session logout and calculate duration
        if 'login_time' in session:
            login_dt = datetime.fromisoformat(session['login_time'])
            logout_dt = datetime.now()
            duration_minutes = int((logout_dt - login_dt).total_seconds() / 60)
            
            # Get the most recent session ID first
            cur.execute("""
                SELECT id FROM user_sessions 
                WHERE user_id = %s AND logout_at IS NULL
                ORDER BY login_at DESC LIMIT 1
            """, (user_id,))
            session_result = cur.fetchone()
            
            if session_result:
                session_id = session_result[0]
                cur.execute("""
                    UPDATE user_sessions SET logout_at = CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul', duration_minutes = %s
                    WHERE id = %s
                """, (duration_minutes, session_id))
            
            # Update total_session_duration
            cur.execute("""
                UPDATE users SET total_session_duration = (
                    SELECT COALESCE(SUM(duration_minutes), 0) FROM user_sessions 
                    WHERE user_id = %s AND duration_minutes IS NOT NULL
                ) WHERE id = %s
            """, (user_id, user_id))
        
        conn.commit()
        cur.close()
        conn.close()
        
        logout_user()
        session.clear()
        return jsonify({"success": True, "message": "Çıkış yapıldı"}), 200
    except Exception as e:
        logger.error(f"Logout error: {str(e)}", exc_info=True)
        return jsonify({"error": "Çıkış yapılırken hata oluştu"}), 500

@app.route("/api/activity/ping", methods=["POST"])
@login_required
def activity_ping():
    """Kullanıcı aktivitesini güncelle (heartbeat)"""
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("UPDATE users SET last_activity_at = CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul' WHERE id = %s", (current_user.id,))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"success": True}), 200
    except Exception as e:
        logger.error(f"Activity ping error: {str(e)}")
        return jsonify({"error": "Aktivite güncellenemedi"}), 500

@app.route("/api/current-user", methods=["GET"])
def get_current_user():
    if current_user.is_authenticated:
        return jsonify({
            "authenticated": True,
            "user": {
                "id": current_user.id,
                "username": current_user.username,
                "role": current_user.role,
                "full_name": current_user.full_name,
                "class_name": current_user.class_name
            }
        }), 200
    return jsonify({"authenticated": False}), 200

@app.route("/api/classes", methods=["GET"])
@login_required
def get_classes():
    """Tüm sınıfları döndür"""
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("SELECT name FROM classes ORDER BY name")
        classes = cur.fetchall()
        
        cur.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "classes": [{"name": c['name']} for c in classes]
        })
    
    except Exception as e:
        logger.error(f"Get classes error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/change-password", methods=["POST"])
@login_required
def change_password():
    data = request.get_json()
    current_password = data.get("current_password")
    new_password = data.get("new_password")
    
    if not current_password or not new_password:
        return jsonify({"error": "Mevcut şifre ve yeni şifre gereklidir"}), 400
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("SELECT password FROM users WHERE id = %s", (current_user.id,))
    user_data = cur.fetchone()
    
    if not user_data or not check_password_hash(user_data['password'], current_password):
        cur.close()
        conn.close()
        return jsonify({"error": "Mevcut şifre hatalı"}), 401
    
    new_hashed_password = generate_password_hash(new_password)
    cur.execute("UPDATE users SET password = %s WHERE id = %s", (new_hashed_password, current_user.id))
    conn.commit()
    cur.close()
    conn.close()
    
    return jsonify({"success": True, "message": "Şifre başarıyla değiştirildi"}), 200

@app.route("/api/admin/users", methods=["GET", "POST"])
@login_required
def admin_users_api():
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    # GET: Tüm kullanıcıları listele
    if request.method == "GET":
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT id, username, role, full_name, class_name, student_no, created_at 
            FROM users 
            ORDER BY 
                CASE role WHEN 'admin' THEN 1 WHEN 'teacher' THEN 2 WHEN 'student' THEN 3 END,
                class_name ASC NULLS LAST,
                student_no ASC NULLS LAST,
                full_name ASC
        """)
        users = cur.fetchall()
        cur.close()
        conn.close()
        
        return jsonify({"users": [{
            "id": u['id'],
            "username": u['username'],
            "role": u['role'],
            "full_name": u['full_name'],
            "class_name": u['class_name'],
            "student_no": u.get('student_no'),
            "created_at": u['created_at'].isoformat()
        } for u in users]})
    
    # POST: Yeni kullanıcı oluştur
    elif request.method == "POST":
        data = request.get_json()
        username = data.get("username")
        password = data.get("password")
        role = data.get("role")
        full_name = data.get("full_name")
        class_name = data.get("class_name")
        
        # Boş string'leri None'a çevir
        if class_name == "":
            class_name = None
        
        if not all([username, password, role, full_name]):
            return jsonify({"error": "Tüm alanları doldurun"}), 400
        
        if role not in ['admin', 'teacher', 'student']:
            return jsonify({"error": "Geçersiz rol"}), 400
        
        if role == 'student' and not class_name:
            return jsonify({"error": "Öğrenci için sınıf bilgisi zorunludur"}), 400
        
        # Admin ve teacher için class_name None olmalı
        if role in ['admin', 'teacher']:
            class_name = None
        
        conn = get_db()
        cur = conn.cursor()
        
        try:
            hashed_password = generate_password_hash(password)
            cur.execute(
                "INSERT INTO users (username, password, role, full_name, class_name) VALUES (%s, %s, %s, %s, %s) RETURNING id",
                (username, hashed_password, role, full_name, class_name)
            )
            user_id = cur.fetchone()[0]
            conn.commit()
            
            return jsonify({
                "success": True, 
                "message": "Kullanıcı oluşturuldu",
                "user_id": user_id
            }), 201
        except psycopg2.IntegrityError:
            conn.rollback()
            return jsonify({"error": "Bu kullanıcı adı zaten kullanılıyor"}), 400
        finally:
            cur.close()
            conn.close()

@app.route("/api/admin/users/<int:user_id>/class", methods=["PUT"])
@login_required
def admin_update_user_class(user_id):
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    data = request.get_json()
    class_name = data.get("class_name")
    
    conn = get_db()
    cur = conn.cursor()
    
    try:
        cur.execute("UPDATE users SET class_name = %s WHERE id = %s", (class_name, user_id))
        conn.commit()
        return jsonify({"success": True, "message": "Sınıf güncellendi"}), 200
    except Exception as e:
        conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()

@app.route("/api/admin/users/<int:user_id>", methods=["PUT"])
@login_required
def admin_update_user(user_id):
    """Admin kullanıcı bilgilerini günceller"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        data = request.get_json()
        username = data.get('username', '').strip()
        full_name = data.get('full_name', '').strip()
        role = data.get('role', '').strip()
        class_name = data.get('class_name', '').strip() if data.get('class_name') else None
        student_no = data.get('student_no', '').strip() if data.get('student_no') else None
        password = data.get('password', '').strip()
        
        if not username or not full_name or not role:
            return jsonify({"error": "Kullanıcı adı, ad soyad ve rol gerekli"}), 400
        
        if role not in ['teacher', 'student', 'admin']:
            return jsonify({"error": "Geçersiz rol"}), 400
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Kullanıcı adı kontrolü (başka kullanıcıda kullanılıyor mu?)
        cur.execute("SELECT id FROM users WHERE username = %s AND id != %s", (username, user_id))
        if cur.fetchone():
            cur.close()
            conn.close()
            return jsonify({"error": "Bu kullanıcı adı zaten kullanılıyor"}), 400
        
        # Şifre güncellemesi varsa hash'le
        if password:
            from werkzeug.security import generate_password_hash
            hashed_password = generate_password_hash(password)
            cur.execute("""
                UPDATE users 
                SET username = %s, full_name = %s, role = %s, class_name = %s, student_no = %s, password = %s
                WHERE id = %s
            """, (username, full_name, role, class_name, student_no, hashed_password, user_id))
        else:
            # Şifre güncellemesi yoksa eski şifreyi koru
            cur.execute("""
                UPDATE users 
                SET username = %s, full_name = %s, role = %s, class_name = %s, student_no = %s
                WHERE id = %s
            """, (username, full_name, role, class_name, student_no, user_id))
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Kullanıcı başarıyla güncellendi"}), 200
        
    except Exception as e:
        logger.error(f"Update user error: {str(e)}")
        return jsonify({"error": "Kullanıcı güncellenirken hata oluştu"}), 500

@app.route("/api/admin/users/<int:user_id>", methods=["DELETE"])
@login_required
def admin_delete_user(user_id):
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    if user_id == current_user.id:
        return jsonify({"error": "Kendi hesabınızı silemezsiniz"}), 400
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Kullanıcı bilgisini al
        cur.execute("SELECT role, username, full_name FROM users WHERE id = %s", (user_id,))
        user = cur.fetchone()
        
        if not user:
            cur.close()
            conn.close()
            return jsonify({"error": "Kullanıcı bulunamadı"}), 404
        
        # Öğretmen ise, bağlı kayıtları sil
        if user['role'] == 'teacher':
            # Öğretmene ait sınavları sil
            cur.execute("DELETE FROM exam_results WHERE exam_id IN (SELECT id FROM exams WHERE teacher_id = %s)", (user_id,))
            cur.execute("DELETE FROM exam_submissions WHERE exam_id IN (SELECT id FROM exams WHERE teacher_id = %s)", (user_id,))
            cur.execute("DELETE FROM exams WHERE teacher_id = %s", (user_id,))
            
            # Öğretmene ait ödevleri sil
            cur.execute("DELETE FROM assignment_submissions WHERE assignment_id IN (SELECT id FROM assignments WHERE teacher_id = %s)", (user_id,))
            cur.execute("DELETE FROM assignments WHERE teacher_id = %s", (user_id,))
            
            # Öğretmene ait duyuruları sil
            cur.execute("DELETE FROM announcements WHERE teacher_id = %s", (user_id,))
            
            # Öğretmene gelen soruları sil
            cur.execute("DELETE FROM student_questions WHERE teacher_id = %s", (user_id,))
        
        # Öğrenci ise, bağlı kayıtları sil
        elif user['role'] == 'student':
            # Öğrencinin sınav sonuçlarını sil
            cur.execute("DELETE FROM exam_results WHERE student_id = %s", (user_id,))
            cur.execute("DELETE FROM exam_submissions WHERE student_id = %s", (user_id,))
            
            # Öğrencinin ödev teslimlerini sil
            cur.execute("DELETE FROM assignment_submissions WHERE student_id = %s", (user_id,))
            
            # Öğrencinin sorularını sil
            cur.execute("DELETE FROM student_questions WHERE student_id = %s", (user_id,))
        
        # Admin ise, admin duyurularını sil
        elif user['role'] == 'admin':
            cur.execute("DELETE FROM teacher_announcements WHERE admin_id = %s", (user_id,))
        
        # Kullanıcıyı sil
        cur.execute("DELETE FROM users WHERE id = %s", (user_id,))
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": f"{user['full_name']} ({user['username']}) ve ilgili tüm kayıtlar silindi"}), 200
    
    except Exception as e:
        conn.rollback()
        cur.close()
        conn.close()
        logger.error(f" Delete user error: {str(e)}")
        return jsonify({"error": f"Kullanıcı silinemedi: {str(e)}"}), 500

@app.route("/api/admin/user-activity", methods=["GET"])
@login_required
def admin_user_activity_api():
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        # Tüm kullanıcıları aktivite bilgileriyle getir
        cur.execute("""
            SELECT 
                u.id,
                u.username,
                u.full_name,
                u.role,
                u.class_name,
                u.last_login_at,
                u.last_logout_at,
                u.last_activity_at,
                u.total_session_duration,
                COUNT(us.id) as total_sessions,
                COALESCE(SUM(us.duration_minutes), 0) as total_duration_minutes
            FROM users u
            LEFT JOIN user_sessions us ON u.id = us.user_id
            WHERE u.role IN ('teacher', 'student')
            GROUP BY u.id, u.username, u.full_name, u.role, u.class_name, u.last_login_at, u.last_logout_at, u.last_activity_at, u.total_session_duration
            ORDER BY u.last_activity_at DESC NULLS LAST
        """)
        
        users = cur.fetchall()
        result = []
        
        for user in users:
            # Aktif durumu hesapla: last_activity_at son 30 dakika içinde güncellediyse aktif
            is_active = False
            if user.get('last_activity_at'):
                # Son 30 dakika içinde aktivite varsa aktif
                time_diff = datetime.now(timezone.utc) - user['last_activity_at'].replace(tzinfo=timezone.utc)
                is_active = time_diff.total_seconds() < 1800  # 30 dakika = 1800 saniye
            # Heartbeat yeni eklendi, son 30 dakika aktivite yoksa otomatik pasif
            
            result.append({
                "id": user['id'],
                "username": user['username'],
                "full_name": user['full_name'],
                "role": user['role'],
                "class_name": user['class_name'],
                "last_login": user['last_login_at'].isoformat() if user['last_login_at'] else None,
                "last_logout": user['last_logout_at'].isoformat() if user['last_logout_at'] else None,
                "last_activity": user['last_activity_at'].isoformat() if user.get('last_activity_at') else None,
                "total_sessions": int(user['total_sessions']) if user['total_sessions'] else 0,
                "total_duration_minutes": int(user['total_duration_minutes']) if user['total_duration_minutes'] else 0,
                "total_duration_hours": round(user['total_duration_minutes'] / 60, 2) if user['total_duration_minutes'] else 0,
                "status": "🟢 Aktif" if is_active else "⚪ Pasif"
            })
        
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "data": result}), 200
    
    except Exception as e:
        logger.error(f"Activity API error: {str(e)}")
        cur.close()
        conn.close()
        return jsonify({"error": f"Aktivite verisi alınamadı: {str(e)}"}), 500

@app.route("/admin/user-activity", methods=["GET"])
@login_required
def admin_user_activity_page():
    if current_user.role != 'admin':
        return redirect('/')
    return render_template("admin_user_activity.html")

@app.route("/api/admin/users/delete-bulk", methods=["POST"])
@login_required
def admin_delete_users_bulk():
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    data = request.get_json()
    user_ids = data.get("user_ids", [])
    
    if not user_ids or not isinstance(user_ids, list):
        return jsonify({"error": "Kullanıcı ID listesi gerekli"}), 400
    
    # Current user'ı listeden çıkar (kendi hesabını silemesin)
    user_ids = [uid for uid in user_ids if uid != current_user.id]
    
    if not user_ids:
        return jsonify({"error": "Silinecek kullanıcı bulunamadı"}), 400
    
    deleted_count = 0
    errors = []
    
    # Her kullanıcıyı kendi transaction'ında sil (partial success için)
    for user_id in user_ids:
        conn = None
        cur = None
        try:
            conn = get_db()
            cur = conn.cursor(cursor_factory=RealDictCursor)
            
            # Kullanıcı bilgisini al
            cur.execute("SELECT role, username, full_name FROM users WHERE id = %s", (user_id,))
            user = cur.fetchone()
            
            if not user:
                errors.append(f"ID {user_id}: Kullanıcı bulunamadı")
                continue
            
            # Admin kullanıcılarını silmeyi engelle
            if user['role'] == 'admin':
                errors.append(f"{user['full_name']}: Admin kullanıcılar silinemez")
                continue
            
            # Öğretmen ise, bağlı kayıtları sil
            if user['role'] == 'teacher':
                cur.execute("DELETE FROM exam_results WHERE exam_id IN (SELECT id FROM exams WHERE teacher_id = %s)", (user_id,))
                cur.execute("DELETE FROM exam_submissions WHERE exam_id IN (SELECT id FROM exams WHERE teacher_id = %s)", (user_id,))
                cur.execute("DELETE FROM exams WHERE teacher_id = %s", (user_id,))
                cur.execute("DELETE FROM assignment_submissions WHERE assignment_id IN (SELECT id FROM assignments WHERE teacher_id = %s)", (user_id,))
                cur.execute("DELETE FROM assignments WHERE teacher_id = %s", (user_id,))
                cur.execute("DELETE FROM announcement_reads WHERE announcement_id IN (SELECT id FROM announcements WHERE teacher_id = %s)", (user_id,))
                cur.execute("DELETE FROM announcements WHERE teacher_id = %s", (user_id,))
                cur.execute("DELETE FROM student_questions WHERE teacher_id = %s", (user_id,))
            
            # Öğrenci ise, bağlı kayıtları sil
            elif user['role'] == 'student':
                cur.execute("DELETE FROM exam_results WHERE student_id = %s", (user_id,))
                cur.execute("DELETE FROM exam_submissions WHERE student_id = %s", (user_id,))
                cur.execute("DELETE FROM assignment_submissions WHERE student_id = %s", (user_id,))
                cur.execute("DELETE FROM announcement_reads WHERE student_id = %s", (user_id,))
                cur.execute("DELETE FROM student_questions WHERE student_id = %s", (user_id,))
                cur.execute("DELETE FROM practice_exams WHERE student_id = %s", (user_id,))
            
            # Kullanıcıyı sil
            cur.execute("DELETE FROM users WHERE id = %s", (user_id,))
            conn.commit()
            deleted_count += 1
            
        except Exception as user_error:
            if conn:
                conn.rollback()
            errors.append(f"ID {user_id}: {str(user_error)}")
            logger.error(f"Error deleting user {user_id}: {str(user_error)}")
        
        finally:
            if cur:
                cur.close()
            if conn:
                conn.close()
    
    # Sonuç döndür
    message = f"✅ {deleted_count} kullanıcı başarıyla silindi"
    if errors:
        message += f"\n⚠️ {len(errors)} hata:\n" + "\n".join(errors[:5])
    
    return jsonify({
        "success": True,
        "message": message,
        "deleted": deleted_count,
        "errors": len(errors)
    }), 200

@app.route("/api/admin/impersonate/<int:user_id>", methods=["POST"])
@login_required
def admin_impersonate_user(user_id):
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    try:
        cur.execute("SELECT id, username, role, full_name, class_name FROM users WHERE id = %s", (user_id,))
        target_user = cur.fetchone()
        
        if not target_user:
            return jsonify({"error": "Kullanıcı bulunamadı"}), 404
        
        user_obj = User(
            target_user['id'],
            target_user['username'],
            target_user['role'],
            target_user['full_name'],
            target_user.get('class_name')
        )
        
        logout_user()
        login_user(user_obj, remember=True)  # Oturum devam etsin
        session.permanent = True  # 30 dakika oturum süresi
        session['login_time'] = datetime.now().isoformat()
        session.modified = True
        
        redirect_url = '/'
        if target_user['role'] == 'teacher':
            redirect_url = '/teacher/dashboard'
        elif target_user['role'] == 'student':
            redirect_url = '/student/dashboard'
        elif target_user['role'] == 'admin':
            redirect_url = '/admin/dashboard'
        
        logger.info(f"✅ Admin {current_user.username} tarafından {target_user['username']} kullanıcısına geçiş yapıldı")
        
        return jsonify({
            "success": True,
            "message": f"{target_user['full_name']} kullanıcısına bağlanıldı",
            "redirect": redirect_url
        }), 200
        
    except Exception as e:
        logger.error(f"Impersonate error: {str(e)}")
        return jsonify({"error": f"Bağlanılamadı: {str(e)}"}), 500
    finally:
        cur.close()
        conn.close()

@app.route("/api/admin/sample-excel", methods=["GET"])
@login_required
def admin_sample_excel():
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    sample_data = {
        'Kullanıcı Adı': ['ahmet.yilmaz', 'ayse.kaya', 'mehmet.demir', 'fatma.ozturk', 'ali.celik'],
        'Şifre': ['ahmet123', 'ayse123', 'mehmet123', 'fatma123', 'ali123'],
        'Rol': ['teacher', 'student', 'student', 'teacher', 'student'],
        'Ad Soyad': ['Ahmet Yılmaz', 'Ayşe Kaya', 'Mehmet Demir', 'Fatma Öztürk', 'Ali Çelik'],
        'Sınıf': ['', '5A', '6C', '', '7E']
    }
    
    df = pd.DataFrame(sample_data)
    
    excel_path = os.path.join(UPLOAD_DIR, 'ornek_kullanicilar.xlsx')
    df.to_excel(excel_path, index=False, engine='openpyxl')
    
    return send_from_directory(UPLOAD_DIR, 'ornek_kullanicilar.xlsx', as_attachment=False)

@app.route("/api/admin/users/sample-excel", methods=["GET"])
@login_required
def admin_users_sample_excel():
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    sample_data = {
        'Kullanıcı Adı': ['ahmet.yilmaz', 'ayse.kaya', 'mehmet.demir', 'fatma.ozturk', 'ali.celik'],
        'Şifre': ['ahmet123', 'ayse123', 'mehmet123', 'fatma123', 'ali123'],
        'Rol': ['teacher', 'student', 'student', 'teacher', 'student'],
        'Ad Soyad': ['Ahmet Yılmaz', 'Ayşe Kaya', 'Mehmet Demir', 'Fatma Öztürk', 'Ali Çelik'],
        'Sınıf': ['', '5A', '6C', '', '7E']
    }
    
    df = pd.DataFrame(sample_data)
    
    excel_path = os.path.join(UPLOAD_DIR, 'ornek_kullanicilar.xlsx')
    df.to_excel(excel_path, index=False, engine='openpyxl')
    
    return send_from_directory(UPLOAD_DIR, 'ornek_kullanicilar.xlsx', as_attachment=False)

@app.route("/api/admin/users/export", methods=["GET"])
@login_required
def admin_export_users():
    """Tüm kullanıcıları Excel'e indir - Şifre = kullanıcı adının ilk 6 hanesi"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Tüm kullanıcıları al (okul numarası dahil)
        cur.execute("""
            SELECT username, full_name, role, class_name, student_no 
            FROM users 
            ORDER BY role, class_name, full_name
        """)
        users = cur.fetchall()
        cur.close()
        conn.close()
        
        # Şifre = kullanıcı adının ilk 6 hanesi
        data = {
            'Kullanıcı Adı': [],
            'Şifre': [],
            'Ad Soyad': [],
            'Sınıf': [],
            'Okul Numarası': [],
            'Rol': []
        }
        
        for user in users:
            username = user['username'] or ''
            data['Kullanıcı Adı'].append(username)
            data['Şifre'].append(username[:6] if len(username) >= 6 else username)
            data['Ad Soyad'].append(user['full_name'])
            data['Sınıf'].append(user['class_name'] or '')
            data['Okul Numarası'].append(user['student_no'] or '')
            data['Rol'].append(user['role'])
        
        df = pd.DataFrame(data)
        
        # Excel dosyasını belleğe yaz
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Kullanıcılar')
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=False,
            download_name=f'kullanicilar_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    
    except Exception as e:
        logger.error(f"Export users error: {str(e)}")
        return jsonify({"error": f"Export hatası: {str(e)}"}), 500

@app.route("/api/admin/users/update-bulk", methods=["POST"])
@login_required
def admin_update_users_bulk():
    """Excel'den kullanıcıları toplu güncelle - sadece eksik/değişen verileri günceller"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        if 'file' not in request.files:
            return jsonify({"error": "Dosya seçilmedi"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "Dosya seçilmedi"}), 400
        
        if not file.filename.endswith('.xlsx'):
            return jsonify({"error": "Sadece .xlsx dosyası yüklenebilir"}), 400
        
        df = pd.read_excel(file, sheet_name='Kullanıcılar')
        
        # Sütun isimlerini normalize et - farklı varyasyonları ve "Unnamed" sütunlarını düzelt
        column_mapping = {}
        unnamed_columns = []
        for col in df.columns:
            col_lower = str(col).lower().strip()
            # Önce unnamed kontrolü
            if col_lower.startswith('unnamed'):
                unnamed_columns.append(col)
            elif 'kullanıcı' in col_lower or 'username' in col_lower:
                column_mapping[col] = 'Kullanıcı Adı'
            elif 'şifre' in col_lower or 'sifre' in col_lower or 'password' in col_lower:
                column_mapping[col] = 'Şifre'
            elif 'sınıf' in col_lower or 'sinif' in col_lower or 'class' in col_lower:
                column_mapping[col] = 'Sınıf'
            elif 'okul' in col_lower or 'numara' in col_lower or 'student' in col_lower:
                column_mapping[col] = 'Okul Numarası'
            elif 'rol' in col_lower or 'role' in col_lower:
                column_mapping[col] = 'Rol'
            elif 'soyad' in col_lower or ('ad' in col_lower and 'ad soyad' in col_lower):
                column_mapping[col] = 'Ad Soyad'
        
        # Unnamed sütunlarını sıraya göre eşleştir (5. sütun = Okul Numarası)
        if 'Okul Numarası' not in column_mapping.values() and unnamed_columns:
            cols = list(df.columns)
            for unnamed_col in unnamed_columns:
                idx = cols.index(unnamed_col)
                # 5. sütun (index 4) genellikle Okul Numarası
                if idx == 4:
                    column_mapping[unnamed_col] = 'Okul Numarası'
                    break
        
        df = df.rename(columns=column_mapping)
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        updated_password = 0
        updated_student_no = 0
        updated_class = 0
        skipped = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                # Kullanıcı adı - sayısal değerlerde .0 sonekini kaldır
                raw_username = row.get('Kullanıcı Adı', '')
                if pd.notna(raw_username):
                    if isinstance(raw_username, float) and raw_username == int(raw_username):
                        username = str(int(raw_username)).strip()
                    else:
                        username = str(raw_username).strip()
                    if username.endswith('.0'):
                        username = username[:-2]
                else:
                    username = ''
                
                if not username:
                    continue
                
                # Kullanıcıyı bul
                cur.execute("SELECT id, password, student_no, class_name FROM users WHERE username = %s", (username,))
                user = cur.fetchone()
                
                if not user:
                    errors.append(f"Satır {index+2}: '{username}' bulunamadı")
                    continue
                
                user_id = user['id']
                updates = []
                params = []
                
                # Şifre güncelleme (boş veya NaN değilse)
                raw_password = row.get('Şifre')
                if pd.notna(raw_password):
                    # Sayısal değerler için .0 sonekini kaldır (Excel float olarak okuyabilir)
                    if isinstance(raw_password, float) and raw_password == int(raw_password):
                        new_password = str(int(raw_password)).strip()
                    else:
                        new_password = str(raw_password).strip()
                    # .0 ile biten sayıları düzelt
                    if new_password.endswith('.0'):
                        new_password = new_password[:-2]
                    if new_password and new_password.lower() != 'nan' and len(new_password) >= 3:
                        new_hash = generate_password_hash(new_password)
                        updates.append("password = %s")
                        params.append(new_hash)
                        updated_password += 1
                
                # Okul numarası güncelleme (boş veya NaN değilse ve mevcut değerden farklıysa)
                raw_student_no = row.get('Okul Numarası')
                current_student_no = user['student_no'] or ''
                if pd.notna(raw_student_no):
                    # Sayısal değerler için .0 sonekini kaldır
                    if isinstance(raw_student_no, float) and raw_student_no == int(raw_student_no):
                        new_student_no = str(int(raw_student_no)).strip()
                    else:
                        new_student_no = str(raw_student_no).strip()
                    if new_student_no.endswith('.0'):
                        new_student_no = new_student_no[:-2]
                    if new_student_no and new_student_no.lower() != 'nan' and new_student_no != current_student_no:
                        updates.append("student_no = %s")
                        params.append(new_student_no)
                        updated_student_no += 1
                
                # Sınıf güncelleme (boş veya NaN değilse ve mevcut değerden farklıysa)
                raw_class = row.get('Sınıf')
                current_class = user['class_name'] or ''
                if pd.notna(raw_class):
                    new_class = str(raw_class).strip()
                    if new_class and new_class.lower() != 'nan' and new_class != current_class:
                        updates.append("class_name = %s")
                        params.append(new_class)
                        updated_class += 1
                
                # Güncelleme yap
                if updates:
                    params.append(user_id)
                    cur.execute(f"UPDATE users SET {', '.join(updates)} WHERE id = %s", params)
                    conn.commit()
                else:
                    skipped += 1
                    
            except Exception as row_error:
                conn.rollback()
                errors.append(f"Satır {index+2}: {str(row_error)}")
                continue
        
        cur.close()
        conn.close()
        
        message = f"✅ Güncelleme tamamlandı!\n"
        message += f"🔐 Şifre güncellenen: {updated_password}\n"
        message += f"🔢 Okul numarası güncellenen: {updated_student_no}\n"
        message += f"📚 Sınıf güncellenen: {updated_class}\n"
        message += f"⏭️ Değişiklik olmayan: {skipped}"
        
        if errors:
            message += f"\n\n⚠️ Hatalar ({len(errors)}):\n"
            for err in errors[:5]:
                message += f"• {err}\n"
            if len(errors) > 5:
                message += f"... ve {len(errors) - 5} hata daha"
        
        return jsonify({
            "success": True,
            "message": message,
            "updated_password": updated_password,
            "updated_student_no": updated_student_no,
            "updated_class": updated_class,
            "skipped": skipped,
            "errors": len(errors)
        })
        
    except Exception as e:
        logger.error(f"Bulk update users error: {str(e)}")
        return jsonify({"error": f"Güncelleme hatası: {str(e)}"}), 500

@app.route("/api/admin/users/reset-all-passwords", methods=["POST"])
@login_required
def admin_reset_all_passwords():
    """Tüm kullanıcıların şifrelerini sıfırla (kullanıcı adının ilk 6 hanesi)"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("SELECT id, username FROM users WHERE role != 'admin'")
        users = cur.fetchall()
        
        reset_count = 0
        for user in users:
            try:
                username = user['username'] or ''
                new_password = username[:6] if len(username) >= 6 else username
                if len(new_password) >= 1:
                    new_hash = generate_password_hash(new_password)
                    cur.execute("UPDATE users SET password = %s WHERE id = %s", (new_hash, user['id']))
                    reset_count += 1
            except Exception as e:
                logger.error(f"Reset password for user {user['id']} error: {str(e)}")
                continue
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "reset_count": reset_count,
            "message": f"{reset_count} kullanıcının şifresi sıfırlandı"
        })
        
    except Exception as e:
        logger.error(f"Reset all passwords error: {str(e)}")
        return jsonify({"error": f"Şifre sıfırlama hatası: {str(e)}"}), 500

@app.route("/api/admin/users/delete-all", methods=["DELETE"])
@login_required
def admin_delete_all_users():
    """Tüm kullanıcıları sil (Admin hariç)"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Admin olmayan kullanıcıları al (silmeden önce ilişkili verileri silelim)
        cur.execute("SELECT id, role FROM users WHERE role != 'admin'")
        users_to_delete = cur.fetchall()
        
        deleted_count = 0
        for user in users_to_delete:
            user_id = user['id']
            role = user['role']
            
            try:
                # Öğretmen ise
                if role == 'teacher':
                    cur.execute("DELETE FROM exam_results WHERE exam_id IN (SELECT id FROM exams WHERE teacher_id = %s)", (user_id,))
                    cur.execute("DELETE FROM exam_submissions WHERE exam_id IN (SELECT id FROM exams WHERE teacher_id = %s)", (user_id,))
                    cur.execute("DELETE FROM exams WHERE teacher_id = %s", (user_id,))
                    cur.execute("DELETE FROM assignment_submissions WHERE assignment_id IN (SELECT id FROM assignments WHERE teacher_id = %s)", (user_id,))
                    cur.execute("DELETE FROM assignments WHERE teacher_id = %s", (user_id,))
                    cur.execute("DELETE FROM announcements WHERE teacher_id = %s", (user_id,))
                    cur.execute("DELETE FROM student_questions WHERE teacher_id = %s", (user_id,))
                
                # Öğrenci ise
                elif role == 'student':
                    cur.execute("DELETE FROM exam_results WHERE student_id = %s", (user_id,))
                    cur.execute("DELETE FROM exam_submissions WHERE student_id = %s", (user_id,))
                    cur.execute("DELETE FROM assignment_submissions WHERE student_id = %s", (user_id,))
                    cur.execute("DELETE FROM student_questions WHERE student_id = %s", (user_id,))
                    cur.execute("DELETE FROM practice_exams WHERE student_id = %s", (user_id,))
                
                # Kullanıcıyı sil
                cur.execute("DELETE FROM users WHERE id = %s", (user_id,))
                conn.commit()
                deleted_count += 1
                
            except Exception as e:
                conn.rollback()
                logger.error(f"Delete user {user_id} error: {str(e)}")
                continue
        
        cur.close()
        conn.close()
        
        logger.info(f"✅ {deleted_count} kullanıcı toplu olarak silindi")
        return jsonify({
            "success": True,
            "message": f"{deleted_count} kullanıcı ve ilgili tüm kayıtlar silindi",
            "deleted_count": deleted_count
        }), 200
    
    except Exception as e:
        logger.error(f"Delete all users error: {str(e)}")
        return jsonify({"error": f"Toplu silme hatası: {str(e)}"}), 500

@app.route("/api/admin/users/upload-excel", methods=["POST"])
@login_required
def admin_upload_excel():
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    if "file" not in request.files:
        return jsonify({"error": "Dosya bulunamadı"}), 400
    
    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "Dosya seçilmedi"}), 400
    
    try:
        df = pd.read_excel(file)
        
        required_columns = ['Kullanıcı Adı', 'Şifre', 'Rol', 'Ad Soyad', 'Sınıf']
        if not all(col in df.columns for col in required_columns):
            return jsonify({"error": f"Excel dosyası şu sütunları içermelidir: {', '.join(required_columns)}"}), 400
        
        conn = get_db()
        cur = conn.cursor()
        
        success_count = 0
        error_count = 0
        errors = []
        
        for index, row in df.iterrows():
            username = str(row['Kullanıcı Adı']).strip()
            password = str(row['Şifre']).strip()
            role = str(row['Rol']).strip().lower()
            full_name = str(row['Ad Soyad']).strip()
            class_name = str(row['Sınıf']).strip() if pd.notna(row['Sınıf']) and str(row['Sınıf']).strip() else None
            
            if role not in ['admin', 'teacher', 'student']:
                errors.append(f"Satır {index+2}: Geçersiz rol '{role}'")
                error_count += 1
                continue
            
            if role == 'student' and not class_name:
                errors.append(f"Satır {index+2}: Öğrenci için sınıf bilgisi zorunludur")
                error_count += 1
                continue
            
            # Admin ve teacher için class_name None olmalı
            if role in ['admin', 'teacher']:
                class_name = None
            
            try:
                hashed_password = generate_password_hash(password)
                cur.execute(
                    "INSERT INTO users (username, password, role, full_name, class_name) VALUES (%s, %s, %s, %s, %s)",
                    (username, hashed_password, role, full_name, class_name)
                )
                conn.commit()
                success_count += 1
            except psycopg2.IntegrityError:
                conn.rollback()
                errors.append(f"Satır {index+2}: '{username}' kullanıcı adı zaten kullanılıyor")
                error_count += 1
            except Exception as e:
                conn.rollback()
                errors.append(f"Satır {index+2}: {str(e)}")
                error_count += 1
        
        cur.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "message": f"{success_count} kullanıcı eklendi, {error_count} hata",
            "success_count": success_count,
            "error_count": error_count,
            "errors": errors
        }), 200
        
    except Exception as e:
        return jsonify({"error": f"Excel dosyası işlenirken hata: {str(e)}"}), 400

@app.route("/download/exam-answer-template", methods=["GET"])
def download_exam_answer_template():
    """Sınav cevap anahtarı için örnek Excel şablonu indir"""
    try:
        # Örnek veri oluştur
        data = {
            'Soru': [1, 2, 3, 4, 5, 6, 7, 8, 9, 10],
            'Cevap': ['A', 'B', 'C', 'D', 'A', 'B', 'C', 'D', 'E', 'A']
        }
        df = pd.DataFrame(data)
        
        # Excel dosyasını belleğe yaz
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Cevap Anahtarı')
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=False,
            download_name='sinav_cevap_anahtari_sablonu.xlsx'
        )
    except Exception as e:
        return jsonify({"error": f"Şablon oluşturulurken hata: {str(e)}"}), 400

@app.route("/teacher/exams/create", methods=["POST"])
@login_required
def create_exam():
    if current_user.role != 'teacher':
        return jsonify({"error": "Sadece öğretmenler sınav oluşturabilir"}), 403
    
    try:
        baslik = request.form["baslik"]
        soru_sayisi = int(request.form["soru_sayisi"])
        baslangic_zamani_str = request.form["baslangic_zamani"]
        bitis_zamani_str = request.form.get("bitis_zamani")  # Bitiş tarihi (opsiyonel)
        
        # Sınav süresini gün/saat/dakika biriminden dakikaya çevir
        # Yeni format: sure_deger + sure_birimi | Eski format: sure_dakika (backward compatibility)
        if "sure_deger" in request.form:
            try:
                sure_deger = int(request.form["sure_deger"])
            except (ValueError, TypeError):
                return jsonify({"error": "Sınav süresi geçersiz"}), 400
                
            sure_birimi = request.form.get("sure_birimi", "dakika")
            
            if sure_birimi == "gun":
                sure_dakika = sure_deger * 24 * 60  # Gün → Dakika
            elif sure_birimi == "saat":
                sure_dakika = sure_deger * 60  # Saat → Dakika
            else:  # dakika
                sure_dakika = sure_deger
        elif "sure_dakika" in request.form:
            # Backward compatibility: Eski sınav oluşturma formatı
            try:
                sure_dakika = int(request.form["sure_dakika"])
            except (ValueError, TypeError):
                return jsonify({"error": "Sınav süresi geçersiz"}), 400
        else:
            return jsonify({"error": "Sınav süresi belirtilmedi"}), 400
        
        # Validation: Süre pozitif olmalı
        if sure_dakika <= 0:
            return jsonify({"error": "Sınav süresi en az 1 dakika olmalıdır"}), 400
        
        timezone_offset = int(request.form.get("timezone_offset", 0))  # Dakika cinsinden
        
        # Frontend'den gelen datetime-local'i parse et ve UTC'ye çevir
        try:
            # datetime-local format: "2025-10-26T15:30" (kullanıcının lokal saati)
            # timezone_offset: GMT+3 için -180 (negatif çünkü UTC'den önde)
            from datetime import timezone
            naive_datetime = datetime.strptime(baslangic_zamani_str, "%Y-%m-%dT%H:%M")
            
            # Kullanıcının lokal saatini UTC'ye çevir
            # Örnek: Kullanıcı Türkiye'de (GMT+3) 15:30 yazarsa
            #        timezone_offset = -180 dakika (negatif çünkü UTC'den +3 saat önde)
            #        UTC saati = 15:30 + (-180) dakika / 60 = 15:30 - 3 saat = 12:30
            # getTimezoneOffset() UTC'den farkı döndürür, bu yüzden direkt çıkarıyoruz
            baslangic_zamani = naive_datetime + timedelta(minutes=timezone_offset)
            # UTC timezone bilgisini ekle
            baslangic_zamani = baslangic_zamani.replace(tzinfo=timezone.utc)
            
            # Bitiş tarihini de işle
            bitis_zamani = None
            if bitis_zamani_str:
                naive_bitis = datetime.strptime(bitis_zamani_str, "%Y-%m-%dT%H:%M")
                bitis_zamani = naive_bitis + timedelta(minutes=timezone_offset)
                bitis_zamani = bitis_zamani.replace(tzinfo=timezone.utc)
                
                # Validasyon: Bitiş tarihi başlangıçtan sonra olmalı
                if bitis_zamani <= baslangic_zamani:
                    return jsonify({"error": "Bitiş tarihi başlangıç tarihinden sonra olmalıdır"}), 400
        except ValueError as e:
            return jsonify({"error": f"Geçersiz tarih formatı: {str(e)}"}), 400
        target_classes_raw = request.form.getlist("target_classes[]")
        target_classes = remove_overlapping_classes(target_classes_raw)

        if not target_classes:
            return jsonify({"error": "En az bir hedef sınıf seçmelisiniz"}), 400
        
        # Ara verilebilir sınav seçeneği
        allow_pause = request.form.get("allow_pause") == "on"

        # Çoklu dosya yükleme desteği
        pdf_files = request.files.getlist("pdf_files[]")
        
        if not pdf_files or len(pdf_files) == 0:
            return jsonify({"error": "Sınav dosyası seçilmedi"}), 400
        
        # Tüm dosyaları kaydet (Object Storage veya lokal)
        saved_filenames = []
        for pdf_file in pdf_files:
            if pdf_file and pdf_file.filename:
                # PDF veya resim dosyası kabul et
                if not (allowed_file(pdf_file.filename, 'pdf') or allowed_file(pdf_file.filename, 'image')):
                    return jsonify({"error": f"Geçersiz dosya: {pdf_file.filename}. Sadece PDF veya resim (JPG, PNG) yüklenebilir"}), 400
                
                # Güvenli dosya adı oluştur
                filename = f"exam_{uuid.uuid4()}_{secure_filename(pdf_file.filename)}"
                
                # Object Storage kullan (varsa)
                if object_storage.enabled:
                    try:
                        object_path = f"exams/{filename}"
                        object_storage.upload_from_file(pdf_file, object_path)
                        saved_filenames.append(object_path)  # Object path kaydet
                        logger.info(f"✅ Sınav dosyası Object Storage'a yüklendi: {object_path}")
                    except Exception as storage_error:
                        logger.error(f"⚠️ Object Storage hatası, lokal sisteme geçiliyor: {storage_error}")
                        # Fallback: Lokal dosya sistemi
                        pdf_file.seek(0)  # Stream'i başa sar (CRITICAL!)
                        file_path = os.path.join(UPLOAD_DIR, filename)
                        pdf_file.save(file_path)
                        saved_filenames.append(filename)
                else:
                    # Object Storage devre dışı - lokal sistemi kullan
                    file_path = os.path.join(UPLOAD_DIR, filename)
                    pdf_file.save(file_path)
                    saved_filenames.append(filename)
                    logger.warning("⚠️ Object Storage kullanılamıyor, lokal sistem kullanılıyor")
        
        if len(saved_filenames) == 0:
            return jsonify({"error": "Geçerli dosya yüklenmedi"}), 400
        
        # Dosya isimlerini JSON array olarak sakla
        pdf_filename = json.dumps(saved_filenames)
        
        # Cevap anahtarını al - Manuel veya Excel
        answer_key_dict = {}
        manual_answers = request.form.get('manual_answers')
        
        if manual_answers:
            # Manuel cevap anahtarı
            try:
                answer_key_dict = json.loads(manual_answers)
                
                # Validasyon: Cevap sayısı kontrolü
                if len(answer_key_dict) != soru_sayisi:
                    return jsonify({"error": f"Cevap sayısı ({len(answer_key_dict)}) soru sayısı ({soru_sayisi}) ile uyuşmuyor"}), 400
                
                # Validasyon: Tüm cevapların A-D arasında olduğunu kontrol et
                valid_answers = {'A', 'B', 'C', 'D'}
                for q_num, answer in answer_key_dict.items():
                    if answer not in valid_answers:
                        return jsonify({"error": f"Soru {q_num} için geçersiz cevap: {answer}. Sadece A, B, C, D kabul edilir"}), 400
                
                # Validasyon: Soru numaralarının 1'den soru_sayisi'na kadar olduğunu kontrol et
                expected_questions = {str(i) for i in range(1, soru_sayisi + 1)}
                actual_questions = set(answer_key_dict.keys())
                if expected_questions != actual_questions:
                    return jsonify({"error": "Cevap anahtarında eksik veya fazla soru numaraları var"}), 400
                    
            except json.JSONDecodeError as e:
                return jsonify({"error": f"Manuel cevap anahtarı okunamadı: {str(e)}"}), 400
        else:
            # Excel'den cevap anahtarı
            cevap_file = request.files.get("cevap_file")
            
            if not cevap_file or not cevap_file.filename:
                return jsonify({"error": "Cevap anahtarı seçilmedi (Excel veya Manuel)"}), 400
            
            if not allowed_file(cevap_file.filename, 'excel'):
                return jsonify({"error": "Sadece Excel dosyası yüklenebilir (xlsx, xls)"}), 400
            
            # Excel'i bellekten oku (disk'e kaydetmeden)
            try:
                cevap_file.seek(0)
                df = pd.read_excel(cevap_file)
                for idx, row in df.iterrows():
                    q_num = str(row.iloc[0])  # İlk sütun soru numarası
                    answer = str(row.iloc[1]).strip().upper()  # İkinci sütun cevap
                    answer_key_dict[q_num] = answer
                logger.info(f"✅ Cevap anahtarı Excel'den okundu: {len(answer_key_dict)} soru")
            except Exception as e:
                return jsonify({"error": f"Cevap anahtarı okunamadı: {str(e)}"}), 400
        
        # Her hedef sınıf için sınav kaydı oluştur
        conn = get_db()
        cur = conn.cursor()
        
        for target_class in target_classes:
            exam_id = str(uuid.uuid4())  # Unique ID üret
            cur.execute(
                """INSERT INTO exams (id, title, pdf_filename, answer_key, question_count, start_time, end_time, duration_minutes, teacher_id, target_class, allow_pause)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                (exam_id, baslik, pdf_filename, json.dumps(answer_key_dict), soru_sayisi, baslangic_zamani, bitis_zamani, sure_dakika, current_user.id, target_class, allow_pause)
            )
        
        conn.commit()
        cur.close()
        conn.close()
        
        # Yetkili kullanıcılar için bildirim gönder (hedef sınıflara)
        if can_send_notification(current_user):
            send_push_notification(
                title="Yeni Sınav Eklendi",
                message=f"{baslik} sınavı eklendi. Hemen inceleyin!",
                url="https://ameo-alanya.com",
                target_classes=target_classes,
                target_role="student"
            )
        
        return jsonify({"success": True, "message": "Sınav oluşturuldu"}), 200
    except Exception as e:
        return jsonify({"error": f"Sınav oluşturulamadı: {str(e)}"}), 500


@app.route("/teacher/assignments/create", methods=["POST"])
@login_required
def create_assignment():
    if current_user.role not in ['teacher', 'admin']:
        return jsonify({"error": "Sadece öğretmenler ödev oluşturabilir"}), 403
    
    try:
        title = request.form["title"]
        description = request.form.get("description", "")
        due_date = request.form["due_date"]
        target_classes_raw = request.form.getlist("target_classes[]")
        target_classes = remove_overlapping_classes(target_classes_raw)
        
        if not target_classes:
            return jsonify({"error": "En az bir hedef sınıf seçmelisiniz"}), 400
        
        # Opsiyonel dosya yükleme (Object Storage veya lokal)
        filename = None
        if "file" in request.files:
            file = request.files["file"]
            if file and file.filename:
                if not allowed_file(file.filename):
                    return jsonify({"error": "Desteklenmeyen dosya formatı"}), 400
                
                base_filename = f"assignment_{uuid.uuid4()}_{secure_filename(file.filename)}"
                
                # Object Storage kullan (varsa)
                if object_storage.enabled:
                    try:
                        object_path = f"assignments/{base_filename}"
                        object_storage.upload_from_file(file, object_path)
                        filename = object_path  # Object path kaydet
                        logger.info(f"✅ Ödev dosyası Object Storage'a yüklendi: {object_path}")
                    except Exception as storage_error:
                        logger.error(f"⚠️ Object Storage hatası, lokal sisteme geçiliyor: {storage_error}")
                        file.seek(0)  # Stream'i başa sar (CRITICAL!)
                        file_path = os.path.join(UPLOAD_DIR, base_filename)
                        file.save(file_path)
                        filename = base_filename
                else:
                    # Object Storage devre dışı - lokal sistemi kullan
                    file_path = os.path.join(UPLOAD_DIR, base_filename)
                    file.save(file_path)
                    filename = base_filename
        
        conn = get_db()
        cur = conn.cursor()
        
        for target_class in target_classes:
            cur.execute(
                """INSERT INTO assignments (title, description, due_date, file_path, teacher_id, target_class)
                   VALUES (%s, %s, %s, %s, %s, %s)""",
                (title, description, due_date, filename, current_user.id, target_class)
            )
        
        conn.commit()
        cur.close()
        conn.close()
        
        # Yetkili kullanıcılar için bildirim gönder (hedef sınıflara)
        if can_send_notification(current_user):
            send_push_notification(
                title="Yeni Ödev Eklendi",
                message=f"{title} ödevi eklendi. Teslim tarihi: {due_date}",
                url="https://ameo-alanya.com",
                target_classes=target_classes,
                target_role="student"
            )
        
        return jsonify({"success": True, "message": "Ödev oluşturuldu"}), 200
    except Exception as e:
        logger.error(f"Assignment creation error: {str(e)}")
        return jsonify({"error": f"Ödev oluşturulamadı: {str(e)}"}), 500

@app.route("/teacher/assignments/<int:assignment_id>/report", methods=["GET"])
@login_required
def get_assignment_report(assignment_id):
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    cur.execute("SELECT teacher_id FROM assignments WHERE id = %s", (assignment_id,))
    assignment = cur.fetchone()
    
    if not assignment:
        cur.close()
        conn.close()
        return jsonify({"error": "Ödev bulunamadı"}), 404
    
    if assignment['teacher_id'] != current_user.id:
        cur.close()
        conn.close()
        return jsonify({"error": "Bu ödeve erişim yetkiniz yok"}), 403
    
    cur.execute(
        """SELECT u.full_name, asub.submitted_at, asub.file_path
           FROM assignment_submissions asub
           JOIN users u ON asub.student_id = u.id
           WHERE asub.assignment_id = %s
           ORDER BY asub.submitted_at DESC""",
        (assignment_id,)
    )
    submissions = cur.fetchall()
    cur.close()
    conn.close()
    
    return jsonify([{
        "student_name": s['full_name'],
        "submitted_at": s['submitted_at'].isoformat(),
        "file_path": s['file_path']
    } for s in submissions])

@app.route("/teacher/announcements/create", methods=["POST"])
@login_required
def create_announcement():
    if current_user.role != 'teacher':
        return jsonify({"error": "Sadece öğretmenler duyuru yapabilir"}), 403
    
    try:
        title = request.form["title"]
        content = request.form.get("content", "")
        video_url = request.form.get("video_url", "").strip() or None
        target_classes_raw = request.form.getlist("target_classes[]")
        target_classes = remove_overlapping_classes(target_classes_raw)
        
        if not target_classes:
            return jsonify({"error": "En az bir hedef sınıf seçmelisiniz"}), 400
        
        filename = None
        file_type = None
        if "file" in request.files:
            file = request.files["file"]
            if file and file.filename:
                if not allowed_file(file.filename):
                    return jsonify({"error": "Desteklenmeyen dosya formatı. İzin verilen: PDF, Resim (JPG, PNG, GIF)"}), 400
                
                base_filename = f"announcement_{uuid.uuid4()}_{secure_filename(file.filename)}"
                
                # Dosya tipi belirle
                ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
                if ext in ['jpg', 'jpeg', 'png', 'gif']:
                    file_type = 'image'
                elif ext == 'pdf':
                    file_type = 'pdf'
                
                # Object Storage kullan (varsa)
                if object_storage.enabled:
                    try:
                        object_path = f"announcements/{base_filename}"
                        object_storage.upload_from_file(file, object_path)
                        filename = object_path  # Object path kaydet
                        logger.info(f"✅ Duyuru dosyası Object Storage'a yüklendi: {object_path}")
                    except Exception as storage_error:
                        logger.error(f"⚠️ Object Storage hatası, lokal sisteme geçiliyor: {storage_error}")
                        file.seek(0)  # Stream'i başa sar (CRITICAL!)
                        file_path = os.path.join(UPLOAD_DIR, base_filename)
                        file.save(file_path)
                        filename = base_filename
                else:
                    # Object Storage devre dışı - lokal sistemi kullan
                    file_path = os.path.join(UPLOAD_DIR, base_filename)
                    file.save(file_path)
                    filename = base_filename
        
        # Her hedef sınıf için duyuru kaydı oluştur
        conn = get_db()
        cur = conn.cursor()
        
        for target_class in target_classes:
            cur.execute(
                """INSERT INTO announcements (title, content, file_path, file_type, video_url, teacher_id, target_class)
                   VALUES (%s, %s, %s, %s, %s, %s, %s)""",
                (title, content, filename, file_type, video_url, current_user.id, target_class)
            )
        
        conn.commit()
        cur.close()
        conn.close()
        
        # Yetkili kullanıcılar için bildirim gönder (hedef sınıflara + öğretmenlere)
        if can_send_notification(current_user):
            # Öğrencilere bildirim
            send_push_notification(
                title="Yeni Duyuru",
                message=f"{title} - {content[:80]}..." if len(content) > 80 else f"{title} - {content}",
                url="https://ameo-alanya.com",
                target_classes=target_classes,
                target_role="student"
            )
            # Admin ise öğretmenlere de bildirim gönder
            if current_user.role == 'admin':
                send_push_notification(
                    title="Yeni Duyuru (Admin)",
                    message=f"{title} - {content[:80]}..." if len(content) > 80 else f"{title} - {content}",
                    url="https://ameo-alanya.com",
                    target_role="teacher"
                )
        
        return jsonify({"success": True, "message": "Duyuru yayınlandı"}), 200
    except Exception as e:
        return jsonify({"error": f"Duyuru yayınlanırken hata: {str(e)}"}), 400

@app.route("/announcements", methods=["GET"])
@login_required
def get_announcements():
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    # Eğer öğrenci ise, sınıf filtrelemesi yap
    if current_user.role == 'student':
        cur.execute("SELECT class_name FROM users WHERE id = %s", (current_user.id,))
        user_data = cur.fetchone()
        student_class = user_data['class_name'] if user_data else None
        
        if not student_class:
            cur.close()
            conn.close()
            return jsonify([])
        
        class_level = student_class[0] if student_class else None
        
        cur.execute(
            """SELECT a.*, u.full_name as teacher_name
               FROM announcements a
               JOIN users u ON a.teacher_id = u.id
               WHERE a.target_class = %s OR a.target_class = %s OR a.target_class = 'tüm_okul'
               ORDER BY a.created_at DESC""",
            (student_class, f'tüm_{class_level}')
        )
    else:
        # Öğretmen veya admin ise tüm duyuruları göster
        cur.execute(
            """SELECT a.*, u.full_name as teacher_name
               FROM announcements a
               JOIN users u ON a.teacher_id = u.id
               ORDER BY a.created_at DESC"""
        )
    
    announcements = cur.fetchall()
    cur.close()
    conn.close()
    
    return jsonify([{
        "id": an['id'],
        "title": an['title'],
        "content": an['content'],
        "file_path": an['file_path'],
        "file_type": an['file_type'],
        "video_url": an.get('video_url'),
        "teacher_name": an['teacher_name'],
        "created_at": an['created_at'].isoformat()
    } for an in announcements])

@app.route("/storage/<path:filepath>")
@login_required
def get_storage_file(filepath):
    """Object Storage'dan dosya serve et"""
    if current_user.role not in ['teacher', 'student', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        data, content_type = get_file_data(f"/storage/{filepath}")
        response = send_file(
            BytesIO(data),
            mimetype=content_type,
            as_attachment=False
        )
        
        # PDF için tarayıcıda göster
        if content_type == 'application/pdf':
            response.headers['Content-Disposition'] = 'inline'
        
        return response
    except FileNotFoundError:
        return jsonify({"error": "Dosya bulunamadı"}), 404
    except Exception as e:
        logger.error(f"Storage file error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/download-file/<path:filename>")
def download_file_direct(filename):
    """APK'da dosya indirmesi için - direkt indirme endpoint'i (session olmadan çalışmalı)
    Önbellek ve sıkıştırma desteği ile optimize edilmiş
    """
    logger.info(f"📥 APK indirme isteği - Dosya: {filename}")
    
    try:
        # Güvenlik kontrolü: Dosya path'ini doğrula
        if '..' in filename or filename.startswith('/'):
            logger.warning(f"❌ Geçersiz path: {filename}")
            return jsonify({"error": "Geçersiz dosya path'i"}), 400
        
        # Dosya uzantısı
        ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
        mime_types = {
            'pdf': 'application/pdf',
            'jpg': 'image/jpeg',
            'jpeg': 'image/jpeg',
            'png': 'image/png',
            'gif': 'image/gif',
            'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'xls': 'application/vnd.ms-excel',
        }
        mimetype = mime_types.get(ext, 'application/octet-stream')
        
        file_data = None
        content_type = mimetype
        base_filename = filename.split('/')[-1]
        
        # Object Storage path mı?
        is_object_storage = '/' in filename
        
        if is_object_storage and object_storage.enabled:
            try:
                file_data, content_type = object_storage.download_as_bytes(filename)
                content_type = content_type or mimetype
            except:
                pass
        
        # Lokal file system
        if file_data is None:
            attached_path = os.path.join('attached_assets', base_filename)
            file_path = os.path.join(UPLOAD_DIR, base_filename)
            
            if os.path.exists(attached_path):
                with open(attached_path, 'rb') as f:
                    file_data = f.read()
            elif os.path.exists(file_path):
                with open(file_path, 'rb') as f:
                    file_data = f.read()
        
        if file_data is None:
            return jsonify({"error": "Dosya bulunamadı"}), 404
        
        # Sıkıştırma kontrolü
        accept_encoding = request.headers.get('Accept-Encoding', '')
        use_gzip = 'gzip' in accept_encoding and len(file_data) > 1024
        
        if use_gzip:
            compressed_data = gzip.compress(file_data, compresslevel=6)
            response = make_response(compressed_data)
            response.headers['Content-Encoding'] = 'gzip'
            response.headers['Content-Length'] = len(compressed_data)
            logger.info(f"📦 Gzip sıkıştırma: {len(file_data)} -> {len(compressed_data)} bytes")
        else:
            response = make_response(file_data)
            response.headers['Content-Length'] = len(file_data)
        
        # Header'lar
        response.headers['Content-Type'] = content_type
        response.headers['Content-Disposition'] = f'attachment; filename="{base_filename}"'
        
        # Önbellek - 1 saat
        response.headers['Cache-Control'] = 'public, max-age=3600'
        response.headers['Vary'] = 'Accept-Encoding'
        
        # CORS
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Expose-Headers'] = 'Content-Disposition, Content-Length'
        
        return response
        
    except Exception as e:
        logger.error(f"Download error: {str(e)}")
        return jsonify({"error": "İndirme hatası"}), 500

@app.route("/view-file/<path:filename>")
def view_file_direct(filename):
    """APK'da dosya görüntüleme için - WebViewer'da PDF/PNG direkt açılacak
    Önbellek ve sıkıştırma desteği ile optimize edilmiş
    """
    logger.info(f"👁️ APK dosya görüntüleme isteği - Dosya: {filename}")
    
    try:
        # Güvenlik kontrolü
        if '..' in filename or filename.startswith('/'):
            logger.warning(f"❌ Geçersiz path: {filename}")
            return jsonify({"error": "Geçersiz dosya path'i"}), 400
        
        # Dosya uzantısı
        ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
        mime_types = {
            'pdf': 'application/pdf',
            'jpg': 'image/jpeg',
            'jpeg': 'image/jpeg',
            'png': 'image/png',
            'gif': 'image/gif',
        }
        mimetype = mime_types.get(ext, 'application/octet-stream')
        
        file_data = None
        content_type = mimetype
        
        # Object Storage path mı?
        is_object_storage = '/' in filename
        
        if is_object_storage and object_storage.enabled:
            try:
                file_data, content_type = object_storage.download_as_bytes(filename)
                content_type = content_type or mimetype
            except:
                pass
        
        # Lokal file system
        if file_data is None:
            base_filename = filename.split('/')[-1]
            attached_path = os.path.join('attached_assets', base_filename)
            file_path = os.path.join(UPLOAD_DIR, base_filename)
            
            if os.path.exists(attached_path):
                with open(attached_path, 'rb') as f:
                    file_data = f.read()
            elif os.path.exists(file_path):
                with open(file_path, 'rb') as f:
                    file_data = f.read()
        
        if file_data is None:
            return jsonify({"error": "Dosya bulunamadı"}), 404
        
        # Sıkıştırma kontrolü - istemci gzip destekliyorsa sıkıştır
        accept_encoding = request.headers.get('Accept-Encoding', '')
        use_gzip = 'gzip' in accept_encoding and len(file_data) > 1024  # 1KB üzeri dosyalarda sıkıştır
        
        if use_gzip:
            # Gzip sıkıştırma
            compressed_data = gzip.compress(file_data, compresslevel=6)
            response = make_response(compressed_data)
            response.headers['Content-Encoding'] = 'gzip'
            response.headers['Content-Length'] = len(compressed_data)
            logger.info(f"📦 Gzip sıkıştırma: {len(file_data)} -> {len(compressed_data)} bytes (%{100-int(len(compressed_data)/len(file_data)*100)} tasarruf)")
        else:
            response = make_response(file_data)
            response.headers['Content-Length'] = len(file_data)
        
        # Temel header'lar
        response.headers['Content-Type'] = content_type
        response.headers['Content-Disposition'] = 'inline'
        
        # Önbellek header'ları - 1 saat önbellek (APK için idealdir)
        response.headers['Cache-Control'] = 'public, max-age=3600, immutable'
        response.headers['ETag'] = f'"{hash(file_data[:1000]) if len(file_data) > 1000 else hash(file_data)}"'
        response.headers['Vary'] = 'Accept-Encoding'
        
        # CORS header'ları (APK erişimi için)
        response.headers['Access-Control-Allow-Origin'] = '*'
        response.headers['Access-Control-Allow-Methods'] = 'GET, OPTIONS'
        
        return response
        
    except Exception as e:
        logger.error(f"View file error: {str(e)}")
        return jsonify({"error": "Dosya görüntüleme hatası"}), 500

@app.route("/apk-file-viewer")
def apk_file_viewer():
    """APK WebViewer'da çalışacak dosya yönetim sayfası"""
    return render_template('apk_viewer.html')

@app.route("/uploads/<path:filename>")
@login_required
def get_file(filename):
    """Basitleştirilmiş dosya sunma - Object Storage veya lokal file system"""
    logger.info(f"🔍 Dosya erişim isteği - User: {current_user.username}, Role: {current_user.role}, Class: {getattr(current_user, 'class_name', 'N/A')}, Dosya: {filename}")
    
    # Sadece teacher ve student erişebilir
    if current_user.role not in ['teacher', 'student', 'admin']:
        logger.warning(f"❌ Yetkisiz rol erişimi: {current_user.role}")
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    # Object Storage path mı? (örn: "exams/exam_uuid.pdf")
    is_object_storage = '/' in filename
    
    # Dosya adını çıkar (path'siz)
    base_filename = filename.split('/')[-1] if '/' in filename else filename
    logger.info(f"📂 Dosya analizi - is_object_storage: {is_object_storage}, base_filename: {base_filename}")
    
    # Lokal file system check (sadece eski dosyalar için)
    if not is_object_storage:
        file_path = os.path.join(UPLOAD_DIR, filename)
        if not os.path.exists(file_path):
            return jsonify({"error": "Dosya bulunamadı"}), 404
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # ÖDEV DOSYALARI - Hedef sınıf kontrolü yap (hem tam path hem basename ile)
        # ÖNEMLİ: Aynı dosya birden fazla sınıfa yüklenmiş olabilir (6A, 6B, 6C, 6D)
        cur.execute("""
            SELECT a.teacher_id, a.target_class
            FROM assignments a 
            WHERE a.file_path = %s OR a.file_path LIKE %s
        """, (filename, f'%{base_filename}'))
        assignments = cur.fetchall()  # TÜM kayıtları al
        if assignments:
            # Öğretmen kontrolü
            if current_user.role == 'teacher':
                for assignment in assignments:
                    if assignment['teacher_id'] == current_user.id:
                        cur.close()
                        conn.close()
                        return serve_file_with_mime(filename)
            # Öğrenci kontrolü
            elif current_user.role == 'student':
                for assignment in assignments:
                    if assignment['target_class'] == current_user.class_name:
                        cur.close()
                        conn.close()
                        return serve_file_with_mime(filename)
        
        # DUYURU DOSYALARI - Hedef sınıf kontrolü yap (hem tam path hem basename ile)
        # ÖNEMLİ: Aynı dosya birden fazla sınıfa yüklenmiş olabilir (6A, 6B, 6C, 6D)
        cur.execute("""
            SELECT an.teacher_id, an.target_class
            FROM announcements an 
            WHERE an.file_path = %s OR an.file_path LIKE %s
        """, (filename, f'%{base_filename}'))
        announcements = cur.fetchall()  # TÜM kayıtları al
        if announcements:
            # Öğretmen kontrolü
            if current_user.role == 'teacher':
                for announcement in announcements:
                    if announcement['teacher_id'] == current_user.id:
                        cur.close()
                        conn.close()
                        return serve_file_with_mime(filename)
            # Öğrenci kontrolü
            elif current_user.role == 'student':
                for announcement in announcements:
                    if announcement['target_class'] == current_user.class_name:
                        cur.close()
                        conn.close()
                        return serve_file_with_mime(filename)
        
        # SINAV DOSYALARI
        # pdf_filename JSON array olarak saklanıyor, içinde tam path veya basename arama yap
        # ÖNEMLİ: Aynı dosya birden fazla sınıfa yüklenmiş olabilir (6A, 6B, 6C, 6D)
        # Bu yüzden fetchall() kullanıp TÜM kayıtları kontrol ediyoruz
        cur.execute("""
            SELECT e.teacher_id, e.target_class, e.title
            FROM exams e 
            WHERE e.pdf_filename::text LIKE %s OR e.pdf_filename::text LIKE %s
        """, (f'%{filename}%', f'%{base_filename}%'))
        exams = cur.fetchall()  # TÜM kayıtları al
        if exams:
            logger.info(f"📝 {len(exams)} sınav kaydı bulundu (çoklu sınıf olabilir)")
            # Öğretmen kontrolü (herhangi bir kayıtta öğretmen kendisi mi?)
            if current_user.role == 'teacher':
                for exam in exams:
                    if exam['teacher_id'] == current_user.id:
                        logger.info(f"✅ Erişim onaylandı (Öğretmen - kendi sınavı)")
                        cur.close()
                        conn.close()
                        return serve_file_with_mime(filename)
            # Öğrenci kontrolü (herhangi bir kayıtta sınıf eşleşiyor mu?)
            elif current_user.role == 'student':
                for exam in exams:
                    if exam['target_class'] == current_user.class_name:
                        logger.info(f"✅ Erişim onaylandı (Öğrenci - sınıf eşleşiyor: {current_user.class_name})")
                        cur.close()
                        conn.close()
                        return serve_file_with_mime(filename)
                # Hiçbir kayıtta sınıf eşleşmedi
                target_classes = [exam['target_class'] for exam in exams]
                logger.warning(f"❌ Sınav erişim reddedildi - User class: {current_user.class_name}, Exam targets: {target_classes}")
        
        # ÖĞRENCİ SORU DOSYALARI (hem tam path hem basename ile)
        cur.execute("""
            SELECT sq.student_id, sq.teacher_id
            FROM student_questions sq 
            WHERE sq.file_path = %s OR sq.file_path LIKE %s
        """, (filename, f'%{base_filename}'))
        question = cur.fetchone()
        if question:
            # Öğrenci kendi dosyasını görebilir
            if current_user.role == 'student' and question['student_id'] == current_user.id:
                cur.close()
                conn.close()
                return serve_file_with_mime(filename)
            # Öğretmen kendisine gönderilen dosyayı görebilir
            elif current_user.role == 'teacher' and question['teacher_id'] == current_user.id:
                cur.close()
                conn.close()
                return serve_file_with_mime(filename)
        
        # ÖDEV TESLİM DOSYALARI (hem tam path hem basename ile)
        cur.execute("""
            SELECT ass.student_id, a.teacher_id
            FROM assignment_submissions ass
            JOIN assignments a ON ass.assignment_id = a.id
            WHERE ass.file_path = %s OR ass.file_path LIKE %s
        """, (filename, f'%{base_filename}'))
        submission = cur.fetchone()
        if submission:
            # Öğrenci kendi dosyasını görebilir
            if current_user.role == 'student' and submission['student_id'] == current_user.id:
                cur.close()
                conn.close()
                return serve_file_with_mime(filename)
            # Öğretmen öğrencisinin dosyasını görebilir
            elif current_user.role == 'teacher' and submission['teacher_id'] == current_user.id:
                cur.close()
                conn.close()
                return serve_file_with_mime(filename)
        
        # Admin her dosyayı görebilir
        if current_user.role == 'admin':
            cur.close()
            conn.close()
            return serve_file_with_mime(filename)
        
        cur.close()
        conn.close()
        return jsonify({"error": "Bu dosyaya erişim yetkiniz yok"}), 403
        
    except Exception as e:
        logger.error(f"File access error: {str(e)}")
        return jsonify({"error": "Dosya erişim hatası"}), 500


def serve_file_with_mime(filename):
    """Dosyayı indir olarak sun - fetch + blob yöntemi için düzgün binary response
    Öğretmen raporlarındaki çalışan pattern kullanılıyor
    """
    # Dosya uzantısına göre MIME type belirle
    ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
    
    mime_types = {
        'pdf': 'application/pdf',
        'jpg': 'image/jpeg',
        'jpeg': 'image/jpeg',
        'png': 'image/png',
        'gif': 'image/gif',
        'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'xls': 'application/vnd.ms-excel',
        'doc': 'application/msword',
        'docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        'txt': 'text/plain'
    }
    
    mimetype = mime_types.get(ext, 'application/octet-stream')
    base_filename = filename.split('/')[-1]
    
    # Object Storage path mı? (örn: "exams/exam_uuid.pdf")
    is_object_storage = '/' in filename
    
    # Dosya verisini al
    file_data = None
    source = None
    
    if is_object_storage and object_storage.enabled:
        try:
            file_data, content_type = object_storage.download_as_bytes(filename)
            source = "Object Storage"
            logger.info(f"📦 Object Storage'dan okundu: {filename}")
        except FileNotFoundError:
            logger.warning(f"⚠️ Object Storage'da dosya bulunamadı: {filename}")
        except Exception as storage_error:
            logger.error(f"⚠️ Object Storage hatası: {storage_error}")
    
    # Lokal file system dene
    if file_data is None:
        attached_path = os.path.join('attached_assets', base_filename)
        file_path = os.path.join(UPLOAD_DIR, base_filename)
        
        if os.path.exists(attached_path):
            with open(attached_path, 'rb') as f:
                file_data = f.read()
            source = "attached_assets"
        elif os.path.exists(file_path):
            with open(file_path, 'rb') as f:
                file_data = f.read()
            source = "uploads"
    
    if file_data is None:
        logger.error(f"❌ Dosya bulunamadı: {filename}")
        return jsonify({"error": "Dosya bulunamadı"}), 404
    
    # Düzgün binary response döndür (öğretmen raporları pattern'i)
    response = send_file(
        BytesIO(file_data),
        mimetype=mimetype,
        as_attachment=True,
        download_name=base_filename
    )
    
    # fetch + blob için gerekli header'lar
    response.headers['Content-Type'] = mimetype
    response.headers['Content-Disposition'] = f'attachment; filename="{base_filename}"'
    response.headers['Content-Length'] = len(file_data)
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Expose-Headers'] = 'Content-Disposition, Content-Length'
    
    logger.info(f"✅ Dosya {source}'dan indirildi: {filename} ({len(file_data)} bytes)")
    return response

def init_database():
    """Veritabanı tablolarını oluşturur (yoksa)"""
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # Users tablosu
        cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                username VARCHAR(100) UNIQUE NOT NULL,
                password VARCHAR(255) NOT NULL,
                role VARCHAR(20) NOT NULL,
                full_name VARCHAR(200) NOT NULL,
                class_name VARCHAR(10),
                student_no VARCHAR(20),
                last_login_at TIMESTAMP,
                last_logout_at TIMESTAMP,
                total_session_duration INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul')
            )
        """)
        
        # student_no sütunu yoksa ekle (mevcut veritabanları için)
        cur.execute("""
            DO $$ 
            BEGIN 
                IF NOT EXISTS (SELECT 1 FROM information_schema.columns 
                               WHERE table_name='users' AND column_name='student_no') THEN
                    ALTER TABLE users ADD COLUMN student_no VARCHAR(20);
                END IF;
            END $$;
        """)
        
        # Exams tablosu
        cur.execute("""
            CREATE TABLE IF NOT EXISTS exams (
                id VARCHAR(50) PRIMARY KEY,
                title VARCHAR(200) NOT NULL,
                question_count INTEGER NOT NULL,
                start_time TIMESTAMP WITH TIME ZONE NOT NULL,
                duration_minutes INTEGER NOT NULL,
                pdf_filename VARCHAR(255) NOT NULL,
                answer_key JSONB NOT NULL,
                target_class VARCHAR(20),
                teacher_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
                created_at TIMESTAMP WITH TIME ZONE DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul')
            )
        """)
        
        # Exam submissions tablosu (öğrenci sınav teslimi)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS exam_submissions (
                id SERIAL PRIMARY KEY,
                exam_id VARCHAR(100) NOT NULL,
                student_id INTEGER NOT NULL,
                answers JSONB NOT NULL,
                score DECIMAL(5,2) NOT NULL,
                submitted_at TIMESTAMP WITH TIME ZONE DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul'),
                FOREIGN KEY (exam_id) REFERENCES exams(id) ON DELETE CASCADE,
                FOREIGN KEY (student_id) REFERENCES users(id) ON DELETE CASCADE,
                UNIQUE(exam_id, student_id)
            )
        """)
        
        # Sınav tablosuna allow_pause sütunu ekle (ara verilebilir sınav)
        try:
            cur.execute("ALTER TABLE exams ADD COLUMN IF NOT EXISTS allow_pause BOOLEAN DEFAULT FALSE")
        except Exception as e:
            logger.info(f"exams.allow_pause sütunu zaten mevcut: {e}")
        
        # Sınav tablosuna end_time sütunu ekle (bitiş tarihi)
        try:
            cur.execute("ALTER TABLE exams ADD COLUMN IF NOT EXISTS end_time TIMESTAMP WITH TIME ZONE")
        except Exception as e:
            logger.info(f"exams.end_time sütunu zaten mevcut: {e}")
        
        # Exam submissions tablosuna status sütunu ekle (draft/submitted)
        try:
            cur.execute("ALTER TABLE exam_submissions ADD COLUMN IF NOT EXISTS status VARCHAR(20) DEFAULT 'submitted'")
            # Mevcut kayıtları 'submitted' olarak işaretle (NULL olanları düzelt)
            cur.execute("UPDATE exam_submissions SET status = 'submitted' WHERE status IS NULL")
        except Exception as e:
            logger.info(f"exam_submissions.status sütunu zaten mevcut: {e}")
        
        # Assignments tablosu
        cur.execute("""
            CREATE TABLE IF NOT EXISTS assignments (
                id SERIAL PRIMARY KEY,
                title VARCHAR(200) NOT NULL,
                description TEXT,
                file_path VARCHAR(255),
                video_url TEXT,
                due_date TIMESTAMP NOT NULL,
                target_class VARCHAR(20),
                teacher_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
                created_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul')
            )
        """)
        
        # Assignment submissions tablosu
        cur.execute("""
            CREATE TABLE IF NOT EXISTS assignment_submissions (
                id SERIAL PRIMARY KEY,
                assignment_id INTEGER REFERENCES assignments(id) ON DELETE CASCADE,
                student_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
                file_path VARCHAR(255) NOT NULL,
                submitted_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul'),
                UNIQUE(assignment_id, student_id)
            )
        """)
        
        # Announcements tablosu
        cur.execute("""
            CREATE TABLE IF NOT EXISTS announcements (
                id SERIAL PRIMARY KEY,
                title VARCHAR(200) NOT NULL,
                content TEXT,
                file_path VARCHAR(255),
                file_type VARCHAR(50),
                video_url TEXT,
                target_class VARCHAR(20),
                teacher_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
                created_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul')
            )
        """)
        
        # Student Questions tablosu
        cur.execute("""
            CREATE TABLE IF NOT EXISTS student_questions (
                id SERIAL PRIMARY KEY,
                student_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
                teacher_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
                question_text TEXT,
                file_path VARCHAR(255),
                answer_text TEXT,
                answered_at TIMESTAMP,
                created_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul')
            )
        """)
        
        # Teacher Announcements tablosu (Admin'den öğretmenlere duyurular)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS teacher_announcements (
                id SERIAL PRIMARY KEY,
                title VARCHAR(200) NOT NULL,
                content TEXT,
                file_path VARCHAR(255),
                video_url TEXT,
                target_teachers TEXT,
                admin_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
                created_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul')
            )
        """)
        
        # Public Announcements tablosu (Ana sayfa duyuruları)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS public_announcements (
                id SERIAL PRIMARY KEY,
                title VARCHAR(200) NOT NULL,
                content TEXT NOT NULL,
                admin_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
                is_active BOOLEAN DEFAULT TRUE,
                created_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul')
            )
        """)
        
        cur.execute("""
            CREATE TABLE IF NOT EXISTS practice_exams (
                id SERIAL PRIMARY KEY,
                student_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
                exam_number INTEGER NOT NULL CHECK (exam_number >= 1 AND exam_number <= 50),
                
                turkce_dogru INTEGER DEFAULT 0 CHECK (turkce_dogru >= 0 AND turkce_dogru <= 20),
                turkce_yanlis INTEGER DEFAULT 0 CHECK (turkce_yanlis >= 0 AND turkce_yanlis <= 20),
                turkce_net DECIMAL(5,2) DEFAULT 0,
                
                matematik_dogru INTEGER DEFAULT 0 CHECK (matematik_dogru >= 0 AND matematik_dogru <= 20),
                matematik_yanlis INTEGER DEFAULT 0 CHECK (matematik_yanlis >= 0 AND matematik_yanlis <= 20),
                matematik_net DECIMAL(5,2) DEFAULT 0,
                
                fen_dogru INTEGER DEFAULT 0 CHECK (fen_dogru >= 0 AND fen_dogru <= 20),
                fen_yanlis INTEGER DEFAULT 0 CHECK (fen_yanlis >= 0 AND fen_yanlis <= 20),
                fen_net DECIMAL(5,2) DEFAULT 0,
                
                sosyal_dogru INTEGER DEFAULT 0 CHECK (sosyal_dogru >= 0 AND sosyal_dogru <= 10),
                sosyal_yanlis INTEGER DEFAULT 0 CHECK (sosyal_yanlis >= 0 AND sosyal_yanlis <= 10),
                sosyal_net DECIMAL(5,2) DEFAULT 0,
                
                ingilizce_dogru INTEGER DEFAULT 0 CHECK (ingilizce_dogru >= 0 AND ingilizce_dogru <= 10),
                ingilizce_yanlis INTEGER DEFAULT 0 CHECK (ingilizce_yanlis >= 0 AND ingilizce_yanlis <= 10),
                ingilizce_net DECIMAL(5,2) DEFAULT 0,
                
                din_dogru INTEGER DEFAULT 0 CHECK (din_dogru >= 0 AND din_dogru <= 10),
                din_yanlis INTEGER DEFAULT 0 CHECK (din_yanlis >= 0 AND din_yanlis <= 10),
                din_net DECIMAL(5,2) DEFAULT 0,
                
                lgs_score DECIMAL(6,2) DEFAULT 0,
                created_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul'),
                
                UNIQUE(student_id, exam_number)
            )
        """)
        
        # Classes tablosu (Dinamik sınıf/grup yönetimi)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS classes (
                id SERIAL PRIMARY KEY,
                name VARCHAR(100) UNIQUE NOT NULL,
                level VARCHAR(10),
                branch VARCHAR(10),
                type VARCHAR(20) DEFAULT 'standard',
                is_active BOOLEAN DEFAULT TRUE,
                created_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul')
            )
        """)
        
        # Lesson Schedules tablosu (Ders ve Kurs Programı)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS lesson_schedules (
                id SERIAL PRIMARY KEY,
                student_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                day_of_week VARCHAR(20) NOT NULL,
                lesson_name VARCHAR(100) NOT NULL,
                course_name VARCHAR(100),
                start_time TIME NOT NULL,
                end_time TIME NOT NULL,
                created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                created_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul'),
                UNIQUE(student_id, day_of_week, start_time)
            )
        """)
        
        # Video URL sütunlarını ekle (eğer yoksa)
        try:
            cur.execute("""
                ALTER TABLE assignments ADD COLUMN IF NOT EXISTS video_url TEXT
            """)
            cur.execute("""
                ALTER TABLE announcements ADD COLUMN IF NOT EXISTS video_url TEXT
            """)
            cur.execute("""
                ALTER TABLE teacher_announcements ADD COLUMN IF NOT EXISTS file_path VARCHAR(255)
            """)
            cur.execute("""
                ALTER TABLE teacher_announcements ADD COLUMN IF NOT EXISTS video_url TEXT
            """)
            cur.execute("""
                ALTER TABLE teacher_announcements ADD COLUMN IF NOT EXISTS target_teachers TEXT
            """)
        except Exception as e:
            logger.info(f"Sütunlar zaten mevcut veya hata: {e}")
        
        # Lesson schedules için index (overlap check optimizasyonu)
        try:
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_lesson_schedules_student_day 
                ON lesson_schedules(student_id, day_of_week)
            """)
        except Exception as e:
            logger.info(f"Index zaten mevcut veya hata: {e}")
        
        # Exam Calendar tablosu (Deneme Sınavı Takvimi)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS exam_calendar (
                id SERIAL PRIMARY KEY,
                exam_date DATE NOT NULL,
                exam_title VARCHAR(255) NOT NULL,
                description TEXT,
                created_by INTEGER REFERENCES users(id) ON DELETE SET NULL,
                created_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul'),
                UNIQUE(exam_date)
            )
        """)
        
        try:
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_exam_calendar_date 
                ON exam_calendar(exam_date)
            """)
        except Exception as e:
            logger.info(f"Index zaten mevcut: {e}")
        
        # Teacher Study Plan tablosu (Öğretmen Ders Çalışma Planı)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS teacher_study_plan (
                id SERIAL PRIMARY KEY,
                student_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                teacher_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                plan_date DATE NOT NULL,
                subject VARCHAR(100) NOT NULL,
                question_count INTEGER NOT NULL CHECK (question_count > 0),
                note TEXT,
                created_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul'),
                updated_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul'),
                UNIQUE(student_id, plan_date, subject)
            )
        """)
        
        try:
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_teacher_study_plan_student 
                ON teacher_study_plan(student_id, plan_date)
            """)
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_teacher_study_plan_teacher 
                ON teacher_study_plan(teacher_id, plan_date)
            """)
        except Exception as e:
            logger.info(f"Index zaten mevcut: {e}")
        
        # Study Plan PDF tablosu (PDF Çalışma Programı)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS study_plan_pdf (
                id SERIAL PRIMARY KEY,
                teacher_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                title VARCHAR(255) NOT NULL,
                file_path VARCHAR(500) NOT NULL,
                target_type VARCHAR(20) NOT NULL CHECK (target_type IN ('class', 'students')),
                target_class VARCHAR(10),
                target_students TEXT,
                created_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul')
            )
        """)
        
        try:
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_study_plan_pdf_teacher 
                ON study_plan_pdf(teacher_id, created_at DESC)
            """)
        except Exception as e:
            logger.info(f"Index zaten mevcut: {e}")
        
        # User Sessions tablosu (Aktivite Takibi)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS user_sessions (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                login_at TIMESTAMP NOT NULL,
                logout_at TIMESTAMP,
                duration_minutes INTEGER,
                ip_address VARCHAR(45),
                user_agent VARCHAR(255)
            )
        """)
        
        # Daily Study Tracking tablosu (Günlük Çalışma Takibi)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS daily_study_tracking (
                id SERIAL PRIMARY KEY,
                student_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                date DATE NOT NULL,
                day_of_week VARCHAR(20) NOT NULL,
                subject VARCHAR(100) NOT NULL,
                note TEXT,
                created_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul'),
                updated_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul'),
                UNIQUE(student_id, date, subject)
            )
        """)
        
        # Gamification: Student Badges (Task #7)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS student_badges (
                id SERIAL PRIMARY KEY,
                student_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                badge_type VARCHAR(50) NOT NULL,
                badge_title VARCHAR(100) NOT NULL,
                badge_description TEXT,
                earned_date TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul')
            )
        """)
        
        # Gamification: Student Points (Task #7)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS student_points (
                student_id INTEGER PRIMARY KEY REFERENCES users(id) ON DELETE CASCADE,
                total_points INTEGER DEFAULT 0,
                exam_count INTEGER DEFAULT 0,
                assignment_count INTEGER DEFAULT 0,
                updated_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul')
            )
        """)
        
        # Report Card Exams tablosu (CSV/Excel upload için)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS report_card_exams (
                id SERIAL PRIMARY KEY,
                exam_name VARCHAR(200) NOT NULL,
                exam_date DATE,
                grade_level INTEGER,
                class_name VARCHAR(20),
                answer_key_a JSON,
                answer_key_b JSON,
                question_counts JSON,
                created_by INTEGER REFERENCES users(id),
                created_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul')
            )
        """)
        
        # Report Card Results tablosu (CSV/Excel upload sonuçları için)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS report_card_results (
                id SERIAL PRIMARY KEY,
                exam_id INTEGER REFERENCES report_card_exams(id) ON DELETE CASCADE,
                student_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
                student_no VARCHAR(50),
                student_name VARCHAR(200),
                class_name VARCHAR(20),
                grade_level INTEGER,
                booklet_type VARCHAR(1),
                subjects JSON,
                totals JSON,
                percentile FLOAT,
                created_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul'),
                UNIQUE(exam_id, student_id)
            )
        """)
        
        # Parent Portal: Parent-Student Relationships (Task #8)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS parent_children (
                id SERIAL PRIMARY KEY,
                parent_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                student_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                relationship VARCHAR(50) DEFAULT 'Veli',
                created_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul'),
                UNIQUE(parent_id, student_id)
            )
        """)
        
        # Parent Portal: Parent-Teacher Messages (Task #8)
        # Table exists from coaching.py - add parent_id column if missing
        cur.execute("""
            CREATE TABLE IF NOT EXISTS parent_messages (
                id SERIAL PRIMARY KEY,
                student_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                parent_name VARCHAR(200),
                parent_email VARCHAR(200),
                parent_phone VARCHAR(50),
                message TEXT NOT NULL,
                teacher_response TEXT,
                status VARCHAR(20) DEFAULT 'unread',
                created_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul'),
                responded_at TIMESTAMP
            )
        """)
        
        # Idempotent migration: Add parent_id column if not exists
        try:
            cur.execute("""
                ALTER TABLE parent_messages 
                ADD COLUMN IF NOT EXISTS parent_id INTEGER REFERENCES users(id) ON DELETE SET NULL
            """)
        except Exception as e:
            logger.info(f"parent_id column already exists or error: {e}")
        
        # Dashboard Widget System (Task #10)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS dashboard_widget_preferences (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                role VARCHAR(20) NOT NULL,
                layout JSON DEFAULT '[]'::json,
                visibility JSON DEFAULT '{}'::json,
                updated_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul'),
                UNIQUE(user_id)
            )
        """)
        
        try:
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_dashboard_widget_preferences_user 
                ON dashboard_widget_preferences(user_id)
            """)
        except Exception as e:
            logger.info(f"Index zaten mevcut: {e}")
        
        try:
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_daily_study_tracking_student 
                ON daily_study_tracking(student_id, date)
            """)
        except Exception as e:
            logger.info(f"Index zaten mevcut: {e}")
        
        # Survey/Anket Tabloları
        cur.execute("""
            CREATE TABLE IF NOT EXISTS surveys (
                id SERIAL PRIMARY KEY,
                title VARCHAR(255) NOT NULL,
                description TEXT,
                created_by INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                creator_role VARCHAR(20) NOT NULL,
                target_role VARCHAR(20) NOT NULL,
                target_class TEXT,
                start_date TIMESTAMP NOT NULL,
                end_date TIMESTAMP NOT NULL,
                status VARCHAR(20) DEFAULT 'active',
                is_anonymous BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul'),
                updated_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul')
            )
        """)
        
        cur.execute("""
            CREATE TABLE IF NOT EXISTS survey_questions (
                id SERIAL PRIMARY KEY,
                survey_id INTEGER NOT NULL REFERENCES surveys(id) ON DELETE CASCADE,
                question_text TEXT NOT NULL,
                question_type VARCHAR(50) NOT NULL,
                options JSONB,
                is_required BOOLEAN DEFAULT TRUE,
                question_order INTEGER NOT NULL,
                created_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul')
            )
        """)
        
        cur.execute("""
            CREATE TABLE IF NOT EXISTS survey_responses (
                id SERIAL PRIMARY KEY,
                survey_id INTEGER NOT NULL REFERENCES surveys(id) ON DELETE CASCADE,
                question_id INTEGER NOT NULL REFERENCES survey_questions(id) ON DELETE CASCADE,
                student_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                answer_text TEXT,
                answer_option VARCHAR(255),
                submitted_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul'),
                UNIQUE(survey_id, question_id, student_id)
            )
        """)
        
        try:
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_surveys_creator 
                ON surveys(created_by, status)
            """)
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_survey_questions_survey 
                ON survey_questions(survey_id)
            """)
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_survey_responses_survey 
                ON survey_responses(survey_id, student_id)
            """)
        except Exception as e:
            logger.info(f"Survey indexes zaten mevcut: {e}")
        
        cur.execute("""
            CREATE TABLE IF NOT EXISTS book_challenges (
                id SERIAL PRIMARY KEY,
                book_title VARCHAR(255) NOT NULL,
                teacher_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                target_class VARCHAR(20),
                questions JSONB NOT NULL,
                status VARCHAR(20) DEFAULT 'active',
                created_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul')
            )
        """)
        
        cur.execute("""
            CREATE TABLE IF NOT EXISTS book_challenge_submissions (
                id SERIAL PRIMARY KEY,
                challenge_id INTEGER NOT NULL REFERENCES book_challenges(id) ON DELETE CASCADE,
                student_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                answers JSONB NOT NULL,
                status VARCHAR(20) DEFAULT 'pending',
                rejection_reason TEXT,
                reviewed_by INTEGER REFERENCES users(id),
                reviewed_at TIMESTAMP,
                submitted_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul'),
                UNIQUE(challenge_id, student_id)
            )
        """)
        
        try:
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_book_challenges_teacher 
                ON book_challenges(teacher_id, status)
            """)
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_book_submissions_student 
                ON book_challenge_submissions(student_id, status)
            """)
        except Exception as e:
            logger.info(f"Book challenge indexes zaten mevcut: {e}")
        
        # Yeni Kitap Kurdu sistemi - 6 sabit sorulu basit sistem
        cur.execute("""
            CREATE TABLE IF NOT EXISTS book_entries (
                id SERIAL PRIMARY KEY,
                student_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                book_title VARCHAR(255) NOT NULL,
                page_count INTEGER NOT NULL,
                answers JSONB NOT NULL,
                status VARCHAR(20) DEFAULT 'pending',
                rejection_reason TEXT,
                reviewed_by INTEGER REFERENCES users(id),
                reviewed_at TIMESTAMP,
                submitted_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul')
            )
        """)
        
        try:
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_book_entries_student 
                ON book_entries(student_id, status)
            """)
        except Exception as e:
            logger.info(f"Book entries index zaten mevcut: {e}")
        
        # Dokümanlar tablosu (Yazılı Örnekleri genişletildi)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS exam_samples (
                id SERIAL PRIMARY KEY,
                teacher_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                title VARCHAR(255) NOT NULL,
                description TEXT,
                sample_type VARCHAR(20) NOT NULL,
                file_path TEXT,
                link_url TEXT,
                target_classes JSONB NOT NULL,
                category VARCHAR(50) DEFAULT 'exam_samples',
                subject VARCHAR(50),
                view_count INTEGER DEFAULT 0,
                download_count INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul')
            )
        """)
        
        # Mevcut tabloya yeni sütunlar ekle
        try:
            cur.execute("ALTER TABLE exam_samples ADD COLUMN IF NOT EXISTS category VARCHAR(50) DEFAULT 'exam_samples'")
            cur.execute("ALTER TABLE exam_samples ADD COLUMN IF NOT EXISTS subject VARCHAR(50)")
            cur.execute("ALTER TABLE exam_samples ADD COLUMN IF NOT EXISTS view_count INTEGER DEFAULT 0")
            cur.execute("ALTER TABLE exam_samples ADD COLUMN IF NOT EXISTS download_count INTEGER DEFAULT 0")
        except Exception as e:
            logger.info(f"exam_samples columns already exist: {e}")
        
        cur.execute("""
            CREATE TABLE IF NOT EXISTS optical_exams (
                id SERIAL PRIMARY KEY,
                exam_name VARCHAR(100) NOT NULL,
                exam_number INTEGER NOT NULL,
                grade_level INTEGER NOT NULL,
                answer_key_a JSONB,
                answer_key_b JSONB,
                question_counts JSONB NOT NULL,
                uploaded_by INTEGER REFERENCES users(id),
                is_published BOOLEAN DEFAULT FALSE,
                school_average DECIMAL(5,2),
                school_std_dev DECIMAL(5,2),
                created_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul')
            )
        """)
        
        cur.execute("""
            CREATE TABLE IF NOT EXISTS optical_student_results (
                id SERIAL PRIMARY KEY,
                optical_exam_id INTEGER NOT NULL REFERENCES optical_exams(id) ON DELETE CASCADE,
                student_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                booklet_type VARCHAR(1) NOT NULL,
                raw_answers JSONB NOT NULL,
                results JSONB NOT NULL,
                total_correct INTEGER DEFAULT 0,
                total_wrong INTEGER DEFAULT 0,
                total_empty INTEGER DEFAULT 0,
                total_net DECIMAL(5,2) DEFAULT 0,
                total_score DECIMAL(5,2) DEFAULT 0,
                school_rank INTEGER,
                class_rank INTEGER,
                created_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul'),
                UNIQUE(optical_exam_id, student_id)
            )
        """)
        
        try:
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_optical_exams_grade 
                ON optical_exams(grade_level, is_published)
            """)
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_optical_results_student 
                ON optical_student_results(student_id, optical_exam_id)
            """)
            cur.execute("""
                CREATE INDEX IF NOT EXISTS idx_optical_results_exam 
                ON optical_student_results(optical_exam_id, school_rank)
            """)
        except Exception as e:
            logger.info(f"Optical indexes zaten mevcut: {e}")
        
        conn.commit()
        
        # Mevcut sınavların grade_level'ini sınav adından düzelt
        try:
            cur.execute("""
                UPDATE report_card_exams 
                SET grade_level = CAST(SUBSTRING(exam_name FROM '^([0-9]+)') AS INTEGER)
                WHERE grade_level IS NULL 
                AND exam_name ~ '^[0-9]+'
            """)
            conn.commit()
        except Exception as e:
            logger.info(f"Grade level güncelleme: {e}")
        
        logger.info("✅ Veritabanı tabloları kontrol edildi/oluşturuldu")
        
        cur.close()
        conn.close()
    except Exception as e:
        logger.info(f"⚠️ Veritabanı oluşturulurken hata: {e}")

def init_admin_user():
    """Varsayılan admin kullanıcısını oluşturur - GÜVENLİ şifre ile"""
    import secrets
    import string
    
    # Güçlü rastgele şifre oluştur (sadece ilk kurulumda)
    def generate_secure_password(length=16):
        alphabet = string.ascii_letters + string.digits + "!@#$%&*"
        return ''.join(secrets.choice(alphabet) for _ in range(length))
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("SELECT * FROM users WHERE username = 'admin' AND role = 'admin'")
        admin = cur.fetchone()
        
        if not admin:
            # Ortam değişkeninden şifre al veya güçlü rastgele şifre oluştur
            admin_password = os.environ.get('ADMIN_PASSWORD', generate_secure_password())
            hashed_password = generate_password_hash(admin_password)
            cur.execute(
                "INSERT INTO users (username, password, role, full_name) VALUES (%s, %s, %s, %s)",
                ('admin', hashed_password, 'admin', 'System Administrator')
            )
            conn.commit()
            logger.info(f"✅ Admin kullanıcısı oluşturuldu. Şifre güvenli olarak ayarlandı.")
            # Güvenlik: Şifreyi loglama!
        else:
            logger.info("✅ Admin kullanıcısı zaten mevcut")
        
        cur.close()
        conn.close()
    except Exception as e:
        logger.info(f"⚠️ Admin kullanıcısı oluşturulurken hata: {e}")

def init_default_classes():
    """Varsayılan sınıfları (5A-8E) oluşturur"""
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Mevcut sınıf sayısını kontrol et
        cur.execute("SELECT COUNT(*) as count FROM classes")
        result = cur.fetchone()
        
        if result and result['count'] == 0:
            # Varsayılan 20 sınıfı ekle (5A-8E)
            default_classes = []
            for level in ['5', '6', '7', '8']:
                for branch in ['A', 'B', 'C', 'D', 'E']:
                    class_name = f"{level}{branch}"
                    default_classes.append((class_name, level, branch, 'standard'))
            
            cur.executemany(
                "INSERT INTO classes (name, level, branch, type) VALUES (%s, %s, %s, %s)",
                default_classes
            )
            conn.commit()
            logger.info(f"✅ {len(default_classes)} varsayılan sınıf oluşturuldu (5A-8E)")
        else:
            logger.info(f"✅ Sınıflar zaten mevcut ({result['count']} adet)")
        
        cur.close()
        conn.close()
    except Exception as e:
        logger.info(f"⚠️ Varsayılan sınıflar oluşturulurken hata: {e}")

# Öğretmen - Kendi sınavlarını listele
@app.route("/teacher/my-exams", methods=["GET"])
@login_required
def get_my_exams():
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute(
        """SELECT id, title, question_count, start_time, end_time, duration_minutes, target_class, created_at
           FROM exams WHERE teacher_id = %s ORDER BY created_at DESC""",
        (current_user.id,)
    )
    exams = cur.fetchall()
    cur.close()
    conn.close()
    
    return jsonify({"exams": [{
        "id": e['id'],
        "title": e['title'],
        "question_count": e['question_count'],
        "start_time": e['start_time'].isoformat() if e['start_time'] else None,
        "end_time": e['end_time'].isoformat() if e.get('end_time') else None,
        "duration_minutes": e['duration_minutes'],
        "target_class": e['target_class']
    } for e in exams]})

# Öğretmen - Sınav sil
@app.route("/teacher/exams/<exam_id>", methods=["DELETE"])
@login_required
def delete_exam(exam_id):
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # Önce bu sınavın öğretmenin sınavı olduğunu doğrula
        cur.execute("SELECT teacher_id FROM exams WHERE id = %s", (exam_id,))
        exam = cur.fetchone()
        
        if not exam or exam[0] != current_user.id:
            cur.close()
            conn.close()
            return jsonify({"error": "Bu sınavı silme yetkiniz yok"}), 403
        
        # Önce sınav sonuçlarını sil (foreign key constraint için)
        cur.execute("DELETE FROM exam_results WHERE exam_id = %s", (exam_id,))
        
        # Sonra sınavı sil
        cur.execute("DELETE FROM exams WHERE id = %s", (exam_id,))
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Sınav ve sonuçları silindi"})
    except Exception as e:
        logger.error(f" Delete exam error: {str(e)}")
        return jsonify({"error": f"Sınav silinemedi: {str(e)}"}), 500

# Öğretmen - Sınav raporu API (kimler yaptı, doğru/yanlış analizi)
@app.route("/teacher/api/exams/<exam_id>/report", methods=["GET"])
@login_required
def get_exam_report(exam_id):
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Sınav bilgilerini al ve yetki kontrolü yap
        cur.execute("""
            SELECT id, title, question_count, start_time, duration_minutes, 
                   target_class, answer_key, teacher_id
            FROM exams WHERE id = %s
        """, (exam_id,))
        exam = cur.fetchone()
        
        if not exam:
            cur.close()
            conn.close()
            return jsonify({"error": "Sınav bulunamadı"}), 404
        
        if exam['teacher_id'] != current_user.id:
            cur.close()
            conn.close()
            return jsonify({"error": "Bu sınava erişim yetkiniz yok"}), 403
        
        # Hedef sınıftaki tüm öğrencileri al
        cur.execute("""
            SELECT id, full_name, username
            FROM users
            WHERE role = 'student' AND class_name = %s
            ORDER BY full_name
        """, (exam['target_class'],))
        students = cur.fetchall()
        
        # Sınav sonuçlarını al
        cur.execute("""
            SELECT student_id, answers, score, submitted_at
            FROM exam_submissions
            WHERE exam_id = %s
        """, (exam_id,))
        submissions = {s['student_id']: s for s in cur.fetchall()}
        
        cur.close()
        conn.close()
        
        # Her öğrenci için rapor oluştur
        student_reports = []
        answer_key = exam['answer_key']  # {"1":"A", "2":"B", ...}
        
        for student in students:
            student_id = student['id']
            submission = submissions.get(student_id)
            
            if submission:
                # Sınavı yaptı - detaylı analiz
                student_answers = submission['answers']  # {"1":"A", "2":"B", ...}
                question_details = []
                correct_count = 0
                wrong_count = 0
                empty_count = 0
                
                for q_num in range(1, exam['question_count'] + 1):
                    q_str = str(q_num)
                    student_answer = student_answers.get(q_str, "")
                    correct_answer = answer_key.get(q_str, "")
                    
                    if not student_answer:
                        status = "empty"
                        empty_count += 1
                    elif student_answer == correct_answer:
                        status = "correct"
                        correct_count += 1
                    else:
                        status = "wrong"
                        wrong_count += 1
                    
                    question_details.append({
                        "question_number": q_num,
                        "student_answer": student_answer or "-",
                        "correct_answer": correct_answer,
                        "status": status
                    })
                
                student_reports.append({
                    "student_id": student_id,
                    "student_name": student['full_name'],
                    "status": "completed",
                    "score": float(submission['score']),
                    "correct_count": correct_count,
                    "wrong_count": wrong_count,
                    "empty_count": empty_count,
                    "submitted_at": submission['submitted_at'].isoformat(),
                    "question_details": question_details
                })
            else:
                # Sınavı yapmadı
                student_reports.append({
                    "student_id": student_id,
                    "student_name": student['full_name'],
                    "status": "not_submitted",
                    "score": 0,
                    "correct_count": 0,
                    "wrong_count": 0,
                    "empty_count": exam['question_count']
                })
        
        # Özet istatistikler
        completed_count = sum(1 for r in student_reports if r['status'] == 'completed')
        not_submitted_count = len(student_reports) - completed_count
        avg_score = sum(r['score'] for r in student_reports if r['status'] == 'completed') / completed_count if completed_count > 0 else 0
        
        return jsonify({
            "exam": {
                "id": exam['id'],
                "title": exam['title'],
                "question_count": exam['question_count'],
                "target_class": exam['target_class'],
                "start_time": exam['start_time'].isoformat()
            },
            "summary": {
                "total_students": len(student_reports),
                "completed": completed_count,
                "not_submitted": not_submitted_count,
                "average_score": round(avg_score, 2)
            },
            "students": student_reports
        })
    
    except Exception as e:
        logger.error(f"Exam report error: {str(e)}")
        return jsonify({"error": f"Rapor alınamadı: {str(e)}"}), 500

# Öğretmen - Kendi ödevlerini listele
@app.route("/teacher/my-assignments", methods=["GET"])
@login_required
def get_my_assignments():
    if current_user.role not in ['teacher', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    # Admin tüm ödevleri görebilir, öğretmen sadece kendisininkileri
    if current_user.role == 'admin':
        cur.execute(
            """SELECT a.id, a.title, a.description, a.due_date, a.target_class, a.created_at, a.teacher_id,
                      u.full_name as teacher_name
               FROM assignments a
               LEFT JOIN users u ON a.teacher_id = u.id
               ORDER BY a.created_at DESC"""
        )
    else:
        cur.execute(
            """SELECT id, title, description, due_date, target_class, created_at, teacher_id
               FROM assignments WHERE teacher_id = %s ORDER BY created_at DESC""",
            (current_user.id,)
        )
    assignments = cur.fetchall()
    cur.close()
    conn.close()
    
    return jsonify({"assignments": [{
        "id": a['id'],
        "title": a['title'],
        "description": a['description'],
        "due_date": a['due_date'].isoformat() if a['due_date'] else None,
        "target_class": a['target_class'],
        "teacher_name": a.get('teacher_name', 'Bilinmiyor')
    } for a in assignments]})

# Öğretmen - Ödev sil
@app.route("/teacher/assignments/<int:assignment_id>", methods=["DELETE"])
@login_required
def delete_assignment(assignment_id):
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Ödevin öğretmene ait olup olmadığını kontrol et
        cur.execute("SELECT * FROM assignments WHERE id = %s AND teacher_id = %s", (assignment_id, current_user.id))
        assignment = cur.fetchone()
        
        if not assignment:
            cur.close()
            conn.close()
            return jsonify({"error": "Ödev bulunamadı veya size ait değil"}), 404
        
        # Önce ödev teslimlerini sil (foreign key constraint)
        cur.execute("DELETE FROM assignment_submissions WHERE assignment_id = %s", (assignment_id,))
        
        # Sonra ödevi sil
        cur.execute("DELETE FROM assignments WHERE id = %s", (assignment_id,))
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Ödev ve teslimler başarıyla silindi"})
    
    except Exception as e:
        logger.error(f" Delete assignment error: {str(e)}")
        return jsonify({"error": f"Ödev silinemedi: {str(e)}"}), 500

# Öğretmen - Kendi duyurularını listele
@app.route("/teacher/my-announcements", methods=["GET"])
@login_required
def get_my_announcements():
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute(
        """SELECT id, title, content, target_class, created_at
           FROM announcements WHERE teacher_id = %s ORDER BY created_at DESC""",
        (current_user.id,)
    )
    announcements = cur.fetchall()
    cur.close()
    conn.close()
    
    return jsonify({
        "announcements": [{
            "id": a['id'],
            "title": a['title'],
            "content": a['content'],
            "target_class": a['target_class'],
            "created_at": a['created_at'].isoformat()
        } for a in announcements]
    })

# Öğretmen - Duyuru sil
@app.route("/teacher/announcements/<int:announcement_id>", methods=["DELETE"])
@login_required
def delete_announcement(announcement_id):
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Duyurunun öğretmene ait olup olmadığını kontrol et
        cur.execute("SELECT * FROM announcements WHERE id = %s AND teacher_id = %s", (announcement_id, current_user.id))
        announcement = cur.fetchone()
        
        if not announcement:
            cur.close()
            conn.close()
            return jsonify({"error": "Duyuru bulunamadı veya size ait değil"}), 404
        
        # Duyuruyu sil
        cur.execute("DELETE FROM announcements WHERE id = %s", (announcement_id,))
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Duyuru başarıyla silindi"})
    
    except Exception as e:
        logger.error(f" Delete announcement error: {str(e)}")
        return jsonify({"error": f"Duyuru silinemedi: {str(e)}"}), 500

# Admin - Son aktiviteler (dashboard için)
@app.route("/api/admin/recent-activity", methods=["GET"])
@login_required
def get_recent_activity():
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim", "success": False}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        activities = []
        
        # Son sınavlar
        cur.execute("""
            SELECT e.title, e.created_at, u.full_name as teacher_name
            FROM exams e
            JOIN users u ON e.teacher_id = u.id
            ORDER BY e.created_at DESC
            LIMIT 3
        """)
        for exam in cur.fetchall():
            activities.append({
                "icon": "📝",
                "color": "blue",
                "title": f"Yeni sınav: {exam['title']}",
                "description": f"{exam['teacher_name']} tarafından oluşturuldu",
                "time": exam['created_at'].strftime("%d.%m.%Y %H:%M") if exam['created_at'] else "",
                "created_at": exam['created_at']
            })
        
        # Son ödevler
        cur.execute("""
            SELECT a.title, a.created_at, u.full_name as teacher_name
            FROM assignments a
            JOIN users u ON a.teacher_id = u.id
            ORDER BY a.created_at DESC
            LIMIT 3
        """)
        for assignment in cur.fetchall():
            activities.append({
                "icon": "📚",
                "color": "green",
                "title": f"Yeni ödev: {assignment['title']}",
                "description": f"{assignment['teacher_name']} tarafından oluşturuldu",
                "time": assignment['created_at'].strftime("%d.%m.%Y %H:%M") if assignment['created_at'] else "",
                "created_at": assignment['created_at']
            })
        
        # Son duyurular
        cur.execute("""
            SELECT a.title, a.created_at, u.full_name as teacher_name
            FROM announcements a
            JOIN users u ON a.teacher_id = u.id
            ORDER BY a.created_at DESC
            LIMIT 3
        """)
        for announcement in cur.fetchall():
            activities.append({
                "icon": "📢",
                "color": "orange",
                "title": f"Yeni duyuru: {announcement['title']}",
                "description": f"{announcement['teacher_name']} tarafından paylaşıldı",
                "time": announcement['created_at'].strftime("%d.%m.%Y %H:%M") if announcement['created_at'] else "",
                "created_at": announcement['created_at']
            })
        
        # Son kayıtlar
        cur.execute("""
            SELECT full_name, role, created_at
            FROM users
            ORDER BY created_at DESC
            LIMIT 3
        """)
        for user in cur.fetchall():
            role_name = {"admin": "Yönetici", "teacher": "Öğretmen", "student": "Öğrenci"}.get(user['role'], user['role'])
            activities.append({
                "icon": "👤",
                "color": "purple",
                "title": f"Yeni kayıt: {user['full_name']}",
                "description": f"{role_name} olarak kaydoldu",
                "time": user['created_at'].strftime("%d.%m.%Y %H:%M") if user['created_at'] else "",
                "created_at": user['created_at']
            })
        
        cur.close()
        conn.close()
        
        # Tarihe göre sırala (en yeni en üstte) - datetime objesine göre
        from datetime import datetime as dt
        activities.sort(key=lambda x: x.get('created_at') or dt.min, reverse=True)
        
        # created_at alanını JSON'dan kaldır (frontend'e gerekmez)
        for activity in activities:
            activity.pop('created_at', None)
        
        return jsonify({"success": True, "activities": activities[:10]})
        
    except Exception as e:
        logger.error(f"Recent activity error: {str(e)}")
        return jsonify({"success": False, "error": str(e), "activities": []})

# Admin - Öğretmen raporları
@app.route("/api/admin/teacher-reports", methods=["GET"])
@login_required
def get_teacher_reports():
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("""
        SELECT 
            u.id,
            u.full_name,
            COUNT(DISTINCT e.id) as exam_count,
            COUNT(DISTINCT a.id) as assignment_count,
            COUNT(DISTINCT an.id) as announcement_count
        FROM users u
        LEFT JOIN exams e ON u.id = e.teacher_id
        LEFT JOIN assignments a ON u.id = a.teacher_id
        LEFT JOIN announcements an ON u.id = an.teacher_id
        WHERE u.role = 'teacher'
        GROUP BY u.id, u.full_name
        ORDER BY u.full_name
    """)
    teachers = cur.fetchall()
    cur.close()
    conn.close()
    
    return jsonify({
        "teachers": [{
            "teacher_id": t['id'],
            "full_name": t['full_name'],
            "exam_count": t['exam_count'],
            "assignment_count": t['assignment_count'],
            "announcement_count": t['announcement_count']
        } for t in teachers]
    })

# Admin - Öğretmen detay endpoint'leri (sınav/ödev/duyuru içerikleri)
@app.route("/api/admin/teacher/<int:teacher_id>/exams", methods=["GET"])
@login_required
def get_teacher_exams_detail(teacher_id):
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("""
        SELECT id, title, target_class, question_count, duration_minutes, start_time, created_at
        FROM exams
        WHERE teacher_id = %s
        ORDER BY created_at DESC
    """, (teacher_id,))
    exams = cur.fetchall()
    cur.close()
    conn.close()
    
    # Datetime objelerini string'e çevir
    exams_list = []
    for exam in exams:
        exam_dict = dict(exam)
        if exam_dict.get('start_time'):
            exam_dict['start_time'] = exam_dict['start_time'].strftime('%Y-%m-%d %H:%M:%S')
        if exam_dict.get('created_at'):
            exam_dict['created_at'] = exam_dict['created_at'].strftime('%Y-%m-%d %H:%M:%S')
        exams_list.append(exam_dict)
    
    return jsonify({"exams": exams_list})

@app.route("/api/admin/teacher/<int:teacher_id>/assignments", methods=["GET"])
@login_required
def get_teacher_assignments_detail(teacher_id):
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("""
        SELECT id, title, description, target_class, due_date, created_at
        FROM assignments
        WHERE teacher_id = %s
        ORDER BY created_at DESC
    """, (teacher_id,))
    assignments = cur.fetchall()
    cur.close()
    conn.close()
    
    # Datetime objelerini string'e çevir
    assignments_list = []
    for assignment in assignments:
        assignment_dict = dict(assignment)
        if assignment_dict.get('due_date'):
            assignment_dict['due_date'] = assignment_dict['due_date'].strftime('%Y-%m-%d %H:%M:%S')
        if assignment_dict.get('created_at'):
            assignment_dict['created_at'] = assignment_dict['created_at'].strftime('%Y-%m-%d %H:%M:%S')
        assignments_list.append(assignment_dict)
    
    return jsonify({"assignments": assignments_list})

@app.route("/api/admin/teacher/<int:teacher_id>/announcements", methods=["GET"])
@login_required
def get_teacher_announcements_detail(teacher_id):
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("""
        SELECT id, title, content, target_class, created_at
        FROM announcements
        WHERE teacher_id = %s
        ORDER BY created_at DESC
    """, (teacher_id,))
    announcements = cur.fetchall()
    cur.close()
    conn.close()
    
    # Datetime objelerini string'e çevir
    announcements_list = []
    for announcement in announcements:
        announcement_dict = dict(announcement)
        if announcement_dict.get('created_at'):
            announcement_dict['created_at'] = announcement_dict['created_at'].strftime('%Y-%m-%d %H:%M:%S')
        announcements_list.append(announcement_dict)
    
    return jsonify({"announcements": announcements_list})

# Admin - Sınav sil
@app.route("/api/admin/exams/<exam_id>", methods=["DELETE"])
@login_required
def delete_exam_admin(exam_id):
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # Önce sınav varını kontrol et
        cur.execute("SELECT id FROM exams WHERE id = %s", (exam_id,))
        exam = cur.fetchone()
        
        if not exam:
            cur.close()
            conn.close()
            return jsonify({"error": "Sınav bulunamadı"}), 404
        
        # Sınavı sil
        cur.execute("DELETE FROM exams WHERE id = %s", (exam_id,))
        conn.commit()
        cur.close()
        conn.close()
        
        logger.info(f"✅ Admin {current_user.username} sınavı sildi: {exam_id}")
        return jsonify({"success": True, "message": "Sınav silindi"})
        
    except Exception as e:
        logger.error(f"Admin exam deletion error: {str(e)}")
        return jsonify({"error": f"Sınav silinemedi: {str(e)}"}), 500

# Admin - Ödev sil
@app.route("/api/admin/assignments/<int:assignment_id>", methods=["DELETE"])
@login_required
def delete_assignment_admin(assignment_id):
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # Önce ödev varlığını kontrol et
        cur.execute("SELECT id FROM assignments WHERE id = %s", (assignment_id,))
        assignment = cur.fetchone()
        
        if not assignment:
            cur.close()
            conn.close()
            return jsonify({"error": "Ödev bulunamadı"}), 404
        
        # Ödevi sil
        cur.execute("DELETE FROM assignments WHERE id = %s", (assignment_id,))
        conn.commit()
        cur.close()
        conn.close()
        
        logger.info(f"✅ Admin {current_user.username} ödevi sildi: {assignment_id}")
        return jsonify({"success": True, "message": "Ödev silindi"})
        
    except Exception as e:
        logger.error(f"Admin assignment deletion error: {str(e)}")
        return jsonify({"error": f"Ödev silinemedi: {str(e)}"}), 500

# Admin - Duyuru sil
@app.route("/api/admin/announcements/<int:announcement_id>", methods=["DELETE"])
@login_required
def delete_announcement_admin(announcement_id):
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # Önce duyuru varlığını kontrol et
        cur.execute("SELECT id FROM announcements WHERE id = %s", (announcement_id,))
        announcement = cur.fetchone()
        
        if not announcement:
            cur.close()
            conn.close()
            return jsonify({"error": "Duyuru bulunamadı"}), 404
        
        # Duyuruyu sil
        cur.execute("DELETE FROM announcements WHERE id = %s", (announcement_id,))
        conn.commit()
        cur.close()
        conn.close()
        
        logger.info(f"✅ Admin {current_user.username} duyuruyu sildi: {announcement_id}")
        return jsonify({"success": True, "message": "Duyuru silindi"})
        
    except Exception as e:
        logger.error(f"Admin announcement deletion error: {str(e)}")
        return jsonify({"error": f"Duyuru silinemedi: {str(e)}"}), 500

# Öğretmen - Sınav sil
@app.route("/api/teacher/exams/<exam_id>", methods=["DELETE"])
@login_required
def delete_exam_teacher(exam_id):
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # Önce sınavın bu öğretmene ait olup olmadığını kontrol et
        cur.execute("SELECT id FROM exams WHERE id = %s AND teacher_id = %s", (exam_id, current_user.id))
        exam = cur.fetchone()
        
        if not exam:
            cur.close()
            conn.close()
            return jsonify({"error": "Sınav bulunamadı veya size ait değil"}), 404
        
        # Sınavı sil
        cur.execute("DELETE FROM exams WHERE id = %s", (exam_id,))
        conn.commit()
        cur.close()
        conn.close()
        
        logger.info(f"✅ Öğretmen {current_user.username} sınavı sildi: {exam_id}")
        return jsonify({"success": True, "message": "Sınav silindi"})
        
    except Exception as e:
        logger.error(f"Teacher exam deletion error: {str(e)}")
        return jsonify({"error": f"Sınav silinemedi: {str(e)}"}), 500

# Öğretmen - Ödev sil
@app.route("/api/teacher/assignments/<int:assignment_id>", methods=["DELETE"])
@login_required
def delete_assignment_teacher(assignment_id):
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # Önce ödevin bu öğretmene ait olup olmadığını kontrol et
        cur.execute("SELECT id FROM assignments WHERE id = %s AND teacher_id = %s", (assignment_id, current_user.id))
        assignment = cur.fetchone()
        
        if not assignment:
            cur.close()
            conn.close()
            return jsonify({"error": "Ödev bulunamadı veya size ait değil"}), 404
        
        # Ödevi sil
        cur.execute("DELETE FROM assignments WHERE id = %s", (assignment_id,))
        conn.commit()
        cur.close()
        conn.close()
        
        logger.info(f"✅ Öğretmen {current_user.username} ödevi sildi: {assignment_id}")
        return jsonify({"success": True, "message": "Ödev silindi"})
        
    except Exception as e:
        logger.error(f"Teacher assignment deletion error: {str(e)}")
        return jsonify({"error": f"Ödev silinemedi: {str(e)}"}), 500

# Öğretmen - Duyuru sil
@app.route("/api/teacher/announcements/<int:announcement_id>", methods=["DELETE"])
@login_required
def delete_announcement_teacher(announcement_id):
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # Önce duyurunun bu öğretmene ait olup olmadığını kontrol et
        cur.execute("SELECT id FROM announcements WHERE id = %s AND teacher_id = %s", (announcement_id, current_user.id))
        announcement = cur.fetchone()
        
        if not announcement:
            cur.close()
            conn.close()
            return jsonify({"error": "Duyuru bulunamadı veya size ait değil"}), 404
        
        # Duyuruyu sil
        cur.execute("DELETE FROM announcements WHERE id = %s", (announcement_id,))
        conn.commit()
        cur.close()
        conn.close()
        
        logger.info(f"✅ Öğretmen {current_user.username} duyuruyu sildi: {announcement_id}")
        return jsonify({"success": True, "message": "Duyuru silindi"})
        
    except Exception as e:
        logger.error(f"Teacher announcement deletion error: {str(e)}")
        return jsonify({"error": f"Duyuru silinemedi: {str(e)}"}), 500

# Öğrenci - Öğretmen listesi
@app.route("/api/teachers", methods=["GET"])
@login_required
def get_teachers_list():
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("""
        SELECT id, full_name FROM users 
        WHERE role = 'teacher' 
        ORDER BY full_name
    """)
    teachers = cur.fetchall()
    cur.close()
    conn.close()
    
    response = jsonify({
        "teachers": [{
            "id": t['id'],
            "full_name": t['full_name']
        } for t in teachers]
    })
    
    # Cache'i engelle - her zaman güncel veri gelsin
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    
    return response

# Öğrenci - Sınav cevap gönder
@app.route("/student/exam/<exam_id>/submit", methods=["POST"])
@login_required
def submit_exam_answers(exam_id):
    if current_user.role != 'student':
        return jsonify({"error": "Sadece öğrenciler sınav gönderebilir"}), 403
    
    try:
        answers = request.get_json().get("answers", {})
        
        if not answers:
            return jsonify({"error": "Cevaplar boş olamaz"}), 400
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Sınav kontrolü
        cur.execute("SELECT * FROM exams WHERE id = %s", (exam_id,))
        exam = cur.fetchone()
        
        if not exam:
            cur.close()
            conn.close()
            return jsonify({"error": "Sınav bulunamadı"}), 404
        
        # Daha önce gönderilmiş mi veya draft var mı kontrol et
        cur.execute("SELECT id, status FROM exam_submissions WHERE exam_id = %s AND student_id = %s", 
                   (exam_id, current_user.id))
        existing = cur.fetchone()
        
        if existing and existing.get('status') == 'submitted':
            cur.close()
            conn.close()
            return jsonify({"error": "Bu sınava zaten cevap gönderdiniz"}), 400
        
        # Cevapları puanla
        answer_key = exam['answer_key']  # JSONB dict: {"1":"A", "2":"B", ...}
        correct_count = 0
        
        # answers: {"1": "A", "2": "B", ...}
        for question_num, student_answer in answers.items():
            correct_answer = answer_key.get(str(question_num))
            if student_answer == correct_answer:
                correct_count += 1
        
        # Puan hesapla (100 üzerinden)
        total_questions = exam['question_count']
        score = (correct_count / total_questions) * 100 if total_questions > 0 else 0
        
        if existing:
            # Draft varsa güncelle ve submitted yap
            cur.execute(
                """UPDATE exam_submissions 
                   SET answers = %s, score = %s, status = 'submitted', submitted_at = NOW()
                   WHERE exam_id = %s AND student_id = %s""",
                (json.dumps(answers), score, exam_id, current_user.id)
            )
        else:
            # Yeni kayıt oluştur
            cur.execute(
                """INSERT INTO exam_submissions (exam_id, student_id, answers, score, status, submitted_at)
                   VALUES (%s, %s, %s, %s, 'submitted', NOW())""",
                (exam_id, current_user.id, json.dumps(answers), score)
            )
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({
            "success": True, 
            "message": "Sınav cevapları gönderildi",
            "score": round(score, 2),
            "correct_count": correct_count,
            "total_questions": total_questions
        }), 200
    except Exception as e:
        logger.error(f"Exam submission error: {str(e)}")
        return jsonify({"error": f"Cevaplar gönderilemedi: {str(e)}"}), 500

# Öğrenci - Ödev teslim et
@app.route("/student/assignments/<int:assignment_id>/submit", methods=["POST"])
@login_required
def submit_assignment(assignment_id):
    if current_user.role != 'student':
        return jsonify({"error": "Sadece öğrenciler ödev teslim edebilir"}), 403
    
    try:
        # Dosya opsiyonel - varsa kaydet (Object Storage veya lokal)
        filename = None
        if "file" in request.files:
            file = request.files["file"]
            if file.filename:
                # Dosya seçilmişse, desteklenen formatta olmalı
                if not allowed_file(file.filename):
                    return jsonify({"error": "Desteklenmeyen dosya formatı. İzin verilen: PDF, JPG, PNG"}), 400
                
                base_filename = f"submission_{uuid.uuid4()}_{secure_filename(file.filename)}"
                
                # Object Storage kullan (varsa)
                if object_storage.enabled:
                    try:
                        object_path = f"submissions/{base_filename}"
                        object_storage.upload_from_file(file, object_path)
                        filename = object_path  # Object path kaydet
                        logger.info(f"✅ Ödev teslimi Object Storage'a yüklendi: {object_path}")
                    except Exception as storage_error:
                        logger.error(f"⚠️ Object Storage hatası, lokal sisteme geçiliyor: {storage_error}")
                        file.seek(0)  # Stream'i başa sar (CRITICAL!)
                        file_path = os.path.join(UPLOAD_DIR, base_filename)
                        file.save(file_path)
                        filename = base_filename
                else:
                    # Object Storage devre dışı - lokal sistemi kullan
                    file_path = os.path.join(UPLOAD_DIR, base_filename)
                    file.save(file_path)
                    filename = base_filename
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Ödev kontrolü
        cur.execute("SELECT * FROM assignments WHERE id = %s", (assignment_id,))
        assignment = cur.fetchone()
        
        if not assignment:
            cur.close()
            conn.close()
            return jsonify({"error": "Ödev bulunamadı"}), 404
        
        # Daha önce teslim edilmiş mi kontrol et
        cur.execute("SELECT id FROM assignment_submissions WHERE assignment_id = %s AND student_id = %s", 
                   (assignment_id, current_user.id))
        existing = cur.fetchone()
        
        if existing:
            # Eski dosyayı sil, yeni dosyayı güncelle
            cur.execute("UPDATE assignment_submissions SET file_path = %s, submitted_at = NOW() WHERE id = %s",
                       (filename, existing['id']))
        else:
            # Yeni kayıt oluştur
            cur.execute(
                """INSERT INTO assignment_submissions (assignment_id, student_id, file_path, submitted_at)
                   VALUES (%s, %s, %s, NOW())""",
                (assignment_id, current_user.id, filename)
            )
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Ödev teslim edildi"}), 200
    except Exception as e:
        logger.error(f"Assignment submission error: {str(e)}")
        return jsonify({"error": f"Ödev teslim edilemedi: {str(e)}"}), 500

# Öğrenci - Sınavlarımı listele
@app.route("/student/my-exams", methods=["GET"])
@login_required
def student_get_my_exams():
    """Öğrenci sınavlarını listele - basitleştirilmiş & hata yakalama ile"""
    if current_user.role not in ['student', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Önce exam_submissions tablosunun var olup olmadığını kontrol et
        cur.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables 
                WHERE table_name = 'exam_submissions'
            ) as table_exists
        """)
        table_check = cur.fetchone()
        
        if table_check and table_check['table_exists']:
            # Tablo varsa is_submitted, allow_pause ve submission_status kontrolü yap
            cur.execute("""
                SELECT e.id, e.title, e.question_count, e.start_time, e.end_time, e.duration_minutes, 
                       e.pdf_filename, e.target_class, e.created_at, 
                       COALESCE(e.allow_pause, FALSE) as allow_pause,
                       EXISTS(SELECT 1 FROM exam_submissions WHERE exam_id = e.id AND student_id = %s AND status = 'submitted') as is_submitted,
                       (SELECT status FROM exam_submissions WHERE exam_id = e.id AND student_id = %s LIMIT 1) as submission_status,
                       CASE 
                           WHEN NOW() < e.start_time THEN 'not_started'
                           WHEN e.end_time IS NOT NULL AND NOW() > e.end_time THEN 'ended'
                           ELSE 'active'
                       END as exam_status
                FROM exams e
                WHERE e.target_class = %s
                ORDER BY e.start_time DESC
            """, (current_user.id, current_user.id, current_user.class_name))
        else:
            # Tablo yoksa basit sorgu
            cur.execute("""
                SELECT e.id, e.title, e.question_count, e.start_time, e.end_time, e.duration_minutes, 
                       e.pdf_filename, e.target_class, e.created_at,
                       FALSE as is_submitted, FALSE as allow_pause, NULL as submission_status,
                       CASE 
                           WHEN NOW() < e.start_time THEN 'not_started'
                           WHEN e.end_time IS NOT NULL AND NOW() > e.end_time THEN 'ended'
                           ELSE 'active'
                       END as exam_status
                FROM exams e
                WHERE e.target_class = %s
                ORDER BY e.start_time DESC
            """, (current_user.class_name,))
        
        exams = cur.fetchall()
        cur.close()
        conn.close()
        
        # JSON serialize edilebilir formata çevir
        exams_list = []
        for exam in exams:
            exam_dict = dict(exam)
            # Datetime nesnelerini string'e çevir
            if exam_dict.get('start_time'):
                exam_dict['start_time'] = exam_dict['start_time'].isoformat()
            if exam_dict.get('end_time'):
                exam_dict['end_time'] = exam_dict['end_time'].isoformat()
            if exam_dict.get('created_at'):
                exam_dict['created_at'] = exam_dict['created_at'].isoformat()
            
            # pdf_filename'i parse et - tek dosya veya çoklu dosya olabilir
            if exam_dict.get('pdf_filename'):
                try:
                    # JSON array olarak parse et
                    exam_dict['pdf_files'] = json.loads(exam_dict['pdf_filename'])
                except:
                    # Eski format - tek dosya string
                    exam_dict['pdf_files'] = [exam_dict['pdf_filename']]
            else:
                exam_dict['pdf_files'] = []
            
            exams_list.append(exam_dict)
        
        return jsonify({"exams": exams_list})
        
    except Exception as e:
        logger.error(f"Student exams error: {str(e)}")
        return jsonify({"error": f"Sınavlar yüklenemedi: {str(e)}"}), 500

# Öğrenci - Sınava başla (sınav sayfası)
@app.route("/student/exam/<exam_id>/take", methods=["GET"])
@login_required
def student_take_exam(exam_id):
    """Öğrenci sınav sayfası - Optik form ile"""
    if current_user.role != 'student':
        return redirect('/login')
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Sınavı getir
        cur.execute("""
            SELECT * FROM exams WHERE id = %s AND target_class = %s
        """, (exam_id, current_user.class_name))
        exam = cur.fetchone()
        
        if not exam:
            cur.close()
            conn.close()
            return "Sınav bulunamadı veya bu sınava katılma yetkiniz yok", 404
        
        # Zaten teslim edilmiş mi kontrol et (status='submitted' olanlar kesin teslim)
        cur.execute("""
            SELECT id, status, answers, started_at FROM exam_submissions WHERE exam_id = %s AND student_id = %s
        """, (exam_id, current_user.id))
        submission = cur.fetchone()
        
        draft_answers = None
        student_started_at = None
        
        if submission:
            if submission.get('status') == 'submitted':
                cur.close()
                conn.close()
                return "Bu sınavı zaten teslim ettiniz!", 403
            elif submission.get('status') == 'draft':
                # Draft var, cevapları yükle
                draft_answers = submission.get('answers', {})
                student_started_at = submission.get('started_at')
        
        # Sınav zamanı kontrolü - Başlangıç ve bitiş tarihi kontrolü
        cur.execute("""
            SELECT 
                CASE 
                    WHEN NOW() < start_time THEN 'not_started'
                    WHEN end_time IS NOT NULL AND NOW() > end_time THEN 'ended'
                    ELSE 'active'
                END as exam_status,
                start_time,
                end_time
            FROM exams WHERE id = %s
        """, (exam_id,))
        status_check = cur.fetchone()
        
        if status_check['exam_status'] == 'not_started':
            cur.close()
            conn.close()
            start_time_str = status_check['start_time'].strftime('%d.%m.%Y %H:%M') if status_check['start_time'] else ''
            return f"Sınav henüz başlamadı! Başlangıç: {start_time_str}", 403
        
        if status_check['exam_status'] == 'ended':
            cur.close()
            conn.close()
            end_time_str = status_check['end_time'].strftime('%d.%m.%Y %H:%M') if status_check['end_time'] else ''
            return f"Sınav süresi dolmuş! Bitiş: {end_time_str}", 403
        
        # Öğrenci sınava ilk kez giriyorsa started_at kaydet
        if not student_started_at:
            if submission:
                # Mevcut draft'ı güncelle
                cur.execute("""
                    UPDATE exam_submissions SET started_at = NOW() WHERE exam_id = %s AND student_id = %s
                """, (exam_id, current_user.id))
            else:
                # Yeni submission oluştur (score=0 başlangıç değeri)
                cur.execute("""
                    INSERT INTO exam_submissions (exam_id, student_id, answers, status, started_at, score)
                    VALUES (%s, %s, '{}', 'draft', NOW(), 0)
                """, (exam_id, current_user.id))
            conn.commit()
            
            # started_at'ı yeniden oku
            cur.execute("""
                SELECT started_at FROM exam_submissions WHERE exam_id = %s AND student_id = %s
            """, (exam_id, current_user.id))
            result = cur.fetchone()
            student_started_at = result['started_at'] if result else None
        
        cur.close()
        conn.close()
        
        # Kalan süreyi hesapla (öğrencinin başlama zamanından itibaren)
        from datetime import datetime, timezone
        now = datetime.now(timezone.utc)
        if student_started_at:
            elapsed_seconds = (now - student_started_at).total_seconds()
            remaining_seconds = max(0, (exam['duration_minutes'] * 60) - elapsed_seconds)
        else:
            remaining_seconds = exam['duration_minutes'] * 60
        
        # pdf_filename'i parse et - çoklu dosya desteği
        pdf_files = []
        if exam['pdf_filename']:
            try:
                pdf_files = json.loads(exam['pdf_filename'])
            except:
                pdf_files = [exam['pdf_filename']]
        
        return render_template('student_take_exam.html', exam=exam, pdf_files=pdf_files, draft_answers=draft_answers, remaining_seconds=int(remaining_seconds))
        
    except Exception as e:
        logger.error(f"Take exam error: {str(e)}")
        return f"Hata: {str(e)}", 500

# Öğrenci - Sınav cevaplarını taslak olarak kaydet
@app.route("/student/exam/<exam_id>/save-draft", methods=["POST"])
@login_required
def save_exam_draft(exam_id):
    """Öğrenci sınav cevaplarını taslak olarak kaydeder - Ara Vererek Sınav Çözme"""
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        data = request.get_json()
        answers = data.get('answers', {})
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Sınavın ara verilebilir olup olmadığını kontrol et
        cur.execute("""
            SELECT id, allow_pause, target_class, start_time, duration_minutes 
            FROM exams WHERE id = %s
        """, (exam_id,))
        exam = cur.fetchone()
        
        if not exam:
            cur.close()
            conn.close()
            return jsonify({"error": "Sınav bulunamadı"}), 404
        
        if exam['target_class'] != current_user.class_name:
            cur.close()
            conn.close()
            return jsonify({"error": "Bu sınava erişim yetkiniz yok"}), 403
        
        if not exam.get('allow_pause'):
            cur.close()
            conn.close()
            return jsonify({"error": "Bu sınav ara verilebilir değil"}), 403
        
        # Sınav süresi kontrolü
        cur.execute("""
            SELECT 
                CASE 
                    WHEN NOW() < start_time THEN 'not_started'
                    WHEN NOW() > start_time + (duration_minutes || ' minutes')::interval THEN 'ended'
                    ELSE 'active'
                END as exam_status
            FROM exams WHERE id = %s
        """, (exam_id,))
        status_check = cur.fetchone()
        
        if status_check['exam_status'] != 'active':
            cur.close()
            conn.close()
            return jsonify({"error": "Sınav süresi aktif değil"}), 403
        
        # Mevcut draft var mı kontrol et
        cur.execute("""
            SELECT id, status FROM exam_submissions WHERE exam_id = %s AND student_id = %s
        """, (exam_id, current_user.id))
        existing = cur.fetchone()
        
        if existing:
            if existing.get('status') == 'submitted':
                cur.close()
                conn.close()
                return jsonify({"error": "Bu sınavı zaten teslim ettiniz"}), 403
            
            # Draft güncelle
            cur.execute("""
                UPDATE exam_submissions 
                SET answers = %s, status = 'draft', submitted_at = NOW()
                WHERE exam_id = %s AND student_id = %s
            """, (json.dumps(answers), exam_id, current_user.id))
        else:
            # Yeni draft oluştur (score=0 geçici)
            cur.execute("""
                INSERT INTO exam_submissions (exam_id, student_id, answers, score, status)
                VALUES (%s, %s, %s, 0, 'draft')
            """, (exam_id, current_user.id, json.dumps(answers)))
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Cevaplarınız kaydedildi"})
        
    except Exception as e:
        logger.error(f"Save draft error: {str(e)}")
        return jsonify({"error": f"Kaydetme hatası: {str(e)}"}), 500

# Öğrenci - Sınavı görüntüle (sadece görüntüleme, cevap veremez)
@app.route("/student/exam/<exam_id>/view", methods=["GET"])
@login_required
def student_view_exam(exam_id):
    """Öğrenci geçmiş sınavları sadece görüntüler (read-only)"""
    if current_user.role != 'student':
        return redirect('/login')
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Sınavı getir
        cur.execute("""
            SELECT * FROM exams WHERE id = %s AND target_class = %s
        """, (exam_id, current_user.class_name))
        exam = cur.fetchone()
        
        if not exam:
            cur.close()
            conn.close()
            return "Sınav bulunamadı veya bu sınava erişim yetkiniz yok", 404
        
        cur.close()
        conn.close()
        
        # pdf_filename'i parse et - çoklu dosya desteği
        pdf_files = []
        if exam['pdf_filename']:
            try:
                pdf_files = json.loads(exam['pdf_filename'])
            except:
                pdf_files = [exam['pdf_filename']]
        
        return render_template('student_view_exam.html', exam=exam, pdf_files=pdf_files)
        
    except Exception as e:
        logger.error(f"View exam error: {str(e)}")
        return f"Hata: {str(e)}", 500

# Öğrenci - Sınavı teslim et
@app.route("/student/exam/<exam_id>/submit", methods=["POST"])
@login_required
def student_submit_exam(exam_id):
    """Öğrenci sınav teslimi - puanlama ile"""
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        data = request.json
        student_answers = data.get('answers', {})
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Sınavı getir
        cur.execute("""
            SELECT * FROM exams WHERE id = %s AND target_class = %s
        """, (exam_id, current_user.class_name))
        exam = cur.fetchone()
        
        if not exam:
            cur.close()
            conn.close()
            return jsonify({"error": "Sınav bulunamadı"}), 404
        
        # Zaten teslim edilmiş mi?
        cur.execute("""
            SELECT id FROM exam_submissions WHERE exam_id = %s AND student_id = %s
        """, (exam_id, current_user.id))
        if cur.fetchone():
            cur.close()
            conn.close()
            return jsonify({"error": "Bu sınavı zaten teslim ettiniz"}), 403
        
        # Cevap anahtarını al ve puanla
        answer_key = exam['answer_key'] or {}
        correct_count = 0
        
        for q_num, correct_answer in answer_key.items():
            student_answer = student_answers.get(str(q_num))
            if student_answer and student_answer.upper() == correct_answer.upper():
                correct_count += 1
        
        # Puan hesapla (100 üzerinden)
        total_questions = exam['question_count']
        score = round((correct_count / total_questions) * 100, 2) if total_questions > 0 else 0
        
        # Teslimi kaydet
        cur.execute("""
            INSERT INTO exam_submissions (exam_id, student_id, answers, score, submitted_at)
            VALUES (%s, %s, %s, %s, NOW())
        """, (exam_id, current_user.id, json.dumps(student_answers), score))
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "score": score,
            "correct_count": correct_count,
            "total_questions": total_questions
        })
        
    except Exception as e:
        logger.error(f"Submit exam error: {str(e)}")
        return jsonify({"error": f"Teslim edilemedi: {str(e)}"}), 500

# Öğrenci - Sınav sonucunu görüntüle
@app.route("/student/exam/<exam_id>/result", methods=["GET"])
@login_required
def student_get_exam_result(exam_id):
    """Öğrenci kendi sınav sonucunu görüntüler"""
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Sınav bilgilerini al
        cur.execute("""
            SELECT id, title, question_count, answer_key
            FROM exams WHERE id = %s AND target_class = %s
        """, (exam_id, current_user.class_name))
        exam = cur.fetchone()
        
        if not exam:
            cur.close()
            conn.close()
            return jsonify({"error": "Sınav bulunamadı"}), 404
        
        # Öğrencinin cevaplarını al
        cur.execute("""
            SELECT answers, score, submitted_at
            FROM exam_submissions
            WHERE exam_id = %s AND student_id = %s
        """, (exam_id, current_user.id))
        submission = cur.fetchone()
        
        cur.close()
        conn.close()
        
        if not submission:
            return jsonify({"error": "Bu sınavı henüz tamamlamadınız"}), 404
        
        # Soru detaylarını hazırla
        student_answers = submission['answers']
        answer_key = exam['answer_key']
        question_details = []
        correct_count = 0
        wrong_count = 0
        empty_count = 0
        
        for q_num in range(1, exam['question_count'] + 1):
            q_str = str(q_num)
            student_answer = student_answers.get(q_str, "")
            correct_answer = answer_key.get(q_str, "")
            
            if not student_answer:
                status = "empty"
                empty_count += 1
            elif student_answer == correct_answer:
                status = "correct"
                correct_count += 1
            else:
                status = "wrong"
                wrong_count += 1
            
            question_details.append({
                "question_number": q_num,
                "student_answer": student_answer or "-",
                "correct_answer": correct_answer,
                "status": status
            })
        
        return jsonify({
            "exam": {
                "title": exam['title'],
                "question_count": exam['question_count']
            },
            "result": {
                "score": float(submission['score']),
                "correct_count": correct_count,
                "wrong_count": wrong_count,
                "empty_count": empty_count,
                "submitted_at": submission['submitted_at'].isoformat()
            },
            "question_details": question_details
        })
        
    except Exception as e:
        logger.error(f"Get exam result error: {str(e)}")
        return jsonify({"error": f"Sonuç alınamadı: {str(e)}"}), 500

# Öğrenci - Ödevlerimi listele
@app.route("/student/my-assignments", methods=["GET"])
@login_required
def student_get_my_assignments():
    if current_user.role not in ['student', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    cur.execute("""
        SELECT a.id, a.title, a.description, a.due_date, a.target_class, a.file_path,
               u.full_name as teacher_name,
               EXISTS(SELECT 1 FROM assignment_submissions WHERE assignment_id = a.id AND student_id = %s) as is_submitted
        FROM assignments a
        LEFT JOIN users u ON a.teacher_id = u.id
        WHERE a.target_class = %s
        ORDER BY a.due_date DESC
    """, (current_user.id, current_user.class_name))
    
    assignments = cur.fetchall()
    cur.close()
    conn.close()
    
    result = []
    for a in assignments:
        result.append({
            "id": a['id'],
            "title": a['title'],
            "description": a['description'],
            "due_date": a['due_date'].isoformat() if a['due_date'] else None,
            "target_class": a['target_class'],
            "file_path": a['file_path'],
            "teacher_name": a['teacher_name'] or 'Öğretmen',
            "is_submitted": a['is_submitted']
        })
    
    return jsonify({"assignments": result})

# Öğrenci - Duyurularımı listele
@app.route("/student/my-announcements", methods=["GET"])
@login_required
def student_get_my_announcements():
    if current_user.role not in ['student', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    cur.execute("""
        SELECT an.*, u.full_name as teacher_name,
               EXISTS(
                   SELECT 1 FROM announcement_reads ar 
                   WHERE ar.announcement_id = an.id AND ar.user_id = %s
               ) as has_read,
               (SELECT read_at FROM announcement_reads ar 
                WHERE ar.announcement_id = an.id AND ar.user_id = %s) as read_at
        FROM announcements an
        JOIN users u ON an.teacher_id = u.id
        WHERE an.target_class = %s
        ORDER BY an.created_at DESC
    """, (current_user.id, current_user.id, current_user.class_name))
    
    announcements = cur.fetchall()
    cur.close()
    conn.close()
    
    return jsonify({"announcements": announcements})

# Duyuru okuma işaretleme (öğrenci ve öğretmen için)
@app.route("/api/announcements/<int:announcement_id>/mark-read", methods=["POST"])
@app.route("/student/announcements/<int:announcement_id>/read", methods=["POST"])
@login_required
def mark_announcement_read(announcement_id):
    if current_user.role not in ['student', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # Duyuru var mı kontrol et
        cur.execute("SELECT id FROM announcements WHERE id = %s", (announcement_id,))
        if not cur.fetchone():
            cur.close()
            conn.close()
            return jsonify({"error": "Duyuru bulunamadı"}), 404
        
        # Okuma kaydını ekle (varsa ignore)
        cur.execute("""
            INSERT INTO announcement_reads (announcement_id, user_id)
            VALUES (%s, %s)
            ON CONFLICT (announcement_id, user_id) DO NOTHING
        """, (announcement_id, current_user.id))
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Duyuru okundu olarak işaretlendi"})
    except Exception as e:
        logger.error(f"Mark announcement read error: {str(e)}")
        return jsonify({"error": f"İşaretleme başarısız: {str(e)}"}), 500

# Duyuru takibi - Kimler okudu? (Öğretmen ve Admin için)
@app.route("/api/announcements/<int:announcement_id>/readers", methods=["GET"])
@login_required
def get_announcement_readers(announcement_id):
    if current_user.role not in ['teacher', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Duyuruyu kontrol et ve yetki doğrula
        cur.execute("SELECT * FROM announcements WHERE id = %s", (announcement_id,))
        announcement = cur.fetchone()
        
        if not announcement:
            cur.close()
            conn.close()
            return jsonify({"error": "Duyuru bulunamadı"}), 404
        
        # Öğretmen sadece kendi duyurusunu görebilir
        if current_user.role == 'teacher' and announcement['teacher_id'] != current_user.id:
            cur.close()
            conn.close()
            return jsonify({"error": "Bu duyurunun takibini görme yetkiniz yok"}), 403
        
        # Hedef sınıftaki tüm öğrencileri al
        cur.execute("""
            SELECT u.id, u.username, u.full_name,
                   EXISTS(
                       SELECT 1 FROM announcement_reads ar 
                       WHERE ar.announcement_id = %s AND ar.user_id = u.id
                   ) as has_read,
                   (SELECT read_at FROM announcement_reads ar 
                    WHERE ar.announcement_id = %s AND ar.user_id = u.id) as read_at
            FROM users u
            WHERE u.class_name = %s AND u.role = 'student'
            ORDER BY u.full_name
        """, (announcement_id, announcement_id, announcement['target_class']))
        
        students = cur.fetchall()
        cur.close()
        conn.close()
        
        # İstatistik hesapla
        total_students = len(students)
        read_count = sum(1 for s in students if s['has_read'])
        unread_count = total_students - read_count
        
        return jsonify({
            "success": True,
            "announcement": {
                "id": announcement['id'],
                "title": announcement['title'],
                "target_class": announcement['target_class']
            },
            "statistics": {
                "total": total_students,
                "read": read_count,
                "unread": unread_count,
                "percentage": round((read_count / total_students * 100) if total_students > 0 else 0, 1)
            },
            "students": [{
                "id": s['id'],
                "username": s['username'],
                "full_name": s['full_name'],
                "has_read": s['has_read'],
                "read_at": s['read_at'].isoformat() if s['read_at'] else None
            } for s in students]
        })
    except Exception as e:
        logger.error(f"Get announcement readers error: {str(e)}")
        return jsonify({"error": f"Takip bilgisi alınamadı: {str(e)}"}), 500

# Ödev takibi - Kimler teslim etti? (Öğretmen ve Admin için)
@app.route("/api/assignments/<int:assignment_id>/submissions", methods=["GET"])
@login_required
def get_assignment_submissions_report(assignment_id):
    """Ödev teslim raporu - hangi öğrenciler teslim etti/etmedi"""
    if current_user.role not in ['teacher', 'admin']:
        return jsonify({"success": False, "error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Ödevi kontrol et ve yetki doğrula
        cur.execute("SELECT * FROM assignments WHERE id = %s", (assignment_id,))
        assignment = cur.fetchone()
        
        if not assignment:
            cur.close()
            conn.close()
            return jsonify({"success": False, "error": "Ödev bulunamadı"}), 404
        
        # Öğretmen sadece kendi ödevini görebilir (admin tümünü görebilir)
        if current_user.role == 'teacher' and assignment['teacher_id'] and assignment['teacher_id'] != current_user.id:
            cur.close()
            conn.close()
            return jsonify({"success": False, "error": "Bu ödevin takibini görme yetkiniz yok"}), 403
        
        # Hedef sınıftaki tüm öğrencileri al
        cur.execute("""
            SELECT u.id, u.username, u.full_name,
                   EXISTS(
                       SELECT 1 FROM assignment_submissions asub 
                       WHERE asub.assignment_id = %s AND asub.student_id = u.id
                   ) as has_submitted,
                   (SELECT submitted_at FROM assignment_submissions asub 
                    WHERE asub.assignment_id = %s AND asub.student_id = u.id) as submitted_at,
                   (SELECT file_path FROM assignment_submissions asub 
                    WHERE asub.assignment_id = %s AND asub.student_id = u.id) as file_path
            FROM users u
            WHERE u.class_name = %s AND u.role = 'student'
            ORDER BY u.full_name
        """, (assignment_id, assignment_id, assignment_id, assignment['target_class']))
        
        students = cur.fetchall()
        cur.close()
        conn.close()
        
        # İstatistik hesapla
        total_students = len(students)
        submitted_count = sum(1 for s in students if s['has_submitted'])
        not_submitted_count = total_students - submitted_count
        
        return jsonify({
            "success": True,
            "assignment": {
                "id": assignment['id'],
                "title": assignment['title'],
                "target_class": assignment['target_class'],
                "due_date": assignment['due_date'].isoformat()
            },
            "statistics": {
                "total": total_students,
                "submitted": submitted_count,
                "not_submitted": not_submitted_count,
                "percentage": round((submitted_count / total_students * 100) if total_students > 0 else 0, 1)
            },
            "students": [{
                "id": s['id'],
                "username": s['username'],
                "full_name": s['full_name'],
                "has_submitted": s['has_submitted'],
                "submitted_at": s['submitted_at'].isoformat() if s['submitted_at'] else None,
                "file_path": s['file_path']
            } for s in students]
        })
    except Exception as e:
        logger.error(f"Get assignment submissions report error: {str(e)}")
        return jsonify({"success": False, "error": f"Takip bilgisi alınamadı: {str(e)}"}), 500

# Öğrenci - Öğretmene soru sor
@app.route("/student/ask-question", methods=["POST"])
@login_required
def ask_question():
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        # FormData ile gönderilen verileri al
        teacher_id = request.form.get('teacher_id')
        question = request.form.get('question')
        
        if not teacher_id:
            return jsonify({"error": "Öğretmen seçiniz"}), 400
        
        if not question or not question.strip():
            return jsonify({"error": "Soru metni boş olamaz"}), 400
        
        # Dosya yükleme (opsiyonel, Object Storage veya lokal)
        file_path = None
        if 'file' in request.files:
            file = request.files['file']
            if file and file.filename:
                if not allowed_file(file.filename):
                    return jsonify({"error": "Desteklenmeyen dosya formatı. İzin verilen: PDF, Word, Excel, Resim"}), 400
                
                # Güvenli dosya adı oluştur
                ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'bin'
                base_filename = f"question_{uuid.uuid4()}.{ext}"
                
                # Object Storage kullan (varsa)
                if object_storage.enabled:
                    try:
                        object_path = f"questions/{base_filename}"
                        object_storage.upload_from_file(file, object_path)
                        file_path = object_path  # Object path kaydet
                        logger.info(f"✅ Soru dosyası Object Storage'a yüklendi: {object_path}")
                    except Exception as storage_error:
                        logger.error(f"⚠️ Object Storage hatası, lokal sisteme geçiliyor: {storage_error}")
                        file.seek(0)  # Stream'i başa sar (CRITICAL!)
                        full_path = os.path.join(UPLOAD_DIR, base_filename)
                        file.save(full_path)
                        file_path = base_filename
                else:
                    # Object Storage devre dışı - lokal sistemi kullan
                    full_path = os.path.join(UPLOAD_DIR, base_filename)
                    file.save(full_path)
                    file_path = base_filename
        
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO student_questions (student_id, teacher_id, question_text, file_path)
            VALUES (%s, %s, %s, %s)
        """, (current_user.id, teacher_id, question, file_path))
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Sorunuz öğretmene iletildi"})
    
    except Exception as e:
        logger.error(f"Ask question error: {str(e)}")
        return jsonify({"error": f"Soru gönderilemedi: {str(e)}"}), 500

# Öğrenci - Kendi sorularını görüntüle
@app.route("/student/my-questions", methods=["GET"])
@login_required
def get_my_questions():
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    cur.execute("""
        SELECT sq.id, sq.question_text, sq.file_path, sq.answer_text, 
               sq.answered_at, sq.created_at, u.full_name as teacher_name
        FROM student_questions sq
        JOIN users u ON sq.teacher_id = u.id
        WHERE sq.student_id = %s
        ORDER BY sq.created_at DESC
    """, (current_user.id,))
    questions = cur.fetchall()
    cur.close()
    conn.close()
    
    return jsonify({"questions": [{
        "id": q['id'],
        "question": q['question_text'],
        "answer": q['answer_text'],
        "file_path": q['file_path'],
        "answered_at": q['answered_at'].isoformat() if q['answered_at'] else None,
        "created_at": q['created_at'].isoformat(),
        "teacher_name": q['teacher_name']
    } for q in questions]})

# Öğretmen - Gelen sorularım sayısı (dashboard için)
@app.route("/teacher/my-questions", methods=["GET"])
@login_required
def get_teacher_questions_count():
    """Öğretmen dashboard için soru sayısı"""
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT COUNT(*) as total,
                   COUNT(*) FILTER (WHERE answer_text IS NULL) as unanswered
            FROM student_questions
            WHERE teacher_id = %s
        """, (current_user.id,))
        result = cur.fetchone()
        
        cur.close()
        conn.close()
        
        return jsonify({
            "total": result['total'] or 0,
            "unanswered": result['unanswered'] or 0
        })
    except Exception as e:
        logger.error(f"Get teacher questions count error: {str(e)}")
        return jsonify({"error": f"Soru sayısı alınamadı: {str(e)}"}), 500

# Öğretmen - Gelen soruları görüntüle
@app.route("/teacher/api/questions", methods=["GET"])
@login_required
def get_teacher_questions():
    """Öğretmen - Öğrenci sorularını listele (API endpoint)"""
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("""
            SELECT sq.id, sq.question_text, sq.file_path, sq.answer_text, 
                   sq.answered_at, sq.created_at, u.full_name as student_name
            FROM student_questions sq
            JOIN users u ON sq.student_id = u.id
            WHERE sq.teacher_id = %s
            ORDER BY sq.answered_at IS NULL DESC, sq.created_at DESC
        """, (current_user.id,))
        questions = cur.fetchall()
        cur.close()
        conn.close()
        
        return jsonify({"questions": [{
            "id": q['id'],
            "question_text": q['question_text'],
            "file_path": q['file_path'],
            "answer_text": q['answer_text'],
            "answered_at": q['answered_at'].isoformat() if q['answered_at'] else None,
            "created_at": q['created_at'].isoformat(),
            "student_name": q['student_name']
        } for q in questions]})
    except Exception as e:
        logger.error(f"Get teacher questions error: {str(e)}")
        return jsonify({"error": f"Sorular yüklenemedi: {str(e)}"}), 500

# Öğretmen - Soruya cevap ver
@app.route("/teacher/questions/<int:question_id>/answer", methods=["POST"])
@login_required
def answer_question(question_id):
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    data = request.get_json()
    answer_text = data.get('answer_text')
    
    if not answer_text:
        return jsonify({"error": "Cevap metni gereklidir"}), 400
    
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        UPDATE student_questions 
        SET answer_text = %s, answered_at = CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul'
        WHERE id = %s AND teacher_id = %s
    """, (answer_text, question_id, current_user.id))
    conn.commit()
    cur.close()
    conn.close()
    
    return jsonify({"success": True, "message": "Cevap gönderildi"})

# Öğretmen - Soruyu sil
@app.route("/teacher/questions/<int:question_id>", methods=["DELETE"])
@login_required
def delete_question(question_id):
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # Sorunun bu öğretmene ait olduğunu kontrol et
        cur.execute("SELECT teacher_id FROM student_questions WHERE id = %s", (question_id,))
        question = cur.fetchone()
        
        if not question:
            cur.close()
            conn.close()
            return jsonify({"error": "Soru bulunamadı"}), 404
        
        if question[0] != current_user.id:
            cur.close()
            conn.close()
            return jsonify({"error": "Bu soruyu silme yetkiniz yok"}), 403
        
        # Soruyu sil
        cur.execute("DELETE FROM student_questions WHERE id = %s", (question_id,))
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Soru silindi"})
    except Exception as e:
        logger.error(f"Delete question error: {str(e)}")
        return jsonify({"error": f"Soru silinemedi: {str(e)}"}), 500

# Admin - Öğretmenlere duyuru tablosu
@app.route("/api/admin/teacher-announcement", methods=["POST"])
@login_required
def send_teacher_announcement():
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    data = request.get_json()
    title = data.get('title')
    content = data.get('content')
    
    if not title or not content:
        return jsonify({"error": "Başlık ve içerik gereklidir"}), 400
    
    # Admin'den gelen duyuruları saklamak için basit bir tablo oluşturabiliriz
    # Veya öğretmenlere e-posta/bildirim gönderebiliriz
    # Şimdilik basit olarak kaydedeceğiz
    
    conn = get_db()
    cur = conn.cursor()
    
    # Teacher announcements tablosu oluştur (ilk kez çağrıldığında)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS teacher_announcements (
            id SERIAL PRIMARY KEY,
            title VARCHAR(200) NOT NULL,
            content TEXT,
            admin_id INTEGER REFERENCES users(id) ON DELETE CASCADE,
            created_at TIMESTAMP DEFAULT (CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul')
        )
    """)
    
    cur.execute(
        """INSERT INTO teacher_announcements (title, content, admin_id)
           VALUES (%s, %s, %s)""",
        (title, content, current_user.id)
    )
    
    conn.commit()
    cur.close()
    conn.close()
    
    return jsonify({"success": True, "message": "Duyuru öğretmenlere gönderildi"})

# Bildirim endpoint'i
@app.route("/api/notifications", methods=["GET"])
@login_required
def get_notifications():
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    notifications = []
    
    # DB'den persistent notifications oku (okunmamışlar)
    cur.execute("""
        SELECT id, message, type, created_at
        FROM notifications
        WHERE user_id = %s AND is_read = FALSE
        ORDER BY created_at DESC
        LIMIT 20
    """, (current_user.id,))
    db_notifications = cur.fetchall()
    for notif in db_notifications:
        notifications.append({
            "id": notif['id'],
            "message": notif['message'],
            "type": notif['type'],
            "created_at": notif['created_at'].isoformat() if notif['created_at'] else None
        })
    
    # Dynamic notifications (backward compatible)
    if current_user.role == 'student':
        cur.execute("SELECT class_name FROM users WHERE id = %s", (current_user.id,))
        user_data = cur.fetchone()
        student_class = user_data['class_name'] if user_data else None
        
        if student_class:
            class_level = student_class[0] if student_class else None
            
            cur.execute("""
                SELECT COUNT(*) as count FROM exams 
                WHERE (target_class = %s OR target_class = %s OR target_class = 'tüm_okul')
                AND created_at > NOW() - INTERVAL '24 hours'
            """, (student_class, f'tüm_{class_level}'))
            result = cur.fetchone()
            new_exams = result['count'] if result else 0
            if new_exams > 0:
                notifications.append({"id": None, "message": f"🎯 {new_exams} yeni sınavınız var!", "type": "exam"})
            
            cur.execute("""
                SELECT COUNT(*) as count FROM assignments 
                WHERE (target_class = %s OR target_class = %s OR target_class = 'tüm_okul')
                AND created_at > NOW() - INTERVAL '24 hours'
            """, (student_class, f'tüm_{class_level}'))
            result = cur.fetchone()
            new_assignments = result['count'] if result else 0
            if new_assignments > 0:
                notifications.append({"id": None, "message": f"📚 {new_assignments} yeni ödeviniz var!", "type": "assignment"})
            
            cur.execute("""
                SELECT COUNT(*) as count FROM announcements 
                WHERE (target_class = %s OR target_class = %s OR target_class = 'tüm_okul')
                AND created_at > NOW() - INTERVAL '24 hours'
            """, (student_class, f'tüm_{class_level}'))
            result = cur.fetchone()
            new_announcements = result['count'] if result else 0
            if new_announcements > 0:
                notifications.append({"id": None, "message": f"📢 {new_announcements} yeni duyuru var!", "type": "announcement"})
    
    elif current_user.role == 'teacher':
        cur.execute("""
            SELECT COUNT(*) as count FROM student_questions 
            WHERE teacher_id = %s AND answer_text IS NULL
        """, (current_user.id,))
        result = cur.fetchone()
        unanswered = result['count'] if result else 0
        if unanswered > 0:
            notifications.append({"id": None, "message": f"💬 {unanswered} yeni mesajınız var!", "type": "message"})
    
    cur.close()
    conn.close()
    
    # Backward compatible: string array + object array
    notification_strings = [n.get('message', n) if isinstance(n, dict) else n for n in notifications]
    
    return jsonify({
        "notifications": notification_strings,  # Legacy string array
        "detailed_notifications": notifications,  # New object array
        "count": len(notifications)
    })

# Bildirim oluştur
@app.route("/api/notifications/create", methods=["POST"])
@login_required
def create_notification():
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    data = request.get_json()
    user_id = data.get('user_id')
    message = data.get('message')
    notif_type = data.get('type', 'info')
    
    if not user_id or not message:
        return jsonify({"error": "user_id ve message gerekli"}), 400
    
    conn = get_db()
    cur = conn.cursor()
    
    cur.execute("""
        INSERT INTO notifications (user_id, message, type, is_read)
        VALUES (%s, %s, %s, FALSE)
    """, (user_id, message, notif_type))
    
    conn.commit()
    cur.close()
    conn.close()
    
    return jsonify({"success": True, "message": "Bildirim oluşturuldu"})

# Bildirimi okundu işaretle
@app.route("/api/notifications/mark_read", methods=["POST"])
@login_required
def mark_notification_read():
    data = request.get_json()
    notif_id = data.get('id')
    
    if not notif_id:
        return jsonify({"error": "Bildirim ID gerekli"}), 400
    
    conn = get_db()
    cur = conn.cursor()
    
    cur.execute("""
        UPDATE notifications
        SET is_read = TRUE
        WHERE id = %s AND user_id = %s
    """, (notif_id, current_user.id))
    
    conn.commit()
    cur.close()
    conn.close()
    
    return jsonify({"success": True})

# =============================================
# GAMIFICATION API (Task #7)
# =============================================

def calculate_badges_for_student(student_id, cur):
    """Belirli bir öğrenci için rozetleri hesapla ve veritabanına ekle"""
    import re
    from decimal import Decimal
    
    # Öğrenci istatistiklerini al
    cur.execute("""
        SELECT 
            COUNT(*) as total_exams,
            COALESCE(MAX(lgs_score), 0) as highest_score,
            COALESCE(AVG(lgs_score), 0) as avg_score
        FROM practice_exams
        WHERE student_id = %s
    """, (student_id,))
    stats = cur.fetchone()
    
    if not stats:
        return []
    
    # Değerleri güvenli şekilde sayıya çevir
    total_exams = int(stats['total_exams']) if stats['total_exams'] else 0
    highest_score = float(stats['highest_score']) if stats['highest_score'] else 0
    avg_score = float(stats['avg_score']) if stats['avg_score'] else 0
    
    if total_exams == 0:
        return []
    
    # Öğrenci sınıf seviyesini al
    cur.execute("SELECT class_name FROM users WHERE id = %s", (student_id,))
    user_data = cur.fetchone()
    student_class = user_data['class_name'] if user_data else None
    
    class_level = None
    if student_class:
        match = re.match(r'(\d+)', student_class)
        if match:
            class_level = match.group(1)
    
    # Kaç kez 1. olduğunu hesapla
    first_place_count = 0
    if class_level:
        cur.execute("""
            WITH exam_ranks AS (
                SELECT 
                    pe.student_id,
                    pe.exam_number,
                    RANK() OVER (PARTITION BY pe.exam_number ORDER BY pe.lgs_score DESC) as rank
                FROM practice_exams pe
                JOIN users u ON pe.student_id = u.id
                WHERE u.class_name LIKE %s
            )
            SELECT COUNT(*) as first_count
            FROM exam_ranks
            WHERE rank = 1 AND student_id = %s
        """, (class_level + '%', student_id))
        result = cur.fetchone()
        first_place_count = result['first_count'] if result else 0
    
    # Gelişim hesapla
    improvement_amount = 0
    if total_exams >= 5:
        cur.execute("""
            WITH first_5 AS (
                SELECT AVG(lgs_score) as avg FROM (
                    SELECT lgs_score FROM practice_exams 
                    WHERE student_id = %s ORDER BY exam_number ASC LIMIT 5
                ) f
            ),
            last_5 AS (
                SELECT AVG(lgs_score) as avg FROM (
                    SELECT lgs_score FROM practice_exams 
                    WHERE student_id = %s ORDER BY exam_number DESC LIMIT 5
                ) l
            )
            SELECT first_5.avg as first_avg, last_5.avg as last_avg FROM first_5, last_5
        """, (student_id, student_id))
        improvement = cur.fetchone()
        if improvement and improvement['first_avg'] and improvement['last_avg']:
            improvement_amount = round(float(improvement['last_avg']) - float(improvement['first_avg']), 1)
    
    # Tamamlanan ödev sayısı (gönderilen ödevler)
    cur.execute("""
        SELECT COUNT(*) as completed_count
        FROM assignment_submissions
        WHERE student_id = %s
    """, (student_id,))
    assignment_result = cur.fetchone()
    completed_assignments = assignment_result['completed_count'] if assignment_result else 0
    
    # Mevcut rozetleri al
    cur.execute("""
        SELECT achievement_type FROM student_achievements WHERE student_id = %s
    """, (student_id,))
    earned = cur.fetchall()
    earned_types = {a['achievement_type'] for a in earned}
    
    new_badges = []
    
    # Deneme rozetleri
    if total_exams >= 10 and 'exam_10' not in earned_types:
        cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'exam_10', '🎯 İlk 10 Deneme', '10 deneme sınavını tamamladın!')", (student_id,))
        new_badges.append('exam_10')
    if total_exams >= 25 and 'exam_25' not in earned_types:
        cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'exam_25', '🔥 25 Deneme Uzmanı', '25 deneme sınavını tamamladın!')", (student_id,))
        new_badges.append('exam_25')
    if total_exams >= 50 and 'exam_50' not in earned_types:
        cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'exam_50', '👑 Deneme Şampiyonu', '50 deneme sınavını tamamladın!')", (student_id,))
        new_badges.append('exam_50')
    
    # Puan rozetleri
    if highest_score >= 400 and 'score_400' not in earned_types:
        cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'score_400', '⭐ 400 Puan Kulübü', '400+ LGS puanına ulaştın!')", (student_id,))
        new_badges.append('score_400')
    if highest_score >= 450 and 'score_450' not in earned_types:
        cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'score_450', '🌟 450 Puan Yıldızı', '450+ LGS puanına ulaştın!')", (student_id,))
        new_badges.append('score_450')
    if highest_score >= 480 and 'score_480' not in earned_types:
        cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'score_480', '💎 Elmas Seviye', '480+ LGS puanına ulaştın! Harika!')", (student_id,))
        new_badges.append('score_480')
    
    # 1. olma rozetleri
    if first_place_count >= 1 and 'first_place_1' not in earned_types:
        cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'first_place_1', '🥇 İlk Birincilik', 'Bir denemede 1. oldun!')", (student_id,))
        new_badges.append('first_place_1')
    if first_place_count >= 3 and 'first_place_3' not in earned_types:
        cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'first_place_3', '🏆 Üçlü Şampiyon', '3 denemede 1. oldun!')", (student_id,))
        new_badges.append('first_place_3')
    if first_place_count >= 5 and 'first_place_5' not in earned_types:
        cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'first_place_5', '👑 Lider', '5 denemede 1. oldun! Süpersin!')", (student_id,))
        new_badges.append('first_place_5')
    if first_place_count >= 10 and 'first_place_10' not in earned_types:
        cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'first_place_10', '🦁 Efsane', '10 denemede 1. oldun! Efsanesin!')", (student_id,))
        new_badges.append('first_place_10')
    
    # Gelişim rozetleri
    if improvement_amount >= 25 and 'improvement_25' not in earned_types:
        cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'improvement_25', '📈 Yükselen Yıldız', 'LGS puanını 25+ puan artırdın!')", (student_id,))
        new_badges.append('improvement_25')
    if improvement_amount >= 50 and 'improvement_50' not in earned_types:
        cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'improvement_50', '🚀 Süper Gelişim', 'LGS puanını 50+ puan artırdın!')", (student_id,))
        new_badges.append('improvement_50')
    if improvement_amount >= 100 and 'improvement_100' not in earned_types:
        cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'improvement_100', '💫 İnanılmaz Atılım', 'LGS puanını 100+ puan artırdın! Harika!')", (student_id,))
        new_badges.append('improvement_100')
    
    # Ödev rozetleri
    if completed_assignments >= 5 and 'assignment_5' not in earned_types:
        cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'assignment_5', '📚 Ödev Sever', '5 ödevi tamamladın!')", (student_id,))
        new_badges.append('assignment_5')
    if completed_assignments >= 15 and 'assignment_15' not in earned_types:
        cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'assignment_15', '📖 Çalışkan Öğrenci', '15 ödevi tamamladın!')", (student_id,))
        new_badges.append('assignment_15')
    if completed_assignments >= 30 and 'assignment_30' not in earned_types:
        cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'assignment_30', '🎓 Ödev Ustası', '30 ödevi tamamladın! Muhteşem!')", (student_id,))
        new_badges.append('assignment_30')
    
    # Kitap Kurdu rozeti (liderlik tablosunda ilk 3'te yer alanlara)
    try:
        cur.execute("""
            WITH book_rankings AS (
                SELECT student_id, COUNT(*) as approved_count,
                       RANK() OVER (ORDER BY COUNT(*) DESC) as rank
                FROM book_challenge_submissions
                WHERE status = 'approved'
                GROUP BY student_id
            )
            SELECT rank FROM book_rankings WHERE student_id = %s
        """, (student_id,))
        book_rank_result = cur.fetchone()
        
        if book_rank_result and book_rank_result['rank'] <= 3 and 'book_worm' not in earned_types:
            cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'book_worm', '📚 Kitap Kurdu', 'Kitap Kurdu liderlik tablosunda ilk 3te yer aldın!')", (student_id,))
            new_badges.append('book_worm')
    except Exception as e:
        logger.info(f"Book worm badge check skipped: {e}")
    
    return new_badges

@app.route("/api/admin/refresh-all-badges", methods=["POST"])
@login_required
def refresh_all_badges():
    """Tüm öğrencilerin rozetlerini yeniden hesapla"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Tüm öğrencileri al
        cur.execute("SELECT id, full_name FROM users WHERE role = 'student'")
        students = cur.fetchall()
        
        total_new_badges = 0
        updated_students = 0
        errors = []
        
        for student in students:
            try:
                new_badges = calculate_badges_for_student(student['id'], cur)
                if new_badges:
                    total_new_badges += len(new_badges)
                    updated_students += 1
            except Exception as student_error:
                errors.append(f"Öğrenci {student['id']}: {str(student_error)}")
                logger.error(f"Badge calculation error for student {student['id']}: {str(student_error)}")
        
        conn.commit()
        cur.close()
        conn.close()
        
        result = {
            "success": True,
            "message": f"{updated_students} öğrenciye toplam {total_new_badges} yeni rozet verildi",
            "updated_students": updated_students,
            "total_new_badges": total_new_badges
        }
        
        if errors:
            result["warnings"] = errors[:5]
        
        return jsonify(result)
    except Exception as e:
        logger.error(f"Refresh badges error: {str(e)}")
        if conn:
            try:
                conn.close()
            except:
                pass
        return jsonify({"error": str(e)}), 500

@app.route("/api/student/badges", methods=["GET"])
@login_required
def get_student_badges():
    if current_user.role not in ['student', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Öğrenci istatistiklerini al
        cur.execute("""
            SELECT 
                COUNT(*) as total_exams,
                MAX(lgs_score) as highest_score,
                AVG(lgs_score) as avg_score
            FROM practice_exams
            WHERE student_id = %s
        """, (current_user.id,))
        stats = cur.fetchone()
        
        # Öğrenci sınıf seviyesini al
        cur.execute("SELECT class_name FROM users WHERE id = %s", (current_user.id,))
        user_data = cur.fetchone()
        student_class = user_data['class_name'] if user_data else None
        
        class_level = None
        if student_class:
            import re
            match = re.match(r'(\d+)', student_class)
            if match:
                class_level = match.group(1)
        
        # Kaç kez 1. olduğunu hesapla
        first_place_count = 0
        if class_level:
            cur.execute("""
                WITH exam_ranks AS (
                    SELECT 
                        pe.student_id,
                        pe.exam_number,
                        RANK() OVER (PARTITION BY pe.exam_number ORDER BY pe.lgs_score DESC) as rank
                    FROM practice_exams pe
                    JOIN users u ON pe.student_id = u.id
                    WHERE u.class_name LIKE %s
                )
                SELECT COUNT(*) as first_count
                FROM exam_ranks
                WHERE rank = 1 AND student_id = %s
            """, (class_level + '%', current_user.id))
            result = cur.fetchone()
            first_place_count = result['first_count'] if result else 0
        
        # Gelişim hesapla (ilk 5 vs son 5) - sadece 5+ deneme varsa
        improvement_amount = 0
        if stats['total_exams'] and stats['total_exams'] >= 5:
            cur.execute("""
                WITH first_5 AS (
                    SELECT AVG(lgs_score) as avg FROM (
                        SELECT lgs_score FROM practice_exams 
                        WHERE student_id = %s ORDER BY exam_number ASC LIMIT 5
                    ) f
                ),
                last_5 AS (
                    SELECT AVG(lgs_score) as avg FROM (
                        SELECT lgs_score FROM practice_exams 
                        WHERE student_id = %s ORDER BY exam_number DESC LIMIT 5
                    ) l
                )
                SELECT first_5.avg as first_avg, last_5.avg as last_avg FROM first_5, last_5
            """, (current_user.id, current_user.id))
            improvement = cur.fetchone()
            if improvement and improvement['first_avg'] and improvement['last_avg']:
                improvement_amount = round(float(improvement['last_avg']) - float(improvement['first_avg']), 1)
        
        # Tamamlanan ödev sayısı (teslim edilmiş tüm ödevler)
        cur.execute("""
            SELECT COUNT(*) as completed_count
            FROM assignment_submissions
            WHERE student_id = %s
        """, (current_user.id,))
        assignment_result = cur.fetchone()
        completed_assignments = assignment_result['completed_count'] if assignment_result else 0
        
        # Mevcut rozetleri al
        cur.execute("""
            SELECT achievement_type, achievement_name, description, earned_at
            FROM student_achievements
            WHERE student_id = %s
            ORDER BY earned_at DESC
        """, (current_user.id,))
        earned = cur.fetchall()
        earned_types = {a['achievement_type'] for a in earned}
        
        # YENİ ROZETLERİ KONTROL ET VE EKLE
        new_badges = []
        
        # 10, 25, 50 Deneme rozetleri
        if stats['total_exams'] and stats['total_exams'] >= 10 and 'exam_10' not in earned_types:
            cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'exam_10', '🎯 İlk 10 Deneme', '10 deneme sınavını tamamladın!')", (current_user.id,))
            new_badges.append('exam_10')
        
        if stats['total_exams'] and stats['total_exams'] >= 25 and 'exam_25' not in earned_types:
            cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'exam_25', '🔥 25 Deneme Uzmanı', '25 deneme sınavını tamamladın!')", (current_user.id,))
            new_badges.append('exam_25')
        
        if stats['total_exams'] and stats['total_exams'] >= 50 and 'exam_50' not in earned_types:
            cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'exam_50', '👑 Deneme Şampiyonu', '50 deneme sınavını tamamladın!')", (current_user.id,))
            new_badges.append('exam_50')
        
        # 400, 450, 480+ Puan rozetleri
        if stats['highest_score'] and stats['highest_score'] >= 400 and 'score_400' not in earned_types:
            cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'score_400', '⭐ 400 Puan Kulübü', '400+ LGS puanına ulaştın!')", (current_user.id,))
            new_badges.append('score_400')
        
        if stats['highest_score'] and stats['highest_score'] >= 450 and 'score_450' not in earned_types:
            cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'score_450', '🌟 450 Puan Yıldızı', '450+ LGS puanına ulaştın!')", (current_user.id,))
            new_badges.append('score_450')
        
        if stats['highest_score'] and stats['highest_score'] >= 480 and 'score_480' not in earned_types:
            cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'score_480', '💎 Elmas Seviye', '480+ LGS puanına ulaştın! Harika!')", (current_user.id,))
            new_badges.append('score_480')
        
        # 1. Olma rozetleri
        if first_place_count >= 1 and 'first_place_1' not in earned_types:
            cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'first_place_1', '🥇 İlk Birincilik', 'Bir denemede 1. oldun!')", (current_user.id,))
            new_badges.append('first_place_1')
        
        if first_place_count >= 3 and 'first_place_3' not in earned_types:
            cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'first_place_3', '🏆 Üçlü Şampiyon', '3 denemede 1. oldun!')", (current_user.id,))
            new_badges.append('first_place_3')
        
        if first_place_count >= 5 and 'first_place_5' not in earned_types:
            cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'first_place_5', '👑 Lider', '5 denemede 1. oldun! Süpersin!')", (current_user.id,))
            new_badges.append('first_place_5')
        
        if first_place_count >= 10 and 'first_place_10' not in earned_types:
            cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'first_place_10', '🦁 Efsane', '10 denemede 1. oldun! Efsanesin!')", (current_user.id,))
            new_badges.append('first_place_10')
        
        # Gelişim rozetleri
        if improvement_amount >= 25 and 'improvement_25' not in earned_types:
            cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'improvement_25', '📈 Yükselen Yıldız', 'LGS puanını 25+ puan artırdın!')", (current_user.id,))
            new_badges.append('improvement_25')
        
        if improvement_amount >= 50 and 'improvement_50' not in earned_types:
            cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'improvement_50', '🚀 Süper Gelişim', 'LGS puanını 50+ puan artırdın!')", (current_user.id,))
            new_badges.append('improvement_50')
        
        if improvement_amount >= 100 and 'improvement_100' not in earned_types:
            cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'improvement_100', '💫 İnanılmaz Atılım', 'LGS puanını 100+ puan artırdın! Harika!')", (current_user.id,))
            new_badges.append('improvement_100')
        
        # Ödev rozetleri
        if completed_assignments >= 5 and 'assignment_5' not in earned_types:
            cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'assignment_5', '📚 Ödev Sever', '5 ödevi tamamladın!')", (current_user.id,))
            new_badges.append('assignment_5')
        
        if completed_assignments >= 15 and 'assignment_15' not in earned_types:
            cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'assignment_15', '📖 Çalışkan Öğrenci', '15 ödevi tamamladın!')", (current_user.id,))
            new_badges.append('assignment_15')
        
        if completed_assignments >= 30 and 'assignment_30' not in earned_types:
            cur.execute("INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description) VALUES (%s, 'assignment_30', '🎓 Ödev Ustası', '30 ödevi tamamladın! Muhteşem!')", (current_user.id,))
            new_badges.append('assignment_30')
        
        conn.commit()
        
        # Güncel rozet listesini al
        cur.execute("""
            SELECT achievement_type as type, achievement_name as title, description, earned_at as earned_date
            FROM student_achievements
            WHERE student_id = %s
            ORDER BY earned_at DESC
        """, (current_user.id,))
        badges = cur.fetchall()
        
        badge_list = []
        for badge in badges:
            badge_list.append({
                "type": badge['type'],
                "title": badge['title'],
                "description": badge['description'],
                "earned_date": badge['earned_date'].isoformat() if badge['earned_date'] else None
            })
        
        cur.close()
        conn.close()
        
        return jsonify({
            "badges": badge_list, 
            "count": len(badge_list),
            "new_badges": new_badges,
            "stats": {
                "total_exams": stats['total_exams'] or 0,
                "highest_score": round(float(stats['highest_score']), 1) if stats['highest_score'] else 0,
                "first_place_count": first_place_count,
                "improvement": improvement_amount,
                "completed_assignments": completed_assignments
            }
        })
    except Exception as e:
        logging.error(f"Get badges error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/student/leaderboard", methods=["GET"])
@login_required
def get_leaderboard():
    """Sadece öğrencinin kendi liderlik başarılarını döndürür - diğer öğrencileri göstermez"""
    if current_user.role not in ['student', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Öğrenci bilgilerini al
        cur.execute("SELECT class_name, full_name FROM users WHERE id = %s", (current_user.id,))
        user_data = cur.fetchone()
        student_class = user_data['class_name'] if user_data else None
        student_name = user_data['full_name'] if user_data else 'Öğrenci'
        
        # Sınıf seviyesini belirle (örn: 7A, 7B -> 7)
        class_level = None
        if student_class:
            import re
            match = re.match(r'(\d+)', student_class)
            if match:
                class_level = match.group(1)
        
        # Öğrencinin kendi istatistiklerini al
        cur.execute("""
            SELECT 
                COUNT(*) as exam_count,
                COALESCE(AVG(lgs_score), 0) as avg_lgs,
                COALESCE(MAX(lgs_score), 0) as max_lgs
            FROM practice_exams
            WHERE student_id = %s
        """, (current_user.id,))
        my_stats = cur.fetchone()
        
        # Kaç kez 1. olduğunu hesapla
        first_place_count = 0
        first_place_exams = []
        
        if class_level and my_stats['exam_count'] and my_stats['exam_count'] > 0:
            cur.execute("""
                WITH exam_ranks AS (
                    SELECT 
                        pe.student_id,
                        pe.exam_number,
                        pe.lgs_score,
                        RANK() OVER (PARTITION BY pe.exam_number ORDER BY pe.lgs_score DESC) as rank
                    FROM practice_exams pe
                    JOIN users u ON pe.student_id = u.id
                    WHERE u.class_name LIKE %s
                )
                SELECT exam_number, lgs_score
                FROM exam_ranks
                WHERE rank = 1 AND student_id = %s
                ORDER BY exam_number DESC
            """, (class_level + '%', current_user.id))
            first_places = cur.fetchall()
            first_place_count = len(first_places)
            first_place_exams = [{"exam": f"Deneme {row['exam_number']}", "score": round(float(row['lgs_score']), 1)} for row in first_places[:5]]
        
        cur.close()
        conn.close()
        
        return jsonify({
            "student_name": student_name,
            "class_name": student_class,
            "first_place_count": first_place_count,
            "first_place_exams": first_place_exams,
            "avg_lgs": round(float(my_stats['avg_lgs']), 1) if my_stats['avg_lgs'] else 0,
            "max_lgs": round(float(my_stats['max_lgs']), 1) if my_stats['max_lgs'] else 0,
            "exam_count": my_stats['exam_count'] or 0,
            "has_leadership": first_place_count > 0
        })
    except Exception as e:
        logging.error(f"Get leaderboard error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# =============================================
# VELİ PORTALI (PARENT PORTAL) - Task #8
# =============================================

@app.route("/api/parent/children", methods=["GET"])
@login_required
def get_parent_children():
    """Velinin çocuklarını listeler"""
    if current_user.role != 'parent':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT u.id, u.full_name, u.class_name, pc.relationship
            FROM parent_children pc
            JOIN users u ON pc.student_id = u.id
            WHERE pc.parent_id = %s
            ORDER BY u.full_name
        """, (current_user.id,))
        
        children = cur.fetchall()
        
        cur.close()
        conn.close()
        return jsonify({
            "children": [{
                "id": c['id'],
                "name": c['full_name'],
                "class": c['class_name'],
                "relationship": c['relationship']
            } for c in children],
            "count": len(children)
        })
    except Exception as e:
        logging.error(f"Get parent children error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/parent/child/<int:child_id>/exams", methods=["GET"])
@login_required
def get_child_exams(child_id):
    """Velinin çocuğunun sınavlarını listeler"""
    if current_user.role != 'parent':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Verify parent owns this child
        cur.execute("""
            SELECT 1 FROM parent_children 
            WHERE parent_id = %s AND student_id = %s
        """, (current_user.id, child_id))
        if not cur.fetchone():
            cur.close()
            conn.close()
            return jsonify({"error": "Bu öğrenciye erişim yetkiniz yok"}), 403
        
        # Get child's exam submissions
        cur.execute("""
            SELECT e.id, e.title, e.question_count, e.start_time, 
                   es.score, es.submitted_at
            FROM exam_submissions es
            JOIN exams e ON es.exam_id = e.id
            WHERE es.student_id = %s
            ORDER BY e.start_time DESC
        """, (child_id,))
        
        exams = cur.fetchall()
        
        cur.close()
        conn.close()
        return jsonify({
            "exams": [{
                "id": ex['id'],
                "title": ex['title'],
                "question_count": ex['question_count'],
                "start_time": ex['start_time'].isoformat() if ex['start_time'] else None,
                "score": float(ex['score']) if ex['score'] else 0,
                "submitted_at": ex['submitted_at'].isoformat() if ex['submitted_at'] else None
            } for ex in exams],
            "count": len(exams)
        })
    except Exception as e:
        logging.error(f"Get child exams error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/parent/child/<int:child_id>/assignments", methods=["GET"])
@login_required
def get_child_assignments(child_id):
    """Velinin çocuğunun ödevlerini listeler"""
    if current_user.role != 'parent':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Verify parent owns this child
        cur.execute("""
            SELECT 1 FROM parent_children 
            WHERE parent_id = %s AND student_id = %s
        """, (current_user.id, child_id))
        if not cur.fetchone():
            cur.close()
            conn.close()
            return jsonify({"error": "Bu öğrenciye erişim yetkiniz yok"}), 403
        
        # Get child's assignment submissions
        cur.execute("""
            SELECT a.id, a.title, a.description, a.due_date, 
                   asub.status, asub.submitted_at, asub.grade
            FROM assignments a
            LEFT JOIN assignment_submissions asub ON a.id = asub.assignment_id AND asub.student_id = %s
            WHERE a.target_class = (SELECT class_name FROM users WHERE id = %s)
            ORDER BY a.due_date DESC
        """, (child_id, child_id))
        
        assignments = cur.fetchall()
        
        cur.close()
        conn.close()
        return jsonify({
            "assignments": [{
                "id": a['id'],
                "title": a['title'],
                "description": a['description'],
                "due_date": a['due_date'].isoformat() if a['due_date'] else None,
                "status": a['status'] or 'not_submitted',
                "submitted_at": a['submitted_at'].isoformat() if a.get('submitted_at') else None,
                "grade": a['grade'] or None
            } for a in assignments],
            "count": len(assignments)
        })
    except Exception as e:
        logging.error(f"Get child assignments error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/parent/messages", methods=["GET"])
@login_required
def get_parent_messages_api():
    """Velinin mesajlarını listeler"""
    if current_user.role != 'parent':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT pm.*, u.full_name as student_name, u.class_name
            FROM parent_messages pm
            JOIN users u ON pm.student_id = u.id
            WHERE pm.parent_id = %s
            ORDER BY pm.created_at DESC
        """, (current_user.id,))
        
        messages = cur.fetchall()
        
        cur.close()
        conn.close()
        return jsonify({
            "messages": [{
                "id": m['id'],
                "student_name": m['student_name'],
                "class_name": m['class_name'],
                "message": m['message'],
                "teacher_response": m.get('teacher_response'),
                "status": m['status'],
                "created_at": m['created_at'].isoformat() if m['created_at'] else None,
                "responded_at": m['responded_at'].isoformat() if m.get('responded_at') else None
            } for m in messages],
            "count": len(messages)
        })
    except Exception as e:
        logging.error(f"Get parent messages error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/parent/send-message", methods=["POST"])
@login_required
def send_parent_message():
    """Veli öğretmene mesaj gönderir"""
    if current_user.role != 'parent':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    data = request.get_json()
    if not data or 'student_id' not in data or 'message' not in data:
        return jsonify({"error": "student_id ve message gerekli"}), 400
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # Verify parent owns this child
        cur.execute("""
            SELECT 1 FROM parent_children 
            WHERE parent_id = %s AND student_id = %s
        """, (current_user.id, data['student_id']))
        if not cur.fetchone():
            cur.close()
            conn.close()
            return jsonify({"error": "Bu öğrenciye erişim yetkiniz yok"}), 403
        
        # Insert message with parent_id
        cur.execute("""
            INSERT INTO parent_messages (student_id, parent_id, message)
            VALUES (%s, %s, %s)
        """, (data['student_id'], current_user.id, data['message']))
        
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"success": True, "message": "Mesaj gönderildi"}), 200
    except Exception as e:
        logging.error(f"Send parent message error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# =============================================
# GLOBAL ARAMA (ADVANCED SEARCH) - Task #9
# =============================================

@app.route("/api/search", methods=["GET"])
@login_required
def global_search():
    """Global arama - kullanıcı, sınav, ödev, duyuru"""
    query = request.args.get('q', '').strip()
    
    # Always return categorical structure
    empty_results = {"users": [], "exams": [], "assignments": [], "announcements": []}
    
    if not query or len(query) < 2:
        return jsonify({"results": empty_results})
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        search_pattern = f"%{query}%"
        results = {"users": [], "exams": [], "assignments": [], "announcements": []}
        
        # Role-based search
        if current_user.role == 'admin':
            # Admin: Search all users
            cur.execute("""
                SELECT id, full_name, role, class_name 
                FROM users 
                WHERE full_name ILIKE %s 
                LIMIT 5
            """, (search_pattern,))
            results["users"] = [{"id": u['id'], "title": u['full_name'], "subtitle": f"{u['role']} {u['class_name'] or ''}".strip(), "url": "/admin/users"} for u in cur.fetchall()]
        
        # Exams (admin sees all, teacher sees own, student sees accessible)
        if current_user.role == 'admin':
            cur.execute("""
                SELECT id, title, start_time 
                FROM exams 
                WHERE title ILIKE %s 
                LIMIT 5
            """, (search_pattern,))
            results["exams"] = [{"id": e['id'], "title": e['title'], "subtitle": e['start_time'].strftime('%d/%m/%Y %H:%M') if e['start_time'] else '', "url": "/admin/exams"} for e in cur.fetchall()]
        elif current_user.role == 'teacher':
            cur.execute("""
                SELECT id, title, start_time 
                FROM exams 
                WHERE title ILIKE %s AND teacher_id = %s 
                LIMIT 5
            """, (search_pattern, current_user.id))
            results["exams"] = [{"id": e['id'], "title": e['title'], "subtitle": e['start_time'].strftime('%d/%m/%Y %H:%M') if e['start_time'] else '', "url": "/teacher/exams"} for e in cur.fetchall()]
        elif current_user.role == 'student':
            cur.execute("""
                SELECT id, title, start_time 
                FROM exams 
                WHERE title ILIKE %s AND target_class = %s 
                LIMIT 5
            """, (search_pattern, current_user.class_name))
            results["exams"] = [{"id": e['id'], "title": e['title'], "subtitle": e['start_time'].strftime('%d/%m/%Y %H:%M') if e['start_time'] else '', "url": "/student/exams"} for e in cur.fetchall()]
        
        # Assignments
        if current_user.role in ['admin', 'teacher']:
            cur.execute("""
                SELECT id, title, due_date 
                FROM assignments 
                WHERE title ILIKE %s 
                LIMIT 5
            """, (search_pattern,))
            url = "/teacher/assignments" if current_user.role == 'teacher' else "/admin/assignments"
            results["assignments"] = [{"id": a['id'], "title": a['title'], "subtitle": a['due_date'].strftime('%d/%m/%Y') if a['due_date'] else '', "url": url} for a in cur.fetchall()]
        elif current_user.role == 'student':
            cur.execute("""
                SELECT id, title, due_date 
                FROM assignments 
                WHERE title ILIKE %s AND target_class = %s 
                LIMIT 5
            """, (search_pattern, current_user.class_name))
            results["assignments"] = [{"id": a['id'], "title": a['title'], "subtitle": a['due_date'].strftime('%d/%m/%Y') if a['due_date'] else '', "url": "/student/assignments"} for a in cur.fetchall()]
        
        # Parent role - search children's accessible content
        if current_user.role == 'parent':
            # Get children's classes
            cur.execute("""
                SELECT DISTINCT u.class_name 
                FROM users u
                JOIN parent_children pc ON u.id = pc.student_id
                WHERE pc.parent_id = %s AND u.class_name IS NOT NULL
            """, (current_user.id,))
            children_classes = [row['class_name'] for row in cur.fetchall()]
            
            # Search children (users)
            cur.execute("""
                SELECT u.id, u.full_name, u.class_name 
                FROM users u
                JOIN parent_children pc ON u.id = pc.student_id
                WHERE pc.parent_id = %s AND u.full_name ILIKE %s
                LIMIT 5
            """, (current_user.id, search_pattern))
            results["users"] = [{"id": u['id'], "title": u['full_name'], "subtitle": f"Öğrenci - {u['class_name']}", "url": "/parent/dashboard"} for u in cur.fetchall()]
            
            # Search exams accessible to children
            if children_classes:
                placeholders = ','.join(['%s'] * len(children_classes))
                cur.execute(f"""
                    SELECT id, title, start_time, target_class
                    FROM exams
                    WHERE title ILIKE %s AND target_class IN ({placeholders})
                    LIMIT 5
                """, (search_pattern, *children_classes))
                results["exams"] = [{"id": e['id'], "title": e['title'], "subtitle": f"{e['target_class']} - {e['start_time'].strftime('%d/%m/%Y') if e['start_time'] else ''}", "url": "/parent/dashboard"} for e in cur.fetchall()]
                
                # Search assignments accessible to children
                cur.execute(f"""
                    SELECT id, title, due_date, target_class
                    FROM assignments
                    WHERE title ILIKE %s AND target_class IN ({placeholders})
                    LIMIT 5
                """, (search_pattern, *children_classes))
                results["assignments"] = [{"id": a['id'], "title": a['title'], "subtitle": f"{a['target_class']} - {a['due_date'].strftime('%d/%m/%Y') if a['due_date'] else ''}", "url": "/parent/dashboard"} for a in cur.fetchall()]
                
                # Search announcements for children's classes
                cur.execute(f"""
                    SELECT id, title, created_at, target_class
                    FROM announcements
                    WHERE (title ILIKE %s OR content ILIKE %s) AND target_class IN ({placeholders})
                    LIMIT 5
                """, (search_pattern, search_pattern, *children_classes))
                results["announcements"] = [{"id": a['id'], "title": a['title'], "subtitle": f"{a['target_class']} - {a['created_at'].strftime('%d/%m/%Y') if a['created_at'] else ''}", "url": "/parent/dashboard"} for a in cur.fetchall()]
        
        # Announcements
        if current_user.role == 'student':
            cur.execute("""
                SELECT id, title, created_at 
                FROM announcements 
                WHERE (title ILIKE %s OR content ILIKE %s) AND target_class = %s 
                LIMIT 5
            """, (search_pattern, search_pattern, current_user.class_name))
            results["announcements"] = [{"id": a['id'], "title": a['title'], "subtitle": a['created_at'].strftime('%d/%m/%Y') if a['created_at'] else '', "url": "/student/announcements"} for a in cur.fetchall()]
        elif current_user.role == 'teacher':
            cur.execute("""
                SELECT id, title, created_at 
                FROM teacher_announcements 
                WHERE title ILIKE %s OR content ILIKE %s 
                LIMIT 5
            """, (search_pattern, search_pattern))
            results["announcements"] = [{"id": a['id'], "title": a['title'], "subtitle": a['created_at'].strftime('%d/%m/%Y') if a['created_at'] else '', "url": "/teacher/announcements"} for a in cur.fetchall()]
        
        cur.close()
        conn.close()
        return jsonify({"results": results})
    except Exception as e:
        logging.error(f"Search error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# =============================================
# DASHBOARD WIDGET SYSTEM - Task #10
# =============================================

# Default widget configurations per role
DEFAULT_WIDGETS = {
    "student": [
        {"id": "exams", "title": "📝 Sınavlarım", "url": "/student/exams", "size": "large", "order": 1},
        {"id": "assignments", "title": "📚 Ödevlerim", "url": "/student/assignments", "size": "large", "order": 2},
        {"id": "announcements", "title": "📢 Duyurular", "url": "/student/announcements", "size": "large", "order": 3},
        {"id": "ask-question", "title": "💬 Soru Sor", "url": "/student/ask-question-page", "size": "large", "order": 4},
        {"id": "badges", "title": "🏅 Rozetlerim", "url": None, "size": "medium", "order": 5},
        {"id": "leaderboard", "title": "📊 Lider Tablosu", "url": None, "size": "medium", "order": 6}
    ],
    "teacher": [
        {"id": "exams", "title": "📝 Sınavlarım", "url": "/teacher/exams", "size": "large", "order": 1},
        {"id": "assignments", "title": "📚 Ödevlerim", "url": "/teacher/assignments", "size": "large", "order": 2},
        {"id": "announcements", "title": "📢 Duyurular", "url": "/teacher/announcements", "size": "large", "order": 3},
        {"id": "questions", "title": "💬 Öğrenci Soruları", "url": "/teacher/questions", "size": "large", "order": 4}
    ],
    "admin": [
        {"id": "users", "title": "👥 Kullanıcı Yönetimi", "url": "/admin/users", "size": "large", "order": 1},
        {"id": "classes", "title": "🏫 Sınıf Yönetimi", "url": "/admin/classes-page", "size": "large", "order": 2},
        {"id": "files", "title": "📁 Dosya Yönetimi", "url": "/admin/files-page", "size": "large", "order": 3},
        {"id": "reports", "title": "📊 Raporlar", "url": "/admin/reports", "size": "large", "order": 4}
    ],
    "parent": [
        {"id": "children", "title": "👨‍👩‍👧‍👦 Çocuklarım", "url": None, "size": "large", "order": 1},
        {"id": "exams", "title": "📝 Sınavlar", "url": "/parent/dashboard", "size": "medium", "order": 2},
        {"id": "assignments", "title": "📚 Ödevler", "url": "/parent/dashboard", "size": "medium", "order": 3},
        {"id": "messages", "title": "💬 Mesajlarım", "url": "/parent/dashboard", "size": "medium", "order": 4}
    ]
}

@app.route("/api/dashboard/widgets", methods=["GET"])
@login_required
def get_dashboard_widgets():
    """Kullanıcının dashboard widget tercihlerini getir"""
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Check if user has custom preferences
        cur.execute("""
            SELECT layout, visibility FROM dashboard_widget_preferences 
            WHERE user_id = %s
        """, (current_user.id,))
        prefs = cur.fetchone()
        
        # Get default widgets for role
        default_widgets = DEFAULT_WIDGETS.get(current_user.role, [])
        
        if prefs:
            # Merge user preferences with defaults
            layout = prefs['layout'] if prefs['layout'] else []
            visibility = prefs['visibility'] if prefs['visibility'] else {}
            
            # Apply visibility settings
            widgets = []
            for widget in default_widgets:
                widget_copy = widget.copy()
                widget_copy['visible'] = visibility.get(widget['id'], True)
                widgets.append(widget_copy)
            
            # Apply custom order if exists
            if layout:
                ordered_widgets = []
                for widget_id in layout:
                    widget = next((w for w in widgets if w['id'] == widget_id), None)
                    if widget:
                        ordered_widgets.append(widget)
                # Add any widgets not in layout (new widgets)
                for widget in widgets:
                    if widget['id'] not in layout:
                        ordered_widgets.append(widget)
                widgets = ordered_widgets
            
            response = {"widgets": widgets, "hasCustomPreferences": True}
        else:
            # Return defaults with all visible
            widgets = [dict(w, visible=True) for w in default_widgets]
            response = {"widgets": widgets, "hasCustomPreferences": False}
        
        cur.close()
        conn.close()
        return jsonify(response)
    except Exception as e:
        logger.error(f"Get widgets error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/dashboard/widgets", methods=["POST"])
@login_required
def save_dashboard_widgets():
    """Kullanıcının dashboard widget tercihlerini kaydet"""
    try:
        data = request.json
        layout = data.get('layout', [])  # Array of widget IDs in order
        visibility = data.get('visibility', {})  # Object: {widgetId: boolean}
        
        # Validate layout is array
        if not isinstance(layout, list):
            return jsonify({"error": "Layout must be an array"}), 400
        
        # Validate visibility is object
        if not isinstance(visibility, dict):
            return jsonify({"error": "Visibility must be an object"}), 400
        
        conn = get_db()
        cur = conn.cursor()
        
        # Upsert preferences
        cur.execute("""
            INSERT INTO dashboard_widget_preferences (user_id, role, layout, visibility, updated_at)
            VALUES (%s, %s, %s, %s, NOW())
            ON CONFLICT (user_id) 
            DO UPDATE SET layout = EXCLUDED.layout, visibility = EXCLUDED.visibility, updated_at = NOW()
        """, (current_user.id, current_user.role, json.dumps(layout), json.dumps(visibility)))
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Widget tercihleri kaydedildi"})
    except Exception as e:
        logger.error(f"Save widgets error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/dashboard/widgets/reset", methods=["POST"])
@login_required
def reset_dashboard_widgets():
    """Kullanıcının dashboard widget tercihlerini sıfırla"""
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # Delete user preferences (will fall back to defaults)
        cur.execute("DELETE FROM dashboard_widget_preferences WHERE user_id = %s", (current_user.id,))
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Widget tercihleri sıfırlandı"})
    except Exception as e:
        logger.error(f"Reset widgets error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# =============================================
# DOSYA YÖNETİMİ (ADMIN)
# =============================================

@app.route("/admin/files", methods=["GET"])
@login_required
def list_files():
    """Admin tüm dosyaları listeler"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    def get_file_info(filename):
        """Dosya bilgilerini al (Object Storage veya yerel)"""
        if not filename:
            return 0, False
        
        # Object Storage dosyası mı kontrol et
        if filename.startswith('/storage/'):
            # Object Storage'da var mı kontrol et
            if object_storage.is_available():
                try:
                    storage_path = filename[len('/storage/'):]
                    # Object Storage'da varlık kontrolü
                    return 0, True  # Boyut Object Storage'dan alınamıyor ama var
                except:
                    return 0, False
            return 0, False
        
        # Yerel dosya yolu oluştur
        if filename.startswith('/uploads/'):
            local_path = filename[1:]  # Baştaki / kaldır
        elif filename.startswith('uploads/'):
            local_path = filename
        else:
            local_path = os.path.join(UPLOAD_DIR, filename)
        
        if os.path.exists(local_path):
            try:
                return os.path.getsize(local_path), True
            except:
                return 0, True
        
        return 0, False
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        files_list = []
        
        # 1. Sınav PDF'leri
        cur.execute("""
            SELECT e.pdf_filename as filename, e.title, e.created_at, 
                   u.full_name as owner, 'Sınav PDF' as file_type, e.id as content_id
            FROM exams e
            JOIN users u ON e.teacher_id = u.id
            WHERE e.pdf_filename IS NOT NULL AND e.pdf_filename != ''
            ORDER BY e.created_at DESC
        """)
        exams = cur.fetchall() or []
        for exam in exams:
            file_size, exists = get_file_info(exam['filename'])
            files_list.append({
                'filename': exam['filename'] or '',
                'title': exam['title'] or 'Başlıksız',
                'type': exam['file_type'],
                'owner': exam['owner'] or 'Bilinmiyor',
                'size': file_size,
                'size_mb': round(file_size / (1024*1024), 2),
                'created_at': exam['created_at'].isoformat() if exam['created_at'] else None,
                'exists': exists,
                'content_id': exam['content_id'],
                'table': 'exams'
            })
        
        # 2. Ödev dosyaları (öğrenci gönderileri)
        cur.execute("""
            SELECT sub.file_path as filename, a.title, sub.submitted_at, 
                   u.full_name as owner, 'Ödev Dosyası' as file_type, sub.id as content_id
            FROM assignment_submissions sub
            JOIN assignments a ON sub.assignment_id = a.id
            JOIN users u ON sub.student_id = u.id
            WHERE sub.file_path IS NOT NULL AND sub.file_path != ''
            ORDER BY sub.submitted_at DESC
        """)
        assignments = cur.fetchall() or []
        for assignment in assignments:
            file_size, exists = get_file_info(assignment['filename'])
            files_list.append({
                'filename': assignment['filename'] or '',
                'title': assignment['title'] or 'Başlıksız',
                'type': assignment['file_type'],
                'owner': assignment['owner'] or 'Bilinmiyor',
                'size': file_size,
                'size_mb': round(file_size / (1024*1024), 2),
                'created_at': assignment['submitted_at'].isoformat() if assignment.get('submitted_at') else None,
                'exists': exists,
                'content_id': assignment['content_id'],
                'table': 'assignment_submissions'
            })
        
        # 3. Duyuru ekleri
        cur.execute("""
            SELECT an.attachment_path as filename, an.title, an.created_at,
                   u.full_name as owner, 'Duyuru Eki' as file_type, an.id as content_id
            FROM announcements an
            JOIN users u ON an.teacher_id = u.id
            WHERE an.attachment_path IS NOT NULL AND an.attachment_path != ''
            ORDER BY an.created_at DESC
        """)
        announcements = cur.fetchall() or []
        for announcement in announcements:
            file_size, exists = get_file_info(announcement['filename'])
            files_list.append({
                'filename': announcement['filename'] or '',
                'title': announcement['title'] or 'Başlıksız',
                'type': announcement['file_type'],
                'owner': announcement['owner'] or 'Bilinmiyor',
                'size': file_size,
                'size_mb': round(file_size / (1024*1024), 2),
                'created_at': announcement['created_at'].isoformat() if announcement.get('created_at') else None,
                'exists': exists,
                'content_id': announcement['content_id'],
                'table': 'announcements'
            })
        
        # 4. Öğrenci soru ekleri
        cur.execute("""
            SELECT sq.file_path as filename, sq.question_text, sq.created_at,
                   u.full_name as owner, 'Öğrenci Sorusu' as file_type, sq.id as content_id
            FROM student_questions sq
            JOIN users u ON sq.student_id = u.id
            WHERE sq.file_path IS NOT NULL AND sq.file_path != ''
            ORDER BY sq.created_at DESC
        """)
        questions = cur.fetchall() or []
        for question in questions:
            file_size, exists = get_file_info(question['filename'])
            question_text = question.get('question_text') or 'Soru'
            title = question_text[:50] + '...' if len(question_text) > 50 else question_text
            files_list.append({
                'filename': question['filename'] or '',
                'title': title,
                'type': question['file_type'],
                'owner': question['owner'] or 'Bilinmiyor',
                'size': file_size,
                'size_mb': round(file_size / (1024*1024), 2),
                'created_at': question['created_at'].isoformat() if question.get('created_at') else None,
                'exists': exists,
                'content_id': question['content_id'],
                'table': 'student_questions'
            })
        
        cur.close()
        conn.close()
        
        return jsonify({"files": files_list})
    
    except Exception as e:
        logger.error(f" File list error: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({"error": f"Dosyalar listelenemedi: {str(e)}"}), 500

@app.route("/admin/files/stats", methods=["GET"])
@login_required
def file_stats():
    """Dosya istatistikleri"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        # Uploads klasöründeki tüm dosyaları tara
        total_size = 0
        file_count = 0
        
        if os.path.exists(UPLOAD_DIR):
            for filename in os.listdir(UPLOAD_DIR):
                file_path = os.path.join(UPLOAD_DIR, filename)
                if os.path.isfile(file_path):
                    total_size += os.path.getsize(file_path)
                    file_count += 1
        
        # PDF ve Excel sayısı
        pdf_count = sum(1 for f in os.listdir(UPLOAD_DIR) if f.lower().endswith('.pdf')) if os.path.exists(UPLOAD_DIR) else 0
        excel_count = sum(1 for f in os.listdir(UPLOAD_DIR) if f.lower().endswith(('.xlsx', '.xls'))) if os.path.exists(UPLOAD_DIR) else 0
        
        return jsonify({
            "total_files": file_count,
            "total_size": total_size,
            "total_size_mb": round(total_size / (1024*1024), 2),
            "total_size_gb": round(total_size / (1024*1024*1024), 3),
            "pdf_count": pdf_count,
            "excel_count": excel_count
        })
    
    except Exception as e:
        logger.error(f" File stats error: {str(e)}")
        return jsonify({"error": f"İstatistikler alınamadı: {str(e)}"}), 500

@app.route("/admin/files/<filename>", methods=["DELETE"])
@login_required
def delete_file(filename):
    """Dosya sil (fiziksel + veritabanı referansını kontrol et)"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Dosya kullanımda mı kontrol et
        used_in = []
        
        # Sınavlarda kullanılıyor mu?
        cur.execute("SELECT COUNT(*) as count FROM exams WHERE pdf_filename = %s", (filename,))
        result = cur.fetchone()
        if result and result['count'] > 0:
            used_in.append(f"Sınav ({result['count']} adet)")
        
        # Ödev gönderilerinde kullanılıyor mu?
        cur.execute("SELECT COUNT(*) as count FROM assignment_submissions WHERE file_path = %s", (filename,))
        result = cur.fetchone()
        if result and result['count'] > 0:
            used_in.append(f"Ödev ({result['count']} adet)")
        
        # Duyurularda kullanılıyor mu?
        cur.execute("SELECT COUNT(*) as count FROM announcements WHERE attachment_path = %s", (filename,))
        result = cur.fetchone()
        if result and result['count'] > 0:
            used_in.append(f"Duyuru ({result['count']} adet)")
        
        # Öğrenci sorularında kullanılıyor mu?
        cur.execute("SELECT COUNT(*) as count FROM student_questions WHERE file_path = %s", (filename,))
        result = cur.fetchone()
        if result and result['count'] > 0:
            used_in.append(f"Soru ({result['count']} adet)")
        
        cur.close()
        conn.close()
        
        if used_in:
            return jsonify({
                "error": "Bu dosya hala kullanımda",
                "used_in": used_in,
                "message": "Dosyayı silmek için önce ilgili içeriği silmelisiniz"
            }), 400
        
        # Dosya kullanımda değilse fiziksel dosyayı sil
        file_path = os.path.join(UPLOAD_DIR, filename)
        if os.path.exists(file_path):
            os.remove(file_path)
            return jsonify({"success": True, "message": "Dosya silindi"})
        else:
            return jsonify({"error": "Dosya bulunamadı"}), 404
    
    except Exception as e:
        logger.error(f" Delete file error: {str(e)}")
        return jsonify({"error": f"Dosya silinemedi: {str(e)}"}), 500

@app.route("/admin/files/unused", methods=["DELETE"])
@login_required
def delete_unused_files():
    """Veritabanında kayıtlı olmayan gereksiz dosyaları sil"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Veritabanında kayıtlı tüm dosyaları topla
        used_files = set()
        
        # Sınav PDF'leri
        cur.execute("SELECT pdf_filename FROM exams WHERE pdf_filename IS NOT NULL AND pdf_filename != ''")
        for row in cur.fetchall():
            used_files.add(row['pdf_filename'])
        
        # Ödev dosyaları
        cur.execute("SELECT file_path FROM assignments WHERE file_path IS NOT NULL AND file_path != ''")
        for row in cur.fetchall():
            used_files.add(row['file_path'])
        
        # Ödev teslimleri
        cur.execute("SELECT file_path FROM assignment_submissions WHERE file_path IS NOT NULL AND file_path != ''")
        for row in cur.fetchall():
            used_files.add(row['file_path'])
        
        # Duyuru ekleri
        cur.execute("SELECT attachment_path FROM announcements WHERE attachment_path IS NOT NULL AND attachment_path != ''")
        for row in cur.fetchall():
            used_files.add(row['attachment_path'])
        
        # Öğrenci soru ekleri
        cur.execute("SELECT file_path FROM student_questions WHERE file_path IS NOT NULL AND file_path != ''")
        for row in cur.fetchall():
            used_files.add(row['file_path'])
        
        cur.close()
        conn.close()
        
        # Upload dizinindeki tüm dosyaları kontrol et
        if not os.path.exists(UPLOAD_DIR):
            return jsonify({"success": True, "deleted_count": 0, "message": "Upload dizini bulunamadı"})
        
        all_files = os.listdir(UPLOAD_DIR)
        deleted_files = []
        total_size_freed = 0
        
        for filename in all_files:
            if filename not in used_files:
                file_path = os.path.join(UPLOAD_DIR, filename)
                if os.path.isfile(file_path):  # Sadece dosyaları sil (dizinleri değil)
                    file_size = os.path.getsize(file_path)
                    os.remove(file_path)
                    deleted_files.append(filename)
                    total_size_freed += file_size
        
        return jsonify({
            "success": True,
            "deleted_count": len(deleted_files),
            "deleted_files": deleted_files[:10],  # İlk 10 dosya adı
            "size_freed_mb": round(total_size_freed / (1024 * 1024), 2),
            "message": f"{len(deleted_files)} kullanılmayan dosya silindi, {round(total_size_freed / (1024 * 1024), 2)} MB alan boşaltıldı"
        })
    
    except Exception as e:
        logger.error(f" Delete unused files error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Gereksiz dosyalar silinemedi: {str(e)}"}), 500

# =============================================
# SINIF YÖNETİMİ (ADMIN)
# =============================================

@app.route("/admin/classes", methods=["GET"])
@login_required
def list_classes():
    """TÜM KULLANICILAR sınıfları görebilir (öğretmen/öğrenci formlarda kullanır)"""
    # GET isteği - Herkes erişebilir (öğretmen sınav oluştururken kullanır)
    # POST/PUT/DELETE - Sadece admin
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT id, name, level, branch, type, is_active, created_at
            FROM classes
            WHERE is_active = TRUE
            ORDER BY 
                CASE 
                    WHEN type = 'standard' THEN 0
                    ELSE 1
                END,
                level, branch, name
        """)
        classes = cur.fetchall()
        
        cur.close()
        conn.close()
        
        return jsonify({"classes": [{
            "id": c['id'],
            "name": c['name'],
            "level": c['level'],
            "branch": c['branch'],
            "type": c['type'],
            "is_active": c['is_active'],
            "created_at": c['created_at'].isoformat() if c['created_at'] else None
        } for c in classes]})
    
    except Exception as e:
        logger.error(f" List classes error: {str(e)}")
        return jsonify({"error": f"Sınıflar listelenemedi: {str(e)}"}), 500

@app.route("/admin/classes", methods=["POST"])
@login_required
def create_class():
    """Admin yeni sınıf oluşturur"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        data = request.get_json()
        name = data.get('name', '').strip()
        level = data.get('level', '').strip()
        branch = data.get('branch', '').strip()
        class_type = data.get('type', 'custom')
        
        if not name:
            return jsonify({"error": "Sınıf adı gerekli"}), 400
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Aynı isimde sınıf var mı kontrol et
        cur.execute("SELECT id FROM classes WHERE name = %s", (name,))
        existing = cur.fetchone()
        
        if existing:
            cur.close()
            conn.close()
            return jsonify({"error": f"{name} adında sınıf zaten mevcut"}), 400
        
        # Yeni sınıf oluştur
        cur.execute("""
            INSERT INTO classes (name, level, branch, type, is_active)
            VALUES (%s, %s, %s, %s, TRUE)
            RETURNING id
        """, (name, level if level else None, branch if branch else None, class_type))
        
        result = cur.fetchone()
        new_id = result['id']
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": f"{name} sınıfı oluşturuldu", "id": new_id})
    
    except Exception as e:
        logger.error(f" Create class error: {str(e)}")
        return jsonify({"error": f"Sınıf oluşturulamadı: {str(e)}"}), 500

@app.route("/admin/classes/<int:class_id>", methods=["PUT"])
@login_required
def update_class(class_id):
    """Admin sınıf bilgilerini günceller"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        data = request.get_json()
        name = data.get('name', '').strip()
        
        if not name:
            return jsonify({"error": "Sınıf adı gerekli"}), 400
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Sınıf mevcut mu kontrol et
        cur.execute("SELECT id, name FROM classes WHERE id = %s", (class_id,))
        existing_class = cur.fetchone()
        
        if not existing_class:
            cur.close()
            conn.close()
            return jsonify({"error": "Sınıf bulunamadı"}), 404
        
        # Başka bir sınıf aynı ismi kullanıyor mu kontrol et
        cur.execute("SELECT id FROM classes WHERE name = %s AND id != %s", (name, class_id))
        duplicate = cur.fetchone()
        
        if duplicate:
            cur.close()
            conn.close()
            return jsonify({"error": f"{name} adında başka bir sınıf zaten mevcut"}), 400
        
        old_name = existing_class['name']
        
        # Sınıf adını güncelle (CASCADE: tüm ilgili tablolardaki referansları da güncelle)
        cur.execute("""
            UPDATE classes 
            SET name = %s
            WHERE id = %s
        """, (name, class_id))
        
        # Öğrencilerin sınıf isimlerini güncelle
        cur.execute("""
            UPDATE users 
            SET class_name = %s
            WHERE class_name = %s AND role = 'student'
        """, (name, old_name))
        
        # Sınavlardaki hedef sınıf isimlerini güncelle
        cur.execute("""
            UPDATE exams 
            SET target_class = %s
            WHERE target_class = %s
        """, (name, old_name))
        
        # Ödevlerdeki hedef sınıf isimlerini güncelle
        cur.execute("""
            UPDATE assignments 
            SET target_class = %s
            WHERE target_class = %s
        """, (name, old_name))
        
        # Duyurulardaki hedef sınıf isimlerini güncelle
        cur.execute("""
            UPDATE announcements 
            SET target_class = %s
            WHERE target_class = %s
        """, (name, old_name))
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": f"Sınıf '{old_name}' → '{name}' olarak güncellendi (tüm kayıtlar güncellendi)"})
    
    except Exception as e:
        if 'conn' in locals():
            conn.rollback()
        logger.error(f" Update class error: {str(e)}")
        return jsonify({"error": f"Sınıf güncellenemedi: {str(e)}"}), 500

@app.route("/admin/classes/<int:class_id>", methods=["DELETE"])
@login_required
def delete_class(class_id):
    """Admin sınıf siler"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Sınıf mevcut mu kontrol et
        cur.execute("SELECT name FROM classes WHERE id = %s", (class_id,))
        class_data = cur.fetchone()
        
        if not class_data:
            cur.close()
            conn.close()
            return jsonify({"error": "Sınıf bulunamadı"}), 404
        
        class_name = class_data['name']
        
        # Sınıf kullanımda mı kontrol et
        used_in = []
        
        # Öğrenciler bu sınıfta mı?
        cur.execute("SELECT COUNT(*) as count FROM users WHERE class_name = %s AND role = 'student'", (class_name,))
        result = cur.fetchone()
        if result and result['count'] > 0:
            used_in.append(f"{result['count']} öğrenci")
        
        # Sınavlarda hedef sınıf olarak kullanılıyor mu?
        cur.execute("SELECT COUNT(*) as count FROM exams WHERE target_class = %s", (class_name,))
        result = cur.fetchone()
        if result and result['count'] > 0:
            used_in.append(f"{result['count']} sınav")
        
        # Ödevlerde kullanılıyor mu?
        cur.execute("SELECT COUNT(*) as count FROM assignments WHERE target_class = %s", (class_name,))
        result = cur.fetchone()
        if result and result['count'] > 0:
            used_in.append(f"{result['count']} ödev")
        
        # Duyurularda kullanılıyor mu?
        cur.execute("SELECT COUNT(*) as count FROM announcements WHERE target_class = %s", (class_name,))
        result = cur.fetchone()
        if result and result['count'] > 0:
            used_in.append(f"{result['count']} duyuru")
        
        if used_in:
            cur.close()
            conn.close()
            return jsonify({
                "error": f"{class_name} sınıfı hala kullanımda",
                "used_in": used_in,
                "message": "Sınıfı silmek için önce ilgili öğrenci/içerikleri kaldırmalısınız"
            }), 400
        
        # Sınıf kullanımda değilse sil
        cur.execute("DELETE FROM classes WHERE id = %s", (class_id,))
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": f"{class_name} sınıfı silindi"})
    
    except Exception as e:
        logger.error(f" Delete class error: {str(e)}")
        return jsonify({"error": f"Sınıf silinemedi: {str(e)}"}), 500

@app.route("/api/admin/classes/<int:class_id>/rename", methods=["PUT"])
@login_required
def rename_class(class_id):
    """Admin bir sınıfın adını değiştirir (öğrencilerin sınıf bilgisi de güncellenir)"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        data = request.get_json()
        new_name = data.get('new_name', '').strip().upper()
        
        if not new_name:
            return jsonify({"error": "Yeni sınıf adı gerekli"}), 400
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Mevcut sınıf bilgisini al
        cur.execute("SELECT name FROM classes WHERE id = %s", (class_id,))
        class_data = cur.fetchone()
        
        if not class_data:
            cur.close()
            conn.close()
            return jsonify({"error": "Sınıf bulunamadı"}), 404
        
        old_name = class_data['name']
        
        # Aynı isimde başka sınıf var mı kontrol et
        cur.execute("SELECT id FROM classes WHERE name = %s AND id != %s", (new_name, class_id))
        existing = cur.fetchone()
        
        if existing:
            cur.close()
            conn.close()
            return jsonify({"error": f"{new_name} adında başka bir sınıf zaten mevcut"}), 400
        
        # Sınıf adını güncelle
        cur.execute("UPDATE classes SET name = %s WHERE id = %s", (new_name, class_id))
        
        # Bu sınıftaki tüm öğrencilerin class_name bilgisini güncelle
        cur.execute("""
            UPDATE users 
            SET class_name = %s 
            WHERE class_name = %s AND role = 'student'
        """, (new_name, old_name))
        
        updated_students = cur.rowcount
        
        conn.commit()
        cur.close()
        conn.close()
        
        logger.info(f"Class renamed: {old_name} -> {new_name}, {updated_students} students updated")
        
        return jsonify({
            "success": True,
            "message": f"Sınıf adı {old_name} -> {new_name} olarak güncellendi",
            "old_name": old_name,
            "new_name": new_name,
            "updated_students": updated_students
        })
    
    except Exception as e:
        logger.error(f"Rename class error: {str(e)}")
        return jsonify({"error": f"Sınıf adı güncellenemedi: {str(e)}"}), 500

@app.route("/admin/classes/<int:class_id>/students", methods=["GET"])
@login_required
def get_class_students(class_id):
    """Bir sınıftaki öğrencileri listele"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Sınıf adını al
        cur.execute("SELECT name FROM classes WHERE id = %s", (class_id,))
        class_data = cur.fetchone()
        
        if not class_data:
            cur.close()
            conn.close()
            return jsonify({"error": "Sınıf bulunamadı"}), 404
        
        class_name = class_data['name']
        
        # Öğrencileri listele
        cur.execute("""
            SELECT id, username, full_name, class_name, created_at
            FROM users
            WHERE class_name = %s AND role = 'student'
            ORDER BY full_name
        """, (class_name,))
        students = cur.fetchall()
        
        cur.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "class_name": class_name,
            "students": [{
                "id": s['id'],
                "username": s['username'],
                "full_name": s['full_name'],
                "class_name": s['class_name'],
                "created_at": s['created_at'].isoformat() if s['created_at'] else None
            } for s in students]
        })
    
    except Exception as e:
        logger.error(f" Get class students error: {str(e)}")
        return jsonify({"error": f"Öğrenciler listelenemedi: {str(e)}"}), 500

@app.route("/admin/students/<int:student_id>/remove-from-class", methods=["PUT"])
@login_required
def remove_student_from_class(student_id):
    """Öğrenciyi sınıftan çıkar (class_name = NULL)"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Öğrenciyi kontrol et
        cur.execute("SELECT username, full_name, class_name FROM users WHERE id = %s AND role = 'student'", (student_id,))
        student = cur.fetchone()
        
        if not student:
            cur.close()
            conn.close()
            return jsonify({"error": "Öğrenci bulunamadı"}), 404
        
        # Sınıftan çıkar
        cur.execute("UPDATE users SET class_name = NULL WHERE id = %s", (student_id,))
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "message": f"{student['full_name']} ({student['username']}) sınıftan çıkarıldı"
        })
    
    except Exception as e:
        logger.error(f" Remove from class error: {str(e)}")
        return jsonify({"error": f"Öğrenci sınıftan çıkarılamadı: {str(e)}"}), 500

@app.route("/admin/students/<int:student_id>/change-class", methods=["PUT"])
@login_required
def change_student_class(student_id):
    """Öğrencinin sınıfını değiştir"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    data = request.get_json()
    new_class_name = data.get('new_class_name')
    
    if not new_class_name:
        return jsonify({"error": "Yeni sınıf adı gerekli"}), 400
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Yeni sınıf var mı kontrol et
        cur.execute("SELECT name FROM classes WHERE name = %s", (new_class_name,))
        if not cur.fetchone():
            cur.close()
            conn.close()
            return jsonify({"error": f"{new_class_name} sınıfı bulunamadı"}), 404
        
        # Öğrenciyi kontrol et
        cur.execute("SELECT username, full_name, class_name FROM users WHERE id = %s AND role = 'student'", (student_id,))
        student = cur.fetchone()
        
        if not student:
            cur.close()
            conn.close()
            return jsonify({"error": "Öğrenci bulunamadı"}), 404
        
        # Sınıfı değiştir
        cur.execute("UPDATE users SET class_name = %s WHERE id = %s", (new_class_name, student_id))
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "message": f"{student['full_name']} ({student['username']}) → {new_class_name} sınıfına taşındı"
        })
    
    except Exception as e:
        logger.error(f" Change class error: {str(e)}")
        return jsonify({"error": f"Sınıf değiştirilemedi: {str(e)}"}), 500

@app.route("/admin/students/<int:student_id>/reset-password", methods=["PUT"])
@login_required
def admin_reset_student_password(student_id):
    """Admin öğrenci şifresini değiştirir"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    data = request.get_json()
    new_password = data.get('new_password')
    
    if not new_password or len(new_password) < 6:
        return jsonify({"error": "Şifre en az 6 karakter olmalı"}), 400
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Öğrenciyi kontrol et
        cur.execute("SELECT username, full_name FROM users WHERE id = %s AND role = 'student'", (student_id,))
        student = cur.fetchone()
        
        if not student:
            cur.close()
            conn.close()
            return jsonify({"error": "Öğrenci bulunamadı"}), 404
        
        # Şifreyi değiştir
        hashed_password = generate_password_hash(new_password)
        cur.execute("UPDATE users SET password = %s WHERE id = %s", (hashed_password, student_id))
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "message": f"{student['full_name']} ({student['username']}) için şifre değiştirildi"
        })
    
    except Exception as e:
        logger.error(f" Reset password error: {str(e)}")
        return jsonify({"error": f"Şifre değiştirilemedi: {str(e)}"}), 500

# =============================================
# SİSTEM AYARLARI (ADMIN)
# =============================================

@app.route("/api/admin/stats", methods=["GET"])
@login_required
def get_admin_stats():
    """Admin dashboard istatistikleri"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Toplam kullanıcı sayısı
        cur.execute("SELECT COUNT(*) as count FROM users")
        total_users = cur.fetchone()['count']
        
        # Öğretmen sayısı
        cur.execute("SELECT COUNT(*) as count FROM users WHERE role = 'teacher'")
        total_teachers = cur.fetchone()['count']
        
        # Öğrenci sayısı
        cur.execute("SELECT COUNT(*) as count FROM users WHERE role = 'student'")
        total_students = cur.fetchone()['count']
        
        # Sınıf sayısı
        cur.execute("SELECT COUNT(*) as count FROM classes")
        total_classes = cur.fetchone()['count']
        
        # Bugün giriş yapan kullanıcı sayısı
        cur.execute("""
            SELECT COUNT(*) as count FROM users 
            WHERE last_login_at IS NOT NULL 
            AND DATE(last_login_at) = CURRENT_DATE
        """)
        today_logins = cur.fetchone()['count']
        
        # Aktif sınavlar (başlangıç zamanı geçmiş, bitiş zamanı geçmemiş veya bugün olan)
        cur.execute("""
            SELECT COUNT(*) as count FROM exams 
            WHERE start_time <= NOW() 
            AND (end_time IS NULL OR end_time >= NOW())
        """)
        active_exams = cur.fetchone()['count']
        
        # Bekleyen ödevler (teslim tarihi geçmemiş)
        cur.execute("""
            SELECT COUNT(*) as count FROM assignments 
            WHERE due_date >= NOW()
        """)
        pending_assignments = cur.fetchone()['count']
        
        # Aktif duyurular (son 7 günde oluşturulan)
        cur.execute("""
            SELECT COUNT(*) as count FROM announcements 
            WHERE created_at >= NOW() - INTERVAL '7 days'
        """)
        active_announcements = cur.fetchone()['count']
        
        cur.close()
        conn.close()
        
        # Dosya sayısı ve boyutu
        total_files = 0
        total_size = 0
        
        if os.path.exists(UPLOAD_DIR):
            for filename in os.listdir(UPLOAD_DIR):
                file_path = os.path.join(UPLOAD_DIR, filename)
                if os.path.isfile(file_path):
                    total_files += 1
                    total_size += os.path.getsize(file_path)
        
        storage_mb = round(total_size / (1024*1024), 2)
        
        return jsonify({
            "success": True,
            "total_users": total_users,
            "total_teachers": total_teachers,
            "total_students": total_students,
            "total_classes": total_classes,
            "total_files": total_files,
            "storage_mb": storage_mb,
            "today_logins": today_logins,
            "active_exams": active_exams,
            "pending_assignments": pending_assignments,
            "active_announcements": active_announcements
        })
    
    except Exception as e:
        logger.error(f"Admin stats error: {str(e)}")
        return jsonify({"error": "İstatistikler alınamadı"}), 500

@app.route("/api/admin/settings", methods=["GET"])
@login_required
def get_admin_settings():
    """Sistem ayarlarını getir"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Kullanıcı sayıları
        cur.execute("SELECT COUNT(*) as count FROM users WHERE role = 'admin'")
        admin_count = cur.fetchone()['count']
        
        cur.execute("SELECT COUNT(*) as count FROM users WHERE role = 'teacher'")
        teacher_count = cur.fetchone()['count']
        
        cur.execute("SELECT COUNT(*) as count FROM users WHERE role = 'student'")
        student_count = cur.fetchone()['count']
        
        # Sınıf sayısı
        cur.execute("SELECT COUNT(*) as count FROM classes")
        class_count = cur.fetchone()['count']
        
        # İçerik sayıları
        cur.execute("SELECT COUNT(*) as count FROM exams")
        exam_count = cur.fetchone()['count']
        
        cur.execute("SELECT COUNT(*) as count FROM assignments")
        assignment_count = cur.fetchone()['count']
        
        cur.execute("SELECT COUNT(*) as count FROM announcements")
        announcement_count = cur.fetchone()['count']
        
        cur.close()
        conn.close()
        
        # Veritabanı bilgileri
        db_info = {
            "host": os.environ.get('PGHOST', 'localhost'),
            "database": os.environ.get('PGDATABASE', 'postgres'),
            "port": os.environ.get('PGPORT', '5432')
        }
        
        return jsonify({
            "users": {
                "admin": admin_count,
                "teacher": teacher_count,
                "student": student_count,
                "total": admin_count + teacher_count + student_count
            },
            "classes": class_count,
            "content": {
                "exams": exam_count,
                "assignments": assignment_count,
                "announcements": announcement_count
            },
            "database": db_info,
            "upload_dir": UPLOAD_DIR,
            "allowed_extensions": ", ".join(ALLOWED_EXTENSIONS)
        })
    
    except Exception as e:
        logger.error(f"Settings error: {str(e)}")
        return jsonify({"error": f"Ayarlar alınamadı: {str(e)}"}), 500

# Admin - Duyuru Listele
@app.route("/admin/announcements/list", methods=["GET"])
@login_required
def admin_list_announcements():
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Öğretmen duyurularını getir
        cur.execute("""
            SELECT *, 'teacher' as announcement_type FROM teacher_announcements 
            ORDER BY created_at DESC
        """)
        teacher_announcements = cur.fetchall()
        
        # Ana sayfa duyurularını getir
        cur.execute("""
            SELECT *, 'public' as announcement_type FROM public_announcements 
            ORDER BY created_at DESC
        """)
        public_announcements = cur.fetchall()
        
        cur.close()
        conn.close()
        
        # Tüm duyuruları birleştir ve tarihe göre sırala
        all_announcements = list(teacher_announcements) + list(public_announcements)
        all_announcements.sort(key=lambda x: x['created_at'], reverse=True)
        
        return jsonify({"announcements": all_announcements})
    
    except Exception as e:
        logger.error(f"Admin announcements list error: {str(e)}")
        return jsonify({"error": f"Duyurular alınamadı: {str(e)}"}), 500

# Admin - Duyuru Oluştur
@app.route("/admin/announcements/create", methods=["POST"])
@login_required
def admin_create_announcement():
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        title = request.form.get('title')
        content = request.form.get('content')
        video_url = request.form.get('video_url')
        target_teachers = request.form.getlist('target_teachers[]')
        
        if not title or not content:
            return jsonify({"error": "Başlık ve içerik zorunludur"}), 400
        
        file_path = None
        if 'file' in request.files:
            file = request.files['file']
            if file and file.filename:
                filename = secure_filename(file.filename)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                unique_filename = f"announcement_{timestamp}_{filename}"
                file_path = os.path.join(UPLOAD_DIR, unique_filename)
                file.save(file_path)
                file_path = unique_filename
        
        target_teachers_str = ','.join(target_teachers) if target_teachers else None
        
        conn = get_db()
        cur = conn.cursor()
        
        cur.execute("""
            INSERT INTO teacher_announcements 
            (title, content, file_path, video_url, target_teachers, admin_id)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (title, content, file_path, video_url, target_teachers_str, current_user.id))
        
        conn.commit()
        cur.close()
        conn.close()
        
        # Yetkili kullanıcılar için bildirim gönder
        if can_send_notification(current_user):
            send_push_notification(
                title="📢 Yeni Öğretmen Duyurusu",
                message="Sizin için yeni bir duyuru yayınlandı. Detaylar için tıklayın.",
                url="https://ameo-alanya.com/teacher/announcements",
                target_role="teacher"
            )
        
        return jsonify({"success": True, "message": "Duyuru oluşturuldu"})
    
    except Exception as e:
        logger.error(f"Admin create announcement error: {str(e)}")
        return jsonify({"error": f"Duyuru oluşturulamadı: {str(e)}"}), 500

# ==================== DENEME TAKİP SİSTEMİ ====================

# HTML ROUTES
@app.route("/admin/practice-exams")
@login_required
def admin_practice_exams_page():
    if current_user.role != 'admin':
        return redirect('/')
    return render_template("admin_practice_exams.html")

@app.route("/admin/exam-calendar")
@login_required
def admin_exam_calendar_page():
    if current_user.role != 'admin':
        return redirect('/')
    return render_template("admin_exam_calendar.html")

@app.route("/admin/teacher-classes")
@login_required
def admin_teacher_classes_page():
    if current_user.role != 'admin':
        return redirect('/')
    return render_template("admin_teacher_classes.html")

@app.route("/teacher/practice-exams")
@login_required
def teacher_practice_exams_page():
    if current_user.role not in ['teacher', 'admin']:
        return redirect('/')
    return render_template("teacher_practice_exams.html")

@app.route("/teacher/ranked-lists")
@login_required
def teacher_ranked_lists_page():
    if current_user.role not in ['teacher', 'admin']:
        return redirect('/')
    return render_template("teacher_ranked_lists.html")

@app.route("/student/practice-exams")
@login_required
def student_practice_exams_page():
    if current_user.role not in ['student', 'admin']:
        return redirect('/')
    return render_template("student_practice_exams.html")

@app.route("/student/study-plan")
@login_required
def student_study_plan_page():
    if current_user.role not in ['student', 'admin']:
        return redirect('/')
    return render_template("student_study_plan.html")

@app.route("/student/question-analysis")
@login_required
def student_question_analysis_page():
    if current_user.role not in ['student', 'admin']:
        return redirect('/')
    return render_template("student_question_analysis.html")

@app.route("/student/daily-tracking")
@login_required
def student_daily_tracking_page():
    if current_user.role not in ['student', 'admin']:
        return redirect('/')
    return render_template("student_daily_tracking.html")

@app.route("/student/exam-calendar")
@login_required
def student_exam_calendar_page():
    if current_user.role not in ['student', 'admin']:
        return redirect('/')
    return render_template("student_exam_calendar.html")

@app.route("/leaderboards")
def leaderboards_page():
    """Yıldızlar ve Enler - Herkes görebilir (giriş yapmadan da erişilebilir)"""
    return render_template("leaderboards.html")

# API ENDPOINTS

# Excel şablon indirme (ESKİ - Tekli)
@app.route("/teacher/practice-exams/download-template")
@login_required
def download_practice_exam_template():
    if current_user.role not in ['teacher', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        template_path = os.path.join('static', 'templates', 'deneme_takip_sablonu.xlsx')
        if not os.path.exists(template_path):
            return jsonify({"error": "Şablon dosyası bulunamadı"}), 404
        response = send_file(template_path, as_attachment=True, download_name='deneme_takip_sablonu.xlsx')
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        return response
    except Exception as e:
        logger.error(f"Template download error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# Excel şablon indirme (YENİ - Toplu - Admin)
@app.route("/admin/practice-exams/download-bulk-template")
@login_required
def download_bulk_practice_exam_template():
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        template_path = os.path.join('static', 'templates', 'toplu_deneme_sablonu.xlsx')
        if not os.path.exists(template_path):
            return jsonify({"error": "Şablon dosyası bulunamadı"}), 404
        response = send_file(template_path, as_attachment=True, download_name='toplu_deneme_sablonu.xlsx')
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        return response
    except Exception as e:
        logger.error(f"Bulk template download error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# Excel şablon indirme (YENİ - Toplu - Teacher)
@app.route("/teacher/practice-exams/download-bulk-template")
@login_required
def download_bulk_practice_exam_template_teacher():
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        template_path = os.path.join('static', 'templates', 'toplu_deneme_sablonu.xlsx')
        if not os.path.exists(template_path):
            return jsonify({"error": "Şablon dosyası bulunamadı"}), 404
        response = send_file(template_path, as_attachment=True, download_name='toplu_deneme_sablonu.xlsx')
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        return response
    except Exception as e:
        logger.error(f"Teacher bulk template download error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# Excel yükleme (Akıllı Eşleşme: Ad Soyad VEYA Öğrenci No)
@app.route("/teacher/practice-exams/upload", methods=["POST"])
@login_required
def upload_practice_exams():
    logger.info(f"📤 Excel yükleme başladı - Kullanıcı: {current_user.username}")
    
    if current_user.role not in ['teacher', 'admin']:
        logger.warning(f"⚠️ Yetkisiz erişim denemesi: {current_user.username}")
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        if 'file' not in request.files:
            logger.warning("⚠️ Dosya seçilmedi - file key yok")
            return jsonify({"error": "Dosya seçilmedi"}), 400
        
        file = request.files['file']
        logger.info(f"📁 Dosya alındı: {file.filename}")
        
        if file.filename == '':
            logger.warning("⚠️ Dosya adı boş")
            return jsonify({"error": "Dosya seçilmedi"}), 400
        
        if not file.filename.endswith('.xlsx'):
            logger.warning(f"⚠️ Yanlış format: {file.filename}")
            return jsonify({"error": "Sadece .xlsx dosyası yüklenebilir"}), 400
        
        # Excel dosyasını oku
        import pandas as pd
        logger.info("📊 Excel dosyası okunuyor...")
        df = pd.read_excel(file, sheet_name='Deneme Takip Şablonu')
        logger.info(f"✅ Excel okundu - {len(df)} satır bulundu")
        
        conn = get_db()
        cur = conn.cursor()
        
        inserted_count = 0
        updated_count = 0
        error_rows = []
        matched_by_name = 0
        matched_by_no = 0
        
        for index, row in df.iterrows():
            try:
                full_name = str(row['Ad Soyad']).strip() if pd.notna(row.get('Ad Soyad')) else ''
                student_no = str(row['Öğrenci No']).strip() if pd.notna(row.get('Öğrenci No')) else ''
                exam_number = int(row['Deneme No'])
                
                student_id = None
                match_method = None
                
                # 1. Önce Ad Soyad ile dene
                if full_name:
                    cur.execute("SELECT id FROM users WHERE role = 'student' AND full_name = %s", (full_name,))
                    result = cur.fetchone()
                    if result:
                        student_id = result[0]
                        match_method = "Ad Soyad"
                        matched_by_name += 1
                
                # 2. Ad Soyad bulunamadıysa username (öğrenci no) ile dene
                if not student_id and student_no:
                    try:
                        cur.execute("SELECT id FROM users WHERE role = 'student' AND username = %s", (student_no,))
                        result = cur.fetchone()
                        if result:
                            student_id = result[0]
                            match_method = "Öğrenci No"
                            matched_by_no += 1
                    except Exception:
                        conn.rollback()
                
                # Hiçbiri eşleşmediyse hata
                if not student_id:
                    error_rows.append(f"Satır {index+2}: '{full_name}' sistemde bulunamadı")
                    continue
                
                # Önce var mı kontrol et
                cur.execute("""
                    SELECT id FROM practice_exams 
                    WHERE student_id = %s AND exam_number = %s
                """, (student_id, exam_number))
                
                existing = cur.fetchone()
                
                if existing:
                    # UPDATE
                    cur.execute("""
                        UPDATE practice_exams SET
                            turkce_dogru = %s, turkce_yanlis = %s, turkce_net = %s,
                            matematik_dogru = %s, matematik_yanlis = %s, matematik_net = %s,
                            fen_dogru = %s, fen_yanlis = %s, fen_net = %s,
                            sosyal_dogru = %s, sosyal_yanlis = %s, sosyal_net = %s,
                            ingilizce_dogru = %s, ingilizce_yanlis = %s, ingilizce_net = %s,
                            din_dogru = %s, din_yanlis = %s, din_net = %s,
                            lgs_score = %s
                        WHERE student_id = %s AND exam_number = %s
                    """, (
                        row['Türkçe Doğru'], row['Türkçe Yanlış'], row['Türkçe Net'],
                        row['Matematik Doğru'], row['Matematik Yanlış'], row['Matematik Net'],
                        row['Fen Doğru'], row['Fen Yanlış'], row['Fen Net'],
                        row['Sosyal Doğru'], row['Sosyal Yanlış'], row['Sosyal Net'],
                        row['İngilizce Doğru'], row['İngilizce Yanlış'], row['İngilizce Net'],
                        row['Din Doğru'], row['Din Yanlış'], row['Din Net'],
                        row['LGS Puanı'],
                        student_id, exam_number
                    ))
                    updated_count += 1
                else:
                    # INSERT
                    cur.execute("""
                        INSERT INTO practice_exams (
                            student_id, exam_number,
                            turkce_dogru, turkce_yanlis, turkce_net,
                            matematik_dogru, matematik_yanlis, matematik_net,
                            fen_dogru, fen_yanlis, fen_net,
                            sosyal_dogru, sosyal_yanlis, sosyal_net,
                            ingilizce_dogru, ingilizce_yanlis, ingilizce_net,
                            din_dogru, din_yanlis, din_net,
                            lgs_score
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        student_id, exam_number,
                        row['Türkçe Doğru'], row['Türkçe Yanlış'], row['Türkçe Net'],
                        row['Matematik Doğru'], row['Matematik Yanlış'], row['Matematik Net'],
                        row['Fen Doğru'], row['Fen Yanlış'], row['Fen Net'],
                        row['Sosyal Doğru'], row['Sosyal Yanlış'], row['Sosyal Net'],
                        row['İngilizce Doğru'], row['İngilizce Yanlış'], row['İngilizce Net'],
                        row['Din Doğru'], row['Din Yanlış'], row['Din Net'],
                        row['LGS Puanı']
                    ))
                    inserted_count += 1
                    conn.commit()
            
            except Exception as row_error:
                conn.rollback()
                error_rows.append(f"Satır {index+2}: {str(row_error)}")
                continue
        cur.close()
        conn.close()
        
        total_rows = len(df)
        success_count = inserted_count + updated_count
        fail_count = len(error_rows)
        
        if fail_count == 0:
            message = f"🎉 {total_rows} kişinin tamamı başarıyla yüklendi!"
        else:
            message = f"✅ {success_count} kişi başarıyla yüklendi, ❌ {fail_count} kişi yüklenemedi"
        
        message += f"\n📊 Detay: {inserted_count} yeni eklendi, {updated_count} güncellendi"
        message += f"\n🔍 Eşleşme: {matched_by_name} Ad Soyad, {matched_by_no} Öğrenci No ile"
        
        if error_rows:
            message += f"\n\n⚠️ Yüklenemeyen kayıtlar:"
            for err in error_rows[:10]:
                message += f"\n• {err}"
            if len(error_rows) > 10:
                message += f"\n... ve {len(error_rows) - 10} kayıt daha"
        
        logger.info(f"✅ Yükleme tamamlandı: {success_count}/{total_rows} başarılı")
        
        # Yetkili kullanıcılar için bildirim gönder
        if can_send_notification(current_user) and success_count > 0:
            # Öğrencilere bildirim
            send_push_notification(
                title="Deneme Sınavı Sonuçları Eklendi",
                message=f"{success_count} öğrencinin deneme sonucu yüklendi. Sonuçlarınızı kontrol edin!",
                url="https://ameo-alanya.com",
                target_role="student"
            )
            # Admin ise öğretmenlere de bildirim gönder
            if current_user.role == 'admin':
                send_push_notification(
                    title="Deneme Sonuçları Yüklendi (Admin)",
                    message=f"{success_count} öğrencinin deneme sonucu sisteme eklendi.",
                    url="https://ameo-alanya.com",
                    target_role="teacher"
                )
        
        return jsonify({
            "success": True,
            "message": message,
            "inserted": inserted_count,
            "updated": updated_count,
            "errors": len(error_rows)
        })
    
    except Exception as e:
        logger.error(f"Practice exam upload error: {str(e)}")
        return jsonify({"error": f"Yükleme hatası: {str(e)}"}), 500

# TOPLU Excel yükleme (Akıllı Eşleşme: Sınıf+Ad Soyad VEYA Öğrenci No)
@app.route("/admin/practice-exams/upload-bulk", methods=["POST"])
@login_required
def upload_bulk_practice_exams():
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        if 'file' not in request.files:
            return jsonify({"error": "Dosya seçilmedi"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "Dosya seçilmedi"}), 400
        
        if not file.filename.endswith('.xlsx'):
            return jsonify({"error": "Sadece .xlsx dosyası yüklenebilir"}), 400
        
        # Güncelleme modu: 'update' = son denemeyi güncelle, 'new' = yeni deneme ekle (varsayılan)
        update_mode = request.form.get('update_mode', 'new')
        
        # Excel dosyasını oku
        import pandas as pd
        df = pd.read_excel(file, sheet_name='Toplu Deneme Yükleme')
        
        conn = get_db()
        cur = conn.cursor()
        
        inserted_count = 0
        updated_count = 0
        skipped_count = 0
        error_rows = []
        not_found_students = []
        matched_by_name = 0
        matched_by_no = 0
        
        for index, row in df.iterrows():
            try:
                class_name = str(row['Sınıf']).strip() if pd.notna(row.get('Sınıf')) else ''
                full_name = str(row['Ad Soyad']).strip() if pd.notna(row.get('Ad Soyad')) else ''
                student_no = str(row['Öğrenci No']).strip() if pd.notna(row.get('Öğrenci No')) else ''
                
                # Deneme No: boş, 0 veya "Otomatik" ise otomatik hesaplanacak
                raw_exam_no = row.get('Deneme No')
                auto_exam_number = False
                
                # NaN veya boş kontrolü
                if pd.isna(raw_exam_no):
                    auto_exam_number = True
                    exam_number = 0
                else:
                    # Sayısal değer kontrolü (0, 0.0 gibi)
                    try:
                        numeric_val = float(raw_exam_no)
                        if numeric_val == 0:
                            auto_exam_number = True
                            exam_number = 0
                        else:
                            exam_number = int(numeric_val)
                            if exam_number < 1 or exam_number > 80:
                                error_rows.append(f"Satır {index+2}: Deneme No 1-80 arasında olmalı (Girilen: {exam_number})")
                                continue
                    except (ValueError, TypeError):
                        # String değer kontrolü ("Otomatik", "Auto" gibi)
                        str_val = str(raw_exam_no).strip().lower()
                        if str_val in ['', 'otomatik', 'auto']:
                            auto_exam_number = True
                            exam_number = 0
                        else:
                            error_rows.append(f"Satır {index+2}: Geçersiz Deneme No değeri: {raw_exam_no}")
                            continue
                
                student_id = None
                
                # 1. Önce Sınıf + Ad Soyad ile dene (tam eşleşme)
                if class_name and full_name:
                    cur.execute("""
                        SELECT id FROM users 
                        WHERE role = 'student' 
                        AND class_name = %s 
                        AND full_name = %s
                    """, (class_name, full_name))
                    result = cur.fetchone()
                    if result:
                        student_id = result[0]
                        matched_by_name += 1
                
                # 2. Sınıf olmadan sadece Ad Soyad ile dene
                if not student_id and full_name:
                    cur.execute("""
                        SELECT id FROM users 
                        WHERE role = 'student' 
                        AND full_name = %s
                    """, (full_name,))
                    result = cur.fetchone()
                    if result:
                        student_id = result[0]
                        matched_by_name += 1
                
                # 3. Bulunamadıysa student_no sütunu ile dene
                if not student_id and student_no:
                    try:
                        cur.execute("SELECT id FROM users WHERE role = 'student' AND student_no = %s", (student_no,))
                        result = cur.fetchone()
                        if result:
                            student_id = result[0]
                            matched_by_no += 1
                    except Exception:
                        conn.rollback()
                
                # 4. Hala bulunamadıysa username ile dene
                if not student_id and student_no:
                    try:
                        cur.execute("SELECT id FROM users WHERE role = 'student' AND username = %s", (student_no,))
                        result = cur.fetchone()
                        if result:
                            student_id = result[0]
                            matched_by_no += 1
                    except Exception:
                        conn.rollback()
                
                # Hiçbiri eşleşmediyse hata ver
                if not student_id:
                    error_rows.append(f"Satır {index+2}: '{full_name}' ({class_name}) sistemde bulunamadı")
                    not_found_students.append({"class_name": class_name, "name": full_name, "student_no": student_no})
                    continue
                
                # Öğrencinin mevcut en yüksek deneme numarasını al
                cur.execute("""
                    SELECT COALESCE(MAX(exam_number), 0) FROM practice_exams 
                    WHERE student_id = %s
                """, (student_id,))
                max_exam_number = cur.fetchone()[0]
                
                # Otomatik deneme numarası hesapla
                if auto_exam_number:
                    if update_mode == 'update' and max_exam_number > 0:
                        # Güncelleme modu: Son denemeyi güncelle
                        exam_number = max_exam_number
                    else:
                        # Yeni ekleme modu: Yeni deneme numarası ata
                        exam_number = max_exam_number + 1
                        if exam_number > 80:
                            error_rows.append(f"Satır {index+2}: '{full_name}' için maksimum 80 deneme sınırına ulaşıldı")
                            skipped_count += 1
                            continue
                
                # Bu öğrencinin kaç denemesi var kontrol et
                cur.execute("""
                    SELECT COUNT(*) FROM practice_exams 
                    WHERE student_id = %s
                """, (student_id,))
                
                exam_count = cur.fetchone()[0]
                
                # Önce bu deneme var mı kontrol et
                cur.execute("""
                    SELECT id FROM practice_exams 
                    WHERE student_id = %s AND exam_number = %s
                """, (student_id, exam_number))
                
                existing = cur.fetchone()
                
                # Doğru/Yanlış değerlerini al ve net hesapla (3 yanlış = 1 doğru götürür)
                def get_val(col_name, default=0):
                    val = row.get(col_name)
                    if pd.isna(val):
                        return default
                    try:
                        return float(val)
                    except:
                        return default
                
                def calc_net(dogru, yanlis):
                    return round(dogru - (yanlis / 3), 2)
                
                turkce_dogru = get_val('Türkçe Doğru')
                turkce_yanlis = get_val('Türkçe Yanlış')
                turkce_net = calc_net(turkce_dogru, turkce_yanlis)
                
                matematik_dogru = get_val('Matematik Doğru')
                matematik_yanlis = get_val('Matematik Yanlış')
                matematik_net = calc_net(matematik_dogru, matematik_yanlis)
                
                fen_dogru = get_val('Fen Doğru')
                fen_yanlis = get_val('Fen Yanlış')
                fen_net = calc_net(fen_dogru, fen_yanlis)
                
                sosyal_dogru = get_val('Sosyal Doğru')
                sosyal_yanlis = get_val('Sosyal Yanlış')
                sosyal_net = calc_net(sosyal_dogru, sosyal_yanlis)
                
                ingilizce_dogru = get_val('İngilizce Doğru')
                ingilizce_yanlis = get_val('İngilizce Yanlış')
                ingilizce_net = calc_net(ingilizce_dogru, ingilizce_yanlis)
                
                din_dogru = get_val('Din Doğru')
                din_yanlis = get_val('Din Yanlış')
                din_net = calc_net(din_dogru, din_yanlis)
                
                lgs_score = get_val('LGS Puanı', None)
                
                if existing:
                    # Güncelle
                    cur.execute("""
                        UPDATE practice_exams SET
                            turkce_dogru = %s, turkce_yanlis = %s, turkce_net = %s,
                            matematik_dogru = %s, matematik_yanlis = %s, matematik_net = %s,
                            fen_dogru = %s, fen_yanlis = %s, fen_net = %s,
                            sosyal_dogru = %s, sosyal_yanlis = %s, sosyal_net = %s,
                            ingilizce_dogru = %s, ingilizce_yanlis = %s, ingilizce_net = %s,
                            din_dogru = %s, din_yanlis = %s, din_net = %s,
                            lgs_score = %s
                        WHERE student_id = %s AND exam_number = %s
                    """, (
                        turkce_dogru, turkce_yanlis, turkce_net,
                        matematik_dogru, matematik_yanlis, matematik_net,
                        fen_dogru, fen_yanlis, fen_net,
                        sosyal_dogru, sosyal_yanlis, sosyal_net,
                        ingilizce_dogru, ingilizce_yanlis, ingilizce_net,
                        din_dogru, din_yanlis, din_net,
                        lgs_score,
                        student_id, exam_number
                    ))
                    updated_count += 1
                elif exam_count >= 80:
                    # Maksimum 80 deneme sınırı
                    error_rows.append(f"Satır {index+2}: '{full_name}' için maksimum 80 deneme sınırına ulaşıldı")
                    skipped_count += 1
                    continue
                else:
                    # Yeni kayıt ekle
                    cur.execute("""
                        INSERT INTO practice_exams (
                            student_id, exam_number,
                            turkce_dogru, turkce_yanlis, turkce_net,
                            matematik_dogru, matematik_yanlis, matematik_net,
                            fen_dogru, fen_yanlis, fen_net,
                            sosyal_dogru, sosyal_yanlis, sosyal_net,
                            ingilizce_dogru, ingilizce_yanlis, ingilizce_net,
                            din_dogru, din_yanlis, din_net,
                            lgs_score
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        student_id, exam_number,
                        turkce_dogru, turkce_yanlis, turkce_net,
                        matematik_dogru, matematik_yanlis, matematik_net,
                        fen_dogru, fen_yanlis, fen_net,
                        sosyal_dogru, sosyal_yanlis, sosyal_net,
                        ingilizce_dogru, ingilizce_yanlis, ingilizce_net,
                        din_dogru, din_yanlis, din_net,
                        lgs_score
                    ))
                    inserted_count += 1
                    conn.commit()
            
            except Exception as row_error:
                conn.rollback()
                error_rows.append(f"Satır {index+2}: {str(row_error)}")
                continue
        cur.close()
        conn.close()
        
        total_rows = len(df)
        success_count = inserted_count + updated_count
        fail_count = len(error_rows)
        
        if fail_count == 0:
            message = f"🎉 {total_rows} kişinin tamamı başarıyla yüklendi!"
        else:
            message = f"✅ {success_count} kişi başarıyla yüklendi, ❌ {fail_count} kişi yüklenemedi"
        
        message += f"\n📊 Detay: {inserted_count} yeni eklendi, {updated_count} güncellendi"
        message += f"\n🔍 Eşleşme: {matched_by_name} Sınıf+Ad Soyad, {matched_by_no} Öğrenci No ile"
        
        if skipped_count > 0:
            message += f"\n⏭️ {skipped_count} kayıt atlandı (80 deneme sınırı)"
        
        if error_rows:
            message += f"\n\n⚠️ Yüklenemeyen kayıtlar:"
            for err in error_rows[:10]:
                message += f"\n• {err}"
            if len(error_rows) > 10:
                message += f"\n... ve {len(error_rows) - 10} kayıt daha"
        
        # Bulunamayan öğrencileri tekil hale getir (aynı öğrenci birden fazla satırda olabilir)
        unique_not_found = []
        seen = set()
        for s in not_found_students:
            key = f"{s['class_name']}-{s['name']}"
            if key not in seen:
                seen.add(key)
                unique_not_found.append(s)
        
        # Yetkili kullanıcılar için bildirim gönder
        if can_send_notification(current_user) and success_count > 0:
            # Öğrencilere bildirim
            send_push_notification(
                title="Deneme Sınavı Sonuçları Eklendi",
                message=f"{success_count} öğrencinin deneme sonucu yüklendi. Sonuçlarınızı kontrol edin!",
                url="https://ameo-alanya.com",
                target_role="student"
            )
            # Admin ise öğretmenlere de bildirim gönder
            if current_user.role == 'admin':
                send_push_notification(
                    title="Deneme Sonuçları Yüklendi (Admin)",
                    message=f"{success_count} öğrencinin deneme sonucu sisteme eklendi.",
                    url="https://ameo-alanya.com",
                    target_role="teacher"
                )
        
        return jsonify({
            "success": True,
            "message": message,
            "inserted": inserted_count,
            "updated": updated_count,
            "skipped": skipped_count,
            "not_found_count": len(unique_not_found),
            "errors": len(error_rows),
            "not_found_students": unique_not_found
        })
    
    except Exception as e:
        import traceback
        logger.error(f"Bulk practice exam upload error: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({"error": f"Toplu yükleme hatası: {str(e)}"}), 500

# Toplu deneme silme (sınıf + deneme numarası)
@app.route("/admin/practice-exams/delete-bulk", methods=["POST"])
@login_required
def delete_bulk_practice_exams():
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        data = request.json
        class_name = data.get('class_name')
        exam_number = data.get('exam_number')
        
        if not class_name or not exam_number:
            return jsonify({"error": "Sınıf ve deneme numarası gerekli"}), 400
        
        try:
            exam_number = int(exam_number)
        except ValueError:
            return jsonify({"error": "Geçersiz deneme numarası"}), 400
        
        conn = get_db()
        cur = conn.cursor()
        
        # Silinecek kayıt sayısını bul
        cur.execute("""
            SELECT COUNT(*) FROM practice_exams pe
            JOIN users u ON pe.student_id = u.id
            WHERE u.class_name = %s AND pe.exam_number = %s
        """, (class_name, exam_number))
        count = cur.fetchone()[0]
        
        if count == 0:
            cur.close()
            conn.close()
            return jsonify({"error": f"{class_name} sınıfında {exam_number}. deneme bulunamadı"}), 404
        
        # Silme işlemi
        cur.execute("""
            DELETE FROM practice_exams 
            WHERE student_id IN (SELECT id FROM users WHERE class_name = %s)
            AND exam_number = %s
        """, (class_name, exam_number))
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "message": f"✅ {class_name} sınıfından {count} öğrencinin {exam_number}. denemesi silindi.",
            "deleted_count": count
        })
    
    except Exception as e:
        logger.error(f"Bulk delete error: {str(e)}")
        return jsonify({"error": f"Silme hatası: {str(e)}"}), 500

# Manuel veri girişi
@app.route("/teacher/practice-exams/add", methods=["POST"])
@login_required
def add_practice_exam():
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        data = request.json
        student_id = data.get('student_id')
        exam_number = data.get('exam_number')
        
        conn = get_db()
        cur = conn.cursor()
        
        # Önce var mı kontrol et
        cur.execute("""
            SELECT id FROM practice_exams 
            WHERE student_id = %s AND exam_number = %s
        """, (student_id, exam_number))
        
        existing = cur.fetchone()
        
        if existing:
            # UPDATE
            cur.execute("""
                UPDATE practice_exams SET
                    turkce_dogru = %s, turkce_yanlis = %s, turkce_net = %s,
                    matematik_dogru = %s, matematik_yanlis = %s, matematik_net = %s,
                    fen_dogru = %s, fen_yanlis = %s, fen_net = %s,
                    sosyal_dogru = %s, sosyal_yanlis = %s, sosyal_net = %s,
                    ingilizce_dogru = %s, ingilizce_yanlis = %s, ingilizce_net = %s,
                    din_dogru = %s, din_yanlis = %s, din_net = %s,
                    lgs_score = %s
                WHERE student_id = %s AND exam_number = %s
            """, (
                data['turkce_dogru'], data['turkce_yanlis'], data['turkce_net'],
                data['matematik_dogru'], data['matematik_yanlis'], data['matematik_net'],
                data['fen_dogru'], data['fen_yanlis'], data['fen_net'],
                data['sosyal_dogru'], data['sosyal_yanlis'], data['sosyal_net'],
                data['ingilizce_dogru'], data['ingilizce_yanlis'], data['ingilizce_net'],
                data['din_dogru'], data['din_yanlis'], data['din_net'],
                data['lgs_score'],
                student_id, exam_number
            ))
            message = "Deneme sonucu güncellendi"
        else:
            # INSERT
            cur.execute("""
                INSERT INTO practice_exams (
                    student_id, exam_number,
                    turkce_dogru, turkce_yanlis, turkce_net,
                    matematik_dogru, matematik_yanlis, matematik_net,
                    fen_dogru, fen_yanlis, fen_net,
                    sosyal_dogru, sosyal_yanlis, sosyal_net,
                    ingilizce_dogru, ingilizce_yanlis, ingilizce_net,
                    din_dogru, din_yanlis, din_net,
                    lgs_score
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                student_id, exam_number,
                data['turkce_dogru'], data['turkce_yanlis'], data['turkce_net'],
                data['matematik_dogru'], data['matematik_yanlis'], data['matematik_net'],
                data['fen_dogru'], data['fen_yanlis'], data['fen_net'],
                data['sosyal_dogru'], data['sosyal_yanlis'], data['sosyal_net'],
                data['ingilizce_dogru'], data['ingilizce_yanlis'], data['ingilizce_net'],
                data['din_dogru'], data['din_yanlis'], data['din_net'],
                data['lgs_score']
            ))
            message = "Deneme sonucu eklendi"
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": message})
    
    except Exception as e:
        logger.error(f"Add practice exam error: {str(e)}")
        return jsonify({"error": f"Kayıt hatası: {str(e)}"}), 500

# Öğrenci listesi (öğretmenin sınıfları)
@app.route("/teacher/api/practice-exams/students")
@login_required
def get_practice_exam_students():
    if current_user.role not in ['teacher', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Tüm öğrencileri getir (öğretmen tüm sınıfları görebilir)
        cur.execute("""
            SELECT id, username, full_name, class_name 
            FROM users 
            WHERE role = 'student'
            ORDER BY class_name, full_name
        """)
        students = cur.fetchall()
        
        cur.close()
        conn.close()
        
        return jsonify({"students": students})
    
    except Exception as e:
        logger.error(f"Get practice exam students error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# Öğrenci deneme sonuçları ve grafikler (öğretmen için)
@app.route("/teacher/api/practice-exams/student/<int:student_id>")
@login_required
def get_student_practice_exams(student_id):
    if current_user.role not in ['teacher', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Öğrenci bilgisi
        cur.execute("SELECT id, full_name, class_name FROM users WHERE id = %s", (student_id,))
        student = cur.fetchone()
        
        if not student:
            return jsonify({"error": "Öğrenci bulunamadı"}), 404
        
        # Tüm deneme sonuçları
        cur.execute("""
            SELECT * FROM practice_exams 
            WHERE student_id = %s 
            ORDER BY exam_number
        """, (student_id,))
        exams = cur.fetchall()
        
        cur.close()
        conn.close()
        
        return jsonify({
            "student": student,
            "exams": exams
        })
    
    except Exception as e:
        logger.error(f"Get student practice exams error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# TEACHER API - Sınıftaki tüm deneme verileri
@app.route("/teacher/api/practice-exams/class/<class_name>")
@login_required
def get_class_practice_exams_teacher(class_name):
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Sınıftaki tüm öğrencilerin tüm denemelerini al
        cur.execute("""
            SELECT 
                pe.*,
                u.full_name,
                u.id as student_id
            FROM practice_exams pe
            JOIN users u ON pe.student_id = u.id
            WHERE u.class_name = %s AND u.role = 'student'
            ORDER BY pe.exam_number, u.full_name
        """, (class_name,))
        
        exams = cur.fetchall()
        
        cur.close()
        conn.close()
        
        return jsonify({"exams": exams, "class_name": class_name})
    
    except Exception as e:
        logger.error(f"Get class practice exams (teacher) error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# TEACHER API - Sınıf ortalaması
@app.route("/teacher/api/practice-exams/class-average/<class_name>")
@login_required
def get_class_average_teacher(class_name):
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Sınıftaki öğrencileri bul
        cur.execute("""
            SELECT id FROM users 
            WHERE class_name = %s AND role = 'student'
        """, (class_name,))
        students = cur.fetchall()
        
        if not students:
            return jsonify({"exams": []})
        
        student_ids = [s['id'] for s in students]
        
        # Tüm denemelerin ortalamasını hesapla
        cur.execute("""
            SELECT 
                exam_number,
                AVG(turkce_net) as turkce_net,
                AVG(matematik_net) as matematik_net,
                AVG(fen_net) as fen_net,
                AVG(sosyal_net) as sosyal_net,
                AVG(ingilizce_net) as ingilizce_net,
                AVG(din_net) as din_net,
                AVG(lgs_score) as lgs_score,
                COUNT(*) as student_count
            FROM practice_exams
            WHERE student_id = ANY(%s)
            GROUP BY exam_number
            ORDER BY exam_number
        """, (student_ids,))
        
        exams = cur.fetchall()
        
        # Float değerleri düzelt
        for exam in exams:
            exam['turkce_net'] = round(float(exam['turkce_net']), 2) if exam['turkce_net'] else 0
            exam['matematik_net'] = round(float(exam['matematik_net']), 2) if exam['matematik_net'] else 0
            exam['fen_net'] = round(float(exam['fen_net']), 2) if exam['fen_net'] else 0
            exam['sosyal_net'] = round(float(exam['sosyal_net']), 2) if exam['sosyal_net'] else 0
            exam['ingilizce_net'] = round(float(exam['ingilizce_net']), 2) if exam['ingilizce_net'] else 0
            exam['din_net'] = round(float(exam['din_net']), 2) if exam['din_net'] else 0
            exam['lgs_score'] = round(float(exam['lgs_score']), 2) if exam['lgs_score'] else 0
        
        cur.close()
        conn.close()
        
        return jsonify({"exams": exams, "class_name": class_name})
    
    except Exception as e:
        logger.error(f"Get class average (teacher) error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# TEACHER API - Sınıf + Ders Bazlı Deneme Performans Analizi
@app.route("/teacher/api/practice-exams/class-subject/<class_name>/<subject>")
@login_required
def get_class_subject_performance(class_name, subject):
    if current_user.role not in ['teacher', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Ders alanı mapping
        subject_map = {
            'turkce': 'turkce_net',
            'matematik': 'matematik_net',
            'fen': 'fen_net',
            'sosyal': 'sosyal_net',
            'ingilizce': 'ingilizce_net',
            'din': 'din_net'
        }
        
        subject_names = {
            'turkce': 'Türkçe',
            'matematik': 'Matematik',
            'fen': 'Fen Bilimleri',
            'sosyal': 'Sosyal Bilgiler',
            'ingilizce': 'İngilizce',
            'din': 'Din Kültürü'
        }
        
        if subject not in subject_map:
            return jsonify({"error": "Geçersiz ders"}), 400
        
        net_column = subject_map[subject]
        subject_name = subject_names[subject]
        
        # Sınıftaki öğrencileri bul
        cur.execute("""
            SELECT id FROM users 
            WHERE class_name = %s AND role = 'student'
        """, (class_name,))
        students = cur.fetchall()
        
        if not students:
            return jsonify({"exams": [], "class_name": class_name, "subject": subject_name})
        
        student_ids = [s['id'] for s in students]
        
        # Her deneme için ders ortalamasını hesapla
        cur.execute(f"""
            SELECT 
                exam_number,
                AVG({net_column}) as net_avg,
                MIN({net_column}) as net_min,
                MAX({net_column}) as net_max,
                COUNT(*) as student_count
            FROM practice_exams
            WHERE student_id = ANY(%s)
            GROUP BY exam_number
            ORDER BY exam_number
        """, (student_ids,))
        
        exams = cur.fetchall()
        
        for exam in exams:
            exam['net_avg'] = round(float(exam['net_avg']), 2) if exam['net_avg'] else 0
            exam['net_min'] = round(float(exam['net_min']), 2) if exam['net_min'] else 0
            exam['net_max'] = round(float(exam['net_max']), 2) if exam['net_max'] else 0
        
        cur.close()
        conn.close()
        
        return jsonify({
            "exams": exams, 
            "class_name": class_name, 
            "subject": subject_name,
            "subject_key": subject
        })
    
    except Exception as e:
        logger.error(f"Get class subject performance error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# TEACHER API - Sınıf + Ders Bazlı PDF Raporu
@app.route("/teacher/api/practice-exams/report/class-subject/<class_name>/<subject>")
@login_required
def get_class_subject_report_pdf(class_name, subject):
    if current_user.role not in ['teacher', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        subject_map = {
            'turkce': ('turkce_net', 'Türkçe', 20),
            'matematik': ('matematik_net', 'Matematik', 20),
            'fen': ('fen_net', 'Fen Bilimleri', 20),
            'sosyal': ('sosyal_net', 'Sosyal Bilgiler', 10),
            'ingilizce': ('ingilizce_net', 'İngilizce', 10),
            'din': ('din_net', 'Din Kültürü', 10)
        }
        
        if subject not in subject_map:
            return jsonify({"error": "Geçersiz ders"}), 400
        
        net_column, subject_name, max_net = subject_map[subject]
        
        # Sınıftaki öğrencileri bul
        cur.execute("""
            SELECT id, full_name FROM users 
            WHERE class_name = %s AND role = 'student'
            ORDER BY full_name
        """, (class_name,))
        students = cur.fetchall()
        
        if not students:
            return jsonify({"error": "Öğrenci bulunamadı"}), 404
        
        student_ids = [s['id'] for s in students]
        student_names = {s['id']: s['full_name'] for s in students}
        
        # Her deneme için ders ortalamasını hesapla
        cur.execute(f"""
            SELECT 
                exam_number,
                AVG({net_column}) as net_avg,
                MIN({net_column}) as net_min,
                MAX({net_column}) as net_max,
                COUNT(*) as student_count
            FROM practice_exams
            WHERE student_id = ANY(%s)
            GROUP BY exam_number
            ORDER BY exam_number
        """, (student_ids,))
        
        exams = cur.fetchall()
        
        # Öğrenci bazlı detaylı veri
        cur.execute(f"""
            SELECT 
                pe.student_id,
                pe.exam_number,
                pe.{net_column} as net
            FROM practice_exams pe
            WHERE pe.student_id = ANY(%s)
            ORDER BY pe.exam_number
        """, (student_ids,))
        
        student_details = cur.fetchall()
        
        cur.close()
        conn.close()
        
        # PDF Oluştur
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30, bottomMargin=30)
        elements = []
        styles = getSampleStyleSheet()
        
        # Header
        elements.extend(create_pdf_header(styles))
        
        # Başlık
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Title'],
            fontSize=16,
            textColor=colors.HexColor('#1e3a5f'),
            alignment=TA_CENTER,
            fontName='DejaVuSans',
            spaceAfter=20
        )
        elements.append(Paragraph(f"{class_name} - {subject_name} Deneme Analizi", title_style))
        elements.append(Spacer(1, 10))
        
        # Tarih
        date_style = ParagraphStyle('DateStyle', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER, fontName='DejaVuSans')
        elements.append(Paragraph(f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y')}", date_style))
        elements.append(Spacer(1, 20))
        
        # Bar grafik oluştur
        if exams:
            fig, ax = plt.subplots(figsize=(10, 5))
            
            exam_numbers = [f"Deneme {e['exam_number']}" for e in exams]
            averages = [round(float(e['net_avg']), 2) if e['net_avg'] else 0 for e in exams]
            
            bar_colors = ['#667eea', '#764ba2', '#f59e0b', '#10b981', '#ef4444', '#3b82f6', '#8b5cf6', '#ec4899']
            colors_list = [bar_colors[i % len(bar_colors)] for i in range(len(exam_numbers))]
            
            bars = ax.bar(exam_numbers, averages, color=colors_list, edgecolor='white', linewidth=1.5)
            
            for bar, avg in zip(bars, averages):
                ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.2, 
                       f'{avg:.2f}', ha='center', va='bottom', fontsize=10, fontweight='bold')
            
            ax.set_ylabel('Net Ortalaması', fontsize=12)
            ax.set_xlabel('Denemeler', fontsize=12)
            ax.set_title(f'{class_name} - {subject_name} Deneme Performansı', fontsize=14, fontweight='bold')
            ax.set_ylim(0, max_net + 2)
            ax.axhline(y=max_net/2, color='red', linestyle='--', alpha=0.5, label=f'Hedef ({max_net/2})')
            ax.legend()
            
            plt.tight_layout()
            
            img_buffer = BytesIO()
            plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
            plt.close()
            img_buffer.seek(0)
            
            img = RLImage(img_buffer, width=450, height=225)
            elements.append(img)
            elements.append(Spacer(1, 20))
        
        # Özet Tablo
        summary_style = ParagraphStyle('Summary', parent=styles['Normal'], fontSize=11, fontName='DejaVuSans')
        elements.append(Paragraph("<b>Deneme Bazlı Özet</b>", summary_style))
        elements.append(Spacer(1, 10))
        
        table_data = [['Deneme No', 'Ortalama Net', 'En Düşük', 'En Yüksek', 'Öğrenci Sayısı']]
        for exam in exams:
            table_data.append([
                f"Deneme {exam['exam_number']}",
                f"{round(float(exam['net_avg']), 2) if exam['net_avg'] else 0}",
                f"{round(float(exam['net_min']), 2) if exam['net_min'] else 0}",
                f"{round(float(exam['net_max']), 2) if exam['net_max'] else 0}",
                str(exam['student_count'])
            ])
        
        table = Table(table_data, colWidths=[80, 80, 70, 70, 90])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'DejaVuSans'),
            ('FONTNAME', (0, 1), (-1, -1), 'DejaVuSans'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e2e8f0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')])
        ]))
        elements.append(table)
        
        # Footer
        elements.append(Spacer(1, 30))
        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER, textColor=colors.gray, fontName='DejaVuSans')
        elements.append(Paragraph(create_pdf_footer_text(), footer_style))
        
        doc.build(elements)
        buffer.seek(0)
        
        # ASCII olmayan karakterleri temizle (filename için)
        safe_subject = subject_name.replace('İ', 'I').replace('ı', 'i').replace('ğ', 'g').replace('ü', 'u').replace('ş', 's').replace('ö', 'o').replace('ç', 'c').replace('Ğ', 'G').replace('Ü', 'U').replace('Ş', 'S').replace('Ö', 'O').replace('Ç', 'C')
        safe_class = class_name.replace('İ', 'I').replace('ı', 'i').replace('ğ', 'g').replace('ü', 'u').replace('ş', 's').replace('ö', 'o').replace('ç', 'c').replace('Ğ', 'G').replace('Ü', 'U').replace('Ş', 'S').replace('Ö', 'O').replace('Ç', 'C')
        filename = f"{safe_class}_{safe_subject}_Deneme_Analizi_{datetime.now().strftime('%Y%m%d')}.pdf"
        
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'inline; filename="{filename}"'
        
        return response
    
    except Exception as e:
        logger.error(f"Class subject report PDF error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# TEACHER API - Sınıftaki öğrencilerin listesi
@app.route("/teacher/api/practice-exams/class-students/<class_name>")
@login_required
def get_class_students_for_practice_teacher(class_name):
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT id, username, full_name, class_name
            FROM users 
            WHERE class_name = %s AND role = 'student'
            ORDER BY full_name
        """, (class_name,))
        students = cur.fetchall()
        
        cur.close()
        conn.close()
        
        return jsonify({"students": students, "class_name": class_name})
    
    except Exception as e:
        logger.error(f"Get class students (teacher) error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# Öğrenci kendi deneme sonuçları
@app.route("/student/api/practice-exams/my-results")
@login_required
def get_my_practice_exams():
    # Admin student_id parametresiyle çağırabilir
    if current_user.role == 'admin':
        student_id = request.args.get('student_id')
        if not student_id:
            return jsonify({"error": "student_id parametresi gerekli"}), 400
        student_id = int(student_id)
    elif current_user.role == 'student':
        student_id = current_user.id
    else:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Önce öğrenci bilgisini al
        cur.execute("SELECT full_name FROM users WHERE id = %s", (student_id,))
        student = cur.fetchone()
        
        if not student:
            cur.close()
            conn.close()
            return jsonify({"error": "Öğrenci bulunamadı"}), 404
        
        cur.execute("""
            SELECT 
                pe.*,
                u.full_name
            FROM practice_exams pe
            JOIN users u ON pe.student_id = u.id
            WHERE pe.student_id = %s 
            ORDER BY pe.exam_number
        """, (student_id,))
        exams = cur.fetchall()
        
        cur.close()
        conn.close()
        
        return jsonify({"exams": exams})
    
    except Exception as e:
        logger.error(f"Get my practice exams error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# ADMIN API - Sınıftaki tüm deneme verileri
@app.route("/admin/api/practice-exams/class/<class_name>")
@login_required
def get_class_practice_exams(class_name):
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Sınıftaki tüm öğrencilerin tüm denemelerini al
        cur.execute("""
            SELECT 
                pe.*,
                u.full_name,
                u.id as student_id
            FROM practice_exams pe
            JOIN users u ON pe.student_id = u.id
            WHERE u.class_name = %s AND u.role = 'student'
            ORDER BY pe.exam_number, u.full_name
        """, (class_name,))
        
        exams = cur.fetchall()
        
        cur.close()
        conn.close()
        
        return jsonify({"exams": exams, "class_name": class_name})
    
    except Exception as e:
        logger.error(f"Get class practice exams error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# ADMIN API - Sınıf ortalaması
@app.route("/admin/api/practice-exams/class-average/<class_name>")
@login_required
def get_class_average(class_name):
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Sınıftaki öğrencileri bul
        cur.execute("""
            SELECT id FROM users 
            WHERE class_name = %s AND role = 'student'
        """, (class_name,))
        students = cur.fetchall()
        
        if not students:
            return jsonify({"exams": []})
        
        student_ids = [s['id'] for s in students]
        
        # Tüm denemelerin ortalamasını hesapla
        cur.execute("""
            SELECT 
                exam_number,
                AVG(turkce_net) as turkce_net,
                AVG(matematik_net) as matematik_net,
                AVG(fen_net) as fen_net,
                AVG(sosyal_net) as sosyal_net,
                AVG(ingilizce_net) as ingilizce_net,
                AVG(din_net) as din_net,
                AVG(lgs_score) as lgs_score,
                COUNT(*) as student_count
            FROM practice_exams
            WHERE student_id = ANY(%s)
            GROUP BY exam_number
            ORDER BY exam_number
        """, (student_ids,))
        
        exams = cur.fetchall()
        
        # Float değerleri düzelt
        for exam in exams:
            exam['turkce_net'] = round(float(exam['turkce_net']), 2) if exam['turkce_net'] else 0
            exam['matematik_net'] = round(float(exam['matematik_net']), 2) if exam['matematik_net'] else 0
            exam['fen_net'] = round(float(exam['fen_net']), 2) if exam['fen_net'] else 0
            exam['sosyal_net'] = round(float(exam['sosyal_net']), 2) if exam['sosyal_net'] else 0
            exam['ingilizce_net'] = round(float(exam['ingilizce_net']), 2) if exam['ingilizce_net'] else 0
            exam['din_net'] = round(float(exam['din_net']), 2) if exam['din_net'] else 0
            exam['lgs_score'] = round(float(exam['lgs_score']), 2) if exam['lgs_score'] else 0
        
        cur.close()
        conn.close()
        
        return jsonify({"exams": exams, "class_name": class_name})
    
    except Exception as e:
        logger.error(f"Get class average error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# ADMIN API - Deneme sil
@app.route("/admin/api/practice-exams/delete/<int:exam_id>", methods=["DELETE"])
@login_required
def delete_practice_exam(exam_id):
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # Denemeyi sil
        cur.execute("DELETE FROM practice_exams WHERE id = %s", (exam_id,))
        
        if cur.rowcount == 0:
            cur.close()
            conn.close()
            return jsonify({"error": "Deneme bulunamadı"}), 404
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Deneme başarıyla silindi"})
    
    except Exception as e:
        logger.error(f"Delete practice exam error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# TEACHER API - Deneme sil
@app.route("/teacher/api/practice-exams/delete/<int:exam_id>", methods=["DELETE"])
@login_required
def delete_practice_exam_teacher(exam_id):
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # Denemeyi sil
        cur.execute("DELETE FROM practice_exams WHERE id = %s", (exam_id,))
        
        if cur.rowcount == 0:
            cur.close()
            conn.close()
            return jsonify({"error": "Deneme bulunamadı"}), 404
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Deneme başarıyla silindi"})
    
    except Exception as e:
        logger.error(f"Delete practice exam (teacher) error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/admin/api/practice-exams/class-students/<class_name>")
@login_required
def get_class_students_for_practice(class_name):
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT id, username, full_name, class_name
            FROM users 
            WHERE class_name = %s AND role = 'student'
            ORDER BY full_name
        """, (class_name,))
        students = cur.fetchall()
        
        cur.close()
        conn.close()
        
        return jsonify({"students": students, "class_name": class_name})
    
    except Exception as e:
        logger.error(f"Get class students error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# ==================== LGS KOÇLUK SİSTEMİ API'LERİ ====================

# 1. HEDEF TAKİP SİSTEMİ API'LERİ

@app.route("/student/api/goals/set", methods=["POST"])
@login_required
def set_student_goal():
    """Öğrenci hedef LGS puanı belirler"""
    if current_user.role != 'student':
        return jsonify({"error": "Sadece öğrenciler hedef belirleyebilir"}), 403
    
    try:
        data = request.get_json()
        target_score = float(data.get('target_score', 0))
        target_date = data.get('target_date')
        
        if target_score < 0 or target_score > 500:
            return jsonify({"error": "Hedef puan 0-500 arasında olmalıdır"}), 400
        
        conn = get_db()
        cur = conn.cursor()
        
        # Upsert (varsa güncelle, yoksa ekle)
        cur.execute("""
            INSERT INTO student_goals (student_id, target_lgs_score, target_date, updated_at)
            VALUES (%s, %s, %s, CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul')
            ON CONFLICT (student_id) 
            DO UPDATE SET 
                target_lgs_score = EXCLUDED.target_lgs_score,
                target_date = EXCLUDED.target_date,
                updated_at = CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul'
        """, (current_user.id, target_score, target_date))
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Hedef başarıyla kaydedildi"})
    
    except Exception as e:
        logger.error(f"Set student goal error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/student/api/goals/get")
@login_required
def get_student_goal():
    """Öğrencinin hedefini ve hedefe kalan mesafeyi döndürür"""
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Hedef bilgisini al
        cur.execute("""
            SELECT target_lgs_score, target_date, created_at, updated_at
            FROM student_goals
            WHERE student_id = %s
        """, (current_user.id,))
        goal = cur.fetchone()
        
        # Mevcut en yüksek ve son LGS puanını al
        cur.execute("""
            SELECT 
                MAX(lgs_score) as highest_score,
                AVG(lgs_score) as avg_score,
                COUNT(*) as total_exams
            FROM practice_exams
            WHERE student_id = %s
        """, (current_user.id,))
        stats = cur.fetchone()
        
        # Son 5 denemenin ortalaması (gelişim trendi için)
        cur.execute("""
            SELECT AVG(lgs_score) as recent_avg
            FROM (
                SELECT lgs_score 
                FROM practice_exams 
                WHERE student_id = %s 
                ORDER BY exam_number DESC 
                LIMIT 5
            ) recent
        """, (current_user.id,))
        recent = cur.fetchone()
        
        cur.close()
        conn.close()
        
        result = {
            "has_goal": goal is not None,
            "goal": goal,
            "stats": stats,
            "recent_avg": float(recent['recent_avg']) if recent and recent['recent_avg'] else 0
        }
        
        # Hedefe kalan mesafe hesapla
        if goal and stats and stats['highest_score']:
            result['distance_to_goal'] = float(goal['target_lgs_score']) - float(stats['highest_score'])
            result['progress_percentage'] = min(100, (float(stats['highest_score']) / float(goal['target_lgs_score'])) * 100)
        
        return jsonify(result)
    
    except Exception as e:
        logger.error(f"Get student goal error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# 2. ZAYIF KONU TESPİTİ API

@app.route("/student/api/weak-subjects")
@login_required
def get_weak_subjects():
    """Öğrencinin zayıf olduğu konuları analiz eder"""
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Son 10 denemenin ortalamasını ders bazında al
        cur.execute("""
            SELECT 
                AVG(turkce_net) as turkce_avg,
                AVG(matematik_net) as matematik_avg,
                AVG(fen_net) as fen_avg,
                AVG(sosyal_net) as sosyal_avg,
                AVG(ingilizce_net) as ingilizce_avg,
                AVG(din_net) as din_avg,
                COUNT(*) as exam_count
            FROM (
                SELECT * FROM practice_exams 
                WHERE student_id = %s 
                ORDER BY exam_number DESC 
                LIMIT 10
            ) recent
        """, (current_user.id,))
        recent_stats = cur.fetchone()
        
        # İlk 10 denemenin ortalaması (gelişim yüzdesi için)
        cur.execute("""
            SELECT 
                AVG(turkce_net) as turkce_avg,
                AVG(matematik_net) as matematik_avg,
                AVG(fen_net) as fen_avg,
                AVG(sosyal_net) as sosyal_avg,
                AVG(ingilizce_net) as ingilizce_avg,
                AVG(din_net) as din_avg
            FROM (
                SELECT * FROM practice_exams 
                WHERE student_id = %s 
                ORDER BY exam_number ASC 
                LIMIT 10
            ) first
        """, (current_user.id,))
        first_stats = cur.fetchone()
        
        cur.close()
        conn.close()
        
        if not recent_stats or recent_stats['exam_count'] == 0:
            return jsonify({"subjects": [], "message": "Henüz deneme verisi yok"})
        
        # Dersleri analiz et
        subjects = [
            {"name": "Türkçe", "recent_avg": float(recent_stats['turkce_avg'] or 0), "max_net": 20, "first_avg": float(first_stats['turkce_avg'] or 0) if first_stats else 0},
            {"name": "Matematik", "recent_avg": float(recent_stats['matematik_avg'] or 0), "max_net": 20, "first_avg": float(first_stats['matematik_avg'] or 0) if first_stats else 0},
            {"name": "Fen", "recent_avg": float(recent_stats['fen_avg'] or 0), "max_net": 20, "first_avg": float(first_stats['fen_avg'] or 0) if first_stats else 0},
            {"name": "Sosyal", "recent_avg": float(recent_stats['sosyal_avg'] or 0), "max_net": 10, "first_avg": float(first_stats['sosyal_avg'] or 0) if first_stats else 0},
            {"name": "İngilizce", "recent_avg": float(recent_stats['ingilizce_avg'] or 0), "max_net": 10, "first_avg": float(first_stats['ingilizce_avg'] or 0) if first_stats else 0},
            {"name": "Din", "recent_avg": float(recent_stats['din_avg'] or 0), "max_net": 10, "first_avg": float(first_stats['din_avg'] or 0) if first_stats else 0}
        ]
        
        # Her ders için performans yüzdesi ve gelişim hesapla
        for subject in subjects:
            subject['performance_pct'] = round((subject['recent_avg'] / subject['max_net']) * 100, 1)
            if subject['first_avg'] > 0:
                subject['improvement_pct'] = round(((subject['recent_avg'] - subject['first_avg']) / subject['first_avg']) * 100, 1)
            else:
                subject['improvement_pct'] = 0
            
            # Zayıflık seviyesi belirleme
            if subject['performance_pct'] < 40:
                subject['level'] = 'zayif'
                subject['recommendation'] = f"{subject['name']} dersine daha fazla odaklanmalısın"
            elif subject['performance_pct'] < 60:
                subject['level'] = 'orta'
                subject['recommendation'] = f"{subject['name']} dersinde ilerleme kaydediyorsun, devam et"
            else:
                subject['level'] = 'iyi'
                subject['recommendation'] = f"{subject['name']} dersinde çok iyi gidiyorsun!"
        
        # En zayıf 3 dersi bul
        subjects_sorted = sorted(subjects, key=lambda x: x['performance_pct'])
        weak_subjects = [s for s in subjects_sorted if s['level'] == 'zayif'][:3]
        
        return jsonify({
            "subjects": subjects,
            "weak_subjects": weak_subjects,
            "exam_count": recent_stats['exam_count']
        })
    
    except Exception as e:
        logger.error(f"Get weak subjects error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# 3. ROZET/BAŞARI SİSTEMİ API

@app.route("/student/api/achievements")
@login_required
def get_student_achievements():
    """Öğrencinin kazandığı rozetleri döndürür"""
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Kazanılmış rozetler
        cur.execute("""
            SELECT achievement_type, achievement_name, description, earned_at
            FROM student_achievements
            WHERE student_id = %s
            ORDER BY earned_at DESC
        """, (current_user.id,))
        earned = cur.fetchall()
        
        # Deneme istatistikleri (yeni rozet kontrolü için)
        cur.execute("""
            SELECT 
                COUNT(*) as total_exams,
                MAX(lgs_score) as highest_score,
                AVG(lgs_score) as avg_score
            FROM practice_exams
            WHERE student_id = %s
        """, (current_user.id,))
        stats = cur.fetchone()
        
        # Gelişim hesaplama (ilk 5 vs son 5 deneme)
        cur.execute("""
            WITH first_5 AS (
                SELECT AVG(lgs_score) as avg FROM (
                    SELECT lgs_score FROM practice_exams 
                    WHERE student_id = %s 
                    ORDER BY exam_number ASC LIMIT 5
                ) f
            ),
            last_5 AS (
                SELECT AVG(lgs_score) as avg FROM (
                    SELECT lgs_score FROM practice_exams 
                    WHERE student_id = %s 
                    ORDER BY exam_number DESC LIMIT 5
                ) l
            )
            SELECT 
                first_5.avg as first_avg,
                last_5.avg as last_avg
            FROM first_5, last_5
        """, (current_user.id, current_user.id))
        improvement = cur.fetchone()
        
        # Yeni rozetleri kontrol et ve ekle
        new_achievements = []
        earned_types = {a['achievement_type'] for a in earned}
        
        # 10 Deneme rozeti
        if stats['total_exams'] >= 10 and 'exam_10' not in earned_types:
            cur.execute("""
                INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description)
                VALUES (%s, 'exam_10', '🎯 İlk 10 Deneme', '10 deneme sınavını tamamladın!')
            """, (current_user.id,))
            new_achievements.append('İlk 10 Deneme')
        
        # 25 Deneme rozeti
        if stats['total_exams'] >= 25 and 'exam_25' not in earned_types:
            cur.execute("""
                INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description)
                VALUES (%s, 'exam_25', '🔥 25 Deneme Uzmanı', '25 deneme sınavını tamamladın!')
            """, (current_user.id,))
            new_achievements.append('25 Deneme Uzmanı')
        
        # 50 Deneme rozeti
        if stats['total_exams'] >= 50 and 'exam_50' not in earned_types:
            cur.execute("""
                INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description)
                VALUES (%s, 'exam_50', '👑 Deneme Şampiyonu', 'Tüm 50 denemeyi tamamladın!')
            """, (current_user.id,))
            new_achievements.append('Deneme Şampiyonu')
        
        # 400+ Puan rozeti
        if stats['highest_score'] and stats['highest_score'] >= 400 and 'score_400' not in earned_types:
            cur.execute("""
                INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description)
                VALUES (%s, 'score_400', '⭐ 400 Puan Kulübü', '400+ LGS puanına ulaştın!')
            """, (current_user.id,))
            new_achievements.append('400 Puan Kulübü')
        
        # Gelişim rozeti (+50 puan artış)
        if improvement and improvement['first_avg'] and improvement['last_avg']:
            improvement_amount = float(improvement['last_avg']) - float(improvement['first_avg'])
            if improvement_amount >= 50 and 'improvement_50' not in earned_types:
                cur.execute("""
                    INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description)
                    VALUES (%s, 'improvement_50', '📈 Süper Gelişim', 'LGS puanını 50+ puan artırdın!')
                """, (current_user.id,))
                new_achievements.append('Süper Gelişim')
        
        conn.commit()
        
        # Güncel rozet listesini yeniden al
        if new_achievements:
            cur.execute("""
                SELECT achievement_type, achievement_name, description, earned_at
                FROM student_achievements
                WHERE student_id = %s
                ORDER BY earned_at DESC
            """, (current_user.id,))
            earned = cur.fetchall()
        
        cur.close()
        conn.close()
        
        return jsonify({
            "achievements": earned,
            "new_achievements": new_achievements,
            "stats": stats
        })
    
    except Exception as e:
        logger.error(f"Get achievements error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# 4. ÖĞRETMEN KOÇLUK NOTLARI API

@app.route("/teacher/api/practice-exams/add-note", methods=["POST"])
@login_required
def add_teacher_note():
    """Öğretmen belirli bir deneme için not ekler"""
    if current_user.role != 'teacher':
        return jsonify({"error": "Sadece öğretmenler not ekleyebilir"}), 403
    
    try:
        data = request.get_json()
        exam_id = data.get('exam_id')
        note = data.get('note', '').strip()
        
        if not exam_id:
            return jsonify({"error": "exam_id gerekli"}), 400
        
        conn = get_db()
        cur = conn.cursor()
        
        # Notu güncelle
        cur.execute("""
            UPDATE practice_exams 
            SET teacher_note = %s 
            WHERE id = %s
        """, (note if note else None, exam_id))
        
        if cur.rowcount == 0:
            cur.close()
            conn.close()
            return jsonify({"error": "Deneme bulunamadı"}), 404
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Not başarıyla kaydedildi"})
    
    except Exception as e:
        logger.error(f"Add teacher note error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# 5. DENEME TAKVİMİ API

@app.route("/api/exam-calendar/<class_name>")
@login_required
def get_exam_calendar(class_name):
    """Sınıfın deneme takvimini döndürür"""
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT 
                ec.*,
                u.full_name as created_by_name
            FROM exam_calendar ec
            LEFT JOIN users u ON ec.created_by = u.id
            WHERE ec.class_name = %s
            ORDER BY ec.exam_number
        """, (class_name,))
        calendar = cur.fetchall()
        
        cur.close()
        conn.close()
        
        return jsonify({"calendar": calendar})
    
    except Exception as e:
        logger.error(f"Get exam calendar error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/teacher/api/exam-calendar/create", methods=["POST"])
@login_required
def create_exam_calendar_entry():
    """Öğretmen deneme takvimi oluşturur"""
    if current_user.role != 'teacher':
        return jsonify({"error": "Sadece öğretmenler takvim oluşturabilir"}), 403
    
    try:
        data = request.get_json()
        class_name = data.get('class_name')
        exam_number = int(data.get('exam_number'))
        exam_date = data.get('exam_date')
        deadline_date = data.get('deadline_date')
        description = data.get('description', '').strip()
        
        if not all([class_name, exam_number, exam_date]):
            return jsonify({"error": "class_name, exam_number ve exam_date gerekli"}), 400
        
        conn = get_db()
        cur = conn.cursor()
        
        # Takvim girişi oluştur
        cur.execute("""
            INSERT INTO exam_calendar (class_name, exam_number, exam_date, deadline_date, description, created_by)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON CONFLICT (class_name, exam_number) 
            DO UPDATE SET 
                exam_date = EXCLUDED.exam_date,
                deadline_date = EXCLUDED.deadline_date,
                description = EXCLUDED.description
        """, (class_name, exam_number, exam_date, deadline_date, description, current_user.id))
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Takvim başarıyla kaydedildi"})
    
    except Exception as e:
        logger.error(f"Create exam calendar error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/teacher/api/exam-calendar/delete/<int:calendar_id>", methods=["DELETE"])
@login_required
def delete_exam_calendar_entry(calendar_id):
    """Öğretmen deneme takvimi girişini siler"""
    if current_user.role != 'teacher':
        return jsonify({"error": "Sadece öğretmenler takvim silebilir"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        cur.execute("DELETE FROM exam_calendar WHERE id = %s", (calendar_id,))
        
        if cur.rowcount == 0:
            cur.close()
            conn.close()
            return jsonify({"error": "Takvim girişi bulunamadı"}), 404
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Takvim girişi silindi"})
    
    except Exception as e:
        logger.error(f"Delete exam calendar error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# ADMIN SINAV TAKVİMİ API'LERİ

@app.route("/api/exam-calendar/month/<int:year>/<int:month>", methods=["GET"])
@login_required
def get_exams_by_month(year, month):
    """Admin - Belirtilen aydaki tüm sınavları getirir"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        from datetime import date
        
        # Ay başlangıcı ve bitişi
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, month + 1, 1)
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT 
                id,
                exam_date as date,
                exam_title as title,
                description,
                created_by,
                created_at
            FROM exam_calendar
            WHERE exam_date >= %s AND exam_date < %s
            ORDER BY exam_date
        """, (start_date, end_date))
        
        exams = cur.fetchall()
        
        cur.close()
        conn.close()
        
        return jsonify({"exams": exams})
    
    except Exception as e:
        logger.error(f"Get exams by month error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/exam-calendar/all", methods=["GET"])
@login_required
def get_all_exam_calendar():
    """Admin, Öğretmen ve Öğrenci - Tüm sınavları getirir (gelecek tarihli)"""
    if current_user.role not in ['admin', 'teacher', 'student']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Öğrenci ise sınıf filtreleme yap
        if current_user.role == 'student' and current_user.class_name:
            student_class = current_user.class_name.upper().strip()
            
            # Hem tire ile hem tiresiz formatı destekle
            # Database'de "7A, 7B" veya "7-A, 7-B" olabilir
            if '-' in student_class:
                class_with_hyphen = student_class
                class_without_hyphen = student_class.replace('-', '')
            else:
                class_without_hyphen = student_class
                if len(student_class) >= 2:
                    class_with_hyphen = student_class[:-1] + '-' + student_class[-1]
                else:
                    class_with_hyphen = student_class
            
            # Her iki format için de pattern'lar oluştur
            patterns = []
            for class_format in [class_with_hyphen, class_without_hyphen]:
                patterns.extend([
                    class_format,                      # Exact match
                    class_format + ', %',              # Start
                    '%, ' + class_format + ', %',      # Middle
                    '%, ' + class_format               # End
                ])
            
            cur.execute("""
                SELECT 
                    id,
                    TO_CHAR(exam_date, 'YYYY-MM-DD') as date,
                    description as title,
                    description,
                    classes,
                    created_by,
                    created_at
                FROM exam_calendar
                WHERE exam_date >= CURRENT_DATE
                    AND (
                        UPPER(classes) LIKE %s
                        OR UPPER(classes) = %s
                        OR UPPER(classes) LIKE %s
                        OR UPPER(classes) LIKE %s
                        OR UPPER(classes) LIKE %s
                        OR UPPER(classes) = %s
                        OR UPPER(classes) LIKE %s
                        OR UPPER(classes) LIKE %s
                        OR UPPER(classes) LIKE %s
                    )
                ORDER BY exam_date ASC
            """, ('%TÜMÜ%', *patterns))
            
            exams = cur.fetchall()
        else:
            # Admin/Öğretmen tüm sınavları görür
            cur.execute("""
                SELECT 
                    id,
                    TO_CHAR(exam_date, 'YYYY-MM-DD') as date,
                    description as title,
                    description,
                    classes,
                    created_by,
                    created_at
                FROM exam_calendar
                WHERE exam_date >= CURRENT_DATE
                ORDER BY exam_date ASC
            """)
            exams = cur.fetchall()
        cur.close()
        conn.close()
        
        return jsonify({"exams": exams})
    
    except Exception as e:
        logger.error(f"Get all exams error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/student/exam-calendar", methods=["GET"])
@login_required
def get_student_exam_calendar():
    """Öğrenci - Sadece kendi sınıfının sınavlarını getirir"""
    if current_user.role not in ['student', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        student_class = current_user.class_name.upper().strip()
        
        # Hem tire ile hem tiresiz formatı destekle
        if '-' in student_class:
            class_with_hyphen = student_class
            class_without_hyphen = student_class.replace('-', '')
        else:
            class_without_hyphen = student_class
            if len(student_class) >= 2:
                class_with_hyphen = student_class[:-1] + '-' + student_class[-1]
            else:
                class_with_hyphen = student_class
        
        # Her iki format için de pattern'lar
        patterns = []
        for class_format in [class_with_hyphen, class_without_hyphen]:
            patterns.extend([
                class_format,
                class_format + ', %',
                '%, ' + class_format + ', %',
                '%, ' + class_format
            ])
        
        cur.execute("""
            SELECT 
                id,
                TO_CHAR(exam_date, 'YYYY-MM-DD') as date,
                description as title,
                description,
                classes,
                created_by,
                created_at
            FROM exam_calendar
            WHERE exam_date >= CURRENT_DATE
                AND (
                    UPPER(classes) LIKE %s
                    OR UPPER(classes) = %s
                    OR UPPER(classes) LIKE %s
                    OR UPPER(classes) LIKE %s
                    OR UPPER(classes) LIKE %s
                    OR UPPER(classes) = %s
                    OR UPPER(classes) LIKE %s
                    OR UPPER(classes) LIKE %s
                    OR UPPER(classes) LIKE %s
                )
            ORDER BY exam_date ASC
        """, ('%TÜMÜ%', *patterns))
        
        exams = cur.fetchall()
        cur.close()
        conn.close()
        
        return jsonify({"exams": exams})
    
    except Exception as e:
        logger.error(f"Get student exams error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/exam-calendar", methods=["POST"])
@login_required
def create_exam_calendar():
    """Admin - Yeni sınav ekler"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        data = request.get_json()
        logger.info(f"Exam creation request: {data}")
        
        exam_date = data.get('date')
        description = data.get('description', '').strip()
        classes = data.get('classes', '')
        
        if not exam_date:
            logger.error("Missing exam_date")
            return jsonify({"error": "Sınav tarihi gerekli"}), 400
        
        if not description:
            logger.error("Missing description")
            return jsonify({"error": "Sınav adı gerekli"}), 400
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Sınav ekle - classes bilgisiyle
        logger.info(f"Inserting exam: date={exam_date}, desc={description}, classes={classes}")
        cur.execute("""
            INSERT INTO exam_calendar (exam_date, description, classes, created_by)
            VALUES (%s, %s, %s, %s)
            RETURNING id, exam_date as date, description as title, description, classes
        """, (exam_date, description, classes, current_user.id))
        
        new_exam = cur.fetchone()
        exam_id = new_exam['id']
        logger.info(f"Exam created successfully: ID={exam_id}")
        
        # Bildirim gönder - İlgili sınıflardaki öğrenciler ve öğretmenlere
        exam_date_formatted = datetime.strptime(exam_date, '%Y-%m-%d').strftime('%d %B %Y')
        notification_message = f"📅 Yeni sınav eklendi: {description} ({exam_date_formatted})"
        
        # Sınıfları parse et
        classes_list = [c.strip() for c in classes.split(',')]
        
        # TÜMÜ varsa tüm öğrenci ve öğretmenlere gönder
        if 'TÜMÜ' in classes_list or 'Tümü' in classes_list or 'tümü' in classes_list:
            # Tüm öğrencilere bildirim
            cur.execute("""
                INSERT INTO notifications (user_id, message, type, is_read)
                SELECT id, %s, 'exam', FALSE
                FROM users
                WHERE role IN ('student', 'teacher')
            """, (notification_message,))
        else:
            # Sadece belirtilen sınıflardaki öğrencilere ve öğretmenlere
            for class_name in classes_list:
                # Öğrencilere
                cur.execute("""
                    INSERT INTO notifications (user_id, message, type, is_read)
                    SELECT id, %s, 'exam', FALSE
                    FROM users
                    WHERE role = 'student' AND class_name = %s
                """, (notification_message, class_name.strip()))
            
            # Tüm öğretmenlere (sınıf fark etmeksizin)
            cur.execute("""
                INSERT INTO notifications (user_id, message, type, is_read)
                SELECT id, %s, 'exam', FALSE
                FROM users
                WHERE role = 'teacher'
            """, (notification_message,))
        
        conn.commit()
        cur.close()
        conn.close()
        
        # Push bildirim gönder - HERKESE (ana sayfa duyurusu gibi)
        try:
            logger.info(f"📅 Sınav takvimi push bildirimi gönderiliyor...")
            
            # Herkese bildirim (target_role olmadan - Total Subscriptions)
            result = send_push_notification(
                title="📅 Yeni Sınav Tarihi",
                message=f"Sınav takvimine yeni sınav eklendi: {description}",
                url="https://ameo-alanya.com"
            )
            logger.info(f"📅 Sınav takvimi bildirimi gönderildi: {result}")
        except Exception as notif_error:
            logger.error(f"📅 Sınav takvimi bildirimi gönderilemedi: {notif_error}")
        
        logger.info(f"Sınav eklendi ve bildirimler gönderildi: {description}")
        
        return jsonify({"success": True, "exam": new_exam})
    
    except Exception as e:
        logger.error(f"Create exam calendar error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/exam-calendar/<int:exam_id>", methods=["DELETE"])
@login_required
def delete_exam_calendar(exam_id):
    """Admin - Sınavı siler"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        cur.execute("DELETE FROM exam_calendar WHERE id = %s", (exam_id,))
        
        if cur.rowcount == 0:
            cur.close()
            conn.close()
            return jsonify({"error": "Sınav bulunamadı"}), 404
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Sınav silindi"})
    
    except Exception as e:
        logger.error(f"Delete exam calendar error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# DENEME SINAVI SORULARI API

def optimize_image_file(file_content, file_ext, max_size_mb=15):
    """
    Görüntü dosyalarını otomatik optimize eder
    - Hedef boyut: ~10-15 MB
    - Kalite: %85-90 (gözle fark edilmez)
    """
    try:
        # Dosya boyutunu kontrol et
        current_size_mb = len(file_content) / (1024 * 1024)
        
        # Zaten küçükse optimize etme
        if current_size_mb <= max_size_mb:
            return file_content
        
        # PIL ile görüntüyü aç
        img = PILImage.open(BytesIO(file_content))
        
        # RGBA'yı RGB'ye çevir (JPEG için)
        if img.mode == 'RGBA' and file_ext in ['jpg', 'jpeg']:
            img = img.convert('RGB')
        
        # Optimize edilmiş çıktı
        output = BytesIO()
        
        # Format belirle
        img_format = 'JPEG' if file_ext in ['jpg', 'jpeg'] else file_ext.upper()
        
        # Kalite ile kaydet (quality=85, optimize=True)
        if img_format == 'JPEG':
            img.save(output, format=img_format, quality=85, optimize=True)
        elif img_format == 'PNG':
            img.save(output, format=img_format, optimize=True, compress_level=6)
        elif img_format == 'WEBP':
            img.save(output, format=img_format, quality=85, method=6)
        else:
            img.save(output, format=img_format, optimize=True)
        
        optimized_content = output.getvalue()
        optimized_size_mb = len(optimized_content) / (1024 * 1024)
        
        # Boyut kontrolü
        if optimized_size_mb > max_size_mb:
            # Daha agresif sıkıştırma (quality=70)
            output = BytesIO()
            if img_format == 'JPEG':
                img.save(output, format=img_format, quality=70, optimize=True)
            elif img_format == 'WEBP':
                img.save(output, format=img_format, quality=70, method=6)
            else:
                img.save(output, format=img_format, optimize=True)
            optimized_content = output.getvalue()
        
        logger.info(f"Image optimized: {current_size_mb:.2f} MB → {len(optimized_content)/(1024*1024):.2f} MB")
        return optimized_content
    
    except Exception as e:
        logger.warning(f"Image optimization failed: {e}, using original")
        return file_content

@app.route("/api/admin/exam-questions", methods=["POST"])
@login_required
def upload_exam_questions():
    """Admin - Deneme sınavı sorularını yükler"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        exam_name = request.form.get('exam_name')
        file = request.files.get('file')
        
        if not exam_name or not file:
            return jsonify({"error": "Deneme adı ve dosya gerekli"}), 400
        
        # Dosya uzantısını kontrol et
        if '.' not in file.filename:
            return jsonify({"error": "Geçersiz dosya"}), 400
        
        file_ext = file.filename.rsplit('.', 1)[1].lower()
        
        # İzin verilen dosya türleri
        ALLOWED_EXTENSIONS = {'pdf', 'jpg', 'jpeg', 'png', 'gif', 'webp'}
        if file_ext not in ALLOWED_EXTENSIONS:
            return jsonify({"error": "Sadece PDF ve resim dosyaları (JPG, PNG, GIF) yüklenebilir"}), 400
        
        # Dosya boyutunu kontrol et (50 MB max)
        file.seek(0, 2)  # Dosya sonuna git
        file_size = file.tell()
        file.seek(0)  # Başa dön
        
        MAX_FILE_SIZE = 50 * 1024 * 1024  # 50 MB
        if file_size > MAX_FILE_SIZE:
            return jsonify({"error": "Dosya boyutu 50 MB'dan küçük olmalıdır"}), 400
        
        file_type = 'pdf' if file_ext == 'pdf' else 'image'
        
        # Dosya içeriğini oku
        file_content = file.read()
        
        # Görüntü dosyalarını otomatik optimize et
        if file_type == 'image':
            file_content = optimize_image_file(file_content, file_ext, max_size_mb=15)
        
        # Benzersiz dosya adı oluştur
        import uuid
        unique_filename = f"exam_questions/{uuid.uuid4()}_{file.filename}"
        
        # Object Storage'a yükle
        object_storage.upload_from_bytes(unique_filename, file_content)
        
        # Veritabanına kaydet
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            INSERT INTO exam_questions (exam_name, file_url, file_type, created_by)
            VALUES (%s, %s, %s, %s)
            RETURNING id, exam_name, file_type, created_at
        """, (exam_name, unique_filename, file_type, current_user.id))
        
        new_question = cur.fetchone()
        
        conn.commit()
        cur.close()
        conn.close()
        
        logger.info(f"Exam questions uploaded: {exam_name}")
        
        return jsonify({"success": True, "question": new_question}), 201
    
    except Exception as e:
        logger.error(f"Upload exam questions error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/admin/exam-questions", methods=["GET"])
@login_required
def get_exam_questions():
    """Admin - Tüm deneme sınavı sorularını getirir"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT id, exam_name, file_type, created_at
            FROM exam_questions
            ORDER BY created_at DESC
        """)
        
        questions = cur.fetchall()
        cur.close()
        conn.close()
        
        return jsonify({"questions": questions})
    
    except Exception as e:
        logger.error(f"Get exam questions error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/admin/exam-questions/<int:question_id>", methods=["DELETE"])
@login_required
def delete_exam_questions(question_id):
    """Admin - Deneme sınavı sorularını siler"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Dosya URL'ini al
        cur.execute("SELECT file_url FROM exam_questions WHERE id = %s", (question_id,))
        result = cur.fetchone()
        
        if not result:
            cur.close()
            conn.close()
            return jsonify({"error": "Soru dosyası bulunamadı"}), 404
        
        file_url = result['file_url']
        
        # Object Storage'dan sil
        try:
            object_storage.delete(file_url)
        except Exception as e:
            logger.warning(f"Object storage delete error: {e}")
        
        # Veritabanından sil
        cur.execute("DELETE FROM exam_questions WHERE id = %s", (question_id,))
        
        conn.commit()
        cur.close()
        conn.close()
        
        logger.info(f"Exam questions deleted: {question_id}")
        
        return jsonify({"success": True, "message": "Soru dosyası silindi"})
    
    except Exception as e:
        logger.error(f"Delete exam questions error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/exam-questions/<int:question_id>/file", methods=["GET"])
@login_required
def view_exam_questions_file(question_id):
    """Öğrenci/Öğretmen/Admin - Soru dosyasını görüntüler"""
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("SELECT file_url, file_type FROM exam_questions WHERE id = %s", (question_id,))
        result = cur.fetchone()
        
        cur.close()
        conn.close()
        
        if not result:
            return jsonify({"error": "Soru dosyası bulunamadı"}), 404
        
        file_url = result['file_url']
        file_type = result['file_type']
        
        # Object Storage'dan dosyayı al
        file_content = object_storage.download_as_bytes(file_url)
        
        # MIME type belirle
        mime_type = 'application/pdf' if file_type == 'pdf' else 'image/jpeg'
        
        from flask import Response
        return Response(file_content, mimetype=mime_type)
    
    except Exception as e:
        logger.error(f"View exam questions file error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/exam-questions", methods=["GET"])
@login_required
def list_all_exam_questions():
    """Öğrenci/Öğretmen/Admin - Tüm deneme sınavı sorularını listeler"""
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT id, exam_name, file_type, created_at
            FROM exam_questions
            ORDER BY exam_name
        """)
        
        questions = cur.fetchall()
        cur.close()
        conn.close()
        
        return jsonify({"questions": questions})
    
    except Exception as e:
        logger.error(f"List exam questions error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# ÖĞRETMEN ÖĞRENCİ LİSTESİ API

@app.route("/api/teacher/my-students")
@login_required
def get_teacher_my_students():
    """Öğretmenin erişebildiği öğrencileri döndürür"""
    if current_user.role not in ['teacher', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Admin tüm öğrencileri görebilir
        if current_user.role == 'admin':
            cur.execute("""
                SELECT 
                    id,
                    full_name,
                    class_name
                FROM users
                WHERE role = 'student'
                ORDER BY class_name, full_name
            """)
        else:
            # Öğretmen sadece kendi öğrencilerini görür
            cur.execute("""
                SELECT DISTINCT
                    u.id,
                    u.full_name,
                    u.class_name
                FROM users u
                WHERE u.role = 'student'
                AND (
                    u.id IN (SELECT student_id FROM teacher_students WHERE teacher_id = %s)
                    OR u.class_name IN (SELECT class_name FROM teacher_classes WHERE teacher_id = %s)
                )
                ORDER BY u.class_name, u.full_name
            """, (current_user.id, current_user.id))
        
        students = cur.fetchall()
        
        cur.close()
        conn.close()
        
        return jsonify({"students": students})
    
    except Exception as e:
        logger.error(f"Get teacher my students error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/teacher/my-classes")
@login_required
def get_teacher_my_classes():
    """Öğretmenin erişebildiği sınıfları döndürür"""
    if current_user.role not in ['teacher', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Admin tüm sınıfları görebilir
        if current_user.role == 'admin':
            cur.execute("""
                SELECT DISTINCT class_name
                FROM users
                WHERE role = 'student' AND class_name IS NOT NULL AND class_name != ''
                ORDER BY class_name
            """)
        else:
            # Öğretmen sadece yetkili olduğu sınıfları görebilir
            cur.execute("""
                SELECT DISTINCT u.class_name
                FROM users u
                WHERE u.role = 'student'
                AND u.class_name IS NOT NULL
                AND u.class_name != ''
                AND (
                    u.id IN (SELECT student_id FROM teacher_students WHERE teacher_id = %s)
                    OR u.class_name IN (SELECT class_name FROM teacher_classes WHERE teacher_id = %s)
                )
                ORDER BY u.class_name
            """, (current_user.id, current_user.id))
        
        classes = cur.fetchall()
        cur.close()
        conn.close()
        
        # class_name listesi döndür
        class_names = [c['class_name'] for c in classes]
        
        return jsonify({"classes": class_names})
    
    except Exception as e:
        logger.error(f"Get teacher my classes error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/teacher/class-students")
@login_required
def get_teacher_class_students():
    """Öğretmenin belirtilen sınıftaki öğrencilerini döndürür"""
    if current_user.role not in ['teacher', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        class_name = request.args.get('class')
        
        if not class_name:
            return jsonify({"error": "Sınıf adı gerekli"}), 400
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Admin tüm öğrencileri görebilir
        if current_user.role == 'admin':
            cur.execute("""
                SELECT 
                    id,
                    full_name,
                    class_name
                FROM users
                WHERE role = 'student' AND class_name = %s
                ORDER BY full_name
            """, (class_name,))
        else:
            # Öğretmen sadece yetkili olduğu sınıftaki öğrencileri görebilir
            cur.execute("""
                SELECT 
                    u.id,
                    u.full_name,
                    u.class_name
                FROM users u
                WHERE u.role = 'student'
                AND u.class_name = %s
                AND (
                    u.id IN (SELECT student_id FROM teacher_students WHERE teacher_id = %s)
                    OR u.class_name IN (SELECT class_name FROM teacher_classes WHERE teacher_id = %s)
                )
                ORDER BY u.full_name
            """, (class_name, current_user.id, current_user.id))
        
        students = cur.fetchall()
        cur.close()
        conn.close()
        
        return jsonify({"students": students})
    
    except Exception as e:
        logger.error(f"Get teacher class students error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# ÖĞRETMEN ÇALIŞMA PLANI API'LERİ

@app.route("/api/teacher-study-plan", methods=["POST"])
@login_required
def create_teacher_study_plan():
    """Öğretmen çalışma planı oluşturur - Tek öğrenci, çoklu öğrenci veya sınıf bazlı"""
    if current_user.role not in ['teacher', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        data = request.get_json()
        plan_date = data.get('date')
        subject = data.get('subject', '').strip()
        question_count = data.get('question_count')
        note = data.get('note', '').strip()
        
        # Seçim tipi: student_id (tek), student_ids (çoklu), class_name (sınıf)
        student_id = data.get('student_id')
        student_ids = data.get('student_ids', [])
        class_name = data.get('class_name', '').strip()
        
        # ID'leri integer'a dönüştür ve validate et
        try:
            if student_id:
                student_id = int(student_id)
            if student_ids:
                student_ids = [int(sid) for sid in student_ids if sid]
        except (ValueError, TypeError):
            return jsonify({"error": "Geçersiz öğrenci ID formatı"}), 400
        
        if not plan_date or not subject or not question_count:
            return jsonify({"error": "Tarih, ders ve soru sayısı gerekli"}), 400
        
        # En az bir hedef seçilmiş olmalı
        if not student_id and not student_ids and not class_name:
            return jsonify({"error": "En az bir öğrenci veya sınıf seçilmeli"}), 400
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Önce öğretmenin yetkili olduğu öğrencileri al
        if current_user.role == 'admin':
            # Admin tüm öğrencilere erişebilir
            cur.execute("""
                SELECT id FROM users WHERE role = 'student'
            """)
        else:
            # Öğretmen sadece kendi öğrencilerine erişebilir
            cur.execute("""
                SELECT DISTINCT u.id
                FROM users u
                WHERE u.role = 'student'
                AND (
                    u.id IN (SELECT student_id FROM teacher_students WHERE teacher_id = %s)
                    OR u.class_name IN (SELECT class_name FROM teacher_classes WHERE teacher_id = %s)
                )
            """, (current_user.id, current_user.id))
        
        authorized_student_ids = [s['id'] for s in cur.fetchall()]
        
        if not authorized_student_ids:
            cur.close()
            conn.close()
            return jsonify({"error": "Atanmış öğrenci bulunamadı"}), 404
        
        target_student_ids = []
        
        # Tek öğrenci
        if student_id:
            if student_id not in authorized_student_ids:
                cur.close()
                conn.close()
                return jsonify({"error": "Bu öğrenciye erişim yetkiniz yok"}), 403
            target_student_ids = [student_id]
        # Çoklu öğrenci
        elif student_ids:
            # Sadece yetkili öğrencileri filtrele
            target_student_ids = [sid for sid in student_ids if sid in authorized_student_ids]
            if not target_student_ids:
                cur.close()
                conn.close()
                return jsonify({"error": "Seçilen öğrencilere erişim yetkiniz yok"}), 403
        # Sınıf bazlı
        elif class_name:
            # Sınıftaki tüm öğrencileri al
            cur.execute("""
                SELECT id FROM users 
                WHERE role = 'student' AND class_name = %s
            """, (class_name,))
            
            students = cur.fetchall()
            class_student_ids = [s['id'] for s in students]
            
            # Sadece yetkili öğrencileri filtrele (intersection)
            target_student_ids = [sid for sid in class_student_ids if sid in authorized_student_ids]
            
            if not target_student_ids:
                cur.close()
                conn.close()
                return jsonify({"error": "Bu sınıfa erişim yetkiniz yok veya sınıfta yetkili öğrenci yok"}), 403
        
        if not target_student_ids:
            cur.close()
            conn.close()
            return jsonify({"error": "Hedef öğrenci bulunamadı"}), 404
        
        # Her öğrenci için plan oluştur - real-time authorization check ile
        created_count = 0
        for sid in target_student_ids:
            try:
                # INSERT öncesi real-time authorization check
                if current_user.role == 'admin':
                    # Admin tüm öğrencilere erişebilir
                    cur.execute("""
                        SELECT 1 FROM users 
                        WHERE id = %s AND role = 'student'
                    """, (sid,))
                else:
                    # Öğretmen: Bu öğrenci hala yetkili mi kontrol et
                    cur.execute("""
                        SELECT 1 FROM users u
                        WHERE u.id = %s AND u.role = 'student'
                        AND (
                            u.id IN (SELECT student_id FROM teacher_students WHERE teacher_id = %s)
                            OR u.class_name IN (SELECT class_name FROM teacher_classes WHERE teacher_id = %s)
                        )
                    """, (sid, current_user.id, current_user.id))
                
                if not cur.fetchone():
                    logger.warning(f"Yetkisiz plan oluşturma girişimi: teacher={current_user.id}, student={sid}")
                    continue
                
                # Yetki doğrulandı, planı oluştur
                cur.execute("""
                    INSERT INTO teacher_study_plan 
                    (student_id, teacher_id, plan_date, subject, question_count, note)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    ON CONFLICT (student_id, plan_date, subject) 
                    DO UPDATE SET 
                        question_count = EXCLUDED.question_count,
                        note = EXCLUDED.note,
                        updated_at = CURRENT_TIMESTAMP AT TIME ZONE 'Europe/Istanbul'
                """, (sid, current_user.id, plan_date, subject, question_count, note))
                created_count += 1
            except Exception as e:
                logger.warning(f"Plan oluşturulurken hata (öğrenci {sid}): {str(e)}")
                continue
        
        conn.commit()
        cur.close()
        conn.close()
        
        logger.info(f"Öğretmen {current_user.id} tarafından {created_count} plan oluşturuldu")
        
        return jsonify({
            "success": True, 
            "message": f"{created_count} öğrenci için plan oluşturuldu",
            "created_count": created_count
        })
    
    except Exception as e:
        logger.error(f"Create teacher study plan error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/teacher-study-plan/my-plans", methods=["GET"])
@login_required
def get_teacher_study_plans():
    """Öğretmenin oluşturduğu planları getirir"""
    if current_user.role not in ['teacher', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT 
                tsp.id,
                tsp.student_id,
                tsp.plan_date as date,
                tsp.subject,
                tsp.question_count,
                tsp.note,
                tsp.created_at,
                u.full_name as student_name,
                u.class_name
            FROM teacher_study_plan tsp
            JOIN users u ON tsp.student_id = u.id
            WHERE tsp.teacher_id = %s
            ORDER BY tsp.plan_date DESC, u.class_name, u.full_name
        """, (current_user.id,))
        
        plans = cur.fetchall()
        
        cur.close()
        conn.close()
        
        return jsonify({"plans": plans})
    
    except Exception as e:
        logger.error(f"Get teacher study plans error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/teacher-study-plan/<int:plan_id>", methods=["DELETE"])
@login_required
def delete_teacher_study_plan(plan_id):
    """Öğretmen planı siler"""
    if current_user.role not in ['teacher', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # Sadece kendi planlarını silebilir
        cur.execute("""
            DELETE FROM teacher_study_plan 
            WHERE id = %s AND teacher_id = %s
        """, (plan_id, current_user.id))
        
        if cur.rowcount == 0:
            cur.close()
            conn.close()
            return jsonify({"error": "Plan bulunamadı veya yetkiniz yok"}), 404
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Plan silindi"})
    
    except Exception as e:
        logger.error(f"Delete teacher study plan error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/student/study-plans", methods=["GET"])
@login_required
def get_student_study_plans():
    """Öğrencinin öğretmen tarafından atanan planlarını getirir"""
    if current_user.role not in ['student', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT 
                tsp.id,
                tsp.plan_date as date,
                tsp.subject,
                tsp.question_count,
                tsp.note,
                tsp.created_at,
                t.full_name as teacher_name
            FROM teacher_study_plan tsp
            JOIN users t ON tsp.teacher_id = t.id
            WHERE tsp.student_id = %s
            ORDER BY tsp.plan_date DESC
        """, (current_user.id,))
        
        plans = cur.fetchall()
        
        cur.close()
        conn.close()
        
        return jsonify({"plans": plans})
    
    except Exception as e:
        logger.error(f"Get student study plans error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# ==================== YENİ PDF ÇALIŞMA PROGRAMI API'LERİ ====================

@app.route("/api/study-plan-pdf/upload", methods=["POST"])
@login_required
def upload_study_plan_pdf():
    """PDF çalışma programı yükle - birden fazla sınıfa atama destekli"""
    if current_user.role not in ['teacher', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        title = request.form.get('title', '').strip()
        target_type = request.form.get('target_type', 'class')
        target_classes_json = request.form.get('target_classes', '[]')
        file = request.files.get('file')
        
        if not title:
            return jsonify({"error": "Başlık gerekli"}), 400
        
        if not file or file.filename == '':
            return jsonify({"error": "PDF dosyası gerekli"}), 400
        
        if not file.filename.lower().endswith('.pdf'):
            return jsonify({"error": "Sadece PDF dosyası yüklenebilir"}), 400
        
        # Birden fazla sınıf seçimi
        target_classes = []
        try:
            target_classes = json.loads(target_classes_json)
            if not target_classes:
                return jsonify({"error": "En az bir sınıf seçilmeli"}), 400
        except:
            return jsonify({"error": "Sınıf listesi geçersiz"}), 400
        
        # Dosyayı kaydet
        file_path = save_uploaded_file(file, "study_plans")
        
        conn = get_db()
        cur = conn.cursor()
        
        # Her sınıf için bir kayıt oluştur
        plan_ids = []
        for target_class in target_classes:
            cur.execute("""
                INSERT INTO study_plan_pdf (teacher_id, title, file_path, target_type, target_class, target_students)
                VALUES (%s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                current_user.id,
                title,
                file_path,
                'class',
                target_class,
                None
            ))
            plan_ids.append(cur.fetchone()[0])
        
        conn.commit()
        cur.close()
        conn.close()
        
        logger.info(f"✅ Çalışma programı yüklendi: {title} - {len(target_classes)} sınıfa by {current_user.username}")
        return jsonify({"success": True, "ids": plan_ids, "message": f"Çalışma programı {len(target_classes)} sınıfa yüklendi"})
    
    except Exception as e:
        logger.error(f"Upload study plan PDF error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/study-plan-pdf/list", methods=["GET"])
@login_required
def list_study_plan_pdfs():
    """Öğretmenin yüklediği PDF çalışma programlarını listele"""
    if current_user.role not in ['teacher', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT id, title, file_path, target_type, target_class, target_students, created_at
            FROM study_plan_pdf
            WHERE teacher_id = %s
            ORDER BY created_at DESC
        """, (current_user.id,))
        
        plans = cur.fetchall()
        cur.close()
        conn.close()
        
        # Öğrenci sayısını hesapla
        for plan in plans:
            if plan['target_type'] == 'students' and plan['target_students']:
                try:
                    students = json.loads(plan['target_students'])
                    plan['student_count'] = len(students)
                except:
                    plan['student_count'] = 0
            else:
                plan['student_count'] = 0
        
        return jsonify(plans)
    
    except Exception as e:
        logger.error(f"List study plan PDFs error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/study-plan-pdf/<int:plan_id>/view", methods=["GET"])
@login_required
def view_study_plan_pdf(plan_id):
    """PDF çalışma programını görüntüle/indir"""
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT sp.*, u.full_name as teacher_name
            FROM study_plan_pdf sp
            JOIN users u ON sp.teacher_id = u.id
            WHERE sp.id = %s
        """, (plan_id,))
        
        plan = cur.fetchone()
        cur.close()
        conn.close()
        
        if not plan:
            return jsonify({"error": "Plan bulunamadı"}), 404
        
        # Yetki kontrolü
        if current_user.role == 'student':
            student_class = current_user.class_name
            if plan['target_type'] == 'class' and plan['target_class'] != student_class:
                return jsonify({"error": "Bu plana erişim yetkiniz yok"}), 403
        
        file_path = plan['file_path']
        
        # get_file_data fonksiyonunu kullan (Object Storage veya yerel)
        file_data = get_file_data(file_path)
        
        if file_data:
            from flask import Response
            return Response(
                file_data,
                mimetype='application/pdf',
                headers={
                    'Content-Disposition': f'inline; filename="{plan["title"]}.pdf"',
                    'Content-Type': 'application/pdf'
                }
            )
        else:
            return jsonify({"error": "Dosya bulunamadı"}), 404
    
    except Exception as e:
        logger.error(f"View study plan PDF error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/study-plan-pdf/<int:plan_id>", methods=["DELETE"])
@login_required
def delete_study_plan_pdf(plan_id):
    """PDF çalışma programını sil"""
    if current_user.role not in ['teacher', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        cur.execute("""
            DELETE FROM study_plan_pdf
            WHERE id = %s AND teacher_id = %s
        """, (plan_id, current_user.id))
        
        if cur.rowcount == 0:
            cur.close()
            conn.close()
            return jsonify({"error": "Plan bulunamadı veya yetkiniz yok"}), 404
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Çalışma programı silindi"})
    
    except Exception as e:
        logger.error(f"Delete study plan PDF error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/student/study-plan-pdfs", methods=["GET"])
@login_required
def get_student_study_plan_pdfs():
    """Öğrenciye atanan PDF çalışma programlarını getir"""
    if current_user.role not in ['student', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        student_class = current_user.class_name
        student_id = current_user.id
        
        # Sınıfa veya öğrenciye özel atanan planları getir
        cur.execute("""
            SELECT 
                sp.id,
                sp.title,
                sp.file_path,
                sp.target_type,
                sp.target_class,
                sp.created_at,
                u.full_name as teacher_name
            FROM study_plan_pdf sp
            JOIN users u ON sp.teacher_id = u.id
            WHERE 
                (sp.target_type = 'class' AND sp.target_class = %s)
                OR (sp.target_type = 'students' AND sp.target_students::text LIKE %s)
            ORDER BY sp.created_at DESC
        """, (student_class, f'%{student_id}%'))
        
        plans = cur.fetchall()
        cur.close()
        conn.close()
        
        # target_students içinde gerçekten bu öğrenci var mı kontrol et
        filtered_plans = []
        for plan in plans:
            if plan['target_type'] == 'class':
                filtered_plans.append(plan)
            else:
                # Öğrenci ID kontrolü
                try:
                    cur2 = get_db().cursor()
                    cur2.execute("SELECT target_students FROM study_plan_pdf WHERE id = %s", (plan['id'],))
                    row = cur2.fetchone()
                    cur2.close()
                    if row and row[0]:
                        student_ids = json.loads(row[0])
                        if student_id in student_ids:
                            filtered_plans.append(plan)
                except:
                    pass
        
        return jsonify(filtered_plans)
    
    except Exception as e:
        logger.error(f"Get student study plan PDFs error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# ÇALIŞMA PLANI PDF API'LERİ (ESKİ)

def get_week_boundaries(date_str=None):
    """Haftanın başlangıç ve bitiş tarihlerini hesaplar (Pazartesi-Pazar)"""
    if date_str:
        try:
            target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return None, None
    else:
        target_date = datetime.now().date()
    
    # Haftanın başı (Pazartesi)
    days_since_monday = target_date.weekday()
    week_start = target_date - timedelta(days=days_since_monday)
    
    # Haftanın sonu (Pazar)
    week_end = week_start + timedelta(days=6)
    
    return week_start, week_end

def build_study_plan_pdf(plans, meta):
    """Çalışma planı PDF'i oluşturur - Türkçe karakter desteği ile"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=50, bottomMargin=50)
    
    styles = getSampleStyleSheet()
    
    # Türkçe karakter desteği için DejaVuSans font kullan
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName='DejaVuSans-Bold',
        fontSize=20,
        textColor=colors.HexColor('#10b981'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    story = []
    
    # Başlık
    story.append(Paragraph(meta['title'], title_style))
    
    # Tarih aralığı
    date_style = ParagraphStyle(
        'DateStyle',
        parent=styles['Normal'],
        fontName='DejaVuSans',
        fontSize=12,
        textColor=colors.HexColor('#666666'),
        spaceAfter=20,
        alignment=TA_CENTER
    )
    story.append(Paragraph(meta['date_range'], date_style))
    
    story.append(Spacer(1, 20))
    
    if not plans:
        no_data_style = ParagraphStyle(
            'NoData',
            parent=styles['Normal'],
            fontName='DejaVuSans',
            fontSize=14,
            textColor=colors.HexColor('#999999'),
            alignment=TA_CENTER
        )
        story.append(Paragraph("Bu hafta için plan bulunamadı.", no_data_style))
    else:
        # Tablo başlıkları
        table_data = [['Tarih', 'Ders', 'Soru Sayısı', 'Not']]
        
        # Plan verileri
        for plan in plans:
            plan_date = plan['date'].strftime('%d.%m.%Y') if isinstance(plan['date'], datetime) else str(plan['date'])
            subject = plan['subject'] or '-'
            question_count = str(plan['question_count']) if plan['question_count'] else '0'
            note = plan['note'][:50] + '...' if plan['note'] and len(plan['note']) > 50 else (plan['note'] or '-')
            
            table_data.append([plan_date, subject, question_count, note])
        
        # Tablo oluştur - DejaVuSans font kullan
        table = Table(table_data, colWidths=[2*inch, 2*inch, 1.5*inch, 2.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'DejaVuSans-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'DejaVuSans'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ]))
        
        story.append(table)
        
        # Özet bilgi
        story.append(Spacer(1, 30))
        summary_style = ParagraphStyle(
            'Summary',
            parent=styles['Normal'],
            fontName='DejaVuSans',
            fontSize=11,
            textColor=colors.HexColor('#666666')
        )
        story.append(Paragraph(f"Toplam {len(plans)} plan bulunmaktadır.", summary_style))
    
    # Footer
    story.append(Spacer(1, 50))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontName='DejaVuSans',
        fontSize=9,
        textColor=colors.HexColor('#999999'),
        alignment=TA_CENTER
    )
    story.append(Paragraph("AMEO - Okul Yönetim Sistemi", footer_style))
    story.append(Paragraph(f"Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}", footer_style))
    
    doc.build(story)
    buffer.seek(0)
    return buffer

@app.route("/api/student/study-plan/pdf", methods=["GET"])
@login_required
def download_student_study_plan_pdf():
    """Öğrencinin aktif hafta çalışma planını PDF olarak indirir"""
    if current_user.role not in ['student', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        # Aktif haftayı hesapla
        week_start, week_end = get_week_boundaries()
        
        if not week_start or not week_end:
            return jsonify({"error": "Tarih hesaplanamadı"}), 400
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Haftalık planları çek
        cur.execute("""
            SELECT 
                tsp.plan_date as date,
                tsp.subject,
                tsp.question_count,
                tsp.note,
                t.full_name as teacher_name
            FROM teacher_study_plan tsp
            JOIN users t ON tsp.teacher_id = t.id
            WHERE tsp.student_id = %s
            AND tsp.plan_date BETWEEN %s AND %s
            ORDER BY tsp.plan_date ASC
        """, (current_user.id, week_start, week_end))
        
        plans = cur.fetchall()
        
        cur.close()
        conn.close()
        
        # PDF meta bilgileri
        meta = {
            'title': f'Haftalık Çalışma Planı - {current_user.full_name}',
            'date_range': f'{week_start.strftime("%d.%m.%Y")} - {week_end.strftime("%d.%m.%Y")}'
        }
        
        # PDF oluştur
        pdf_buffer = build_study_plan_pdf(plans, meta)
        
        filename = f"calisma_plani_{week_start.strftime('%Y%m%d')}.pdf"
        
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=False,
            download_name=filename
        )
    
    except Exception as e:
        logger.error(f"Student PDF generation error: {str(e)}")
        return jsonify({"error": "PDF oluşturulamadı"}), 500

@app.route("/api/teacher/study-plan/pdf", methods=["GET"])
@login_required
def download_teacher_study_plan_pdf():
    """Öğretmenin seçtiği hafta çalışma planlarını PDF olarak indirir"""
    if current_user.role not in ['teacher', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    # Teacher role kontrolü - admin için farklı davranış
    if current_user.role == 'admin':
        teacher_id = request.args.get('teacher_id')
        if not teacher_id:
            return jsonify({"error": "Admin için teacher_id parametresi gerekli"}), 400
    else:
        teacher_id = current_user.id
    
    try:
        # Hafta parametresini al
        week_param = request.args.get('week')
        week_start, week_end = get_week_boundaries(week_param)
        
        if not week_start or not week_end:
            return jsonify({"error": "Geçersiz tarih formatı (YYYY-MM-DD bekleniyor)"}), 400
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Öğretmen bilgisini al
        cur.execute("SELECT full_name FROM users WHERE id = %s", (teacher_id,))
        teacher = cur.fetchone()
        if not teacher:
            return jsonify({"error": "Öğretmen bulunamadı"}), 404
        
        # Öğretmenin o haftaki tüm planları
        cur.execute("""
            SELECT 
                tsp.plan_date as date,
                tsp.subject,
                tsp.question_count,
                tsp.note,
                s.full_name as student_name,
                COUNT(*) OVER (PARTITION BY tsp.plan_date) as plans_per_day
            FROM teacher_study_plan tsp
            JOIN users s ON tsp.student_id = s.id
            WHERE tsp.teacher_id = %s
            AND tsp.plan_date BETWEEN %s AND %s
            ORDER BY tsp.plan_date ASC, s.full_name ASC
        """, (teacher_id, week_start, week_end))
        
        plans = cur.fetchall()
        
        cur.close()
        conn.close()
        
        # PDF meta bilgileri
        meta = {
            'title': f'Haftalık Çalışma Planları - {teacher["full_name"]}',
            'date_range': f'{week_start.strftime("%d.%m.%Y")} - {week_end.strftime("%d.%m.%Y")}'
        }
        
        # PDF oluştur
        pdf_buffer = build_study_plan_pdf(plans, meta)
        
        filename = f"ogretmen_planlari_{week_start.strftime('%Y%m%d')}.pdf"
        
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=False,
            download_name=filename
        )
    
    except Exception as e:
        logger.error(f"Teacher PDF generation error: {str(e)}")
        return jsonify({"error": "PDF oluşturulamadı"}), 500

# SORU ANALİZİ API'LERİ

@app.route("/api/question-asks/practice-exams/my-exams", methods=["GET"])
@login_required
def get_student_practice_exams_for_questions():
    """Öğrencinin deneme sınavlarını soru analizi için döndürür"""
    if current_user.role not in ['student', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Öğrencinin tüm deneme sınavları
        cur.execute("""
            SELECT 
                id,
                exam_number,
                CAST(lgs_score AS FLOAT) as lgs_score,
                CAST((turkce_net + matematik_net + fen_net + sosyal_net + ingilizce_net + din_net) AS FLOAT) as total_net,
                created_at
            FROM practice_exams
            WHERE student_id = %s
            ORDER BY created_at DESC
        """, (current_user.id,))
        
        exams = cur.fetchall()
        
        cur.close()
        conn.close()
        
        return jsonify({"exams": exams})
    
    except Exception as e:
        logger.error(f"Get student practice exams for questions error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/question-asks/teachers/my-teachers", methods=["GET"])
@login_required  
def get_student_teachers_for_questions():
    """Öğrencinin öğretmenlerini döndürür"""
    if current_user.role not in ['student', 'admin']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Öğrencinin sınıfının öğretmenleri
        cur.execute("""
            SELECT DISTINCT
                u.id,
                u.full_name
            FROM users u
            INNER JOIN teacher_classes tc ON tc.teacher_id = u.id
            WHERE u.role = 'teacher'
            AND tc.class_name = (SELECT class_name FROM users WHERE id = %s)
            ORDER BY u.full_name
        """, (current_user.id,))
        
        teachers = cur.fetchall()
        
        cur.close()
        conn.close()
        
        return jsonify({"teachers": teachers})
    
    except Exception as e:
        logger.error(f"Get student teachers for questions error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# 6. İSTATİSTİKSEL ANALİZ API

@app.route("/student/api/class-comparison")
@login_required
def get_class_comparison():
    """Öğrencinin performansını sınıf ortalamasıyla karşılaştırır"""
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Öğrenci bilgisi
        cur.execute("SELECT class_name FROM users WHERE id = %s", (current_user.id,))
        user_info = cur.fetchone()
        
        if not user_info or not user_info['class_name']:
            return jsonify({"error": "Sınıf bilgisi bulunamadı"}), 404
        
        class_name = user_info['class_name']
        
        # Öğrencinin ortalama puanı
        cur.execute("""
            SELECT 
                AVG(lgs_score) as my_avg,
                MAX(lgs_score) as my_max,
                COUNT(*) as my_total
            FROM practice_exams
            WHERE student_id = %s
        """, (current_user.id,))
        my_stats = cur.fetchone()
        
        # Sınıf ortalaması
        cur.execute("""
            SELECT 
                AVG(pe.lgs_score) as class_avg,
                MAX(pe.lgs_score) as class_max,
                COUNT(DISTINCT pe.student_id) as student_count
            FROM practice_exams pe
            JOIN users u ON pe.student_id = u.id
            WHERE u.class_name = %s AND u.role = 'student'
        """, (class_name,))
        class_stats = cur.fetchone()
        
        # Sınıf sıralaması (son 5 deneme ortalamasına göre)
        cur.execute("""
            WITH student_avgs AS (
                SELECT 
                    pe.student_id,
                    u.full_name,
                    AVG(pe.lgs_score) as avg_score
                FROM (
                    SELECT * FROM practice_exams 
                    WHERE student_id IN (
                        SELECT id FROM users WHERE class_name = %s AND role = 'student'
                    )
                    ORDER BY exam_number DESC
                ) pe
                JOIN users u ON pe.student_id = u.id
                GROUP BY pe.student_id, u.full_name
                HAVING COUNT(*) >= 1
            )
            SELECT 
                ROW_NUMBER() OVER (ORDER BY avg_score DESC) as rank,
                student_id,
                full_name,
                avg_score
            FROM student_avgs
        """, (class_name,))
        rankings = cur.fetchall()
        
        # Öğrencinin sırası
        my_rank = None
        for r in rankings:
            if r['student_id'] == current_user.id:
                my_rank = r['rank']
                break
        
        cur.close()
        conn.close()
        
        return jsonify({
            "my_stats": my_stats,
            "class_stats": class_stats,
            "my_rank": my_rank,
            "total_students": len(rankings),
            "class_name": class_name,
            "above_average": float(my_stats['my_avg'] or 0) > float(class_stats['class_avg'] or 0) if my_stats and class_stats else False
        })
    
    except Exception as e:
        logger.error(f"Get class comparison error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/teacher/api/class-analysis/<class_name>")
@login_required
def get_class_analysis(class_name):
    """Öğretmen için sınıf analizi - en iyi/en zayıf öğrenciler, ders bazlı performans"""
    if current_user.role != 'teacher':
        return jsonify({"error": "Sadece öğretmenler sınıf analizi görebilir"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # En iyi 5 öğrenci (ortalama LGS puanına göre)
        cur.execute("""
            SELECT 
                u.id,
                u.full_name,
                AVG(pe.lgs_score) as avg_score,
                MAX(pe.lgs_score) as max_score,
                COUNT(*) as exam_count
            FROM practice_exams pe
            JOIN users u ON pe.student_id = u.id
            WHERE u.class_name = %s AND u.role = 'student'
            GROUP BY u.id, u.full_name
            HAVING COUNT(*) >= 3
            ORDER BY avg_score DESC
            LIMIT 5
        """, (class_name,))
        top_students = cur.fetchall()
        
        # En zayıf 5 öğrenci
        cur.execute("""
            SELECT 
                u.id,
                u.full_name,
                AVG(pe.lgs_score) as avg_score,
                MAX(pe.lgs_score) as max_score,
                COUNT(*) as exam_count
            FROM practice_exams pe
            JOIN users u ON pe.student_id = u.id
            WHERE u.class_name = %s AND u.role = 'student'
            GROUP BY u.id, u.full_name
            HAVING COUNT(*) >= 3
            ORDER BY avg_score ASC
            LIMIT 5
        """, (class_name,))
        bottom_students = cur.fetchall()
        
        # Ders bazlı sınıf performansı
        cur.execute("""
            SELECT 
                AVG(turkce_net) as turkce_avg,
                AVG(matematik_net) as matematik_avg,
                AVG(fen_net) as fen_avg,
                AVG(sosyal_net) as sosyal_avg,
                AVG(ingilizce_net) as ingilizce_avg,
                AVG(din_net) as din_avg,
                AVG(lgs_score) as lgs_avg
            FROM practice_exams pe
            JOIN users u ON pe.student_id = u.id
            WHERE u.class_name = %s AND u.role = 'student'
        """, (class_name,))
        subject_performance = cur.fetchone()
        
        cur.close()
        conn.close()
        
        # Performansları yüzdeye çevir
        subjects_analysis = []
        if subject_performance:
            subjects = [
                ("Türkçe", float(subject_performance['turkce_avg'] or 0), 20),
                ("Matematik", float(subject_performance['matematik_avg'] or 0), 20),
                ("Fen", float(subject_performance['fen_avg'] or 0), 20),
                ("Sosyal", float(subject_performance['sosyal_avg'] or 0), 10),
                ("İngilizce", float(subject_performance['ingilizce_avg'] or 0), 10),
                ("Din", float(subject_performance['din_avg'] or 0), 10)
            ]
            for name, avg, max_net in subjects:
                subjects_analysis.append({
                    "name": name,
                    "avg_net": round(avg, 2),
                    "max_net": max_net,
                    "percentage": round((avg / max_net) * 100, 1)
                })
        
        return jsonify({
            "top_students": top_students,
            "bottom_students": bottom_students,
            "subjects_analysis": subjects_analysis,
            "class_name": class_name
        })
    
    except Exception as e:
        logger.error(f"Get class analysis error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# ==================== LGS KOÇLUK SİSTEMİ API'LERİ BİTİŞ ====================

# PDF RAPOR SİSTEMİ

def create_practice_exam_chart(exams, title, filename):
    """Grafik oluştur ve dosyaya kaydet - BAR CHART"""
    import numpy as np
    
    exam_numbers = [f"D{e['exam_number']}" for e in exams]
    n_exams = len(exam_numbers)
    
    # 3x3 grid: 6 ders + 1 LGS (7 grafik)
    fig = plt.figure(figsize=(14, 10))
    
    subjects = [
        ('Türkçe', 'turkce_net', '#ef4444', 20, 1),
        ('Matematik', 'matematik_net', '#3b82f6', 20, 2),
        ('Fen', 'fen_net', '#10b981', 20, 3),
        ('Sosyal', 'sosyal_net', '#f59e0b', 10, 4),
        ('İngilizce', 'ingilizce_net', '#8b5cf6', 10, 5),
        ('Din', 'din_net', '#ec4899', 10, 6)
    ]
    
    # Her ders için bar chart
    for subject_name, field, color, max_val, pos in subjects:
        ax = plt.subplot(3, 3, pos)
        values = [float(e[field]) for e in exams]
        
        bars = ax.bar(exam_numbers, values, color=color, alpha=0.8, edgecolor=color, linewidth=1.5)
        
        # Bar üzerinde değer göster
        for bar, val in zip(bars, values):
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                   f'{val:.1f}', ha='center', va='bottom', fontsize=8, fontweight='bold')
        
        ax.set_ylabel('Net', fontsize=9)
        ax.set_title(f'{subject_name} ({max_val} soru)', fontsize=10, fontweight='bold')
        ax.set_ylim(0, max_val)
        ax.grid(axis='y', alpha=0.3)
        ax.set_axisbelow(True)
    
    # LGS Puan Bar Chart (daha büyük, 7. pozisyon)
    ax_lgs = plt.subplot(3, 3, 7)
    lgs_scores = [float(e['lgs_score']) for e in exams]
    
    # Renk gradyanı: düşük->kırmızı, yüksek->yeşil
    colors_lgs = []
    for score in lgs_scores:
        if score >= 400:
            colors_lgs.append('#10b981')  # Yeşil
        elif score >= 300:
            colors_lgs.append('#3b82f6')  # Mavi
        elif score >= 200:
            colors_lgs.append('#f59e0b')  # Sarı
        else:
            colors_lgs.append('#ef4444')  # Kırmızı
    
    bars_lgs = ax_lgs.bar(exam_numbers, lgs_scores, color=colors_lgs, alpha=0.8, edgecolor='black', linewidth=1.5)
    
    # LGS bar üzerinde değer göster
    for bar, val in zip(bars_lgs, lgs_scores):
        height = bar.get_height()
        ax_lgs.text(bar.get_x() + bar.get_width()/2., height,
                   f'{val:.1f}', ha='center', va='bottom', fontsize=9, fontweight='bold')
    
    ax_lgs.set_ylabel('LGS Puanı', fontsize=10)
    ax_lgs.set_title('LGS Puan Gelişimi', fontsize=11, fontweight='bold')
    ax_lgs.set_ylim(0, 500)
    ax_lgs.grid(axis='y', alpha=0.3)
    ax_lgs.set_axisbelow(True)
    
    # Ana başlık
    fig.suptitle(f'{title} - Deneme Sınav Gelişim Raporu', fontsize=14, fontweight='bold', y=0.98)
    
    plt.tight_layout(rect=[0, 0, 1, 0.96])
    plt.savefig(filename, dpi=150, bbox_inches='tight')
    plt.close()

@app.route("/admin/api/practice-exams/report/class/<class_name>")
@app.route("/teacher/api/practice-exams/report/class/<class_name>")
@login_required
def generate_class_report_pdf(class_name):
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Sınıftaki öğrencileri bul
        cur.execute("""
            SELECT id FROM users 
            WHERE class_name = %s AND role = 'student'
        """, (class_name,))
        students = cur.fetchall()
        
        if not students:
            return jsonify({"error": "Bu sınıfta öğrenci bulunamadı"}), 404
        
        student_ids = [s['id'] for s in students]
        
        # Sınıf ortalaması
        cur.execute("""
            SELECT 
                exam_number,
                AVG(turkce_net) as turkce_net,
                AVG(matematik_net) as matematik_net,
                AVG(fen_net) as fen_net,
                AVG(sosyal_net) as sosyal_net,
                AVG(ingilizce_net) as ingilizce_net,
                AVG(din_net) as din_net,
                AVG(lgs_score) as lgs_score,
                COUNT(*) as student_count
            FROM practice_exams
            WHERE student_id = ANY(%s)
            GROUP BY exam_number
            ORDER BY exam_number
        """, (student_ids,))
        
        exams = cur.fetchall()
        
        if not exams:
            return jsonify({"error": "Bu sınıf için deneme verisi bulunamadı"}), 404
        
        # PDF oluştur
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        story = []
        styles = getSampleStyleSheet()
        
        # Logo ve okul adı header'ı
        story.extend(create_pdf_header(styles))
        
        # Başlık
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#667eea'), alignment=TA_CENTER, fontName='DejaVuSans')
        story.append(Paragraph(f"{class_name} Sınıfı - LGS Deneme Gelişim Raporu", title_style))
        story.append(Spacer(1, 15))
        
        # Tarih
        date_style = ParagraphStyle('Date', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER)
        story.append(Paragraph(f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}", date_style))
        story.append(Spacer(1, 15))
        
        # Özet bilgiler
        summary_data = [
            ['Sınıf', class_name],
            ['Öğrenci Sayısı', str(len(students))],
            ['Toplam Deneme', str(len(exams))],
            ['Ortalama LGS Puanı', f"{sum([float(e['lgs_score']) for e in exams]) / len(exams):.2f}"]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'DejaVuSans-Bold'),
            ('FONTNAME', (1, 0), (-1, -1), 'DejaVuSans'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 30))
        
        # Grafik oluştur ve ekle
        try:
            chart_filename = f'/tmp/class_chart_{class_name}_{datetime.now().timestamp()}.png'
            create_practice_exam_chart(exams, f"{class_name} Sınıfı", chart_filename)
            
            from reportlab.platypus import Image as RLImage
            img = RLImage(chart_filename, width=6.5*inch, height=5.2*inch)
            story.append(img)
            story.append(PageBreak())
        except Exception as chart_error:
            logger.error(f"Chart error: {chart_error}")
            story.append(Paragraph("Grafik oluşturulamadı", styles['Normal']))
            story.append(Spacer(1, 20))
        
        # Detaylı tablo
        story.append(Paragraph("Deneme Bazında Detaylı Sonuçlar", styles['Heading2']))
        story.append(Spacer(1, 10))
        
        table_data = [['Deneme', 'Türkçe', 'Matematik', 'Fen', 'Sosyal', 'İngilizce', 'Din', 'LGS']]
        
        for exam in exams:
            table_data.append([
                str(exam['exam_number']),
                f"{float(exam['turkce_net']):.2f}",
                f"{float(exam['matematik_net']):.2f}",
                f"{float(exam['fen_net']):.2f}",
                f"{float(exam['sosyal_net']):.2f}",
                f"{float(exam['ingilizce_net']):.2f}",
                f"{float(exam['din_net']):.2f}",
                f"{float(exam['lgs_score']):.2f}"
            ])
        
        detail_table = Table(table_data, colWidths=[0.7*inch]*8)
        detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'DejaVuSans-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'DejaVuSans'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(detail_table)
        
        doc.build(story)
        cur.close()
        conn.close()
        
        buffer.seek(0)
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=False,
            download_name=f'{class_name}_deneme_raporu_{datetime.now().strftime("%Y%m%d")}.pdf'
        )
    
    except Exception as e:
        logger.error(f"Generate class report error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/teacher/api/practice-exams/report/grade-level/<grade_level>/<exam_no>/<subject>")
@login_required
def generate_grade_level_report_pdf(grade_level, exam_no, subject):
    """Sınıf seviyesi karşılaştırma PDF raporu"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Bu seviyedeki tüm sınıfları bul
        grade_classes = [f"{grade_level}{letter}" for letter in ['A', 'B', 'C', 'D', 'E']]
        
        subject_names = {
            'turkce': 'Türkçe',
            'matematik': 'Matematik',
            'fen': 'Fen Bilimleri',
            'sosyal': 'Sosyal Bilgiler',
            'ingilizce': 'İngilizce',
            'din': 'Din Kültürü',
            'lgs_score': 'LGS Puanı'
        }
        
        subject_field = 'lgs_score' if subject == 'lgs_score' else f'{subject}_net'
        
        class_data = []
        
        for class_name in grade_classes:
            cur.execute("""
                SELECT id FROM users 
                WHERE class_name = %s AND role = 'student'
            """, (class_name,))
            students = cur.fetchall()
            
            if not students:
                continue
                
            student_ids = [s['id'] for s in students]
            
            cur.execute(f"""
                SELECT AVG({subject_field}) as avg_value, COUNT(*) as count
                FROM practice_exams
                WHERE student_id = ANY(%s) AND exam_number = %s
            """, (student_ids, int(exam_no)))
            
            result = cur.fetchone()
            if result and result['avg_value'] is not None:
                class_data.append({
                    'class_name': class_name,
                    'avg_value': float(result['avg_value']),
                    'student_count': result['count']
                })
        
        cur.close()
        conn.close()
        
        if not class_data:
            return jsonify({"error": "Bu deneme için veri bulunamadı"}), 404
        
        # PDF oluştur
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        story = []
        styles = getSampleStyleSheet()
        
        # Logo ve okul adı header'ı
        story.extend(create_pdf_header(styles))
        
        # Başlık
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#667eea'), alignment=TA_CENTER, fontName='DejaVuSans')
        story.append(Paragraph(f"{grade_level}. Sınıflar - Deneme {exam_no} - {subject_names.get(subject, subject)} Karşılaştırması", title_style))
        story.append(Spacer(1, 15))
        
        # Tarih
        date_style = ParagraphStyle('Date', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER)
        story.append(Paragraph(f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}", date_style))
        story.append(Spacer(1, 20))
        
        # Özet bilgiler
        unit = 'puan' if subject == 'lgs_score' else 'net'
        summary_data = [
            ['Sınıf Seviyesi', f'{grade_level}. Sınıflar'],
            ['Deneme No', f'Deneme {exam_no}'],
            ['Karşılaştırılan Ders', subject_names.get(subject, subject)],
            ['Toplam Sınıf Sayısı', str(len(class_data))]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'DejaVuSans-Bold'),
            ('FONTNAME', (1, 0), (-1, -1), 'DejaVuSans'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 30))
        
        # Sınıf karşılaştırma tablosu
        story.append(Paragraph("Sınıf Bazında Sonuçlar", styles['Heading2']))
        story.append(Spacer(1, 10))
        
        table_data = [['Sınıf', f'Ortalama {subject_names.get(subject, subject)}', 'Öğrenci Sayısı', 'Sıralama']]
        
        sorted_classes = sorted(class_data, key=lambda x: x['avg_value'], reverse=True)
        
        for i, cls in enumerate(sorted_classes, 1):
            table_data.append([
                cls['class_name'],
                f"{cls['avg_value']:.2f} {unit}",
                str(cls['student_count']),
                f"{i}."
            ])
        
        detail_table = Table(table_data, colWidths=[1.5*inch, 2.5*inch, 1.5*inch, 1*inch])
        detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'DejaVuSans-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'DejaVuSans'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(detail_table)
        story.append(Spacer(1, 20))
        
        # En iyi ve en düşük performans
        if sorted_classes:
            best = sorted_classes[0]
            worst = sorted_classes[-1]
            
            analysis_style = ParagraphStyle('Analysis', parent=styles['Normal'], fontSize=11)
            story.append(Paragraph(f"<b>En Yüksek Ortalama:</b> {best['class_name']} ({best['avg_value']:.2f} {unit})", analysis_style))
            story.append(Paragraph(f"<b>En Düşük Ortalama:</b> {worst['class_name']} ({worst['avg_value']:.2f} {unit})", analysis_style))
            
            if len(sorted_classes) > 1:
                avg_all = sum(c['avg_value'] for c in sorted_classes) / len(sorted_classes)
                story.append(Paragraph(f"<b>Genel Ortalama:</b> {avg_all:.2f} {unit}", analysis_style))
        
        doc.build(story)
        
        buffer.seek(0)
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=False,
            download_name=f'sinif_{grade_level}_deneme_{exam_no}_{subject}_karsilastirma_{datetime.now().strftime("%Y%m%d")}.pdf'
        )
    
    except Exception as e:
        logger.error(f"Generate grade level report error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/student/api/practice-exams/report/my-pdf")
@login_required
def generate_my_report_pdf():
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    return generate_student_report_pdf(current_user.id)

@app.route("/teacher/api/practice-exams/report/student/<int:student_id>")
@login_required
def generate_student_report_pdf(student_id):
    if current_user.role not in ['admin', 'teacher', 'student']:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    if current_user.role == 'student' and current_user.id != student_id:
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Öğrenci bilgisi
        cur.execute("SELECT id, full_name, class_name FROM users WHERE id = %s", (student_id,))
        student = cur.fetchone()
        
        if not student:
            return jsonify({"error": "Öğrenci bulunamadı"}), 404
        
        # Deneme sonuçları
        cur.execute("""
            SELECT * FROM practice_exams 
            WHERE student_id = %s 
            ORDER BY exam_number
        """, (student_id,))
        exams = cur.fetchall()
        
        if not exams:
            return jsonify({"error": "Bu öğrenci için deneme verisi bulunamadı"}), 404
        
        # PDF oluştur
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        story = []
        styles = getSampleStyleSheet()
        
        # Logo ve okul adı header'ı
        story.extend(create_pdf_header(styles))
        
        # Başlık
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#10b981'), alignment=TA_CENTER, fontName='DejaVuSans')
        story.append(Paragraph(f"LGS Deneme Gelişim Raporu", title_style))
        story.append(Spacer(1, 15))
        
        # Öğrenci bilgileri
        student_info = ParagraphStyle('StudentInfo', parent=styles['Normal'], fontSize=12, alignment=TA_CENTER, textColor=colors.HexColor('#374151'), fontName='DejaVuSans')
        story.append(Paragraph(f"<b>{student['full_name']}</b> - {student['class_name']}", student_info))
        story.append(Spacer(1, 8))
        story.append(Paragraph(f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}", styles['Normal']))
        story.append(Spacer(1, 15))
        
        # Özet istatistikler
        lgs_scores = [float(e['lgs_score']) for e in exams]
        summary_data = [
            ['Toplam Deneme', str(len(exams))],
            ['Ortalama LGS Puanı', f"{sum(lgs_scores) / len(lgs_scores):.2f}"],
            ['En Yüksek Puan', f"{max(lgs_scores):.2f}"],
            ['En Düşük Puan', f"{min(lgs_scores):.2f}"]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'DejaVuSans-Bold'),
            ('FONTNAME', (1, 0), (-1, -1), 'DejaVuSans'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 30))
        
        # Grafik (güvenli mode - hata olursa skip et)
        try:
            chart_filename = f'/tmp/student_chart_{student_id}_{datetime.now().timestamp()}.png'
            create_practice_exam_chart(exams, student['full_name'], chart_filename)
            import os
            if os.path.exists(chart_filename):
                img = Image(chart_filename, width=6.5*inch, height=5.2*inch)
                story.append(img)
                story.append(PageBreak())
        except Exception as chart_err:
            logger.warning(f"Chart generation skipped: {str(chart_err)}")
            story.append(Paragraph("📊 (Grafik gösterimi kısmen devre dışı)", styles['Normal']))
            story.append(Spacer(1, 20))
        
        # Detaylı tablo
        story.append(Paragraph("Deneme Bazında Detaylı Sonuçlar", styles['Heading2']))
        story.append(Spacer(1, 10))
        
        table_data = [['Deneme', 'Türkçe', 'Mat', 'Fen', 'Sosyal', 'İng', 'Din', 'LGS']]
        
        for exam in exams:
            table_data.append([
                str(exam['exam_number']),
                f"{float(exam['turkce_net']):.1f}",
                f"{float(exam['matematik_net']):.1f}",
                f"{float(exam['fen_net']):.1f}",
                f"{float(exam['sosyal_net']):.1f}",
                f"{float(exam['ingilizce_net']):.1f}",
                f"{float(exam['din_net']):.1f}",
                f"{float(exam['lgs_score']):.1f}"
            ])
        
        detail_table = Table(table_data, colWidths=[0.7*inch]*8)
        detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'DejaVuSans-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'DejaVuSans'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(detail_table)
        
        doc.build(story)
        cur.close()
        conn.close()
        
        buffer.seek(0)
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=False,
            download_name=f'{student["full_name"]}_deneme_raporu_{datetime.now().strftime("%Y%m%d")}.pdf'
        )
    
    except Exception as e:
        logger.error(f"Generate student report error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# ==================== RANKING SİSTEMİ ====================

@app.route("/student/api/practice-exams/rankings")
@login_required
def get_practice_exam_rankings():
    """Öğrencinin deneme rankinglerini döndür (sınıf ve okul)"""
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Öğrencinin sınıfını bul
        cur.execute("SELECT class_name FROM users WHERE id = %s", (current_user.id,))
        user = cur.fetchone()
        
        if not user:
            return jsonify({"error": "Kullanıcı bulunamadı"}), 404
        
        class_name = user['class_name']
        
        # Deneme sonuçları
        cur.execute("""
            SELECT * FROM practice_exams 
            WHERE student_id = %s 
            ORDER BY exam_number
        """, (current_user.id,))
        exams = cur.fetchall()
        
        rankings = {}
        
        for exam in exams:
            exam_num = exam['exam_number']
            rankings[exam_num] = {}
            
            # LGS PUAN RANKİ
            # Sınıf içi rank
            cur.execute("""
                SELECT COUNT(*) as rank
                FROM practice_exams pe
                JOIN users u ON pe.student_id = u.id
                WHERE u.class_name = %s AND pe.exam_number = %s
                AND pe.lgs_score > %s
            """, (class_name, exam_num, exam['lgs_score']))
            class_rank_lgs = cur.fetchone()['rank'] + 1
            
            # Okul içi rank
            cur.execute("""
                SELECT COUNT(*) as rank
                FROM practice_exams
                WHERE exam_number = %s AND lgs_score > %s
            """, (exam_num, exam['lgs_score']))
            school_rank_lgs = cur.fetchone()['rank'] + 1
            
            rankings[exam_num]['lgs'] = {
                'class_rank': class_rank_lgs,
                'school_rank': school_rank_lgs,
                'score': float(exam['lgs_score'])
            }
            
            # HER DERS İÇİN RANK (Güvenli - whitelist ile)
            subjects_map = {
                'turkce': ('turkce_net', exam['turkce_net']),
                'matematik': ('matematik_net', exam['matematik_net']),
                'fen': ('fen_net', exam['fen_net']),
                'sosyal': ('sosyal_net', exam['sosyal_net']),
                'ingilizce': ('ingilizce_net', exam['ingilizce_net']),
                'din': ('din_net', exam['din_net'])
            }
            
            for subject_key, (column_name, net_score) in subjects_map.items():
                net_score = float(net_score)
                
                # Sınıf rank - parametrize sütun adı
                class_rank_query = f"""
                    SELECT COUNT(*) as rank
                    FROM practice_exams pe
                    JOIN users u ON pe.student_id = u.id
                    WHERE u.class_name = %s AND pe.exam_number = %s
                    AND pe.{column_name} > %s
                """
                cur.execute(class_rank_query, (class_name, exam_num, net_score))
                class_rank = cur.fetchone()['rank'] + 1
                
                # Okul rank - parametrize sütun adı
                school_rank_query = f"""
                    SELECT COUNT(*) as rank
                    FROM practice_exams
                    WHERE exam_number = %s AND {column_name} > %s
                """
                cur.execute(school_rank_query, (exam_num, net_score))
                school_rank = cur.fetchone()['rank'] + 1
                
                rankings[exam_num][subject_key] = {
                    'class_rank': class_rank,
                    'school_rank': school_rank,
                    'net': net_score
                }
        
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "rankings": rankings})
    
    except Exception as e:
        logger.error(f"Get rankings error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# ==================== DENEME TAKİP SİSTEMİ SONU ====================

# ==================== GÜNLÜK ÇALIŞMA TAKİBİ PDF RAPORU ====================

@app.route("/teacher/daily-tracking/pdf")
@login_required
def generate_daily_tracking_pdf():
    """Öğretmen günlük çalışma takibi PDF raporu oluşturur"""
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    view_type = request.args.get('view', 'week')
    date_str = request.args.get('date')
    
    if not date_str:
        return jsonify({"error": "Tarih parametresi gerekli"}), 400
    
    try:
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({"error": "Geçersiz tarih formatı"}), 400
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Öğretmenin öğrencilerini bul
        cur.execute("""
            SELECT DISTINCT u.id, u.full_name, u.class_name
            FROM users u
            WHERE u.role = 'student'
            AND (
                u.id IN (SELECT student_id FROM teacher_students WHERE teacher_id = %s)
                OR u.class_name IN (SELECT class_name FROM teacher_classes WHERE teacher_id = %s)
            )
            ORDER BY u.class_name, u.full_name
        """, (current_user.id, current_user.id))
        
        students = cur.fetchall()
        
        if not students:
            return jsonify({"error": "Öğrenci bulunamadı"}), 404
        
        # Tarih aralığını hesapla
        if view_type == 'week':
            weekday = selected_date.weekday()
            start_date = selected_date - timedelta(days=weekday)
            end_date = start_date + timedelta(days=6)
            period_text = "Haftalık"
        else:
            start_date = selected_date.replace(day=1)
            if selected_date.month == 12:
                end_date = selected_date.replace(year=selected_date.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                end_date = selected_date.replace(month=selected_date.month + 1, day=1) - timedelta(days=1)
            period_text = "Aylık"
        
        # Tüm kayıtları tek query ile çek (N+1 optimizasyonu)
        student_ids = [s['id'] for s in students]
        
        cur.execute("""
            SELECT 
                dst.id, dst.student_id, dst.date, dst.day_of_week, 
                dst.subject, dst.note, dst.created_at, dst.updated_at
            FROM daily_study_tracking dst
            WHERE dst.student_id = ANY(%s)
            AND dst.date BETWEEN %s AND %s
            ORDER BY dst.student_id, dst.date ASC, dst.subject ASC
        """, (student_ids, start_date, end_date))
        
        all_records = cur.fetchall()
        total_records = len(all_records)
        
        records_by_student = {}
        for record in all_records:
            sid = record['student_id']
            if sid not in records_by_student:
                records_by_student[sid] = []
            records_by_student[sid].append(record)
        
        students_data = []
        for student in students:
            students_data.append({
                "id": student['id'],
                "name": student['full_name'],
                "class": student['class_name'],
                "records": records_by_student.get(student['id'], [])
            })
        
        # PDF oluştur
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        story = []
        styles = getSampleStyleSheet()
        
        # Logo ve okul adı header'ı
        story.extend(create_pdf_header(styles))
        
        # Başlık
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#2563eb'), alignment=TA_CENTER, fontName='DejaVuSans')
        story.append(Paragraph(f"{period_text} Günlük Çalışma Takip Raporu", title_style))
        story.append(Spacer(1, 15))
        
        # Tarih ve öğretmen bilgisi
        date_style = ParagraphStyle('Date', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER)
        story.append(Paragraph(f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}", date_style))
        story.append(Paragraph(f"Öğretmen: {current_user.full_name}", date_style))
        story.append(Paragraph(f"Dönem: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}", date_style))
        story.append(Spacer(1, 15))
        
        # Özet bilgiler
        summary_data = [
            ['Toplam Öğrenci', str(len(students))],
            ['Toplam Kayıt', str(total_records)],
            ['Ortalama Kayıt/Öğrenci', f"{total_records / len(students):.1f}" if students else "0"]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'DejaVuSans-Bold'),
            ('FONTNAME', (1, 0), (-1, -1), 'DejaVuSans'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        
        story.append(summary_table)
        story.append(Spacer(1, 30))
        
        # Her öğrenci için detaylar
        for student_data in students_data:
            story.append(Paragraph(f"<b>{student_data['name']}</b> ({student_data['class']})", styles['Heading3']))
            story.append(Spacer(1, 10))
            
            if not student_data['records']:
                story.append(Paragraph("Bu dönemde kayıt yok", styles['Normal']))
                story.append(Spacer(1, 20))
                continue
            
            # Öğrenci kayıtları tablosu
            table_data = [['Tarih', 'Gün', 'Ders', 'Not']]
            
            for record in student_data['records']:
                note_text = record['note'][:100] + ('...' if len(record['note']) > 100 else '')
                
                table_data.append([
                    record['date'].strftime('%d.%m'),
                    record['day_of_week'],
                    record['subject'],
                    note_text
                ])
            
            student_table = Table(table_data, colWidths=[0.8*inch, 1*inch, 1.2*inch, 3.5*inch])
            student_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (2, -1), 'CENTER'),
                ('ALIGN', (3, 0), (3, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'DejaVuSans-Bold'),
                ('FONTNAME', (0, 1), (-1, -1), 'DejaVuSans'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
            ]))
            
            story.append(student_table)
            story.append(Spacer(1, 20))
        
        doc.build(story)
        cur.close()
        conn.close()
        
        buffer.seek(0)
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=False,
            download_name=f'gunluk_takip_raporu_{start_date.strftime("%Y%m%d")}_{datetime.now().strftime("%H%M%S")}.pdf'
        )
    
    except Exception as e:
        logger.error(f"Generate daily tracking PDF error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# ==================== GÜNLÜK ÇALIŞMA TAKİBİ PDF RAPORU SONU ====================

# ==================== ADMIN EK SİLME ENDPOINTLERİ ====================

@app.route("/admin/announcements/<int:announcement_id>", methods=["DELETE"])
@login_required
def admin_delete_announcement(announcement_id):
    """Admin duyuru siler"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # Duyuruyu sil (doğru tablo: teacher_announcements)
        cur.execute("DELETE FROM teacher_announcements WHERE id = %s", (announcement_id,))
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Duyuru silindi"})
    except Exception as e:
        logger.error(f"Delete announcement error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/admin/public-announcements/<int:announcement_id>", methods=["DELETE"])
@login_required
def admin_delete_public_announcement(announcement_id):
    """Admin ana sayfa duyurusu siler"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # Ana sayfa duyurusunu sil
        cur.execute("DELETE FROM public_announcements WHERE id = %s", (announcement_id,))
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Ana sayfa duyurusu silindi"})
    except Exception as e:
        logger.error(f"Delete public announcement error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/admin/practice-exams/<int:exam_id>", methods=["DELETE"])
@login_required
def admin_delete_practice_exam(exam_id):
    """Admin deneme sınavı siler"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # Önce ilgili skorları sil
        cur.execute("DELETE FROM exam_scores WHERE exam_id = %s", (exam_id,))
        
        # Sonra sınavı sil
        cur.execute("DELETE FROM student_exams WHERE id = %s", (exam_id,))
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Deneme sınavı silindi"})
    except Exception as e:
        logger.error(f"Delete practice exam error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# ==================== ADMIN EK SİLME ENDPOINTLERİ SONU ====================

# ==================== ANA SAYFA DUYURU SİSTEMİ ====================

@app.route("/api/public-announcements", methods=["GET"])
def get_public_announcements():
    """Ana sayfa için aktif duyuruları getir"""
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT id, title, content, video_url, created_at 
            FROM public_announcements 
            WHERE is_active = TRUE 
            ORDER BY created_at DESC 
            LIMIT 5
        """)
        
        announcements = cur.fetchall()
        cur.close()
        conn.close()
        
        # Datetime'ları serialize et
        announcements_list = []
        for ann in announcements:
            ann_dict = dict(ann)
            if ann_dict.get('created_at'):
                ann_dict['created_at'] = ann_dict['created_at'].isoformat()
            announcements_list.append(ann_dict)
        
        return jsonify({"success": True, "announcements": announcements_list})
    except Exception as e:
        logger.error(f"Public announcements error: {str(e)}")
        return jsonify({"success": False, "error": str(e), "announcements": []}), 500

@app.route("/admin/announcements/create-public", methods=["POST"])
@login_required
def admin_create_public_announcement():
    """Admin - Ana sayfa duyurusu oluştur"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        title = request.form.get('title')
        content = request.form.get('content')
        video_url = request.form.get('video_url', '').strip() or None
        
        if not title or not content:
            return jsonify({"error": "Başlık ve içerik zorunludur"}), 400
        
        conn = get_db()
        cur = conn.cursor()
        
        cur.execute("""
            INSERT INTO public_announcements (title, content, video_url, admin_id, is_active)
            VALUES (%s, %s, %s, %s, TRUE)
        """, (title, content, video_url, current_user.id))
        
        conn.commit()
        cur.close()
        conn.close()
        
        # Push bildirim gönder - herkese
        send_push_notification(
            title="📢 Yeni Ana Sayfa Duyurusu",
            message="AMEO'da yeni bir duyuru yayınlandı. Detaylar için tıklayın.",
            url="https://ameo-alanya.com"
        )
        
        # Öğretmenlere özel bildirim
        send_push_notification(
            title="📢 Yeni Ana Sayfa Duyurusu",
            message="AMEO'da yeni bir duyuru yayınlandı. Kontrol etmeyi unutmayın.",
            url="https://ameo-alanya.com",
            target_role="teacher"
        )
        
        return jsonify({"success": True, "message": "Ana sayfa duyurusu yayınlandı"})
    except Exception as e:
        logger.error(f"Create public announcement error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# ==================== ANA SAYFA DUYURU SİSTEMİ SONU ====================

# ==================== MANUEL BİLDİRİM GÖNDERME ====================

@app.route("/api/send-notification", methods=["POST"])
@login_required
def api_send_notification():
    """Manuel push bildirim gönder (sadece yetkili kullanıcılar)"""
    if not can_send_notification(current_user):
        return jsonify({"error": "Bildirim gönderme yetkiniz yok"}), 403
    
    try:
        data = request.get_json()
        title = data.get('title')
        message = data.get('message')
        url = data.get('url', 'https://ameo-alanya.com')
        target_classes = data.get('target_classes')
        target_role = data.get('target_role')
        
        if not title or not message:
            return jsonify({"error": "Başlık ve mesaj zorunludur"}), 400
        
        logger.info(f"📢 Manuel bildirim gönderiliyor: {title}, hedef_sınıflar={target_classes}, hedef_rol={target_role}")
        
        success = send_push_notification(
            title=title, 
            message=message, 
            url=url,
            target_classes=target_classes,
            target_role=target_role
        )
        
        if success:
            return jsonify({"success": True, "message": "Bildirim gönderildi"})
        else:
            return jsonify({"error": "Bildirim gönderilemedi"}), 500
    except Exception as e:
        logger.error(f"Send notification error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/can-send-notification", methods=["GET"])
@login_required
def api_can_send_notification():
    """Kullanıcının bildirim gönderme yetkisini kontrol et"""
    return jsonify({"can_send": can_send_notification(current_user)})

# ==================== ÖĞRETMEN-ÖĞRENCİ ATAMA SİSTEMİ ====================

# SINIF BAZLI ATAMA (YENİ)
@app.route("/api/admin/teacher-classes/assign", methods=["POST"])
@login_required
def admin_assign_classes():
    """Admin öğretmenlere sınıf atar (birden fazla)"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    data = request.get_json()
    teacher_id = data.get('teacher_id')
    class_names = data.get('class_names', [])
    
    if not teacher_id or not class_names:
        return jsonify({"error": "Öğretmen ve sınıf bilgileri gerekli"}), 400
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # Öğretmen kontrolü
        cur.execute("SELECT role FROM users WHERE id = %s", (teacher_id,))
        teacher = cur.fetchone()
        if not teacher or teacher[0] != 'teacher':
            return jsonify({"error": "Geçersiz öğretmen"}), 400
        
        assigned = 0
        for class_name in class_names:
            try:
                cur.execute("""
                    INSERT INTO teacher_classes (teacher_id, class_name)
                    VALUES (%s, %s)
                    ON CONFLICT (teacher_id, class_name) DO NOTHING
                """, (teacher_id, class_name))
                if cur.rowcount > 0:
                    assigned += 1
            except:
                pass
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "message": f"{assigned} sınıf atandı"
        })
    except Exception as e:
        logger.error(f"Assign classes error: {e}")
        return jsonify({"error": "Atama başarısız"}), 500

@app.route("/api/admin/teacher-classes/<int:teacher_id>", methods=["GET"])
@login_required
def get_teacher_classes(teacher_id):
    """Öğretmene atanmış sınıfları getir"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT class_name
            FROM teacher_classes
            WHERE teacher_id = %s
            ORDER BY class_name
        """, (teacher_id,))
        
        classes = cur.fetchall()
        cur.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "classes": [dict(c) for c in classes]
        })
    except Exception as e:
        logger.error(f"Get teacher classes error: {e}")
        return jsonify({"error": "Yükleme başarısız"}), 500

@app.route("/api/teacher-classes", methods=["GET"])
@login_required
def teacher_get_my_classes():
    """Öğretmen kendi atanmış sınıflarını getir"""
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT class_name
            FROM teacher_classes
            WHERE teacher_id = %s
            ORDER BY class_name
        """, (current_user.id,))
        
        classes = cur.fetchall()
        cur.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "classes": [dict(c) for c in classes]
        })
    except Exception as e:
        logger.error(f"Teacher get classes error: {e}")
        return jsonify({"error": "Yükleme başarısız"}), 500

@app.route("/api/admin/teacher-classes/remove", methods=["POST"])
@login_required
def admin_remove_class():
    """Admin öğretmenden sınıf atamasını kaldırır"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    data = request.get_json()
    teacher_id = data.get('teacher_id')
    class_name = data.get('class_name')
    
    if not teacher_id or not class_name:
        return jsonify({"error": "Öğretmen ve sınıf bilgileri gerekli"}), 400
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        cur.execute("""
            DELETE FROM teacher_classes
            WHERE teacher_id = %s AND class_name = %s
        """, (teacher_id, class_name))
        
        removed = cur.rowcount > 0
        conn.commit()
        cur.close()
        conn.close()
        
        if removed:
            return jsonify({
                "success": True,
                "message": "Sınıf ataması kaldırıldı"
            })
        else:
            return jsonify({"error": "Atama bulunamadı"}), 404
    except Exception as e:
        logger.error(f"Remove class error: {e}")
        return jsonify({"error": "Kaldırma başarısız"}), 500

# BİREYSEL ATAMA (MEVCUT - ÖZEL DURUMLAR İÇİN)
@app.route("/api/admin/teacher-students/assign", methods=["POST"])
@login_required
def admin_assign_students():
    """Admin öğretmenlere öğrenci atar"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    data = request.get_json()
    teacher_id = data.get('teacher_id')
    student_ids = data.get('student_ids', [])
    
    if not teacher_id or not student_ids:
        return jsonify({"error": "Öğretmen ve öğrenci bilgileri gerekli"}), 400
    
    try:
        conn = get_db()
        cur = conn.cursor()
        
        # Öğretmen kontrolü
        cur.execute("SELECT role FROM users WHERE id = %s", (teacher_id,))
        teacher = cur.fetchone()
        if not teacher or teacher[0] != 'teacher':
            return jsonify({"error": "Geçersiz öğretmen"}), 400
        
        assigned = 0
        for student_id in student_ids:
            try:
                cur.execute("""
                    INSERT INTO teacher_students (teacher_id, student_id)
                    VALUES (%s, %s)
                    ON CONFLICT (teacher_id, student_id) DO NOTHING
                """, (teacher_id, student_id))
                assigned += 1
            except:
                pass
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "message": f"{assigned} öğrenci atandı"
        })
    except Exception as e:
        logger.error(f"Assign students error: {e}")
        return jsonify({"error": "Atama başarısız"}), 500

@app.route("/api/admin/teacher-students/<int:teacher_id>", methods=["GET"])
@login_required
def get_teacher_students(teacher_id):
    """Öğretmene atanmış öğrencileri getir"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT u.id, u.full_name, u.username, u.class_name
            FROM users u
            JOIN teacher_students ts ON ts.student_id = u.id
            WHERE ts.teacher_id = %s
            ORDER BY u.class_name, u.full_name
        """, (teacher_id,))
        
        students = cur.fetchall()
        cur.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "students": [dict(s) for s in students]
        })
    except Exception as e:
        logger.error(f"Get teacher students error: {e}")
        return jsonify({"error": "Yükleme başarısız"}), 500

@app.route("/api/students", methods=["GET"])
@login_required
def get_students_list():
    """Öğrenci listesini getirir (role tabanlı filtreleme)"""
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        if current_user.role == 'admin':
            # Admin tüm öğrencileri görebilir
            cur.execute("""
                SELECT id, username, full_name, class_name
                FROM users
                WHERE role = 'student'
                ORDER BY class_name, full_name
            """)
        elif current_user.role == 'teacher':
            # Öğretmen sadece atanmış sınıflarındaki veya direkt atanmış öğrencileri görebilir
            cur.execute("""
                SELECT DISTINCT u.id, u.username, u.full_name, u.class_name
                FROM users u
                WHERE u.role = 'student' AND (
                    u.class_name IN (
                        SELECT class_name FROM teacher_classes WHERE teacher_id = %s
                    )
                    OR u.id IN (
                        SELECT student_id FROM teacher_students WHERE teacher_id = %s
                    )
                )
                ORDER BY u.class_name, u.full_name
            """, (current_user.id, current_user.id))
        else:
            return jsonify({"error": "Yetkisiz erişim"}), 403
        
        students = cur.fetchall()
        cur.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "students": [dict(s) for s in students]
        })
    except Exception as e:
        logger.error(f"Get students error: {e}")
        return jsonify({"error": "Yükleme başarısız"}), 500

@app.route("/admin/api/teachers", methods=["GET"])
@login_required
def admin_get_teachers_list():
    """Admin için öğretmen listesi"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT id, username, full_name
            FROM users
            WHERE role = 'teacher'
            ORDER BY full_name
        """)
        
        teachers = cur.fetchall()
        cur.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "teachers": [dict(t) for t in teachers]
        })
    except Exception as e:
        logger.error(f"Get teachers error: {e}")
        return jsonify({"error": "Yükleme başarısız"}), 500

@app.route("/admin/api/classes", methods=["GET"])
@login_required
def admin_get_classes_list():
    """Admin için sınıf listesi"""
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT name as class_name
            FROM classes
            WHERE is_active = TRUE
            ORDER BY name
        """)
        
        classes = cur.fetchall()
        cur.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "classes": [dict(c) for c in classes]
        })
    except Exception as e:
        logger.error(f"Get classes error: {e}")
        return jsonify({"error": "Yükleme başarısız"}), 500
# ==================== ÖĞRETMEN-ÖĞRENCİ ATAMA SİSTEMİ SONU ====================

# ==================== LEADERBOARD API (DENEMENİN YILDIZLARI VE YÜKSELENLERİ) ====================

@app.route("/api/leaderboards/stars", methods=["GET"])
def get_leaderboard_stars():
    """Denemenin Yıldızları - Ortalama puana göre sıralama"""
    grade = request.args.get('grade', '5')
    
    # Permission kontrolü: Öğrenci ise sadece kendi sınıf seviyesini görebilir
    current_user_id = None
    current_user_role = None
    allowed_grades = ['5', '6', '7', '8']  # Default: tüm seviyeler
    
    if current_user.is_authenticated:
        current_user_id = current_user.id
        current_user_role = current_user.role
        
        if current_user.role == 'student' and current_user.class_name:
            # Öğrenci: Sadece kendi sınıf seviyesi
            student_grade = current_user.class_name[0]  # "5A" -> "5"
            allowed_grades = [student_grade]
            grade = student_grade  # Force to own grade level
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT 
                u.id,
                u.full_name,
                u.class_name,
                COUNT(pe.id) as exam_count,
                ROUND(AVG(pe.lgs_score), 2) as avg_score,
                ROUND(MIN(pe.lgs_score), 2) as min_score,
                ROUND(MAX(pe.lgs_score), 2) as max_score
            FROM users u
            INNER JOIN practice_exams pe ON u.id = pe.student_id
            WHERE u.role = 'student' 
                AND u.class_name IS NOT NULL
                AND u.class_name LIKE %s
                AND pe.lgs_score IS NOT NULL
            GROUP BY u.id, u.full_name, u.class_name
            HAVING COUNT(pe.id) > 0
            ORDER BY avg_score DESC
        """, (grade + '%',))
        
        students = cur.fetchall()
        cur.close()
        conn.close()
        
        leaderboard = []
        for idx, student in enumerate(students, 1):
            avg = float(student['avg_score']) if student['avg_score'] else 0
            min_score = float(student['min_score']) if student['min_score'] else 0
            max_score = float(student['max_score']) if student['max_score'] else 0
            
            if avg >= 400:
                color = 'green'
            elif avg >= 350:
                color = 'yellow'
            else:
                color = 'red'
            
            leaderboard.append({
                'rank': idx,
                'student_id': student['id'],
                'student_name': student['full_name'],
                'class_name': student['class_name'],
                'exam_count': student['exam_count'],
                'avg_score': avg,
                'min_score': min_score,
                'max_score': max_score,
                'color': color
            })
        
        return jsonify({
            "success": True,
            "grade": grade,
            "leaderboard": leaderboard,
            "current_user_id": current_user_id,
            "current_user_role": current_user_role,
            "allowed_grades": allowed_grades
        })
    
    except Exception as e:
        logger.error(f"Leaderboard stars error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/leaderboards/risers", methods=["GET"])
def get_leaderboard_risers():
    """Denemenin Yükselenleri - Son iki deneme arasındaki puan farkına göre"""
    grade = request.args.get('grade', '5')
    
    # Permission kontrolü: Öğrenci ise sadece kendi sınıf seviyesini görebilir
    current_user_id = None
    current_user_role = None
    allowed_grades = ['5', '6', '7', '8']  # Default: tüm seviyeler
    
    if current_user.is_authenticated:
        current_user_id = current_user.id
        current_user_role = current_user.role
        
        if current_user.role == 'student' and current_user.class_name:
            # Öğrenci: Sadece kendi sınıf seviyesi
            student_grade = current_user.class_name[0]  # "5A" -> "5"
            allowed_grades = [student_grade]
            grade = student_grade  # Force to own grade level
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            WITH student_exams AS (
                SELECT 
                    u.id,
                    u.full_name,
                    u.class_name,
                    pe.exam_number,
                    pe.lgs_score,
                    ROW_NUMBER() OVER (PARTITION BY u.id ORDER BY pe.created_at DESC) as rn
                FROM users u
                INNER JOIN practice_exams pe ON u.id = pe.student_id
                WHERE u.role = 'student'
                    AND u.class_name IS NOT NULL
                    AND u.class_name LIKE %s
                    AND pe.lgs_score IS NOT NULL
            ),
            latest_two AS (
                SELECT 
                    id,
                    full_name,
                    class_name,
                    MAX(CASE WHEN rn = 1 THEN lgs_score END) as latest_score,
                    MAX(CASE WHEN rn = 2 THEN lgs_score END) as previous_score,
                    MAX(CASE WHEN rn = 1 THEN exam_number END) as latest_exam
                FROM student_exams
                WHERE rn <= 2
                GROUP BY id, full_name, class_name
                HAVING MAX(CASE WHEN rn = 2 THEN lgs_score END) IS NOT NULL
            )
            SELECT 
                id,
                full_name,
                class_name,
                latest_score,
                previous_score,
                ROUND(latest_score - previous_score, 2) as score_change,
                latest_exam
            FROM latest_two
            ORDER BY score_change DESC
        """, (grade + '%',))
        
        students = cur.fetchall()
        cur.close()
        conn.close()
        
        leaderboard = []
        for idx, student in enumerate(students, 1):
            change = float(student['score_change']) if student['score_change'] else 0
            
            if change >= 25:
                color = 'green'
            elif change >= 15:
                color = 'yellow'
            else:
                color = 'red'
            
            leaderboard.append({
                'rank': idx,
                'student_id': student['id'],
                'student_name': student['full_name'],
                'class_name': student['class_name'],
                'latest_score': float(student['latest_score']) if student['latest_score'] else 0,
                'previous_score': float(student['previous_score']) if student['previous_score'] else 0,
                'score_change': change,
                'color': color
            })
        
        return jsonify({
            "success": True,
            "grade": grade,
            "leaderboard": leaderboard,
            "current_user_id": current_user_id,
            "current_user_role": current_user_role,
            "allowed_grades": allowed_grades
        })
    
    except Exception as e:
        logger.error(f"Leaderboard risers error: {e}")
        return jsonify({"error": str(e)}), 500

# PDF Export Helper Function
def build_leaderboard_pdf(leaderboard_type, grade, leaderboard_data):
    """Leaderboard listesi için PDF oluşturur"""
    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.units import inch
    from datetime import datetime
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    story = []
    styles = getSampleStyleSheet()
    
    # Başlık stili
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName='DejaVuSans-Bold',
        fontSize=18,
        textColor=colors.HexColor('#667eea'),
        alignment=TA_CENTER,
        spaceAfter=6
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontName='DejaVuSans',
        fontSize=11,
        textColor=colors.grey,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    
    # Başlık
    if leaderboard_type == 'stars':
        title = f"⭐ Denemenin Yıldızları - {grade}. Sınıflar"
        subtitle = "Ortalama Puana Göre Sıralama"
    else:
        title = f"🚀 Denemenin Yükselenleri - {grade}. Sınıflar"
        subtitle = "Son İki Deneme Arasındaki Gelişim"
    
    story.append(Paragraph(title, title_style))
    story.append(Paragraph(subtitle, subtitle_style))
    story.append(Paragraph(f"Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}", subtitle_style))
    story.append(Spacer(1, 0.3*inch))
    
    # Tablo verisi
    if leaderboard_type == 'stars':
        table_data = [['Sıra', 'Öğrenci', 'Sınıf', 'Deneme\nSayısı', 'En\nDüşük', 'Ortalama', 'En\nYüksek']]
        for student in leaderboard_data:
            table_data.append([
                str(student['rank']),
                student['student_name'],
                student['class_name'],
                str(student['exam_count']),
                str(student['min_score']),
                str(student['avg_score']),
                str(student['max_score'])
            ])
        col_widths = [0.6*inch, 2*inch, 0.8*inch, 0.9*inch, 0.9*inch, 1*inch, 1*inch]
    else:
        table_data = [['Sıra', 'Öğrenci', 'Sınıf', 'Son Puan', 'Değişim']]
        for student in leaderboard_data:
            change_str = f"+{student['score_change']}" if student['score_change'] > 0 else str(student['score_change'])
            table_data.append([
                str(student['rank']),
                student['student_name'],
                student['class_name'],
                str(student['latest_score']),
                change_str
            ])
        col_widths = [0.8*inch, 2.5*inch, 1*inch, 1.2*inch, 1.2*inch]
    
    # Tablo oluştur
    table = Table(table_data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'DejaVuSans-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTNAME', (0, 1), (-1, -1), 'DejaVuSans'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')])
    ]))
    
    # Sıra numaralarına renk ekle (ilk 3'e özel)
    for idx, student in enumerate(leaderboard_data, 1):
        row_idx = idx
        if student['rank'] == 1:
            table.setStyle(TableStyle([('BACKGROUND', (0, row_idx), (0, row_idx), colors.HexColor('#ffd700'))]))  # Gold
        elif student['rank'] == 2:
            table.setStyle(TableStyle([('BACKGROUND', (0, row_idx), (0, row_idx), colors.HexColor('#c0c0c0'))]))  # Silver
        elif student['rank'] == 3:
            table.setStyle(TableStyle([('BACKGROUND', (0, row_idx), (0, row_idx), colors.HexColor('#cd7f32'))]))  # Bronze
    
    story.append(table)
    story.append(Spacer(1, 0.3*inch))
    
    # Footer
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontName='DejaVuSans',
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    story.append(Paragraph("AMEO LMS - Öğrenci Başarı Takip Sistemi", footer_style))
    
    doc.build(story)
    buffer.seek(0)
    return buffer


@app.route("/api/leaderboards/stars/pdf", methods=["GET"])
@login_required
def export_stars_pdf():
    """Denemenin Yıldızları PDF Export (Sadece Admin/Teacher)"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Bu işlem için yetkiniz yok"}), 403
    
    grade = request.args.get('grade', '5')
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT 
                u.id,
                u.full_name,
                u.class_name,
                COUNT(pe.id) as exam_count,
                ROUND(AVG(pe.lgs_score), 2) as avg_score,
                ROUND(MIN(pe.lgs_score), 2) as min_score,
                ROUND(MAX(pe.lgs_score), 2) as max_score
            FROM users u
            INNER JOIN practice_exams pe ON u.id = pe.student_id
            WHERE u.role = 'student' 
                AND u.class_name IS NOT NULL
                AND u.class_name LIKE %s
                AND pe.lgs_score IS NOT NULL
            GROUP BY u.id, u.full_name, u.class_name
            HAVING COUNT(pe.id) > 0
            ORDER BY avg_score DESC
        """, (grade + '%',))
        
        students = cur.fetchall()
        cur.close()
        conn.close()
        
        leaderboard = []
        for idx, student in enumerate(students, 1):
            avg = float(student['avg_score']) if student['avg_score'] else 0
            min_score = float(student['min_score']) if student['min_score'] else 0
            max_score = float(student['max_score']) if student['max_score'] else 0
            
            if avg >= 400:
                color = 'green'
            elif avg >= 350:
                color = 'yellow'
            else:
                color = 'red'
            
            leaderboard.append({
                'rank': idx,
                'student_name': student['full_name'],  # Tam isim (maskeleme YOK)
                'class_name': student['class_name'],
                'exam_count': student['exam_count'],
                'avg_score': avg,
                'min_score': min_score,
                'max_score': max_score,
                'color': color
            })
        
        if not leaderboard:
            return jsonify({"error": "Bu sınıf için veri bulunamadı"}), 404
        
        # PDF oluştur
        pdf_buffer = build_leaderboard_pdf('stars', grade, leaderboard)
        
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=False,
            download_name=f'yildizlar_{grade}_sinif_{datetime.now().strftime("%Y%m%d")}.pdf'
        )
    
    except Exception as e:
        logger.error(f"Stars PDF export error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/leaderboards/risers/pdf", methods=["GET"])
@login_required
def export_risers_pdf():
    """Denemenin Yükselenleri PDF Export (Sadece Admin/Teacher)"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Bu işlem için yetkiniz yok"}), 403
    
    grade = request.args.get('grade', '5')
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            WITH student_exams AS (
                SELECT 
                    u.id,
                    u.full_name,
                    u.class_name,
                    pe.exam_number,
                    pe.lgs_score,
                    ROW_NUMBER() OVER (PARTITION BY u.id ORDER BY pe.created_at DESC) as rn
                FROM users u
                INNER JOIN practice_exams pe ON u.id = pe.student_id
                WHERE u.role = 'student'
                    AND u.class_name IS NOT NULL
                    AND u.class_name LIKE %s
                    AND pe.lgs_score IS NOT NULL
            ),
            latest_two AS (
                SELECT 
                    id,
                    full_name,
                    class_name,
                    MAX(CASE WHEN rn = 1 THEN lgs_score END) as latest_score,
                    MAX(CASE WHEN rn = 2 THEN lgs_score END) as previous_score,
                    MAX(CASE WHEN rn = 1 THEN exam_number END) as latest_exam
                FROM student_exams
                WHERE rn <= 2
                GROUP BY id, full_name, class_name
                HAVING MAX(CASE WHEN rn = 2 THEN lgs_score END) IS NOT NULL
            )
            SELECT 
                id,
                full_name,
                class_name,
                latest_score,
                previous_score,
                ROUND(latest_score - previous_score, 2) as score_change,
                latest_exam
            FROM latest_two
            ORDER BY score_change DESC
        """, (grade + '%',))
        
        students = cur.fetchall()
        cur.close()
        conn.close()
        
        leaderboard = []
        for idx, student in enumerate(students, 1):
            change = float(student['score_change']) if student['score_change'] else 0
            
            if change >= 25:
                color = 'green'
            elif change >= 15:
                color = 'yellow'
            else:
                color = 'red'
            
            leaderboard.append({
                'rank': idx,
                'student_name': student['full_name'],  # Tam isim (maskeleme YOK)
                'class_name': student['class_name'],
                'latest_score': float(student['latest_score']) if student['latest_score'] else 0,
                'previous_score': float(student['previous_score']) if student['previous_score'] else 0,
                'score_change': change,
                'color': color
            })
        
        if not leaderboard:
            return jsonify({"error": "Bu sınıf için veri bulunamadı"}), 404
        
        # PDF oluştur
        pdf_buffer = build_leaderboard_pdf('risers', grade, leaderboard)
        
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=False,
            download_name=f'yukselenler_{grade}_sinif_{datetime.now().strftime("%Y%m%d")}.pdf'
        )
    
    except Exception as e:
        logger.error(f"Risers PDF export error: {e}")
        return jsonify({"error": str(e)}), 500

# ==================== LEADERBOARD API SONU ====================

# ==================== SURVEY/ANKET YÖNETİMİ ====================

@app.route('/admin/surveys')
@login_required
def admin_surveys():
    """Admin anket yönetim sayfası"""
    if current_user.role != 'admin':
        return redirect(url_for('index'))
    return render_template('admin_surveys.html')

@app.route('/teacher/surveys')
@login_required
def teacher_surveys():
    """Öğretmen anket yönetim sayfası"""
    if current_user.role not in ['admin', 'teacher']:
        return redirect(url_for('index'))
    return render_template('teacher_surveys.html')

@app.route('/student/error-report')
@login_required
def student_error_report():
    """Öğrenci hata karnesi sayfası"""
    if current_user.role != 'student':
        return redirect(url_for('index'))
    return render_template('student_error_report.html')

@app.route('/student/report-cards')
@login_required
def student_report_cards():
    """Öğrenci karne analizi sayfası - kendi verilerini görebilir"""
    if current_user.role != 'student':
        return redirect(url_for('index'))
    return render_template('student_report_cards.html')


@app.route('/student/surveys')
@login_required
def student_surveys():
    """Öğrenci anket sayfası"""
    if current_user.role != 'student':
        return redirect(url_for('index'))
    return render_template('student_surveys.html')

# Survey API Endpoints

@app.route('/api/surveys/create', methods=['POST'])
@login_required
def create_survey():
    """Yeni anket oluştur (Admin/Teacher)"""
    try:
        if current_user.role not in ['admin', 'teacher']:
            return jsonify({"error": "Yetkisiz erişim"}), 403
        
        data = request.json
        title = data.get('title')
        description = data.get('description', '')
        target_role = data.get('target_role', 'student')
        # Çoklu sınıf desteği - array olarak gelir, JSON olarak saklanır
        target_classes = data.get('target_classes', [])
        # Eski format uyumluluğu (tek sınıf)
        if not target_classes and data.get('target_class'):
            target_classes = [data.get('target_class')]
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        is_anonymous = data.get('is_anonymous', False)
        questions = data.get('questions', [])
        
        if not title or not start_date or not end_date or not questions:
            return jsonify({"error": "Eksik alan"}), 400
        
        if not target_classes:
            return jsonify({"error": "En az bir hedef sınıf seçmelisiniz"}), 400
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # target_class sütununa JSON array olarak kaydet
        target_class_json = json.dumps(target_classes)
        
        # Anket oluştur
        cur.execute("""
            INSERT INTO surveys (title, description, created_by, creator_role, target_role, target_class, start_date, end_date, is_anonymous)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (title, description, current_user.id, current_user.role, target_role, target_class_json, start_date, end_date, is_anonymous))
        
        survey_id = cur.fetchone()['id']
        
        # Soruları ekle
        for idx, q in enumerate(questions, 1):
            cur.execute("""
                INSERT INTO survey_questions (survey_id, question_text, question_type, options, is_required, question_order)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (survey_id, q['text'], q['type'], json.dumps(q.get('options', [])), q.get('required', True), idx))
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "survey_id": survey_id}), 201
    
    except Exception as e:
        logger.error(f"Survey creation error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/surveys/list', methods=['GET'])
@login_required
def list_surveys():
    """Ankletleri listele (role-based)"""
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        if current_user.role in ['admin', 'teacher']:
            # Admin/Teacher kendi oluşturduğu anketleri görür
            cur.execute("""
                SELECT s.*, 
                    (SELECT COUNT(*) FROM survey_questions WHERE survey_id = s.id) as question_count,
                    (SELECT COUNT(DISTINCT student_id) FROM survey_responses WHERE survey_id = s.id) as response_count
                FROM surveys s
                WHERE s.created_by = %s
                ORDER BY s.created_at DESC
            """, (current_user.id,))
        else:
            # Student kendine gönderilen anketleri görür
            # target_class artık JSON array olarak saklanıyor, öğrencinin sınıfının bu array içinde olup olmadığını kontrol et
            cur.execute("""
                SELECT s.*, 
                    (SELECT COUNT(*) FROM survey_questions WHERE survey_id = s.id) as question_count,
                    (SELECT COUNT(*) FROM survey_responses WHERE survey_id = s.id AND student_id = %s) as my_responses
                FROM surveys s
                WHERE s.target_role = 'student'
                    AND (s.target_class IS NULL OR s.target_class::jsonb ? %s)
                    AND s.status = 'active'
                    AND NOW() BETWEEN s.start_date AND s.end_date
                ORDER BY s.created_at DESC
            """, (current_user.id, current_user.class_name))
        
        surveys = cur.fetchall()
        cur.close()
        conn.close()
        
        return jsonify({"surveys": surveys}), 200
    
    except Exception as e:
        logger.error(f"Survey list error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/surveys/<int:survey_id>', methods=['GET'])
@login_required
def get_survey_details(survey_id):
    """Anket detaylarını getir"""
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Anket bilgisi
        cur.execute("SELECT * FROM surveys WHERE id = %s", (survey_id,))
        survey = cur.fetchone()
        
        if not survey:
            return jsonify({"error": "Anket bulunamadı"}), 404
        
        # Soruları getir
        cur.execute("""
            SELECT * FROM survey_questions 
            WHERE survey_id = %s 
            ORDER BY question_order
        """, (survey_id,))
        questions = cur.fetchall()
        
        cur.close()
        conn.close()
        
        return jsonify({"survey": survey, "questions": questions}), 200
    
    except Exception as e:
        logger.error(f"Survey details error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/surveys/<int:survey_id>/submit', methods=['POST'])
@login_required
def submit_survey_response(survey_id):
    """Anket yanıtı gönder (Student)"""
    try:
        if current_user.role != 'student':
            return jsonify({"error": "Yetkisiz erişim"}), 403
        
        data = request.json
        responses = data.get('responses', [])
        
        if not responses:
            return jsonify({"error": "Yanıt bulunamadı"}), 400
        
        conn = get_db()
        cur = conn.cursor()
        
        # Yanıtları kaydet
        for r in responses:
            cur.execute("""
                INSERT INTO survey_responses (survey_id, question_id, student_id, answer_text, answer_option)
                VALUES (%s, %s, %s, %s, %s)
                ON CONFLICT (survey_id, question_id, student_id) 
                DO UPDATE SET answer_text = EXCLUDED.answer_text, answer_option = EXCLUDED.answer_option
            """, (survey_id, r['question_id'], current_user.id, r.get('answer_text'), r.get('answer_option')))
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True}), 200
    
    except Exception as e:
        logger.error(f"Survey submit error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/surveys/<int:survey_id>/results', methods=['GET'])
@login_required
def get_survey_results(survey_id):
    """Anket sonuçlarını getir (Admin/Teacher)"""
    try:
        if current_user.role not in ['admin', 'teacher']:
            return jsonify({"error": "Yetkisiz erişim"}), 403
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Anket bilgisi
        cur.execute("SELECT * FROM surveys WHERE id = %s AND created_by = %s", (survey_id, current_user.id))
        survey = cur.fetchone()
        
        if not survey:
            return jsonify({"error": "Anket bulunamadı"}), 404
        
        # Soruları ve yanıtları getir
        cur.execute("""
            SELECT 
                q.id, q.question_text, q.question_type, q.options,
                COUNT(DISTINCT r.student_id) as response_count
            FROM survey_questions q
            LEFT JOIN survey_responses r ON q.id = r.question_id
            WHERE q.survey_id = %s
            GROUP BY q.id, q.question_text, q.question_type, q.options
            ORDER BY q.question_order
        """, (survey_id,))
        questions = cur.fetchall()
        
        results = []
        for q in questions:
            # Her soru için yanıtları getir
            cur.execute("""
                SELECT answer_text, answer_option, COUNT(*) as count
                FROM survey_responses
                WHERE question_id = %s
                GROUP BY answer_text, answer_option
                ORDER BY count DESC
            """, (q['id'],))
            answers = cur.fetchall()
            
            results.append({
                'question': q,
                'answers': answers
            })
        
        cur.close()
        conn.close()
        
        return jsonify({"survey": survey, "results": results}), 200
    
    except Exception as e:
        logger.error(f"Survey results error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/surveys/<int:survey_id>/delete', methods=['DELETE'])
@login_required
def delete_survey(survey_id):
    """Anket sil (Admin/Teacher)"""
    try:
        if current_user.role not in ['admin', 'teacher']:
            return jsonify({"error": "Yetkisiz erişim"}), 403
        
        conn = get_db()
        cur = conn.cursor()
        
        cur.execute("DELETE FROM surveys WHERE id = %s AND created_by = %s", (survey_id, current_user.id))
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True}), 200
    
    except Exception as e:
        logger.error(f"Survey delete error: {e}")
        return jsonify({"error": str(e)}), 500

# ==================== SURVEY/ANKET YÖNETİMİ SONU ====================

# ==================== PDF RAPOR FONKSİYONLARI ====================

def build_activity_report_pdf(activity_data):
    """Kullanıcı aktivite raporu için PDF oluşturur"""
    import io
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.units import inch
    from datetime import datetime
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch)
    story = []
    styles = getSampleStyleSheet()
    
    # Başlık stili
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName='DejaVuSans-Bold',
        fontSize=20,
        textColor=colors.HexColor('#667eea'),
        alignment=TA_CENTER,
        spaceAfter=6
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontName='DejaVuSans',
        fontSize=12,
        textColor=colors.grey,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    
    # Başlık
    story.append(Paragraph("📊 Kullanıcı Aktivite Raporu", title_style))
    story.append(Paragraph(f"Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}", subtitle_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Tablo verisi
    table_data = [['Kullanıcı Adı', 'Tam Adı', 'Rol', 'Sınıf', 'Durum', 'Son Giriş', 'Son Çıkış', 'Oturum', 'Süre']]
    
    for user in activity_data:
        last_login = user.get('last_login', '')
        if last_login:
            try:
                last_login = datetime.fromisoformat(last_login).strftime('%d.%m %H:%M')
            except:
                last_login = '-'
        else:
            last_login = '-'
        
        last_logout = user.get('last_logout', '')
        if last_logout:
            try:
                last_logout = datetime.fromisoformat(last_logout).strftime('%d.%m %H:%M')
            except:
                last_logout = '-'
        else:
            last_logout = '-'
        
        role_map = {'admin': 'Admin', 'teacher': 'Öğretmen', 'student': 'Öğrenci'}
        role_text = role_map.get(user.get('role', ''), user.get('role', ''))
        
        status_icon = '🟢' if 'Aktif' in user.get('status', '') else '⚪'
        
        table_data.append([
            user.get('username', ''),
            user.get('full_name', ''),
            role_text,
            user.get('class_name', '') or '-',
            status_icon,
            last_login,
            last_logout,
            str(user.get('total_sessions', 0)),
            f"{user.get('total_duration_hours', 0)}s {user.get('total_duration_minutes', 0)}d"
        ])
    
    # Tablo oluştur
    col_widths = [1.2*inch, 1.8*inch, 1*inch, 0.7*inch, 0.6*inch, 1*inch, 1*inch, 0.7*inch, 0.9*inch]
    table = Table(table_data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'DejaVuSans-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTNAME', (0, 1), (-1, -1), 'DejaVuSans'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')])
    ]))
    
    story.append(table)
    story.append(Spacer(1, 0.3*inch))
    
    # Footer
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontName='DejaVuSans',
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    story.append(Paragraph("AMEO LMS - Kullanıcı Aktivite İzleme Sistemi", footer_style))
    
    doc.build(story)
    buffer.seek(0)
    return buffer


def build_survey_report_pdf(survey, questions_with_stats):
    """Anket sonuçları için renkli PDF rapor oluşturur"""
    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.units import inch
    from datetime import datetime
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch, leftMargin=0.7*inch, rightMargin=0.7*inch)
    story = []
    styles = getSampleStyleSheet()
    
    # Özel stiller
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName='DejaVuSans-Bold',
        fontSize=18,
        textColor=colors.HexColor('#667eea'),
        alignment=TA_CENTER,
        spaceAfter=6
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontName='DejaVuSans',
        fontSize=11,
        textColor=colors.grey,
        alignment=TA_CENTER,
        spaceAfter=15
    )
    
    question_style = ParagraphStyle(
        'Question',
        parent=styles['Normal'],
        fontName='DejaVuSans-Bold',
        fontSize=12,
        textColor=colors.HexColor('#333333'),
        spaceAfter=10,
        leftIndent=10
    )
    
    # Başlık
    story.append(Paragraph(f"📋 {survey['title']}", title_style))
    if survey.get('description'):
        story.append(Paragraph(survey['description'], subtitle_style))
    story.append(Paragraph(f"Oluşturulma: {datetime.now().strftime('%d.%m.%Y %H:%M')}", subtitle_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Anket bilgileri
    start_date_str = '-'
    if survey.get('start_date'):
        try:
            start_date_str = datetime.fromisoformat(survey['start_date']).strftime('%d.%m.%Y')
        except (ValueError, TypeError):
            start_date_str = '-'
    
    end_date_str = '-'
    if survey.get('end_date'):
        try:
            end_date_str = datetime.fromisoformat(survey['end_date']).strftime('%d.%m.%Y')
        except (ValueError, TypeError):
            end_date_str = '-'
    
    info_data = [
        ['Başlangıç', 'Bitiş', 'Hedef', 'Yanıt Sayısı'],
        [
            start_date_str,
            end_date_str,
            survey.get('target_class') or 'Tüm Sınıflar',
            str(survey.get('response_count', 0))
        ]
    ]
    
    info_table = Table(info_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f0f0f0')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#666666')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'DejaVuSans-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTNAME', (0, 1), (-1, -1), 'DejaVuSans'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8)
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Sorular ve cevaplar
    for idx, q_data in enumerate(questions_with_stats, 1):
        q = q_data['question']
        answers = q_data['answers']
        
        # Soru başlığı
        story.append(Paragraph(f"Soru {idx}: {q['question_text']}", question_style))
        story.append(Spacer(1, 0.1*inch))
        
        # Cevap tipi
        if q['question_type'] == 'text':
            # Metin cevaplar - liste şeklinde
            if answers:
                answer_data = [['Cevap', 'Adet']]
                for ans in answers[:10]:  # İlk 10 cevap
                    text = ans.get('answer_text', '-')
                    if len(text) > 50:
                        text = text[:50] + '...'
                    answer_data.append([text, str(ans.get('count', 0))])
                
                ans_table = Table(answer_data, colWidths=[5*inch, 1*inch])
                ans_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e3f2fd')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1976d2')),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'DejaVuSans-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('FONTNAME', (0, 1), (-1, -1), 'DejaVuSans'),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fafafa')])
                ]))
                story.append(ans_table)
            else:
                story.append(Paragraph("Henüz cevap yok", subtitle_style))
        
        elif q['question_type'] == 'choice':
            # Çoktan seçmeli - istatistikli tablo
            if answers:
                total_responses = sum(ans.get('count', 0) for ans in answers)
                answer_data = [['Seçenek', 'Adet', 'Yüzde', 'Grafik']]
                
                colors_list = [
                    colors.HexColor('#4caf50'),
                    colors.HexColor('#2196f3'),
                    colors.HexColor('#ff9800'),
                    colors.HexColor('#f44336'),
                    colors.HexColor('#9c27b0')
                ]
                
                for ans_idx, ans in enumerate(answers):
                    count = ans.get('count', 0)
                    option = ans.get('answer_option', '-')
                    percentage = (count / total_responses * 100) if total_responses > 0 else 0
                    bar_width = int(percentage / 5)  # Her 5% için bir █
                    bar = '█' * bar_width
                    
                    answer_data.append([
                        option,
                        str(count),
                        f"{percentage:.1f}%",
                        bar
                    ])
                
                ans_table = Table(answer_data, colWidths=[2.5*inch, 0.8*inch, 1*inch, 1.7*inch])
                ans_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e8f5e9')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#2e7d32')),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('ALIGN', (1, 0), (2, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'DejaVuSans-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('FONTNAME', (0, 1), (-1, -1), 'DejaVuSans'),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fafafa')])
                ]))
                
                # Her satıra farklı renk ekle
                for i, ans in enumerate(answers, 1):
                    color = colors_list[i % len(colors_list)]
                    ans_table.setStyle(TableStyle([
                        ('TEXTCOLOR', (3, i), (3, i), color)
                    ]))
                
                story.append(ans_table)
            else:
                story.append(Paragraph("Henüz cevap yok", subtitle_style))
        
        elif q['question_type'] == 'rating':
            # Puanlama - ortalama ve dağılım
            if answers:
                total_responses = sum(ans.get('count', 0) for ans in answers)
                total_points = sum(int(ans.get('answer_text', 0)) * ans.get('count', 0) for ans in answers if ans.get('answer_text', '').isdigit())
                avg_rating = total_points / total_responses if total_responses > 0 else 0
                
                # Ortalama göster
                avg_text = f"⭐ Ortalama Puan: {avg_rating:.2f} / 5.0 ({total_responses} yanıt)"
                story.append(Paragraph(avg_text, subtitle_style))
                
                # Dağılım tablosu
                rating_data = [['Puan', 'Adet', 'Yüzde', 'Grafik']]
                for rating in [5, 4, 3, 2, 1]:
                    count = next((ans.get('count', 0) for ans in answers if ans.get('answer_text') == str(rating)), 0)
                    percentage = (count / total_responses * 100) if total_responses > 0 else 0
                    stars = '⭐' * rating
                    bar_width = int(percentage / 5)
                    bar = '█' * bar_width
                    
                    rating_data.append([
                        stars,
                        str(count),
                        f"{percentage:.1f}%",
                        bar
                    ])
                
                rating_table = Table(rating_data, colWidths=[1.5*inch, 0.8*inch, 1*inch, 2.7*inch])
                rating_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fff3e0')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#e65100')),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('ALIGN', (1, 0), (2, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'DejaVuSans-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('FONTNAME', (0, 1), (-1, -1), 'DejaVuSans'),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#fafafa')])
                ]))
                story.append(rating_table)
            else:
                story.append(Paragraph("Henüz cevap yok", subtitle_style))
        
        story.append(Spacer(1, 0.25*inch))
    
    # Footer
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontName='DejaVuSans',
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    story.append(Spacer(1, 0.2*inch))
    story.append(Paragraph("AMEO LMS - Anket Yönetim Sistemi", footer_style))
    
    doc.build(story)
    buffer.seek(0)
    return buffer


@app.route('/api/admin/user-activity/pdf', methods=['GET'])
@login_required
def export_user_activity_pdf():
    """Kullanıcı aktivite raporunu PDF olarak indir (Sadece Admin)"""
    if current_user.role != 'admin':
        return jsonify({"error": "Bu işlem için yetkiniz yok"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Tüm kullanıcıları getir
        cur.execute("""
            SELECT 
                u.id,
                u.username,
                u.full_name,
                u.role,
                u.class_name,
                u.last_login_at as last_login,
                u.last_logout_at as last_logout,
                COALESCE(
                    (SELECT COUNT(*) FROM user_sessions WHERE user_id = u.id),
                    0
                ) as total_sessions,
                0 as total_duration_hours,
                0 as total_duration_minutes
            FROM users u
            ORDER BY 
                CASE 
                    WHEN u.role = 'admin' THEN 1
                    WHEN u.role = 'teacher' THEN 2
                    WHEN u.role = 'student' THEN 3
                END,
                u.full_name
        """)
        
        users = cur.fetchall()
        cur.close()
        conn.close()
        
        # Durum belirle
        activity_data = []
        for user in users:
            is_active = user['last_logout'] is None or (user['last_login'] and user['last_login'] > user['last_logout'])
            
            activity_data.append({
                'id': user['id'],
                'username': user['username'],
                'full_name': user['full_name'],
                'role': user['role'],
                'class_name': user['class_name'],
                'status': '🟢 Aktif' if is_active else '⚪ Pasif',
                'last_login': user['last_login'].isoformat() if user['last_login'] else None,
                'last_logout': user['last_logout'].isoformat() if user['last_logout'] else None,
                'total_sessions': user['total_sessions'],
                'total_duration_hours': user['total_duration_hours'],
                'total_duration_minutes': user['total_duration_minutes']
            })
        
        # PDF oluştur
        pdf_buffer = build_activity_report_pdf(activity_data)
        
        return send_file(
            pdf_buffer,
            mimetype='application/pdf',
            as_attachment=False,
            download_name=f'kullanici_aktivite_raporu_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf'
        )
    
    except Exception as e:
        logger.error(f"Activity PDF export error: {e}")
        return jsonify({"error": str(e)}), 500


@app.route('/api/surveys/<int:survey_id>/report/pdf', methods=['GET'])
@login_required
def export_survey_report_excel(survey_id):
    """Anket sonuçlarını Excel olarak indir (Admin/Teacher)"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Bu işlem için yetkiniz yok"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Anket bilgisi
        cur.execute("""
            SELECT * FROM surveys WHERE id = %s AND created_by = %s
        """, (survey_id, current_user.id))
        survey = cur.fetchone()
        
        if not survey:
            cur.close()
            conn.close()
            return jsonify({"error": "Anket bulunamadı"}), 404
        
        # Soruları getir (sıralı)
        cur.execute("""
            SELECT id, question_text, question_type, question_order
            FROM survey_questions
            WHERE survey_id = %s
            ORDER BY question_order
        """, (survey_id,))
        questions = cur.fetchall()
        
        # Tüm yanıtları getir (öğrenci bazlı)
        cur.execute("""
            SELECT 
                r.student_id,
                u.full_name,
                u.class_name,
                r.question_id,
                COALESCE(r.answer_option, r.answer_text) as answer
            FROM survey_responses r
            JOIN users u ON r.student_id = u.id
            WHERE r.survey_id = %s
            ORDER BY u.class_name, u.full_name, r.question_id
        """, (survey_id,))
        responses = cur.fetchall()
        
        cur.close()
        conn.close()
        
        # Öğrenci bazlı cevapları grupla
        student_answers = {}
        for r in responses:
            sid = r['student_id']
            if sid not in student_answers:
                student_answers[sid] = {
                    'full_name': r['full_name'],
                    'class_name': r['class_name'],
                    'answers': {}
                }
            student_answers[sid]['answers'][r['question_id']] = r['answer']
        
        # Excel oluştur
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Anket Sonuçları"
        
        # Başlık stili
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Başlık satırı: Sıra, İsim, Sınıf, Soru1, Soru2, ...
        headers = ["Sıra", "İsim", "Sınıf"]
        for i, q in enumerate(questions, 1):
            headers.append(f"{i}. Soru")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        # Soru metinlerini 2. satıra yaz (referans için)
        ws.cell(row=2, column=1, value="")
        ws.cell(row=2, column=2, value="")
        ws.cell(row=2, column=3, value="")
        for col, q in enumerate(questions, 4):
            cell = ws.cell(row=2, column=col, value=q['question_text'][:50] + "..." if len(q['question_text']) > 50 else q['question_text'])
            cell.font = Font(italic=True, size=9)
            cell.alignment = Alignment(wrap_text=True)
        
        # Veri satırları
        row_num = 3
        for idx, (sid, data) in enumerate(student_answers.items(), 1):
            ws.cell(row=row_num, column=1, value=idx).border = thin_border
            ws.cell(row=row_num, column=2, value=data['full_name']).border = thin_border
            ws.cell(row=row_num, column=3, value=data['class_name']).border = thin_border
            
            for col, q in enumerate(questions, 4):
                answer = data['answers'].get(q['id'], "-")
                cell = ws.cell(row=row_num, column=col, value=answer)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True)
            
            row_num += 1
        
        # Sütun genişlikleri
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 10
        for col in range(4, 4 + len(questions)):
            ws.column_dimensions[chr(64 + col) if col <= 26 else 'A' + chr(64 + col - 26)].width = 20
        
        # Buffer'a kaydet
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'anket_sonuclari_{survey_id}_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        )
    
    except Exception as e:
        logger.error(f"Survey Excel export error: {e}")
        return jsonify({"error": str(e)}), 500

# ==================== PDF RAPOR FONKSİYONLARI SONU ====================

# ==================== ÖĞRETMEN RAPORLAR ====================

@app.route("/teacher/reports")
@login_required
def teacher_reports():
    """Öğretmen raporlar sayfası"""
    if current_user.role != 'teacher':
        return redirect('/ameo_kullanıcı_giriş')
    return render_template('teacher_reports.html')

@app.route("/teacher/api/reports/data")
@login_required
def get_report_data():
    """Sınıf raporları verilerini döndür - 5 sütun format (Doğru-Yanlış-Net-Sınıf Sırası-Okul Sırası)"""
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        class_name = request.args.get('class')
        exam_number = request.args.get('exam', type=int)
        
        if not class_name or not exam_number:
            return jsonify({"error": "Sınıf ve deneme gerekli"}), 400
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Sınıf öğrencileri
        cur.execute("""
            SELECT id, full_name FROM users 
            WHERE class_name = %s AND role = 'student'
            ORDER BY full_name
        """, (class_name,))
        students = cur.fetchall()
        
        result_students = []
        
        for student in students:
            # Öğrencinin deneme sonuçları
            cur.execute("""
                SELECT * FROM practice_exams
                WHERE student_id = %s AND exam_number = %s
            """, (student['id'], exam_number))
            exam_result = cur.fetchone()
            
            if not exam_result:
                continue
            
            subjects = ['turkce', 'matematik', 'fen', 'sosyal', 'ingilizce', 'din']
            
            student_data = {
                'student_name': student['full_name'],
                'lgs_score': float(exam_result['lgs_score'])
            }
            
            # Her ders için 5 sütun: doğru, yanlış, net, sınıf sırası, okul sırası
            for subject in subjects:
                dogru_col = f'{subject}_dogru'
                yanlis_col = f'{subject}_yanlis'
                net_col = f'{subject}_net'
                
                dogru = int(exam_result[dogru_col]) if exam_result[dogru_col] else 0
                yanlis = int(exam_result[yanlis_col]) if exam_result[yanlis_col] else 0
                net = float(exam_result[net_col]) if exam_result[net_col] else 0.0
                
                # Sınıf sıralaması - Whitelist validation
                if net_col not in ['turkce_net', 'matematik_net', 'fen_net', 'sosyal_net', 'ingilizce_net', 'din_net']:
                    continue
                    
                cur.execute(f"""
                    SELECT COUNT(*) as rank
                    FROM practice_exams pe
                    JOIN users u ON pe.student_id = u.id
                    WHERE u.class_name = %s AND pe.exam_number = %s
                    AND pe.{net_col} > %s
                """, (class_name, exam_number, net))
                class_rank = cur.fetchone()['rank'] + 1
                
                # Okul sıralaması (tüm okulda)
                cur.execute(f"""
                    SELECT COUNT(*) as rank
                    FROM practice_exams
                    WHERE exam_number = %s AND {net_col} > %s
                """, (exam_number, net))
                school_rank = cur.fetchone()['rank'] + 1
                
                student_data[f'{subject}_dogru'] = dogru
                student_data[f'{subject}_yanlis'] = yanlis
                student_data[f'{subject}_net'] = net
                student_data[f'{subject}_class_rank'] = class_rank
                student_data[f'{subject}_school_rank'] = school_rank
            
            result_students.append(student_data)
        
        cur.close()
        conn.close()
        
        # PDF indirme için güvenli token oluştur (APK session bypass için)
        pdf_token = generate_pdf_token(current_user.id, class_name, exam_number)
        
        return jsonify({
            "success": True,
            "class_name": class_name,
            "exam_number": exam_number,
            "students": result_students,
            "pdf_token": pdf_token
        })
    
    except Exception as e:
        logger.error(f"Report data error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/teacher/api/reports/pdf")
def download_report_pdf():
    """Sınıf raporu PDF'ini indir - Token tabanlı yetkilendirme (APK uyumlu)"""
    try:
        class_name = request.args.get('class')
        exam_number = request.args.get('exam', type=int)
        token = request.args.get('token')
        
        if not class_name or not exam_number:
            return jsonify({"error": "Sınıf ve deneme gerekli"}), 400
        
        # Yetkilendirme: Ya session ya da token ile
        teacher_id = None
        
        # 1. Session ile giriş kontrolü (web tarayıcı)
        if current_user.is_authenticated and current_user.role == 'teacher':
            teacher_id = current_user.id
        # 2. Token ile kontrol (APK)
        elif token:
            token_data = verify_pdf_token(token)
            if token_data and token_data.get('class_name') == class_name and token_data.get('exam_number') == exam_number:
                teacher_id = token_data.get('teacher_id')
        
        if not teacher_id:
            return jsonify({"error": "Yetkisiz"}), 403
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Öğrencileri al
        cur.execute("""
            SELECT u.id, u.full_name FROM users u
            WHERE u.class_name = %s AND u.role = 'student'
            ORDER BY u.full_name
        """, (class_name,))
        students = cur.fetchall()
        
        if not students:
            cur.close()
            conn.close()
            return jsonify({"error": "Bu sınıfta öğrenci yok"}), 404
        
        student_ids = [s['id'] for s in students]
        
        # Tüm deneme verilerini al
        cur.execute("""
            SELECT * FROM practice_exams 
            WHERE student_id = ANY(%s) AND exam_number = %s
            ORDER BY student_id
        """, (student_ids, exam_number))
        exams = cur.fetchall()
        
        if not exams:
            cur.close()
            conn.close()
            return jsonify({"error": "Bu sınıf/deneme için veri yok"}), 404
        
        # Exam datalarını student_id'ye göre index et
        exam_by_student = {e['student_id']: e for e in exams}
        
        # Tüm öğrencilerin net değerleri
        cur.execute("""
            SELECT 
                turkce_net, matematik_net, fen_net, sosyal_net, 
                ingilizce_net, din_net
            FROM practice_exams
            WHERE exam_number = %s
        """, (exam_number,))
        all_scores = cur.fetchall()
        
        # Sınıftaki öğrencilerin net değerleri
        cur.execute("""
            SELECT 
                turkce_net, matematik_net, fen_net, sosyal_net, 
                ingilizce_net, din_net
            FROM practice_exams pe
            WHERE pe.exam_number = %s 
            AND pe.student_id = ANY(%s)
        """, (exam_number, student_ids))
        class_scores = cur.fetchall()
        
        cur.close()
        conn.close()
        
        # PDF oluştur - TÜRKÇE KARAKTER DESTEĞİ İLE
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=10, leftMargin=10, topMargin=15, bottomMargin=15)
        story = []
        styles = getSampleStyleSheet()
        
        # Başlık - DejaVuSans font kullan (Türkçe karakter desteği)
        title_style = ParagraphStyle(
            'CustomTitle', 
            parent=styles['Heading1'], 
            fontSize=12, 
            textColor=colors.HexColor('#1f2937'), 
            alignment=TA_CENTER,
            fontName='DejaVuSans-Bold'
        )
        story.append(Paragraph(f"{class_name} Sinifi - Deneme {exam_number} Raporu", title_style))
        story.append(Spacer(1, 8))
        
        # Ders renkleri - Her ders için farklı renk
        subject_colors = {
            'turkce': colors.HexColor('#FED7AA'),      # Turuncu
            'matematik': colors.HexColor('#A7F3D0'),   # Yeşil
            'fen': colors.HexColor('#BAE6FD'),         # Mavi
            'sosyal': colors.HexColor('#F5D0A9'),      # Terrakota
            'ingilizce': colors.HexColor('#F0A4D0'),   # Pembe
            'din': colors.HexColor('#E0E7FF')          # Mor
        }
        
        # Para stili - Wrapped ve centered
        cell_style = ParagraphStyle(
            'CellStyle',
            parent=styles['Normal'],
            fontSize=3.5,
            fontName='DejaVuSans',
            alignment=1,  # CENTER
            wordWrap='CJK',
            leading=3
        )
        
        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Normal'],
            fontSize=3.2,
            fontName='DejaVuSans-Bold',
            alignment=1,
            wordWrap='CJK',
            leading=3
        )
        
        # Tablo başlığı
        table_data = []
        header_row = []
        
        # Öğrenci adı başlığı
        header_row.append(Paragraph('Ogrenci<br/>Ad Soyad', header_style))
        
        # Her ders için başlık: D=Dogru, Y=Yanlis, N=Net, SS=Sinif Sirasi, OS=Okul Sirasi
        subjects_info = [
            ('turkce', 'Turkce'),
            ('matematik', 'Matematik'),
            ('fen', 'Fen'),
            ('sosyal', 'Sosyal'),
            ('ingilizce', 'Ingilizce'),
            ('din', 'Din')
        ]
        
        for subject_key, subject_display in subjects_info:
            # D, Y, N, SS, OS için normal text (multi-line)
            for col_name in ['D', 'Y', 'N', 'SS', 'OS']:
                if col_name in ['D', 'Y', 'N']:
                    col_display = {'D': 'Dogru', 'Y': 'Yanlis', 'N': 'Net'}[col_name]
                    text = f'{subject_display}<br/>{col_display}'
                else:
                    col_display = {'SS': 'Sinif<br/>Sirasi', 'OS': 'Okul<br/>Sirasi'}[col_name]
                    text = f'{subject_display}<br/>{col_display}'
                header_row.append(Paragraph(text, header_style))
        
        # LGS başlığı
        header_row.append(Paragraph('LGS', header_style))
        table_data.append(header_row)
        
        # Tablo verileri
        subject_data = [
            ('turkce', 'turkce_dogru', 'turkce_yanlis', 'turkce_net'),
            ('matematik', 'matematik_dogru', 'matematik_yanlis', 'matematik_net'),
            ('fen', 'fen_dogru', 'fen_yanlis', 'fen_net'),
            ('sosyal', 'sosyal_dogru', 'sosyal_yanlis', 'sosyal_net'),
            ('ingilizce', 'ingilizce_dogru', 'ingilizce_yanlis', 'ingilizce_net'),
            ('din', 'din_dogru', 'din_yanlis', 'din_net')
        ]
        
        # LGS puanına göre sıralama (en yüksek ilk)
        students_sorted = sorted(
            students, 
            key=lambda s: float(exam_by_student.get(s['id'], {}).get('lgs_score', 0)) if exam_by_student.get(s['id']) else 0.0,
            reverse=True
        )
        
        for student in students_sorted:
            exam = exam_by_student.get(student['id'])
            if not exam:
                continue
            
            row = [Paragraph(student['full_name'], cell_style)]
            
            for subject_name, dogru_col, yanlis_col, net_col in subject_data:
                dogru = int(exam[dogru_col]) if exam[dogru_col] else 0
                yanlis = int(exam[yanlis_col]) if exam[yanlis_col] else 0
                net = float(exam[net_col]) if exam[net_col] else 0.0
                
                # Sınıf sırası
                class_rank = sum(1 for s in class_scores if float(s[net_col] or 0) > net) + 1
                # Okul sırası
                school_rank = sum(1 for s in all_scores if float(s[net_col] or 0) > net) + 1
                
                row.append(Paragraph(str(dogru), cell_style))
                row.append(Paragraph(str(yanlis), cell_style))
                row.append(Paragraph(f"{net:.1f}", cell_style))
                row.append(Paragraph(str(class_rank), cell_style))
                row.append(Paragraph(str(school_rank), cell_style))
            
            lgs = int(exam['lgs_score']) if exam['lgs_score'] else 0
            row.append(Paragraph(str(lgs), cell_style))
            table_data.append(row)
        
        # Tablo stili - Ders grupları için farklı renkler
        table_style = TableStyle([
            # Başlık satırı
            ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (0, 0), colors.whitesmoke),
            
            # Her ders grubu 5 sütun + başlık sütunu(0)
            # Türkçe: 1-5
            ('BACKGROUND', (1, 0), (5, 0), colors.HexColor('#FED7AA')),
            ('BACKGROUND', (1, 1), (5, -1), colors.HexColor('#FEF3C7')),
            # Matematik: 6-10
            ('BACKGROUND', (6, 0), (10, 0), colors.HexColor('#A7F3D0')),
            ('BACKGROUND', (6, 1), (10, -1), colors.HexColor('#DCFCE7')),
            # Fen: 11-15
            ('BACKGROUND', (11, 0), (15, 0), colors.HexColor('#BAE6FD')),
            ('BACKGROUND', (11, 1), (15, -1), colors.HexColor('#E0F2FE')),
            # Sosyal: 16-20
            ('BACKGROUND', (16, 0), (20, 0), colors.HexColor('#F5D0A9')),
            ('BACKGROUND', (16, 1), (20, -1), colors.HexColor('#FEE2D5')),
            # İngilizce: 21-25
            ('BACKGROUND', (21, 0), (25, 0), colors.HexColor('#F0A4D0')),
            ('BACKGROUND', (21, 1), (25, -1), colors.HexColor('#FCE7F3')),
            # Din: 26-30
            ('BACKGROUND', (26, 0), (30, 0), colors.HexColor('#E0E7FF')),
            ('BACKGROUND', (26, 1), (30, -1), colors.HexColor('#F3F4F6')),
            # LGS: 31
            ('BACKGROUND', (31, 0), (31, 0), colors.HexColor('#1E40AF')),
            ('TEXTCOLOR', (31, 0), (31, 0), colors.whitesmoke),
            ('BACKGROUND', (31, 1), (31, -1), colors.HexColor('#DBEAFE')),
            
            # Genel stil
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'DejaVuSans-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'DejaVuSans'),
            ('LEFTPADDING', (0, 0), (-1, -1), 1),
            ('RIGHTPADDING', (0, 0), (-1, -1), 1),
            ('TOPPADDING', (0, 0), (-1, -1), 1),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
            ('GRID', (0, 0), (-1, -1), 0.2, colors.grey),
            ('ROWHEIGHT', (0, 0), (-1, 0), 35),  # Header row height - daha geniş
            ('ROWHEIGHT', (0, 1), (-1, -1), 18),  # Data rows - daha geniş
        ])
        
        # Sütun genişliklerini optimize et
        # Tüm sütunları eşit genişlikte yap
        num_cols = len(table_data[0]) if table_data else 1
        col_width = (10.5 * inch) / num_cols  # A4 landscape genişliği (margins hariç)
        col_widths = [col_width] * num_cols
        
        table = Table(table_data, colWidths=col_widths, style=table_style, repeatRows=1)
        story.append(table)
        
        doc.build(story)
        buffer.seek(0)
        
        # APK PNG açabiliyor ama PDF açamıyor - PNG'ye convert etmeyi dene
        try:
            # Temp PDF file oluştur
            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as pdf_file:
                pdf_path = pdf_file.name
                buffer.seek(0)
                pdf_file.write(buffer.read())
                pdf_file.flush()
            
            # PNG path
            png_path = pdf_path.replace('.pdf', '.png')
            
            # pdftoppm ile convert et (poppler-utils)
            try:
                subprocess.run(
                    ['pdftoppm', '-png', '-singlefile', pdf_path, png_path.replace('.png', '')],
                    capture_output=True,
                    timeout=10,
                    check=True
                )
                
                # PNG'yi oku ve döndür
                if os.path.exists(png_path):
                    with open(png_path, 'rb') as f:
                        png_buffer = BytesIO(f.read())
                    png_buffer.seek(0)
                    
                    # Temp dosyaları temizle
                    os.unlink(pdf_path)
                    if os.path.exists(png_path):
                        os.unlink(png_path)
                    
                    return send_file(
                        png_buffer,
                        mimetype='image/png',
                        as_attachment=False,
                        download_name=f'{class_name}_Deneme{exam_number}_{datetime.now().strftime("%Y%m%d")}.png'
                    )
            except (subprocess.CalledProcessError, FileNotFoundError, subprocess.TimeoutExpired):
                # pdftoppm yoksa PDF döndür
                logger.warning("pdftoppm not found, returning PDF instead")
                pass
            
            # Cleanup
            if os.path.exists(pdf_path):
                os.unlink(pdf_path)
            
            # Fallback: PDF döndür
            buffer.seek(0)
            return send_file(
                buffer,
                mimetype='application/pdf',
                as_attachment=False,
                download_name=f'{class_name}_Deneme{exam_number}_{datetime.now().strftime("%Y%m%d")}.pdf'
            )
        
        except Exception as e:
            logger.error(f"PNG conversion failed: {str(e)}")
            # En son fallback: PDF döndür
            buffer.seek(0)
            return send_file(
                buffer,
                mimetype='application/pdf',
                as_attachment=False,
                download_name=f'{class_name}_Deneme{exam_number}_{datetime.now().strftime("%Y%m%d")}.pdf'
            )
    
    except Exception as e:
        logger.error(f"Teacher reports PDF: {str(e)}", exc_info=True)
        return jsonify({"error": str(e)}), 500

# ==================== ÖĞRETMEN RAPORLAR SONU ====================

# ==================== KİTAP KURDU SİSTEMİ ====================

@app.route("/teacher/book-worm")
@login_required
def teacher_book_worm():
    if current_user.role != 'teacher':
        return redirect('/login')
    return render_template("teacher_book_worm.html", user=current_user)

@app.route("/student/book-worm")
@login_required
def student_book_worm():
    if current_user.role != 'student':
        return redirect('/login')
    return render_template("student_book_worm.html", user=current_user)

@app.route("/teacher/exam-samples")
@login_required
def teacher_exam_samples():
    if current_user.role != 'teacher':
        return redirect('/login')
    return render_template("teacher_exam_samples.html", user=current_user)

@app.route("/student/exam-samples")
@login_required
def student_exam_samples():
    if current_user.role != 'student':
        return redirect('/login')
    return render_template("student_exam_samples.html", user=current_user)

@app.route("/teacher/documents")
@login_required
def teacher_documents():
    if current_user.role != 'teacher':
        return redirect('/login')
    return render_template("teacher_documents.html", user=current_user)

@app.route("/student/documents")
@login_required
def student_documents():
    if current_user.role != 'student':
        return redirect('/login')
    return render_template("student_documents.html", user=current_user)

@app.route("/api/teacher/exam-samples", methods=["GET", "POST"])
@login_required
def api_teacher_exam_samples():
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    if request.method == "GET":
        cur.execute("""
            SELECT es.*, u.full_name as teacher_name
            FROM exam_samples es
            JOIN users u ON es.teacher_id = u.id
            WHERE es.teacher_id = %s
            ORDER BY es.created_at DESC
        """, (current_user.id,))
        samples = cur.fetchall()
        cur.close()
        conn.close()
        return jsonify(samples)
    
    # POST - yeni örnek ekle
    title = request.form.get('title')
    description = request.form.get('description', '')
    sample_type = request.form.get('sample_type')
    target_classes = request.form.getlist('target_classes')
    
    if not title or not sample_type or not target_classes:
        cur.close()
        conn.close()
        return jsonify({"error": "Başlık, tür ve en az bir sınıf seçilmelidir"}), 400
    
    file_path = None
    link_url = None
    
    if sample_type == 'pdf':
        if 'file' not in request.files:
            cur.close()
            conn.close()
            return jsonify({"error": "PDF dosyası seçilmedi"}), 400
        
        file = request.files['file']
        if file.filename == '':
            cur.close()
            conn.close()
            return jsonify({"error": "Dosya seçilmedi"}), 400
        
        if file and file.filename.lower().endswith('.pdf'):
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            unique_filename = f"exam_sample_{current_user.id}_{timestamp}_{filename}"
            
            upload_folder = os.path.join('static', 'uploads', 'exam_samples')
            os.makedirs(upload_folder, exist_ok=True)
            
            file_path = os.path.join(upload_folder, unique_filename)
            file.save(file_path)
            file_path = '/' + file_path
        else:
            cur.close()
            conn.close()
            return jsonify({"error": "Sadece PDF dosyaları yüklenebilir"}), 400
    
    elif sample_type == 'link':
        link_url = request.form.get('link_url')
        if not link_url:
            cur.close()
            conn.close()
            return jsonify({"error": "Link URL gerekli"}), 400
    
    cur.execute("""
        INSERT INTO exam_samples (teacher_id, title, description, sample_type, file_path, link_url, target_classes)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        RETURNING id
    """, (current_user.id, title, description, sample_type, file_path, link_url, json.dumps(target_classes)))
    
    new_id = cur.fetchone()['id']
    conn.commit()
    cur.close()
    conn.close()
    
    return jsonify({"success": True, "id": new_id, "message": "Yazılı örneği başarıyla eklendi"})

@app.route("/api/teacher/exam-samples/<int:sample_id>", methods=["DELETE"])
@login_required
def api_delete_exam_sample(sample_id):
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    cur.execute("SELECT * FROM exam_samples WHERE id = %s AND teacher_id = %s", (sample_id, current_user.id))
    sample = cur.fetchone()
    
    if not sample:
        cur.close()
        conn.close()
        return jsonify({"error": "Yazılı örneği bulunamadı"}), 404
    
    # PDF dosyasını sil
    if sample['file_path'] and os.path.exists(sample['file_path'].lstrip('/')):
        os.remove(sample['file_path'].lstrip('/'))
    
    cur.execute("DELETE FROM exam_samples WHERE id = %s", (sample_id,))
    conn.commit()
    cur.close()
    conn.close()
    
    return jsonify({"success": True, "message": "Yazılı örneği silindi"})

@app.route("/api/student/exam-samples", methods=["GET"])
@login_required
def api_student_exam_samples():
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    student_class = current_user.class_name
    if not student_class:
        return jsonify([])
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    cur.execute("""
        SELECT es.id, es.title, es.description, es.sample_type, es.file_path, es.link_url, 
               es.created_at, u.full_name as teacher_name
        FROM exam_samples es
        JOIN users u ON es.teacher_id = u.id
        WHERE es.target_classes::jsonb ? %s
        ORDER BY es.created_at DESC
    """, (student_class,))
    
    samples = cur.fetchall()
    cur.close()
    conn.close()
    
    return jsonify(samples)

# ==================== DOKÜMANLAR SİSTEMİ ====================

@app.route("/api/teacher/documents", methods=["GET", "POST"])
@login_required
def api_teacher_documents():
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    if request.method == "GET":
        cur.execute("""
            SELECT es.*, u.full_name as teacher_name
            FROM exam_samples es
            JOIN users u ON es.teacher_id = u.id
            WHERE es.teacher_id = %s
            ORDER BY es.created_at DESC
        """, (current_user.id,))
        documents = cur.fetchall()
        cur.close()
        conn.close()
        return jsonify(documents)
    
    title = request.form.get('title')
    description = request.form.get('description', '')
    sample_type = request.form.get('sample_type')
    target_classes = request.form.getlist('target_classes')
    category = request.form.get('category', 'exam_samples')
    subject = request.form.get('subject', '')
    
    if not title or not sample_type or not target_classes:
        cur.close()
        conn.close()
        return jsonify({"error": "Başlık, tür ve en az bir sınıf seçilmelidir"}), 400
    
    file_path = None
    link_url = None
    
    if sample_type == 'pdf':
        if 'file' not in request.files:
            cur.close()
            conn.close()
            return jsonify({"error": "PDF dosyası seçilmedi"}), 400
        
        file = request.files['file']
        if file.filename == '':
            cur.close()
            conn.close()
            return jsonify({"error": "Dosya seçilmedi"}), 400
        
        if file and file.filename.lower().endswith('.pdf'):
            filename = secure_filename(file.filename)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            unique_filename = f"doc_{current_user.id}_{timestamp}_{filename}"
            
            upload_folder = os.path.join('static', 'uploads', 'documents')
            os.makedirs(upload_folder, exist_ok=True)
            
            file_path = os.path.join(upload_folder, unique_filename)
            file.save(file_path)
            file_path = '/' + file_path
        else:
            cur.close()
            conn.close()
            return jsonify({"error": "Sadece PDF dosyaları yüklenebilir"}), 400
    
    elif sample_type == 'link':
        link_url = request.form.get('link_url')
        if not link_url:
            cur.close()
            conn.close()
            return jsonify({"error": "Link URL gerekli"}), 400
    
    cur.execute("""
        INSERT INTO exam_samples (teacher_id, title, description, sample_type, file_path, link_url, target_classes, category, subject)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        RETURNING id
    """, (current_user.id, title, description, sample_type, file_path, link_url, json.dumps(target_classes), category, subject if subject else None))
    
    new_id = cur.fetchone()['id']
    conn.commit()
    cur.close()
    conn.close()
    
    return jsonify({"success": True, "id": new_id, "message": "Doküman başarıyla eklendi"})

@app.route("/api/teacher/documents/<int:doc_id>", methods=["DELETE"])
@login_required
def api_delete_document(doc_id):
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    cur.execute("SELECT * FROM exam_samples WHERE id = %s AND teacher_id = %s", (doc_id, current_user.id))
    doc = cur.fetchone()
    
    if not doc:
        cur.close()
        conn.close()
        return jsonify({"error": "Doküman bulunamadı"}), 404
    
    if doc['file_path'] and os.path.exists(doc['file_path'].lstrip('/')):
        os.remove(doc['file_path'].lstrip('/'))
    
    cur.execute("DELETE FROM exam_samples WHERE id = %s", (doc_id,))
    conn.commit()
    cur.close()
    conn.close()
    
    return jsonify({"success": True, "message": "Doküman silindi"})

@app.route("/api/student/documents", methods=["GET"])
@login_required
def api_student_documents():
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz erişim"}), 403
    
    student_class = current_user.class_name
    if not student_class:
        return jsonify([])
    
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    cur.execute("""
        SELECT es.id, es.title, es.description, es.sample_type, es.file_path, es.link_url, 
               es.created_at, es.category, es.subject, es.view_count, u.full_name as teacher_name
        FROM exam_samples es
        JOIN users u ON es.teacher_id = u.id
        WHERE es.target_classes::jsonb ? %s
        ORDER BY es.created_at DESC
    """, (student_class,))
    
    documents = cur.fetchall()
    cur.close()
    conn.close()
    
    return jsonify(documents)

@app.route("/api/student/documents/<int:doc_id>/view", methods=["POST"])
@login_required
def api_document_view(doc_id):
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("UPDATE exam_samples SET view_count = COALESCE(view_count, 0) + 1 WHERE id = %s", (doc_id,))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"success": True})
    except:
        return jsonify({"success": False}), 500

# ==================== DOKÜMANLAR SİSTEMİ SONU ====================

@app.route("/api/teacher/book-challenges", methods=["GET"])
@login_required
def get_teacher_book_challenges():
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT bc.*, 
                   (SELECT COUNT(*) FROM book_challenge_submissions WHERE challenge_id = bc.id) as total_submissions,
                   (SELECT COUNT(*) FROM book_challenge_submissions WHERE challenge_id = bc.id AND status = 'pending') as pending_count,
                   (SELECT COUNT(*) FROM book_challenge_submissions WHERE challenge_id = bc.id AND status = 'approved') as approved_count
            FROM book_challenges bc
            WHERE bc.teacher_id = %s
            ORDER BY bc.created_at DESC
        """, (current_user.id,))
        
        challenges = cur.fetchall()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "challenges": challenges})
    except Exception as e:
        logger.error(f"Book challenges fetch error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/teacher/book-challenges", methods=["POST"])
@login_required
def create_book_challenge():
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        data = request.get_json()
        book_title = data.get('book_title', '').strip()
        questions = data.get('questions', [])
        target_class = data.get('target_class', '')
        
        if not book_title:
            return jsonify({"error": "Kitap adı gerekli"}), 400
        
        if not questions or len(questions) == 0:
            return jsonify({"error": "En az 1 soru ekleyin"}), 400
        
        if len(questions) > 10:
            return jsonify({"error": "En fazla 10 soru ekleyebilirsiniz"}), 400
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            INSERT INTO book_challenges (book_title, teacher_id, target_class, questions, status)
            VALUES (%s, %s, %s, %s, 'active')
            RETURNING id
        """, (book_title, current_user.id, target_class, json.dumps(questions)))
        
        result = cur.fetchone()
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Kitap yarışması oluşturuldu", "id": result['id']})
    except Exception as e:
        logger.error(f"Book challenge create error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/teacher/book-challenges/<int:challenge_id>", methods=["DELETE"])
@login_required
def delete_book_challenge(challenge_id):
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("DELETE FROM book_challenges WHERE id = %s AND teacher_id = %s", (challenge_id, current_user.id))
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Kitap yarışması silindi"})
    except Exception as e:
        logger.error(f"Book challenge delete error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/teacher/book-challenges/<int:challenge_id>/submissions", methods=["GET"])
@login_required
def get_challenge_submissions(challenge_id):
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT bcs.*, u.full_name as student_name, u.class_name as student_class,
                   bc.book_title, bc.questions
            FROM book_challenge_submissions bcs
            JOIN users u ON bcs.student_id = u.id
            JOIN book_challenges bc ON bcs.challenge_id = bc.id
            WHERE bcs.challenge_id = %s AND bc.teacher_id = %s
            ORDER BY bcs.submitted_at DESC
        """, (challenge_id, current_user.id))
        
        submissions = cur.fetchall()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "submissions": submissions})
    except Exception as e:
        logger.error(f"Challenge submissions fetch error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/teacher/book-submissions/<int:submission_id>/review", methods=["POST"])
@login_required
def review_book_submission(submission_id):
    if current_user.role != 'teacher':
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        data = request.get_json()
        action = data.get('action')
        rejection_reason = data.get('rejection_reason', '')
        
        if action not in ['approve', 'reject']:
            return jsonify({"error": "Geçersiz işlem"}), 400
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT bcs.*, bc.teacher_id 
            FROM book_challenge_submissions bcs
            JOIN book_challenges bc ON bcs.challenge_id = bc.id
            WHERE bcs.id = %s
        """, (submission_id,))
        
        submission = cur.fetchone()
        
        if not submission or submission['teacher_id'] != current_user.id:
            cur.close()
            conn.close()
            return jsonify({"error": "Yetkisiz"}), 403
        
        status = 'approved' if action == 'approve' else 'rejected'
        
        cur.execute("""
            UPDATE book_challenge_submissions
            SET status = %s, rejection_reason = %s, reviewed_by = %s, reviewed_at = NOW()
            WHERE id = %s
        """, (status, rejection_reason if action == 'reject' else None, current_user.id, submission_id))
        
        conn.commit()
        cur.close()
        conn.close()
        
        message = "Cevap kabul edildi" if action == 'approve' else "Cevap reddedildi"
        return jsonify({"success": True, "message": message})
    except Exception as e:
        logger.error(f"Book submission review error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/student/book-challenges", methods=["GET"])
@login_required
def get_student_book_challenges():
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        student_class = current_user.class_name or ''
        grade_level = student_class[0] if student_class else ''
        
        cur.execute("""
            SELECT bc.*, u.full_name as teacher_name,
                   bcs.id as submission_id, bcs.status as my_status, bcs.rejection_reason
            FROM book_challenges bc
            JOIN users u ON bc.teacher_id = u.id
            LEFT JOIN book_challenge_submissions bcs ON bc.id = bcs.challenge_id AND bcs.student_id = %s
            WHERE bc.status = 'active' 
              AND (bc.target_class = '' OR bc.target_class IS NULL OR bc.target_class = %s OR bc.target_class = %s)
            ORDER BY bc.created_at DESC
        """, (current_user.id, student_class, grade_level))
        
        challenges = cur.fetchall()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "challenges": challenges})
    except Exception as e:
        logger.error(f"Student book challenges fetch error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/student/book-challenges/<int:challenge_id>/submit", methods=["POST"])
@login_required
def submit_book_answers(challenge_id):
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        data = request.get_json()
        answers = data.get('answers', [])
        
        if not answers:
            return jsonify({"error": "Cevaplar gerekli"}), 400
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("SELECT * FROM book_challenges WHERE id = %s AND status = 'active'", (challenge_id,))
        challenge = cur.fetchone()
        
        if not challenge:
            cur.close()
            conn.close()
            return jsonify({"error": "Yarışma bulunamadı"}), 404
        
        cur.execute("""
            SELECT * FROM book_challenge_submissions 
            WHERE challenge_id = %s AND student_id = %s
        """, (challenge_id, current_user.id))
        
        existing = cur.fetchone()
        
        if existing:
            if existing['status'] == 'approved':
                cur.close()
                conn.close()
                return jsonify({"error": "Bu yarışmaya zaten katıldınız ve kabul edildiniz"}), 400
            
            cur.execute("""
                UPDATE book_challenge_submissions
                SET answers = %s, status = 'pending', rejection_reason = NULL, submitted_at = NOW()
                WHERE id = %s
            """, (json.dumps(answers), existing['id']))
        else:
            cur.execute("""
                INSERT INTO book_challenge_submissions (challenge_id, student_id, answers, status)
                VALUES (%s, %s, %s, 'pending')
            """, (challenge_id, current_user.id, json.dumps(answers)))
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Cevaplarınız gönderildi"})
    except Exception as e:
        logger.error(f"Book answers submit error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/book-worm/leaderboard", methods=["GET"])
@login_required
def get_book_worm_leaderboard():
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT u.id, u.full_name, u.class_name, 
                   COUNT(bcs.id) as approved_count
            FROM users u
            JOIN book_challenge_submissions bcs ON u.id = bcs.student_id
            WHERE bcs.status = 'approved'
            GROUP BY u.id, u.full_name, u.class_name
            ORDER BY approved_count DESC
        """)
        
        leaderboard = cur.fetchall()
        cur.close()
        conn.close()
        
        for i, student in enumerate(leaderboard):
            student['rank'] = i + 1
        
        return jsonify({"success": True, "leaderboard": leaderboard})
    except Exception as e:
        logger.error(f"Book worm leaderboard error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/book-worm/my-stats", methods=["GET"])
@login_required
def get_my_book_worm_stats():
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT 
                (SELECT COUNT(*) FROM book_challenge_submissions WHERE student_id = %s AND status = 'approved') as approved_count,
                (SELECT COUNT(*) FROM book_challenge_submissions WHERE student_id = %s AND status = 'pending') as pending_count,
                (SELECT COUNT(*) FROM book_challenge_submissions WHERE student_id = %s AND status = 'rejected') as rejected_count
        """, (current_user.id, current_user.id, current_user.id))
        
        stats = cur.fetchone()
        
        cur.execute("""
            SELECT COUNT(*) + 1 as my_rank
            FROM (
                SELECT student_id, COUNT(*) as cnt
                FROM book_challenge_submissions
                WHERE status = 'approved'
                GROUP BY student_id
                HAVING COUNT(*) > (
                    SELECT COUNT(*) FROM book_challenge_submissions 
                    WHERE student_id = %s AND status = 'approved'
                )
            ) ranked
        """, (current_user.id,))
        
        rank_result = cur.fetchone()
        stats['my_rank'] = rank_result['my_rank'] if stats['approved_count'] > 0 else None
        
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "stats": stats})
    except Exception as e:
        logger.error(f"Book worm stats error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# ==================== YENİ KİTAP KURDU SİSTEMİ (6 SABİT SORU) ====================

# Sabit 6 soru
BOOK_ENTRY_QUESTIONS = [
    "Okuduğunuz kitabın ismi nedir?",
    "Okuduğunuz kitap kaç sayfadır?",
    "Kitapta en beğendiğiniz bölümü en az 5 cümle ile yazınız.",
    "Kitabın yazarı siz olsaydınız bu kitabın sonunu nasıl bitirirdiniz? En az 5 cümle ile yazınız.",
    "Bu kitaptan çıkardığınız dersler nelerdir? En az 5 cümle ile yazınız.",
    "Okuduğunuz kitabın hikaye unsurlarını yazınız."
]

@app.route("/api/book-entries/questions", methods=["GET"])
@login_required
def get_book_entry_questions():
    """Sabit 6 soruyu döndür"""
    return jsonify({"success": True, "questions": BOOK_ENTRY_QUESTIONS})

ACTIVE_SEMESTER = "2025-2026-2"

@app.route("/api/student/book-entries", methods=["GET"])
@login_required
def get_student_book_entries():
    """Öğrencinin kendi kitaplarını getir - sadece aktif dönem"""
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT id, book_title, page_count, answers, status, rejection_reason, submitted_at
            FROM book_entries
            WHERE student_id = %s AND semester = %s
            ORDER BY submitted_at DESC
        """, (current_user.id, ACTIVE_SEMESTER))
        
        entries = cur.fetchall()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "entries": entries})
    except Exception as e:
        logger.error(f"Get student book entries error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/student/book-entries", methods=["POST"])
@login_required
def submit_book_entry():
    """Öğrenci yeni kitap girişi gönderir"""
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        data = request.get_json()
        answers = data.get('answers', [])
        
        if len(answers) != 6:
            return jsonify({"error": "Tüm 6 soru cevaplanmalı"}), 400
        
        book_title = answers[0].strip()
        try:
            page_count = int(answers[1])
        except:
            return jsonify({"error": "Sayfa sayısı sayı olmalı"}), 400
        
        if not book_title:
            return jsonify({"error": "Kitap adı gerekli"}), 400
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            INSERT INTO book_entries (student_id, book_title, page_count, answers, status, semester)
            VALUES (%s, %s, %s, %s, 'pending', %s)
            RETURNING id
        """, (current_user.id, book_title, page_count, json.dumps(answers), ACTIVE_SEMESTER))
        
        new_id = cur.fetchone()['id']
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Kitap girişi gönderildi", "id": new_id})
    except Exception as e:
        logger.error(f"Submit book entry error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/student/book-entries/stats", methods=["GET"])
@login_required
def get_student_book_stats():
    """Öğrencinin kitap istatistiklerini getir"""
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT 
                (SELECT COUNT(*) FROM book_entries WHERE student_id = %s AND status = 'approved' AND semester = %s) as approved_count,
                (SELECT COUNT(*) FROM book_entries WHERE student_id = %s AND status = 'pending' AND semester = %s) as pending_count,
                (SELECT COALESCE(SUM(page_count), 0) FROM book_entries WHERE student_id = %s AND status = 'approved' AND semester = %s) as total_pages
        """, (current_user.id, ACTIVE_SEMESTER, current_user.id, ACTIVE_SEMESTER, current_user.id, ACTIVE_SEMESTER))
        
        stats = cur.fetchone()
        
        # Sıralama hesapla - sadece aktif dönem
        cur.execute("""
            SELECT COUNT(*) + 1 as my_rank
            FROM (
                SELECT student_id, COUNT(*) as cnt
                FROM book_entries
                WHERE status = 'approved' AND semester = %s
                GROUP BY student_id
                HAVING COUNT(*) > (
                    SELECT COUNT(*) FROM book_entries 
                    WHERE student_id = %s AND status = 'approved' AND semester = %s
                )
            ) ranked
        """, (ACTIVE_SEMESTER, current_user.id, ACTIVE_SEMESTER))
        
        rank_result = cur.fetchone()
        stats['my_rank'] = rank_result['my_rank'] if stats['approved_count'] > 0 else None
        
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "stats": stats})
    except Exception as e:
        logger.error(f"Get book stats error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/teacher/book-entries", methods=["GET"])
@login_required
def get_teacher_book_entries():
    """Öğretmenin sınıflarındaki bekleyen kitap girişlerini getir"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        class_filter = request.args.get('class_name', '')
        status_filter = request.args.get('status', 'pending')
        semester_filter = request.args.get('semester', ACTIVE_SEMESTER)
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        query = """
            SELECT be.id, be.book_title, be.page_count, be.answers, be.status, 
                   be.rejection_reason, be.submitted_at, be.semester,
                   u.full_name as student_name, u.class_name
            FROM book_entries be
            JOIN users u ON be.student_id = u.id
            WHERE be.semester = %s
        """
        params = [semester_filter]
        
        if class_filter:
            query += " AND u.class_name = %s"
            params.append(class_filter)
        
        if status_filter:
            query += " AND be.status = %s"
            params.append(status_filter)
        
        query += " ORDER BY be.submitted_at DESC"
        
        cur.execute(query, params)
        entries = cur.fetchall()
        
        # Mevcut dönemleri getir (arşiv seçeneği için)
        cur.execute("SELECT DISTINCT semester FROM book_entries ORDER BY semester DESC")
        semesters = [row['semester'] for row in cur.fetchall()]
        
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "entries": entries, "semesters": semesters, "active_semester": ACTIVE_SEMESTER})
    except Exception as e:
        logger.error(f"Get teacher book entries error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/teacher/book-entries/<int:entry_id>/approve", methods=["POST"])
@login_required
def approve_book_entry(entry_id):
    """Kitap girişini onayla"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            UPDATE book_entries 
            SET status = 'approved', reviewed_by = %s, reviewed_at = NOW()
            WHERE id = %s
            RETURNING student_id
        """, (current_user.id, entry_id))
        
        result = cur.fetchone()
        if not result:
            cur.close()
            conn.close()
            return jsonify({"error": "Giriş bulunamadı"}), 404
        
        # Rozet kontrolü - ilk 3'te ise rozet ver
        student_id = result['student_id']
        check_and_award_book_badge(cur, student_id)
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Kitap girişi onaylandı"})
    except Exception as e:
        logger.error(f"Approve book entry error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/teacher/book-entries/<int:entry_id>/reject", methods=["POST"])
@login_required
def reject_book_entry(entry_id):
    """Kitap girişini reddet"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        data = request.get_json()
        reason = data.get('reason', 'Gerekçe belirtilmedi')
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            UPDATE book_entries 
            SET status = 'rejected', rejection_reason = %s, reviewed_by = %s, reviewed_at = NOW()
            WHERE id = %s
        """, (reason, current_user.id, entry_id))
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Kitap girişi reddedildi"})
    except Exception as e:
        logger.error(f"Reject book entry error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/teacher/book-entries/<int:entry_id>/delete", methods=["DELETE"])
@login_required
def delete_book_entry(entry_id):
    """Kitap girişini sil"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("DELETE FROM book_entries WHERE id = %s", (entry_id,))
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "message": "Kitap girişi silindi"})
    except Exception as e:
        logger.error(f"Delete book entry error: {str(e)}")
        return jsonify({"error": str(e)}), 500

def check_and_award_book_badge(cur, student_id):
    """İlk 3'te ise rozet ver"""
    try:
        cur.execute("""
            SELECT student_id, COUNT(*) as cnt,
                   RANK() OVER (ORDER BY COUNT(*) DESC) as rnk
            FROM book_entries
            WHERE status = 'approved'
            GROUP BY student_id
        """)
        rankings = cur.fetchall()
        
        for r in rankings:
            if r['student_id'] == student_id and r['rnk'] <= 3:
                # Rozet var mı kontrol et
                cur.execute("""
                    SELECT id FROM student_achievements 
                    WHERE student_id = %s AND achievement_type = 'book_worm_new'
                """, (student_id,))
                if not cur.fetchone():
                    cur.execute("""
                        INSERT INTO student_achievements (student_id, achievement_type, achievement_name, description)
                        VALUES (%s, 'book_worm_new', '📚 Kitap Kurdu', 'Kitap Kurdu liderlik tablosunda ilk 3te yer aldın!')
                    """, (student_id,))
                break
    except Exception as e:
        logger.error(f"Book badge check error: {e}")

@app.route("/api/book-entries/leaderboard", methods=["GET"])
@login_required
def get_book_entries_leaderboard():
    """Yeni kitap kurdu liderlik tablosu - sadece aktif dönem"""
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        semester = request.args.get('semester', ACTIVE_SEMESTER)
        
        cur.execute("""
            SELECT u.id, u.full_name, u.class_name, 
                   COUNT(be.id) as approved_count,
                   COALESCE(SUM(be.page_count), 0) as total_pages
            FROM users u
            JOIN book_entries be ON u.id = be.student_id
            WHERE be.status = 'approved' AND be.semester = %s
            GROUP BY u.id, u.full_name, u.class_name
            ORDER BY approved_count DESC, total_pages DESC
        """, (semester,))
        
        leaderboard = cur.fetchall()
        cur.close()
        conn.close()
        
        for i, student in enumerate(leaderboard):
            student['rank'] = i + 1
        
        return jsonify({"success": True, "leaderboard": leaderboard, "semester": semester})
    except Exception as e:
        logger.error(f"Book entries leaderboard error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/book-entries/student/<int:student_id>/details", methods=["GET"])
@login_required
def get_student_book_details(student_id):
    """Öğrencinin okuduğu kitapların detayları"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT u.full_name, u.class_name
            FROM users u WHERE u.id = %s
        """, (student_id,))
        student = cur.fetchone()
        
        if not student:
            cur.close()
            conn.close()
            return jsonify({"error": "Öğrenci bulunamadı"}), 404
        
        cur.execute("""
            SELECT id, book_title, page_count, submitted_at
            FROM book_entries
            WHERE student_id = %s AND status = 'approved'
            ORDER BY submitted_at DESC
        """, (student_id,))
        
        books = cur.fetchall()
        cur.close()
        conn.close()
        
        return jsonify({
            "success": True, 
            "student": student,
            "books": books,
            "total_books": len(books),
            "total_pages": sum(b['page_count'] for b in books)
        })
    except Exception as e:
        logger.error(f"Get student book details error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/teacher/book-entries/report/<class_name>/pdf", methods=["GET"])
@login_required
def get_book_entries_class_report_pdf(class_name):
    """Sınıf bazlı kitap kurdu PDF raporu"""
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT u.full_name, u.class_name,
                   COUNT(be.id) as book_count,
                   COALESCE(SUM(be.page_count), 0) as total_pages,
                   STRING_AGG(be.book_title || ' (' || be.page_count || ' sayfa)', ', ' ORDER BY be.submitted_at DESC) as books
            FROM users u
            LEFT JOIN book_entries be ON u.id = be.student_id AND be.status = 'approved'
            WHERE u.class_name = %s AND u.role = 'student'
            GROUP BY u.id, u.full_name, u.class_name
            ORDER BY book_count DESC, u.full_name
        """, (class_name,))
        
        students = cur.fetchall()
        cur.close()
        conn.close()
        
        if not students:
            return jsonify({"error": "Bu sınıfta öğrenci bulunamadı"}), 404
        
        # PDF oluştur
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        story = []
        styles = getSampleStyleSheet()
        
        # Header
        story.extend(create_pdf_header(styles))
        
        # Başlık
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#667eea'), alignment=TA_CENTER, fontName='DejaVuSans')
        story.append(Paragraph(f"{class_name} Sınıfı - Kitap Kurdu Raporu", title_style))
        story.append(Spacer(1, 15))
        
        # Tarih
        date_style = ParagraphStyle('Date', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER)
        story.append(Paragraph(f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}", date_style))
        story.append(Spacer(1, 20))
        
        # Özet
        total_books = sum(s['book_count'] for s in students)
        total_pages = sum(s['total_pages'] for s in students)
        readers_count = sum(1 for s in students if s['book_count'] > 0)
        
        summary_data = [
            ['Sınıf', class_name],
            ['Toplam Öğrenci', str(len(students))],
            ['Kitap Okuyan Öğrenci', str(readers_count)],
            ['Toplam Okunan Kitap', str(total_books)],
            ['Toplam Sayfa', str(total_pages)]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 3*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f3f4f6')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'DejaVuSans-Bold'),
            ('FONTNAME', (1, 0), (-1, -1), 'DejaVuSans'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Öğrenci listesi
        story.append(Paragraph("Öğrenci Bazlı Kitap Listesi", styles['Heading2']))
        story.append(Spacer(1, 10))
        
        # Kitap hücresi için stil
        book_cell_style = ParagraphStyle(
            'BookCell',
            parent=styles['Normal'],
            fontSize=8,
            fontName='DejaVuSans',
            leading=10,
            wordWrap='CJK'
        )
        
        name_cell_style = ParagraphStyle(
            'NameCell',
            parent=styles['Normal'],
            fontSize=9,
            fontName='DejaVuSans',
            leading=11
        )
        
        header_style = ParagraphStyle(
            'HeaderCell',
            parent=styles['Normal'],
            fontSize=9,
            fontName='DejaVuSans-Bold',
            textColor=colors.whitesmoke
        )
        
        table_data = [[
            Paragraph('Öğrenci Adı', header_style),
            Paragraph('Kitap', header_style),
            Paragraph('Sayfa', header_style),
            Paragraph('Okunan Kitaplar', header_style)
        ]]
        
        for s in students:
            books_text = s['books'] if s['books'] else '-'
            # Kitap listesini satır satır göster (virgülle ayrılmış)
            if books_text != '-':
                books_list = books_text.split(', ')
                books_formatted = '<br/>'.join([f"• {b}" for b in books_list])
            else:
                books_formatted = '-'
            
            table_data.append([
                Paragraph(s['full_name'], name_cell_style),
                Paragraph(str(s['book_count']), name_cell_style),
                Paragraph(str(s['total_pages']), name_cell_style),
                Paragraph(books_formatted, book_cell_style)
            ])
        
        detail_table = Table(table_data, colWidths=[1.6*inch, 0.6*inch, 0.6*inch, 3.7*inch])
        detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (2, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#fefce8')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#fefce8'), colors.white])
        ]))
        story.append(detail_table)
        
        doc.build(story)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=False,
            download_name=f'{class_name}_kitap_kurdu_raporu_{datetime.now().strftime("%Y%m%d")}.pdf'
        )
    except Exception as e:
        logger.error(f"Book entries PDF report error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# ==================== YENİ KİTAP KURDU SİSTEMİ SONU ====================

# ==================== OPTİK FORM YÖNETİMİ ====================

@app.route("/admin/optical-management")
@login_required
def admin_optical_management():
    if current_user.role != 'admin':
        return redirect(url_for('dashboard'))
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT name FROM classes ORDER BY name")
        classes = cur.fetchall()
        cur.close()
        conn.close()
        return render_template('admin_optical_management.html', classes=classes)
    except:
        return render_template('admin_optical_management.html', classes=[])

@app.route("/api/admin/optical/exams", methods=["GET"])
@login_required
def get_optical_exams():
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT oe.*, 
                   (SELECT COUNT(*) FROM optical_student_results WHERE optical_exam_id = oe.id) as result_count
            FROM optical_exams oe
            ORDER BY oe.created_at DESC
        """)
        exams = cur.fetchall()
        
        total = len(exams)
        published = len([e for e in exams if e['is_published']])
        draft = total - published
        total_results = sum(e['result_count'] for e in exams)
        
        cur.close()
        conn.close()
        
        return jsonify({
            "success": True, 
            "exams": exams,
            "stats": {
                "total": total,
                "published": published,
                "draft": draft,
                "total_results": total_results
            }
        })
    except Exception as e:
        logger.error(f"Optical exams fetch error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/admin/optical/exams", methods=["POST"])
@login_required
def create_optical_exam():
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        exam_name = request.form.get('exam_name')
        grade_level = int(request.form.get('grade_level'))
        
        question_counts = {
            'turkce': int(request.form.get('turkce', 20)),
            'matematik': int(request.form.get('matematik', 20)),
            'fen': int(request.form.get('fen', 20)),
            'sosyal': int(request.form.get('sosyal', 10)),
            'din': int(request.form.get('din', 10)),
            'ingilizce': int(request.form.get('ingilizce', 10))
        }
        
        if grade_level in [5, 6]:
            question_counts['turkce'] = 15
            question_counts['matematik'] = 15
            question_counts['fen'] = 15
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT COALESCE(MAX(exam_number), 0) + 1 as next_num
            FROM optical_exams WHERE grade_level = %s
        """, (grade_level,))
        exam_number = cur.fetchone()['next_num']
        
        answer_key_a = None
        answer_key_b = None
        
        if 'answer_key_a' in request.files:
            file_a = request.files['answer_key_a']
            if file_a.filename:
                answer_key_a = parse_answer_key_excel(file_a, question_counts)
        
        if 'answer_key_b' in request.files:
            file_b = request.files['answer_key_b']
            if file_b.filename:
                answer_key_b = parse_answer_key_excel(file_b, question_counts)
        
        cur.execute("""
            INSERT INTO optical_exams (exam_name, exam_number, grade_level, question_counts, 
                                       answer_key_a, answer_key_b, uploaded_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (exam_name, exam_number, grade_level, json.dumps(question_counts),
              json.dumps(answer_key_a) if answer_key_a else None,
              json.dumps(answer_key_b) if answer_key_b else None,
              current_user.id))
        
        exam_id = cur.fetchone()['id']
        
        if 'results_file' in request.files:
            results_file = request.files['results_file']
            if results_file.filename:
                upload_optical_results_file(cur, conn, exam_id, results_file, 
                                            answer_key_a, answer_key_b, question_counts)
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "exam_id": exam_id, "exam_number": exam_number})
    except Exception as e:
        logger.error(f"Create optical exam error: {str(e)}")
        return jsonify({"error": str(e)}), 500

def parse_answer_key_excel(file, question_counts):
    try:
        df = pd.read_excel(file)
        answer_key = {}
        
        subjects = ['turkce', 'matematik', 'fen', 'sosyal', 'din', 'ingilizce']
        subject_names = ['Türkçe', 'Matematik', 'Fen', 'Sosyal', 'Din', 'İngilizce']
        
        for i, subject in enumerate(subjects):
            answers = []
            count = question_counts.get(subject, 10)
            
            for col in df.columns:
                if subject_names[i].lower() in col.lower() or subject in col.lower():
                    row_data = df[col].dropna().tolist()
                    if row_data:
                        answers = [str(a).upper().strip() for a in row_data[:count]]
                        break
            
            if not answers:
                for q in range(1, count + 1):
                    col_name = f"{subject_names[i]}_{q}"
                    if col_name in df.columns:
                        val = df[col_name].iloc[0] if len(df) > 0 else ''
                        answers.append(str(val).upper().strip() if pd.notna(val) else '')
            
            if not answers:
                answers = [''] * count
            
            answer_key[subject] = answers[:count]
        
        return answer_key
    except Exception as e:
        logger.error(f"Parse answer key error: {str(e)}")
        return None

def upload_optical_results_file(cur, conn, exam_id, file, answer_key_a, answer_key_b, question_counts):
    try:
        df = pd.read_excel(file)
        
        df.columns = [str(col).strip() for col in df.columns]
        
        uploaded = 0
        errors = []
        
        for idx, row in df.iterrows():
            try:
                student_no = str(row.get('Öğrenci No', row.get('OgrenciNo', ''))).strip()
                full_name = str(row.get('Ad Soyad', row.get('AdSoyad', ''))).strip().upper()
                class_name = str(row.get('Sınıf', row.get('Sinif', ''))).strip().upper()
                booklet = str(row.get('Kitapçık', row.get('Kitapcik', 'A'))).strip().upper()
                
                if booklet not in ['A', 'B']:
                    booklet = 'A'
                
                cur.execute("""
                    SELECT id, class_name FROM users 
                    WHERE role = 'student' AND (
                        (UPPER(class_name) = %s AND UPPER(full_name) = %s)
                        OR username = %s
                    )
                """, (class_name, full_name, student_no))
                
                student = cur.fetchone()
                
                if not student:
                    errors.append(f"Satır {idx+2}: '{full_name}' ({class_name}) bulunamadı")
                    continue
                
                raw_answers = {}
                subjects = ['turkce', 'matematik', 'fen', 'sosyal', 'din', 'ingilizce']
                subject_prefixes = ['T', 'M', 'F', 'S', 'D', 'I']
                
                for i, subject in enumerate(subjects):
                    answers = []
                    prefix = subject_prefixes[i]
                    count = question_counts.get(subject, 10)
                    
                    for q in range(1, count + 1):
                        col_name = f"{prefix}{q}"
                        val = row.get(col_name, '')
                        if pd.notna(val):
                            answers.append(str(val).upper().strip())
                        else:
                            answers.append('')
                    
                    raw_answers[subject] = answers
                
                answer_key = answer_key_a if booklet == 'A' else (answer_key_b or answer_key_a)
                
                results = calculate_student_results(raw_answers, answer_key, question_counts)
                
                cur.execute("""
                    INSERT INTO optical_student_results 
                    (optical_exam_id, student_id, booklet_type, raw_answers, results,
                     total_correct, total_wrong, total_empty, total_net)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (optical_exam_id, student_id) DO UPDATE SET
                        booklet_type = EXCLUDED.booklet_type,
                        raw_answers = EXCLUDED.raw_answers,
                        results = EXCLUDED.results,
                        total_correct = EXCLUDED.total_correct,
                        total_wrong = EXCLUDED.total_wrong,
                        total_empty = EXCLUDED.total_empty,
                        total_net = EXCLUDED.total_net
                """, (exam_id, student['id'], booklet, json.dumps(raw_answers), 
                      json.dumps(results['by_subject']),
                      results['total_correct'], results['total_wrong'], 
                      results['total_empty'], results['total_net']))
                
                uploaded += 1
                
            except Exception as e:
                errors.append(f"Satır {idx+2}: {str(e)}")
        
        if uploaded > 0:
            calculate_rankings_and_scores(cur, exam_id)
        
        return uploaded, errors
        
    except Exception as e:
        logger.error(f"Upload optical results error: {str(e)}")
        raise e

def calculate_student_results(raw_answers, answer_key, question_counts):
    results = {
        'by_subject': {},
        'total_correct': 0,
        'total_wrong': 0,
        'total_empty': 0,
        'total_net': 0
    }
    
    subjects = ['turkce', 'matematik', 'fen', 'sosyal', 'din', 'ingilizce']
    weights = {'turkce': 4, 'matematik': 4, 'fen': 4, 'sosyal': 1, 'din': 1, 'ingilizce': 1}
    
    for subject in subjects:
        student_answers = raw_answers.get(subject, [])
        correct_answers = answer_key.get(subject, []) if answer_key else []
        
        correct = 0
        wrong = 0
        empty = 0
        
        count = question_counts.get(subject, 10)
        
        for i in range(count):
            student_ans = student_answers[i] if i < len(student_answers) else ''
            correct_ans = correct_answers[i] if i < len(correct_answers) else ''
            
            if not student_ans or student_ans == '-':
                empty += 1
            elif student_ans == correct_ans:
                correct += 1
            else:
                wrong += 1
        
        net = correct - (wrong / 3)
        
        results['by_subject'][subject] = {
            'correct': correct,
            'wrong': wrong,
            'empty': empty,
            'net': round(net, 2),
            'weight': weights[subject]
        }
        
        results['total_correct'] += correct
        results['total_wrong'] += wrong
        results['total_empty'] += empty
        results['total_net'] += net * weights[subject]
    
    results['total_net'] = round(results['total_net'], 2)
    
    return results

def calculate_rankings_and_scores(cur, exam_id):
    cur.execute("""
        SELECT AVG(total_net) as avg_net, STDDEV(total_net) as std_net,
               MAX(total_net) as max_net
        FROM optical_student_results WHERE optical_exam_id = %s
    """, (exam_id,))
    
    stats = cur.fetchone()
    avg_net = float(stats['avg_net'] or 0)
    std_net = float(stats['std_net'] or 1)
    max_net = float(stats['max_net'] or 1)
    
    if std_net == 0:
        std_net = 1
    
    cur.execute("""
        UPDATE optical_exams 
        SET school_average = %s, school_std_dev = %s
        WHERE id = %s
    """, (avg_net, std_net, exam_id))
    
    cur.execute("""
        SELECT osr.id, osr.student_id, osr.total_net, u.class_name
        FROM optical_student_results osr
        JOIN users u ON osr.student_id = u.id
        WHERE osr.optical_exam_id = %s
        ORDER BY osr.total_net DESC
    """, (exam_id,))
    
    results = cur.fetchall()
    
    for rank, result in enumerate(results, 1):
        total_net = float(result['total_net'] or 0)
        score = 250 + (250 * (total_net - avg_net) / max(max_net - avg_net, 1))
        score = max(0, min(500, score))
        
        cur.execute("""
            UPDATE optical_student_results 
            SET school_rank = %s, total_score = %s
            WHERE id = %s
        """, (rank, round(score, 2), result['id']))
    
    cur.execute("""
        SELECT DISTINCT u.class_name
        FROM optical_student_results osr
        JOIN users u ON osr.student_id = u.id
        WHERE osr.optical_exam_id = %s
    """, (exam_id,))
    
    classes = [row['class_name'] for row in cur.fetchall()]
    
    for class_name in classes:
        cur.execute("""
            SELECT osr.id FROM optical_student_results osr
            JOIN users u ON osr.student_id = u.id
            WHERE osr.optical_exam_id = %s AND u.class_name = %s
            ORDER BY osr.total_net DESC
        """, (exam_id, class_name))
        
        class_results = cur.fetchall()
        
        for rank, result in enumerate(class_results, 1):
            cur.execute("""
                UPDATE optical_student_results SET class_rank = %s WHERE id = %s
            """, (rank, result['id']))

@app.route("/api/admin/optical/exams/<int:exam_id>/results", methods=["POST"])
@login_required
def upload_optical_exam_results(exam_id):
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        if 'results_file' not in request.files:
            return jsonify({"error": "Dosya gerekli"}), 400
        
        file = request.files['results_file']
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("SELECT * FROM optical_exams WHERE id = %s", (exam_id,))
        exam = cur.fetchone()
        
        if not exam:
            cur.close()
            conn.close()
            return jsonify({"error": "Sınav bulunamadı"}), 404
        
        answer_key_a = json.loads(exam['answer_key_a']) if exam['answer_key_a'] else None
        answer_key_b = json.loads(exam['answer_key_b']) if exam['answer_key_b'] else None
        question_counts = json.loads(exam['question_counts'])
        
        uploaded, errors = upload_optical_results_file(cur, conn, exam_id, file,
                                                       answer_key_a, answer_key_b, question_counts)
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({
            "success": True, 
            "uploaded": uploaded,
            "errors": errors[:20]
        })
        
    except Exception as e:
        logger.error(f"Upload results error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/admin/optical/exams/<int:exam_id>/analysis", methods=["GET"])
@login_required
def get_optical_exam_analysis(exam_id):
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("SELECT * FROM optical_exams WHERE id = %s", (exam_id,))
        exam = cur.fetchone()
        
        if not exam:
            cur.close()
            conn.close()
            return jsonify({"error": "Sınav bulunamadı"}), 404
        
        cur.execute("""
            SELECT u.class_name, 
                   COUNT(*) as count,
                   AVG(osr.total_net) as average,
                   MAX(osr.total_net) as max,
                   MIN(osr.total_net) as min
            FROM optical_student_results osr
            JOIN users u ON osr.student_id = u.id
            WHERE osr.optical_exam_id = %s
            GROUP BY u.class_name
            ORDER BY u.class_name
        """, (exam_id,))
        
        class_stats = {}
        for row in cur.fetchall():
            class_stats[row['class_name']] = {
                'count': row['count'],
                'average': float(row['average']) if row['average'] else 0,
                'max': float(row['max']) if row['max'] else 0,
                'min': float(row['min']) if row['min'] else 0
            }
        
        cur.execute("""
            SELECT COUNT(*) as total, AVG(total_net) as avg
            FROM optical_student_results WHERE optical_exam_id = %s
        """, (exam_id,))
        
        overall = cur.fetchone()
        
        cur.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "exam": exam,
            "analysis": {
                "class_stats": class_stats,
                "total_students": overall['total'],
                "school_average": float(overall['avg']) if overall['avg'] else 0
            }
        })
        
    except Exception as e:
        logger.error(f"Optical analysis error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/admin/optical/exams/<int:exam_id>/publish", methods=["POST"])
@login_required
def publish_optical_exam(exam_id):
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("SELECT * FROM optical_exams WHERE id = %s", (exam_id,))
        exam = cur.fetchone()
        
        if not exam:
            cur.close()
            conn.close()
            return jsonify({"error": "Sınav bulunamadı"}), 404
        
        cur.execute("""
            SELECT osr.*, u.class_name, u.full_name
            FROM optical_student_results osr
            JOIN users u ON osr.student_id = u.id
            WHERE osr.optical_exam_id = %s
        """, (exam_id,))
        
        results = cur.fetchall()
        
        for result in results:
            result_data = json.loads(result['results'])
            
            subjects_data = {}
            for subject, data in result_data.items():
                subjects_data[subject] = {
                    'dogru': data['correct'],
                    'yanlis': data['wrong'],
                    'bos': data['empty'],
                    'net': data['net']
                }
            
            cur.execute("""
                INSERT INTO practice_exams 
                (student_id, exam_number, exam_date, turkce_dogru, turkce_yanlis, turkce_net,
                 mat_dogru, mat_yanlis, mat_net, fen_dogru, fen_yanlis, fen_net,
                 sosyal_dogru, sosyal_yanlis, sosyal_net, din_dogru, din_yanlis, din_net,
                 ing_dogru, ing_yanlis, ing_net, toplam_net, puan)
                VALUES (%s, %s, CURRENT_DATE, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (student_id, exam_number) DO UPDATE SET
                    turkce_dogru = EXCLUDED.turkce_dogru,
                    turkce_yanlis = EXCLUDED.turkce_yanlis,
                    turkce_net = EXCLUDED.turkce_net,
                    mat_dogru = EXCLUDED.mat_dogru,
                    mat_yanlis = EXCLUDED.mat_yanlis,
                    mat_net = EXCLUDED.mat_net,
                    fen_dogru = EXCLUDED.fen_dogru,
                    fen_yanlis = EXCLUDED.fen_yanlis,
                    fen_net = EXCLUDED.fen_net,
                    sosyal_dogru = EXCLUDED.sosyal_dogru,
                    sosyal_yanlis = EXCLUDED.sosyal_yanlis,
                    sosyal_net = EXCLUDED.sosyal_net,
                    din_dogru = EXCLUDED.din_dogru,
                    din_yanlis = EXCLUDED.din_yanlis,
                    din_net = EXCLUDED.din_net,
                    ing_dogru = EXCLUDED.ing_dogru,
                    ing_yanlis = EXCLUDED.ing_yanlis,
                    ing_net = EXCLUDED.ing_net,
                    toplam_net = EXCLUDED.toplam_net,
                    puan = EXCLUDED.puan
            """, (
                result['student_id'], exam['exam_number'],
                subjects_data.get('turkce', {}).get('dogru', 0),
                subjects_data.get('turkce', {}).get('yanlis', 0),
                subjects_data.get('turkce', {}).get('net', 0),
                subjects_data.get('matematik', {}).get('dogru', 0),
                subjects_data.get('matematik', {}).get('yanlis', 0),
                subjects_data.get('matematik', {}).get('net', 0),
                subjects_data.get('fen', {}).get('dogru', 0),
                subjects_data.get('fen', {}).get('yanlis', 0),
                subjects_data.get('fen', {}).get('net', 0),
                subjects_data.get('sosyal', {}).get('dogru', 0),
                subjects_data.get('sosyal', {}).get('yanlis', 0),
                subjects_data.get('sosyal', {}).get('net', 0),
                subjects_data.get('din', {}).get('dogru', 0),
                subjects_data.get('din', {}).get('yanlis', 0),
                subjects_data.get('din', {}).get('net', 0),
                subjects_data.get('ingilizce', {}).get('dogru', 0),
                subjects_data.get('ingilizce', {}).get('yanlis', 0),
                subjects_data.get('ingilizce', {}).get('net', 0),
                result['total_net'],
                result['total_score']
            ))
        
        cur.execute("UPDATE optical_exams SET is_published = TRUE WHERE id = %s", (exam_id,))
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "published": len(results)})
        
    except Exception as e:
        logger.error(f"Publish optical exam error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/admin/optical/templates/<template_type>")
@login_required
def download_optical_template(template_type):
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        output = io.BytesIO()
        
        if template_type.startswith('answer_key'):
            is_56 = '56' in template_type
            counts = {'turkce': 15, 'matematik': 15, 'fen': 15} if is_56 else {'turkce': 20, 'matematik': 20, 'fen': 20}
            counts.update({'sosyal': 10, 'din': 10, 'ingilizce': 10})
            
            data = {}
            subjects = ['Türkçe', 'Matematik', 'Fen', 'Sosyal', 'Din', 'İngilizce']
            subject_keys = ['turkce', 'matematik', 'fen', 'sosyal', 'din', 'ingilizce']
            
            max_q = max(counts.values())
            
            for i, subject in enumerate(subjects):
                count = counts[subject_keys[i]]
                data[subject] = [''] * max_q
                for q in range(count):
                    data[subject][q] = ''
            
            df = pd.DataFrame(data)
            df.index = [f"Soru {i+1}" for i in range(max_q)]
            df.to_excel(output, index=True)
            
            filename = f"cevap_anahtari_sablonu_{'5-6' if is_56 else '7-8'}_sinif.xlsx"
            
        else:
            is_56 = '56' in template_type
            counts = {'turkce': 15, 'matematik': 15, 'fen': 15} if is_56 else {'turkce': 20, 'matematik': 20, 'fen': 20}
            counts.update({'sosyal': 10, 'din': 10, 'ingilizce': 10})
            
            columns = ['Öğrenci No', 'Ad Soyad', 'Sınıf', 'Kitapçık']
            
            prefixes = [('T', 'turkce'), ('M', 'matematik'), ('F', 'fen'), 
                       ('S', 'sosyal'), ('D', 'din'), ('I', 'ingilizce')]
            
            for prefix, key in prefixes:
                for q in range(1, counts[key] + 1):
                    columns.append(f"{prefix}{q}")
            
            df = pd.DataFrame(columns=columns)
            df.loc[0] = ['12345', 'ÖRNEK ÖĞRENCİ', '8A', 'A'] + [''] * (len(columns) - 4)
            df.to_excel(output, index=False)
            
            filename = f"sonuc_yukleme_sablonu_{'5-6' if is_56 else '7-8'}_sinif.xlsx"
        
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Download template error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/admin/optical/forms/pdf")
@login_required
def download_optical_forms_pdf():
    if current_user.role != 'admin':
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        from reportlab.lib.units import mm
        from reportlab.lib.colors import black, white, gray, lightgrey
        
        class_name = request.args.get('class_name')
        grade = request.args.get('grade', '7')
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT id, username, full_name, class_name FROM users
            WHERE role = 'student' AND class_name = %s
            ORDER BY full_name
        """, (class_name,))
        
        students = cur.fetchall()
        cur.close()
        conn.close()
        
        if not students:
            return jsonify({"error": "Bu sınıfta öğrenci bulunamadı"}), 404
        
        is_56 = grade in ['5', '6']
        counts = {'turkce': 15, 'matematik': 15, 'fen': 15} if is_56 else {'turkce': 20, 'matematik': 20, 'fen': 20}
        counts.update({'sosyal': 10, 'din': 10, 'ingilizce': 10})
        
        buffer = BytesIO()
        from reportlab.pdfgen import canvas
        c = canvas.Canvas(buffer, pagesize=A4)
        page_width, page_height = A4
        
        for student in students:
            student_no = str(student['username']).zfill(4)
            full_name = student['full_name']
            student_class = student['class_name']
            
            c.setFont("Helvetica-Bold", 14)
            c.drawCentredString(page_width/2, page_height - 15*mm, "AYSE MELAHAT ERKIN ORTAOKULU")
            c.setFont("Helvetica-Bold", 11)
            c.drawCentredString(page_width/2, page_height - 21*mm, "LGS DENEME SINAVI OPTIK FORMU")
            
            c.setStrokeColor(black)
            c.setLineWidth(0.5)
            c.rect(10*mm, page_height - 50*mm, page_width - 20*mm, 26*mm)
            
            c.setFont("Helvetica-Bold", 9)
            c.drawString(15*mm, page_height - 32*mm, "AD SOYAD:")
            c.setFont("Helvetica", 10)
            c.drawString(38*mm, page_height - 32*mm, full_name.upper())
            
            c.setFont("Helvetica-Bold", 9)
            c.drawString(15*mm, page_height - 40*mm, "OGRENCI NO:")
            c.setFont("Helvetica-Bold", 12)
            c.drawString(42*mm, page_height - 40*mm, student_no)
            
            c.setFont("Helvetica-Bold", 9)
            c.drawString(75*mm, page_height - 32*mm, "SINIF:")
            c.setFont("Helvetica-Bold", 11)
            c.drawString(88*mm, page_height - 32*mm, student_class)
            
            c.setFont("Helvetica-Bold", 9)
            c.drawString(75*mm, page_height - 40*mm, "KITAPCIK:")
            
            booklet_x = 98*mm
            booklet_y = page_height - 39*mm
            bubble_r = 2.5*mm
            
            c.circle(booklet_x, booklet_y, bubble_r, stroke=1, fill=0)
            c.setFont("Helvetica-Bold", 7)
            c.drawCentredString(booklet_x, booklet_y - 1.5*mm, "A")
            
            c.circle(booklet_x + 12*mm, booklet_y, bubble_r, stroke=1, fill=0)
            c.drawCentredString(booklet_x + 12*mm, booklet_y - 1.5*mm, "B")
            
            no_grid_x = 130*mm
            no_grid_y = page_height - 30*mm
            cell_w = 6*mm
            cell_h = 3.5*mm
            
            c.setFont("Helvetica-Bold", 6)
            c.drawString(no_grid_x, no_grid_y + 3*mm, "OGRENCI NO KODLAMA")
            
            for col, digit in enumerate(student_no):
                for row in range(10):
                    cx = no_grid_x + col * cell_w + cell_w/2
                    cy = no_grid_y - row * cell_h
                    
                    if str(row) == digit:
                        c.setFillColor(black)
                        c.circle(cx, cy, 1.5*mm, stroke=0, fill=1)
                        c.setFillColor(black)
                    else:
                        c.circle(cx, cy, 1.5*mm, stroke=1, fill=0)
                    
                    if col == 0:
                        c.setFont("Helvetica", 5)
                        c.drawString(no_grid_x - 4*mm, cy - 1*mm, str(row))
            
            subjects = [
                ('TURKCE', counts['turkce'], 12*mm),
                ('MATEMATIK', counts['matematik'], 72*mm),
                ('FEN BILIMLERI', counts['fen'], 132*mm),
            ]
            
            answer_section_y = page_height - 58*mm
            row_height = 5.5*mm
            bubble_radius = 2*mm
            
            for subject_name, q_count, x_start in subjects:
                c.setStrokeColor(black)
                c.setLineWidth(0.3)
                c.rect(x_start - 2*mm, answer_section_y - q_count * row_height - 5*mm, 52*mm, q_count * row_height + 12*mm)
                
                c.setFont("Helvetica-Bold", 8)
                c.drawString(x_start, answer_section_y, f"{subject_name}")
                
                c.setFont("Helvetica-Bold", 6)
                options = ['A', 'B', 'C', 'D']
                for opt_idx, opt in enumerate(options):
                    c.drawCentredString(x_start + 12*mm + opt_idx * 8*mm, answer_section_y - 4*mm, opt)
                
                bubble_start_y = answer_section_y - 10*mm
                
                for q in range(1, q_count + 1):
                    q_y = bubble_start_y - (q-1) * row_height
                    
                    c.setFont("Helvetica", 6)
                    c.drawString(x_start, q_y - 1*mm, f"{q:2d}.")
                    
                    for opt_idx in range(4):
                        opt_x = x_start + 12*mm + opt_idx * 8*mm
                        c.circle(opt_x, q_y, bubble_radius, stroke=1, fill=0)
            
            subjects2 = [
                ('SOSYAL/INKILAP', counts['sosyal'], 12*mm),
                ('DIN KULTURU', counts['din'], 72*mm),
                ('INGILIZCE', counts['ingilizce'], 132*mm),
            ]
            
            if is_56:
                answer_section_y2 = page_height - 165*mm
            else:
                answer_section_y2 = page_height - 185*mm
            
            for subject_name, q_count, x_start in subjects2:
                c.setStrokeColor(black)
                c.setLineWidth(0.3)
                c.rect(x_start - 2*mm, answer_section_y2 - q_count * row_height - 5*mm, 52*mm, q_count * row_height + 12*mm)
                
                c.setFont("Helvetica-Bold", 8)
                c.drawString(x_start, answer_section_y2, f"{subject_name}")
                
                c.setFont("Helvetica-Bold", 6)
                options = ['A', 'B', 'C', 'D']
                for opt_idx, opt in enumerate(options):
                    c.drawCentredString(x_start + 12*mm + opt_idx * 8*mm, answer_section_y2 - 4*mm, opt)
                
                bubble_start_y = answer_section_y2 - 10*mm
                
                for q in range(1, q_count + 1):
                    q_y = bubble_start_y - (q-1) * row_height
                    
                    c.setFont("Helvetica", 6)
                    c.drawString(x_start, q_y - 1*mm, f"{q:2d}.")
                    
                    for opt_idx in range(4):
                        opt_x = x_start + 12*mm + opt_idx * 8*mm
                        c.circle(opt_x, q_y, bubble_radius, stroke=1, fill=0)
            
            c.showPage()
        
        c.save()
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"optik_form_{class_name}.pdf"
        )
        
    except Exception as e:
        logger.error(f"Generate optical PDF error: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500

@app.route("/api/admin/optical/exams/<int:exam_id>/ranked-list")
@login_required
def download_ranked_list(exam_id):
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        list_type = request.args.get('type', 'school')
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("SELECT * FROM optical_exams WHERE id = %s", (exam_id,))
        exam = cur.fetchone()
        
        if not exam:
            return jsonify({"error": "Sınav bulunamadı"}), 404
        
        order_by = "osr.school_rank" if list_type == 'school' else "u.class_name, osr.class_rank"
        
        cur.execute(f"""
            SELECT u.full_name, u.class_name, osr.results, osr.total_net, 
                   osr.total_score, osr.school_rank, osr.class_rank
            FROM optical_student_results osr
            JOIN users u ON osr.student_id = u.id
            WHERE osr.optical_exam_id = %s
            ORDER BY {order_by}
        """, (exam_id,))
        
        results = cur.fetchall()
        cur.close()
        conn.close()
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=30, bottomMargin=30)
        
        styles = getSampleStyleSheet()
        try:
            pdfmetrics.registerFont(TTFont('DejaVu', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
            title_style = ParagraphStyle('Title', fontName='DejaVu', fontSize=14, alignment=1)
            header_style = ParagraphStyle('Header', fontName='DejaVu', fontSize=8, alignment=1)
        except:
            title_style = styles['Heading1']
            header_style = styles['Normal']
        
        elements = []
        elements.append(Paragraph(f"{exam['exam_name']} - Sirali Liste", title_style))
        elements.append(Spacer(1, 20))
        
        headers = ['Sira', 'Ad Soyad', 'Sinif', 'Tur', 'Mat', 'Fen', 'Sos', 'Din', 'Ing', 'Net', 'Puan']
        data = [headers]
        
        for result in results:
            res = json.loads(result['results']) if result['results'] else {}
            rank = result['school_rank'] if list_type == 'school' else result['class_rank']
            
            row = [
                str(rank),
                result['full_name'][:25],
                result['class_name'],
                f"{res.get('turkce', {}).get('net', 0):.1f}",
                f"{res.get('matematik', {}).get('net', 0):.1f}",
                f"{res.get('fen', {}).get('net', 0):.1f}",
                f"{res.get('sosyal', {}).get('net', 0):.1f}",
                f"{res.get('din', {}).get('net', 0):.1f}",
                f"{res.get('ingilizce', {}).get('net', 0):.1f}",
                f"{result['total_net']:.1f}",
                f"{result['total_score']:.0f}"
            ]
            data.append(row)
        
        table = Table(data, colWidths=[30, 150, 40, 40, 40, 40, 40, 40, 40, 50, 50])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(table)
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"sirali_liste_{exam['exam_name']}_{list_type}.pdf"
        )
        
    except Exception as e:
        logger.error(f"Download ranked list error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/admin/optical/exams/<int:exam_id>/failed-questions")
@login_required
def download_failed_questions(exam_id):
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("SELECT * FROM optical_exams WHERE id = %s", (exam_id,))
        exam = cur.fetchone()
        
        if not exam:
            return jsonify({"error": "Sınav bulunamadı"}), 404
        
        question_counts = json.loads(exam['question_counts'])
        answer_key = json.loads(exam['answer_key_a']) if exam['answer_key_a'] else {}
        
        cur.execute("""
            SELECT osr.raw_answers, u.class_name
            FROM optical_student_results osr
            JOIN users u ON osr.student_id = u.id
            WHERE osr.optical_exam_id = %s
        """, (exam_id,))
        
        all_results = cur.fetchall()
        cur.close()
        conn.close()
        
        class_question_stats = {}
        
        for result in all_results:
            class_name = result['class_name']
            raw_answers = json.loads(result['raw_answers'])
            
            if class_name not in class_question_stats:
                class_question_stats[class_name] = {}
            
            for subject, answers in raw_answers.items():
                if subject not in class_question_stats[class_name]:
                    class_question_stats[class_name][subject] = {'total': 0, 'correct': {}}
                
                class_question_stats[class_name][subject]['total'] += 1
                correct_answers = answer_key.get(subject, [])
                
                for i, ans in enumerate(answers):
                    if i not in class_question_stats[class_name][subject]['correct']:
                        class_question_stats[class_name][subject]['correct'][i] = 0
                    
                    if i < len(correct_answers) and ans == correct_answers[i]:
                        class_question_stats[class_name][subject]['correct'][i] += 1
        
        failed_questions = []
        
        for class_name, subjects in class_question_stats.items():
            for subject, data in subjects.items():
                total = data['total']
                if total == 0:
                    continue
                    
                for q_index, correct_count in data['correct'].items():
                    success_rate = (correct_count / total) * 100
                    if success_rate < 50:
                        failed_questions.append({
                            'class': class_name,
                            'subject': subject,
                            'question': q_index + 1,
                            'success_rate': success_rate
                        })
        
        failed_questions.sort(key=lambda x: (x['class'], x['subject'], x['question']))
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30, bottomMargin=30)
        
        styles = getSampleStyleSheet()
        try:
            pdfmetrics.registerFont(TTFont('DejaVu', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
            title_style = ParagraphStyle('Title', fontName='DejaVu', fontSize=14, alignment=1)
        except:
            title_style = styles['Heading1']
        
        elements = []
        elements.append(Paragraph(f"{exam['exam_name']} - Basarisiz Sorular (%50 Alti)", title_style))
        elements.append(Spacer(1, 20))
        
        if not failed_questions:
            elements.append(Paragraph("Tebrikler! %50 altinda basari orani olan soru bulunmamaktadir.", styles['Normal']))
        else:
            data = [['Sinif', 'Ders', 'Soru No', 'Basari %']]
            
            for fq in failed_questions:
                data.append([
                    fq['class'],
                    fq['subject'].title(),
                    str(fq['question']),
                    f"%{fq['success_rate']:.1f}"
                ])
            
            table = Table(data, colWidths=[80, 100, 80, 80])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.red),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ]))
            
            elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"basarisiz_sorular_{exam['exam_name']}.pdf"
        )
        
    except Exception as e:
        logger.error(f"Download failed questions error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# ==================== KARNE PDF ANALİZ SİSTEMİ ====================

@app.route("/api/admin/optical/analyze-report-cards", methods=["POST"])
@login_required
def analyze_report_card_pdfs():
    """Liste PDF'lerini analiz edip ders bazlı başarı raporu oluştur"""
    logger.info("=== LİSTE ANALİZ BAŞLADI ===")
    
    if current_user.role not in ['admin', 'teacher']:
        logger.warning("Yetkisiz erişim denemesi")
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        from PyPDF2 import PdfReader
        from collections import defaultdict
        import re
        
        logger.info("Adım 1: Dosyalar alınıyor")
        files = request.files.getlist('files')
        logger.info(f"Adım 2: {len(files)} dosya alındı")
        if not files:
            return jsonify({"error": "PDF dosyası yüklenmedi"}), 400
        
        for f in files:
            logger.info(f"Dosya: {f.filename}")
        
        # Soru sayıları (7-8. sınıf ve 5-6. sınıf için)
        soru_sayilari_78 = {'turkce': 20, 'sosyal': 10, 'din': 10, 'ingilizce': 10, 'matematik': 20, 'fen': 20}
        soru_sayilari_56 = {'turkce': 15, 'sosyal': 10, 'din': 10, 'ingilizce': 10, 'matematik': 15, 'fen': 15}
        
        ders_isimleri = {
            'turkce': 'Türkçe',
            'sosyal': 'Sosyal Bilgiler',
            'din': 'Din Kültürü',
            'ingilizce': 'İngilizce',
            'matematik': 'Matematik',
            'fen': 'Fen Bilimleri'
        }
        
        # Tüm sonuçları topla
        all_results = {
            'sinif': '',
            'sinav_adi': '',
            'ogrenci_sayisi': 0,
            'sube_ogrenci': defaultdict(int),
            'ders_basari': defaultdict(lambda: {'dogru': 0, 'yanlis': 0, 'toplam': 0}),
            'sube_ders_basari': defaultdict(lambda: defaultdict(lambda: {'dogru': 0, 'yanlis': 0, 'toplam': 0})),
            'okul_ortalama': {},
            'genel_ortalama': {},
            'il_ortalama': {}
        }
        
        for file in files:
            if not file.filename.lower().endswith('.pdf'):
                continue
            
            try:
                reader = PdfReader(file)
                full_text = ""
                
                for page in reader.pages:
                    text = page.extract_text()
                    if text:
                        full_text += text + "\n"
                
                logger.info(f"PDF metin uzunluğu: {len(full_text)}")
                
                # Sınav adını bul
                sinav_match = re.search(r'(\d)\.\s*SINIF\s+([^0-9]+?)\s*-\s*(\d+)', full_text)
                if sinav_match:
                    sinif_seviye = int(sinav_match.group(1))
                    all_results['sinif'] = f"{sinif_seviye}. Sınıf"
                    all_results['sinav_adi'] = f"{sinif_seviye}. SINIF {sinav_match.group(2).strip()} - {sinav_match.group(3)}"
                    soru_sayilari = soru_sayilari_56 if sinif_seviye in [5, 6] else soru_sayilari_78
                else:
                    soru_sayilari = soru_sayilari_78
                    sinif_match = re.search(r'(\d)\.\s*SINIF', full_text)
                    if sinif_match:
                        sinif_seviye = int(sinif_match.group(1))
                        all_results['sinif'] = f"{sinif_seviye}. Sınıf"
                        soru_sayilari = soru_sayilari_56 if sinif_seviye in [5, 6] else soru_sayilari_78
                
                logger.info(f"Sınıf: {all_results['sinif']}, Sınav: {all_results['sinav_adi']}")
                
                # Okul ortalamasını bul - Liste formatında
                okul_ort_match = re.search(r'Okul\s+Ortalaması\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)', full_text)
                if okul_ort_match:
                    all_results['okul_ortalama'] = {
                        'turkce': {'d': float(okul_ort_match.group(1)), 'y': float(okul_ort_match.group(2)), 'n': float(okul_ort_match.group(3))},
                        'sosyal': {'d': float(okul_ort_match.group(4)), 'y': float(okul_ort_match.group(5)), 'n': float(okul_ort_match.group(6))},
                        'din': {'d': float(okul_ort_match.group(7)), 'y': float(okul_ort_match.group(8)), 'n': float(okul_ort_match.group(9))},
                        'ingilizce': {'d': float(okul_ort_match.group(10)), 'y': float(okul_ort_match.group(11)), 'n': float(okul_ort_match.group(12))},
                        'matematik': {'d': float(okul_ort_match.group(13)), 'y': float(okul_ort_match.group(14)), 'n': float(okul_ort_match.group(15))},
                        'fen': {'d': float(okul_ort_match.group(16)), 'y': float(okul_ort_match.group(17)), 'n': float(okul_ort_match.group(18))}
                    }
                    logger.info(f"Okul ortalaması bulundu")
                
                # Helper function to parse compressed Y+N values like '217.33' -> Y=2, N=17.33
                def parse_compressed_yn(value):
                    if '.' not in str(value):
                        return None, None
                    value = str(value)
                    dot_pos = value.index('.')
                    int_part = value[:dot_pos]
                    decimal_part = value[dot_pos:]
                    for n_int_len in [2, 1]:
                        if len(int_part) >= n_int_len:
                            n_int = int_part[-n_int_len:]
                            y_part = int_part[:-n_int_len] if len(int_part) > n_int_len else "0"
                            n_val = float(n_int + decimal_part)
                            y_val = int(y_part) if y_part else 0
                            if -5 <= n_val <= 25:
                                return y_val, n_val
                    return None, None
                
                # Process line by line
                lines = full_text.split('\n')
                for line in lines:
                    # Find class/section pattern: "8 / B" or "8/B"
                    class_match = re.search(r'(\d)\s*/\s*([A-E])', line)
                    if not class_match:
                        continue
                    
                    sinif_num = class_match.group(1)
                    sube_letter = class_match.group(2)
                    sube = f"{sinif_num}/{sube_letter}"
                    
                    # Get text after class/section
                    after_class = line[class_match.end():]
                    
                    # Extract all numbers
                    numbers = re.findall(r'-?[\d.]+', after_class)
                    
                    if len(numbers) < 12:  # At least 6 subjects * 2 values each
                        continue
                    
                    try:
                        # Parse 6 subjects
                        subjects = []
                        i = 0
                        for subj_idx in range(6):
                            if i >= len(numbers):
                                break
                            
                            d = int(float(numbers[i]))
                            i += 1
                            
                            if i >= len(numbers):
                                break
                            
                            # Check if next number is compressed (contains decimal) or separate Y
                            if '.' in numbers[i]:
                                y, n = parse_compressed_yn(numbers[i])
                                if y is None:
                                    y, n = 0, float(numbers[i])
                                i += 1
                            else:
                                y = int(float(numbers[i]))
                                i += 1
                                if i < len(numbers) and '.' in numbers[i]:
                                    n = float(numbers[i])
                                    i += 1
                                else:
                                    n = 0
                            
                            subjects.append({'d': d, 'y': y})
                        
                        if len(subjects) < 6:
                            continue
                        
                        # Successfully parsed a student
                        all_results['ogrenci_sayisi'] += 1
                        all_results['sube_ogrenci'][sube] += 1
                        
                        ders_keys = ['turkce', 'sosyal', 'din', 'ingilizce', 'matematik', 'fen']
                        for idx, ders_key in enumerate(ders_keys):
                            d_val = subjects[idx]['d']
                            y_val = subjects[idx]['y']
                            
                            all_results['ders_basari'][ders_key]['dogru'] += d_val
                            all_results['ders_basari'][ders_key]['yanlis'] += y_val
                            all_results['ders_basari'][ders_key]['toplam'] += soru_sayilari[ders_key]
                            
                            all_results['sube_ders_basari'][sube][ders_key]['dogru'] += d_val
                            all_results['sube_ders_basari'][sube][ders_key]['yanlis'] += y_val
                            all_results['sube_ders_basari'][sube][ders_key]['toplam'] += soru_sayilari[ders_key]
                        
                    except (ValueError, IndexError) as e:
                        logger.error(f"Öğrenci parse hatası: {e}")
                        continue
                    
            except Exception as e:
                logger.error(f"PDF okuma hatası: {str(e)}")
                import traceback
                logger.error(traceback.format_exc())
                continue
        
        logger.info(f"Adım 3: Toplam {all_results['ogrenci_sayisi']} öğrenci bulundu")
        
        if all_results['ogrenci_sayisi'] == 0:
            return jsonify({"error": "PDF'lerden öğrenci verisi çıkarılamadı. Lütfen 'Öğrenci Listesi - Puan Sıralı' formatında PDF yükleyin."}), 400
        
        logger.info("Adım 4: PDF raporu oluşturuluyor")
        # PDF raporu oluştur
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30, bottomMargin=30, leftMargin=40, rightMargin=40)
        
        styles = getSampleStyleSheet()
        try:
            pdfmetrics.registerFont(TTFont('DejaVu', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
            pdfmetrics.registerFont(TTFont('DejaVu-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
            title_style = ParagraphStyle('Title', fontName='DejaVu-Bold', fontSize=16, alignment=1, spaceAfter=20)
            subtitle_style = ParagraphStyle('Subtitle', fontName='DejaVu-Bold', fontSize=12, alignment=1, spaceAfter=10)
            normal_style = ParagraphStyle('Normal', fontName='DejaVu', fontSize=10)
            header_style = ParagraphStyle('Header', fontName='DejaVu-Bold', fontSize=11, spaceAfter=10, spaceBefore=15)
        except:
            title_style = styles['Heading1']
            subtitle_style = styles['Heading2']
            normal_style = styles['Normal']
            header_style = styles['Heading3']
        
        elements = []
        
        # Başlık
        elements.append(Paragraph("KARNE PDF ANALİZ RAPORU", title_style))
        elements.append(Paragraph(f"{all_results['sinif']} - {all_results['sinav_adi']}", subtitle_style))
        elements.append(Spacer(1, 10))
        
        # Özet bilgiler
        elements.append(Paragraph(f"Toplam Öğrenci: {all_results['ogrenci_sayisi']}", normal_style))
        sube_text = ", ".join([f"{s}: {c}" for s, c in sorted(all_results['sube_ogrenci'].items())])
        elements.append(Paragraph(f"Şubeler: {sube_text}", normal_style))
        elements.append(Spacer(1, 15))
        
        # DERS BAZLI BAŞARI
        elements.append(Paragraph("DERS BAZLI BASARI ORANLARI", header_style))
        
        ders_data = [['Ders', 'Basari %', 'Durum']]
        dusuk_dersler = []
        
        for ders_key in ['turkce', 'sosyal', 'din', 'ingilizce', 'matematik', 'fen']:
            stats = all_results['ders_basari'][ders_key]
            if stats['toplam'] > 0:
                basari = (stats['dogru'] / stats['toplam']) * 100
                durum = 'OK' if basari >= 50 else 'DUSUK'
                ders_data.append([ders_isimleri[ders_key], f"%{basari:.1f}", durum])
                if basari < 50:
                    dusuk_dersler.append({'ders': ders_isimleri[ders_key], 'basari': basari})
        
        if len(ders_data) > 1:
            table = Table(ders_data, colWidths=[150, 80, 80])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f9ff')]),
            ]))
            elements.append(table)
        
        elements.append(Spacer(1, 15))
        
        # %50 ALTINDA DERSLER
        elements.append(Paragraph("%50 ALTINDA BASARI GOSTEREN DERSLER", header_style))
        
        if dusuk_dersler:
            dusuk_dersler.sort(key=lambda x: x['basari'])
            dusuk_data = [['Ders', 'Basari %', 'Durum']]
            for d in dusuk_dersler:
                dusuk_data.append([d['ders'], f"%{d['basari']:.1f}", 'ACIL MUDAHALE'])
            
            table = Table(dusuk_data, colWidths=[150, 80, 100])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc2626')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#fef2f2'), colors.HexColor('#fee2e2')]),
            ]))
            elements.append(table)
        else:
            elements.append(Paragraph("Tum derslerde basari orani %50 uzerinde!", normal_style))
        
        elements.append(Spacer(1, 15))
        
        # ŞUBE BAZLI ANALİZ
        elements.append(Paragraph("SUBE BAZLI DERS BASARILARI", header_style))
        
        for sube in sorted(all_results['sube_ders_basari'].keys()):
            elements.append(Paragraph(f"{sube} Subesi ({all_results['sube_ogrenci'][sube]} ogrenci):", normal_style))
            
            sube_ders_data = []
            for ders_key in ['turkce', 'sosyal', 'din', 'ingilizce', 'matematik', 'fen']:
                stats = all_results['sube_ders_basari'][sube][ders_key]
                if stats['toplam'] > 0:
                    basari = (stats['dogru'] / stats['toplam']) * 100
                    durum = 'OK' if basari >= 50 else 'DUSUK'
                    sube_ders_data.append(f"  {ders_isimleri[ders_key]}: %{basari:.1f} ({durum})")
            
            if sube_ders_data:
                elements.append(Paragraph("<br/>".join(sube_ders_data), normal_style))
            elements.append(Spacer(1, 10))
        
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"karne_analiz_raporu_{all_results['sinif'].replace(' ', '_')}.pdf"
        )
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        logger.error(f"Karne analiz hatası: {str(e)}")
        logger.error(f"Traceback: {error_trace}")
        print(f"KARNE HATA: {error_trace}")
        return jsonify({"error": f"Hata: {str(e)}"}), 500

@app.route("/api/student/report-card")
@login_required
def get_student_report_card():
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT * FROM practice_exams 
            WHERE student_id = %s
            ORDER BY exam_number
        """, (current_user.id,))
        
        exams = cur.fetchall()
        
        if not exams:
            cur.close()
            conn.close()
            return jsonify({
                "success": True,
                "report_card": None,
                "message": "Henüz deneme sınavı sonucu bulunmuyor"
            })
        
        total_exams = len(exams)
        
        subject_totals = {
            'turkce': {'dogru': 0, 'yanlis': 0, 'net': 0},
            'matematik': {'dogru': 0, 'yanlis': 0, 'net': 0},
            'fen': {'dogru': 0, 'yanlis': 0, 'net': 0},
            'sosyal': {'dogru': 0, 'yanlis': 0, 'net': 0},
            'din': {'dogru': 0, 'yanlis': 0, 'net': 0},
            'ingilizce': {'dogru': 0, 'yanlis': 0, 'net': 0}
        }
        
        scores = []
        exam_details = []
        
        for exam in exams:
            score = float(exam.get('puan', 0) or 0)
            scores.append(score)
            
            subject_totals['turkce']['dogru'] += exam.get('turkce_dogru', 0) or 0
            subject_totals['turkce']['yanlis'] += exam.get('turkce_yanlis', 0) or 0
            subject_totals['turkce']['net'] += float(exam.get('turkce_net', 0) or 0)
            
            subject_totals['matematik']['dogru'] += exam.get('mat_dogru', 0) or 0
            subject_totals['matematik']['yanlis'] += exam.get('mat_yanlis', 0) or 0
            subject_totals['matematik']['net'] += float(exam.get('mat_net', 0) or 0)
            
            subject_totals['fen']['dogru'] += exam.get('fen_dogru', 0) or 0
            subject_totals['fen']['yanlis'] += exam.get('fen_yanlis', 0) or 0
            subject_totals['fen']['net'] += float(exam.get('fen_net', 0) or 0)
            
            subject_totals['sosyal']['dogru'] += exam.get('sosyal_dogru', 0) or 0
            subject_totals['sosyal']['yanlis'] += exam.get('sosyal_yanlis', 0) or 0
            subject_totals['sosyal']['net'] += float(exam.get('sosyal_net', 0) or 0)
            
            subject_totals['din']['dogru'] += exam.get('din_dogru', 0) or 0
            subject_totals['din']['yanlis'] += exam.get('din_yanlis', 0) or 0
            subject_totals['din']['net'] += float(exam.get('din_net', 0) or 0)
            
            subject_totals['ingilizce']['dogru'] += exam.get('ing_dogru', 0) or 0
            subject_totals['ingilizce']['yanlis'] += exam.get('ing_yanlis', 0) or 0
            subject_totals['ingilizce']['net'] += float(exam.get('ing_net', 0) or 0)
            
            exam_details.append({
                'exam_number': exam['exam_number'],
                'score': score,
                'total_net': float(exam.get('toplam_net', 0) or 0)
            })
        
        subject_averages = {}
        for subject, totals in subject_totals.items():
            subject_averages[subject] = {
                'avg_dogru': round(totals['dogru'] / total_exams, 1),
                'avg_yanlis': round(totals['yanlis'] / total_exams, 1),
                'avg_net': round(totals['net'] / total_exams, 2)
            }
        
        avg_score = sum(scores) / total_exams
        max_score = max(scores)
        min_score = min(scores)
        
        if len(scores) >= 3:
            last_3 = scores[-3:]
            first_3 = scores[:3]
            trend = 'up' if sum(last_3) / 3 > sum(first_3) / 3 else 'down' if sum(last_3) / 3 < sum(first_3) / 3 else 'stable'
        else:
            trend = 'stable'
        
        best_exam = max(exam_details, key=lambda x: x['score'])
        
        strong_subjects = []
        weak_subjects = []
        
        subject_names = {
            'turkce': 'Türkçe', 'matematik': 'Matematik', 'fen': 'Fen Bilimleri',
            'sosyal': 'Sosyal/İnkılap', 'din': 'Din Kültürü', 'ingilizce': 'İngilizce'
        }
        
        for subject, avg in subject_averages.items():
            if avg['avg_net'] >= 10:
                strong_subjects.append(subject_names[subject])
            elif avg['avg_net'] <= 5:
                weak_subjects.append(subject_names[subject])
        
        cur.close()
        conn.close()
        
        return jsonify({
            "success": True,
            "report_card": {
                "student_name": current_user.full_name,
                "class_name": current_user.class_name,
                "total_exams": total_exams,
                "avg_score": round(avg_score, 1),
                "max_score": round(max_score, 1),
                "min_score": round(min_score, 1),
                "trend": trend,
                "best_exam": best_exam,
                "subject_averages": subject_averages,
                "exam_details": exam_details,
                "strong_subjects": strong_subjects,
                "weak_subjects": weak_subjects
            }
        })
        
    except Exception as e:
        logger.error(f"Student report card error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/student/report-card/<int:exam_number>")
@login_required
def get_student_single_report_card(exam_number):
    """Öğrenci tekil deneme karnesi - önceki deneme karşılaştırması ve tavsiyeler"""
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Seçili denemeyi al
        cur.execute("""
            SELECT * FROM practice_exams 
            WHERE student_id = %s AND exam_number = %s
        """, (current_user.id, exam_number))
        current_exam = cur.fetchone()
        
        if not current_exam:
            cur.close()
            conn.close()
            return jsonify({"success": False, "error": "Deneme bulunamadı"}), 404
        
        # Önceki denemeyi al (karşılaştırma için)
        cur.execute("""
            SELECT * FROM practice_exams 
            WHERE student_id = %s AND exam_number < %s
            ORDER BY exam_number DESC LIMIT 1
        """, (current_user.id, exam_number))
        prev_exam = cur.fetchone()
        
        cur.close()
        conn.close()
        
        # Ders bilgilerini hazırla
        subjects = {
            'turkce': {'name': 'Türkçe', 'max': 20, 'net': float(current_exam.get('turkce_net', 0) or 0)},
            'matematik': {'name': 'Matematik', 'max': 20, 'net': float(current_exam.get('matematik_net', 0) or 0)},
            'fen': {'name': 'Fen Bilimleri', 'max': 20, 'net': float(current_exam.get('fen_net', 0) or 0)},
            'sosyal': {'name': 'Sosyal Bilgiler', 'max': 10, 'net': float(current_exam.get('sosyal_net', 0) or 0)},
            'din': {'name': 'Din Kültürü', 'max': 10, 'net': float(current_exam.get('din_net', 0) or 0)},
            'ingilizce': {'name': 'İngilizce', 'max': 10, 'net': float(current_exam.get('ingilizce_net', 0) or 0)}
        }
        
        recommendations = []
        strong_subjects = []
        weak_subjects = []
        
        for key, data in subjects.items():
            net = data['net']
            max_net = data['max']
            name = data['name']
            percentage = (net / max_net) * 100 if max_net > 0 else 0
            
            # Önceki denemeyle karşılaştırma
            change = None
            if prev_exam:
                prev_net = float(prev_exam.get(f'{key}_net', 0) or 0)
                change = net - prev_net
            
            subjects[key]['change'] = change
            
            # Durum belirleme
            if percentage >= 70:
                subjects[key]['status'] = 'strong'
                strong_subjects.append(name)
            elif percentage >= 40:
                subjects[key]['status'] = 'medium'
            else:
                subjects[key]['status'] = 'weak'
                weak_subjects.append(name)
                recommendations.append(f"{name} dersinde daha fazla çalışmalısın (Net: {net:.1f}/{max_net})")
            
            # Düşüş varsa tavsiye ekle
            if change is not None and change < -1:
                recommendations.append(f"{name} dersinde önceki denemeye göre {abs(change):.1f} net düşüş var")
        
        # Genel tavsiyeler
        lgs_score = float(current_exam.get('lgs_score', 0) or 0)
        if lgs_score < 300:
            recommendations.append("LGS puanını yükseltmek için tüm derslere düzenli çalış")
        elif lgs_score < 400:
            recommendations.append("İyi gidiyorsun! Zayıf derslerine odaklanarak puanını artırabilirsin")
        else:
            recommendations.append("Harika bir performans! Aynı tempoda devam et")
        
        return jsonify({
            "success": True,
            "report": {
                "exam_number": exam_number,
                "lgs_score": lgs_score,
                "subjects": subjects,
                "recommendations": recommendations,
                "strong_subjects": strong_subjects,
                "weak_subjects": weak_subjects,
                "has_previous": prev_exam is not None
            }
        })
        
    except Exception as e:
        logger.error(f"Single report card error: {str(e)}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/student/karne-token/<int:exam_number>")
@login_required
def get_karne_token(exam_number):
    """APK için karne PDF token'ı oluştur"""
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz"}), 403
    
    import hashlib
    token = hashlib.sha256(f"{current_user.id}-karne-{exam_number}-ameo2024".encode()).hexdigest()[:16]
    
    return jsonify({
        "success": True,
        "pdf_url": f"/api/student/report-card/{exam_number}/pdf?user_id={current_user.id}&token={token}"
    })

@app.route("/api/student/report-card/<int:exam_number>/pdf")
def download_single_report_card_pdf(exam_number):
    """Zenginlestirilmis tekil deneme karnesi PDF - grafikler, siralamalar, detayli analiz"""
    # APK için token bazlı erişim veya normal login
    user_id = request.args.get('user_id', type=int)
    token = request.args.get('token')
    
    if user_id and token:
        # Token kontrolü - basit hash kontrolü
        import hashlib
        expected_token = hashlib.sha256(f"{user_id}-karne-{exam_number}-ameo2024".encode()).hexdigest()[:16]
        if token != expected_token:
            return jsonify({"error": "Geçersiz token"}), 403
        # Token geçerli, kullanıcıyı bul
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT * FROM users WHERE id = %s AND role = 'student'", (user_id,))
        user_data = cur.fetchone()
        cur.close()
        conn.close()
        if not user_data:
            return jsonify({"error": "Kullanıcı bulunamadı"}), 404
        # User objesi oluştur
        class TempUser:
            def __init__(self, data):
                self.id = data['id']
                self.full_name = data['full_name']
                self.class_name = data['class_name']
                self.role = data['role']
        temp_user = TempUser(user_data)
    elif current_user.is_authenticated:
        if current_user.role != 'student':
            return jsonify({"error": "Yetkisiz"}), 403
        temp_user = current_user
    else:
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        from matplotlib.patches import FancyBboxPatch
        import numpy as np
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        # Seçili denemeyi al
        cur.execute("""
            SELECT * FROM practice_exams 
            WHERE student_id = %s AND exam_number = %s
        """, (temp_user.id, exam_number))
        current_exam = cur.fetchone()
        
        if not current_exam:
            cur.close()
            conn.close()
            return jsonify({"error": "Deneme bulunamadı"}), 404
        
        # Önceki denemeyi al
        cur.execute("""
            SELECT * FROM practice_exams 
            WHERE student_id = %s AND exam_number < %s
            ORDER BY exam_number DESC LIMIT 1
        """, (temp_user.id, exam_number))
        prev_exam = cur.fetchone()
        
        # Toplam net hesaplama SQL ifadesi
        toplam_net_calc = "(COALESCE(pe.turkce_net, 0) + COALESCE(pe.matematik_net, 0) + COALESCE(pe.fen_net, 0) + COALESCE(pe.sosyal_net, 0) + COALESCE(pe.din_net, 0) + COALESCE(pe.ingilizce_net, 0))"
        
        # Okul sıralaması
        cur.execute(f"""
            SELECT pe.student_id, pe.lgs_score, 
                   {toplam_net_calc} as toplam_net,
                   ROW_NUMBER() OVER (ORDER BY pe.lgs_score DESC) as school_rank,
                   COUNT(*) OVER () as school_total
            FROM practice_exams pe
            JOIN users u ON pe.student_id = u.id
            WHERE pe.exam_number = %s
        """, (exam_number,))
        all_school = cur.fetchall()
        
        school_rank = 0
        school_total = len(all_school)
        for r in all_school:
            if r['student_id'] == temp_user.id:
                school_rank = r['school_rank']
                break
        
        # Sınıf sıralaması
        cur.execute(f"""
            SELECT pe.student_id, pe.lgs_score, 
                   {toplam_net_calc} as toplam_net,
                   ROW_NUMBER() OVER (ORDER BY pe.lgs_score DESC) as class_rank,
                   COUNT(*) OVER () as class_total
            FROM practice_exams pe
            JOIN users u ON pe.student_id = u.id
            WHERE pe.exam_number = %s AND u.class_name = %s
        """, (exam_number, temp_user.class_name))
        class_results = cur.fetchall()
        
        class_rank = 0
        class_total = len(class_results)
        for r in class_results:
            if r['student_id'] == temp_user.id:
                class_rank = r['class_rank']
                break
        
        # Sınıf ve okul ortalamaları
        cur.execute("""
            SELECT AVG(pe.lgs_score) as class_avg, 
                   AVG(COALESCE(pe.turkce_net, 0) + COALESCE(pe.matematik_net, 0) + COALESCE(pe.fen_net, 0) + COALESCE(pe.sosyal_net, 0) + COALESCE(pe.din_net, 0) + COALESCE(pe.ingilizce_net, 0)) as class_net_avg
            FROM practice_exams pe
            JOIN users u ON pe.student_id = u.id
            WHERE pe.exam_number = %s AND u.class_name = %s
        """, (exam_number, temp_user.class_name))
        class_avg_data = cur.fetchone()
        class_lgs_avg = float(class_avg_data['class_avg'] or 0) if class_avg_data else 0
        
        cur.execute("""
            SELECT AVG(lgs_score) as school_avg, 
                   AVG(COALESCE(turkce_net, 0) + COALESCE(matematik_net, 0) + COALESCE(fen_net, 0) + COALESCE(sosyal_net, 0) + COALESCE(din_net, 0) + COALESCE(ingilizce_net, 0)) as school_net_avg
            FROM practice_exams WHERE exam_number = %s
        """, (exam_number,))
        school_avg_data = cur.fetchone()
        school_lgs_avg = float(school_avg_data['school_avg'] or 0) if school_avg_data else 0
        
        # Öğrencinin tüm denemeleri (trend için)
        cur.execute("""
            SELECT exam_number, lgs_score, 
                   (COALESCE(turkce_net, 0) + COALESCE(matematik_net, 0) + COALESCE(fen_net, 0) + COALESCE(sosyal_net, 0) + COALESCE(din_net, 0) + COALESCE(ingilizce_net, 0)) as toplam_net
            FROM practice_exams
            WHERE student_id = %s ORDER BY exam_number
        """, (temp_user.id,))
        all_exams = cur.fetchall()
        
        cur.close()
        conn.close()
        
        # PDF Oluştur
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=25, bottomMargin=25, leftMargin=30, rightMargin=30)
        
        styles = getSampleStyleSheet()
        try:
            pdfmetrics.registerFont(TTFont('DejaVu', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
            pdfmetrics.registerFont(TTFont('DejaVuBold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
            title_style = ParagraphStyle('Title', fontName='DejaVuBold', fontSize=18, alignment=1, spaceAfter=10, textColor=colors.HexColor('#1e3a5f'))
            subtitle_style = ParagraphStyle('Subtitle', fontName='DejaVu', fontSize=11, alignment=1, spaceAfter=8, textColor=colors.HexColor('#4a5568'))
            section_style = ParagraphStyle('Section', fontName='DejaVuBold', fontSize=12, spaceAfter=8, textColor=colors.HexColor('#2563eb'))
            normal_style = ParagraphStyle('Normal', fontName='DejaVu', fontSize=10, spaceAfter=4)
            rec_style = ParagraphStyle('Rec', fontName='DejaVu', fontSize=9, leftIndent=15, spaceAfter=3, textColor=colors.HexColor('#059669'))
            warning_style = ParagraphStyle('Warning', fontName='DejaVu', fontSize=9, leftIndent=15, spaceAfter=3, textColor=colors.HexColor('#dc2626'))
        except:
            title_style = styles['Heading1']
            subtitle_style = styles['Heading2']
            section_style = styles['Heading3']
            normal_style = styles['Normal']
            rec_style = styles['Normal']
            warning_style = styles['Normal']
        
        elements = []
        
        # Başlık
        elements.append(Paragraph("AYSE MELAHAT ERKIN ORTAOKULU", title_style))
        elements.append(Paragraph(f"{exam_number}. DENEME SINAVI KARNESI", subtitle_style))
        elements.append(Paragraph(f"{temp_user.full_name} - {temp_user.class_name}", subtitle_style))
        elements.append(Spacer(1, 10))
        
        lgs_score = float(current_exam.get('lgs_score', 0) or 0)
        # Toplam net'i hesapla (sütun olmadığı için)
        toplam_net = (
            float(current_exam.get('turkce_net', 0) or 0) +
            float(current_exam.get('matematik_net', 0) or 0) +
            float(current_exam.get('fen_net', 0) or 0) +
            float(current_exam.get('sosyal_net', 0) or 0) +
            float(current_exam.get('din_net', 0) or 0) +
            float(current_exam.get('ingilizce_net', 0) or 0)
        )
        
        # Ana Özet Tablosu (renkli)
        summary_data = [
            ['LGS PUANI', 'TOPLAM NET', 'SINIF SIRALAMASI', 'OKUL SIRALAMASI'],
            [f'{lgs_score:.1f}', f'{toplam_net:.1f}', f'{class_rank}/{class_total}', f'{school_rank}/{school_total}']
        ]
        summary_table = Table(summary_data, colWidths=[120, 100, 120, 120])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('BACKGROUND', (0, 1), (0, 1), colors.HexColor('#3b82f6')),
            ('BACKGROUND', (1, 1), (1, 1), colors.HexColor('#10b981')),
            ('BACKGROUND', (2, 1), (2, 1), colors.HexColor('#f59e0b')),
            ('BACKGROUND', (3, 1), (3, 1), colors.HexColor('#8b5cf6')),
            ('TEXTCOLOR', (0, 1), (-1, 1), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, 1), 14),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#1e3a5f')),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 15))
        
        # Karşılaştırma tablosu
        comparison_data = [
            ['', 'SENIN PUANIN', 'SINIF ORT.', 'OKUL ORT.'],
            ['LGS Puani', f'{lgs_score:.1f}', f'{class_lgs_avg:.1f}', f'{school_lgs_avg:.1f}']
        ]
        comparison_table = Table(comparison_data, colWidths=[100, 120, 120, 120])
        diff_class = lgs_score - class_lgs_avg
        diff_school = lgs_score - school_lgs_avg
        comparison_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#374151')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db')),
            ('BACKGROUND', (1, 1), (1, 1), colors.HexColor('#dbeafe')),
        ]))
        elements.append(comparison_table)
        elements.append(Spacer(1, 5))
        
        # Ortalama karşılaştırma notu
        if diff_class > 0:
            elements.append(Paragraph(f"Sinif ortalamasinin {diff_class:.1f} puan ustundesin!", rec_style))
        else:
            elements.append(Paragraph(f"Sinif ortalamasinin {abs(diff_class):.1f} puan altindasin.", warning_style))
        elements.append(Spacer(1, 10))
        
        # Ders Performans Grafiği oluştur
        subjects_data = [
            ('Turkce', 'turkce_net', 20),
            ('Matematik', 'matematik_net', 20),
            ('Fen', 'fen_net', 20),
            ('Sosyal', 'sosyal_net', 10),
            ('Din K.', 'din_net', 10),
            ('Ingilizce', 'ingilizce_net', 10)
        ]
        
        fig, axes = plt.subplots(1, 2, figsize=(10, 3.5))
        
        # Sol grafik: Ders bazlı performans çubuğu
        subject_names = [s[0] for s in subjects_data]
        nets = [float(current_exam.get(s[1], 0) or 0) for s in subjects_data]
        max_vals = [s[2] for s in subjects_data]
        percentages = [(n/m)*100 for n, m in zip(nets, max_vals)]
        
        bar_colors = ['#ef4444' if p < 40 else '#f59e0b' if p < 70 else '#10b981' for p in percentages]
        bars = axes[0].barh(subject_names, percentages, color=bar_colors, height=0.6)
        axes[0].set_xlim(0, 100)
        axes[0].set_xlabel('Basari Yuzdesi (%)', fontsize=9)
        axes[0].set_title('Ders Bazli Performans', fontsize=11, fontweight='bold', color='#1e3a5f')
        axes[0].axvline(x=70, color='#10b981', linestyle='--', alpha=0.7, label='Hedef %70')
        axes[0].axvline(x=40, color='#f59e0b', linestyle='--', alpha=0.7, label='Minimum %40')
        
        for i, (bar, pct, net) in enumerate(zip(bars, percentages, nets)):
            axes[0].text(pct + 2, bar.get_y() + bar.get_height()/2, f'{net:.1f}', va='center', fontsize=8)
        
        # Sağ grafik: LGS puan trendi (renkli sütun grafik)
        if len(all_exams) > 1:
            exam_nums = [e['exam_number'] for e in all_exams]
            lgs_scores = [float(e['lgs_score'] or 0) for e in all_exams]
            # Renk gradyanı - düşük puanlar kırmızı, yüksek puanlar yeşil
            max_score = max(lgs_scores) if lgs_scores else 500
            min_score = min(lgs_scores) if lgs_scores else 0
            bar_colors = []
            for score in lgs_scores:
                if max_score == min_score:
                    bar_colors.append('#3b82f6')
                else:
                    ratio = (score - min_score) / (max_score - min_score)
                    if ratio < 0.33:
                        bar_colors.append('#ef4444')  # Kırmızı
                    elif ratio < 0.66:
                        bar_colors.append('#f59e0b')  # Turuncu
                    else:
                        bar_colors.append('#10b981')  # Yeşil
            
            bars = axes[1].bar([str(n) for n in exam_nums], lgs_scores, color=bar_colors, edgecolor='#1e3a5f', linewidth=1)
            axes[1].set_xlabel('Deneme No', fontsize=9)
            axes[1].set_ylabel('LGS Puani', fontsize=9)
            axes[1].set_title('LGS Puan Gelisimi', fontsize=11, fontweight='bold', color='#1e3a5f')
            # Her sütunun üzerine değer yaz
            for bar, score in zip(bars, lgs_scores):
                axes[1].text(bar.get_x() + bar.get_width()/2, bar.get_height() + 5, f'{score:.0f}', 
                            ha='center', va='bottom', fontsize=8, fontweight='bold')
            axes[1].set_ylim(0, max(lgs_scores) * 1.15 if lgs_scores else 500)
        else:
            # Tek deneme varsa net dağılımı sütun grafik
            bar_colors_single = ['#ef4444', '#3b82f6', '#10b981', '#f59e0b', '#8b5cf6', '#06b6d4']
            bars = axes[1].bar(subject_names, nets, color=bar_colors_single, edgecolor='#1e3a5f', linewidth=1)
            axes[1].set_ylabel('Net', fontsize=9)
            axes[1].set_title('Ders Bazli Net Dagilimi', fontsize=11, fontweight='bold', color='#1e3a5f')
            for bar, net in zip(bars, nets):
                axes[1].text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.3, f'{net:.1f}', 
                            ha='center', va='bottom', fontsize=8, fontweight='bold')
            axes[1].tick_params(axis='x', rotation=45)
        
        plt.tight_layout()
        
        # Grafiği buffer'a kaydet
        chart_buffer = io.BytesIO()
        plt.savefig(chart_buffer, format='png', dpi=150, bbox_inches='tight', facecolor='white')
        plt.close()
        chart_buffer.seek(0)
        
        chart_img = Image(chart_buffer, width=480, height=170)
        elements.append(chart_img)
        elements.append(Spacer(1, 10))
        
        # Detaylı Ders Tablosu
        elements.append(Paragraph("DERS BAZLI DETAYLI ANALIZ", section_style))
        
        table_data = [['Ders', 'Net', 'Maks', 'Yuzde', 'Degisim', 'Durum']]
        recommendations = []
        strong_subjects = []
        weak_subjects = []
        
        for name, key, max_val in subjects_data:
            net = float(current_exam.get(key, 0) or 0)
            percentage = (net / max_val) * 100 if max_val > 0 else 0
            
            change_str = '-'
            if prev_exam:
                prev_net = float(prev_exam.get(key, 0) or 0)
                change = net - prev_net
                if change > 0:
                    change_str = f'+{change:.1f}'
                elif change < 0:
                    change_str = f'{change:.1f}'
                    if change < -1:
                        recommendations.append(f"{name} dersinde {abs(change):.1f} net dusus var, bu dersi gozden gecir.")
                else:
                    change_str = '='
            
            if percentage >= 70:
                status = 'GUCLU'
                strong_subjects.append(name)
            elif percentage >= 40:
                status = 'ORTA'
            else:
                status = 'ZAYIF'
                weak_subjects.append(name)
                recommendations.append(f"{name} dersine daha fazla zaman ayirmalisin.")
            
            table_data.append([name, f'{net:.1f}', str(max_val), f'%{percentage:.0f}', change_str, status])
        
        table = Table(table_data, colWidths=[80, 50, 50, 55, 60, 60])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
        ]))
        
        # Durum sütunu renklendirme
        for i, row in enumerate(table_data[1:], 1):
            if row[5] == 'GUCLU':
                table.setStyle(TableStyle([('BACKGROUND', (5, i), (5, i), colors.HexColor('#dcfce7'))]))
                table.setStyle(TableStyle([('TEXTCOLOR', (5, i), (5, i), colors.HexColor('#166534'))]))
            elif row[5] == 'ZAYIF':
                table.setStyle(TableStyle([('BACKGROUND', (5, i), (5, i), colors.HexColor('#fee2e2'))]))
                table.setStyle(TableStyle([('TEXTCOLOR', (5, i), (5, i), colors.HexColor('#991b1b'))]))
        
        elements.append(table)
        elements.append(Spacer(1, 15))
        
        # Yüzdelik dilim
        if school_total > 0:
            percentile = ((school_total - school_rank) / school_total) * 100
            elements.append(Paragraph(f"YUZDELIK DILIM: Okulun en iyi %{percentile:.0f}'lik diliminde yer aliyorsun.", section_style))
            elements.append(Spacer(1, 10))
        
        # Güçlü ve Zayıf Dersler
        col_data = []
        if strong_subjects:
            col_data.append(['GUCLU DERSLER', ', '.join(strong_subjects)])
        if weak_subjects:
            col_data.append(['GELISTIRILECEK DERSLER', ', '.join(weak_subjects)])
        
        if col_data:
            strength_table = Table(col_data, colWidths=[150, 300])
            strength_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#10b981')),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.white),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d1d5db')),
            ]))
            if len(col_data) > 1:
                strength_table.setStyle(TableStyle([('BACKGROUND', (0, 1), (0, 1), colors.HexColor('#ef4444'))]))
            elements.append(strength_table)
            elements.append(Spacer(1, 15))
        
        # Tavsiyeler ve Hedefler
        elements.append(Paragraph("ONERILER VE HEDEFLER", section_style))
        
        if lgs_score < 300:
            recommendations.append("LGS puanini yukseltmek icin tum derslere duzenli calis.")
            recommendations.append("Gunluk en az 2 saat ders calismasi hedefle.")
        elif lgs_score < 400:
            recommendations.append("Iyi gidiyorsun! Zayif derslerine odaklanarak puanini artirabilirsin.")
            recommendations.append("Konu eksiklerini gidermek icin video dersler izle.")
        else:
            recommendations.append("Harika bir performans! Ayni tempoda devam et.")
            recommendations.append("Yari cozdugun konulara odaklanarak %100'e ulasmaya calis.")
        
        for rec in recommendations:
            elements.append(Paragraph(f"* {rec}", rec_style))
        
        elements.append(Spacer(1, 15))
        
        # Alt bilgi
        from datetime import datetime
        elements.append(Paragraph(f"Olusturulma Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}", 
                                  ParagraphStyle('Footer', fontSize=8, alignment=2, textColor=colors.gray)))
        
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=False,
            download_name=f'karne_deneme_{exam_number}_{temp_user.full_name.replace(" ", "_")}.pdf'
        )
        
    except Exception as e:
        logger.error(f"Single report card PDF error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/student/report-card/pdf")
@login_required
def download_student_report_card_pdf():
    if current_user.role != 'student':
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        cur.execute("""
            SELECT * FROM practice_exams 
            WHERE student_id = %s
            ORDER BY exam_number
        """, (current_user.id,))
        
        exams = cur.fetchall()
        cur.close()
        conn.close()
        
        if not exams:
            return jsonify({"error": "Henüz deneme sınavı sonucu bulunmuyor"}), 404
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30, bottomMargin=30)
        
        styles = getSampleStyleSheet()
        try:
            pdfmetrics.registerFont(TTFont('DejaVu', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
            title_style = ParagraphStyle('Title', fontName='DejaVu', fontSize=16, alignment=1, spaceAfter=20)
            subtitle_style = ParagraphStyle('Subtitle', fontName='DejaVu', fontSize=12, alignment=1, spaceAfter=10)
            normal_style = ParagraphStyle('Normal', fontName='DejaVu', fontSize=10)
        except:
            title_style = styles['Heading1']
            subtitle_style = styles['Heading2']
            normal_style = styles['Normal']
        
        elements = []
        
        elements.append(Paragraph("OGRENCI DENEME SINAVI KARNESI", title_style))
        elements.append(Paragraph(f"{current_user.full_name} - {current_user.class_name}", subtitle_style))
        elements.append(Spacer(1, 20))
        
        total_exams = len(exams)
        scores = [float(e.get('lgs_score', 0) or 0) for e in exams]
        avg_score = sum(scores) / total_exams if scores else 0
        max_score = max(scores) if scores else 0
        
        summary_data = [
            ['Toplam Deneme', 'Ortalama Puan', 'En Yuksek Puan'],
            [str(total_exams), f'{avg_score:.1f}', f'{max_score:.1f}']
        ]
        
        summary_table = Table(summary_data, colWidths=[150, 150, 150])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
        
        elements.append(Paragraph("Deneme Bazli Sonuclar", subtitle_style))
        
        exam_headers = ['Deneme', 'Tur', 'Mat', 'Fen', 'Sos', 'Din', 'Ing', 'LGS']
        exam_data = [exam_headers]
        
        for exam in exams:
            row = [
                str(exam['exam_number']),
                f"{float(exam.get('turkce_net', 0) or 0):.1f}",
                f"{float(exam.get('matematik_net', 0) or 0):.1f}",
                f"{float(exam.get('fen_net', 0) or 0):.1f}",
                f"{float(exam.get('sosyal_net', 0) or 0):.1f}",
                f"{float(exam.get('din_net', 0) or 0):.1f}",
                f"{float(exam.get('ingilizce_net', 0) or 0):.1f}",
                f"{float(exam.get('lgs_score', 0) or 0):.0f}"
            ]
            exam_data.append(row)
        
        exam_table = Table(exam_data, colWidths=[40, 45, 45, 45, 45, 45, 45, 50])
        exam_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        elements.append(exam_table)
        
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=False,
            download_name=f"karne_{current_user.full_name.replace(' ', '_')}.pdf"
        )
        
    except Exception as e:
        logger.error(f"Student report card PDF error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/teacher/ranked-list/<int:exam_number>")
@login_required
def get_teacher_ranked_list(exam_number):
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        list_type = request.args.get('type', 'school')
        class_filter = request.args.get('class_name', '')
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        if list_type == 'class' and class_filter:
            cur.execute("""
                SELECT pe.*, u.full_name, u.class_name
                FROM practice_exams pe
                JOIN users u ON pe.student_id = u.id
                WHERE pe.exam_number = %s AND u.class_name = %s
                ORDER BY pe.toplam_net DESC
            """, (exam_number, class_filter))
        else:
            cur.execute("""
                SELECT pe.*, u.full_name, u.class_name
                FROM practice_exams pe
                JOIN users u ON pe.student_id = u.id
                WHERE pe.exam_number = %s
                ORDER BY pe.toplam_net DESC
            """, (exam_number,))
        
        results = cur.fetchall()
        
        for i, result in enumerate(results, 1):
            result['rank'] = i
        
        cur.close()
        conn.close()
        
        return jsonify({"success": True, "results": results})
        
    except Exception as e:
        logger.error(f"Teacher ranked list error: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/teacher/ranked-list/<int:exam_number>/pdf")
@login_required
def download_teacher_ranked_list_pdf(exam_number):
    if current_user.role not in ['admin', 'teacher']:
        return jsonify({"error": "Yetkisiz"}), 403
    
    try:
        list_type = request.args.get('type', 'school')
        grade_level = request.args.get('grade_level', '')
        
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=30, bottomMargin=30, leftMargin=20, rightMargin=20)
        
        styles = getSampleStyleSheet()
        try:
            pdfmetrics.registerFont(TTFont('DejaVu', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
            pdfmetrics.registerFont(TTFont('DejaVu-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
            title_style = ParagraphStyle('Title', fontName='DejaVu-Bold', fontSize=14, alignment=1, spaceAfter=10)
            subtitle_style = ParagraphStyle('Subtitle', fontName='DejaVu-Bold', fontSize=12, alignment=1, spaceAfter=5)
        except:
            title_style = styles['Heading1']
            subtitle_style = styles['Heading2']
        
        elements = []
        
        # Logo ve başlık ekle
        header_elements = create_pdf_header(styles)
        elements.extend(header_elements)
        elements.append(Spacer(1, 10))
        
        if list_type == 'class':
            # Şube bazlı sıralama - her şube kendi içinde sıralanır
            if grade_level:
                # Belirli sınıf seviyesi (5, 6, 7, 8)
                cur.execute("""
                    SELECT pe.*, u.full_name, u.class_name,
                           (COALESCE(pe.turkce_net, 0) + COALESCE(pe.matematik_net, 0) + 
                            COALESCE(pe.fen_net, 0) + COALESCE(pe.sosyal_net, 0) + 
                            COALESCE(pe.din_net, 0) + COALESCE(pe.ingilizce_net, 0)) as toplam_net
                    FROM practice_exams pe
                    JOIN users u ON pe.student_id = u.id
                    WHERE pe.exam_number = %s AND u.class_name LIKE %s
                    ORDER BY u.class_name, toplam_net DESC
                """, (exam_number, f"{grade_level}%"))
            else:
                # Tüm sınıflar
                cur.execute("""
                    SELECT pe.*, u.full_name, u.class_name,
                           (COALESCE(pe.turkce_net, 0) + COALESCE(pe.matematik_net, 0) + 
                            COALESCE(pe.fen_net, 0) + COALESCE(pe.sosyal_net, 0) + 
                            COALESCE(pe.din_net, 0) + COALESCE(pe.ingilizce_net, 0)) as toplam_net
                    FROM practice_exams pe
                    JOIN users u ON pe.student_id = u.id
                    WHERE pe.exam_number = %s
                    ORDER BY u.class_name, toplam_net DESC
                """, (exam_number,))
            
            results = cur.fetchall()
            
            elements.append(Paragraph(f"{exam_number}. Deneme Sıralı Listesi - Şube Bazlı", title_style))
            elements.append(Spacer(1, 15))
            
            # Şubelere göre grupla
            class_groups = {}
            for r in results:
                cn = r['class_name']
                if cn not in class_groups:
                    class_groups[cn] = []
                class_groups[cn].append(r)
            
            # Her şube için ayrı tablo
            for class_name in sorted(class_groups.keys()):
                class_results = class_groups[class_name]
                
                elements.append(Paragraph(f"📚 {class_name} Sınıfı", subtitle_style))
                elements.append(Spacer(1, 5))
                
                headers = ['Sıra', 'Ad Soyad', 'Tür', 'Mat', 'Fen', 'Sos', 'Din', 'İng', 'Net', 'Puan']
                data = [headers]
                
                for i, result in enumerate(class_results, 1):
                    row = [
                        str(i),
                        result['full_name'][:25],
                        f"{result.get('turkce_net', 0) or 0:.1f}",
                        f"{result.get('matematik_net', 0) or 0:.1f}",
                        f"{result.get('fen_net', 0) or 0:.1f}",
                        f"{result.get('sosyal_net', 0) or 0:.1f}",
                        f"{result.get('din_net', 0) or 0:.1f}",
                        f"{result.get('ingilizce_net', 0) or 0:.1f}",
                        f"{result.get('toplam_net', 0) or 0:.1f}",
                        f"{result.get('lgs_score', 0) or 0:.0f}"
                    ]
                    data.append(row)
                
                table = Table(data, colWidths=[30, 150, 40, 40, 40, 40, 40, 40, 50, 50])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, -1), 'DejaVu'),
                    ('FONTNAME', (0, 0), (-1, 0), 'DejaVu-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
                ]))
                
                elements.append(table)
                elements.append(Spacer(1, 15))
            
            title_suffix = "sube_bazli"
        else:
            # Okul geneli sıralama
            if grade_level:
                cur.execute("""
                    SELECT pe.*, u.full_name, u.class_name,
                           (COALESCE(pe.turkce_net, 0) + COALESCE(pe.matematik_net, 0) + 
                            COALESCE(pe.fen_net, 0) + COALESCE(pe.sosyal_net, 0) + 
                            COALESCE(pe.din_net, 0) + COALESCE(pe.ingilizce_net, 0)) as toplam_net
                    FROM practice_exams pe
                    JOIN users u ON pe.student_id = u.id
                    WHERE pe.exam_number = %s AND u.class_name LIKE %s
                    ORDER BY toplam_net DESC
                """, (exam_number, f"{grade_level}%"))
                title_suffix = f"sinif_{grade_level}"
            else:
                cur.execute("""
                    SELECT pe.*, u.full_name, u.class_name,
                           (COALESCE(pe.turkce_net, 0) + COALESCE(pe.matematik_net, 0) + 
                            COALESCE(pe.fen_net, 0) + COALESCE(pe.sosyal_net, 0) + 
                            COALESCE(pe.din_net, 0) + COALESCE(pe.ingilizce_net, 0)) as toplam_net
                    FROM practice_exams pe
                    JOIN users u ON pe.student_id = u.id
                    WHERE pe.exam_number = %s
                    ORDER BY toplam_net DESC
                """, (exam_number,))
                title_suffix = "okul_geneli"
            
            results = cur.fetchall()
            
            grade_text = f"{grade_level}. Sınıflar" if grade_level else "Okul Geneli"
            elements.append(Paragraph(f"{exam_number}. Deneme Sıralı Listesi - {grade_text}", title_style))
            elements.append(Spacer(1, 15))
            
            headers = ['Sıra', 'Ad Soyad', 'Sınıf', 'Tür', 'Mat', 'Fen', 'Sos', 'Din', 'İng', 'Net', 'Puan']
            data = [headers]
            
            for i, result in enumerate(results, 1):
                row = [
                    str(i),
                    result['full_name'][:25],
                    result['class_name'],
                    f"{result.get('turkce_net', 0) or 0:.1f}",
                    f"{result.get('matematik_net', 0) or 0:.1f}",
                    f"{result.get('fen_net', 0) or 0:.1f}",
                    f"{result.get('sosyal_net', 0) or 0:.1f}",
                    f"{result.get('din_net', 0) or 0:.1f}",
                    f"{result.get('ingilizce_net', 0) or 0:.1f}",
                    f"{result.get('toplam_net', 0) or 0:.1f}",
                    f"{result.get('lgs_score', 0) or 0:.0f}"
                ]
                data.append(row)
            
            table = Table(data, colWidths=[30, 150, 40, 40, 40, 40, 40, 40, 40, 50, 50])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), 'DejaVu'),
                ('FONTNAME', (0, 0), (-1, 0), 'DejaVu-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
            ]))
            
            elements.append(table)
        
        cur.close()
        conn.close()
        
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"sirali_liste_{exam_number}_deneme_{title_suffix}.pdf"
        )
        
    except Exception as e:
        logger.error(f"Teacher ranked list PDF error: {str(e)}")
        return jsonify({"error": str(e)}), 500

# ==================== OPTİK FORM YÖNETİMİ SONU ====================

# ==================== MODÜLER BLUEPRINT'LER ====================
from routes.study_plan import study_plan_bp
from routes.question_analysis import question_analysis_bp
from routes.coaching import coaching_bp
from routes.schedule import schedule_bp
from routes.question_asks import question_asks_bp
from routes.daily_tracking import daily_tracking_bp
# from routes.exam_calendar import exam_calendar_bp  # Disabled - using app.py endpoints instead
from routes.teacher_study_plan import teacher_study_plan_bp
from routes.report_cards import report_cards_bp, init_object_storage

# Object Storage'ı report_cards modülüne aktar
init_object_storage(object_storage)

app.register_blueprint(study_plan_bp)
app.register_blueprint(question_analysis_bp)
app.register_blueprint(coaching_bp)
app.register_blueprint(schedule_bp)
app.register_blueprint(question_asks_bp)
app.register_blueprint(daily_tracking_bp)
# app.register_blueprint(exam_calendar_bp)  # Disabled - using app.py endpoints instead
app.register_blueprint(teacher_study_plan_bp)
app.register_blueprint(report_cards_bp)
# ==================== MODÜLER BLUEPRINT'LER SONU ====================

# Uygulama başlarken veritabanını initialize et
with app.app_context():
    init_database()
    
    init_admin_user()
    
    init_default_classes()

if __name__ == "__main__":
    # Production'da debug=False, development'ta debug=True
    debug_mode = os.environ.get('FLASK_ENV') != 'production'
    app.run(host="0.0.0.0", port=5000, debug=debug_mode)

# Temp test logging
import sys
